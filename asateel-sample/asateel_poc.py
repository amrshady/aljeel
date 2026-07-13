#!/usr/bin/env python3
from __future__ import annotations

import base64
import argparse
import hashlib
import json
import math
import os
import platform
import re
import shutil
import subprocess
import sys
import time
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
from collections import Counter
from copy import copy
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill
from rapidfuzz import fuzz, process

ROOT = Path("/home/clawdbot/.openclaw/workspace/aljeel")
ASATEEL_PIPELINE_VERSION = "2026-07-13.2"
SAMPLE_ROOT = ROOT / "asateel-sample"
PDF_ROOT = SAMPLE_ROOT / "_pdfs"
OUT_DIR = SAMPLE_ROOT / "_poc_out"
RENDER_DIR = OUT_DIR / "_rendered"
CACHE_DIR = OUT_DIR / "_cache"
MASTER_XLSX = ROOT / "qc/master-data/Aljeel_Lookups-v2.xlsx"
JAWAL_TEMPLATE_XLSX = ROOT / "batches/jawal-J26-640/output/Spreadsheet-J26-640-FILLED-v30.xlsx"
ENTRY_FILES = [
    SAMPLE_ROOT / "_allocation" / "Entry-1.xlsm",
    SAMPLE_ROOT / "_allocation" / "Entry-2.xlsm",
]
DEFAULT_EXPENSES_FORMAT_XLSX = SAMPLE_ROOT / "_allocation" / "Central-11-2026.xlsx"
DEFAULT_SO_DETAIL_XLSX = ROOT / "reference" / "SO_Detail_Labadi_1_R21_AA.xlsx"
DEFAULT_PROJECT_ALLOCATION_LOOKUP = ROOT / "pipelines" / "lookups" / "asateel_projects_labadi_v1.json"
STANDARD_ALLOCATION_MODE = "standard"
PROJECT_ALLOCATION_MODE = "projects-labadi-v1"

GL_ACCOUNT = "61500027"
GL_FALLBACK_DESC = "Transportation/Freight Expense"
COMPANY = "03"
DEFAULT_LOCATION = "20100"
SUPPLIER_NAME = "شركة اساطيل الطريق للنقل البري"
BUSINESS_UNIT = "Al Jeel Medical BU"
GEMINI_BASE_URL = os.environ.get(
    "GEMINI_BASE_URL",
    "https://gateway.ai.cloudflare.com/v1/3724a3e71944b366a39b3735aa117a58/accord-aljeel-ap/google-ai-studio/v1beta",
)
GEMINI_MODEL_CASCADE = ["gemini-pro-latest", "gemini-2.5-pro", "gemini-2.5-flash"]
VAT_RATE = 0.15
_PRINTED_GEMINI_HOST = False

SAMPLES = [
    ("PROJECTS", "مشاريع 13-2026", ["03048", "03049", "03050", "03051", "03052"]),
    ("ADMIN", "اداره 8-2026", ["03063", "03064", "03065", "03066", "03088"]),
    ("CENTRAL", "وسطي 11-2026", ["03041", "03042", "03043", "03044", "03045"]),
]

FOLDER_NAMES = {label: folder for label, folder, _ in SAMPLES}

COLUMNS = [
    "folder", "invoice_no", "invoice_date", "line_no", "description", "reference",
    "dispatch_ref", "supply_order", "line_amount", "vat", "total", "GL",
    "Agency_code", "Agency_name", "Division", "Cost_Center", "Cost_Center_name",
    "allocation_source", "split_method", "confidence", "Row_Status", "notes",
]

DEBUG_HEADERS = [
    "Row Status",
    "Allocation Source",
    "Agency Resolve Method",
    "Additional Information",
    "SO_Detail Agency",
    "SO_Detail Salesperson",
    "Supplier Sheet Agency",
    "Manpower Home Agency",
    "SO_Detail vs Supplier Discrepancy",
    "Home Agency Discrepancy",
    "Supplier JQ Count",
    "Extracted Brand(s)",
    "Extracted Salesperson",
    "Agency Match Confidence",
    "Manpower Cluster Found (Y/N)",
    "Split Method",
    "Reference(المرجع)",
    "Dispatch Ref",
    "Supply Order",
    "VAT Amount",
    "Notes",
    "Folder",
    "Invoice Hint",
    "Line No",
    "Resolved Source Raw",
    "Resolved Match Method",
    "Status Reason",
    "Extraction Notes",
    "Header Subtotal",
    "Header Total",
    "Amount Basis",
    "Location Assumption",
    "Account Lookup Note",
    "Trace PDF",
    "Agent Method",
    "Project Allocation Mode",
    "Project Lookup Status",
    "Project Lookup Explanation",
]

SEGMENT_WIDTHS = {
    "company": 2,
    "location": 5,
    "account": 8,
    "cost_center": 6,
    "div": 3,
    "solution": 5,
    "agency": 5,
    "project": 5,
    "intercompany": 2,
    "future": 6,
}


def _load_env_key(name: str) -> str:
    val = os.environ.get(name, "")
    if val:
        return val
    env_path = Path("/home/clawdbot/.openclaw/.env")
    if env_path.exists():
        for line in env_path.read_text(encoding="utf-8", errors="ignore").splitlines():
            line = line.strip()
            if line.startswith(name + "="):
                return line.split("=", 1)[1].strip().strip("'\"")
    return ""


def _clean(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and math.isnan(v):
        return ""
    return str(v).strip()


def _code(v: Any, width: int | None = None) -> str:
    s = _clean(v)
    if not s:
        return ""
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    if width and s.isdigit():
        s = s.zfill(width)
    return s


def _money(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return round(float(v), 2)
    s = str(v)
    s = s.replace(",", "").replace("SAR", "").replace("ر.س", "")
    s = re.sub(r"[^\d.\-]", "", s)
    if not s:
        return None
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def _date_str(v: Any) -> str:
    if not v:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    try:
        dt = pd.to_datetime(v, errors="coerce")
        if pd.isna(dt):
            return _clean(v)
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return _clean(v)


def _oracle_date_str(v: Any) -> str:
    """Render an invoice date in Oracle's required month/day/year form."""
    if not v:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%m/%d/%Y")
    try:
        dt = pd.to_datetime(v, errors="coerce")
        if pd.isna(dt):
            return _clean(v)
        return dt.strftime("%m/%d/%Y")
    except Exception:
        return _clean(v)


def _sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _file_fingerprint(path: Path | str | None) -> dict[str, Any] | None:
    if not path:
        return None
    p = Path(path).expanduser()
    if not p.exists() or not p.is_file():
        return None
    stat = p.stat()
    return {
        "name": p.name,
        "path": str(p.resolve()),
        "sha256": _sha256_file(p),
        "size": stat.st_size,
        "mtime": datetime.fromtimestamp(stat.st_mtime, timezone.utc).isoformat(),
    }


def _pdf_manifest_fingerprint(pdf_paths: list[Path]) -> dict[str, Any]:
    resolved = sorted(Path(p).expanduser().resolve() for p in pdf_paths)
    manifest_lines = []
    for path in resolved:
        stat = path.stat()
        manifest_lines.append(f"{path.name}\t{stat.st_size}")
    manifest = "\n".join(manifest_lines) + ("\n" if manifest_lines else "")
    return {
        "count": len(resolved),
        "sha256": hashlib.sha256(manifest.encode("utf-8")).hexdigest(),
        "manifest": manifest_lines,
    }


def input_fingerprints(
    expenses_format: Path | str | None,
    so_detail: Path | str | None,
    pdf_paths: list[Path],
) -> dict[str, Any]:
    return {
        "expenses_format_xlsx": _file_fingerprint(expenses_format),
        "so_detail_xlsx": _file_fingerprint(so_detail),
        "lookups_master_xlsx": _file_fingerprint(MASTER_XLSX),
        "pdf_manifest": _pdf_manifest_fingerprint(pdf_paths),
    }


def _git_sha() -> str:
    try:
        out = subprocess.check_output(
            ["git", "rev-parse", "--short", "HEAD"],
            cwd=ROOT,
            stderr=subprocess.DEVNULL,
            text=True,
            timeout=5,
        ).strip()
        return out or "nogit"
    except Exception:
        return "nogit"


def provenance(args: argparse.Namespace, fingerprints: dict[str, Any]) -> dict[str, Any]:
    return {
        "pipeline_version": ASATEEL_PIPELINE_VERSION,
        "git_sha": _git_sha(),
        "run_timestamp_utc": datetime.now(timezone.utc).isoformat(),
        "python_version": platform.python_version(),
        "cli_args": {
            "folder": getattr(args, "folder", None),
            "full": bool(getattr(args, "full", False)),
            "pdf_dir": str(Path(getattr(args, "pdf_dir", "")).expanduser().resolve()) if getattr(args, "pdf_dir", "") else "",
            "expenses_format": str(Path(getattr(args, "expenses_format", "")).expanduser().resolve()) if getattr(args, "expenses_format", "") else "",
            "so_detail": str(Path(getattr(args, "so_detail", "")).expanduser().resolve()) if getattr(args, "so_detail", "") else "",
            "allocation_mode": getattr(args, "allocation_mode", STANDARD_ALLOCATION_MODE),
            "project_allocation_lookup": (
                str(Path(getattr(args, "project_allocation_lookup", "")).expanduser().resolve())
                if getattr(args, "allocation_mode", STANDARD_ALLOCATION_MODE) == PROJECT_ALLOCATION_MODE
                and getattr(args, "project_allocation_lookup", "")
                else ""
            ),
        },
        "input_fingerprints_key": "input_fingerprints",
        "input_fingerprints": fingerprints,
    }


@dataclass
class Lookups:
    agencies: list[dict[str, str]]
    manpower: list[dict[str, str]]
    manpower_by_emp_no: dict[str, dict[str, str]]
    agency_to_cluster: dict[str, dict[str, str]]
    agency_name_by_code: dict[str, str]
    cc_name_by_code: dict[str, str]
    cc_code_by_name: dict[str, str]
    cc_div_by_code: dict[str, str]
    div_name_by_code: dict[str, str]
    div_code_by_name: dict[str, str]
    solution_code_by_description: dict[str, str]
    solution_name_by_code: dict[str, str]
    employee_names: list[str]
    account_name_by_code: dict[str, str]


class SoDetailIndex(dict[str, dict[str, Any]]):
    """A loaded SO_Detail presence index, including the valid empty-export case."""

    loaded = True


BRAND_ALIAS_REMAP = {
    "3m": "Solvento",
    "biofire": "BMX",
    "biomerieux": "BMX",
}

SUPPLIER_AGENCY_ALIAS_REMAP = {
    "3m": "Solventum",
    "solvento": "Solventum",
    "biofire": "BMX",
    "biomerieux": "BMX",
    "getingegroup": "Getinge",
}

AGENCY_CORPORATE_WORDS = {
    "group",
    "co",
    "company",
    "inc",
    "ltd",
    "llc",
    "medical",
    "systems",
    "scientific",
}


def _norm_key(v: Any) -> str:
    s = unicodedata.normalize("NFKD", _clean(v)).casefold()
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return re.sub(r"[^0-9a-z]+", "", s)


def _norm_text(v: Any) -> str:
    return re.sub(r"\s+", " ", _clean(v).casefold()).strip()


def _norm_agency_words(v: Any, strip_corporate: bool = False) -> str:
    text = re.sub(r"[^0-9a-z]+", " ", unicodedata.normalize("NFKD", _clean(v)).casefold())
    words = [w for w in text.split() if w]
    if strip_corporate:
        words = [w for w in words if w not in AGENCY_CORPORATE_WORDS]
    return " ".join(words)


def _project_override_supplier_match(
    supplier_match: dict[str, Any],
    project_lookup: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    """Apply the exact, pre-resolved project mapping to a supplier line."""
    if str(ROOT) not in sys.path:
        sys.path.insert(0, str(ROOT))
    from scripts.asateel_project_allocation import resolve_override

    updated = dict(supplier_match)
    audit = resolve_override(
        project_lookup,
        agency_code=supplier_match.get("agency_code"),
        agency_name=supplier_match.get("agency") or supplier_match.get("agency_name"),
        employee_no=supplier_match.get("employee_number"),
        employee_name=supplier_match.get("employee_name"),
    )
    agency_after = audit.get("agency_after") or {}
    employee_after = audit.get("employee_after") or {}
    if agency_after:
        updated["agency_code"] = agency_after.get("code", "")
        updated["agency_name"] = agency_after.get("name", "")
        updated["agency_resolve_method"] = f"{PROJECT_ALLOCATION_MODE}:{audit.get('agency_match_method') or 'unresolved'}"
        updated["agency_resolve_score"] = 1.0
    if employee_after:
        updated["employee_number"] = employee_after.get("employee_no", "")
        updated["employee_name"] = employee_after.get("employee_name", "")
    updated["_project_allocation_audit"] = audit
    return updated, audit


def load_lookups() -> Lookups:
    wb = openpyxl.load_workbook(MASTER_XLSX, read_only=True, data_only=True)
    account_name_by_code = {}
    for row in wb["Account"].iter_rows(min_row=2, values_only=True):
        code = _code(row[0], 8)
        name = _clean(row[1] if len(row) > 1 else "")
        if code and name:
            account_name_by_code[code] = name

    agencies = []
    agency_name_by_code = {}
    for row in wb["Agency"].iter_rows(min_row=2, values_only=True):
        code = _code(row[0], 5)
        name = _clean(row[1])
        if code and name:
            agencies.append({"code": code, "name": name})
            agency_name_by_code[code] = name

    cc_name_by_code = {}
    cc_code_by_name = {}
    cc_div_by_code = {}
    div_name_by_code = {}
    div_code_by_name = {}
    for row in wb["DIV"].iter_rows(min_row=2, values_only=True):
        div_code_by_name[_clean(row[1]).casefold()] = _code(row[0])

    for row in wb["Manpower"].iter_rows(min_row=2, values_only=True):
        div_code = _code(row[8])
        div_name = _clean(row[9])
        cc = _code(row[12])
        cc_name = _clean(row[13])
        if div_code and div_name:
            div_name_by_code.setdefault(div_code, div_name)
        if cc and cc_name:
            cc_name_by_code[cc] = cc_name
            cc_code_by_name.setdefault(cc_name.casefold(), cc)
        if cc and div_code:
            cc_div_by_code.setdefault(cc, div_code)

    solution_code_by_description = {}
    solution_name_by_code = {}
    for row in wb["Solution"].iter_rows(min_row=2, values_only=True):
        code = _code(row[0], 5)
        name = _clean(row[1] if len(row) > 1 else "")
        if code and name:
            solution_name_by_code[code] = name
            solution_code_by_description[_norm_text(name)] = code

    manpower = []
    manpower_by_emp_no = {}
    agency_to_cluster = {}
    employee_names = []
    for row in wb["Manpower"].iter_rows(min_row=2, values_only=True):
        emp = {
            "emp_no": _code(row[0]),
            "name": _clean(row[2]),
            "arabic_name": _clean(row[3]),
            "division_code": _code(row[8]),
            "division": _clean(row[9]),
            "agency_code": _code(row[10], 5),
            "agency_name": _clean(row[11]),
            "cost_center": _code(row[12]),
            "cost_center_name": _clean(row[13]),
            "allocation_status": _clean(row[14]),
            "solution": _clean(row[15]),
        }
        if emp["name"]:
            employee_names.append(emp["name"])
        if emp["arabic_name"]:
            employee_names.append(emp["arabic_name"])
        if emp["emp_no"] or emp["name"]:
            manpower.append(emp)
        if emp["emp_no"]:
            manpower_by_emp_no[emp["emp_no"]] = emp
        key = emp["agency_name"].casefold()
        if key and key not in agency_to_cluster and emp["division"] and emp["cost_center"]:
            agency_to_cluster[key] = emp
    wb.close()
    return Lookups(
        agencies,
        manpower,
        manpower_by_emp_no,
        agency_to_cluster,
        agency_name_by_code,
        cc_name_by_code,
        cc_code_by_name,
        cc_div_by_code,
        div_name_by_code,
        div_code_by_name,
        solution_code_by_description,
        solution_name_by_code,
        employee_names,
        account_name_by_code,
    )


def render_pdf(pdf_path: Path) -> list[Path]:
    if not shutil.which("pdftoppm"):
        raise RuntimeError("pdftoppm not found")
    target = RENDER_DIR / pdf_path.parent.name / pdf_path.stem
    target.mkdir(parents=True, exist_ok=True)
    existing = sorted(target.glob("page-*.png"))
    if existing:
        return existing
    prefix = target / "page"
    subprocess.run(
        ["pdftoppm", "-r", "150", "-png", str(pdf_path), str(prefix)],
        check=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )
    return sorted(target.glob("page-*.png"))


def build_prompt(folder: str, invoice_hint: str, lookups: Lookups) -> str:
    agency_lines = "\n".join(f"{a['code']}\t{a['name']}" for a in lookups.agencies)
    return f"""
You are extracting and allocating one full multi-page Asateel transport invoice PDF for Aljeel.

Invoice folder={folder}; expected invoice number hint={invoice_hint}.

Rules:
- Page 1 may be blank to OCR but is visible in these images. Read all pages.
- Extract invoice header: invoice_number, invoice_date, vendor_vat, buyer_vat, subtotal, vat, total.
- Extract EVERY invoice line from the ZATCA tax invoice, including 3+ line invoices. Do not collapse lines.
  For each line capture description/truck type, reference (المرجع), dispatch_ref (كشف التخريج),
  supply_order (أمر التوريد), unit_price, qty, line_subtotal.
- The sum of line_subtotal values must reconcile to invoice subtotal where the document provides line amounts.
- Search the whole PDF for allocation signals:
  1. BRAND/principal on an Al Jeel goods-receipt note, especially handwritten PROJECTS/ADMIN pages.
  2. SALES PERSON name on embedded Al Jeel sales/tax invoices, especially CENTRAL pages.
- For each extracted signal, match brand/principal ONLY to one row from the agency list below, or no_match.
  Do not invent an agency. Return the selected code and canonical name exactly as listed.
- Preserve original brand text in raw. If the document shows 3M, Biofire, BioMerieux, bioMérieux,
  or bio merieux, return that original raw text; downstream code will normalize the alias.
- For CENTRAL multi-line invoices, if separate embedded invoices/salesperson evidence maps to separate lines,
  put line-level allocation_signal on each line. Every extracted line must be represented even if the
  allocation signal is none/no_match.

Return JSON only:
{{
  "invoice_number": "03041",
  "invoice_date": "YYYY-MM-DD or raw",
  "vendor_vat": "",
  "buyer_vat": "",
  "subtotal": 0,
  "vat": 0,
  "total": 0,
  "signals": [
    {{"source": "brand|salesperson", "raw": "", "matched_agency_code": "", "matched_agency_name": "", "confidence": 0.0, "evidence": ""}}
  ],
  "lines": [
    {{"line_no": 1, "description": "", "reference": "", "dispatch_ref": "", "supply_order": "", "unit_price": 0, "qty": 1, "line_subtotal": 0,
      "allocation_signal": {{"source": "brand|salesperson|none", "raw": "", "matched_agency_code": "", "matched_agency_name": "", "confidence": 0.0, "evidence": ""}}
    }}
  ],
  "extraction_notes": []
}}

Agency list, code then canonical name:
{agency_lines}
""".strip()


def _extract_gemini_text(data: dict[str, Any]) -> str:
    cand = (data.get("candidates") or [{}])[0]
    content = cand.get("content") or {}
    parts = content.get("parts") or [{}]
    text = parts[0].get("text", "")
    text = re.sub(r"^```(?:json)?\s*", "", text.strip())
    text = re.sub(r"\s*```$", "", text)
    text = text.strip()
    if not text.startswith("{"):
        match = re.search(r"\{.*\}", text, re.DOTALL)
        if match:
            text = match.group()
    return text


def _is_retryable_gemini_error(exc: Exception) -> bool:
    if isinstance(exc, urllib.error.HTTPError):
        return exc.code == 429 or 500 <= exc.code <= 599
    if isinstance(exc, TimeoutError):
        return True
    msg = str(exc).lower()
    return any(token in msg for token in ["timed out", "timeout", "429", "quota", "rate", "500", "502", "503", "504"])


def call_gemini_vision(parts: list[dict[str, Any]], model: str) -> tuple[str, dict[str, Any]]:
    global _PRINTED_GEMINI_HOST
    url = f"{GEMINI_BASE_URL}/models/{model}:generateContent?key={_load_env_key('GEMINI_API_KEY')}"
    host = urllib.parse.urlparse(url).hostname or ""
    if not _PRINTED_GEMINI_HOST:
        print(f"[gemini route] host={host} base_url={GEMINI_BASE_URL}", flush=True)
        _PRINTED_GEMINI_HOST = True
    payload = {
        "contents": [{"parts": parts}],
        "generationConfig": {
            "temperature": 0.0,
            "topP": 1.0,
            "maxOutputTokens": 32768,
            "responseMimeType": "application/json",
        },
    }
    req = urllib.request.Request(
        url,
        data=json.dumps(payload).encode(),
        headers={"Content-Type": "application/json", "User-Agent": "AlJeel-Asateel-POC/1.0"},
    )
    with urllib.request.urlopen(req, timeout=180) as resp:
        data = json.loads(resp.read())
    return _extract_gemini_text(data), data


def _strip_json_wrapper(text: str) -> str:
    s = text.lstrip("\ufeff").strip()
    fence = re.search(r"```(?:json)?\s*(.*?)\s*```", s, flags=re.I | re.S)
    if fence:
        s = fence.group(1).strip()
    first_positions = [pos for pos in (s.find("{"), s.find("[")) if pos >= 0]
    if first_positions:
        s = s[min(first_positions):]
    return s.strip()


def _first_balanced_json_value(text: str) -> str:
    s = _strip_json_wrapper(text)
    starts = [(pos, ch) for ch, pos in (("{", s.find("{")), ("[", s.find("["))) if pos >= 0]
    if not starts:
        return s
    start, opener = min(starts)
    closer_for = {"{": "}", "[": "]"}
    stack: list[str] = []
    in_string = False
    escape = False
    for idx in range(start, len(s)):
        ch = s[idx]
        if in_string:
            if escape:
                escape = False
            elif ch == "\\":
                escape = True
            elif ch == "\"":
                in_string = False
            continue
        if ch == "\"":
            in_string = True
        elif ch in "{[":
            stack.append(closer_for[ch])
        elif ch in "}]":
            if not stack or ch != stack[-1]:
                continue
            stack.pop()
            if not stack:
                return s[start:idx + 1].strip()
    return s[start:].strip()


def _trim_stray_trailing_closers(text: str) -> str:
    s = text.strip()
    while s:
        stack: list[str] = []
        in_string = False
        escape = False
        bad_at: int | None = None
        for idx, ch in enumerate(s):
            if in_string:
                if escape:
                    escape = False
                elif ch == "\\":
                    escape = True
                elif ch == "\"":
                    in_string = False
                continue
            if ch == "\"":
                in_string = True
            elif ch in "{[":
                stack.append("}" if ch == "{" else "]")
            elif ch in "}]":
                if not stack or ch != stack[-1]:
                    bad_at = idx
                    break
                stack.pop()
        if bad_at is not None and not s[bad_at:].strip(" \t\r\n]}"):
            s = s[:bad_at].rstrip()
            continue
        return s
    return s


def _repair_closers_to_first_json_value(text: str) -> str:
    s = _strip_json_wrapper(text)
    starts = [(pos, ch) for ch, pos in (("{", s.find("{")), ("[", s.find("["))) if pos >= 0]
    if not starts:
        return s
    start, _opener = min(starts)
    closer_for = {"{": "}", "[": "]"}
    stack: list[str] = []
    out: list[str] = []
    in_string = False
    escape = False
    for idx in range(start, len(s)):
        ch = s[idx]
        if in_string:
            out.append(ch)
            if escape:
                escape = False
            elif ch == "\\":
                escape = True
            elif ch == "\"":
                in_string = False
            continue
        if ch == "\"":
            in_string = True
            out.append(ch)
        elif ch in "{[":
            stack.append(closer_for[ch])
            out.append(ch)
        elif ch in "}]":
            if not stack:
                if not s[idx:].strip(" \t\r\n]}"):
                    break
                continue
            expected = stack.pop()
            out.append(expected if ch != expected else ch)
            if not stack:
                break
        else:
            out.append(ch)
    while stack:
        out.append(stack.pop())
    return "".join(out).strip()


def _mark_json_salvaged(value: Any) -> Any:
    if isinstance(value, dict):
        notes = value.get("extraction_notes")
        if not isinstance(notes, list):
            notes = [] if notes in (None, "") else [str(notes)]
        if "json_salvaged=true" not in notes:
            notes.append("json_salvaged=true")
        value["extraction_notes"] = notes
    return value


def _salvage_json(text: str) -> Any:
    s = _strip_json_wrapper(text)
    first_err: Exception
    try:
        return _mark_json_salvaged(json.loads(s))
    except json.JSONDecodeError as exc:
        first_err = exc

    candidates: list[str] = []
    if isinstance(first_err, json.JSONDecodeError) and first_err.msg == "Extra data":
        candidates.append(s[:first_err.pos].strip())
    candidates.append(_first_balanced_json_value(s))
    candidates.append(_repair_closers_to_first_json_value(s))

    first_brace = s.find("{")
    last_brace = s.rfind("}")
    if first_brace >= 0 and last_brace > first_brace:
        greedy = s[first_brace:last_brace + 1].strip()
        candidates.append(greedy)
        candidates.append(_repair_closers_to_first_json_value(greedy))
        candidates.append(re.sub(r",\s*([}\]])", r"\1", greedy))
        trimmed = _trim_stray_trailing_closers(greedy)
        candidates.append(trimmed)
        candidates.append(re.sub(r",\s*([}\]])", r"\1", trimmed))

    seen: set[str] = set()
    last_err: Exception = first_err
    for cand in candidates:
        cand = cand.strip()
        if not cand or cand in seen:
            continue
        seen.add(cand)
        try:
            return _mark_json_salvaged(json.loads(cand))
        except Exception as exc:
            last_err = exc
    raise last_err


def gemini_extract(
    pdf_path: Path,
    folder: str,
    invoice_hint: str,
    lookups: Lookups,
    force: bool = False,
    cache_tag: str | None = None,
) -> dict[str, Any]:
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    if cache_tag:
        cache_path = CACHE_DIR / f"{folder}__{cache_tag}__{invoice_hint}.json"
    else:
        cache_path = CACHE_DIR / f"{folder}__{invoice_hint}.json"
    if cache_path.exists() and not force:
        # The per-batch cache is the authoritative extraction record for repeatable runs.
        return json.loads(cache_path.read_text(encoding="utf-8"))

    gemini_key = _load_env_key("GEMINI_API_KEY")
    if not gemini_key:
        raise RuntimeError("GEMINI_API_KEY not found")
    os.environ["GEMINI_API_KEY"] = gemini_key
    os.environ.pop("GOOGLE_API_KEY", None)
    pages = render_pdf(pdf_path)
    contents: list[dict[str, Any]] = []
    for img in pages:
        contents.append({
            "inline_data": {
                "mime_type": "image/png",
                "data": base64.b64encode(img.read_bytes()).decode("ascii"),
            }
        })
    contents.append({"text": build_prompt(folder, invoice_hint, lookups)})

    t0 = time.time()
    retry_delays = [2, 5, 10]
    raw_text = ""
    raw_response: dict[str, Any] = {}
    used_model = ""
    last_exc: Exception | None = None
    for model in GEMINI_MODEL_CASCADE:
        for attempt in range(1, len(retry_delays) + 2):
            try:
                raw_text, raw_response = call_gemini_vision(contents, model)
                used_model = model
                last_exc = None
                break
            except Exception as exc:
                last_exc = exc
                if not _is_retryable_gemini_error(exc) or attempt > len(retry_delays):
                    print(f"[model fallback] {folder} {invoice_hint}: {model} failed: {exc}", flush=True)
                    break
                delay = retry_delays[attempt - 1]
                print(f"[retry {attempt}/3] {folder} {invoice_hint} {model}: {exc}; sleeping {delay}s", flush=True)
                time.sleep(delay)
        if used_model:
            break
    if not used_model:
        raise RuntimeError(f"Gemini extraction failed via Cloudflare gateway: {last_exc}")
    payload = {
        "model": used_model,
        "model_cascade": GEMINI_MODEL_CASCADE,
        "gateway_base_url": GEMINI_BASE_URL,
        "gateway_host": urllib.parse.urlparse(GEMINI_BASE_URL).hostname or "",
        "pdf": str(pdf_path),
        "rendered_pages": [str(p) for p in pages],
        "elapsed_seconds": round(time.time() - t0, 2),
        "raw_text": raw_text,
    }
    if raw_response.get("usageMetadata"):
        payload["usage_metadata"] = raw_response["usageMetadata"]
    try:
        payload["extraction"] = json.loads(raw_text)
    except Exception:
        try:
            payload["extraction"] = _salvage_json(raw_text)
        except Exception:
            payload["extraction"] = {"parse_error": True, "raw": raw_text}
    cache_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return payload


def _find_employee(raw: Any, lookups: Lookups) -> dict[str, str] | None:
    name = _clean(raw)
    if not name:
        return None
    match = process.extractOne(name, lookups.employee_names, scorer=fuzz.WRatio)
    if not match or match[1] < 85:
        return None
    matched_name = match[0]
    for emp in lookups.manpower:
        if emp["name"] == matched_name or emp["arabic_name"] == matched_name:
            return emp
    return None


def _match_agency_reference(raw: Any, agencies: list[dict[str, str]], method_prefix: str) -> dict[str, Any] | None:
    """Resolve supplier text against one explicit agency value/description table."""
    name = _clean(raw)
    if not name:
        return None
    raw_code = _code(name, 5)
    norm_text = _norm_agency_words(name)
    norm_stripped = _norm_agency_words(name, strip_corporate=True)
    for agency in agencies:
        if raw_code.isdigit() and raw_code == _code(agency.get("code"), 5):
            return {
                "code": _code(agency.get("code"), 5),
                "name": _clean(agency.get("name")),
                "score": 1.0,
                "method": f"{method_prefix}_value",
                "raw": name,
            }
    by_norm: dict[str, dict[str, str]] = {}
    by_stripped: dict[str, dict[str, str]] = {}
    for agency in agencies:
        by_norm.setdefault(_norm_agency_words(agency.get("name")), agency)
        by_stripped.setdefault(_norm_agency_words(agency.get("name"), strip_corporate=True), agency)
    if norm_text in by_norm:
        agency = by_norm[norm_text]
        return {
            "code": _code(agency.get("code"), 5), "name": _clean(agency.get("name")),
            "score": 1.0, "method": f"{method_prefix}_exact", "raw": name,
        }
    if norm_stripped and norm_stripped in by_stripped:
        agency = by_stripped[norm_stripped]
        return {
            "code": _code(agency.get("code"), 5), "name": _clean(agency.get("name")),
            "score": 0.99, "method": f"{method_prefix}_exact", "raw": name,
        }
    for agency in agencies:
        agency_norm = _norm_agency_words(agency.get("name"), strip_corporate=True)
        if norm_stripped and agency_norm and (
            norm_stripped.startswith(agency_norm) or agency_norm.startswith(norm_stripped)
        ):
            return {
                "code": _code(agency.get("code"), 5), "name": _clean(agency.get("name")),
                "score": 0.95, "method": f"{method_prefix}_substring", "raw": name,
            }
    return None


def _resolve_supplier_agency_text(
    raw: Any,
    lookups: Lookups,
    workbook_agencies: list[dict[str, str]] | None = None,
) -> dict[str, Any]:
    name = _clean(raw)
    empty = {"code": "", "name": name, "score": 0.0, "method": "unresolved", "raw": name}
    if not name:
        return empty

    # The Expenses Format workbook is the authority for its own validation
    # values. Fall back to the shared master only if that local table cannot
    # resolve the supplier sub-line text.
    workbook_match = _match_agency_reference(name, workbook_agencies or [], "workbook_agency_reference")
    if workbook_match:
        shared_name = lookups.agency_name_by_code.get(workbook_match["code"])
        if shared_name:
            workbook_match["name"] = shared_name
        return workbook_match

    norm_text = _norm_agency_words(name)
    norm_stripped = _norm_agency_words(name, strip_corporate=True)
    by_norm: dict[str, dict[str, str]] = {}
    by_stripped: dict[str, dict[str, str]] = {}
    for agency in lookups.agencies:
        agency_norm = _norm_agency_words(agency["name"])
        agency_stripped = _norm_agency_words(agency["name"], strip_corporate=True)
        by_norm.setdefault(agency_norm, agency)
        by_stripped.setdefault(agency_stripped, agency)

    alias_target = SUPPLIER_AGENCY_ALIAS_REMAP.get(_norm_key(name))
    if alias_target:
        alias_resolved = _resolve_supplier_agency_text(alias_target, lookups)
        if alias_resolved.get("code"):
            alias_resolved["method"] = "alias"
            alias_resolved["raw"] = name
            return alias_resolved

    if norm_text in by_norm:
        agency = by_norm[norm_text]
        return {"code": agency["code"], "name": agency["name"], "score": 1.0, "method": "exact", "raw": name}
    if norm_stripped and norm_stripped in by_stripped:
        agency = by_stripped[norm_stripped]
        return {"code": agency["code"], "name": agency["name"], "score": 0.99, "method": "exact", "raw": name}

    for agency in lookups.agencies:
        agency_norm = _norm_agency_words(agency["name"], strip_corporate=True)
        if norm_stripped and agency_norm and (norm_stripped.startswith(agency_norm) or agency_norm.startswith(norm_stripped)):
            return {"code": agency["code"], "name": agency["name"], "score": 0.95, "method": "substring", "raw": name}
        if norm_stripped and agency_norm and (norm_stripped in agency_norm or agency_norm in norm_stripped):
            return {"code": agency["code"], "name": agency["name"], "score": 0.93, "method": "substring", "raw": name}

    raw_tokens = set(norm_stripped.split())
    best_agency: dict[str, str] | None = None
    best_score = 0.0
    for agency in lookups.agencies:
        agency_tokens = set(_norm_agency_words(agency["name"], strip_corporate=True).split())
        if not raw_tokens or not agency_tokens:
            continue
        score = len(raw_tokens & agency_tokens) / max(len(raw_tokens), len(agency_tokens))
        if score > best_score:
            best_score = score
            best_agency = agency
    if best_agency and best_score >= 0.9:
        return {
            "code": best_agency["code"],
            "name": best_agency["name"],
            "score": round(best_score, 3),
            "method": "token",
            "raw": name,
        }

    return empty


def _resolve_agency_name(raw: Any, lookups: Lookups) -> tuple[str, str, float]:
    supplier_resolved = _resolve_supplier_agency_text(raw, lookups)
    if supplier_resolved.get("code"):
        return supplier_resolved["code"], supplier_resolved["name"], float(supplier_resolved.get("score") or 0.0)
    name = _clean(raw)
    if not name:
        return "", "", 0.0
    candidates = [a["name"] for a in lookups.agencies]
    match = process.extractOne(name, candidates, scorer=fuzz.WRatio)
    if match and match[1] >= 85:
        matched = match[0]
        return next(a["code"] for a in lookups.agencies if a["name"] == matched), matched, round(match[1] / 100, 3)
    return "", name, 0.0


def _resolved_from_manpower(emp: dict[str, str], source: str, raw: Any, match_method: str) -> dict[str, Any]:
    allocation_status = _clean(emp.get("allocation_status"))
    return {
        "source": source,
        "raw": _clean(raw),
        "agency_code": emp.get("agency_code", ""),
        "agency_name": emp.get("agency_name", ""),
        "division_code": emp.get("division_code", ""),
        "division": emp.get("division", ""),
        "cost_center": emp.get("cost_center", ""),
        "cost_center_name": emp.get("cost_center_name", ""),
        "confidence": 1.0 if source == "manpower_empno" else 0.9,
        "status_reason": allocation_status if allocation_status and allocation_status != "Can Be used" else "",
        "match_method": match_method,
        "agency_resolve_method": "manpower_fallback",
        "brand_remap_from": "",
        "brand_remap_to": "",
        "allocation_status": allocation_status,
        "emp_no": emp.get("emp_no", ""),
    }


def _find_employee_by_no(raw: Any, lookups: Lookups) -> dict[str, str] | None:
    emp_no = _code(raw)
    if not emp_no:
        return None
    return lookups.manpower_by_emp_no.get(emp_no)


def _solution_from_text(raw: Any, lookups: Lookups) -> tuple[str, str, str]:
    text = _clean(raw)
    if not text:
        return "00000", "General", "Supplier Expenses Format Solution text blank"
    code = lookups.solution_code_by_description.get(_norm_text(text))
    if not code:
        return "00000", "General", f"Supplier Solution '{text}' not matched to Solution lookup"
    return code, lookups.solution_name_by_code.get(code, text), ""


def _brand_alias_target(signal: dict[str, Any]) -> tuple[str, str]:
    fields = [
        signal.get("raw"),
        signal.get("matched_agency_name"),
    ]
    for field in fields:
        norm = _norm_key(field)
        if norm in BRAND_ALIAS_REMAP:
            return _clean(field), BRAND_ALIAS_REMAP[norm]
    return "", ""


def resolve_signal(signal: dict[str, Any], lookups: Lookups) -> dict[str, Any]:
    source = _clean(signal.get("source")).lower() or "none"
    raw = _clean(signal.get("raw"))
    out = {
        "source": source if source in {"brand", "salesperson"} else "none",
        "raw": raw,
        "agency_code": "",
        "agency_name": "",
        "division_code": "",
        "division": "",
        "cost_center": "",
        "cost_center_name": "",
        "confidence": float(signal.get("confidence") or 0),
        "status_reason": "",
        "match_method": "",
        "agency_resolve_method": "manpower_fallback",
        "brand_remap_from": "",
        "brand_remap_to": "",
    }
    if out["source"] == "salesperson" and raw:
        emp_match = _find_employee(raw, lookups)
        if emp_match:
            resolved = _resolved_from_manpower(
                emp_match,
                "salesperson",
                raw,
                f"salesperson_fuzzy:{emp_match['name'] or emp_match['arabic_name']}",
            )
            resolved["confidence"] = max(out["confidence"], 0.85)
            return resolved
        out["status_reason"] = "salesperson not matched to Manpower"
        return out

    code = _code(signal.get("matched_agency_code"), 5)
    name = _clean(signal.get("matched_agency_name"))
    remap_from, remap_to = _brand_alias_target(signal)
    if remap_to:
        code = ""
        name = remap_to
        out["brand_remap_from"] = remap_from
        out["brand_remap_to"] = remap_to
    if code and code in lookups.agency_name_by_code:
        name = lookups.agency_name_by_code[code]
    elif name:
        resolved_code, resolved_name, score = _resolve_agency_name(name, lookups)
        if resolved_code and score >= 0.85:
            name = resolved_name
            code = resolved_code
            out["confidence"] = max(out["confidence"], score)

    if not code or not name:
        out["status_reason"] = "no agency match"
        return out
    out["agency_code"] = code
    out["agency_name"] = name
    out["match_method"] = "brand_alias_remap" if remap_to else "brand_llm_agency_list"
    out["agency_resolve_method"] = "manpower_fallback"
    cluster = lookups.agency_to_cluster.get(name.casefold())
    if not cluster:
        out["status_reason"] = "agency has no Manpower New agency cluster"
        return out
    out.update({
        "division": cluster["division"],
        "division_code": cluster["division_code"],
        "cost_center": cluster["cost_center"],
        "cost_center_name": cluster["cost_center_name"],
    })
    cc_div_code = lookups.cc_div_by_code.get(out["cost_center"])
    if cc_div_code:
        out["division_code"] = cc_div_code
        out["division"] = lookups.div_name_by_code.get(cc_div_code, out["division"])
    return out


def normalize_extraction(value: Any) -> dict[str, Any]:
    if isinstance(value, list):
        if len(value) == 1 and isinstance(value[0], dict):
            return value[0]
        return {"invoice_number": "", "lines": [], "extraction_notes": [f"Unexpected list extraction with {len(value)} items"]}
    if isinstance(value, dict) and value.get("parse_error") and value.get("raw"):
        try:
            return normalize_extraction(_salvage_json(str(value.get("raw") or "")))
        except Exception:
            pass
        return {"invoice_number": "", "lines": [], "extraction_notes": ["Gemini JSON parse error"]}
    if isinstance(value, dict):
        return value
    return {"invoice_number": "", "lines": [], "extraction_notes": ["Gemini returned non-object extraction"]}


def classify(
    extraction: dict[str, Any],
    resolved: dict[str, Any],
    notes: list[str],
    *,
    scan_available: bool = True,
) -> str:
    inv = _clean(extraction.get("invoice_number"))
    subtotal = _money(extraction.get("subtotal"))
    vat = _money(extraction.get("vat"))
    total = _money(extraction.get("total"))
    allocation_status = _clean(resolved.get("allocation_status"))
    if allocation_status and allocation_status != "Can Be used":
        notes.append(f"RED: Manpower allocation status: {allocation_status}")
        return "RED"
    if scan_available and (not inv or total is None):
        notes.append("RED: invoice number or total not extractable")
        return "RED"
    if scan_available and subtotal is not None and vat is not None and abs((subtotal + vat) - total) > 1.0:
        notes.append(f"RED: totals do not foot ({subtotal}+{vat}!={total})")
        return "RED"
    if not resolved.get("division") or not resolved.get("cost_center") or not resolved.get("agency_name"):
        notes.append(resolved.get("status_reason") or "YELLOW: unresolved allocation")
        return "YELLOW"
    if float(resolved.get("confidence") or 0) < 0.9:
        notes.append("YELLOW: confidence below 0.90")
        return "YELLOW"
    return "GREEN"


def account_description(lookups: Lookups) -> tuple[str, str]:
    desc = lookups.account_name_by_code.get(_code(GL_ACCOUNT, 8))
    if desc:
        return desc, ""
    return GL_FALLBACK_DESC, f"{GL_ACCOUNT} is not in Aljeel_Lookups-v2 Account sheet"


def _build_gl_description(row_fields: dict[str, Any], include_tail: bool = True) -> str:
    """Mirror Jawal's canonical GL Description format for Asateel row fields."""

    def _part(v: Any) -> str:
        s = _clean(v)
        return s if s and s != "#N/A" else "—"

    parts = [
        _part(row_fields.get("GL")),
        _part(row_fields.get("Cost Name")),
        _part(row_fields.get("Contribution")),
        _part(row_fields.get("Solution Name")),
        _part(row_fields.get("Agency Name")),
    ]
    if include_tail:
        parts.extend(["00000", "00", "000000"])
    return " · ".join(parts)


def _segment(v: Any, width: int, default: str = "") -> str:
    s = _code(v, width)
    if not s and default:
        s = _code(default, width)
    return s


def build_distribution_combination(row: dict[str, Any]) -> str:
    parts = [
        _segment(row.get("Company"), SEGMENT_WIDTHS["company"], COMPANY),
        _segment(DEFAULT_LOCATION, SEGMENT_WIDTHS["location"], DEFAULT_LOCATION),
        _segment(row.get("Account"), SEGMENT_WIDTHS["account"], GL_ACCOUNT),
        _segment(row.get("Cost Center"), SEGMENT_WIDTHS["cost_center"]),
        _segment(row.get("DIV"), SEGMENT_WIDTHS["div"]),
        _segment(row.get("Solution"), SEGMENT_WIDTHS["solution"], "00000"),
        _segment(row.get("Agency"), SEGMENT_WIDTHS["agency"], "00000"),
        _segment(row.get("Project"), SEGMENT_WIDTHS["project"], "00000"),
        _segment(row.get("Intercompany"), SEGMENT_WIDTHS["intercompany"], "00"),
        _segment(row.get("Future 1"), SEGMENT_WIDTHS["future"], "000000"),
    ]
    return "-".join(parts)


def _valid_allocation_signal(sig: dict[str, Any]) -> bool:
    source = _clean(sig.get("source")).lower()
    code = _code(sig.get("matched_agency_code"), 5)
    name = _clean(sig.get("matched_agency_name")).casefold()
    return source in {"brand", "salesperson"} and code not in {"", "00000"} and name not in {"", "general", "no_match"}


def expand_distribution_lines(lines: list[dict[str, Any]], ext: dict[str, Any]) -> tuple[list[dict[str, Any]], bool]:
    signals = []
    seen = set()
    for sig in ext.get("signals") or []:
        if not isinstance(sig, dict) or not _valid_allocation_signal(sig):
            continue
        key = (_clean(sig.get("source")).lower(), _code(sig.get("matched_agency_code"), 5), _clean(sig.get("raw")).casefold())
        if key not in seen:
            seen.add(key)
            signals.append(sig)

    if len(lines) == 1 and len(signals) > 1:
        base = dict(lines[0])
        expanded = []
        for idx, sig in enumerate(signals, start=1):
            ln = dict(base)
            ln["line_no"] = idx
            ln["allocation_signal"] = sig
            ln["_expanded_from_multi_signal"] = True
            expanded.append(ln)
        return expanded, True
    return lines, False


def normalize_exclusive_amount(amount: float | None, subtotal: float | None, total: float | None, notes: list[str]) -> tuple[float | None, str]:
    if amount is None:
        return None, ""
    if subtotal and abs(amount - subtotal) <= 1:
        return amount, "excl"
    if total and abs(amount - total) <= 1:
        notes.append("Line amount appeared VAT-inclusive; divided by 1.15 for Oracle line amount")
        return round(amount / (1 + VAT_RATE), 2), "incl_to_excl"
    return amount, "excl_assumed"


def _canonical_jq(raw: Any, *, allow_bare: bool = True) -> str:
    text = _clean(raw).upper()
    m = re.search(r"\bJQ\s*-\s*(\d+)\b", text)
    if m:
        return f"JQ-{m.group(1).zfill(8)}"
    if allow_bare and re.fullmatch(r"\d+", text):
        return f"JQ-{text.zfill(8)}"
    return ""


def _sperson_id(raw: Any) -> str:
    text = _clean(raw)
    m = re.match(r"^(\d+)", text)
    return m.group(1) if m else text


def _split_jqs(raw: Any, *, allow_bare: bool = True) -> list[str]:
    out = []
    seen = set()
    text = _clean(raw)
    for m in re.finditer(r"\bJQ\s*-\s*(\d+)\b", text, flags=re.I):
        jq = f"JQ-{m.group(1).zfill(8)}"
        if jq not in seen:
            seen.add(jq)
            out.append(jq)
    if allow_bare and not out and re.fullmatch(r"\d+", text):
        out.append(f"JQ-{text.zfill(8)}")
    return out


def _split_supplier_amount(amount: float | None, count: int, index: int) -> tuple[float | None, str]:
    if amount is None or count <= 1:
        return amount, "supplier_row_amount"
    cents = int(round(float(amount) * 100))
    base = cents // count
    remainder = cents - (base * count)
    unit_cents = base + (remainder if index == count else 0)
    basis = "supplier_row_amount_per_jq"
    if index == count and remainder:
        basis += "_cent_remainder_to_last"
    return round(unit_cents / 100, 2), basis


def _supplier_resolve_allocation(
    rec: dict[str, Any],
    lookups: Lookups,
    workbook_agencies: list[dict[str, str]] | None = None,
) -> dict[str, str]:
    agency_resolved = _resolve_supplier_agency_text(rec.get("agency"), lookups, workbook_agencies)
    agency_code = agency_resolved.get("code", "")
    agency_name = agency_resolved.get("name", "")
    cluster = lookups.agency_to_cluster.get((agency_name or _clean(rec.get("agency"))).casefold())
    div_raw = _clean(rec.get("division"))
    cc_raw = _clean(rec.get("cost_center"))
    cc_code = _code(cc_raw)
    if not cc_code or not cc_code.isdigit():
        cc_code = lookups.cc_code_by_name.get(cc_raw.casefold(), "")
    if cluster:
        cc_code = cc_code or cluster["cost_center"]
    div_code = lookups.cc_div_by_code.get(cc_code, "")
    if not div_code and cluster:
        div_code = cluster["division_code"]
    return {
        "agency_code": agency_code,
        "agency_name": agency_name or _clean(rec.get("agency")),
        "agency_resolve_method": agency_resolved.get("method", "unresolved"),
        "agency_resolve_score": agency_resolved.get("score", 0.0),
        "division_code": div_code,
        "division": lookups.div_name_by_code.get(div_code, div_raw),
        "cost_center": cc_code,
        "cost_center_name": lookups.cc_name_by_code.get(cc_code, cc_raw),
    }


def load_so_detail(path: Path) -> dict[str, dict[str, Any]]:
    if not path or not path.exists():
        return {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Sheet1" not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb["Sheet1"]
    headers = {
        _clean(cell.value): idx
        for idx, cell in enumerate(ws[5], start=1)
        if _clean(cell.value)
    }
    # LOCKED: SO_Detail validates only that a canonical JQ exists. Agency,
    # salesperson, and organization columns must not influence allocation.
    required = ["ORDER_NUMBER"]
    missing = [name for name in required if name not in headers]
    if missing:
        wb.close()
        raise RuntimeError(f"SO_Detail missing required columns: {', '.join(missing)}")

    by_jq_rows: dict[str, list[dict[str, Any]]] = {}
    for ridx, row in enumerate(ws.iter_rows(min_row=6, values_only=True), start=6):
        def val(name: str) -> Any:
            pos = headers[name] - 1
            return row[pos] if pos < len(row) else ""

        raw_jq = _clean(val("ORDER_NUMBER"))
        jq = _canonical_jq(raw_jq)
        if not jq:
            continue
        by_jq_rows.setdefault(jq, []).append({
            "jq": jq,
            "raw_jq": raw_jq,
            "row": ridx,
        })
    wb.close()

    index: SoDetailIndex = SoDetailIndex()
    for jq, rows_for_jq in by_jq_rows.items():
        index[jq] = {
            "jq": jq,
            "raw_jq": rows_for_jq[0]["raw_jq"],
            "rows": rows_for_jq,
            "duplicate_row_count": max(0, len(rows_for_jq) - 1),
        }
    return index


def _header_column(ws: Any, aliases: set[str], fallback: int) -> int:
    """Return a 1-based column from the labelled row 8, with known-layout fallback."""
    for cell in ws[8]:
        if _norm_key(cell.value) in aliases:
            return cell.column
    return fallback


def _supplier_description_invoice(raw: Any) -> str:
    """Extract the invoice explicitly repeated in a supplier description."""
    match = re.search(r"(?:^|/)\s*(\d{4,5})\s*$", _clean(raw))
    return _code(match.group(1), 5) if match else ""


def _workbook_agency_reference(wb: Any) -> list[dict[str, str]]:
    sheet_name = next((name for name in wb.sheetnames if name.strip().casefold() == "agency"), "")
    if not sheet_name:
        return []
    ws = wb[sheet_name]
    value_col = description_col = 0
    header_row = 0
    for ridx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10)), start=1):
        labels = {_norm_key(cell.value): cell.column for cell in row if _clean(cell.value)}
        if "value" in labels and "description" in labels:
            value_col, description_col, header_row = labels["value"], labels["description"], ridx
            break
    if not value_col or not description_col:
        return []
    agencies = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        code = _code(row[value_col - 1] if value_col <= len(row) else "", 5)
        name = _clean(row[description_col - 1] if description_col <= len(row) else "")
        if code and name:
            agencies.append({"code": code, "name": name})
    return agencies


def load_expenses_format(path: Path, lookups: Lookups) -> dict[str, list[dict[str, Any]]]:
    if not path or not path.exists():
        return {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Expenses Format" not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb["Expenses Format"]
    workbook_agencies = _workbook_agency_reference(wb)
    description_col = _header_column(ws, {"descriptioncomments", "description", "comments"}, 26)
    columns = {
        "invoice": _header_column(ws, {"invoicenumber"}, 14),
        "jq": _header_column(ws, {"ofjq", "jq", "jqnumber"}, 24),
        "employee_name": _header_column(ws, {"employeename"}, 25),
        # This header is merged across four columns, while supplier data rows
        # can anchor their own merged value in any column within that span.
        "description": tuple(range(description_col, description_col + 4)),
        "agency": _header_column(ws, {"agency"}, 32),
        "division": _header_column(ws, {"division"}, 33),
        "solution": _header_column(ws, {"solution"}, 35),
        "cost_center": _header_column(ws, {"costcenter"}, 36),
        "amount": _header_column(ws, {"amount"}, 37),
        "invoice_amount": _header_column(ws, {"invoiceamount"}, 16),
        "invoice_date": _header_column(ws, {"invoicedate"}, 17),
        "employee_number": _header_column(ws, {"employeenumber", "employeeno"}, 41),
    }

    def cell(vals: list[Any], name: str) -> Any:
        positions = columns[name]
        if isinstance(positions, int):
            positions = (positions,)
        for column in positions:
            pos = column - 1
            value = vals[pos] if pos < len(vals) else ""
            if _clean(value):
                return value
        return ""

    index: dict[str, list[dict[str, Any]]] = {}
    current_inv = ""
    for ridx, row in enumerate(ws.iter_rows(min_row=9, values_only=True), start=9):
        vals = list(row)
        header_inv = _code(cell(vals, "invoice"), 5)
        if header_inv:
            current_inv = header_inv
        description = _clean(cell(vals, "description"))
        description_inv = _supplier_description_invoice(description)
        # Supplier allocation rows repeat the invoice in DESCRIPTION / Comments.
        # Prefer that row-level reference when it conflicts with a mistyped header;
        # unlike header fill-down, it cannot bleed into the following invoice.
        row_inv = description_inv or current_inv
        if not row_inv:
            continue
        jq = _clean(cell(vals, "jq"))
        parsed_jqs = _split_jqs(jq, allow_bare=False)
        employee_name = _clean(cell(vals, "employee_name"))
        agency = _clean(cell(vals, "agency"))
        division = _clean(cell(vals, "division"))
        solution = _clean(cell(vals, "solution"))
        cost_center = _clean(cell(vals, "cost_center"))
        if not lookups.solution_code_by_description.get(_norm_text(solution)):
            alt_solution = cost_center
            if lookups.solution_code_by_description.get(_norm_text(alt_solution)):
                solution = alt_solution
        amount = _money(cell(vals, "amount"))
        invoice_amount = _money(cell(vals, "invoice_amount"))
        invoice_date = cell(vals, "invoice_date")
        employee_number = _code(cell(vals, "employee_number"))
        emp = _find_employee(employee_name, lookups)
        if not employee_number and emp:
            employee_number = emp["emp_no"]
        rec_base = {
            "row": ridx,
            "invoice_no": row_inv,
            "description": description,
            "jq": jq,
            "_source_jq_cell": jq,
            "employee_name": employee_name,
            "employee_number": employee_number,
            "agency": agency,
            "division": division,
            "solution": solution,
            "cost_center": cost_center,
            "amount": amount,
            "invoice_amount": invoice_amount,
            "invoice_date": invoice_date,
            "_supplier_row_amount": amount,
        }
        solution_code, solution_name, solution_note = _solution_from_text(solution, lookups)
        rec_base["solution_code"] = solution_code
        rec_base["solution_name"] = solution_name
        rec_base["solution_note"] = solution_note
        jqs = parsed_jqs or ([_canonical_jq(jq, allow_bare=False)] if _canonical_jq(jq, allow_bare=False) else [jq])
        for jq_index, jq_token in enumerate(jqs, start=1):
            rec = dict(rec_base)
            rec["jq"] = jq_token
            rec["_jq_count"] = len(jqs) if parsed_jqs else 0
            rec["_jq_index"] = jq_index if parsed_jqs else 0
            rec["amount"], rec["_amount_basis"] = _split_supplier_amount(amount, len(jqs), jq_index)
            rec.update(_supplier_resolve_allocation(rec, lookups, workbook_agencies))
            index.setdefault(row_inv, []).append(rec)
    wb.close()
    return index


def supplier_jq_units_for_invoice(invoice_no: Any, supplier_index: dict[str, list[dict[str, Any]]]) -> list[dict[str, Any]]:
    units = []
    for rec in supplier_index.get(_code(invoice_no, 5), []):
        jq = _canonical_jq(rec.get("jq"), allow_bare=False)
        if not jq.startswith("JQ-"):
            continue
        out = dict(rec)
        out["jq"] = jq
        out["_match_method"] = "supplier_jq_unit"
        out["_amount_match"] = True
        units.append(out)
    return units


def match_supplier_line(
    invoice_no: Any,
    line_index: int,
    line_amount: float | None,
    supplier_index: dict[str, list[dict[str, Any]]],
    used_supplier_rows: set[int],
) -> dict[str, Any] | None:
    candidates = supplier_index.get(_code(invoice_no, 5), [])
    if not candidates:
        return None
    if line_index - 1 < len(candidates):
        rec = candidates[line_index - 1]
        if rec["row"] not in used_supplier_rows:
            used_supplier_rows.add(rec["row"])
            supplier_amount = _money(rec.get("amount"))
            amount_match = line_amount is not None and supplier_amount is not None and abs(line_amount - supplier_amount) <= 1
            out = dict(rec)
            out["_match_method"] = "line_order_amount" if amount_match else "line_order"
            out["_amount_match"] = amount_match
            return out
    amount_matches = [
        rec for rec in candidates
        if rec["row"] not in used_supplier_rows
        and line_amount is not None
        and rec.get("amount") is not None
        and abs(float(line_amount) - float(rec["amount"])) <= 1
    ]
    if amount_matches:
        rec = amount_matches[0]
        used_supplier_rows.add(rec["row"])
        out = dict(rec)
        out["_match_method"] = "amount_only"
        out["_amount_match"] = True
        return out
    return None


def supplier_inherited_allocation(
    invoice_no: Any,
    supplier_index: dict[str, list[dict[str, Any]]],
    preferred_jq: Any = "",
    *,
    allow_invoice_fallback: bool = True,
) -> dict[str, Any] | None:
    """Return an unambiguous supplier CC/DIV/Solution block, never agency."""
    candidates = supplier_index.get(_code(invoice_no, 5), [])

    preferred_jq_key = _canonical_jq(preferred_jq, allow_bare=False)
    if preferred_jq_key:
        preferred = [
            rec for rec in candidates
            if _canonical_jq(rec.get("jq"), allow_bare=False) == preferred_jq_key
        ]
        if preferred:
            candidates = preferred
        elif not allow_invoice_fallback:
            return None

    blocks: dict[tuple[str, str, str, str, str, str], dict[str, Any]] = {}
    for rec in candidates:
        block = (
            _code(rec.get("division_code")),
            _clean(rec.get("division")),
            _code(rec.get("cost_center")),
            _clean(rec.get("cost_center_name")),
            _code(rec.get("solution_code"), 5) or "00000",
            _clean(rec.get("solution_name")) or "General",
        )
        if block[0] and block[2]:
            blocks.setdefault(block, rec)
    if len(blocks) != 1:
        return None
    rec = next(iter(blocks.values()))
    return {
        "source": "supplier_expenses_format_allocation_inherited",
        "raw": "supplier allocation inheritance",
        "agency_code": "",
        "agency_name": "",
        "division_code": rec.get("division_code", ""),
        "division": rec.get("division", ""),
        "cost_center": rec.get("cost_center", ""),
        "cost_center_name": rec.get("cost_center_name", ""),
        "solution_code": rec.get("solution_code", "") or "00000",
        "solution_name": rec.get("solution_name", "") or "General",
        "confidence": 1.0,
        "status_reason": "supplier agency is not stated for this preserved sub-line",
        "match_method": "supplier_invoice_allocation_inheritance",
        "agency_resolve_method": "unresolved",
    }


def _additional_info(supplier_line: dict[str, Any] | None, display_jq: Any = "") -> str:
    if not supplier_line:
        return ""
    emp_no = _code(supplier_line.get("employee_number")).strip()
    jq = _clean(display_jq).strip() or _clean(supplier_line.get("jq")).strip()
    if emp_no and jq:
        return f"{emp_no}.{jq}"
    return emp_no or jq


def _agency_display(code: Any, name: Any) -> str:
    code_s = _code(code, 5)
    name_s = _clean(name)
    if code_s and name_s:
        return f"{code_s}/{name_s}"
    return code_s or name_s


def _extract_pdf_jqs(line: dict[str, Any], extraction: dict[str, Any]) -> list[str]:
    fields = [
        line.get("description"),
        line.get("reference"),
        line.get("dispatch_ref"),
        line.get("supply_order"),
        extraction.get("raw"),
    ]
    found: list[str] = []
    seen: set[str] = set()
    for field in fields:
        for jq in _split_jqs(field, allow_bare=False):
            key = _canonical_jq(jq, allow_bare=False)
            if key not in seen:
                seen.add(key)
                found.append(key)
    return found


def sample_files() -> list[tuple[str, str, Path]]:
    out = []
    for label, folder, invs in SAMPLES:
        for inv in invs:
            path = PDF_ROOT / folder / f"{inv}_0001.pdf"
            if not path.exists():
                raise FileNotFoundError(path)
            out.append((label, inv, path))
    return out


def _invoice_sort_key(path: Path) -> int:
    m = re.search(r"(\d{5})_0001\.pdf$", path.name)
    return int(m.group(1)) if m else 10**9


def pdf_dir_cache_tag(pdf_dir: str | None) -> str | None:
    if not pdf_dir:
        return None
    path = Path(pdf_dir).expanduser()
    tag = path.name
    if tag.lower() == "src" and path.parent.name:
        tag = path.parent.name
    return re.sub(r"[^A-Za-z0-9_.-]+", "-", tag).strip("-") or "pdf-dir"


def folder_files(folder_arg: str | None, full: bool, pdf_dir: str | None = None) -> list[tuple[str, str, Path]]:
    if pdf_dir:
        label = folder_arg or "PDF_DIR"
        paths = sorted(Path(pdf_dir).expanduser().glob("*_0001.pdf"), key=_invoice_sort_key)
        out = []
        for path in paths:
            m = re.search(r"(\d{5})_0001\.pdf$", path.name)
            if m:
                out.append((label, m.group(1), path))
        return out

    if not folder_arg:
        return sample_files()

    labels = list(FOLDER_NAMES) if folder_arg == "ALL" else [folder_arg]
    out = []
    for label in labels:
        folder = FOLDER_NAMES[label]
        if full:
            folder_path = PDF_ROOT / folder
            paths = sorted(folder_path.glob("*_0001.pdf"), key=_invoice_sort_key)
            for path in paths:
                m = re.search(r"(\d{5})_0001\.pdf$", path.name)
                if m:
                    out.append((label, m.group(1), path))
        else:
            invs = next(invs for sample_label, _, invs in SAMPLES if sample_label == label)
            for inv in invs:
                path = PDF_ROOT / folder / f"{inv}_0001.pdf"
                if not path.exists():
                    raise FileNotFoundError(path)
                out.append((label, inv, path))
    return out


def build_rows(
    extractions: list[dict[str, Any]],
    lookups: Lookups,
    supplier_index: dict[str, list[dict[str, Any]]] | None = None,
    so_detail_index: dict[str, dict[str, Any]] | None = None,
    available_pdf_invoice_numbers: set[str] | None = None,
    allocation_mode: str = STANDARD_ALLOCATION_MODE,
    project_lookup: dict[str, Any] | None = None,
    project_master_fallback: bool = False,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if allocation_mode not in {STANDARD_ALLOCATION_MODE, PROJECT_ALLOCATION_MODE}:
        raise ValueError(f"unsupported Asateel allocation mode: {allocation_mode}")
    if allocation_mode == PROJECT_ALLOCATION_MODE and not project_lookup:
        raise ValueError("projects-labadi-v1 mode requires a validated project lookup")
    rows = []
    supplier_index = supplier_index or {}
    if so_detail_index is None:
        so_detail_index = {}
    so_detail_enabled = bool(so_detail_index) or bool(getattr(so_detail_index, "loaded", False))
    trace: dict[str, Any] = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "invoices": [],
        "supplier_expenses_format_loaded_invoices": len(supplier_index),
        "so_detail_loaded_jqs": len(so_detail_index),
        "allocation_mode": allocation_mode,
        "project_lookup_statistics": (project_lookup or {}).get("statistics", {}),
    }
    gl_desc, account_note = account_description(lookups)
    # A staged PDF is authoritative evidence that the invoice was supplied, even
    # when extraction fails or reads a different invoice number.  Filename hints
    # are normalized by the same five-digit rule as the master index.
    pdf_invoice_numbers = {
        _code(item.get("invoice_hint"), 5)
        for item in extractions
        if _code(item.get("invoice_hint"), 5)
    }
    if available_pdf_invoice_numbers is None:
        # Direct callers historically pass only the selected folder's
        # extractions. Treat PDFs staged in sibling supplier folders as present.
        available_pdf_invoice_numbers = {
            match.group(1)
            for path in PDF_ROOT.rglob("*_0001.pdf")
            if (match := re.search(r"(\d{5})_0001\.pdf$", path.name))
        }
    pdf_invoice_numbers.update(_code(inv, 5) for inv in available_pdf_invoice_numbers)
    work_items = list(extractions)
    for missing_invoice in sorted(set(supplier_index) - pdf_invoice_numbers):
        master_records = supplier_index[missing_invoice]
        master_header = master_records[0] if master_records else {}
        work_items.append({
            "folder": "MASTER_FALLBACK",
            "invoice_hint": missing_invoice,
            "pdf": "",
            "master_fallback": True,
            "master_invoice_amount": master_header.get("invoice_amount"),
            "master_invoice_date": master_header.get("invoice_date"),
            "payload": {
                "pdf": "",
                "extraction": {
                    "invoice_number": missing_invoice,
                    "invoice_date": master_header.get("invoice_date"),
                    "lines": [{"line_no": 1}],
                },
            },
        })
    trace["missing_pdf_invoices"] = sorted(set(supplier_index) - pdf_invoice_numbers)

    for item in work_items:
        folder = item["folder"]
        is_project_invoice = folder == "PROJECTS" or (bool(item.get("master_fallback")) and project_master_fallback)
        expected_inv = item["invoice_hint"]
        payload = item["payload"]
        master_fallback = bool(item.get("master_fallback"))
        if item.get("error"):
            err = _clean(item.get("error"))
            trace["invoices"].append({
                "folder": folder,
                "expected_invoice": expected_inv,
                "pdf": item.get("pdf"),
                "error": err,
            })
            row = {
                "folder": folder,
                "invoice_no": expected_inv,
                "invoice_date": "",
                "line_no": 1,
                "description": "",
                "reference": "",
                "dispatch_ref": "",
                "supply_order": "",
                "line_amount": 0.0,
                "vat": 0.0,
                "total": 0.0,
                "GL": GL_ACCOUNT,
                "Agency_code": "",
                "Agency_name": "",
                "Division": "",
                "Cost_Center": "",
                "Cost_Center_name": "",
                "allocation_source": "none",
                "split_method": "n/a",
                "confidence": 0,
                "Row_Status": "RED",
                "notes": err,
                "_resolved_full": {"source": "none", "raw": "", "status_reason": err},
                "*Invoice Header Identifier": 1,
                "*Business Unit": BUSINESS_UNIT,
                "*Invoice Number": expected_inv,
                "*Invoice Currency": "SAR",
                "*Invoice Amount": 0.0,
                "*Invoice Date": "",
                "**Supplier[..]": SUPPLIER_NAME,
                "**Supplier Number": "",
                "*Supplier Site[..]": "",
                "Invoice Type": "Standard",
                "Description": "",
                "*Type": "Item",
                "*Amount": 0.0,
                "Tax Classification Code[..]": "KSA VAT STANDARD",
                "Employee No": "",
                "Company": COMPANY,
                "Location": DEFAULT_LOCATION,
                "Account": GL_ACCOUNT,
                "GL": gl_desc,
                "Cost Center": "",
                "Cost Name": "",
                "DIV": "",
                "Contribution": "",
                "Solution": "00000",
                "Solution Name": "General",
                "Agency": "00000",
                "Agency Name": "",
                "Project": "00000",
                "Intercompany": "00",
                "Future 1": "000000",
                "GL Description": gl_desc,
                "_header_subtotal": 0.0,
                "_header_total": 0.0,
                "_amount_basis": "error_placeholder",
                "_trace_pdf": item.get("pdf"),
                "_extracted_brands": "",
                "_extracted_salespeople": "",
                "_extraction_notes": err,
                "_account_note": account_note,
                "_supplier_match": None,
                "_additional_information": "",
                "_supplier_allocation_action": "none",
            }
            row["GL Description"] = _build_gl_description(row)
            row["Distribution Combination[..]"] = build_distribution_combination(row)
            rows.append(row)
            continue
        ext = normalize_extraction(payload.get("extraction") or {})
        brand_values = []
        for sig in ext.get("signals") or []:
            if isinstance(sig, dict) and _clean(sig.get("source")).lower() == "brand" and _clean(sig.get("raw")):
                brand_values.append(_clean(sig.get("raw")))
        for ln in ext.get("lines") or []:
            sig = ln.get("allocation_signal") if isinstance(ln, dict) else None
            if isinstance(sig, dict) and _clean(sig.get("source")).lower() == "brand" and _clean(sig.get("raw")):
                brand_values.append(_clean(sig.get("raw")))
        extracted_brands = "; ".join(dict.fromkeys(brand_values))
        extracted_salespeople = "; ".join(
            _clean(sig.get("raw"))
            for sig in ext.get("signals") or []
            if isinstance(sig, dict) and _clean(sig.get("source")).lower() == "salesperson" and _clean(sig.get("raw"))
        )
        lines = ext.get("lines") or []
        if not lines:
            lines = [{"line_no": 1, "description": "", "line_subtotal": ext.get("subtotal")}]
        lines, expanded_multi_signal = expand_distribution_lines(lines, ext)

        invoice_signal = None
        for sig in ext.get("signals") or []:
            if _clean(sig.get("source")).lower() in {"brand", "salesperson"}:
                invoice_signal = sig
                break

        line_resolutions = []
        for ln in lines:
            sig = ln.get("allocation_signal") or invoice_signal or {"source": "none", "raw": ""}
            if _clean(sig.get("source")).lower() == "none" and invoice_signal:
                sig = invoice_signal
            line_resolutions.append(resolve_signal(sig, lookups))

        subtotal = _money(ext.get("subtotal"))
        vat_total = _money(ext.get("vat")) or 0.0
        total = _money(ext.get("total")) or 0.0
        amount_basis_by_line: list[str] = []
        clean_amounts = []
        for ln in lines:
            amount_notes: list[str] = []
            amount, basis = normalize_exclusive_amount(_money(ln.get("line_subtotal")), subtotal, total, amount_notes)
            clean_amounts.append(amount)
            amount_basis_by_line.append("; ".join([basis] + amount_notes).strip("; "))
        have_all_line_amounts = all(v is not None for v in clean_amounts)
        distinct_allocs = {
            (r.get("agency_code"), r.get("agency_name"), r.get("division"), r.get("cost_center"))
            for r in line_resolutions
            if r.get("agency_name") or r.get("raw")
        }
        if expanded_multi_signal:
            split_method = "even"
        elif len(lines) > 1 and len(distinct_allocs) > 1:
            split_method = "per_line" if have_all_line_amounts else "even"
        elif len(lines) > 1:
            split_method = "per_line" if have_all_line_amounts else "even"
        else:
            split_method = "n/a"

        even_amount = round((subtotal or 0.0) / len(lines), 2) if lines else 0.0
        if expanded_multi_signal:
            clean_amounts = [even_amount for _ in lines]
            have_all_line_amounts = True
            amount_basis_by_line = ["even_split_excl"] * len(lines)

        used_supplier_rows: set[int] = set()
        invoice_no = _code(ext.get("invoice_number"), 5) or expected_inv
        supplier_jq_units = supplier_jq_units_for_invoice(invoice_no, supplier_index)
        paired_signal_sources = [
            _clean((ln.get("allocation_signal") or {}).get("source")).lower()
            for ln in lines
        ]
        preserve_paired_signal_rows = bool(
            expanded_multi_signal
            and len(supplier_jq_units) > 1
            and len(lines) == 2 * len(supplier_jq_units)
            and paired_signal_sources[::2] == ["brand"] * len(supplier_jq_units)
            and paired_signal_sources[1::2] == ["salesperson"] * len(supplier_jq_units)
        )
        invoice_trace = {
            "folder": folder,
            "expected_invoice": expected_inv,
            "pdf": payload.get("pdf"),
            "extraction": ext,
            "line_resolutions": line_resolutions,
            "supplier_matches": [],
            "supplier_jq_unit_count": len(supplier_jq_units),
        }
        trace["invoices"].append(invoice_trace)

        # Multi-signal PDF rows are preserved as evidence rows. Supplier JQs
        # provide their CC/DIV/Solution inheritance below; they must not collapse
        # those rows or reintroduce SO_Detail-driven agency splitting.
        if supplier_jq_units and not preserve_paired_signal_rows:
            output_units = []
            for idx, unit in enumerate(supplier_jq_units, start=1):
                source_idx = min(idx - 1, len(lines) - 1)
                base_unit = {
                    "ln": lines[source_idx],
                    "resolved": line_resolutions[source_idx],
                    "line_amount": _money(unit.get("amount")) or 0.0,
                    "amount_basis": unit.get("_amount_basis") or "supplier_row_amount",
                    "supplier_match": unit,
                    "line_no": idx,
                    "split_method": "per_jq",
                }
                so_detail = so_detail_index.get(_canonical_jq(unit.get("jq")))
                # Supplier sub-lines are already the allocation units. SO_Detail
                # is validation-only and must never expand or replace them.
                base_unit["so_detail_group"] = so_detail
                output_units.append(base_unit)
        else:
            output_units = []
            for idx, ln in enumerate(lines, start=1):
                output_units.append({
                    "ln": ln,
                    "resolved": line_resolutions[idx - 1],
                    "line_amount": clean_amounts[idx - 1] if have_all_line_amounts else even_amount,
                    "amount_basis": amount_basis_by_line[idx - 1],
                    "supplier_match": None,
                    "supplier_inheritance_jq": (
                        supplier_jq_units[(idx - 1) // 2].get("jq", "")
                        if preserve_paired_signal_rows and idx % 2 == 1
                        else ""
                    ),
                    "line_no": ln.get("line_no") or idx,
                    "split_method": split_method,
                })

        for idx, unit in enumerate(output_units, start=1):
            ln = unit["ln"]
            resolved = dict(unit["resolved"])
            notes = list(ext.get("extraction_notes") or [])
            if master_fallback:
                notes.append(
                    "PDF MISSING — allocated from Expenses-Format master only; "
                    "invoice total NOT verified against scan"
                )
            line_amount = unit["line_amount"]
            supplier_match = unit.get("supplier_match")
            if not supplier_match and not ln.get("_expanded_from_multi_signal"):
                supplier_match = match_supplier_line(
                    invoice_no,
                    idx,
                    line_amount,
                    supplier_index,
                    used_supplier_rows,
                )
            project_audit: dict[str, Any] | None = None
            if (
                allocation_mode == PROJECT_ALLOCATION_MODE
                and is_project_invoice
                and supplier_match
            ):
                supplier_match, project_audit = _project_override_supplier_match(supplier_match, project_lookup or {})
            resolved_source = _clean(resolved.get("source")).lower()
            resolved_agency_code = _code(resolved.get("agency_code"), 5)
            supplier_inheritance_jq = unit.get("supplier_inheritance_jq", "")
            can_inherit_supplier_allocation = (
                resolved_source == "brand"
                and resolved_agency_code not in {"", "00000"}
            )
            inherited_supplier_allocation = (
                None
                if supplier_match or not can_inherit_supplier_allocation
                else supplier_inherited_allocation(
                    invoice_no,
                    supplier_index,
                    supplier_inheritance_jq,
                    allow_invoice_fallback=not supplier_inheritance_jq,
                )
            )
            supplier_action = "none"
            so_detail_group = unit.get("so_detail_group")
            unit_jqs = []
            if supplier_match and _canonical_jq(supplier_match.get("jq")):
                unit_jqs.append(_canonical_jq(supplier_match.get("jq")))
            unit_jqs.extend(_extract_pdf_jqs(ln, ext))
            unit_jq = next((jq for jq in unit_jqs if jq), "")
            if not so_detail_group and unit_jq:
                so_detail_group = so_detail_index.get(unit_jq)
            if supplier_match:
                invoice_trace["supplier_matches"].append({
                    "line_no": unit["line_no"],
                    "supplier_row": supplier_match.get("row"),
                    "match_method": supplier_match.get("_match_method"),
                    "amount_match": supplier_match.get("_amount_match"),
                    "jq": supplier_match.get("jq"),
                    "parsed_jq": supplier_match.get("jq"),
                    "source_jq_cell": supplier_match.get("_source_jq_cell"),
                    "jq_index": supplier_match.get("_jq_index"),
                    "jq_count": supplier_match.get("_jq_count"),
                    "employee_number": supplier_match.get("employee_number"),
                    "agency": supplier_match.get("agency"),
                    "agency_code": supplier_match.get("agency_code"),
                    "agency_resolve_method": supplier_match.get("agency_resolve_method"),
                    "manpower_home_agency": "",
                    "home_agency_discrepancy": "",
                    "solution": supplier_match.get("solution"),
                    "solution_code": supplier_match.get("solution_code"),
                    "project_allocation": project_audit,
                })
                if not supplier_match.get("_amount_match"):
                    notes.append(
                        f"Supplier Expenses Format matched by line order; supplier amount {supplier_match.get('amount')} differs from PDF line amount {line_amount}"
                    )
            elif _extract_pdf_jqs(ln, ext):
                notes.append("JQ found in PDF but missing from Supplier Expenses Format; PDF signal not used for allocation")
            manpower_emp = _find_employee_by_no(supplier_match.get("employee_number"), lookups) if supplier_match else None
            supplier_sheet_agency = _agency_display(
                supplier_match.get("agency_code") if supplier_match else "",
                supplier_match.get("agency_name") if supplier_match else "",
            )
            manpower_home_agency = _agency_display(
                manpower_emp.get("agency_code") if manpower_emp else "",
                manpower_emp.get("agency_name") if manpower_emp else "",
            )
            supplier_home_agency_discrepancy = ""
            so_detail_agency = ""
            so_detail_salesperson = ""
            so_detail_supplier_discrepancy = ""
            if supplier_match and supplier_match.get("employee_number") and not manpower_emp:
                notes.append(f"Supplier employee number {supplier_match.get('employee_number')} not found in Manpower")
            if project_audit:
                notes.append(project_audit.get("explanation", ""))
            supplier_has_allocation = bool(
                supplier_match
                and supplier_match.get("agency_code")
                and supplier_match.get("division_code")
                and supplier_match.get("cost_center")
            )
            supplier_allocation = {
                "source": PROJECT_ALLOCATION_MODE if project_audit and project_audit.get("agency_after") else "supplier_expenses_format",
                "raw": _clean(supplier_match.get("agency")) if supplier_match else "",
                "agency_code": supplier_match.get("agency_code") if supplier_match else "",
                "agency_name": supplier_match.get("agency_name") if supplier_match else "",
                "division_code": supplier_match.get("division_code") if supplier_match else "",
                "division": supplier_match.get("division") if supplier_match else "",
                "cost_center": supplier_match.get("cost_center") if supplier_match else "",
                "cost_center_name": supplier_match.get("cost_center_name") if supplier_match else "",
                "solution_code": supplier_match.get("solution_code") if supplier_match else "",
                "solution_name": supplier_match.get("solution_name") if supplier_match else "",
                "confidence": 1.0 if supplier_match and supplier_match.get("_match_method") == "line_order_amount" else 0.85,
                "status_reason": "",
                "match_method": supplier_match.get("_match_method") if supplier_match else "",
                "agency_resolve_method": supplier_match.get("agency_resolve_method") if supplier_match else "",
            }
            if supplier_has_allocation:
                resolved = supplier_allocation
                supplier_action = "supplier_override"
                notes.append("Full allocation block used from Supplier Expenses Format")
                if (
                    manpower_emp
                    and _code(manpower_emp.get("agency_code"), 5)
                    and _code(manpower_emp.get("agency_code"), 5) != _code(supplier_match.get("agency_code"), 5)
                    and not (project_audit and project_audit.get("status") == "applied")
                ):
                    supplier_home_agency_discrepancy = "Y"
                    notes.append(
                        f"Supplier agency {supplier_sheet_agency} differs from Manpower home agency {manpower_home_agency}; supplier sheet used"
                    )
            elif supplier_match:
                supplier_action = "supplier_unresolved"
                resolved = supplier_allocation
                resolved["source"] = "supplier_expenses_format_unresolved"
                resolved["status_reason"] = "supplier agency text unresolved" if supplier_match.get("agency") and not supplier_match.get("agency_code") else "supplier allocation incomplete"
                notes.append(
                    f"Supplier Agency '{supplier_match.get('agency')}' did not resolve to Agency lookup; supplier sheet not silently replaced by Manpower"
                )
            elif inherited_supplier_allocation:
                supplier_action = "supplier_allocation_inherited"
                resolved = dict(inherited_supplier_allocation)
                notes.append(
                    "Supplier CC/DIV/Solution inherited for preserved sub-line; agency left blank because no line-specific supplier agency was stated"
                )
            else:
                # OCR brand/salesperson signals remain evidence only. Allocation
                # authority is the Expenses Format supplier sub-line.
                resolved = dict(supplier_allocation)
                resolved["source"] = "supplier_expenses_format_missing"
                resolved["status_reason"] = "supplier allocation line missing"
            if so_detail_enabled and unit_jq and unit_jq not in so_detail_index:
                notes.append("JQ not in SO_Detail export")
            jq_display = ""
            if supplier_match:
                jq_key = _canonical_jq(supplier_match.get("jq"))
                jq_display = _clean((so_detail_index.get(jq_key) or {}).get("raw_jq"))
            additional_info = _additional_info(supplier_match, jq_display)
            if supplier_match and not additional_info:
                notes.append("Employee Number and JQ unavailable for matched Supplier Expenses Format line")
            elif not supplier_match:
                notes.append("No Supplier Expenses Format line matched for Employee Number/JQ")
            solution_code = "00000"
            solution_name = "General"
            solution_source = supplier_match or inherited_supplier_allocation
            if solution_source:
                solution_code = _code(solution_source.get("solution_code"), 5) or "00000"
                solution_name = _clean(solution_source.get("solution_name")) or "General"
                if solution_source.get("solution_note"):
                    notes.append(solution_source.get("solution_note"))
            else:
                notes.append("No Supplier Expenses Format line matched for Solution")
            status = classify(ext, resolved, notes, scan_available=not master_fallback)
            if project_audit and project_audit.get("status") != "applied" and status == "GREEN":
                status = "YELLOW"
            if supplier_action in {"supplier_override", "supplier_unresolved"}:
                status = "YELLOW" if status == "GREEN" and supplier_home_agency_discrepancy else status
            if supplier_action == "filled" and supplier_match and supplier_match.get("_match_method") != "line_order_amount" and status == "GREEN":
                status = "YELLOW"
            if not supplier_match and _extract_pdf_jqs(ln, ext) and status == "GREEN":
                status = "YELLOW"
            is_warehouse_cc = (
                _code(resolved.get("cost_center")) == "140040"
                or _clean(resolved.get("cost_center_name")).casefold() == "warehouse"
            )
            if so_detail_enabled and not so_detail_group and (unit_jq or supplier_match) and not is_warehouse_cc:
                if status == "GREEN":
                    status = "YELLOW"
            if master_fallback:
                status = "RED"
            line_vat = round((line_amount or 0) * VAT_RATE, 2)
            line_total = round((line_amount or 0) + line_vat, 2)
            if len(output_units) == 1 and not supplier_jq_units:
                line_vat = vat_total
                line_total = total
            is_green = status == "GREEN"
            has_resolved_segments = bool(resolved.get("agency_name") and resolved.get("division") and resolved.get("cost_center"))
            keep_allocation = is_green or supplier_action in {"filled", "manpower_empno", "supplier_agency", "supplier_override", "supplier_allocation_inherited"} or (
                has_resolved_segments and _clean(resolved.get("allocation_status")) and _clean(resolved.get("allocation_status")) != "Can Be used"
            )
            agency_code = resolved.get("agency_code") if keep_allocation else ""
            agency_name = resolved.get("agency_name") if keep_allocation else ""
            division_code = resolved.get("division_code") if keep_allocation else ""
            division_name = resolved.get("division") if keep_allocation else ""
            cost_center = resolved.get("cost_center") if keep_allocation else ""
            cost_center_name = resolved.get("cost_center_name") if keep_allocation else ""
            if not is_green and not keep_allocation:
                notes.append(resolved.get("status_reason") or "allocation unresolved")
            debug_notes = [n for n in notes if _clean(n)]
            if account_note:
                debug_notes.append(account_note)
            debug_notes.append(f"Location confirmed as {DEFAULT_LOCATION}")
            if resolved.get("brand_remap_from") and resolved.get("brand_remap_to"):
                debug_notes.append(f"Brand remapped {resolved.get('brand_remap_from')} -> {resolved.get('brand_remap_to')}")
            description = " ".join(
                p for p in [
                    _clean(ln.get("description")),
                    _clean(ln.get("reference")),
                    _clean(ln.get("dispatch_ref")),
                ]
                if p
            )
            row = {
                "folder": folder,
                "invoice_no": invoice_no,
                "invoice_date": _date_str(ext.get("invoice_date")),
                "line_no": unit["line_no"],
                "description": _clean(ln.get("description")),
                "reference": _clean(ln.get("reference")),
                "dispatch_ref": _clean(ln.get("dispatch_ref")),
                "supply_order": _clean(ln.get("supply_order")),
                "line_amount": line_amount,
                "vat": line_vat,
                "total": line_total,
                "GL": GL_ACCOUNT,
                "Agency_code": agency_code,
                "Agency_name": agency_name,
                "Division": division_name,
                "Cost_Center": cost_center,
                "Cost_Center_name": cost_center_name,
                "allocation_source": resolved.get("source") if resolved.get("raw") else "none",
                "agency_resolve_method": resolved.get("agency_resolve_method") or "manpower_fallback",
                "split_method": unit["split_method"],
                "confidence": resolved.get("confidence") or 0,
                "Row_Status": status,
                "notes": "; ".join(_clean(n) for n in debug_notes if _clean(n)),
                "_resolved_full": resolved,
                "*Invoice Header Identifier": 1,
                "*Business Unit": BUSINESS_UNIT,
                "*Invoice Number": invoice_no,
                "*Invoice Currency": "SAR",
                "*Invoice Amount": item.get("master_invoice_amount") if master_fallback else total,
                "*Invoice Date": _oracle_date_str(ext.get("invoice_date")),
                "**Supplier[..]": SUPPLIER_NAME,
                "**Supplier Number": "",
                "*Supplier Site[..]": "",
                "Invoice Type": "Standard",
                "Description": description,
                "*Type": "Item",
                "*Amount": line_amount,
                "Tax Classification Code[..]": "KSA VAT STANDARD",
                "Employee No": _code(supplier_match.get("employee_number")) if supplier_match else "",
                "Company": COMPANY,
                "Location": DEFAULT_LOCATION,
                "Account": GL_ACCOUNT,
                "GL": gl_desc,
                "Cost Center": cost_center,
                "Cost Name": cost_center_name,
                "DIV": division_code,
                "Contribution": division_name,
                "Solution": solution_code,
                "Solution Name": solution_name,
                "Agency": agency_code or "00000",
                "Agency Name": agency_name,
                "Project": "00000",
                "Intercompany": "00",
                "Future 1": "000000",
                "GL Description": "",
                "_header_subtotal": "" if master_fallback else subtotal,
                "_header_total": "" if master_fallback else total,
                "_amount_basis": unit["amount_basis"],
                "_trace_pdf": "" if master_fallback else payload.get("pdf"),
                "_extracted_brands": "" if master_fallback else extracted_brands,
                "_extracted_salespeople": "" if master_fallback else extracted_salespeople,
                "_extraction_notes": "" if master_fallback else "; ".join(_clean(n) for n in ext.get("extraction_notes") or [] if _clean(n)),
                "_exception_category": "MISSING_PDF" if master_fallback else "",
                "_account_note": account_note,
                "_supplier_match": supplier_match,
                "_additional_information": additional_info,
                "_supplier_allocation_action": supplier_action,
                "_supplier_sheet_agency": supplier_sheet_agency,
                "_manpower_home_agency": manpower_home_agency,
                "_so_detail_agency": so_detail_agency,
                "_so_detail_salesperson": so_detail_salesperson,
                "_so_detail_supplier_discrepancy": so_detail_supplier_discrepancy,
                "_so_detail_jq": unit_jq,
                "_supplier_home_agency_discrepancy": supplier_home_agency_discrepancy,
                "_supplier_jq_count": supplier_match.get("_jq_count") if supplier_match else "",
                "_project_allocation_audit": project_audit,
            }
            row["GL Description"] = _build_gl_description(row)
            row["Distribution Combination[..]"] = build_distribution_combination(row)
            if invoice_trace["supplier_matches"]:
                invoice_trace["supplier_matches"][-1]["manpower_home_agency"] = manpower_home_agency
                invoice_trace["supplier_matches"][-1]["home_agency_discrepancy"] = supplier_home_agency_discrepancy
            if status != "GREEN":
                row["notes"] = row["notes"] or resolved.get("status_reason") or "review required"
            rows.append(row)
    trace["project_lookup_status_counts"] = dict(Counter(
        _clean((row.get("_project_allocation_audit") or {}).get("status"))
        for row in rows
        if row.get("_project_allocation_audit")
    ))
    invoice_serials: dict[str, int] = {}
    for row in rows:
        invoice_no = _clean(row.get("*Invoice Number"))
        first_invoice_row = invoice_no not in invoice_serials
        if first_invoice_row:
            invoice_serials[invoice_no] = len(invoice_serials) + 1
        row["*Invoice Header Identifier"] = invoice_serials[invoice_no]
        if not first_invoice_row:
            row["*Invoice Amount"] = ""
            row["*Invoice Date"] = ""
    return rows, trace


def _row_style(status: str):
    sys.path.insert(0, str(ROOT / "scripts"))
    try:
        from excel_styling import get_row_style
    except Exception:
        fills = {
            "GREEN": (PatternFill("solid", fgColor="C6EFCE"), Font(color="006100")),
            "YELLOW": (PatternFill("solid", fgColor="FFEB9C"), Font(color="9C5700")),
            "RED": (PatternFill("solid", fgColor="FFC7CE"), Font(color="9C0006")),
        }
        get_row_style = lambda status: fills.get(status, fills["GREEN"])
    return get_row_style(status)


def style_workbook(path: Path) -> None:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    status_col = header.index("Row_Status") + 1
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        status = row[status_col - 1].value or "GREEN"
        fill, font = _row_style(status)
        for cell in row:
            cell.fill = fill
            cell.font = font
    for col in ws.columns:
        width = min(max(len(str(c.value or "")) for c in col) + 2, 42)
        ws.column_dimensions[col[0].column_letter].width = width
    wb.save(path)
    wb.close()


def write_excel(rows: list[dict[str, Any]], path: Path) -> None:
    ref_wb = openpyxl.load_workbook(JAWAL_TEMPLATE_XLSX)
    ref_ws = ref_wb["Sheet"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"

    oracle_headers = [ref_ws.cell(3, c).value for c in range(1, 33)]
    headers = oracle_headers + DEBUG_HEADERS
    header_count = len(headers)
    if header_count < 59:
        raise RuntimeError(f"Oracle output must have at least 59 columns, got {header_count}")

    for r in range(1, 4):
        for c in range(1, header_count + 1):
            src = ref_ws.cell(r, c)
            dst = ws.cell(r, c)
            dst.value = src.value
            if src.has_style:
                dst.font = copy(src.font)
                dst.fill = copy(src.fill)
                dst.border = copy(src.border)
                dst.alignment = copy(src.alignment)
                dst.number_format = src.number_format
                dst.protection = copy(src.protection)
    for merged in ref_ws.merged_cells.ranges:
        if merged.min_row == 2 and merged.max_row == 2:
            ws.merge_cells(str(merged))
    ws["AG2"] = "DEBUG (delete before upload)"
    for c, header in enumerate(headers, start=1):
        ws.cell(3, c).value = header
        if c >= 33:
            src = ref_ws.cell(3, 33)
            dst = ws.cell(3, c)
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)

    for c in range(1, header_count + 1):
        letter = openpyxl.utils.get_column_letter(c)
        ws.column_dimensions[letter].width = ref_ws.column_dimensions[letter].width or 14

    for ridx, row in enumerate(rows, start=4):
        debug_values = {
            "Row Status": row.get("Row_Status"),
            "Allocation Source": row.get("allocation_source"),
            "Agency Resolve Method": row.get("agency_resolve_method") or (row.get("_resolved_full") or {}).get("agency_resolve_method", ""),
            "Additional Information": row.get("_additional_information"),
            "SO_Detail Agency": row.get("_so_detail_agency"),
            "SO_Detail Salesperson": row.get("_so_detail_salesperson"),
            "Supplier Sheet Agency": row.get("_supplier_sheet_agency"),
            "Manpower Home Agency": row.get("_manpower_home_agency"),
            "SO_Detail vs Supplier Discrepancy": row.get("_so_detail_supplier_discrepancy"),
            "Home Agency Discrepancy": row.get("_supplier_home_agency_discrepancy"),
            "Supplier JQ Count": row.get("_supplier_jq_count"),
            "Extracted Brand(s)": row.get("_extracted_brands"),
            "Extracted Salesperson": row.get("_extracted_salespeople"),
            "Agency Match Confidence": row.get("confidence"),
            "Manpower Cluster Found (Y/N)": "Y" if row.get("Cost Center") and row.get("DIV") else "N",
            "Split Method": row.get("split_method"),
            "Reference(المرجع)": row.get("reference"),
            "Dispatch Ref": row.get("dispatch_ref"),
            "Supply Order": row.get("supply_order"),
            "VAT Amount": row.get("vat"),
            "Notes": row.get("notes"),
            "Folder": row.get("folder"),
            "Invoice Hint": row.get("invoice_no"),
            "Line No": row.get("line_no"),
            "Resolved Source Raw": (row.get("_resolved_full") or {}).get("raw", ""),
            "Resolved Match Method": (row.get("_resolved_full") or {}).get("match_method", ""),
            "Status Reason": (row.get("_resolved_full") or {}).get("status_reason", ""),
            "Extraction Notes": row.get("_extraction_notes"),
            "Header Subtotal": row.get("_header_subtotal"),
            "Header Total": row.get("_header_total"),
            "Amount Basis": row.get("_amount_basis"),
            "Location Assumption": f"confirmed {DEFAULT_LOCATION}",
            "Account Lookup Note": row.get("_account_note"),
            "Trace PDF": row.get("_trace_pdf"),
            "Agent Method": "asateel_poc_cached_gemini",
            "Project Allocation Mode": (row.get("_project_allocation_audit") or {}).get("mode", ""),
            "Project Lookup Status": (row.get("_project_allocation_audit") or {}).get("status", ""),
            "Project Lookup Explanation": (row.get("_project_allocation_audit") or {}).get("explanation", ""),
        }
        for c, header in enumerate(headers, start=1):
            value = row.get(header, debug_values.get(header, ""))
            cell = ws.cell(ridx, c)
            cell.value = value
            if header == "*Invoice Date":
                cell.number_format = "mm/dd/yyyy"

        fill, font = _row_style(row.get("Row_Status") or "GREEN")
        for c in range(1, header_count + 1):
            cell = ws.cell(ridx, c)
            cell.fill = copy(fill)
            cell.font = copy(font)

    ws.freeze_panes = "A4"
    last_col = openpyxl.utils.get_column_letter(header_count)
    ws.auto_filter.ref = f"A3:{last_col}{max(ws.max_row, 3)}"
    ref_wb.close()
    wb.save(path)
    wb.close()


def parse_distribution(combo: Any) -> dict[str, str]:
    s = _clean(combo)
    if not s:
        return {}
    parts = [p.strip() for p in re.split(r"[-.]", s) if p.strip()]
    if len(parts) < 10:
        return {"raw": s}
    return {
        "raw": s,
        "account": parts[2],
        "cost_center": parts[3],
        "division": parts[4],
        "solution": parts[5],
        "agency": parts[6],
    }


def load_answer_key() -> list[dict[str, Any]]:
    rows = []
    for path in ENTRY_FILES:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=False)
        ws = wb["Invoices"]
        current_inv = ""
        current_amount = None
        current_date = ""
        for ridx, row in enumerate(ws.iter_rows(min_row=9, values_only=True), start=9):
            vals = list(row)
            inv = _code(vals[7] if len(vals) > 7 else "", 5)
            if inv:
                current_inv = inv
                current_amount = _money(vals[9] if len(vals) > 9 else None)
                current_date = _date_str(vals[10] if len(vals) > 10 else "")
            if not current_inv:
                continue
            line_amount = _money(vals[82] if len(vals) > 82 else None)
            combo = vals[98] if len(vals) > 98 else None
            dist = parse_distribution(combo)
            if dist.get("account") or line_amount is not None:
                rows.append({
                    "source_file": path.name,
                    "row": ridx,
                    "invoice_no": current_inv,
                    "invoice_amount": current_amount,
                    "invoice_date": current_date,
                    "line_amount": line_amount,
                    "distribution": _clean(combo),
                    **dist,
                })
        wb.close()
    return rows


def _division_compare_value(v: Any, lookups: Lookups, resolved: dict[str, Any] | None = None) -> str:
    if resolved and resolved.get("division_code"):
        return _code(resolved.get("division_code"))
    s = _code(v)
    if s.isdigit():
        return s
    return lookups.div_code_by_name.get(_clean(v).casefold(), s)


def validate(rows: list[dict[str, Any]], lookups: Lookups) -> dict[str, Any]:
    answer_rows = load_answer_key()
    sample_invs = {_code(r["invoice_no"], 5) for r in rows}
    answer_by_inv: dict[str, list[dict[str, Any]]] = {}
    for r in answer_rows:
        if r["invoice_no"] in sample_invs:
            answer_by_inv.setdefault(r["invoice_no"], []).append(r)

    details = []
    invs = sorted(sample_invs)
    agency_hit = cc_hit = div_hit = all_hit = distribution_hit = amount_hit = comparable = 0
    source_breakdown: dict[str, dict[str, int]] = {}
    for inv in invs:
        poc_rows = [r for r in rows if _code(r["invoice_no"], 5) == inv]
        poc_rows = sorted(
            poc_rows,
            key=lambda r: int(bool(r.get("Agency_code") or r.get("Cost_Center") or r.get("Division"))),
            reverse=True,
        )
        answers = answer_by_inv.get(inv, [])
        matched = []
        unused = answers.copy()
        for pr in poc_rows:
            candidates = unused or answers
            best = None
            if candidates:
                amount = _money(pr.get("line_amount"))
                with_amount = [a for a in candidates if amount is not None and a.get("line_amount") is not None and abs(a["line_amount"] - amount) <= 1]
                pool = with_amount or candidates
                def score_answer(ans: dict[str, Any]) -> tuple[int, int]:
                    resolved = pr.get("_resolved_full") if isinstance(pr.get("_resolved_full"), dict) else {}
                    return (
                        int(_code(pr.get("Agency_code"), 5) == _code(ans.get("agency"), 5))
                        + int(_code(pr.get("Cost_Center")) == _code(ans.get("cost_center")))
                        + int(_division_compare_value(pr.get("Division"), lookups, resolved) == _code(ans.get("division"))),
                        int(amount is not None and ans.get("line_amount") is not None and abs(ans["line_amount"] - amount) <= 1),
                    )
                best = max(pool, key=score_answer)
                if best in unused:
                    unused.remove(best)
            if best and best.get("account"):
                comparable += 1
                ah = _code(pr.get("Agency_code"), 5) == _code(best.get("agency"), 5)
                ch = _code(pr.get("Cost_Center")) == _code(best.get("cost_center"))
                resolved = pr.get("_resolved_full") if isinstance(pr.get("_resolved_full"), dict) else {}
                dh = _division_compare_value(pr.get("Division"), lookups, resolved) == _code(best.get("division"))
                combo_match = _clean(pr.get("Distribution Combination[..]")) == _clean(best.get("distribution"))
                amount_match = (
                    _money(pr.get("line_amount")) is not None
                    and best.get("line_amount") is not None
                    and abs((_money(pr.get("line_amount")) or 0) - best["line_amount"]) <= 0.01
                )
                agency_hit += int(ah)
                cc_hit += int(ch)
                div_hit += int(dh)
                all_hit += int(ah and ch and dh)
                distribution_hit += int(combo_match)
                amount_hit += int(amount_match)
                source = _clean(pr.get("allocation_source")) or "none"
                source_stats = source_breakdown.setdefault(
                    source,
                    {"lines": 0, "agency": 0, "cost_center": 0, "division": 0, "distribution": 0, "amount": 0},
                )
                source_stats["lines"] += 1
                source_stats["agency"] += int(ah)
                source_stats["cost_center"] += int(ch)
                source_stats["division"] += int(dh)
                source_stats["distribution"] += int(combo_match)
                source_stats["amount"] += int(amount_match)
                matched.append({
                    "line_no": pr.get("line_no"),
                    "poc_amount": pr.get("line_amount"),
                    "answer_amount": best.get("line_amount"),
                    "poc_agency": pr.get("Agency_code"),
                    "answer_agency": best.get("agency"),
                    "agency_match": ah,
                    "poc_cost_center": pr.get("Cost_Center"),
                    "answer_cost_center": best.get("cost_center"),
                    "cost_center_match": ch,
                    "poc_division": _division_compare_value(pr.get("Division"), lookups, resolved),
                    "answer_division": best.get("division"),
                    "division_match": dh,
                    "distribution_match": combo_match,
                    "amount_match": amount_match,
                    "answer_distribution": best.get("distribution"),
                })
            else:
                matched.append({"line_no": pr.get("line_no"), "no_answer_key": True})
        details.append({"invoice_no": inv, "matches": matched, "answer_rows": len(answers)})

    return {
        "answer_rows_loaded": len(answer_rows),
        "comparable_rows": comparable,
        "agency_hits": agency_hit,
        "cost_center_hits": cc_hit,
        "division_hits": div_hit,
        "all_segment_hits": all_hit,
        "distribution_hits": distribution_hit,
        "amount_hits": amount_hit,
        "agency_rate": round(agency_hit / comparable, 3) if comparable else 0,
        "cost_center_rate": round(cc_hit / comparable, 3) if comparable else 0,
        "division_rate": round(div_hit / comparable, 3) if comparable else 0,
        "all_segment_rate": round(all_hit / comparable, 3) if comparable else 0,
        "distribution_rate": round(distribution_hit / comparable, 3) if comparable else 0,
        "amount_rate": round(amount_hit / comparable, 3) if comparable else 0,
        "source_breakdown": source_breakdown,
        "details": details,
    }


def validate_output_headers(path: Path) -> list[str]:
    out_wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ref_wb = openpyxl.load_workbook(JAWAL_TEMPLATE_XLSX, read_only=True, data_only=True)
    out_ws = out_wb["Sheet"]
    ref_ws = ref_wb["Sheet"]
    diffs = []
    for c in range(1, 33):
        out_val = out_ws.cell(3, c).value
        ref_val = ref_ws.cell(3, c).value
        if out_val != ref_val:
            diffs.append(f"{openpyxl.utils.get_column_letter(c)}: output={out_val!r} reference={ref_val!r}")
    out_wb.close()
    ref_wb.close()
    return diffs


def print_summary(
    rows: list[dict[str, Any]],
    validation: dict[str, Any],
    header_diffs: list[str],
    xlsx_path: Path,
    json_path: Path,
) -> None:
    print("\nASATEEL POC SUMMARY")
    print("===================")
    invoice_numbers = sorted({_code(r["invoice_no"], 5) for r in rows})
    for inv in invoice_numbers:
        inv_rows = [r for r in rows if _code(r["invoice_no"], 5) == inv]
        sources = []
        green = yellow = red = 0
        for r in inv_rows:
            source = _clean(r.get("allocation_source")) or "none"
            if source not in sources:
                sources.append(source)
            green += int(r.get("Row_Status") == "GREEN")
            yellow += int(r.get("Row_Status") == "YELLOW")
            red += int(r.get("Row_Status") == "RED")
        print(
            f"{inv_rows[0]['folder']} {inv}: lines={len(inv_rows)} | "
            f"sources={','.join(sources) or 'none'} | GREEN={green} YELLOW={yellow} RED={red}"
        )

    green = sum(1 for r in rows if r.get("Row_Status") == "GREEN")
    yellow = sum(1 for r in rows if r.get("Row_Status") == "YELLOW")
    red = sum(1 for r in rows if r.get("Row_Status") == "RED")
    fully_resolved = 0
    for inv in invoice_numbers:
        inv_rows = [r for r in rows if _code(r["invoice_no"], 5) == inv]
        fully_resolved += int(bool(inv_rows) and all(r.get("Row_Status") == "GREEN" for r in inv_rows))
    invoice_total_sum = 0.0
    seen_invoices = set()
    for r in rows:
        inv = _code(r.get("*Invoice Number"), 5)
        if inv and inv not in seen_invoices:
            seen_invoices.add(inv)
            invoice_total_sum += _money(r.get("*Invoice Amount")) or 0.0
    line_excl_sum = sum((_money(r.get("*Amount")) or 0.0) for r in rows)

    print("\nTOTALS")
    print("======")
    multi_line_invoices = sum(
        1 for inv in invoice_numbers
        if len([r for r in rows if _code(r["invoice_no"], 5) == inv]) > 1
    )
    print(f"Invoices processed: {len(invoice_numbers)}")
    print(f"Multi-line invoices detected: {multi_line_invoices}")
    print(f"Distribution rows written: {len(rows)}")
    print(f"GREEN/YELLOW/RED rows: {green}/{yellow}/{red}")
    print(f"Invoices fully resolved vs needing review: {fully_resolved}/{len(invoice_numbers) - fully_resolved}")
    print(f"Sum col E invoice totals, dedup per invoice: {round(invoice_total_sum, 2)}")
    print(f"Sum col M line excl-VAT amounts: {round(line_excl_sum, 2)}")

    inv_03317_rows = [r for r in rows if _code(r.get("invoice_no"), 5) == "03317"]
    print("\n03317 SMOKE CHECK")
    print("=================")
    if inv_03317_rows:
        print(f"Lines captured: {len(inv_03317_rows)}")
        for r in inv_03317_rows:
            print(
                f"- line {r.get('line_no')}: ref={_clean(r.get('reference')) or _clean(r.get('description'))} | "
                f"source={_clean(r.get('allocation_source')) or 'none'} | "
                f"agency={_clean(r.get('Agency'))}/{_clean(r.get('Agency Name')) or 'unresolved'} | "
                f"cc={_clean(r.get('Cost Center')) or 'blank'} | div={_clean(r.get('DIV')) or 'blank'} | "
                f"solution={_clean(r.get('Solution'))}/{_clean(r.get('Solution Name')) or 'General'} | "
                f"amount={r.get('line_amount')} | "
                f"additional_info={_clean(r.get('_additional_information')) or 'blank'}"
            )
    else:
        print("03317 was not included in this run.")

    inv_03041_rows = [r for r in rows if _code(r.get("invoice_no"), 5) == "03041"]
    print("\n03041 SMOKE CHECK")
    print("=================")
    if inv_03041_rows:
        print(f"Lines captured: {len(inv_03041_rows)}")
        for r in inv_03041_rows:
            print(
                f"- line {r.get('line_no')}: ref={_clean(r.get('reference')) or _clean(r.get('description'))} | "
                f"source={_clean(r.get('allocation_source')) or 'none'} | "
                f"agency={_clean(r.get('Agency'))}/{_clean(r.get('Agency Name')) or 'unresolved'} | "
                f"agency_method={_clean(r.get('agency_resolve_method')) or 'blank'} | "
                f"cc={_clean(r.get('Cost Center')) or 'blank'} | div={_clean(r.get('DIV')) or 'blank'} | "
                f"solution={_clean(r.get('Solution'))}/{_clean(r.get('Solution Name')) or 'General'} | "
                f"amount={r.get('line_amount')} | "
                f"additional_info={_clean(r.get('_additional_information')) or 'blank'}"
            )
    else:
        print("03041 was not included in this run.")

    remap_counts = {"3M->Solvento": 0, "Biofire->BMX": 0, "Biomerieux->BMX": 0}
    supplier_matched = supplier_filled = supplier_corroborated = 0
    addl_full = 0
    location_ok = 0
    source_counts: dict[str, int] = {}
    need_allocate_red = 0
    solution_resolved = 0
    agency_method_counts: dict[str, int] = {}
    agency_unresolved = 0
    supplier_agency_unresolved = 0
    for r in rows:
        source = _clean(r.get("allocation_source")) or "none"
        source_counts[source] = source_counts.get(source, 0) + 1
        agency_method = _clean(r.get("agency_resolve_method")) or "manpower_fallback"
        agency_method_counts[agency_method] = agency_method_counts.get(agency_method, 0) + 1
        agency_unresolved += int(_code(r.get("Agency"), 5) in {"", "00000"})
        resolved = r.get("_resolved_full") if isinstance(r.get("_resolved_full"), dict) else {}
        remap_from = _norm_key(resolved.get("brand_remap_from"))
        remap_to = _clean(resolved.get("brand_remap_to"))
        if remap_from == "3m" and remap_to == "Solvento":
            remap_counts["3M->Solvento"] += 1
        elif remap_from == "biofire" and remap_to == "BMX":
            remap_counts["Biofire->BMX"] += 1
        elif remap_from == "biomerieux" and remap_to == "BMX":
            remap_counts["Biomerieux->BMX"] += 1
        supplier_match = r.get("_supplier_match")
        supplier_matched += int(bool(supplier_match))
        action = _clean(r.get("_supplier_allocation_action"))
        supplier_filled += int(action == "filled")
        supplier_corroborated += int(action == "corroborated")
        supplier_line = supplier_match if isinstance(supplier_match, dict) else {}
        addl_full += int(bool(_code(supplier_line.get("employee_number"))) and bool(_clean(supplier_line.get("jq"))))
        combo_parts = _clean(r.get("Distribution Combination[..]")).split("-")
        location_ok += int(_code(r.get("Location"), 5) == DEFAULT_LOCATION and len(combo_parts) > 1 and combo_parts[1] == DEFAULT_LOCATION)
        allocation_status = _clean(resolved.get("allocation_status"))
        need_allocate_red += int(bool(r.get("Row_Status") == "RED" and allocation_status and allocation_status != "Can Be used"))
        solution_resolved += int(_code(r.get("Solution"), 5) not in {"", "00000"})
        supplier_line = r.get("_supplier_match") if isinstance(r.get("_supplier_match"), dict) else {}
        supplier_agency_unresolved += int(bool(_clean(supplier_line.get("agency"))) and not _code(supplier_line.get("agency_code"), 5))

    print("\nREFINEMENT METRICS")
    print("==================")
    print(
        "Brand-remap hits: "
        f"3M->Solvento={remap_counts['3M->Solvento']}, "
        f"Biofire->BMX={remap_counts['Biofire->BMX']}, "
        f"Biomerieux->BMX={remap_counts['Biomerieux->BMX']}"
    )
    print(f"Supplier Expenses-Format match rate: {supplier_matched}/{len(rows)} lines")
    print(f"Supplier allocations filled/corroborated: {supplier_filled}/{supplier_corroborated}")
    print(f"Additional Info coverage emp_no+JQ: {addl_full}/{len(rows)} rows")
    print(f"Location=20100 rows: {location_ok}/{len(rows)}")
    print(f"Solution resolved rows: {solution_resolved}/{len(rows)}")
    print(f"Need-to-allocate/charge-to RED rows: {need_allocate_red}")
    print("Allocation source distribution: " + ", ".join(f"{k}={v}" for k, v in sorted(source_counts.items())))
    print("Agency resolve-method distribution: " + ", ".join(f"{k}={v}" for k, v in sorted(agency_method_counts.items())))
    print(f"Agency unresolved rows: {agency_unresolved}/{len(rows)}")
    print(f"Supplier Agency text still unmatched to lookup: {supplier_agency_unresolved}/{len(rows)}")

    print("\nLEGACY COMPARISON VS ENTRY SHEETS (NON-AUTHORITATIVE)")
    print("====================================================")
    comp = validation["comparable_rows"]
    print(f"Comparable distribution rows: {comp}")
    print(f"Agency:      {validation['agency_hits']}/{comp} = {validation['agency_rate']:.1%}")
    print(f"Agency delta vs v3 baseline 78.3%: {validation['agency_rate'] - 0.783:+.1%}")
    print(f"Cost Center: {validation['cost_center_hits']}/{comp} = {validation['cost_center_rate']:.1%}")
    print(f"Division:    {validation['division_hits']}/{comp} = {validation['division_rate']:.1%}")
    print(f"All 3:       {validation['all_segment_hits']}/{comp} = {validation['all_segment_rate']:.1%}")
    print(f"Full combo:  {validation['distribution_hits']}/{comp} = {validation['distribution_rate']:.1%}")
    print(f"Amount:      {validation['amount_hits']}/{comp} = {validation['amount_rate']:.1%}")
    if validation.get("source_breakdown"):
        print("\nPer-source hit rates:")
        for source, stats in sorted(validation["source_breakdown"].items()):
            total = stats["lines"] or 1
            print(
                f"- {source}: lines={stats['lines']} | "
                f"Agency={stats['agency']}/{stats['lines']} {stats['agency'] / total:.1%} | "
                f"CC={stats['cost_center']}/{stats['lines']} {stats['cost_center'] / total:.1%} | "
                f"DIV={stats['division']}/{stats['lines']} {stats['division'] / total:.1%} | "
                f"Combo={stats['distribution']}/{stats['lines']} {stats['distribution'] / total:.1%} | "
                f"Amount={stats['amount']}/{stats['lines']} {stats['amount'] / total:.1%}"
            )
    misses = []
    for d in validation["details"]:
        for m in d["matches"]:
            if m.get("no_answer_key"):
                misses.append(f"{d['invoice_no']} line {m.get('line_no')}: no answer row")
            elif not (m.get("agency_match") and m.get("cost_center_match") and m.get("division_match")):
                misses.append(
                    f"{d['invoice_no']} line {m.get('line_no')}: "
                    f"POC {m.get('poc_agency')}/{m.get('poc_division')}/{m.get('poc_cost_center')} vs "
                    f"Entry {m.get('answer_agency')}/{m.get('answer_division')}/{m.get('answer_cost_center')}"
                )
    if misses:
        print("\nMisses / review reasons:")
        for m in misses[:30]:
            print(f"- {m}")

    errored = [
        (_code(r.get("invoice_no"), 5), r.get("notes"))
        for r in rows
        if r.get("Row_Status") == "RED" and r.get("_amount_basis") == "error_placeholder"
    ]
    print("\nEXTRACTION ERRORS")
    print("=================")
    if errored:
        for inv, reason in errored:
            print(f"- {inv}: {reason}")
    else:
        print("None")

    print("\nHEADER VALIDATION A..AF")
    print("=======================")
    if header_diffs:
        for diff in header_diffs:
            print(f"- {diff}")
    else:
        print("A..AF header row strings match the reference character-for-character.")

    print("\nOUTPUT FILES")
    print("============")
    print(f"XLSX: {xlsx_path}")
    print(f"JSON: {json_path}")

    print("\nAssumptions:")
    print(f"- Location is hard-set to {DEFAULT_LOCATION} for every Asateel row.")
    print("- Supplier Number and Supplier Site are unknown in the cached extraction, so they are blank.")
    print(f"- Account {GL_ACCOUNT} is used as requested; fallback GL description is '{GL_FALLBACK_DESC}' if absent from the Account lookup.")
    print("- RED/YELLOW unresolved allocations keep Cost Center, DIV, and Agency Name blank; Agency segment defaults to 00000.")


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Asateel POC Oracle-Fusion workbook generator")
    parser.add_argument("--folder", choices=["PROJECTS", "ADMIN", "CENTRAL", "ALL"], help="Folder label to process")
    parser.add_argument("--pdf-dir", default="", help="Optional directory of *_0001.pdf files to process with the folder label")
    parser.add_argument("--full", action="store_true", help="Enumerate every *_0001.pdf in the selected folder")
    parser.add_argument("--out-suffix", default="", help="Optional label appended to output filenames")
    parser.add_argument("--refresh-cache", action="store_true", help="Force fresh Gemini extraction and overwrite cache")
    parser.add_argument(
        "--expenses-format",
        default=str(DEFAULT_EXPENSES_FORMAT_XLSX),
        help="Supplier Expenses Format workbook used for employee/JQ and allocation enrichment",
    )
    parser.add_argument(
        "--so-detail",
        default=str(DEFAULT_SO_DETAIL_XLSX),
        help="SO_Detail export used only to validate canonical JQ existence",
    )
    parser.add_argument(
        "--allocation-mode",
        choices=[STANDARD_ALLOCATION_MODE, PROJECT_ALLOCATION_MODE],
        default=STANDARD_ALLOCATION_MODE,
        help="Separate Labadi override mode; applies only to PROJECTS invoices",
    )
    parser.add_argument(
        "--project-allocation-lookup",
        default=str(DEFAULT_PROJECT_ALLOCATION_LOOKUP),
        help="Normalized deterministic lookup for projects-labadi-v1 mode",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv if argv is not None else sys.argv[1:])
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    lookups = load_lookups()
    supplier_index = load_expenses_format(Path(args.expenses_format), lookups)
    so_detail_index = load_so_detail(Path(args.so_detail))
    project_lookup = None
    if args.allocation_mode == PROJECT_ALLOCATION_MODE:
        if args.folder not in {"PROJECTS", "ALL"}:
            raise ValueError("projects-labadi-v1 mode requires --folder PROJECTS or ALL")
        if str(ROOT) not in sys.path:
            sys.path.insert(0, str(ROOT))
        from scripts.asateel_project_allocation import load_lookup
        project_lookup = load_lookup(Path(args.project_allocation_lookup))
    extracted = []
    cache_tag = pdf_dir_cache_tag(args.pdf_dir)
    files = folder_files(args.folder, args.full, args.pdf_dir)
    for idx, (folder, inv, pdf_path) in enumerate(files, start=1):
        print(f"[extract {idx}/{len(files)}] {folder} {inv}", flush=True)
        try:
            payload = gemini_extract(pdf_path, folder, inv, lookups, force=args.refresh_cache, cache_tag=cache_tag)
            extracted.append({"folder": folder, "invoice_hint": inv, "pdf": str(pdf_path), "payload": payload})
        except Exception as exc:
            print(f"[extract error] {folder} {inv}: {exc}", flush=True)
            extracted.append({
                "folder": folder,
                "invoice_hint": inv,
                "pdf": str(pdf_path),
                "payload": {},
                "error": str(exc),
            })

    rows, trace = build_rows(
        extracted,
        lookups,
        supplier_index,
        so_detail_index,
        allocation_mode=args.allocation_mode,
        project_lookup=project_lookup,
        project_master_fallback=args.folder == "PROJECTS",
    )
    trace["supplier_expenses_format_path"] = str(Path(args.expenses_format))
    trace["so_detail_path"] = str(Path(args.so_detail))
    trace["project_allocation_lookup_path"] = (
        str(Path(args.project_allocation_lookup)) if project_lookup else ""
    )
    validation = validate(rows, lookups)
    trace["validation"] = validation

    stamp = "2026-06-23"
    suffix = f"-{args.out_suffix}" if args.out_suffix else ""
    xlsx_path = OUT_DIR / f"asateel-poc-oracle{suffix}-{stamp}.xlsx"
    json_path = OUT_DIR / f"asateel-poc-trace{suffix}-{stamp}.json"
    write_excel(rows, xlsx_path)
    header_diffs = validate_output_headers(xlsx_path)
    trace["header_validation"] = {
        "a_to_af_match_reference": not header_diffs,
        "diffs": header_diffs,
    }
    json_path.write_text(json.dumps(trace, ensure_ascii=False, indent=2, default=str), encoding="utf-8")

    print_summary(rows, validation, header_diffs, xlsx_path, json_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
