#!/usr/bin/env python3
"""Build an Excel diff report between Asateel v3 output and Finance entries."""

from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUR_PATH = ROOT / "asateel-sample/_poc_out/asateel-poc-oracle-CENTRAL-full-v3-2026-06-20.xlsx"
ENTRY_PATHS = [
    ROOT / "asateel-sample/_allocation/Entry-1.xlsm",
    ROOT / "asateel-sample/_allocation/Entry-2.xlsm",
]
SUPPLIER_PATH = ROOT / "asateel-sample/_allocation/Central-11-2026.xlsx"
MANPOWER_PATH = ROOT / "qc/master-data/Aljeel_Lookups-v2.xlsx"
OUT_PATH = ROOT / "asateel-sample/_poc_out/asateel-diff-vs-finance-v3-2026-06-20.xlsx"


HEADER_FILL = "1E40AF"
HEADER_FONT = "FFFFFF"
BORDER_COLOR = "E5E7EB"
MISMATCH_FILL = "F9D2D2"  # #DC2626 at roughly 15% tint on white.
GROUP_FILL = "DBEAFE"


FIELDS = ["Agency", "CC", "DIV", "Solution", "Combo"]


@dataclass
class OurLine:
    invoice: str
    line_no: int
    description: str
    amount: Decimal | None
    combo: str
    employee_no: str
    allocation_source: str
    agency: str
    agency_name: str
    cc: str
    div: str
    solution: str
    extracted_brands: str
    notes: str


@dataclass
class KeyLine:
    invoice: str
    line_no: int
    amount: Decimal | None
    combo: str
    agency: str
    cc: str
    div: str
    solution: str


@dataclass
class SupplierLine:
    invoice: str
    line_no: int
    amount: Decimal | None
    jq: str
    employee_name: str
    employee_no: str


@dataclass
class Match:
    our: OurLine
    key: KeyLine | None
    supplier: SupplierLine | None
    mp: dict[str, str]
    mismatch_fields: list[str]
    home_agency_vs_brand: str


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def norm_code(value: Any, width: int | None = None) -> str:
    text = clean(value)
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    if width and text.isdigit():
        return text.zfill(width)
    return text


def norm_invoice(value: Any) -> str:
    return norm_code(value, 5)


def norm_name(value: Any) -> str:
    return " ".join(clean(value).lower().split())


def to_decimal(value: Any) -> Decimal | None:
    text = clean(value).replace(",", "")
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def amount_close(a: Decimal | None, b: Decimal | None) -> bool:
    if a is None or b is None:
        return False
    return abs(a - b) <= Decimal("0.01")


def split_combo(combo: str) -> dict[str, str]:
    parts = clean(combo).split("-")
    return {
        "cc": parts[3] if len(parts) > 3 else "",
        "div": parts[4] if len(parts) > 4 else "",
        "solution": parts[5] if len(parts) > 5 else "",
        "agency": parts[6] if len(parts) > 6 else "",
    }


def header_map(ws, header_row: int) -> dict[str, int]:
    return {clean(ws.cell(header_row, c).value): c for c in range(1, ws.max_column + 1) if clean(ws.cell(header_row, c).value)}


def load_our_lines() -> list[OurLine]:
    wb = load_workbook(OUR_PATH, read_only=False, data_only=True)
    ws = wb["Sheet"]
    headers = header_map(ws, 3)
    rows: list[OurLine] = []
    invoice_counts: Counter[str] = Counter()
    for r in range(4, ws.max_row + 1):
        invoice = norm_invoice(ws.cell(r, 3).value)
        combo = clean(ws.cell(r, 14).value)
        amount = to_decimal(ws.cell(r, 13).value)
        if not invoice or not combo:
            continue
        invoice_counts[invoice] += 1
        rows.append(
            OurLine(
                invoice=invoice,
                line_no=invoice_counts[invoice],
                description=clean(ws.cell(r, 11).value),
                amount=amount,
                combo=combo,
                employee_no=norm_code(ws.cell(r, 16).value),
                allocation_source=clean(ws.cell(r, headers.get("Allocation Source", 34)).value),
                agency=norm_code(ws.cell(r, 27).value, 5),
                agency_name=clean(ws.cell(r, 28).value),
                cc=norm_code(ws.cell(r, 21).value),
                div=norm_code(ws.cell(r, 23).value),
                solution=norm_code(ws.cell(r, 25).value, 5),
                extracted_brands=clean(ws.cell(r, headers.get("Extracted Brand(s)", 36)).value),
                notes=clean(ws.cell(r, headers.get("Notes", 45)).value),
            )
        )
    wb.close()
    return rows


def load_key_lines() -> dict[str, list[KeyLine]]:
    by_invoice: dict[str, list[KeyLine]] = defaultdict(list)
    for path in ENTRY_PATHS:
        wb = load_workbook(path, read_only=False, data_only=True, keep_vba=True)
        ws = wb["Invoices"]
        headers = header_map(ws, 8)
        invoice_col = headers.get("*Invoice Number", 8)
        amount_col = 83
        combo_col = 99
        current_invoice = ""
        for r in range(9, ws.max_row + 1):
            raw_invoice = ws.cell(r, invoice_col).value
            if clean(raw_invoice):
                current_invoice = norm_invoice(raw_invoice)
            combo = clean(ws.cell(r, combo_col).value)
            if not current_invoice or not combo:
                continue
            parts = split_combo(combo)
            line_no = len(by_invoice[current_invoice]) + 1
            by_invoice[current_invoice].append(
                KeyLine(
                    invoice=current_invoice,
                    line_no=line_no,
                    amount=to_decimal(ws.cell(r, amount_col).value),
                    combo=combo,
                    agency=norm_code(parts["agency"], 5),
                    cc=norm_code(parts["cc"]),
                    div=norm_code(parts["div"]),
                    solution=norm_code(parts["solution"], 5),
                )
            )
        wb.close()
    return by_invoice


def load_supplier_lines() -> dict[str, list[SupplierLine]]:
    wb = load_workbook(SUPPLIER_PATH, read_only=True, data_only=True)
    ws = wb["Expenses Format"]
    by_invoice: dict[str, list[SupplierLine]] = defaultdict(list)
    current_invoice = ""
    for row in ws.iter_rows(min_row=9, max_col=41, values_only=True):
        raw_invoice = row[13]
        if clean(raw_invoice):
            current_invoice = norm_invoice(raw_invoice)
        if not current_invoice:
            continue
        has_line_data = any(clean(row[c - 1]) for c in (24, 25, 37, 41))
        if not has_line_data:
            continue
        line_no = len(by_invoice[current_invoice]) + 1
        by_invoice[current_invoice].append(
            SupplierLine(
                invoice=current_invoice,
                line_no=line_no,
                amount=to_decimal(row[36]),
                jq=clean(row[23]),
                employee_name=clean(row[24]),
                employee_no=norm_code(row[40]),
            )
        )
    wb.close()
    return by_invoice


def load_manpower() -> tuple[dict[str, dict[str, str]], dict[str, dict[str, str]]]:
    wb = load_workbook(MANPOWER_PATH, read_only=True, data_only=True)
    ws = wb["Manpower"]
    by_emp: dict[str, dict[str, str]] = {}
    by_name: dict[str, dict[str, str]] = {}
    for row_values in ws.iter_rows(min_row=2, max_col=16, values_only=True):
        row = {
            "emp_no": norm_code(row_values[0]),
            "name": clean(row_values[2]),
            "div": norm_code(row_values[8]),
            "new_division": clean(row_values[9]),
            "agency_code": norm_code(row_values[10], 5),
            "agency_name": clean(row_values[11]),
            "cc": norm_code(row_values[12]),
            "cc_name": clean(row_values[13]),
            "alloc_flag": clean(row_values[14]),
            "solution": norm_code(row_values[15], 5),
        }
        if row["emp_no"]:
            by_emp[row["emp_no"]] = row
        if row["name"]:
            by_name[norm_name(row["name"])] = row
    wb.close()
    return by_emp, by_name


def pick_by_amount(lines: list[Any], used: set[int], amount: Decimal | None) -> int | None:
    for idx, line in enumerate(lines):
        if idx not in used and amount_close(amount, line.amount):
            return idx
    return None


def build_matches() -> list[Match]:
    our_lines = load_our_lines()
    key_by_invoice = load_key_lines()
    supplier_by_invoice = load_supplier_lines()
    mp_by_emp, mp_by_name = load_manpower()

    used_key: dict[str, set[int]] = defaultdict(set)
    used_supplier: dict[str, set[int]] = defaultdict(set)
    matches: list[Match] = []

    for our in our_lines:
        key_lines = key_by_invoice.get(our.invoice, [])
        key_idx: int | None = None
        if our.line_no - 1 < len(key_lines) and our.line_no - 1 not in used_key[our.invoice]:
            key_idx = our.line_no - 1
        else:
            key_idx = pick_by_amount(key_lines, used_key[our.invoice], our.amount)
        key = key_lines[key_idx] if key_idx is not None else None
        if key_idx is not None:
            used_key[our.invoice].add(key_idx)

        supplier_lines = supplier_by_invoice.get(our.invoice, [])
        sup_idx: int | None = None
        if our.line_no - 1 < len(supplier_lines) and our.line_no - 1 not in used_supplier[our.invoice]:
            sup_idx = our.line_no - 1
        else:
            sup_idx = pick_by_amount(supplier_lines, used_supplier[our.invoice], our.amount)
        supplier = supplier_lines[sup_idx] if sup_idx is not None else None
        if sup_idx is not None:
            used_supplier[our.invoice].add(sup_idx)

        lookup_emp = supplier.employee_no if supplier and supplier.employee_no else our.employee_no
        lookup_name = supplier.employee_name if supplier else ""
        mp = mp_by_emp.get(lookup_emp, {}) if lookup_emp else {}
        if not mp and lookup_name:
            mp = mp_by_name.get(norm_name(lookup_name), {})

        mismatches: list[str] = []
        if key is None:
            mismatches = FIELDS.copy()
        else:
            comparisons = {
                "Agency": our.agency == key.agency,
                "CC": our.cc == key.cc,
                "DIV": our.div == key.div,
                "Solution": our.solution == key.solution,
                "Combo": our.combo == key.combo,
            }
            mismatches = [field for field, ok in comparisons.items() if not ok]

        home_agency_vs_brand = ""
        if key and mp and mp.get("agency_code") == our.agency and our.agency != key.agency:
            home_agency_vs_brand = "Y"

        matches.append(Match(our=our, key=key, supplier=supplier, mp=mp, mismatch_fields=mismatches, home_agency_vs_brand=home_agency_vs_brand))
    return matches


def yes_no(value: bool) -> str:
    return "Y" if value else "N"


def row_values(match: Match) -> list[Any]:
    our = match.our
    key = match.key
    supplier = match.supplier
    mp = match.mp
    employee_no = supplier.employee_no if supplier and supplier.employee_no else our.employee_no
    employee_name = supplier.employee_name if supplier else ""
    field_set = set(match.mismatch_fields)
    return [
        our.invoice,
        our.line_no,
        our.description,
        float(our.amount) if our.amount is not None else None,
        our.allocation_source,
        employee_no,
        employee_name,
        yes_no(bool(mp)),
        mp.get("name", ""),
        mp.get("agency_code", ""),
        mp.get("agency_name", ""),
        mp.get("div", ""),
        mp.get("new_division", ""),
        mp.get("cc", ""),
        mp.get("cc_name", ""),
        mp.get("solution", ""),
        mp.get("alloc_flag", ""),
        match.home_agency_vs_brand,
        our.agency,
        our.agency_name,
        our.cc,
        our.div,
        our.solution,
        our.combo,
        key.agency if key else "",
        key.cc if key else "",
        key.div if key else "",
        key.solution if key else "",
        key.combo if key else "",
        yes_no("Agency" not in field_set),
        yes_no("CC" not in field_set),
        yes_no("DIV" not in field_set),
        yes_no("Solution" not in field_set),
        yes_no("Combo" not in field_set),
        ", ".join(match.mismatch_fields),
        our.extracted_brands,
        our.notes,
    ]


DETAIL_HEADERS = [
    "Invoice",
    "Line",
    "Description",
    "Amount",
    "Allocation Source",
    "Employee No",
    "Employee Name",
    "MP Emp Found(Y/N)",
    "MP Name",
    "MP Agency Code",
    "MP Agency Name",
    "MP DIV",
    "MP New Division",
    "MP Cost Center",
    "MP CC Name",
    "MP Solution",
    "MP Alloc Flag",
    "HomeAgencyVsBrand",
    "Our Agency",
    "Our Agency Name",
    "Our CC",
    "Our DIV",
    "Our Solution",
    "Our Full Combo",
    "Key Agency",
    "Key CC",
    "Key DIV",
    "Key Solution",
    "Key Full Combo",
    "Agency Match",
    "CC Match",
    "DIV Match",
    "Solution Match",
    "Combo Match",
    "Mismatch Fields",
    "Extracted Brand(s)",
    "Notes",
]


def add_detail_sheet(wb: Workbook, name: str, matches: list[Match]) -> None:
    ws = wb.create_sheet(name)
    ws.append(DETAIL_HEADERS)
    for match in matches:
        ws.append(row_values(match))

    style_sheet(ws, freeze="A2", auto_filter=True)
    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["G"].width = 28
    ws.column_dimensions["I"].width = 28
    ws.column_dimensions["K"].width = 22
    ws.column_dimensions["M"].width = 20
    ws.column_dimensions["O"].width = 24
    ws.column_dimensions["X"].width = 58
    ws.column_dimensions["AC"].width = 58
    ws.column_dimensions["AI"].width = 28
    ws.column_dimensions["AJ"].width = 24
    ws.column_dimensions["AK"].width = 90

    red_fill = PatternFill("solid", fgColor=MISMATCH_FILL)
    our_cols = {"Agency": 19, "CC": 21, "DIV": 22, "Solution": 23, "Combo": 24}
    for row_idx in range(2, ws.max_row + 1):
        mismatch_text = clean(ws.cell(row_idx, 35).value)
        mismatches = {part.strip() for part in mismatch_text.split(",") if part.strip()}
        for field, col in our_cols.items():
            if field in mismatches:
                ws.cell(row_idx, col).fill = red_fill


def style_sheet(ws, freeze: str | None = None, auto_filter: bool = False) -> None:
    thin = Side(style="thin", color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name="Inter", size=10)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(name="Inter", size=10, bold=True, color=HEADER_FONT)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if freeze:
        ws.freeze_panes = freeze
    if auto_filter and ws.max_row > 1 and ws.max_column > 1:
        ws.auto_filter.ref = ws.dimensions


def add_summary_sheet(wb: Workbook, matches: list[Match], differences: list[Match]) -> None:
    ws = wb.create_sheet("Summary", 0)
    total = len(matches)
    matched_lines = [m for m in matches if m.key is not None]
    matched = len(matched_lines)
    field_counts = {field: sum(1 for m in matched_lines if field in m.mismatch_fields) for field in FIELDS}
    no_mp = sum(1 for m in matched_lines if not m.mp)
    home_brand = sum(1 for m in differences if m.home_agency_vs_brand == "Y")

    rows = [
        ["Metric", "Value"],
        ["Total lines", total],
        ["Matched lines", matched],
        ["Differences sheet rows", len(differences)],
        ["Agency disagreements", field_counts["Agency"]],
        ["CC disagreements", field_counts["CC"]],
        ["DIV disagreements", field_counts["DIV"]],
        ["Solution disagreements", field_counts["Solution"]],
        ["Full-combo disagreements", field_counts["Combo"]],
        ["HomeAgencyVsBrand differences", home_brand],
        ["Lines with no Manpower match", no_mp],
        ["Amount-split note", "Amount-split is intentionally excluded from the diff per operator decision; Amount is shown for context."],
        [],
        ["Top Our Agency -> Key Agency substitutions", "Count", "Example Invoice"],
    ]
    for row in rows:
        ws.append(row)

    agency_subs: dict[tuple[str, str], list[str]] = defaultdict(list)
    for match in matches:
        if match.key and match.our.agency != match.key.agency:
            agency_subs[(match.our.agency, match.key.agency)].append(match.our.invoice)
    for (our_agency, key_agency), invoices in sorted(agency_subs.items(), key=lambda item: (-len(item[1]), item[0]))[:20]:
        ws.append([f"{our_agency} -> {key_agency}", len(invoices), invoices[0]])

    style_sheet(ws)
    for cell in ws[14]:
        cell.fill = PatternFill("solid", fgColor=GROUP_FILL)
        cell.font = Font(name="Inter", size=10, bold=True, color="111827")
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18


def write_workbook(matches: list[Match]) -> None:
    matched_lines = [m for m in matches if m.key is not None]
    differences = [m for m in matched_lines if m.mismatch_fields]
    differences.sort(key=lambda m: (-len(m.mismatch_fields), m.our.invoice, m.our.line_no))
    all_lines = sorted(matched_lines, key=lambda m: (m.our.invoice, m.our.line_no))

    wb = Workbook()
    wb.remove(wb.active)
    add_summary_sheet(wb, matches, differences)
    add_detail_sheet(wb, "Differences", differences)
    add_detail_sheet(wb, "All Lines", all_lines)
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)


def print_report(matches: list[Match]) -> None:
    matched_lines = [m for m in matches if m.key is not None]
    differences = [m for m in matched_lines if m.mismatch_fields]
    field_counts = {field: sum(1 for m in matched_lines if field in m.mismatch_fields) for field in FIELDS}
    agency_subs: dict[tuple[str, str], list[str]] = defaultdict(list)
    for match in matched_lines:
        if match.key and match.our.agency != match.key.agency:
            agency_subs[(match.our.agency, match.key.agency)].append(match.our.invoice)
    home_brand = sum(1 for m in differences if m.home_agency_vs_brand == "Y")
    no_mp = sum(1 for m in matched_lines if not m.mp)

    print(f"Differences sheet rows: {len(differences)}")
    print(f"Total OUR lines: {len(matches)}")
    print(f"Matched lines: {len(matched_lines)}")
    print("Breakdown by field:")
    for field in FIELDS:
        print(f"  {field}: {field_counts[field]}")
    print("Top 10 agency substitutions (ours -> booked):")
    for (our_agency, key_agency), invoices in sorted(agency_subs.items(), key=lambda item: (-len(item[1]), item[0]))[:10]:
        print(f"  {our_agency} -> {key_agency}: {len(invoices)} (example {invoices[0]})")
    print(f"HomeAgencyVsBrand diffs: {home_brand}")
    print(f"Lines with no Manpower match: {no_mp}")
    print(f"Final xlsx path: {OUT_PATH}")


def main() -> None:
    matches = build_matches()
    write_workbook(matches)
    print_report(matches)


if __name__ == "__main__":
    main()
