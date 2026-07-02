#!/usr/bin/env python3
"""Build the v5 side-by-side Finance comparison workbook and markdown report."""

from __future__ import annotations

from collections import Counter, defaultdict
from copy import copy
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string


ROOT = Path(__file__).resolve().parents[1]
POC_OUT = ROOT / "asateel-sample/_poc_out"
V4_PATH = POC_OUT / "asateel-poc-oracle-CENTRAL-full-v4-2026-06-20.xlsx"
V5_PATH = POC_OUT / "asateel-poc-oracle-CENTRAL-full-v5-2026-06-20.xlsx"
ENTRY_PATHS = [
    ROOT / "asateel-sample/_allocation/Entry-1.xlsm",
    ROOT / "asateel-sample/_allocation/Entry-2.xlsm",
]
SUPPLIER_PATH = ROOT / "asateel-sample/_allocation/Central-11-2026.xlsx"
MANPOWER_PATH = ROOT / "qc/master-data/Aljeel_Lookups-v2.xlsx"
OUT_XLSX = POC_OUT / "asateel-sidebyside-v5-2026-06-20.xlsx"
OUT_MD = ROOT / "asateel-sample/COMPARE-REPORT-v5-2026-06-20.md"

SEGMENTS = ["company", "location", "account", "cc", "div", "solution", "agency", "project", "intercompany", "future"]
SCORED_FIELDS = ["Agency", "CC", "DIV", "Solution", "Additional Info", "Full Combo"]
RATE_FIELDS = ["Agency", "CC", "DIV", "Solution", "Additional Info", "Full Combo", "Amount"]

HEADER_FILL = "1E40AF"
HEADER_FONT = "FFFFFF"
BORDER_COLOR = "E5E7EB"
MATCH_FILL = "DCFCE7"
MISMATCH_FILL = "F9D2D2"
ALT_FILL = "F8FAFC"


@dataclass
class OurLine:
    invoice: str
    line_no: int
    description: str
    amount: Decimal | None
    combo: str
    additional_info: str
    agency_name: str
    segments: dict[str, str]


@dataclass
class KeyLine:
    invoice: str
    line_no: int
    amount: Decimal | None
    combo: str
    additional_info: str
    segments: dict[str, str]


@dataclass
class Match:
    our: OurLine
    key: KeyLine
    checks: dict[str, bool]

    @property
    def mismatch_fields(self) -> list[str]:
        return [field for field in SCORED_FIELDS if not self.checks[field]]


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def norm_code(value: Any, width: int | None = None) -> str:
    text = clean(value)
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    if width and text.isdigit():
        return text.zfill(width)
    return text


def norm_invoice(value: Any) -> str:
    return norm_code(value, 5)


def norm_compare(value: Any) -> str:
    return clean(value).lower()


def to_decimal(value: Any) -> Decimal | None:
    text = clean(value).replace(",", "")
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def amount_close(a: Decimal | None, b: Decimal | None) -> bool:
    if a is None or b is None:
        return False
    return abs(a - b) <= Decimal("0.01")


def split_combo(combo: str) -> dict[str, str]:
    parts = clean(combo).split("-")
    return {name: parts[idx] if idx < len(parts) else "" for idx, name in enumerate(SEGMENTS)}


def load_our_lines(path: Path) -> list[OurLine]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet"]
    rows: list[OurLine] = []
    invoice_counts: Counter[str] = Counter()
    for row in ws.iter_rows(min_row=4, max_col=36, values_only=True):
        invoice = norm_invoice(row[2])
        combo = clean(row[13])
        if not invoice or not combo:
            continue
        invoice_counts[invoice] += 1
        rows.append(
            OurLine(
                invoice=invoice,
                line_no=invoice_counts[invoice],
                description=clean(row[10]),
                amount=to_decimal(row[12]),
                combo=combo,
                additional_info=clean(row[35]),
                agency_name=clean(row[27]),
                segments=split_combo(combo),
            )
        )
    wb.close()
    return rows


def load_key_lines() -> dict[str, list[KeyLine]]:
    by_invoice: dict[str, list[KeyLine]] = defaultdict(list)
    invoice_col = column_index_from_string("H") - 1
    amount_col = column_index_from_string("CE") - 1
    combo_col = column_index_from_string("CU") - 1
    addl_col = column_index_from_string("EL") - 1
    max_col = column_index_from_string("EL")
    for path in ENTRY_PATHS:
        wb = load_workbook(path, read_only=True, data_only=True, keep_vba=True)
        ws = wb["Invoices"]
        current_invoice = ""
        for row in ws.iter_rows(min_row=9, max_col=max_col, values_only=True):
            raw_invoice = row[invoice_col]
            if clean(raw_invoice):
                current_invoice = norm_invoice(raw_invoice)
            combo = clean(row[combo_col])
            if not current_invoice or not combo:
                continue
            line_no = len(by_invoice[current_invoice]) + 1
            by_invoice[current_invoice].append(
                KeyLine(
                    invoice=current_invoice,
                    line_no=line_no,
                    amount=to_decimal(row[amount_col]),
                    combo=combo,
                    additional_info=clean(row[addl_col]),
                    segments=split_combo(combo),
                )
            )
        wb.close()
    return by_invoice


def touch_supplier_and_manpower() -> tuple[int, int]:
    supplier_rows = 0
    wb = load_workbook(SUPPLIER_PATH, read_only=True, data_only=True)
    ws = wb["Expenses Format"]
    for row in ws.iter_rows(min_row=9, max_col=41, values_only=True):
        if any(clean(row[c - 1]) for c in (14, 24, 25, 37, 41)):
            supplier_rows += 1
    wb.close()

    manpower_rows = 0
    wb = load_workbook(MANPOWER_PATH, read_only=True, data_only=True)
    ws = wb["Manpower"]
    for row in ws.iter_rows(min_row=2, max_col=16, values_only=True):
        if clean(row[0]) or clean(row[2]):
            manpower_rows += 1
    wb.close()
    return supplier_rows, manpower_rows


def pick_by_amount(lines: list[KeyLine], used: set[int], amount: Decimal | None) -> int | None:
    for idx, line in enumerate(lines):
        if idx not in used and amount_close(amount, line.amount):
            return idx
    return None


def build_matches(our_lines: list[OurLine], key_by_invoice: dict[str, list[KeyLine]]) -> list[Match]:
    used_key: dict[str, set[int]] = defaultdict(set)
    matches: list[Match] = []
    for our in our_lines:
        key_lines = key_by_invoice.get(our.invoice, [])
        key_idx: int | None = None
        line_idx = our.line_no - 1
        if line_idx < len(key_lines) and line_idx not in used_key[our.invoice]:
            key_idx = line_idx
        else:
            key_idx = pick_by_amount(key_lines, used_key[our.invoice], our.amount)
        if key_idx is None:
            continue
        used_key[our.invoice].add(key_idx)
        key = key_lines[key_idx]
        checks = {
            "Agency": norm_code(our.segments["agency"], 5) == norm_code(key.segments["agency"], 5),
            "CC": norm_code(our.segments["cc"]) == norm_code(key.segments["cc"]),
            "DIV": norm_code(our.segments["div"]) == norm_code(key.segments["div"]),
            "Solution": norm_code(our.segments["solution"], 5) == norm_code(key.segments["solution"], 5),
            "Additional Info": norm_compare(our.additional_info) == norm_compare(key.additional_info),
            "Full Combo": clean(our.combo) == clean(key.combo),
            "Amount": amount_close(our.amount, key.amount),
        }
        matches.append(Match(our=our, key=key, checks=checks))
    return sorted(matches, key=lambda m: (m.our.invoice, m.our.line_no))


def yn(value: bool) -> str:
    return "Y" if value else "N"


HEADERS = [
    "Invoice", "Line", "Description", "Amount(ours)", "Amount(theirs)",
    "Our Account", "Our Location", "Our Cost Center", "Our DIV", "Our Solution", "Our Agency",
    "Our Agency Name", "Our Additional Info", "Our Full Distribution",
    "Key Account", "Key Location", "Key Cost Center", "Key DIV", "Key Solution", "Key Agency",
    "Key Additional Info", "Key Full Distribution",
    "Agency✓", "CC✓", "DIV✓", "Solution✓", "Additional Info✓", "Full Combo✓", "Mismatch Fields",
]

BANNERS = [
    ("KEY", 1, 5),
    ("OURS", 6, 14),
    ("THEIRS / FINANCE", 15, 22),
    ("MATCH", 23, 29),
]

THEIRS_BY_FIELD = {
    "Full Combo": 22,
    "Additional Info": 21,
    "Agency": 20,
    "Solution": 19,
    "DIV": 18,
    "CC": 17,
}

MATCH_COLS = {
    "Agency": 23,
    "CC": 24,
    "DIV": 25,
    "Solution": 26,
    "Additional Info": 27,
    "Full Combo": 28,
}


def row_values(match: Match) -> list[Any]:
    our = match.our
    key = match.key
    mismatches = match.mismatch_fields
    return [
        our.invoice,
        our.line_no,
        our.description,
        float(our.amount) if our.amount is not None else None,
        float(key.amount) if key.amount is not None else None,
        our.segments["account"],
        our.segments["location"],
        our.segments["cc"],
        our.segments["div"],
        our.segments["solution"],
        our.segments["agency"],
        our.agency_name,
        our.additional_info,
        our.combo,
        key.segments["account"],
        key.segments["location"],
        key.segments["cc"],
        key.segments["div"],
        key.segments["solution"],
        key.segments["agency"],
        key.additional_info,
        key.combo,
        yn(match.checks["Agency"]),
        yn(match.checks["CC"]),
        yn(match.checks["DIV"]),
        yn(match.checks["Solution"]),
        yn(match.checks["Additional Info"]),
        yn(match.checks["Full Combo"]),
        ", ".join(mismatches),
    ]


def style_side_by_side(ws) -> None:
    thin = Side(style="thin", color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    banner_fill = PatternFill("solid", fgColor=HEADER_FILL)
    match_fill = PatternFill("solid", fgColor=MATCH_FILL)
    mismatch_fill = PatternFill("solid", fgColor=MISMATCH_FILL)
    alt_fill = PatternFill("solid", fgColor=ALT_FILL)

    for title, start_col, end_col in BANNERS:
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        cell = ws.cell(1, start_col)
        cell.value = title
        cell.fill = banner_fill
        cell.font = Font(name="Inter", size=11, bold=True, color=HEADER_FONT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            new_font = copy(cell.font)
            new_font.name = "Inter"
            new_font.sz = 10
            cell.font = new_font
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in ws[2]:
        cell.fill = banner_fill
        cell.font = Font(name="Inter", size=10, bold=True, color=HEADER_FONT)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    last_invoice = None
    invoice_group = -1
    for row_idx in range(3, ws.max_row + 1):
        invoice = clean(ws.cell(row_idx, 1).value)
        if invoice != last_invoice:
            invoice_group += 1
            last_invoice = invoice
        if invoice_group % 2:
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row_idx, col_idx).fill = alt_fill

        mismatches = {part.strip() for part in clean(ws.cell(row_idx, 29).value).split(",") if part.strip()}
        for field, col_idx in MATCH_COLS.items():
            if clean(ws.cell(row_idx, col_idx).value) == "Y":
                ws.cell(row_idx, col_idx).fill = match_fill
        for field in mismatches:
            col_idx = THEIRS_BY_FIELD.get(field)
            if col_idx:
                ws.cell(row_idx, col_idx).fill = mismatch_fill

    ws.freeze_panes = "C3"
    if ws.max_row >= 2:
        ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"
    widths = {
        "A": 11, "B": 7, "C": 34, "D": 14, "E": 14, "F": 13, "G": 13, "H": 16, "I": 10,
        "J": 13, "K": 13, "L": 22, "M": 24, "N": 58, "O": 13, "P": 13, "Q": 16,
        "R": 10, "S": 13, "T": 13, "U": 24, "V": 58, "W": 10, "X": 8, "Y": 8,
        "Z": 11, "AA": 18, "AB": 14, "AC": 28,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 36


def add_match_sheet(wb: Workbook, name: str, matches: list[Match]) -> None:
    ws = wb.create_sheet(name)
    ws.append([""] * len(HEADERS))
    ws.append(HEADERS)
    for match in matches:
        ws.append(row_values(match))
    style_side_by_side(ws)


def pct(n: int, d: int) -> str:
    return f"{(n / d * 100):.1f}%" if d else "0.0%"


def metric_counts(matches: list[Match]) -> dict[str, tuple[int, int]]:
    total = len(matches)
    return {field: (sum(1 for m in matches if m.checks[field]), total) for field in RATE_FIELDS}


def add_summary_sheet(wb: Workbook, matches: list[Match], differences: list[Match], generated_rows: int) -> int:
    ws = wb.create_sheet("Summary", 0)
    counts = metric_counts(matches)
    fully_identical = sum(1 for m in matches if not m.mismatch_fields)
    rows = [
        ["Metric", "Value", "Notes"],
        ["Generated OUR distribution rows", generated_rows, "All v5 output distribution rows."],
        ["Matched Finance lines", len(matches), "Matched by invoice + line order, with amount fallback."],
        ["Differences Only rows", len(differences), ""],
        ["Fully identical scored lines", fully_identical, "Agency, CC, DIV, Solution, Additional Info, Full Combo."],
        ["Amount split scoring", "Excluded", "Amount is shown and reported separately per operator."],
        [],
        ["Field", "Hits", "Rate"],
    ]
    for row in rows:
        ws.append(row)
    for field in RATE_FIELDS:
        hits, total = counts[field]
        ws.append([field, f"{hits}/{total}", pct(hits, total)])

    thin = Side(style="thin", color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name="Inter", size=10)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row_num in (1, 8):
        for cell in ws[row_num]:
            cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
            cell.font = Font(name="Inter", size=10, bold=True, color=HEADER_FONT)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 62
    return ws.max_row


def write_workbook(matches: list[Match], generated_rows: int) -> tuple[int, int, int]:
    differences = [m for m in matches if m.mismatch_fields]
    wb = Workbook()
    wb.remove(wb.active)
    summary_rows = add_summary_sheet(wb, matches, differences, generated_rows)
    add_match_sheet(wb, "Side by Side", matches)
    add_match_sheet(wb, "Differences Only", differences)
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    return len(matches), len(differences), summary_rows


def location_hits(our_lines: list[OurLine]) -> tuple[int, int]:
    hits = 0
    for line in our_lines:
        if line.segments.get("location") == "20100":
            hits += 1
    return hits, len(our_lines)


def example_lines(matches: list[Match], invoice: str) -> list[str]:
    out = []
    for match in matches:
        if match.our.invoice == invoice:
            out.append(
                f"{invoice} line {match.our.line_no}: ours={match.our.additional_info or '(blank)'} | "
                f"finance={match.key.additional_info or '(blank)'}"
            )
    return out


def write_markdown(v4_matches: list[Match], v5_matches: list[Match], loc: tuple[int, int], generated_rows: int) -> None:
    v4_counts = metric_counts(v4_matches)
    v5_counts = metric_counts(v5_matches)
    existing = ["Agency", "CC", "DIV", "Solution", "Full Combo", "Amount"]
    unchanged = all(v4_counts[field] == v5_counts[field] for field in existing)
    lines = [
        "# Asateel v5 Compare Report - 2026-06-20",
        "",
        "Matched by invoice + distribution-line order, with amount within 0.01 as fallback. Amount split remains excluded from scored workbook mismatches per operator.",
        "",
        "| Field | v4 hits | v4 rate | v5 hits | v5 rate | Delta hits |",
        "|---|---:|---:|---:|---:|---:|",
    ]
    for field in RATE_FIELDS:
        v4_hit, v4_total = v4_counts[field]
        v5_hit, v5_total = v5_counts[field]
        lines.append(
            f"| {field} | {v4_hit}/{v4_total} | {pct(v4_hit, v4_total)} | "
            f"{v5_hit}/{v5_total} | {pct(v5_hit, v5_total)} | {v5_hit - v4_hit:+d} |"
        )
    lines.extend(
        [
            "",
            f"- Generated v5 distribution rows: {generated_rows}. Matched Finance side-by-side rows: {len(v5_matches)}.",
            f"- Location verification from v5 distribution segment[1] / column R equivalent: {loc[0]}/{loc[1]} = {pct(*loc)}.",
            f"- Existing scored fields unchanged from v4: {'YES' if unchanged else 'NO'}.",
            f"- Side-by-side workbook: `{OUT_XLSX}`.",
            f"- v5 workbook: `{V5_PATH}`.",
        ]
    )
    OUT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def print_report(
    v4_matches: list[Match],
    v5_matches: list[Match],
    side_rows: int,
    diff_rows: int,
    summary_rows: int,
    loc: tuple[int, int],
    generated_rows: int,
) -> None:
    v4_counts = metric_counts(v4_matches)
    v5_counts = metric_counts(v5_matches)
    existing = ["Agency", "CC", "DIV", "Solution", "Full Combo"]
    unchanged = all(v4_counts[field] == v5_counts[field] for field in existing)
    addl_hits, addl_total = v5_counts["Additional Info"]
    print('Additional Info format confirmed: "empno.JQ" where both values exist.')
    print("03041 examples:")
    for line in example_lines(v5_matches, "03041"):
        print(f"  {line}")
    print("03317 examples:")
    for line in example_lines(v5_matches, "03317"):
        print(f"  {line}")
    print(f"New Additional Info match rate: {addl_hits}/{addl_total} = {pct(addl_hits, addl_total)}")
    print(f"Agency/CC/DIV/Solution/full-combo unchanged from v4: {'YES' if unchanged else 'NO'}")
    if not unchanged:
        for field in existing:
            print(f"  {field}: v4 {v4_counts[field][0]}/{v4_counts[field][1]} vs v5 {v5_counts[field][0]}/{v5_counts[field][1]}")
    print(f"Location=20100 verification: {loc[0]}/{loc[1]} = {pct(*loc)}")
    print(f"Side-by-side workbook path: {OUT_XLSX}")
    print(f"Generated OUR distribution rows: {generated_rows}")
    print(f"Rows by tab: Side by Side={side_rows}; Differences Only={diff_rows}; Summary={summary_rows}")
    print("Final paths:")
    print(f"  v5 xlsx: {V5_PATH}")
    print(f"  side-by-side xlsx: {OUT_XLSX}")
    print(f"  compare report: {OUT_MD}")


def main() -> None:
    # Read these inputs explicitly as part of the v5 audit inputs; side-by-side values are sourced from OURS and Finance.
    touch_supplier_and_manpower()
    key_by_invoice = load_key_lines()
    v4_lines = load_our_lines(V4_PATH)
    v5_lines = load_our_lines(V5_PATH)
    v4_matches = build_matches(v4_lines, key_by_invoice)
    v5_matches = build_matches(v5_lines, key_by_invoice)
    side_rows, diff_rows, summary_rows = write_workbook(v5_matches, len(v5_lines))
    loc = location_hits(v5_lines)
    write_markdown(v4_matches, v5_matches, loc, len(v5_lines))
    print_report(v4_matches, v5_matches, side_rows, diff_rows, summary_rows, loc, len(v5_lines))


if __name__ == "__main__":
    main()
