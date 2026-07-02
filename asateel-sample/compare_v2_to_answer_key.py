#!/usr/bin/env python3
from __future__ import annotations

import math
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parent
POC_XLSX = ROOT / "_poc_out" / "asateel-poc-oracle-CENTRAL-full-v2-2026-06-20.xlsx"
ENTRY_FILES = [ROOT / "_allocation" / "Entry-1.xlsm", ROOT / "_allocation" / "Entry-2.xlsm"]
REPORT_PATH = ROOT / "COMPARE-REPORT-2026-06-20.md"
AMOUNT_TOLERANCE = 0.01


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip()


def code(value: Any, width: int | None = None) -> str:
    text = clean(value)
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    if width and text.isdigit():
        text = text.zfill(width)
    return text


def money(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return round(float(value), 2)
    text = str(value).replace(",", "").replace("SAR", "").replace("ر.س", "")
    text = re.sub(r"[^\d.\-]", "", text)
    if not text:
        return None
    try:
        return round(float(text), 2)
    except ValueError:
        return None


def parse_distribution(value: Any) -> dict[str, str]:
    raw = clean(value)
    if not raw:
        return {"raw": ""}
    parts = [part.strip() for part in re.split(r"[-.]", raw)]
    if len(parts) < 10:
        return {"raw": raw}
    return {
        "raw": raw,
        "company": parts[0],
        "location": parts[1],
        "account": parts[2],
        "cost_center": parts[3],
        "division": parts[4],
        "solution": parts[5],
        "agency": parts[6],
        "project": parts[7],
        "intercompany": parts[8],
        "future": parts[9],
    }


def pct(hit: int, total: int) -> str:
    if not total:
        return "0.0%"
    return f"{hit / total:.1%}"


def load_poc_rows() -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(POC_XLSX, read_only=True, data_only=True)
    ws = wb["Sheet"]
    rows: list[dict[str, Any]] = []
    for ridx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        inv = code(row[2] if len(row) > 2 else "", 5)
        amount = money(row[12] if len(row) > 12 else None)
        combo = clean(row[13] if len(row) > 13 else "")
        if not inv and amount is None and not combo:
            continue
        dist = parse_distribution(combo)
        rows.append(
            {
                "source_file": POC_XLSX.name,
                "row": ridx,
                "invoice_no": inv,
                "line_no": len([r for r in rows if r["invoice_no"] == inv]) + 1,
                "amount": amount,
                "distribution": combo,
                "cost_center": code(row[20] if len(row) > 20 else ""),
                "division": code(row[22] if len(row) > 22 else ""),
                "agency": code(row[26] if len(row) > 26 else "", 5),
                "agency_name": clean(row[27] if len(row) > 27 else ""),
                "allocation_source": clean(row[33] if len(row) > 33 else "") or "none",
                "additional_information": clean(row[34] if len(row) > 34 else ""),
                "row_status": clean(row[32] if len(row) > 32 else ""),
                "dist_cost_center": code(dist.get("cost_center", "")),
                "dist_division": code(dist.get("division", "")),
                "dist_agency": code(dist.get("agency", ""), 5),
            }
        )
    wb.close()
    return rows


def load_answer_rows() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for path in ENTRY_FILES:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=False)
        ws = wb["Invoices"]
        current_inv = ""
        current_amount = None
        for ridx, row in enumerate(ws.iter_rows(min_row=9, values_only=True), start=9):
            inv = code(row[7] if len(row) > 7 else "", 5)
            if inv:
                current_inv = inv
                current_amount = money(row[9] if len(row) > 9 else None)
            if not current_inv:
                continue
            line_amount = money(row[82] if len(row) > 82 else None)
            combo = clean(row[98] if len(row) > 98 else "")
            dist = parse_distribution(combo)
            if line_amount is None and not dist.get("account"):
                continue
            rows.append(
                {
                    "source_file": path.name,
                    "row": ridx,
                    "invoice_no": current_inv,
                    "invoice_amount": current_amount,
                    "line_no": len([r for r in rows if r["invoice_no"] == current_inv]) + 1,
                    "amount": line_amount,
                    "distribution": combo,
                    "cost_center": code(dist.get("cost_center", "")),
                    "division": code(dist.get("division", "")),
                    "agency": code(dist.get("agency", ""), 5),
                    "solution": code(dist.get("solution", "")),
                }
            )
        wb.close()
    return rows


def same_amount(left: float | None, right: float | None) -> bool:
    return left is not None and right is not None and abs(left - right) <= AMOUNT_TOLERANCE


def match_invoice_rows(poc_rows: list[dict[str, Any]], answer_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    unused = answer_rows.copy()
    matches: list[dict[str, Any]] = []
    for idx, ours in enumerate(poc_rows):
        best = None
        match_method = "none"
        if idx < len(answer_rows) and answer_rows[idx] in unused:
            ordered = answer_rows[idx]
            if same_amount(ours["amount"], ordered["amount"]) or not any(same_amount(ours["amount"], a["amount"]) for a in unused):
                best = ordered
                match_method = "line_order"
        if best is None:
            amount_candidates = [a for a in unused if same_amount(ours["amount"], a["amount"])]
            if amount_candidates:
                best = amount_candidates[0]
                match_method = "amount"
        if best is None and unused:
            best = unused[0]
            match_method = "line_order_fallback"
        if best in unused:
            unused.remove(best)
        matches.append({"ours": ours, "answer": best, "match_method": match_method})
    for answer in unused:
        matches.append({"ours": None, "answer": answer, "match_method": "answer_unmatched"})
    return matches


def build_comparison() -> dict[str, Any]:
    poc_rows = load_poc_rows()
    answer_rows = load_answer_rows()
    poc_by_inv: dict[str, list[dict[str, Any]]] = defaultdict(list)
    answer_by_inv: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in poc_rows:
        poc_by_inv[row["invoice_no"]].append(row)
    for row in answer_rows:
        answer_by_inv[row["invoice_no"]].append(row)

    all_invoices = sorted(set(poc_by_inv) | set(answer_by_inv))
    matches = []
    for inv in all_invoices:
        matches.extend(match_invoice_rows(poc_by_inv.get(inv, []), answer_by_inv.get(inv, [])))

    comparable = [m for m in matches if m["ours"] is not None and m["answer"] is not None]
    stats = {
        "agency": 0,
        "cost_center": 0,
        "division": 0,
        "distribution": 0,
        "amount": 0,
    }
    field_mismatches = []
    source_breakdown: dict[str, Counter[str]] = defaultdict(Counter)
    pair_patterns = Counter()
    for match in comparable:
        ours = match["ours"]
        answer = match["answer"]
        source = ours["allocation_source"] or "none"
        checks = {
            "Agency": (ours["agency"], answer["agency"]),
            "CostCenter": (ours["cost_center"], answer["cost_center"]),
            "DIV": (ours["division"], answer["division"]),
        }
        if ours["agency"] == answer["agency"]:
            stats["agency"] += 1
        if ours["cost_center"] == answer["cost_center"]:
            stats["cost_center"] += 1
        if ours["division"] == answer["division"]:
            stats["division"] += 1
        if ours["distribution"] == answer["distribution"]:
            stats["distribution"] += 1
        if same_amount(ours["amount"], answer["amount"]):
            stats["amount"] += 1

        source_breakdown[source]["lines"] += 1
        for stat_name, field_name in [("agency", "Agency"), ("cost_center", "CostCenter"), ("division", "DIV")]:
            source_breakdown[source][f"{stat_name}_hit"] += int(checks[field_name][0] == checks[field_name][1])
        source_breakdown[source]["combo_hit"] += int(ours["distribution"] == answer["distribution"])
        source_breakdown[source]["amount_hit"] += int(same_amount(ours["amount"], answer["amount"]))

        for field, (left, right) in checks.items():
            if left != right:
                field_mismatches.append(
                    {
                        "invoice": ours["invoice_no"],
                        "line": ours["line_no"],
                        "field": field,
                        "ours": left,
                        "answer": right,
                        "source": source,
                    }
                )
                pair_patterns[(field, source, left or "(blank)", right or "(blank)")] += 1

    return {
        "poc_rows": poc_rows,
        "answer_rows": answer_rows,
        "matches": matches,
        "comparable": comparable,
        "stats": stats,
        "field_mismatches": field_mismatches,
        "source_breakdown": source_breakdown,
        "pair_patterns": pair_patterns,
        "poc_by_inv": poc_by_inv,
        "answer_by_inv": answer_by_inv,
    }


def md_table(headers: list[str], rows: list[list[Any]]) -> list[str]:
    lines = ["| " + " | ".join(headers) + " |", "| " + " | ".join(["---"] * len(headers)) + " |"]
    for row in rows:
        escaped = [clean(cell).replace("|", "\\|") for cell in row]
        lines.append("| " + " | ".join(escaped) + " |")
    return lines


def build_report(data: dict[str, Any]) -> str:
    comparable = data["comparable"]
    total = len(comparable)
    stats = data["stats"]
    poc_by_inv = data["poc_by_inv"]
    answer_by_inv = data["answer_by_inv"]
    poc_invs = set(poc_by_inv)
    answer_invs = set(answer_by_inv)
    missing_in_key = sorted(poc_invs - answer_invs)
    missing_in_ours = sorted(answer_invs - poc_invs)
    matched_poc_lines = sum(1 for m in data["matches"] if m["ours"] is not None and m["answer"] is not None)
    unmatched_poc_lines = sum(1 for m in data["matches"] if m["ours"] is not None and m["answer"] is None)
    unmatched_answer_lines = sum(1 for m in data["matches"] if m["ours"] is None and m["answer"] is not None)

    lines = [
        "# Asateel v2 POC vs Finance Entry Answer Key",
        "",
        f"- POC output: `{POC_XLSX.relative_to(ROOT)}`",
        "- Answer key: `_allocation/Entry-1.xlsm`, `_allocation/Entry-2.xlsm`",
        "- Matching: invoice number, then line order; if line order amount did not align, amount within 0.01 was used as fallback.",
        "",
        "## Overall Hit Rates",
        "",
    ]
    lines.extend(
        md_table(
            ["Field", "Hits", "Total", "Rate"],
            [
                ["Agency", stats["agency"], total, pct(stats["agency"], total)],
                ["Cost Center", stats["cost_center"], total, pct(stats["cost_center"], total)],
                ["DIV", stats["division"], total, pct(stats["division"], total)],
                ["Full Distribution Combination", stats["distribution"], total, pct(stats["distribution"], total)],
                ["Line Amount", stats["amount"], total, pct(stats["amount"], total)],
            ],
        )
    )
    lines.extend(
        [
            "",
            "## Coverage",
            "",
            f"- POC invoices / lines: {len(poc_invs)} / {len(data['poc_rows'])}",
            f"- Answer-key invoices / lines: {len(answer_invs)} / {len(data['answer_rows'])}",
            f"- Matched POC lines: {matched_poc_lines}",
            f"- POC lines without answer-key row: {unmatched_poc_lines}",
            f"- Answer-key lines without POC row: {unmatched_answer_lines}",
            f"- Invoices in POC but not answer key: {', '.join(missing_in_key) if missing_in_key else 'none'}",
            f"- Invoices in answer key but not POC: {', '.join(missing_in_ours) if missing_in_ours else 'none'}",
            "",
            "## Mismatch Table",
            "",
            f"Total Agency/CC/DIV disagreements: {len(data['field_mismatches'])}. Showing first 60.",
            "",
        ]
    )
    mismatch_rows = [
        [m["invoice"], m["line"], m["field"], m["ours"], m["answer"], m["source"]]
        for m in data["field_mismatches"][:60]
    ]
    lines.extend(md_table(["Invoice", "Line", "Field", "Ours", "Answer Key", "Allocation Source"], mismatch_rows or [["none", "", "", "", "", ""]]))

    lines.extend(["", "## Mismatches by Allocation Source", ""])
    source_rows = []
    ordered_sources = ["brand", "salesperson", "supplier_expenses_format", "none"]
    ordered_sources.extend(
        source for source in sorted(data["source_breakdown"])
        if source not in set(ordered_sources)
    )
    for source in ordered_sources:
        counts = data["source_breakdown"].get(source, Counter())
        line_count = counts["lines"]
        if not line_count:
            source_rows.append([source, 0, 0, "0/0 0.0%", "0/0 0.0%", "0/0 0.0%", "0/0 0.0%", "0/0 0.0%"])
            continue
        segment_misses = (
            (line_count - counts["agency_hit"])
            + (line_count - counts["cost_center_hit"])
            + (line_count - counts["division_hit"])
        )
        source_rows.append(
            [
                source,
                line_count,
                segment_misses,
                f"{counts['agency_hit']}/{line_count} {pct(counts['agency_hit'], line_count)}",
                f"{counts['cost_center_hit']}/{line_count} {pct(counts['cost_center_hit'], line_count)}",
                f"{counts['division_hit']}/{line_count} {pct(counts['division_hit'], line_count)}",
                f"{counts['combo_hit']}/{line_count} {pct(counts['combo_hit'], line_count)}",
                f"{counts['amount_hit']}/{line_count} {pct(counts['amount_hit'], line_count)}",
            ]
        )
    lines.extend(md_table(["Source", "Lines", "Segment Misses", "Agency", "CC", "DIV", "Full Combo", "Amount"], source_rows))

    lines.extend(["", "## 03317 Detail", ""])
    inv_03317 = [m for m in data["matches"] if (m["ours"] or m["answer"])["invoice_no"] == "03317"]
    detail_rows = []
    for match in inv_03317:
        ours = match["ours"] or {}
        answer = match["answer"] or {}
        detail_rows.append(
            [
                ours.get("line_no", ""),
                match["match_method"],
                ours.get("amount", ""),
                answer.get("amount", ""),
                ours.get("cost_center", ""),
                answer.get("cost_center", ""),
                ours.get("division", ""),
                answer.get("division", ""),
                ours.get("agency", ""),
                answer.get("agency", ""),
                ours.get("distribution", ""),
                answer.get("distribution", ""),
                ours.get("allocation_source", ""),
            ]
        )
    lines.extend(
        md_table(
            ["Line", "Match", "Our Amt", "Key Amt", "Our CC", "Key CC", "Our DIV", "Key DIV", "Our Agency", "Key Agency", "Our Combo", "Key Combo", "Source"],
            detail_rows or [["not present", "", "", "", "", "", "", "", "", "", "", "", ""]],
        )
    )

    lines.extend(["", "## Top Systematic Error Patterns", ""])
    top_patterns = data["pair_patterns"].most_common(12)
    pattern_rows = [[count, field, source, ours, answer] for (field, source, ours, answer), count in top_patterns]
    lines.extend(md_table(["Count", "Field", "Source", "Ours", "Answer Key"], pattern_rows or [["none", "", "", "", ""]]))

    observations = []
    if top_patterns:
        sources = Counter()
        fields = Counter()
        for (field, source, _ours, _answer), count in data["pair_patterns"].items():
            sources[source] += count
            fields[field] += count
        weakest_source, weakest_source_count = sources.most_common(1)[0]
        weakest_field, weakest_field_count = fields.most_common(1)[0]
        observations.append(
            f"- The weakest path by raw segment disagreements is `{weakest_source}` ({weakest_source_count} Agency/CC/DIV misses)."
        )
        observations.append(
            f"- The most frequent disagreed segment is `{weakest_field}` ({weakest_field_count} misses)."
        )
    combo_misses = total - stats["distribution"]
    amount_misses = total - stats["amount"]
    observations.extend(
        [
            f"- Full-combo misses ({combo_misses}) are dominated by nonzero finance solution segments where the POC wrote `00000`; a canonical brand/salesperson-to-solution rule is needed before full combinations can score well.",
            f"- Amount mismatches ({amount_misses}) show that the POC PDF-derived line split often does not mirror finance's Entry split; the Supplier Expenses Format allocation amounts need to be the primary line source when finance uses them.",
            "- Supplier Expenses Format rows mostly carry the right Agency/CC, but DIV is often different, especially `110 -> 192` and `260 -> 170`; the rule needs the Entry/lookup DIV, not the generic source DIV.",
            "- Salesperson rows are the weakest source because unresolved or mismapped salespeople lead to blank/00000 or wrong Agency/CC/DIV; closing this needs the canonical salesperson-to-manpower mapping plus finance's manual override cases.",
        ]
    )
    lines.extend(observations)
    lines.append("")
    return "\n".join(lines)


def main() -> None:
    data = build_comparison()
    report = build_report(data)
    REPORT_PATH.write_text(report, encoding="utf-8")

    total = len(data["comparable"])
    stats = data["stats"]
    print("Asateel v2 POC vs finance Entry answer key")
    print(f"Report: {REPORT_PATH}")
    print(f"Matched lines: {total}")
    print(f"Agency: {stats['agency']}/{total} = {pct(stats['agency'], total)}")
    print(f"Cost Center: {stats['cost_center']}/{total} = {pct(stats['cost_center'], total)}")
    print(f"DIV: {stats['division']}/{total} = {pct(stats['division'], total)}")
    print(f"Full combo: {stats['distribution']}/{total} = {pct(stats['distribution'], total)}")
    print(f"Amount: {stats['amount']}/{total} = {pct(stats['amount'], total)}")
    print(f"Agency/CC/DIV mismatches: {len(data['field_mismatches'])}")


if __name__ == "__main__":
    main()
