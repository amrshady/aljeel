#!/usr/bin/env python3
"""Conservative, auditable Jawal five-segment accuracy scorer.

Truth is read only from the actual Oracle upload combination.  In particular, this
module never reconstructs a truth combination from decomposed helper columns.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass, asdict
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable
from xml.etree import ElementTree

from openpyxl import load_workbook

SEGMENTS = ("account", "cost_center", "div", "solution", "agency")
SEGMENT_POSITIONS = (2, 3, 4, 5, 6)
VALID_COMBINATION = re.compile(r"^\s*[^-]+(?:\s*-\s*[^-]+){9}\s*$")
TEN_PLUS_DIGITS = re.compile(r"(?<!\d)(\d{10,})(?!\d)")
SPACE = re.compile(r"\s+")
NON_ALNUM = re.compile(r"[^a-z0-9]+")

DEFAULTS = {
    "J26-1080": (
        "batches/jawal-J26-1080/output/Spreadsheet-J26-1080-FILLED-v30-SPLIT.xlsx",
        "/home/clawdbot/.openclaw/media/inbound/J26-1080_Ready_to_upload---b121c957-a3cc-4c8e-9b5b-f62a30ca7036.xlsx",
    ),
    "J26-1108": (
        "batches/jawal-J26-1108/output/Spreadsheet-J26-1108-FILLED-v30-SPLIT.xlsx",
        "/home/clawdbot/.openclaw/media/inbound/Copy_of_Spreadsheet-J26-1108-FILLED-v30-SPLIT---48c951a2-1983-4d1c-9155-eefd1a14dba9.xlsx",
    ),
    "J26-1116": (
        "batches/jawal-J26-1116/output/Spreadsheet-J26-1116-FILLED-v30-SPLIT.xlsx",
        "/home/clawdbot/.openclaw/media/inbound/Copy_of_Spreadsheet-J26-1116-FILLED-v30_1---07303b4b-2313-4544-beaf-35d96fcd72ca.xlsx",
    ),
}


@dataclass(frozen=True)
class Row:
    side: str
    sheet: str
    excel_row: int
    description: str
    amount: str
    combination: str
    segments: tuple[str, ...] | None
    ticket_core: str | None
    employee: str

    @property
    def covered(self) -> bool:
        return self.segments is not None

    @property
    def row_id(self) -> str:
        return f"{self.side}:{self.sheet}:{self.excel_row}"


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return SPACE.sub(" ", str(value).strip())


def norm_segment(value: Any) -> str:
    """Preserve blank as blank: blank is deliberately not equal to zero."""
    value = clean(value)
    if not value:
        return ""
    if re.fullmatch(r"[+-]?\d+(?:\.0+)?", value):
        value = value.split(".", 1)[0]
        sign = "-" if value.startswith("-") else ""
        digits = value.lstrip("+-0") or "0"
        return sign + digits
    return value.casefold()


def parse_combination(value: Any) -> tuple[str, ...] | None:
    raw = clean(value)
    if not raw or not VALID_COMBINATION.fullmatch(raw):
        return None
    parts = [part.strip() for part in raw.split("-")]
    if len(parts) != 10 or any(not part for part in parts):
        return None
    return tuple(norm_segment(parts[pos]) for pos in SEGMENT_POSITIONS)


def ticket_core(value: Any) -> str | None:
    matches = TEN_PLUS_DIGITS.findall(clean(value))
    return matches[-1][-10:] if matches else None


def norm_amount(value: Any) -> str:
    raw = clean(value).replace(",", "")
    if not raw:
        return ""
    try:
        return format(Decimal(raw).quantize(Decimal("0.01")), "f")
    except InvalidOperation:
        return raw.casefold()


def norm_description(value: Any) -> str:
    return NON_ALNUM.sub(" ", clean(value).casefold()).strip()


def header_text(value: Any) -> str:
    return SPACE.sub(" ", clean(value)).casefold()


def find_upload_sheet(path: Path, side: str, batch: str) -> tuple[Any, int, dict[str, int], str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    candidates = []
    for ws in wb.worksheets:
        for row_no, values in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12), values_only=True), 1):
            headers = {header_text(v): i for i, v in enumerate(values) if clean(v)}
            dist = next((i for h, i in headers.items() if h.startswith("distribution combination")), None)
            gl = headers.get("gl") if side == "truth" and batch == "J26-1080" else None
            combo = dist if dist is not None else gl
            if combo is not None:
                source = "Distribution Combination[..]" if dist is not None else "GL"
                candidates.append((ws, row_no, headers, source, combo))
    if len(candidates) != 1:
        raise ValueError(f"{path}: expected exactly one Oracle upload sheet, found {len(candidates)}")
    ws, row_no, headers, source, combo = candidates[0]
    headers["__combination__"] = combo
    return ws, row_no, headers, source


def find_col(headers: dict[str, int], predicates: Iterable[str]) -> int | None:
    for wanted in predicates:
        for name, index in headers.items():
            if wanted in name:
                return index
    return None


def load_rows(path: Path, side: str, batch: str) -> tuple[list[Row], dict[str, Any]]:
    ws, header_row, headers, source = find_upload_sheet(path, side, batch)
    combo_col = headers["__combination__"]
    desc_col = find_col(headers, ("description", "ticket no."))
    amount_col = find_col(headers, ("*amount", "taxable amt."))
    employee_col = find_col(headers, ("employee no", "employee no."))
    primary_col = find_col(headers, ("*invoice header identifier", "sl. #"))
    if primary_col is None:
        primary_col = 0
    rows = []
    for excel_row, values in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
        # An upload/invoice data line is identified structurally, never by Account/GL.
        if primary_col >= len(values) or not clean(values[primary_col]):
            continue
        def at(index: int | None) -> Any:
            return values[index] if index is not None and index < len(values) else None
        description = clean(at(desc_col))
        combination = clean(at(combo_col))
        rows.append(Row(
            side=side, sheet=ws.title, excel_row=excel_row,
            description=description, amount=norm_amount(at(amount_col)),
            combination=combination, segments=parse_combination(combination),
            ticket_core=ticket_core(description), employee=clean(at(employee_col)),
        ))
    return rows, {
        "sheet": ws.title, "header_row": header_row, "combination_source": source,
        "data_rows": len(rows), "combination_covered_rows": sum(r.covered for r in rows),
    }


def pair_multiset(left: list[Row], right: list[Row], method: str) -> tuple[list[tuple[Row, Row, str]], list[Row], list[Row]]:
    """Pair deterministic multisets, aligning equal segment tuples before misses."""
    key = lambda r: (r.segments is None, r.segments or (), r.employee, r.excel_row)
    left, right = sorted(left, key=key), sorted(right, key=key)
    pairs = [(a, b, method) for a, b in zip(left, right)]
    return pairs, left[len(pairs):], right[len(pairs):]


def align(truth: list[Row], produced: list[Row]) -> dict[str, Any]:
    remaining_t = {r.row_id: r for r in truth}
    remaining_p = {r.row_id: r for r in produced}
    pairs: list[tuple[Row, Row, str]] = []
    ambiguous: list[dict[str, Any]] = []

    def consume(trows: list[Row], prows: list[Row], method: str) -> None:
        new, _, _ = pair_multiset(trows, prows, method)
        for trow, prow, how in new:
            pairs.append((trow, prow, how))
            remaining_t.pop(trow.row_id, None)
            remaining_p.pop(prow.row_id, None)

    # Stage 1a: a unique, valid ticket core is sufficient; amount is audited.
    tg, pg = defaultdict(list), defaultdict(list)
    for row in remaining_t.values():
        if row.ticket_core: tg[row.ticket_core].append(row)
    for row in remaining_p.values():
        if row.ticket_core: pg[row.ticket_core].append(row)
    for core in sorted(tg.keys() & pg.keys()):
        if len(tg[core]) == len(pg[core]) == 1:
            consume(tg[core], pg[core], "ticket_core_unique")

    # Stage 1b: repeats use amount as a tie-breaker and retain multiplicity.
    tg, pg = defaultdict(list), defaultdict(list)
    for row in remaining_t.values():
        if row.ticket_core: tg[(row.ticket_core, row.amount)].append(row)
    for row in remaining_p.values():
        if row.ticket_core: pg[(row.ticket_core, row.amount)].append(row)
    for key in sorted(tg.keys() & pg.keys()):
        consume(tg[key], pg[key], "ticket_core_amount_multiset")

    # Stage 2: ticketless hotel/train/nonstandard lines require exact normalized
    # description + amount. It is intentionally available only without a valid core.
    tg, pg = defaultdict(list), defaultdict(list)
    for row in remaining_t.values():
        if not row.ticket_core: tg[(norm_description(row.description), row.amount)].append(row)
    for row in remaining_p.values():
        if not row.ticket_core: pg[(norm_description(row.description), row.amount)].append(row)
    for key in sorted((k for k in tg.keys() & pg.keys() if k[0])):
        consume(tg[key], pg[key], "description_amount_multiset")

    # Surface unresolved many-to-many candidate cores as ambiguity; never guess.
    tg, pg = defaultdict(list), defaultdict(list)
    for row in remaining_t.values():
        if row.ticket_core: tg[row.ticket_core].append(row)
    for row in remaining_p.values():
        if row.ticket_core: pg[row.ticket_core].append(row)
    for core in sorted(tg.keys() & pg.keys()):
        ambiguous.append({"ticket_core": core, "truth_rows": [r.row_id for r in tg[core]],
                          "produced_rows": [r.row_id for r in pg[core]],
                          "reason": "ticket core repeats but amount/multiplicity does not resolve conservatively"})

    return {"pairs": pairs, "unmatched_truth": list(remaining_t.values()),
            "unmatched_produced": list(remaining_p.values()), "ambiguous": ambiguous}


def file_evidence(path: Path) -> dict[str, Any]:
    stat = path.stat()
    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    core = {"created": None, "modified": None}
    with zipfile.ZipFile(path) as archive:
        xml = ElementTree.fromstring(archive.read("docProps/core.xml"))
        for element in xml:
            local = element.tag.rsplit("}", 1)[-1]
            if local in core:
                core[local] = element.text
    return {"path": str(path.resolve()), "size_bytes": stat.st_size, "sha256": digest,
            "filesystem_mtime_utc": datetime.fromtimestamp(stat.st_mtime, timezone.utc).isoformat(),
            "workbook_core_timestamps": core}


def safe_row(row: Row) -> dict[str, Any]:
    return {"row_id": row.row_id, "excel_row": row.excel_row, "ticket_core": row.ticket_core,
            "amount": row.amount, "description": row.description, "employee": row.employee,
            "covered": row.covered, "combination": row.combination}


def ratio(n: int, d: int) -> dict[str, Any]:
    return {"numerator": n, "denominator": d, "percent": round(100 * n / d, 4) if d else None}


def score_batch(batch: str, produced_path: Path, truth_path: Path) -> dict[str, Any]:
    truth, truth_info = load_rows(truth_path, "truth", batch)
    produced, produced_info = load_rows(produced_path, "produced", batch)
    aligned = align(truth, produced)
    pairs = aligned.pop("pairs")
    eligible = [(t, p, m) for t, p, m in pairs if t.covered and p.covered]
    segment_hits = {name: 0 for name in SEGMENTS}
    exact = 0
    comparisons = []
    for trow, prow, method in eligible:
        hits = [a == b for a, b in zip(trow.segments or (), prow.segments or ())]
        for name, hit in zip(SEGMENTS, hits): segment_hits[name] += int(hit)
        exact += int(all(hits))
        comparisons.append({"truth_row": trow.row_id, "produced_row": prow.row_id,
                            "method": method, "segment_equal": dict(zip(SEGMENTS, hits)),
                            "truth_segments": dict(zip(SEGMENTS, trow.segments or ())),
                            "produced_segments": dict(zip(SEGMENTS, prow.segments or ()))})
    methods = Counter(method for _, _, method in pairs)
    # End-to-end is anchored to the full truth universe. Unmatched/ambiguous truth
    # rows and truth rows without a valid combination remain denominator misses.
    e2e_denominator = len(truth)
    ambiguous_truth_ids = {row_id for group in aligned["ambiguous"] for row_id in group["truth_rows"]}
    ambiguous_produced_ids = {row_id for group in aligned["ambiguous"] for row_id in group["produced_rows"]}
    result = {
        "batch": batch,
        "inputs": {"truth": file_evidence(truth_path), "produced": file_evidence(produced_path)},
        "truth_ingest": truth_info, "produced_ingest": produced_info,
        "coverage": {
            "truth": ratio(truth_info["combination_covered_rows"], len(truth)),
            "produced": ratio(produced_info["combination_covered_rows"], len(produced)),
        },
        "alignment": {
            "matched_rows": len(pairs), "eligible_covered_matched_rows": len(eligible),
            "unmatched_truth_rows": len(aligned["unmatched_truth"]),
            "unmatched_produced_rows": len(aligned["unmatched_produced"]),
            "ambiguous_groups": len(aligned["ambiguous"]),
            "ambiguous_truth_rows": len(ambiguous_truth_ids),
            "ambiguous_produced_rows": len(ambiguous_produced_ids),
            "method_counts": dict(sorted(methods.items())),
        },
        "covered_matched_metrics": {
            "all_5_exact": ratio(exact, len(eligible)),
            "segments": {name: ratio(segment_hits[name], len(eligible)) for name in SEGMENTS},
        },
        "end_to_end": {
            "scoring_universe": "all truth rows; unmatched, ambiguous, uncovered, or extra truth rows are padded misses",
            "all_5_exact": ratio(exact, e2e_denominator),
            "segments": {name: ratio(segment_hits[name], e2e_denominator) for name in SEGMENTS},
        },
        "unmatched_truth": [safe_row(r) for r in aligned["unmatched_truth"]],
        "unmatched_produced": [safe_row(r) for r in aligned["unmatched_produced"]],
        "ambiguous": aligned["ambiguous"], "comparisons": comparisons,
    }
    return result


def markdown(report: dict[str, Any]) -> str:
    lines = [f"# Jawal Accuracy v2 — {report['report_date_utc']}", "",
             "Truth provenance is limited to the supplied files and their recorded metadata. "
             "No download event is inferred: `AuditEvent` does not log `DOWNLOAD`, so the absolute latest downloaded file cannot be proven.", ""]
    for b in report["batches"]:
        cm, ee, al = b["covered_matched_metrics"], b["end_to_end"], b["alignment"]
        lines += [f"## {b['batch']}", "",
                  f"Truth rows: {b['truth_ingest']['data_rows']}; produced rows: {b['produced_ingest']['data_rows']}. "
                  f"Truth source: `{b['truth_ingest']['combination_source']}`; produced source: `{b['produced_ingest']['combination_source']}`.", "",
                  f"Coverage — truth: **{b['coverage']['truth']['numerator']}/{b['coverage']['truth']['denominator']} ({b['coverage']['truth']['percent']}%)**; "
                  f"produced: **{b['coverage']['produced']['numerator']}/{b['coverage']['produced']['denominator']} ({b['coverage']['produced']['percent']}%)**.", "",
                  f"Alignment — matched {al['matched_rows']}, eligible covered+matched {al['eligible_covered_matched_rows']}, "
                  f"unmatched truth {al['unmatched_truth_rows']}, unmatched produced {al['unmatched_produced_rows']}, "
                  f"ambiguous {al['ambiguous_truth_rows']} truth/{al['ambiguous_produced_rows']} produced rows in {al['ambiguous_groups']} groups. "
                  f"Methods: `{json.dumps(al['method_counts'], sort_keys=True)}`.", "",
                  "| Metric | Covered + matched | End-to-end padded |", "|---|---:|---:|",
                  f"| All 5 exact | {cm['all_5_exact']['numerator']}/{cm['all_5_exact']['denominator']} ({cm['all_5_exact']['percent']}%) | {ee['all_5_exact']['numerator']}/{ee['all_5_exact']['denominator']} ({ee['all_5_exact']['percent']}%) |"]
        for name in SEGMENTS:
            a, z = cm["segments"][name], ee["segments"][name]
            lines.append(f"| {name.replace('_', ' ').title()} | {a['numerator']}/{a['denominator']} ({a['percent']}%) | {z['numerator']}/{z['denominator']} ({z['percent']}%) |")
        lines += ["", "### Input evidence", ""]
        for side in ("truth", "produced"):
            e = b["inputs"][side]
            lines.append(f"- {side.title()}: `{e['path']}`; {e['size_bytes']} bytes; mtime `{e['filesystem_mtime_utc']}`; core `{json.dumps(e['workbook_core_timestamps'], sort_keys=True)}`; SHA256 `{e['sha256']}`")
        lines += ["", "Detailed comparisons, unmatched rows, and ambiguity records are in the companion JSON report.", ""]
    lines += ["## Known limitations", "",
              "- Matching deliberately refuses fuzzy passenger/route matching. Ticketless rows require exact normalized description and amount.",
              "- Workbook core timestamps and filesystem mtimes establish file evidence, not download history or human approval chronology.",
              "- End-to-end uses the complete truth-row universe. Unmatched, ambiguous, uncovered, or extra truth rows are misses; unmatched produced rows are separately audited.", ""]
    return "\n".join(lines)


def self_test() -> None:
    assert norm_segment("") == "" and norm_segment("00000") == "0"
    assert norm_segment("") != norm_segment("00000")
    assert parse_combination("03-40100-60307021-160014-170-10017-10072-00000-00-000000") == ("60307021", "160014", "170", "10017", "10072")
    assert parse_combination("03-40100-60307021") is None
    assert ticket_core("prefix 065 4860349330") == "4860349330"
    assert ticket_core("hotel without ticket") is None
    assert norm_amount("1,234.5") == "1234.50"
    print("score_5seg.py self-test: PASS (7 assertions)")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--self-test", action="store_true")
    parser.add_argument("--date", default="2026-08-04")
    parser.add_argument("--out-dir", type=Path, default=Path(__file__).resolve().parent)
    args = parser.parse_args()
    if args.self_test:
        self_test()
        return 0
    root = Path(__file__).resolve().parents[2]
    batches = [score_batch(batch, root / produced, Path(truth)) for batch, (produced, truth) in DEFAULTS.items()]
    report = {"schema_version": "jawal-accuracy-v2", "report_date_utc": args.date,
              "generated_at_utc": datetime.now(timezone.utc).isoformat(),
              "provenance_statement": "State evidence only; no download event inferred. AuditEvent does not log DOWNLOAD, so absolute latest downloaded cannot be proven.",
              "batches": batches}
    args.out_dir.mkdir(parents=True, exist_ok=True)
    stem = f"jawal-accuracy-v2-{args.date}"
    json_path, md_path = args.out_dir / f"{stem}.json", args.out_dir / f"{stem}.md"
    json_path.write_text(json.dumps(report, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    md_path.write_text(markdown(report), encoding="utf-8")
    print(json.dumps({b["batch"]: {"coverage": b["coverage"], "alignment": b["alignment"],
                      "covered_matched": b["covered_matched_metrics"], "end_to_end": b["end_to_end"]}
                      for b in batches}, indent=2))
    print(f"wrote {json_path}\nwrote {md_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
