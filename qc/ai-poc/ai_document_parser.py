#!/usr/bin/env python3
"""
AI Document Parser POC — replaces hand-coded column detection with LLM extraction.

Reads an invoice xlsx (any AlJeel layout — Details flat, Sheet wide-merged, future variants),
sends the first 30-50 rows of cell content to Gemini, gets back a structured row table.

Schema-validated against Pydantic. Cached by xlsx SHA256.
"""

import openpyxl
import hashlib
import json
import os
import sys
import time
import urllib.request
import urllib.error
from pathlib import Path
from typing import Optional, List
from datetime import datetime, date

GEMINI_BASE_URL = os.environ.get("GEMINI_BASE_URL", "https://gateway.ai.cloudflare.com/v1/3724a3e71944b366a39b3735aa117a58/accord-aljeel-ap/google-ai-studio/v1beta")

# ---------------------------------------------------------------------------
# Pydantic-lite schema validation (no external deps)
# ---------------------------------------------------------------------------

REQUIRED_FIELDS = [
    "sl_no", "issue_date", "ref_no", "ticket_no", "passenger",
    "route", "service_date", "taxable_amount", "vat_pct",
    "vat_amount", "total_amount", "notes"
]

def _validate_row(row: dict, row_idx: int) -> tuple[bool, str]:
    """Validate a single parsed row. Returns (valid, error_msg)."""
    for f in REQUIRED_FIELDS:
        if f not in row:
            return False, f"Row {row_idx}: missing field '{f}'"
    # sl_no should be a number or string-of-number
    sl = row.get("sl_no")
    if sl is None or str(sl).strip() == "":
        return False, f"Row {row_idx}: sl_no is empty"
    # ticket_no may be empty: hotel/event/train rows carry a hotel voucher
    # ref (e.g. "26-808") in the invoice's Ticket/Hotel column, not a
    # 10-digit airline ticket, so the model correctly emits "" for them.
    # The key must still be present (checked in REQUIRED_FIELDS above).
    return True, ""

def _validate_schema(rows: list) -> tuple[bool, List[str]]:
    errors = []
    if not isinstance(rows, list):
        return False, ["Response is not a JSON array"]
    for i, row in enumerate(rows):
        ok, err = _validate_row(row, i + 1)
        if not ok:
            errors.append(err)
    return len(errors) == 0, errors


# ---------------------------------------------------------------------------
# Env loader (mirrors location_resolver.py)
# ---------------------------------------------------------------------------

def _load_env_key(name: str) -> str:
    val = os.environ.get(name, "")
    if not val:
        env_path = Path("/home/clawdbot/.openclaw/.env")
        if env_path.exists():
            for line in env_path.read_text(errors="replace").splitlines():
                if line.startswith(f"{name}="):
                    val = line.split("=", 1)[1].strip().strip('"').strip("'")
                    break
    return val


# ---------------------------------------------------------------------------
# Gemini call (reuses same HTTP pattern as location_resolver.py L7.7)
# ---------------------------------------------------------------------------

_GEMINI_MODEL_CASCADE = ["gemini-2.5-flash", "gemini-2.5-pro", "gemini-pro-latest"]

def _call_gemini_model(prompt_text: str, api_key: str, model: str) -> tuple[Optional[dict], bool]:
    """Returns (response_json, is_unavailable). is_unavailable is True on HTTP 503/UNAVAILABLE."""
    url = (
        f"{GEMINI_BASE_URL}/"
        f"models/{model}:generateContent?key={api_key}"
    )
    payload = json.dumps({
        "contents": [{"parts": [{"text": prompt_text}]}],
        "generationConfig": {
            "temperature": 0.0,
            "maxOutputTokens": 65536,
        }
    }).encode("utf-8")
    for attempt in range(3):
        req = urllib.request.Request(url, data=payload, headers={"Content-Type": "application/json", "User-Agent": "AlJeel-AP-Agent/1.0"})
        try:
            with urllib.request.urlopen(req, timeout=180) as resp:
                return json.loads(resp.read()), False
        except urllib.error.HTTPError as e:
            body = e.read().decode(errors="replace")[:300]
            if e.code == 503 or "UNAVAILABLE" in body:
                return None, True
            print(f"    [gemini/{model}] HTTP {e.code}: {body}", file=sys.stderr)
            if attempt == 2:
                return None, False
            time.sleep(2 ** attempt)
        except Exception as e:
            print(f"    [gemini/{model}] Error: {e}", file=sys.stderr)
            if attempt == 2:
                return None, False
            time.sleep(2 ** attempt)
    return None, False


def _call_gemini(prompt_text: str) -> tuple[Optional[str], Optional[str]]:
    """Call Gemini, return (raw_text_response, model_version_used)."""
    api_key = _load_env_key("GEMINI_API_KEY")
    if not api_key:
        raise RuntimeError("GEMINI_API_KEY not found in env or .env file")
    for idx, model in enumerate(_GEMINI_MODEL_CASCADE):
        print(f"  [ai_parser] Trying model: {model} ...", file=sys.stderr)
        result, unavailable = _call_gemini_model(prompt_text, api_key, model)
        if result:
            # Extract text from response
            try:
                text = result["candidates"][0]["content"]["parts"][0]["text"]
                model_ver = result.get("modelVersion") or model
                return text, model_ver
            except (KeyError, IndexError) as e:
                print(f"    [gemini/{model}] Unexpected response shape: {e}", file=sys.stderr)
        elif unavailable and idx < len(_GEMINI_MODEL_CASCADE) - 1:
            print(f"  [ai_parser] Model {model} returned 503 — waiting 15s before trying next model...", file=sys.stderr)
            time.sleep(15)
    return None, None


# ---------------------------------------------------------------------------
# xlsx → compact text grid
# ---------------------------------------------------------------------------

def _xlsx_sha256(xlsx_path: Path) -> str:
    h = hashlib.sha256()
    with open(xlsx_path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _xlsx_to_grid_text(xlsx_path: Path, max_data_rows: int = 120) -> tuple[str, int]:
    """
    Convert xlsx to a compact markdown-table-like text representation.
    Returns (grid_text, data_row_count).
    Works for both J26-788 (wide-merged Sheet) and J26-640 (Details flat sheet).
    """
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

    # Sheet selection: prefer 'Details' (J26-640 format), else first sheet
    if "Details" in wb.sheetnames:
        ws = wb["Details"]
        sheet_used = "Details"
    else:
        ws = wb.active
        sheet_used = ws.title

    rows = list(ws.iter_rows(values_only=True))
    max_col = ws.max_column

    # Find header row: first row where col1 matches Sl# pattern
    header_row_idx = None
    for i, row in enumerate(rows):
        first_cell = str(row[0] or "")
        if "Sl" in first_cell or "رقم" in first_cell:
            header_row_idx = i
            break
        # Also catch if col 0 is "1" and the row before has headers
        if str(row[0] or "").strip() == "1" and i > 0:
            # check previous row for header signals
            prev = rows[i - 1]
            prev_str = str(prev[0] or "") + str(prev[1] or "")
            if "Sl" in prev_str or "رقم" in prev_str or "Issue" in prev_str or "Date" in prev_str:
                header_row_idx = i - 1
                break

    if header_row_idx is None:
        # Fallback: find row where most cells have non-numeric content
        for i, row in enumerate(rows[:30]):
            non_empty = [c for c in row if c is not None]
            if len(non_empty) >= 4:
                header_row_idx = i
                break
        if header_row_idx is None:
            header_row_idx = 0

    header_row = rows[header_row_idx]
    data_rows = rows[header_row_idx + 1:]

    # Filter out empty rows and totals rows (no Sl# in first col)
    data_rows_clean = []
    for row in data_rows:
        first = str(row[0] or "").strip()
        if not first:
            continue
        # Stop at totals / summary rows
        first_lower = first.lower()
        if any(kw in first_lower for kw in ["total", "grand", "إجمالي", "مجموع"]):
            break
        data_rows_clean.append(row)
        if len(data_rows_clean) >= max_data_rows:
            break

    # Build compact grid: only include columns that have data in at least one row
    # Collect active column indices (1-indexed)
    active_cols = set()
    for cell_val in header_row:
        pass
    for ci in range(max_col):
        has_data = any(str(row[ci] if ci < len(row) else "").strip() for row in data_rows_clean)
        if has_data:
            active_cols.add(ci)
        # Also include header columns
        if ci < len(header_row) and header_row[ci] is not None:
            active_cols.add(ci)
    active_cols = sorted(active_cols)

    def fmt_cell(v) -> str:
        if v is None:
            return ""
        if isinstance(v, (int, float)):
            if isinstance(v, float) and v == int(v):
                return str(int(v))
            return str(v)
        s = str(v).strip().replace("\n", " ")
        return s[:40]  # truncate very long cells

    # Build header
    header_cells = [fmt_cell(header_row[ci] if ci < len(header_row) else None) for ci in active_cols]
    lines = ["| " + " | ".join(header_cells) + " |"]
    lines.append("| " + " | ".join(["---"] * len(active_cols)) + " |")
    for row in data_rows_clean:
        cells = [fmt_cell(row[ci] if ci < len(row) else None) for ci in active_cols]
        lines.append("| " + " | ".join(cells) + " |")

    grid_text = f"Sheet: {sheet_used}\n" + "\n".join(lines)
    return grid_text, len(data_rows_clean)


# ---------------------------------------------------------------------------
# Prompt builder
# ---------------------------------------------------------------------------

_EXTRACTION_PROMPT = """You are a structured data extractor for Arabic/English airline invoice spreadsheets.

Extract every invoice data row from the table below. Return ONLY a JSON array, no markdown, no explanation.

For each row return exactly these fields:
- sl_no: the row sequence number (integer or string)
- issue_date: YYYY-MM-DD string (normalize from any Excel/ISO date format)
- ref_no: reference/PO number string (may be empty string "")
- ticket_no: 10-digit ticket number ONLY, no airline prefix (e.g. "6905428827" from "065 6905428827")
- passenger: passenger name as string
- route: IATA route string (e.g. "RUH JED" or "RUH AJF RUH"), empty string if not present
- service_date: YYYY-MM-DD string (the actual travel/service date, NOT the invoice issue date)
- taxable_amount: float (the pre-tax amount, 0.0 if blank or VAT-zero row)
- vat_pct: float (percentage, e.g. 15.0, 0.0 if blank)
- vat_amount: float (0.0 if blank)
- total_amount: float (total invoice amount including VAT)
- notes: any notes/comments in the row, empty string "" if none

Rules:
- Skip header rows, blank rows, subtotal rows, grand-total rows
- For ticket_no: extract ONLY the last 10 digits. Input may be "065 6905428827" → output "6905428827"
- For dates: normalize to YYYY-MM-DD regardless of Excel number format or ISO string
- Do NOT invent data. If a field has no value, use "" for strings and 0.0 for numbers
- Return the rows in the same order as the input table

INPUT TABLE:
{grid_text}

Return ONLY a compact JSON array (no whitespace between elements, each object on one line). No backticks, no commentary. Example format: [{{"sl_no":1,"issue_date":"2026-05-01",...}},{{"sl_no":2,...}}]
"""


# ---------------------------------------------------------------------------
# Main parser
# ---------------------------------------------------------------------------

def parse_invoice_xlsx(
    xlsx_path: Path,
    cache_dir: Optional[Path] = None,
    force: bool = False
) -> dict:
    """
    Parse an invoice xlsx using AI. Returns:
    {
      "rows": [...],
      "model_version": "...",
      "cost_usd": float or None,
      "latency_sec": float,
      "input_tokens": int or None,
      "output_tokens": int or None,
      "data_rows_sent": int,
      "cache_hit": bool,
      "sha256": str
    }
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"xlsx not found: {xlsx_path}")

    sha = _xlsx_sha256(xlsx_path)

    if cache_dir is None:
        cache_dir = xlsx_path.parent / "ai-poc-cache"
    cache_dir = Path(cache_dir)
    cache_dir.mkdir(parents=True, exist_ok=True)
    cache_file = cache_dir / f"{sha}.json"

    if cache_file.exists() and not force:
        print(f"  [ai_parser] Cache HIT: {sha[:16]}...", file=sys.stderr)
        with open(cache_file) as f:
            cached = json.load(f)
        cached["cache_hit"] = True
        return cached

    print(f"  [ai_parser] Building grid from {xlsx_path.name}...", file=sys.stderr)
    grid_text, data_row_count = _xlsx_to_grid_text(xlsx_path)
    prompt = _EXTRACTION_PROMPT.format(grid_text=grid_text)

    print(f"  [ai_parser] Grid: {data_row_count} data rows, prompt ~{len(prompt)} chars", file=sys.stderr)
    print(f"  [ai_parser] Calling Gemini...", file=sys.stderr)
    t0 = time.time()
    raw_text, model_ver = _call_gemini(prompt)
    latency = time.time() - t0

    if raw_text is None:
        raise RuntimeError("All Gemini models failed — cannot parse invoice")

    # Robust JSON extraction: find outermost [ ... ] in the full raw response
    # This handles all fence variations, code blocks, leading/trailing text
    start = raw_text.find("[")
    end = raw_text.rfind("]")
    if start == -1:
        raise ValueError("No JSON array in response (no [ found)")
    if end == -1:
        # Response truncated before closing ] — repair by closing at last complete object
        import sys as _sys
        print("  [ai_parser] WARNING: truncated response. Repairing...", file=_sys.stderr)
        last_close = raw_text.rfind("}")
        if last_close == -1:
            raise ValueError("Response truncated: no complete objects found.")
        json_text = raw_text[start:last_close+1] + "\n]"
        print(f"  [ai_parser] Repaired: {last_close+1}/{len(raw_text)} chars", file=_sys.stderr)
    else:
        json_text = raw_text[start:end+1]

    try:
        rows = json.loads(json_text)
    except json.JSONDecodeError as e:
        raise ValueError(f"JSON parse error: {e}\nFirst 500 chars: {json_text[:500]}")

    # Schema validation
    valid, errors = _validate_schema(rows)
    if not valid:
        print(f"  [ai_parser] WARNING: Schema validation errors ({len(errors)}):", file=sys.stderr)
        for err in errors[:10]:
            print(f"    {err}", file=sys.stderr)

    # Cost estimation: Gemini 2.5 Pro pricing ~$1.25/1M input, $10/1M output (as of 2026)
    input_chars = len(prompt)
    output_chars = len(raw_text)
    # Rough token estimate: 4 chars/token
    input_tok = input_chars // 4
    output_tok = output_chars // 4
    cost_usd = (input_tok * 1.25 + output_tok * 10.0) / 1_000_000

    result = {
        "rows": rows,
        "model_version": model_ver,
        "cost_usd": round(cost_usd, 5),
        "latency_sec": round(latency, 2),
        "input_tokens": input_tok,
        "output_tokens": output_tok,
        "data_rows_sent": data_row_count,
        "cache_hit": False,
        "sha256": sha,
        "parsed_at": datetime.utcnow().isoformat() + "Z",
        "validation_errors": errors if not valid else [],
    }

    with open(cache_file, "w") as f:
        json.dump(result, f, indent=2, ensure_ascii=False)
    print(f"  [ai_parser] Cached to {cache_file.name}", file=sys.stderr)

    return result


# ---------------------------------------------------------------------------
# CLI entrypoint
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser(description="AI Document Parser POC for AlJeel invoices")
    ap.add_argument("xlsx", help="Invoice xlsx path")
    ap.add_argument("--cache-dir", default=None)
    ap.add_argument("--force", action="store_true", help="Ignore cache, re-parse")
    ap.add_argument("--raw-out", default=None, help="Write raw parsed rows JSON to this file")
    args = ap.parse_args()

    result = parse_invoice_xlsx(
        Path(args.xlsx),
        cache_dir=Path(args.cache_dir) if args.cache_dir else None,
        force=args.force
    )

    print(f"\n=== AI Document Parser POC Results ===")
    print(f"  File:          {args.xlsx}")
    print(f"  SHA256:        {result['sha256'][:16]}...")
    print(f"  Cache hit:     {result['cache_hit']}")
    print(f"  Model:         {result['model_version']}")
    print(f"  Rows extracted:{len(result['rows'])}")
    print(f"  Data rows sent:{result['data_rows_sent']}")
    print(f"  Latency:       {result['latency_sec']}s")
    print(f"  Cost (est.):   ${result['cost_usd']:.5f}")
    print(f"  Input tokens:  ~{result['input_tokens']}")
    print(f"  Output tokens: ~{result['output_tokens']}")
    if result.get("validation_errors"):
        print(f"  Validation:    {len(result['validation_errors'])} errors")
        for e in result["validation_errors"][:5]:
            print(f"    - {e}")
    else:
        print(f"  Validation:    ✓ all rows valid")

    # Check service_date coverage
    rows_with_svc = sum(1 for r in result["rows"] if r.get("service_date") and r["service_date"] != "")
    print(f"  service_date:  {rows_with_svc}/{len(result['rows'])} rows populated")

    if args.raw_out:
        Path(args.raw_out).write_text(json.dumps(result, indent=2, ensure_ascii=False))
        print(f"  Raw output:    {args.raw_out}")
