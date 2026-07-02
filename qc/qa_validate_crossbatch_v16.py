#!/usr/bin/env python3
"""
Independent QA validation script for AlJeel AP pipeline cross-batch benchmark claims.
Written from scratch — does NOT use score_against_truth.py.

Claims to validate:
| Batch    | Rows | Baseline All-5 | v16 All-5 | v16 CC gaps | v16 Agency gaps |
|----------|------|----------------|-----------|-------------|-----------------|
| J26-550  |  72  | 80.6%          | 80.6%     | 5           | 5               |
| J26-589  | 129  | 62.0%          | 65.9%     | 7           | 11              |
| J26-593  | 160  | 74.4%          | 81.2%     | 22          | 27              |
| J26-640  | 117  | 100.0%         | 100.0%    | 0           | 0               |
"""

import re
import sys
import openpyxl
from pathlib import Path
from typing import Optional

# ─── Normalization ────────────────────────────────────────────────────────────

def norm_val(v) -> str:
    """Normalize a cell value: strip whitespace, remove leading zeros, treat None/0/empty as '0'."""
    if v is None:
        return "0"
    s = str(v).strip()
    if s == "" or s.lower() in ("none", "n/a", "na"):
        return "0"
    # Try to strip leading zeros from numeric-looking strings
    # But keep strings like "A1234" as-is
    try:
        # If it's a float/int representation
        f = float(s)
        if f == int(f):
            s = str(int(f))
        else:
            s = str(f)
    except (ValueError, TypeError):
        # Not numeric — just remove any leading zeros if all digits
        if re.match(r'^\d+$', s):
            s = str(int(s))
    return s if s else "0"

def extract_ticket(val) -> Optional[str]:
    """Extract a canonical ticket identifier from a cell value.
    
    Supports:
    - 10-digit numbers (e.g. 1234567890)
    - 26-NNN pattern (e.g. 26-550, 26-589)
    - Strings containing either pattern
    Returns the matched group or None.
    """
    if val is None:
        return None
    s = str(val).strip()
    # Try 10-digit number first
    m = re.search(r'\b(\d{10})\b', s)
    if m:
        return m.group(1)
    # Try 26-NNN pattern (batch-specific ticket like 26-550-0001 or just 26-550)
    m = re.search(r'(26-\d{3,}(?:-\d+)*)', s)
    if m:
        return m.group(1)
    # Fallback: any long number (>= 7 digits)
    m = re.search(r'\b(\d{7,})\b', s)
    if m:
        return m.group(1)
    return None

# ─── Truth file reader ─────────────────────────────────────────────────────────

def read_truth(path: str) -> dict:
    """
    Read truth file (AlJeel ground truth - Details tab).
    
    Column layout (1-indexed, as told):
      1=Sl#, 4=Ticket, 17=Account, 19=CC, 21=DIV, 23=Solution, 25=Agency
    
    Returns dict: ticket_key -> {account, cc, div, solution, agency, row_num, raw_ticket}
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    
    # Find the Details sheet
    sheet_name = None
    for name in wb.sheetnames:
        if name.lower() == "details":
            sheet_name = name
            break
    if sheet_name is None:
        # Fall back to first sheet
        sheet_name = wb.sheetnames[0]
        print(f"  [WARN] No 'Details' sheet found in {path}, using '{sheet_name}'")
    
    ws = wb[sheet_name]
    rows_data = {}
    header_found = False
    header_row = None
    
    # Scan all rows
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not any(c is not None for c in row):
            continue
        
        # Try to detect header row by looking for "Sl#" or similar in col 1
        if not header_found:
            col1 = str(row[0]).strip().lower() if row[0] is not None else ""
            if col1 in ("sl#", "sl", "serial", "#", "no", "sno", "s.no"):
                header_found = True
                header_row = row_idx
                print(f"  [INFO] Truth header found at row {row_idx}: {[str(c)[:20] for c in row[:5]]}")
                continue
            # Also check if col 1 looks like a number (data row without explicit header)
            # Some sheets start data immediately
            if re.match(r'^\d+$', col1):
                header_found = True
                header_row = row_idx
                print(f"  [INFO] Truth data starts at row {row_idx} (no header detected)")
                # Process this row as data too
        
        if not header_found:
            continue
        
        # Extract fields (0-indexed from tuple)
        # Col 1=Sl# (idx 0), 4=Ticket (idx 3), 17=Account (idx 16), 
        # 19=CC (idx 18), 21=DIV (idx 20), 23=Solution (idx 22), 25=Agency (idx 24)
        def safe_get(r, idx):
            return r[idx] if idx < len(r) else None
        
        sl = safe_get(row, 0)
        ticket_raw = safe_get(row, 3)
        account = safe_get(row, 16)
        cc = safe_get(row, 18)
        div = safe_get(row, 20)
        solution = safe_get(row, 22)
        agency = safe_get(row, 24)
        
        # Skip rows with no Sl# or no ticket
        if sl is None or ticket_raw is None:
            continue
        
        # Skip header-like rows
        sl_str = str(sl).strip().lower()
        if sl_str in ("sl#", "sl", "serial", "#", "no", "sno") or not re.match(r'^\d+', sl_str):
            continue
        
        ticket_key = extract_ticket(ticket_raw)
        if ticket_key is None:
            # Try raw normalization
            ticket_key = norm_val(ticket_raw)
            if ticket_key == "0":
                continue
        
        rows_data[ticket_key] = {
            "account": norm_val(account),
            "cc": norm_val(cc),
            "div": norm_val(div),
            "solution": norm_val(solution),
            "agency": norm_val(agency),
            "row_num": row_idx,
            "raw_ticket": ticket_raw,
            "raw_account": account,
            "raw_cc": cc,
            "raw_div": div,
            "raw_solution": solution,
            "raw_agency": agency,
        }
    
    wb.close()
    print(f"  [INFO] Truth: {len(rows_data)} rows loaded from {Path(path).name}")
    return rows_data


# ─── Pipeline output reader ────────────────────────────────────────────────────

def read_pipeline(path: str) -> dict:
    """
    Read v16 pipeline output FILLED xlsx.
    
    Find header row containing "*Invoice Header Identifier" (or similar),
    then extract columns by header name:
      Description=col 11 (idx 10), Account=19 (idx 18), CC=21 (idx 20),
      DIV=23 (idx 22), Solution=25 (idx 24), Agency=27 (idx 26)
    
    Also try to find ticket from Description column.
    
    Returns dict: ticket_key -> {account, cc, div, solution, agency, row_num, raw_desc}
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    
    rows_data = {}
    header_row_idx = None
    col_map = {}  # name -> 0-based column index
    
    all_rows = list(ws.iter_rows(values_only=True))
    
    # Find header row
    for row_idx, row in enumerate(all_rows):
        row_str = " ".join(str(c).lower() for c in row if c is not None)
        if "invoice header identifier" in row_str or "*invoice" in row_str:
            header_row_idx = row_idx
            # Build column map
            for col_i, cell in enumerate(row):
                if cell is not None:
                    key = str(cell).strip().lower().lstrip("*")
                    col_map[key] = col_i
            print(f"  [INFO] Pipeline header found at row {row_idx + 1}")
            print(f"  [INFO] Column map (first 30): {dict(list(col_map.items())[:30])}")
            break
    
    if header_row_idx is None:
        print(f"  [WARN] No header row found in pipeline output {path}")
        # Try positional fallback
        header_row_idx = 0
    
    # Map known field names to column indices
    # The task says: Description=11, Account=19, CC=21, DIV=23, Solution=25, Agency=27 (1-indexed)
    # = indices 10, 18, 20, 22, 24, 26 (0-indexed)
    POSITIONAL_FALLBACK = {
        "description": 10,
        "account": 18,
        "cc": 20,
        "div": 22,
        "solution": 24,
        "agency": 26,
    }
    
    def get_col(name: str) -> int:
        """Look up column index by name, fall back to positional."""
        # Try various name variants
        for variant in [name, name.lower(), name.upper(), f"*{name}"]:
            if variant.lower().lstrip("*") in col_map:
                return col_map[variant.lower().lstrip("*")]
        return POSITIONAL_FALLBACK.get(name.lower(), -1)
    
    desc_col = get_col("description")
    account_col = get_col("account")
    cc_col = get_col("cc")
    div_col = get_col("div")
    solution_col = get_col("solution")
    agency_col = get_col("agency")
    
    print(f"  [INFO] Pipeline column indices: desc={desc_col}, account={account_col}, cc={cc_col}, div={div_col}, sol={solution_col}, agency={agency_col}")
    
    # Process data rows
    for row_idx in range(header_row_idx + 1, len(all_rows)):
        row = all_rows[row_idx]
        if not any(c is not None for c in row):
            continue
        
        def safe_get(r, idx):
            return r[idx] if 0 <= idx < len(r) else None
        
        desc = safe_get(row, desc_col)
        account = safe_get(row, account_col)
        cc = safe_get(row, cc_col)
        div = safe_get(row, div_col)
        solution = safe_get(row, solution_col)
        agency = safe_get(row, agency_col)
        
        # Extract ticket from description
        ticket_key = extract_ticket(desc)
        if ticket_key is None:
            # Try other columns for ticket
            for col_idx in range(min(15, len(row))):
                ticket_key = extract_ticket(safe_get(row, col_idx))
                if ticket_key:
                    break
        
        if ticket_key is None:
            continue
        
        rows_data[ticket_key] = {
            "account": norm_val(account),
            "cc": norm_val(cc),
            "div": norm_val(div),
            "solution": norm_val(solution),
            "agency": norm_val(agency),
            "row_num": row_idx + 1,
            "raw_desc": desc,
            "raw_account": account,
            "raw_cc": cc,
            "raw_div": div,
            "raw_solution": solution,
            "raw_agency": agency,
        }
    
    wb.close()
    print(f"  [INFO] Pipeline: {len(rows_data)} rows loaded from {Path(path).name}")
    return rows_data


# ─── Scoring ───────────────────────────────────────────────────────────────────

FIELDS = ["account", "cc", "div", "solution", "agency"]

def score_batch(truth: dict, pipeline: dict, batch_name: str) -> dict:
    """
    Compare truth vs pipeline for a batch.
    Returns scoring summary.
    """
    # Find matched tickets
    truth_keys = set(truth.keys())
    pipeline_keys = set(pipeline.keys())
    matched_keys = truth_keys & pipeline_keys
    only_truth = truth_keys - pipeline_keys
    only_pipeline = pipeline_keys - truth_keys
    
    print(f"\n  [MATCH] {batch_name}: truth={len(truth_keys)}, pipeline={len(pipeline_keys)}, "
          f"matched={len(matched_keys)}, only_truth={len(only_truth)}, only_pipeline={len(only_pipeline)}")
    
    if only_truth:
        sample = list(only_truth)[:5]
        print(f"  [WARN] Keys in truth but not pipeline (sample): {sample}")
    if only_pipeline:
        sample = list(only_pipeline)[:5]
        print(f"  [WARN] Keys in pipeline but not truth (sample): {sample}")
    
    # Per-field counters
    field_matches = {f: 0 for f in FIELDS}
    all5_match = 0
    cc_gaps = []
    agency_gaps = []
    mismatches = []  # For spot-check
    
    for key in sorted(matched_keys):
        t = truth[key]
        p = pipeline[key]
        
        row_fields_match = {}
        for f in FIELDS:
            row_fields_match[f] = (t[f] == p[f])
            if row_fields_match[f]:
                field_matches[f] += 1
        
        if all(row_fields_match.values()):
            all5_match += 1
        else:
            if not row_fields_match["cc"]:
                cc_gaps.append(key)
            if not row_fields_match["agency"]:
                agency_gaps.append(key)
            mismatches.append({
                "key": key,
                "fields": row_fields_match,
                "truth": {f: t[f] for f in FIELDS},
                "pipeline": {f: p[f] for f in FIELDS},
                "truth_raw": {f: t.get(f"raw_{f}") for f in FIELDS},
                "pipeline_raw": {f: p.get(f"raw_{f}") for f in FIELDS},
                "truth_row": t["row_num"],
                "pipeline_row": p["row_num"],
            })
    
    n = len(matched_keys)
    all5_pct = (all5_match / n * 100) if n > 0 else 0
    
    result = {
        "batch": batch_name,
        "truth_rows": len(truth_keys),
        "pipeline_rows": len(pipeline_keys),
        "matched_rows": n,
        "all5_match": all5_match,
        "all5_pct": all5_pct,
        "cc_gaps": len(cc_gaps),
        "agency_gaps": len(agency_gaps),
        "cc_gap_keys": cc_gaps,
        "agency_gap_keys": agency_gaps,
        "field_matches": field_matches,
        "mismatches": mismatches,
        "only_truth": list(only_truth)[:10],
        "only_pipeline": list(only_pipeline)[:10],
    }
    return result


# ─── Spot-check ───────────────────────────────────────────────────────────────

def spot_check(truth: dict, pipeline: dict, batch_name: str, scoring: dict):
    """Print detailed field-by-field comparison for selected rows."""
    print(f"\n{'='*70}")
    print(f"SPOT CHECK: {batch_name}")
    print(f"{'='*70}")
    
    matched = set(truth.keys()) & set(pipeline.keys())
    
    # Pick a "perfect" row (all 5 match)
    perfect_key = None
    for key in sorted(matched):
        t = truth[key]
        p = pipeline[key]
        if all(t[f] == p[f] for f in FIELDS):
            perfect_key = key
            break
    
    # Pick a CC mismatch row
    cc_mismatch_key = scoring["cc_gap_keys"][0] if scoring["cc_gap_keys"] else None
    
    # Pick an agency mismatch row
    agency_mismatch_key = scoring["agency_gap_keys"][0] if scoring["agency_gap_keys"] else None
    
    for label, key in [("PERFECT (all 5 match)", perfect_key),
                        ("CC MISMATCH", cc_mismatch_key),
                        ("AGENCY MISMATCH", agency_mismatch_key)]:
        if key is None or key not in truth or key not in pipeline:
            print(f"\n  [{label}] No suitable row found")
            continue
        
        t = truth[key]
        p = pipeline[key]
        print(f"\n  [{label}] Ticket: {key}")
        print(f"  Truth row: {t['row_num']}, Pipeline row: {p['row_num']}")
        print(f"  {'Field':<12} {'Truth (raw)':<30} {'Truth (norm)':<20} {'Pipeline (raw)':<30} {'Pipeline (norm)':<20} {'Match'}")
        print(f"  {'-'*115}")
        for f in FIELDS:
            match_sym = "✓" if t[f] == p[f] else "✗"
            tr = str(t.get(f"raw_{f}"))[:28]
            tn = t[f][:18]
            pr = str(p.get(f"raw_{f}"))[:28]
            pn = p[f][:18]
            print(f"  {f:<12} {tr:<30} {tn:<20} {pr:<30} {pn:<20} {match_sym}")


# ─── Main ─────────────────────────────────────────────────────────────────────

BATCHES = [
    {
        "name": "J26-550",
        "truth": "/home/clawdbot/.openclaw/workspace/aljeel/batches/jawal-J26-550/J26-550.xlsx",
        "pipeline": "/home/clawdbot/.openclaw/workspace/aljeel/batches/jawal-J26-550/output/Spreadsheet-J26-550-FILLED-v16.xlsx",
        "claimed_rows": 72,
        "claimed_baseline_all5": 80.6,
        "claimed_v16_all5": 80.6,
        "claimed_cc_gaps": 5,
        "claimed_agency_gaps": 5,
    },
    {
        "name": "J26-589",
        "truth": "/home/clawdbot/.openclaw/workspace/aljeel/batches/jawal-J26-589/J26-589.xlsx",
        "pipeline": "/home/clawdbot/.openclaw/workspace/aljeel/batches/jawal-J26-589/output/Spreadsheet-J26-589-FILLED-v16.xlsx",
        "claimed_rows": 129,
        "claimed_baseline_all5": 62.0,
        "claimed_v16_all5": 65.9,
        "claimed_cc_gaps": 7,
        "claimed_agency_gaps": 11,
    },
    {
        "name": "J26-593",
        "truth": "/home/clawdbot/.openclaw/workspace/aljeel/batches/jawal-J26-593/J26-593.xlsx",
        "pipeline": "/home/clawdbot/.openclaw/workspace/aljeel/batches/jawal-J26-593/output/Spreadsheet-J26-593-FILLED-v16.xlsx",
        "claimed_rows": 160,
        "claimed_baseline_all5": 74.4,
        "claimed_v16_all5": 81.2,
        "claimed_cc_gaps": 22,
        "claimed_agency_gaps": 27,
    },
    {
        "name": "J26-640",
        "truth": "/home/clawdbot/.openclaw/workspace/aljeel/batches/jawal-J26-640/J26-640.xlsx",
        "pipeline": "/home/clawdbot/.openclaw/workspace/aljeel/batches/jawal-J26-640/output/Spreadsheet-J26-640-FILLED-v16.xlsx",
        "claimed_rows": 117,
        "claimed_baseline_all5": 100.0,
        "claimed_v16_all5": 100.0,
        "claimed_cc_gaps": 0,
        "claimed_agency_gaps": 0,
    },
]

def main():
    results = []
    all_truth = {}
    all_pipeline = {}
    
    for batch in BATCHES:
        print(f"\n{'='*70}")
        print(f"PROCESSING BATCH: {batch['name']}")
        print(f"{'='*70}")
        
        print(f"\nReading truth: {batch['truth']}")
        truth = read_truth(batch["truth"])
        all_truth[batch["name"]] = truth
        
        print(f"\nReading pipeline: {batch['pipeline']}")
        pipeline = read_pipeline(batch["pipeline"])
        all_pipeline[batch["name"]] = pipeline
        
        scoring = score_batch(truth, pipeline, batch["name"])
        results.append((batch, scoring))
    
    # Print summary table
    print(f"\n\n{'='*90}")
    print("SCORING SUMMARY")
    print(f"{'='*90}")
    print(f"{'Batch':<12} {'T.Rows':<8} {'P.Rows':<8} {'Matched':<9} {'All-5':<8} {'All-5%':<10} {'CC gaps':<10} {'Agency gaps':<12}")
    print(f"{'-'*90}")
    
    for batch, s in results:
        print(f"{s['batch']:<12} {s['truth_rows']:<8} {s['pipeline_rows']:<8} {s['matched_rows']:<9} "
              f"{s['all5_match']:<8} {s['all5_pct']:<10.1f} {s['cc_gaps']:<10} {s['agency_gaps']:<12}")
    
    # Compare against claims
    print(f"\n\n{'='*90}")
    print("CLAIM VALIDATION")
    print(f"{'='*90}")
    
    for batch, s in results:
        print(f"\nBatch: {s['batch']}")
        print(f"  Rows:        claimed={batch['claimed_rows']}, actual_truth={s['truth_rows']}, matched={s['matched_rows']}")
        
        rows_ok = abs(s['truth_rows'] - batch['claimed_rows']) <= 1
        v16_pct_ok = abs(s['all5_pct'] - batch['claimed_v16_all5']) <= 1.0
        cc_ok = abs(s['cc_gaps'] - batch['claimed_cc_gaps']) <= 1
        agency_ok = abs(s['agency_gaps'] - batch['claimed_agency_gaps']) <= 1
        
        print(f"  All-5%:      claimed={batch['claimed_v16_all5']}%, actual={s['all5_pct']:.1f}% → {'✓ OK' if v16_pct_ok else '✗ MISMATCH'}")
        print(f"  CC gaps:     claimed={batch['claimed_cc_gaps']}, actual={s['cc_gaps']} → {'✓ OK' if cc_ok else '✗ MISMATCH'}")
        print(f"  Agency gaps: claimed={batch['claimed_agency_gaps']}, actual={s['agency_gaps']} → {'✓ OK' if agency_ok else '✗ MISMATCH'}")
        print(f"  Rows:        claimed={batch['claimed_rows']}, actual={s['truth_rows']} → {'✓ OK' if rows_ok else '✗ MISMATCH'}")
    
    # Spot checks
    print(f"\n\n{'='*90}")
    print("SPOT CHECKS")
    print(f"{'='*90}")
    
    # J26-640 (should be 100%)
    spot_check(all_truth["J26-640"], all_pipeline["J26-640"], "J26-640 (100% batch)", 
               [s for b, s in results if s["batch"] == "J26-640"][0])
    
    # J26-593 (most CC gaps)
    spot_check(all_truth["J26-593"], all_pipeline["J26-593"], "J26-593 (CC & agency gaps)", 
               [s for b, s in results if s["batch"] == "J26-593"][0])
    
    # Return results for report generation
    return results, all_truth, all_pipeline


if __name__ == "__main__":
    results, all_truth, all_pipeline = main()
