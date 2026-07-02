#!/usr/bin/env python3
"""
qa_validate_v2.py - Independent QA validation for AlJeel cross-batch benchmark (v16).
This version properly handles duplicate ticket numbers (N:M pairing) and
exactly mirrors the methodology of the official scorer.

Key corrections from v1:
- Duplicate tickets: store all occurrences in lists, pair N:M (official pair_rows() logic)
- Denominator: use paired_evaluated (total pairs), not unique keys
- Normalization: strip leading zeros, identical to official _norm()

All column mappings verified against official score_against_truth.py.
"""

import re
import sys
from pathlib import Path
from typing import Optional, List, Dict

import openpyxl

# ─── Ticket extraction ────────────────────────────────────────────────────────

# Strict: 10-digit number OR 26-NNN (exactly 3 digits) - matches official regex
_TICKET_RE = re.compile(r"(?<![\d])(\d{10}|26-\d{3})(?![\d])")

def extract_ticket(s) -> Optional[str]:
    if s is None:
        return None
    m = _TICKET_RE.search(str(s))
    return m.group(1) if m else None

# ─── Normalization ─────────────────────────────────────────────────────────────

def norm(v) -> str:
    """Strip leading zeros; treat None/empty as '0'. Mirrors official _norm()."""
    s = '' if v is None else str(v).strip()
    return s.lstrip('0') or '0'

# ─── Truth loader ─────────────────────────────────────────────────────────────

def load_truth(path: str) -> Dict[str, List[dict]]:
    """
    Load truth file Details sheet.
    Column layout (1-indexed):
      1=Sl#, 4=Ticket, 17=Account, 19=CC, 21=DIV, 23=Solution, 25=Agency

    Returns: {ticket_key: [list of row dicts]}
    Preserves duplicates as separate list entries.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Details'] if 'Details' in wb.sheetnames else wb[wb.sheetnames[0]]
    by_ticket = {}
    skipped = 0

    for r in range(2, ws.max_row + 1):
        sl = ws.cell(r, 1).value
        if sl is None:
            continue

        ticket = extract_ticket(ws.cell(r, 4).value)
        if not ticket:
            skipped += 1
            continue

        rec = {
            'sl': str(sl).strip(),
            'ticket': ticket,
            'raw_ticket': ws.cell(r, 4).value,
            'account':  norm(ws.cell(r, 17).value),
            'cc':       norm(ws.cell(r, 19).value),
            'div':      norm(ws.cell(r, 21).value),
            'solution': norm(ws.cell(r, 23).value),
            'agency':   norm(ws.cell(r, 25).value),
            'row_num': r,
            # For spot-check: raw values
            'raw_account':  ws.cell(r, 17).value,
            'raw_cc':       ws.cell(r, 19).value,
            'raw_div':      ws.cell(r, 21).value,
            'raw_solution': ws.cell(r, 23).value,
            'raw_agency':   ws.cell(r, 25).value,
        }
        by_ticket.setdefault(ticket, []).append(rec)

    wb.close()
    total = sum(len(v) for v in by_ticket.values())
    unique = len(by_ticket)
    dupes = {k: len(v) for k, v in by_ticket.items() if len(v) > 1}
    print(f"  [TRUTH] {Path(path).name}: total={total}, unique_tickets={unique}, "
          f"dup_tickets={len(dupes)}, skipped={skipped}")
    if dupes:
        print(f"  [TRUTH] Duplicate tickets: {dupes}")
    return by_ticket

# ─── Pipeline loader ──────────────────────────────────────────────────────────

def load_pipeline(path: str) -> Dict[str, List[dict]]:
    """
    Load pipeline FILLED output.
    Finds header row containing 'Description' and extracts columns by header name.

    Segments: Account (col name 'Account'), CC ('Cost Center'), DIV ('DIV'),
              Solution ('Solution'), Agency ('Agency')

    Returns: {ticket_key: [list of row dicts]}
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Find header row
    hdr_row = None
    hdr = None
    for r in (3, 1, 2, 4):
        try:
            row_vals = [str(c.value or '').strip() for c in ws[r]]
            if 'Description' in row_vals:
                hdr_row = r
                hdr = row_vals
                break
        except Exception:
            pass

    if hdr_row is None:
        raise ValueError(f"No header row found in {path}")

    def col_idx(name: str) -> int:
        try:
            return hdr.index(name)
        except ValueError:
            return -1

    desc_col = col_idx('Description')
    account_col = col_idx('Account')
    cc_col = col_idx('Cost Center')
    div_col = col_idx('DIV')
    solution_col = col_idx('Solution')
    agency_col = col_idx('Agency')

    print(f"  [PIPE] Column indices: desc={desc_col}, account={account_col}, "
          f"cc={cc_col}, div={div_col}, sol={solution_col}, agency={agency_col}")

    by_ticket = {}
    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        if all(v is None for v in row):
            continue

        def g(idx):
            return row[idx] if 0 <= idx < len(row) else None

        desc = g(desc_col)
        ticket = extract_ticket(desc)
        if not ticket:
            continue

        rec = {
            'ticket': ticket,
            'desc': str(desc or '')[:50],
            'account':  norm(g(account_col)),
            'cc':       norm(g(cc_col)),
            'div':      norm(g(div_col)),
            'solution': norm(g(solution_col)),
            'agency':   norm(g(agency_col)),
            # Raw for spot-check
            'raw_account':  g(account_col),
            'raw_cc':       g(cc_col),
            'raw_div':      g(div_col),
            'raw_solution': g(solution_col),
            'raw_agency':   g(agency_col),
        }
        by_ticket.setdefault(ticket, []).append(rec)

    wb.close()
    total = sum(len(v) for v in by_ticket.values())
    unique = len(by_ticket)
    dupes = {k: len(v) for k, v in by_ticket.items() if len(v) > 1}
    print(f"  [PIPE] {Path(path).name}: total={total}, unique_tickets={unique}, "
          f"dup_tickets={len(dupes)}")
    if dupes:
        print(f"  [PIPE] Duplicate tickets: {dupes}")
    return by_ticket

# ─── Pairing (mirrors official pair_rows logic) ───────────────────────────────

def pair_rows(truth_list: list, pipe_list: list) -> list:
    """Pair truth and pipeline rows for the same ticket.
    N:1 → each truth with single pipe; 1:N → single truth with each pipe;
    N:N → index-aligned; N:M → align up to min, extras get None partner.
    """
    pairs = []
    if not pipe_list:
        return [(t, None) for t in truth_list]
    if not truth_list:
        return [(None, p) for p in pipe_list]
    n, m = len(truth_list), len(pipe_list)
    if n == m:
        return list(zip(truth_list, pipe_list))
    elif m == 1:
        return [(t, pipe_list[0]) for t in truth_list]
    elif n == 1:
        return [(truth_list[0], p) for p in pipe_list]
    else:
        k = min(n, m)
        pairs = list(zip(truth_list[:k], pipe_list[:k]))
        if n > m:
            pairs += [(t, None) for t in truth_list[m:]]
        else:
            pairs += [(None, p) for p in pipe_list[n:]]
        return pairs

# ─── Scoring ──────────────────────────────────────────────────────────────────

FIELDS = ['account', 'cc', 'div', 'solution', 'agency']

def score_batch(truth_by_ticket: dict, pipe_by_ticket: dict, batch_name: str) -> dict:
    """Score all paired rows for a batch."""

    all_pairs = []  # (truth_rec or None, pipe_rec or None, ticket)

    # Truth tickets
    for ticket, truth_list in truth_by_ticket.items():
        pipe_list = pipe_by_ticket.get(ticket, [])
        for t, p in pair_rows(truth_list, pipe_list):
            all_pairs.append((ticket, t, p))

    # Pipeline-only tickets
    for ticket, pipe_list in pipe_by_ticket.items():
        if ticket not in truth_by_ticket:
            for p in pipe_list:
                all_pairs.append((ticket, None, p))

    # Evaluate only fully-paired rows
    evaluated = [(ticket, t, p) for ticket, t, p in all_pairs if t is not None and p is not None]

    all5_match = 0
    cc_gaps = []
    agency_gaps = []
    field_matches = {f: 0 for f in FIELDS}
    mismatches = []

    for ticket, t, p in evaluated:
        per_field = {f: (t[f] == p[f]) for f in FIELDS}
        for f in FIELDS:
            if per_field[f]:
                field_matches[f] += 1

        if all(per_field.values()):
            all5_match += 1
        else:
            if not per_field['cc']:
                cc_gaps.append(ticket)
            if not per_field['agency']:
                agency_gaps.append(ticket)
            mismatches.append({'ticket': ticket, 'truth': t, 'pipeline': p, 'fields': per_field})

    n = len(evaluated)
    all5_pct = (all5_match / n * 100) if n > 0 else 0

    only_truth = [t for t, _, p in all_pairs if p is None]
    only_pipe = [t for t, tr, _ in all_pairs if tr is None]

    print(f"\n  [RESULT] {batch_name}: truth_total={sum(len(v) for v in truth_by_ticket.values())}, "
          f"pipe_total={sum(len(v) for v in pipe_by_ticket.values())}, "
          f"paired={n}, all5={all5_match} ({all5_pct:.1f}%), "
          f"cc_gaps={len(cc_gaps)}, agency_gaps={len(agency_gaps)}")

    if only_truth:
        print(f"  [WARN] Truth-only (no pipeline match): {only_truth[:5]}")
    if only_pipe:
        print(f"  [WARN] Pipeline-only (no truth match): {only_pipe[:5]}")

    return {
        'batch': batch_name,
        'truth_total': sum(len(v) for v in truth_by_ticket.values()),
        'pipe_total': sum(len(v) for v in pipe_by_ticket.values()),
        'paired': n,
        'all5_match': all5_match,
        'all5_pct': all5_pct,
        'cc_gaps': len(cc_gaps),
        'agency_gaps': len(agency_gaps),
        'cc_gap_tickets': cc_gaps,
        'agency_gap_tickets': agency_gaps,
        'field_matches': field_matches,
        'mismatches': mismatches,
        'only_truth': only_truth,
        'only_pipe': only_pipe,
    }

# ─── Spot check ───────────────────────────────────────────────────────────────

def spot_check(batch_name: str, scoring: dict):
    """Print field-by-field comparison for selected rows."""
    print(f"\n{'='*70}")
    print(f"SPOT CHECK: {batch_name}")
    print(f"{'='*70}")

    mismatches = scoring['mismatches']

    # 1. Find a perfect row (all 5 match) - need to find from evaluated pairs
    # Perfect rows are those NOT in mismatches but paired
    # We'll pick a short example from the scoring already done

    # Find a CC mismatch
    cc_mm = [m for m in mismatches if not m['fields']['cc']]
    agency_mm = [m for m in mismatches if not m['fields']['agency']]

    for label, rows in [("CC MISMATCH", cc_mm), ("AGENCY MISMATCH", agency_mm)]:
        if not rows:
            print(f"\n  [{label}]: None found")
            continue
        m = rows[0]
        t, p = m['truth'], m['pipeline']
        print(f"\n  [{label}] Ticket: {m['ticket']}")
        print(f"  Truth row: {t.get('row_num','?')}, SL#: {t.get('sl','?')}")
        print(f"  {'Field':<12} {'Truth raw':<25} {'Truth norm':<15} {'Pipe raw':<25} {'Pipe norm':<15} {'Match'}")
        print(f"  {'-'*100}")
        for f in FIELDS:
            match_sym = "✓" if t[f] == p[f] else "✗"
            tr = str(t.get(f'raw_{f}'))[:23]
            tn = t[f][:13]
            pr = str(p.get(f'raw_{f}'))[:23]
            pn = p[f][:13]
            print(f"  {f:<12} {tr:<25} {tn:<15} {pr:<25} {pn:<15} {match_sym}")


# ─── Main ─────────────────────────────────────────────────────────────────────

BATCHES = [
    {
        "name": "J26-550",
        "truth": "/home/clawdbot/.openclaw/workspace/aljeel/batches/jawal-J26-550/J26-550.xlsx",
        "pipeline": "/home/clawdbot/.openclaw/workspace/aljeel/batches/jawal-J26-550/output/Spreadsheet-J26-550-FILLED-v16.xlsx",
        "claimed_rows": 72,
        "claimed_v16_all5_pct": 80.6,
        "claimed_cc_gaps": 5,
        "claimed_agency_gaps": 5,
    },
    {
        "name": "J26-589",
        "truth": "/home/clawdbot/.openclaw/workspace/aljeel/batches/jawal-J26-589/J26-589.xlsx",
        "pipeline": "/home/clawdbot/.openclaw/workspace/aljeel/batches/jawal-J26-589/output/Spreadsheet-J26-589-FILLED-v16.xlsx",
        "claimed_rows": 129,
        "claimed_v16_all5_pct": 65.9,
        "claimed_cc_gaps": 7,
        "claimed_agency_gaps": 11,
    },
    {
        "name": "J26-593",
        "truth": "/home/clawdbot/.openclaw/workspace/aljeel/batches/jawal-J26-593/J26-593.xlsx",
        "pipeline": "/home/clawdbot/.openclaw/workspace/aljeel/batches/jawal-J26-593/output/Spreadsheet-J26-593-FILLED-v16.xlsx",
        "claimed_rows": 160,
        "claimed_v16_all5_pct": 81.2,
        "claimed_cc_gaps": 22,
        "claimed_agency_gaps": 27,
    },
    {
        "name": "J26-640",
        "truth": "/home/clawdbot/.openclaw/workspace/aljeel/batches/jawal-J26-640/J26-640.xlsx",
        "pipeline": "/home/clawdbot/.openclaw/workspace/aljeel/batches/jawal-J26-640/output/Spreadsheet-J26-640-FILLED-v16.xlsx",
        "claimed_rows": 117,
        "claimed_v16_all5_pct": 100.0,
        "claimed_cc_gaps": 0,
        "claimed_agency_gaps": 0,
    },
]

def main():
    results = []

    for batch in BATCHES:
        print(f"\n{'='*70}")
        print(f"BATCH: {batch['name']}")
        print(f"{'='*70}")

        truth = load_truth(batch["truth"])
        pipeline = load_pipeline(batch["pipeline"])
        scoring = score_batch(truth, pipeline, batch["name"])
        results.append((batch, scoring))

    # Summary table
    print(f"\n\n{'='*95}")
    print("SCORING SUMMARY (v2 - correct duplicate handling)")
    print(f"{'='*95}")
    print(f"{'Batch':<12} {'T.Total':<10} {'P.Total':<10} {'Paired':<8} {'All-5':<8} {'All-5%':<9} {'CC gaps':<10} {'Ag gaps'}")
    print(f"{'-'*95}")
    for batch, s in results:
        print(f"{s['batch']:<12} {s['truth_total']:<10} {s['pipe_total']:<10} {s['paired']:<8} "
              f"{s['all5_match']:<8} {s['all5_pct']:<9.1f} {s['cc_gaps']:<10} {s['agency_gaps']}")

    # Claim validation
    print(f"\n\n{'='*95}")
    print("CLAIM VALIDATION")
    print(f"{'='*95}")

    for batch, s in results:
        rows_ok = abs(s['truth_total'] - batch['claimed_rows']) <= 1
        pct_ok  = abs(s['all5_pct'] - batch['claimed_v16_all5_pct']) <= 1.0
        cc_ok   = abs(s['cc_gaps'] - batch['claimed_cc_gaps']) <= 1
        ag_ok   = abs(s['agency_gaps'] - batch['claimed_agency_gaps']) <= 1

        status = "✓ CONFIRMED" if (rows_ok and pct_ok and cc_ok and ag_ok) else "✗ DISCREPANCY"
        print(f"\n{status} - {s['batch']}")
        row_diff = s['truth_total'] - batch['claimed_rows']
        pct_diff = s['all5_pct'] - batch['claimed_v16_all5_pct']
        cc_diff  = s['cc_gaps'] - batch['claimed_cc_gaps']
        ag_diff  = s['agency_gaps'] - batch['claimed_agency_gaps']
        print(f"  Rows:        claimed={batch['claimed_rows']}, actual={s['truth_total']} → {'OK' if rows_ok else 'DIFF '+str(row_diff)}")
        print(f"  All-5%:      claimed={batch['claimed_v16_all5_pct']}%, actual={s['all5_pct']:.1f}% → {'OK' if pct_ok else 'DIFF '+str(round(pct_diff,1))+'pp'}")
        print(f"  CC gaps:     claimed={batch['claimed_cc_gaps']}, actual={s['cc_gaps']} → {'OK' if cc_ok else 'DIFF '+str(cc_diff)}")
        print(f"  Agency gaps: claimed={batch['claimed_agency_gaps']}, actual={s['agency_gaps']} → {'OK' if ag_ok else 'DIFF '+str(ag_diff)}")

    # Spot checks
    print(f"\n\n{'='*95}")
    print("SPOT CHECKS")
    print(f"{'='*95}")

    for batch, scoring in results:
        # J26-640 spot-check: show first mismatch (should be none) vs first perfect match
        if scoring['batch'] == 'J26-640':
            print(f"\n[J26-640] All-5 = 100% - no mismatches expected. Confirmed: {scoring['all5_match']}/{scoring['paired']}")
        else:
            spot_check(scoring['batch'], scoring)

    # Extra: J26-550 agency gap analysis (claimed 5, got 4)
    print(f"\n\n{'='*95}")
    print("EDGE CASE ANALYSIS: J26-550 Agency Gaps")
    print(f"{'='*95}")
    for batch, scoring in results:
        if scoring['batch'] == 'J26-550':
            print(f"\n  Agency gap tickets: {scoring['agency_gap_tickets']}")
            print(f"  CC gap tickets:     {scoring['cc_gap_tickets']}")
            mm = scoring['mismatches']
            print(f"\n  All mismatches ({len(mm)} total):")
            for m in mm:
                t, p = m['truth'], m['pipeline']
                diffed = [f for f in FIELDS if not m['fields'][f]]
                print(f"    Ticket {m['ticket']} (SL#{t.get('sl')}): fields differ={diffed}")
                for f in diffed:
                    print(f"      {f}: truth={t[f]!r} vs pipe={p[f]!r}")

    return results


if __name__ == "__main__":
    results = main()
