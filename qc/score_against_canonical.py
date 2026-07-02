#!/usr/bin/env python3
"""
score_against_canonical.py — robust canonical comparison utility for AlJeel AP pipeline.

CRITICAL: This script matches rows by TICKET NUMBER (stable identifier), NOT by Sl#/row position.
The pipeline's output row order is NOT guaranteed to match the canonical file's row order — they
sort differently. Comparing by row index will produce false mismatches even when the data is correct.

Lesson from 2026-05-22 v15.6/v15.7 debug:
  - First scoring attempt used Sl# (1:1 row index) → reported 84.3% location match (false low)
  - Re-scored by ticket → actual was 97.4%
  - Difference: 18 false-positive mismatches caused by sort-order drift, NOT real bugs.

Usage:
  python3 score_against_canonical.py <pipeline_xlsx> <canonical_xlsx>

Optional flags:
  --json          Output JSON instead of human-readable summary
  --threshold     Print PASS/FAIL against per-segment minimum match rates
  --verbose       Show every mismatch row

The canonical file's structure (AlJeel-prepared):
  - Sheet 'Sheet1' (or 'Details' for older versions)
  - Row 1 = headers
  - Row 2+ = data
  - Col 0 (A): Sl. #
  - Col 3 (D): Ticket Number (sometimes "065 6905264364" — extract 10-digit)
  - Col 4 (E): Pax Name
  - Col 5 (F): Route
  - Col 14 (O): Company
  - Col 15 (P): location
  - Col 16 (Q): Account
  - Col 18 (S): Cost Center
  - Col 20 (U): DIV
  - Col 22 (W): Solution
  - Col 24 (Y): Agency

The pipeline output structure (v15.x):
  - Default sheet (Sheet1)
  - Row 3 = headers (rows 1-2 are metadata/block labels)
  - Row 4+ = data
  - Headers found by NAME, not position (resilient to column reorderings):
    Description, Company, Location, Account, Cost Center, DIV, Solution, Agency
  - Ticket lives inside 'Description' as a 10-digit or 26-NNN suffix
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(2)


# ──────────────────────────────────────────────────────────────────────────────
# Ticket extraction (stable identifier across both files)
# ──────────────────────────────────────────────────────────────────────────────

# 10-digit airline ticket OR 26-NNN voucher (Aljeel sponsorship reference)
_TICKET_RE = re.compile(r"(?<![\d])(\d{10}|26-\d{3})(?![\d])")


def extract_ticket(*sources) -> str | None:
    """Find ticket number in any of the given strings. Returns None if no match."""
    for s in sources:
        if s is None:
            continue
        text = str(s)
        m = _TICKET_RE.search(text)
        if m:
            return m.group(1)
    return None


# ──────────────────────────────────────────────────────────────────────────────
# Canonical loader (AlJeel-prepared format)
# ──────────────────────────────────────────────────────────────────────────────

def load_canonical(path: Path) -> dict[str, dict]:
    """Load canonical xlsx into {ticket: {segments + metadata}} keyed by ticket."""
    wb = openpyxl.load_workbook(path, data_only=True)
    # Try common sheet names
    ws = None
    for sn in ('Sheet1', 'Details'):
        if sn in wb.sheetnames:
            ws = wb[sn]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    by_ticket: dict[str, dict] = {}
    skipped_no_ticket = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        # Ticket lives in col D (idx 3), but also in pax/route/notes
        ticket = extract_ticket(
            row[3] if len(row) > 3 else None,
            row[4] if len(row) > 4 else None,
            row[12] if len(row) > 12 else None,
        )
        if not ticket:
            skipped_no_ticket += 1
            continue

        by_ticket[ticket] = {
            'sl': str(row[0]).strip(),
            'pax': str(row[4] or '').strip()[:40] if len(row) > 4 else '',
            'route': str(row[5] or '').strip()[:30] if len(row) > 5 else '',
            'company': str(row[14] or '').strip() if len(row) > 14 else '',
            'location': str(row[15] or '').strip() if len(row) > 15 else '',
            'account': str(row[16] or '').strip() if len(row) > 16 else '',
            'cc': str(row[18] or '').strip() if len(row) > 18 else '',
            'div': str(row[20] or '').strip() if len(row) > 20 else '',
            'solution': str(row[22] or '').strip() if len(row) > 22 else '',
            'agency': str(row[24] or '').strip() if len(row) > 24 else '',
        }

    return by_ticket, skipped_no_ticket


# ──────────────────────────────────────────────────────────────────────────────
# Pipeline output loader (v15.x format, header-name based)
# ──────────────────────────────────────────────────────────────────────────────

def load_pipeline(path: Path) -> dict[str, dict]:
    """Load pipeline xlsx into {ticket: {segments}} keyed by ticket."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Headers are in row 3 (rows 1-2 are metadata/block-label rows)
    hdr_row_idx = None
    for candidate in (3, 1, 2):
        try:
            hdr = [str(c.value or '').strip() for c in ws[candidate]]
            if 'Description' in hdr and 'Location' in hdr:
                hdr_row_idx = candidate
                break
        except Exception:
            continue
    if hdr_row_idx is None:
        raise ValueError(f"Could not find header row in {path}")

    hdr = [str(c.value or '').strip() for c in ws[hdr_row_idx]]

    def col(name: str) -> int | None:
        try:
            return hdr.index(name)
        except ValueError:
            return None

    cols = {
        'description': col('Description'),
        'company': col('Company'),
        'location': col('Location'),
        'account': col('Account'),
        'cc': col('Cost Center'),
        'div': col('DIV'),
        'solution': col('Solution'),
        'agency': col('Agency'),
    }
    missing = [k for k, v in cols.items() if v is None]
    if missing:
        raise ValueError(f"Pipeline xlsx missing columns: {missing}")

    by_ticket: dict[str, dict] = {}
    skipped_no_ticket = 0
    duplicates: list[str] = []
    for r_idx, row in enumerate(ws.iter_rows(min_row=hdr_row_idx + 1, values_only=True), start=1):
        if all(v is None for v in row):
            continue
        desc = row[cols['description']]
        ticket = extract_ticket(desc)
        if not ticket:
            skipped_no_ticket += 1
            continue
        if ticket in by_ticket:
            duplicates.append(ticket)
            continue  # keep first occurrence

        by_ticket[ticket] = {
            'row_idx': r_idx,
            'desc': str(desc)[:40] if desc else '',
            'company': str(row[cols['company']] or '').strip(),
            'location': str(row[cols['location']] or '').strip(),
            'account': str(row[cols['account']] or '').strip(),
            'cc': str(row[cols['cc']] or '').strip(),
            'div': str(row[cols['div']] or '').strip(),
            'solution': str(row[cols['solution']] or '').strip(),
            'agency': str(row[cols['agency']] or '').strip(),
        }

    return by_ticket, skipped_no_ticket, duplicates


# ──────────────────────────────────────────────────────────────────────────────
# Scoring
# ──────────────────────────────────────────────────────────────────────────────

SEGMENTS = ['company', 'location', 'account', 'cc', 'div', 'solution', 'agency']


def _norm(v: str) -> str:
    """Normalize segment value: strip leading zeros, treat empty as '0'."""
    return (v or '').lstrip('0') or '0'


def score(canonical: dict, pipeline: dict) -> dict:
    """Score pipeline against canonical, return full result dict."""
    shared = set(canonical.keys()) & set(pipeline.keys())
    only_canonical = set(canonical.keys()) - set(pipeline.keys())
    only_pipeline = set(pipeline.keys()) - set(canonical.keys())

    per_segment = {s: {'match': 0, 'mismatch': 0, 'mismatches': []} for s in SEGMENTS}
    full_match = 0
    full_match_excl_loc = 0

    for ticket in shared:
        c = canonical[ticket]
        p = pipeline[ticket]
        row_all_match = True
        row_all_match_excl_loc = True
        for s in SEGMENTS:
            cv = _norm(c[s])
            pv = _norm(p[s])
            if cv == pv:
                per_segment[s]['match'] += 1
            else:
                per_segment[s]['mismatch'] += 1
                per_segment[s]['mismatches'].append({
                    'ticket': ticket,
                    'pax': c['pax'],
                    'route': c['route'],
                    'canonical': cv,
                    'pipeline': pv,
                })
                row_all_match = False
                if s != 'location':
                    row_all_match_excl_loc = False
        if row_all_match:
            full_match += 1
        if row_all_match_excl_loc:
            full_match_excl_loc += 1

    n = len(shared)
    return {
        'rows_canonical': len(canonical),
        'rows_pipeline': len(pipeline),
        'rows_shared': n,
        'only_canonical': sorted(only_canonical),
        'only_pipeline': sorted(only_pipeline),
        'per_segment': {
            s: {
                'match': per_segment[s]['match'],
                'mismatch': per_segment[s]['mismatch'],
                'pct': round(per_segment[s]['match'] / n * 100, 1) if n else 0.0,
                'mismatches': per_segment[s]['mismatches'],
            }
            for s in SEGMENTS
        },
        'full_match': full_match,
        'full_match_pct': round(full_match / n * 100, 1) if n else 0.0,
        'full_match_excl_loc': full_match_excl_loc,
        'full_match_excl_loc_pct': round(full_match_excl_loc / n * 100, 1) if n else 0.0,
    }


# ──────────────────────────────────────────────────────────────────────────────
# Output formatting
# ──────────────────────────────────────────────────────────────────────────────

def print_human(result: dict, verbose: bool, thresholds: dict | None) -> int:
    """Pretty-print the scoring result. Returns 0 if all thresholds pass, 1 otherwise."""
    print(f"Canonical rows: {result['rows_canonical']}")
    print(f"Pipeline rows:  {result['rows_pipeline']}")
    print(f"Shared (by ticket): {result['rows_shared']}")

    if result['only_canonical']:
        print(f"⚠️  In canonical but not in pipeline ({len(result['only_canonical'])}): {result['only_canonical'][:10]}{'...' if len(result['only_canonical']) > 10 else ''}")
    if result['only_pipeline']:
        print(f"⚠️  In pipeline but not in canonical ({len(result['only_pipeline'])}): {result['only_pipeline'][:10]}{'...' if len(result['only_pipeline']) > 10 else ''}")

    print()
    print(f"{'Segment':<11} {'Match':>6} {'Mismatch':>9} {'Match %':>9} {'Threshold':>10}")
    failed = False
    for s in SEGMENTS:
        d = result['per_segment'][s]
        thr = thresholds.get(s) if thresholds else None
        thr_str = f"{thr}%" if thr is not None else '—'
        status = ''
        if thr is not None:
            if d['pct'] < thr:
                status = ' FAIL'
                failed = True
            else:
                status = ' PASS'
        print(f"  {s:<9} {d['match']:>6} {d['mismatch']:>9} {d['pct']:>8.1f}% {thr_str:>10}{status}")

    print()
    print(f"Full-row exact match:    {result['full_match']}/{result['rows_shared']} = {result['full_match_pct']:.1f}%")
    print(f"Match excluding location: {result['full_match_excl_loc']}/{result['rows_shared']} = {result['full_match_excl_loc_pct']:.1f}%")

    if verbose:
        for s in SEGMENTS:
            d = result['per_segment'][s]
            if d['mismatches']:
                print(f"\n=== {s} mismatches ({len(d['mismatches'])}) ===")
                for m in d['mismatches']:
                    print(f"  Ticket {m['ticket']} {m['pax']:<30} route={m['route']:<25} canonical={m['canonical']} pipeline={m['pipeline']}")

    return 1 if failed else 0


def main() -> int:
    p = argparse.ArgumentParser(description="Score AlJeel AP pipeline output against canonical")
    p.add_argument('pipeline_xlsx', type=Path, help="Pipeline output xlsx (v15.x format)")
    p.add_argument('canonical_xlsx', type=Path, help="Canonical AlJeel-prepared xlsx")
    p.add_argument('--json', action='store_true', help="Output JSON instead of human-readable")
    p.add_argument('--verbose', action='store_true', help="Show every mismatch row")
    p.add_argument('--threshold-location', type=float, default=None)
    p.add_argument('--threshold-account', type=float, default=None)
    p.add_argument('--threshold-cc', type=float, default=None)
    p.add_argument('--threshold-div', type=float, default=None)
    p.add_argument('--threshold-solution', type=float, default=None)
    p.add_argument('--threshold-agency', type=float, default=None)
    p.add_argument('--threshold-company', type=float, default=None)
    p.add_argument('--check-catches', type=Path, default=None,
                   help="v15.11: JSON file with expected catch counts/categories to verify")
    args = p.parse_args()

    if not args.pipeline_xlsx.exists():
        print(f"ERROR: pipeline xlsx not found: {args.pipeline_xlsx}", file=sys.stderr)
        return 2
    if not args.canonical_xlsx.exists():
        print(f"ERROR: canonical xlsx not found: {args.canonical_xlsx}", file=sys.stderr)
        return 2

    canonical, c_skipped = load_canonical(args.canonical_xlsx)
    pipeline, p_skipped, p_dupes = load_pipeline(args.pipeline_xlsx)

    if c_skipped > 0:
        print(f"⚠️  Canonical: {c_skipped} rows skipped (no extractable ticket)", file=sys.stderr)
    if p_skipped > 0:
        print(f"⚠️  Pipeline:  {p_skipped} rows skipped (no extractable ticket)", file=sys.stderr)
    if p_dupes:
        print(f"⚠️  Pipeline:  duplicate ticket numbers (first occurrence kept): {p_dupes}", file=sys.stderr)

    result = score(canonical, pipeline)

    thresholds = {}
    for s in SEGMENTS:
        v = getattr(args, f'threshold_{s}', None)
        if v is not None:
            thresholds[s] = v

    # v15.11: optional catch verification
    if args.check_catches:
        import json as _json
        from pathlib import Path as _Path
        from collections import Counter as _Counter
        batch_dir = args.pipeline_xlsx.parent
        within_file = batch_dir / "catches-within-batch.json"
        cross_file = batch_dir / "catches-cross-batch.json"
        expected = _json.loads(args.check_catches.read_text())
        actual_catches = []
        for f in (within_file, cross_file):
            if f.exists():
                actual_catches.extend(_json.loads(f.read_text()))
        actual_counts = dict(_Counter(c["category"] for c in actual_catches))
        actual_strict = sum(1 for c in actual_catches
                            if c.get("category") == "DUP_ROUTE_STRICT" and c.get("tier") == "STRICT")
        actual_soft = sum(1 for c in actual_catches
                          if c.get("category") == "DUP_ROUTE_STRICT" and c.get("tier") == "SOFT")
        actual_no_high = sum(1 for c in actual_catches
                             if c.get("category") == "NO_APPROVAL" and c.get("severity") == "HIGH")
        print("\n=== CATCH VERIFICATION (v15.11) ===")
        print(f"  Actual catch counts: {actual_counts}")
        print(f"  DUP_ROUTE_STRICT STRICT: {actual_strict}")
        print(f"  DUP_ROUTE_STRICT SOFT: {actual_soft}")
        print(f"  NO_APPROVAL HIGH: {actual_no_high}")
        exp_counts = expected.get("catch_counts", {})
        all_ok = True
        for cat, exp_count in exp_counts.items():
            actual = actual_counts.get(cat, 0)
            ok = actual == exp_count
            status = "✓" if ok else "✗"
            print(f"  {status} {cat}: expected {exp_count}, got {actual}")
            if not ok:
                all_ok = False
        if all_ok:
            print("  ✅ All catch counts match expected")
        else:
            print("  ⚠️  Some catch counts differ from expected")

    if args.json:
        # Drop verbose mismatch lists from default JSON output unless --verbose
        compact = {
            'rows_canonical': result['rows_canonical'],
            'rows_pipeline': result['rows_pipeline'],
            'rows_shared': result['rows_shared'],
            'only_canonical_count': len(result['only_canonical']),
            'only_pipeline_count': len(result['only_pipeline']),
            'per_segment': {
                s: {k: v for k, v in result['per_segment'][s].items() if k != 'mismatches' or args.verbose}
                for s in SEGMENTS
            },
            'full_match': result['full_match'],
            'full_match_pct': result['full_match_pct'],
            'full_match_excl_loc': result['full_match_excl_loc'],
            'full_match_excl_loc_pct': result['full_match_excl_loc_pct'],
        }
        print(json.dumps(compact, indent=2, default=str))
        return 0
    else:
        return print_human(result, args.verbose, thresholds if thresholds else None)


if __name__ == '__main__':
    sys.exit(main())
