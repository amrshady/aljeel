#!/usr/bin/env python3
"""Format-aware, deterministic Jawal workbook scorer.

The scorer is read-only.  It supports the legacy J26-640 ``Details`` truth
layout and the clerk-reviewed J26-1108 Oracle-template layout.
"""
from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Literal

import openpyxl


SCORER_SCHEMA_VERSION = "jawal-truth-scorer/v2"
FIVE_SEGS = ("account", "cc", "div", "solution", "agency")
ALL_FIELDS = FIVE_SEGS + ("emp_no",)
_GDS_RE = re.compile(r"(?<!\d)(\d{10,})(?!\d)")
_VOUCHER_RE = re.compile(r"(?<![\w-])(26-\d{3,})(?!\d)", re.I)
_WS_RE = re.compile(r"\s+")
_PUNCT_RE = re.compile(r"[^\w\s]+", re.UNICODE)


def _header_name(value: object) -> str:
    return _WS_RE.sub(" ", str(value or "").strip()).casefold()


HEADER_ALIASES = {
    "description": ("description",),
    "amount": ("*amount", "amount", "inv. amt. incl. vat"),
    "distribution": ("distribution combination[..]",),
    "emp_no": ("employee no", "emp no new"),
    "company": ("company",),
    "location": ("location",),
    "account": ("account",),
    "cc": ("cost center",),
    "div": ("div",),
    "solution": ("solution",),
    "agency": ("agency",),
    "invoice_ref": ("invoice ref no", "ref. no.", "رقم المرجع ref. no."),
    "opex_serial": ("opex serial",),
    "opex_details": ("opex allocation details",),
    "ticket": ("ticket no.", "رقم التذكرة/ الفندق ticket no."),
    "passenger": ("passenger name", "اسم الراكب passenger name"),
    "route": ("route", "خط السير route"),
    "note": ("notes",),
    "sl": ("sl. #", "الرقم sl. #"),
    "method": ("agent method", "resolution layer"),
    "flags": ("qc catches",),
}


@dataclass(frozen=True)
class WorkbookLayout:
    path: Path
    sheet_name: str
    header_row: int
    columns: dict[str, int]  # one-based Excel indices
    profile: str


@dataclass
class BaseRow:
    excel_row: int
    description: str = ""
    amount: Decimal | None = None
    emp_no: str = ""
    employee_set: frozenset[str] = frozenset()
    company: str = ""
    location: str = ""
    account: str = ""
    cc: str = ""
    div: str = ""
    solution: str = ""
    agency: str = ""
    invoice_ref: str = ""
    opex_serial: str = ""
    identifier: str | None = None
    method: str = ""
    flags: str = ""
    note: str = ""
    pax: str = ""
    route: str = ""


@dataclass
class TruthRow(BaseRow):
    kind: Literal["sponsorship", "travel"] = "travel"


@dataclass
class PipelineRow(BaseRow):
    pass


@dataclass
class PairedRow:
    key: tuple[Any, ...]
    truth: TruthRow | None
    pipeline: PipelineRow | None
    kind: str
    method: str


@dataclass
class PairingResult:
    pairs: list[PairedRow] = field(default_factory=list)
    truth_only_groups: list[str] = field(default_factory=list)
    pipeline_only_groups: list[str] = field(default_factory=list)
    method_counts: Counter[str] = field(default_factory=Counter)
    multiplicity_counts: Counter[str] = field(default_factory=Counter)
    direct_pairs: int = 0
    virtual_sponsorship_allocations: int = 0
    missing_employees: int = 0
    extra_employees: int = 0
    ambiguous_groups: int = 0
    amount_sum_mismatches: int = 0

    def merge(self, other: "PairingResult") -> None:
        self.pairs.extend(other.pairs)
        self.truth_only_groups.extend(other.truth_only_groups)
        self.pipeline_only_groups.extend(other.pipeline_only_groups)
        self.method_counts.update(other.method_counts)
        self.multiplicity_counts.update(other.multiplicity_counts)
        self.direct_pairs += other.direct_pairs
        self.virtual_sponsorship_allocations += other.virtual_sponsorship_allocations
        self.missing_employees += other.missing_employees
        self.extra_employees += other.extra_employees
        self.ambiguous_groups += other.ambiguous_groups
        self.amount_sum_mismatches += other.amount_sum_mismatches


def discover_columns(path: Path) -> WorkbookLayout:
    """Discover a scoring layout by normalized, unambiguous header text."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheets = [wb["Details"]] if "Details" in wb.sheetnames else [wb.active]
    candidates: list[tuple[Any, int, dict[str, int]]] = []
    alias_lookup = {
        alias: logical for logical, aliases in HEADER_ALIASES.items() for alias in aliases
    }
    for ws in sheets:
        for row_no in range(1, min(ws.max_row, 12) + 1):
            found: dict[str, list[int]] = defaultdict(list)
            for col_no, cell in enumerate(ws[row_no], 1):
                logical = alias_lookup.get(_header_name(cell.value))
                if logical:
                    found[logical].append(col_no)
            # Optional diagnostic aliases may legitimately coexist (for example
            # Resolution Layer and Agent Method).  Prefer the right-most/newer
            # diagnostic column; scoring headers must remain unambiguous.
            ambiguous = {
                name: cols for name, cols in found.items()
                if len(cols) > 1 and name not in {"method"}
            }
            if ambiguous:
                raise ValueError(f"Ambiguous headers in {path} row {row_no}: {ambiguous}")
            cols = {name: values[-1] if name == "method" else values[0] for name, values in found.items()}
            required = {"emp_no", "company", "location", "account", "cc", "div", "solution", "agency"}
            if required <= cols.keys() and ({"description", "amount"} <= cols.keys() or "ticket" in cols):
                candidates.append((ws, row_no, cols))
    if len(candidates) != 1:
        raise ValueError(f"Expected one scoring header in {path}, discovered {len(candidates)}")
    ws, header_row, cols = candidates[0]
    profile = "j26-1108" if header_row == 3 and "invoice_ref" in cols and "description" in cols else "j26-640"
    expected = {
        "j26-1108": {"description": 11, "amount": 13, "emp_no": 18 if cols["emp_no"] == 18 else 16},
    }.get(profile)
    if expected:
        mismatched = {key: (expected[key], cols.get(key)) for key in expected if cols.get(key) != expected[key]}
        if mismatched:
            raise ValueError(f"Unexpected {profile} column indices in {path}: {mismatched}")
    return WorkbookLayout(path, ws.title, header_row, cols, profile)


def _norm(v: object) -> str:
    """Normalize a segment while preserving the distinction between blank and zero."""
    if v is None:
        return ""
    s = str(v).strip()
    if not s:
        return ""
    if re.fullmatch(r"[+-]?\d+(?:\.0+)?", s):
        try:
            return str(int(Decimal(s)))
        except InvalidOperation:
            pass
    return s.lstrip("0") or "0"


def _norm_text(v: object) -> str:
    return _WS_RE.sub(" ", str(v or "").strip()).casefold()


def _norm_ref(v: object) -> str:
    s = str(v or "").strip()
    return re.sub(r"\.0$", "", s)


def _norm_amount(v: object) -> Decimal | None:
    if v is None or str(v).strip() == "":
        return None
    try:
        return Decimal(str(v).replace(",", "")).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def _norm_emp(v: object) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.casefold() in {"", "-", "nan", "none"}:
        return ""
    if re.fullmatch(r"\d+\.0+", s):
        return s.split(".", 1)[0]
    return s


def normalize_employee_set(value: object) -> frozenset[str]:
    return frozenset(emp for part in re.split(r"\s*,\s*", str(value or "")) if (emp := _norm_emp(part)))


def extract_line_identifier(description: object) -> str | None:
    text = str(description or "")
    match = _GDS_RE.search(text)
    if match:
        return match.group(1)
    match = _VOUCHER_RE.search(text)
    return match.group(1).upper() if match else None


def extract_ticket(*sources: object) -> str | None:  # legacy public alias
    for source in sources:
        if identifier := extract_line_identifier(source):
            return identifier
    return None


def _cell(row: tuple[Any, ...], columns: dict[str, int], name: str) -> object:
    col = columns.get(name)
    return row[col - 1] if col and col <= len(row) else None


def _load_rows(path: Path, truth: bool, profile: str = "auto") -> tuple[list[BaseRow], WorkbookLayout]:
    layout = discover_columns(path)
    if profile != "auto" and layout.profile != profile:
        raise ValueError(f"Expected truth profile {profile}, discovered {layout.profile}")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[layout.sheet_name]
    result: list[BaseRow] = []
    for excel_row, values in enumerate(ws.iter_rows(min_row=layout.header_row + 1, values_only=True), layout.header_row + 1):
        if not any(value is not None for value in values):
            continue
        description = str(_cell(values, layout.columns, "description") or "").strip()
        if "ticket" in layout.columns:
            ticket_value = _cell(values, layout.columns, "ticket")
            passenger = str(_cell(values, layout.columns, "passenger") or "").strip()
            identifier = extract_ticket(ticket_value, passenger)
            # Details rows are keyed by their serial number; footer/lookup rows are not data.
            if _cell(values, layout.columns, "sl") is None:
                continue
            description = description or passenger
        else:
            passenger = description.split(" - ", 1)[0]
            identifier = extract_line_identifier(description)
        emp_value = _cell(values, layout.columns, "emp_no")
        common = dict(
            excel_row=excel_row,
            description=description,
            amount=_norm_amount(_cell(values, layout.columns, "amount")),
            emp_no=_norm_emp(emp_value),
            employee_set=normalize_employee_set(emp_value),
            company=_norm(_cell(values, layout.columns, "company")),
            location=_norm(_cell(values, layout.columns, "location")),
            account=_norm(_cell(values, layout.columns, "account")),
            cc=_norm(_cell(values, layout.columns, "cc")),
            div=_norm(_cell(values, layout.columns, "div")),
            solution=_norm(_cell(values, layout.columns, "solution")),
            agency=_norm(_cell(values, layout.columns, "agency")),
            invoice_ref=_norm_ref(_cell(values, layout.columns, "invoice_ref")),
            opex_serial=_norm_text(_cell(values, layout.columns, "opex_serial")),
            identifier=identifier,
            method=str(_cell(values, layout.columns, "method") or "").strip(),
            flags=str(_cell(values, layout.columns, "flags") or "").strip(),
            note=str(_cell(values, layout.columns, "note") or "").strip(),
            pax=passenger,
            route=str(_cell(values, layout.columns, "route") or "").strip(),
        )
        if truth:
            rec = TruthRow(**common)
            rec.kind = classify_truth_row(rec)
        else:
            rec = PipelineRow(**common)
        result.append(rec)
    return result, layout


def load_truth(path: Path, profile: str = "auto") -> list[TruthRow]:
    rows, _ = _load_rows(Path(path), True, profile)
    return [row for row in rows if isinstance(row, TruthRow)]


def load_pipeline(path: Path) -> list[PipelineRow]:
    rows, _ = _load_rows(Path(path), False)
    return [row for row in rows if isinstance(row, PipelineRow)]


def classify_truth_row(row: TruthRow) -> Literal["sponsorship", "travel"]:
    return "sponsorship" if row.account == "60307021" else "travel"


def _logical_description(row: BaseRow) -> str:
    text = _norm_text(row.description)
    prefixes = (_norm_text(row.opex_serial), _norm_text(_norm_ref(row.invoice_ref)))
    for prefix in prefixes:
        if prefix:
            match = re.match(rf"^{re.escape(prefix)}[\s\-–—:|/]+", text, re.I)
            if match:
                text = text[match.end():]
                break
    return _WS_RE.sub(" ", _PUNCT_RE.sub(" ", text)).strip()


def sponsorship_group_key(row: BaseRow) -> tuple[str, str, str, str]:
    return ("sponsorship", _norm_text(row.opex_serial), _logical_description(row), _norm_ref(row.invoice_ref).casefold())


def _key_text(key: tuple[Any, ...]) -> str:
    return "|".join("" if value is None else str(value) for value in key)


def _multiplicity(n: int, m: int) -> str:
    if n == m == 1:
        return "1:1"
    if n > 1 and m == 1:
        return "N:1"
    if n == 1 and m > 1:
        return "1:N"
    return "N:M"


def _pair_group(key: tuple[Any, ...], truths: list[TruthRow], pipes: list[PipelineRow], kind: str) -> PairingResult:
    result = PairingResult()
    n, m = len(truths), len(pipes)
    mult = _multiplicity(n, m)
    result.multiplicity_counts[mult] += 1
    if n == m == 1:
        result.pairs.append(PairedRow(key, truths[0], pipes[0], kind, "direct"))
        result.method_counts["direct"] += 1
        result.direct_pairs += 1
        return result
    if m == 1:
        for truth in truths:
            result.pairs.append(PairedRow(key, truth, pipes[0], kind, "shared-pipeline"))
            result.method_counts["shared-pipeline"] += 1
        return result
    if n == 1:
        for pipe in pipes:
            result.pairs.append(PairedRow(key, truths[0], pipe, kind, "shared-truth"))
            result.method_counts["shared-truth"] += 1
        return result

    remaining_t = list(sorted(truths, key=lambda row: row.excel_row))
    remaining_p = list(sorted(pipes, key=lambda row: row.excel_row))
    # Exact amount multiset matches are deliberately independent of scored fields.
    for truth in list(remaining_t):
        match = next((pipe for pipe in remaining_p if truth.amount is not None and pipe.amount == truth.amount), None)
        if match is not None:
            result.pairs.append(PairedRow(key, truth, match, kind, "amount-exact"))
            result.method_counts["amount-exact"] += 1
            remaining_t.remove(truth)
            remaining_p.remove(match)
    for truth, pipe in zip(remaining_t, remaining_p):
        result.pairs.append(PairedRow(key, truth, pipe, kind, "row-order-residual"))
        result.method_counts["row-order-residual"] += 1
    for truth in remaining_t[len(remaining_p):]:
        result.pairs.append(PairedRow(key, truth, None, kind, "unmatched-truth"))
    for pipe in remaining_p[len(remaining_t):]:
        result.pairs.append(PairedRow(key, None, pipe, kind, "unmatched-pipeline"))
    return result


def pair_travel_rows(truth_rows: list[TruthRow], pipeline_rows: list[PipelineRow]) -> PairingResult:
    result = PairingResult()
    truth_groups: dict[tuple[Any, ...], list[TruthRow]] = defaultdict(list)
    pipe_groups: dict[tuple[Any, ...], list[PipelineRow]] = defaultdict(list)
    ticketless: list[TruthRow] = []
    for row in truth_rows:
        (truth_groups[("travel", row.identifier)].append(row) if row.identifier else ticketless.append(row))
    for row in pipeline_rows:
        if row.identifier:
            pipe_groups[("travel", row.identifier)].append(row)

    # Ticketless rows remain loaded.  Prefer the ordinary composite identity,
    # then apply the narrow malformed-description reference+amount fallback.
    ticketless_pipes = [row for row in pipeline_rows if not row.identifier]
    composite_pipes: dict[tuple[str, Decimal | None, str], list[PipelineRow]] = defaultdict(list)
    for row in ticketless_pipes:
        composite_pipes[(row.invoice_ref, row.amount, _logical_description(row))].append(row)
    fallback_pipes: dict[tuple[str, Decimal | None], list[PipelineRow]] = defaultdict(list)
    for row in pipeline_rows:
        fallback_pipes[(row.invoice_ref, row.amount)].append(row)
    fallback_truth_counts = Counter((row.invoice_ref, row.amount) for row in ticketless)
    claimed_pipe_rows: set[int] = set()
    for truth in ticketless:
        composite = (truth.invoice_ref, truth.amount, _logical_description(truth))
        composite_matches = composite_pipes.get(composite, [])
        fallback = (truth.invoice_ref, truth.amount)
        matches = fallback_pipes.get(fallback, [])
        if len(composite_matches) == 1:
            pipe = composite_matches[0]
            key = ("travel-composite", truth.invoice_ref, str(truth.amount), composite[2])
            result.pairs.append(PairedRow(key, truth, pipe, "travel", "composite-identity"))
            result.method_counts["composite-identity"] += 1
        elif truth.invoice_ref and fallback_truth_counts[fallback] == 1 and len(matches) == 1:
            pipe = matches[0]
            key = ("travel-fallback", truth.invoice_ref, str(truth.amount))
            result.pairs.append(PairedRow(key, truth, pipe, "travel", "unique-reference-amount"))
            result.method_counts["unique-reference-amount"] += 1
        else:
            pipe = None
        if pipe is not None:
            result.multiplicity_counts["1:1"] += 1
            result.direct_pairs += 1
            claimed_pipe_rows.add(pipe.excel_row)
            if pipe.identifier:
                group = pipe_groups.get(("travel", pipe.identifier), [])
                pipe_groups[("travel", pipe.identifier)] = [p for p in group if p.excel_row != pipe.excel_row]
        if pipe is None:
            key = ("travel-fallback", truth.invoice_ref, str(truth.amount))
            result.truth_only_groups.append(_key_text(key))
            result.pairs.append(PairedRow(key, truth, None, "travel", "ambiguous-fallback"))
            if len(matches) > 1 or fallback_truth_counts[fallback] > 1:
                result.ambiguous_groups += 1

    for pipe in ticketless_pipes:
        if pipe.excel_row not in claimed_pipe_rows:
            key = ("travel-composite", pipe.invoice_ref, str(pipe.amount), _logical_description(pipe))
            result.pipeline_only_groups.append(_key_text(key))
            result.pairs.append(PairedRow(key, None, pipe, "travel", "unmatched-pipeline"))

    for key in sorted(set(truth_groups) | set(pipe_groups), key=_key_text):
        truths, pipes = truth_groups.get(key, []), pipe_groups.get(key, [])
        if not truths and not pipes:
            continue
        if not truths:
            result.pipeline_only_groups.append(_key_text(key))
            result.pairs.extend(PairedRow(key, None, pipe, "travel", "unmatched-pipeline") for pipe in pipes)
        elif not pipes:
            result.truth_only_groups.append(_key_text(key))
            result.pairs.extend(PairedRow(key, truth, None, "travel", "unmatched-truth") for truth in truths)
        else:
            result.merge(_pair_group(key, truths, pipes, "travel"))
    return result


def pair_sponsorship_rows(truth_rows: list[TruthRow], pipeline_rows: list[PipelineRow]) -> PairingResult:
    result = PairingResult()
    truth_groups: dict[tuple[str, str, str, str], list[TruthRow]] = defaultdict(list)
    pipe_groups: dict[tuple[str, str, str, str], list[PipelineRow]] = defaultdict(list)
    for row in truth_rows:
        truth_groups[sponsorship_group_key(row)].append(row)
    for row in pipeline_rows:
        pipe_groups[sponsorship_group_key(row)].append(row)
    for key in sorted(set(truth_groups) | set(pipe_groups), key=_key_text):
        truths, pipes = truth_groups.get(key, []), pipe_groups.get(key, [])
        if not truths:
            result.pipeline_only_groups.append(_key_text(key))
            continue
        if not pipes:
            result.truth_only_groups.append(_key_text(key))
            result.missing_employees += sum(bool(row.emp_no) for row in truths)
            result.pairs.extend(PairedRow(key, truth, None, "sponsorship", "unmatched-truth") for truth in truths)
            continue
        result.multiplicity_counts[_multiplicity(len(truths), len(pipes))] += 1
        truth_sum = sum((row.amount or Decimal(0)) for row in truths)
        pipe_sum = sum((row.amount or Decimal(0)) for row in pipes)
        if truth_sum != pipe_sum:
            result.amount_sum_mismatches += 1
        allocations: dict[str, list[PipelineRow]] = defaultdict(list)
        for pipe in sorted(pipes, key=lambda row: row.excel_row):
            result.virtual_sponsorship_allocations += len(pipe.employee_set)
            for employee in sorted(pipe.employee_set):
                allocations[employee].append(pipe)
        for truth in sorted(truths, key=lambda row: row.excel_row):
            matches = allocations.get(truth.emp_no, [])
            if truth.emp_no and matches:
                pipe = matches.pop(0)
                result.pairs.append(PairedRow(key, truth, pipe, "sponsorship", "virtual-employee"))
                result.method_counts["virtual-employee"] += 1
            else:
                result.missing_employees += bool(truth.emp_no)
                result.pairs.append(PairedRow(key, truth, None if not pipes else pipes[0], "sponsorship", "missing-employee"))
                result.method_counts["missing-employee"] += 1
        extras = sum(len(rows) for rows in allocations.values())
        result.extra_employees += extras
    return result


def pair_rows_by_policy(truth_rows: list[TruthRow], pipeline_rows: list[PipelineRow]) -> PairingResult:
    sponsorship_truth = [row for row in truth_rows if classify_truth_row(row) == "sponsorship"]
    travel_truth = [row for row in truth_rows if classify_truth_row(row) == "travel"]
    sponsorship_keys = {sponsorship_group_key(row) for row in sponsorship_truth}
    sponsorship_pipe = [row for row in pipeline_rows if sponsorship_group_key(row) in sponsorship_keys]
    sponsorship_pipe_ids = {id(row) for row in sponsorship_pipe}
    travel_pipe = [row for row in pipeline_rows if id(row) not in sponsorship_pipe_ids]
    result = pair_sponsorship_rows(sponsorship_truth, sponsorship_pipe)
    result.merge(pair_travel_rows(travel_truth, travel_pipe))
    return result


def _pair_legacy_j26_640(truth_rows: list[TruthRow], pipeline_rows: list[PipelineRow]) -> PairingResult:
    """Preserve ticket-based J26-640 pairing while using the new scorer."""
    result = PairingResult()
    truths: dict[str, list[TruthRow]] = defaultdict(list)
    pipes: dict[str, list[PipelineRow]] = defaultdict(list)
    for row in truth_rows:
        if row.identifier:
            truths[row.identifier].append(row)
    for row in pipeline_rows:
        if row.identifier:
            pipes[row.identifier].append(row)
    for identifier in sorted(set(truths) | set(pipes)):
        key = ("legacy", identifier)
        truth_group, pipe_group = truths.get(identifier, []), pipes.get(identifier, [])
        if not truth_group:
            result.pipeline_only_groups.append(_key_text(key))
            result.pairs.extend(PairedRow(key, None, row, "travel", "unmatched-pipeline") for row in pipe_group)
        elif not pipe_group:
            result.truth_only_groups.append(_key_text(key))
            result.pairs.extend(PairedRow(key, row, None, row.kind, "unmatched-truth") for row in truth_group)
        else:
            result.merge(_pair_group(key, truth_group, pipe_group, truth_group[0].kind))
    return result


def _empty_fields(details: bool = False) -> dict[str, dict[str, Any]]:
    return {name: {"match": 0, "mismatch": 0, **({"mismatches": []} if details else {})} for name in ALL_FIELDS}


def score_pairs(pairing: PairingResult) -> dict[str, object]:
    per_field = _empty_fields(details=True)
    breakouts = {
        "sponsorship": {"n": 0, "per_field": _empty_fields(), "full5": 0, "full5_emp": 0},
        "travel": {"n": 0, "per_field": _empty_fields(), "full5": 0, "full5_emp": 0},
    }
    full5 = full5_emp = 0
    off_by_1: Counter[str] = Counter()
    account_breakout: dict[str, dict[str, int]] = defaultdict(lambda: {"evaluated": 0, "full5": 0, "full5_emp": 0})
    evaluated = 0
    for pair in pairing.pairs:
        truth, pipe = pair.truth, pair.pipeline
        if truth is None or pipe is None:
            continue
        evaluated += 1
        bucket = breakouts[pair.kind]
        bucket["n"] += 1
        account = account_breakout[truth.account]
        account["evaluated"] += 1
        diffs: list[str] = []
        for name in ALL_FIELDS:
            if name == "emp_no":
                matched = truth.emp_no in pipe.employee_set if truth.emp_no else not pipe.employee_set
                pipeline_value = ",".join(sorted(pipe.employee_set))
            else:
                matched = getattr(truth, name) == getattr(pipe, name)
                pipeline_value = getattr(pipe, name)
            tally = "match" if matched else "mismatch"
            per_field[name][tally] += 1
            bucket["per_field"][name][tally] += 1
            if not matched:
                diffs.append(name)
                per_field[name]["mismatches"].append({
                    "key": _key_text(pair.key), "truth": getattr(truth, name), "pipeline": pipeline_value,
                    "truth_row": truth.excel_row, "pipeline_row": pipe.excel_row,
                })
        if not any(name in diffs for name in FIVE_SEGS):
            full5 += 1
            bucket["full5"] += 1
            account["full5"] += 1
        if not diffs:
            full5_emp += 1
            bucket["full5_emp"] += 1
            account["full5_emp"] += 1
        if len(diffs) == 1:
            off_by_1[diffs[0]] += 1
    truth_rows = sum(pair.truth is not None for pair in pairing.pairs)
    end_to_end = {
        "denominator": truth_rows,
        "full5": full5,
        "full5_emp": full5_emp,
    }
    return {
        "scorer_schema_version": SCORER_SCHEMA_VERSION,
        "paired_evaluated": evaluated,
        "logical_virtual_evaluated": evaluated,
        "truth_rows": truth_rows,
        "per_field": per_field,
        "full5": full5,
        "full5_emp": full5_emp,
        "off_by_1": dict(sorted(off_by_1.items())),
        "sponsorship": breakouts["sponsorship"],
        "travel": breakouts["travel"],
        "account_breakout": dict(sorted(account_breakout.items())),
        "end_to_end": end_to_end,
        "pairing_integrity": {
            "direct_1_to_1_pairs": pairing.direct_pairs,
            "multiplicity_groups": dict(sorted(pairing.multiplicity_counts.items())),
            "pairing_methods": dict(sorted(pairing.method_counts.items())),
            "virtual_sponsorship_allocations": pairing.virtual_sponsorship_allocations,
            "missing_employees": pairing.missing_employees,
            "extra_employees": pairing.extra_employees,
            "ambiguous_groups": pairing.ambiguous_groups,
            "amount_sum_mismatches": pairing.amount_sum_mismatches,
        },
        "only_truth_tickets": sorted(set(pairing.truth_only_groups)),
        "only_pipe_tickets": sorted(set(pairing.pipeline_only_groups)),
        "truth_only_logical_groups": sorted(set(pairing.truth_only_groups)),
        "pipeline_only_logical_groups": sorted(set(pairing.pipeline_only_groups)),
    }


def score_workbooks(pipeline_path: Path, truth_path: Path, truth_profile: str = "auto") -> dict[str, object]:
    truths = load_truth(Path(truth_path), truth_profile)
    pipelines = load_pipeline(Path(pipeline_path))
    discovered_truth_profile = discover_columns(Path(truth_path)).profile
    pairing = _pair_legacy_j26_640(truths, pipelines) if discovered_truth_profile == "j26-640" else pair_rows_by_policy(truths, pipelines)
    result = score_pairs(pairing)
    result["truth_rows"] = len(truths)
    result["pipe_rows"] = len(pipelines)
    result["truth_employee_coverage"] = sum(bool(row.emp_no) for row in truths)
    result["truth_sponsorship_rows"] = sum(row.kind == "sponsorship" for row in truths)
    result["truth_sponsorship_employee_coverage"] = sum(row.kind == "sponsorship" and bool(row.emp_no) for row in truths)
    result["truth_ticketless_rows"] = sum(row.identifier is None for row in truths)
    return result


def score(truth_rows: Any, pipeline_rows: Any) -> dict[str, object]:
    """Compatibility entry point accepting row lists (or legacy ticket maps)."""
    if isinstance(truth_rows, dict):
        truth_rows = [row for rows in truth_rows.values() for row in rows]
        pipeline_rows = [row for rows in pipeline_rows.values() for row in rows]
        raise TypeError("Legacy dictionary records are no longer supported; use load_truth/load_pipeline")
    return score_pairs(pair_rows_by_policy(truth_rows, pipeline_rows))


def pct(value: int, denominator: int) -> str:
    return f"{value / denominator * 100:.1f}%" if denominator else "n/a"


def render_md(result: dict[str, Any], pipe_path: Path, truth_path: Path) -> str:
    batch_match = re.search(r"J26-\d+", pipe_path.name, re.I) or re.search(r"J26-\d+", truth_path.name, re.I)
    batch = batch_match.group(0).upper() if batch_match else pipe_path.stem
    n = result["paired_evaluated"]
    lines = [
        f"# {batch} Semantic Scoring vs Truth", "",
        f"- Pipeline output: `{pipe_path}`", f"- Truth: `{truth_path}`",
        f"- Truth physical rows: {result['truth_rows']}", f"- Pipeline physical rows: {result['pipe_rows']}",
        f"- Paired/logical evaluated: {n}", "", "## Headline", "",
        f"- **All-5-segments exact:** {result['full5']}/{n} = **{pct(result['full5'], n)}**",
        f"- **All-5 + emp_no exact:** {result['full5_emp']}/{n} = **{pct(result['full5_emp'], n)}**", "",
        "## Per-field match rate", "", "| Field | Match | Mismatch | Match % |", "|---|---:|---:|---:|",
    ]
    for name in ALL_FIELDS:
        values = result["per_field"][name]
        lines.append(f"| `{name}` | {values['match']} | {values['mismatch']} | {pct(values['match'], values['match'] + values['mismatch'])} |")
    lines.extend(["", "## Pairing integrity", "", "```json", json.dumps(result["pairing_integrity"], indent=2, sort_keys=True), "```", ""])
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("pipeline_xlsx")
    parser.add_argument("truth_xlsx")
    parser.add_argument("--out", default=None, help="Write markdown report to this path")
    parser.add_argument("--json", dest="json_out", default=None, help="Write JSON result to this path")
    args = parser.parse_args()
    pipe_path, truth_path = Path(args.pipeline_xlsx), Path(args.truth_xlsx)
    result = score_workbooks(pipe_path, truth_path)
    report = render_md(result, pipe_path, truth_path)
    if args.out:
        Path(args.out).write_text(report, encoding="utf-8")
        print(f"wrote {args.out}")
    else:
        print(report)
    if args.json_out:
        Path(args.json_out).write_text(json.dumps(result, indent=2, default=str), encoding="utf-8")
        print(f"wrote {args.json_out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
