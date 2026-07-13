#!/usr/bin/env python3
"""Deterministic Labadi allocation lookup for Asateel project invoices.

The normal Asateel pipeline does not import or call this module unless the
explicit ``projects-labadi-v1`` allocation mode is selected.
"""
from __future__ import annotations

import hashlib
import json
import re
import unicodedata
from pathlib import Path
from typing import Any

import openpyxl


SCHEMA_VERSION = 1
MODE = "projects-labadi-v1"
DEFAULT_LOOKUP_PATH = (
    Path(__file__).resolve().parents[1]
    / "pipelines"
    / "lookups"
    / "asateel_projects_labadi_v1.json"
)


class ProjectLookupValidationError(ValueError):
    """Raised when source or normalized lookup data is uncertain."""


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def code(value: Any, width: int | None = None) -> str:
    text = clean(value)
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    if width and text.isdigit():
        return text.zfill(width)
    return text


def normalize_alias(value: Any) -> str:
    """Normalize Arabic/English text for exact alias matching only."""
    text = unicodedata.normalize("NFKD", clean(value)).casefold()
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"[^0-9a-z\u0600-\u06ff]+", "", text)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _master_indexes(master_path: Path) -> tuple[dict[str, dict[str, str]], dict[str, dict[str, str]], dict[str, list[str]], dict[str, list[str]]]:
    wb = openpyxl.load_workbook(master_path, read_only=True, data_only=True)
    try:
        agencies: dict[str, dict[str, str]] = {}
        agency_aliases: dict[str, list[str]] = {}
        for row in wb["Agency"].iter_rows(min_row=2, values_only=True):
            agency_code = code(row[0] if row else "", 5)
            name = clean(row[1] if len(row) > 1 else "")
            if agency_code and name:
                agencies[agency_code] = {"code": agency_code, "name": name}
                agency_aliases.setdefault(normalize_alias(name), []).append(agency_code)

        employees: dict[str, dict[str, str]] = {}
        employee_aliases: dict[str, list[str]] = {}
        for row in wb["Manpower"].iter_rows(min_row=2, values_only=True):
            emp_no = code(row[0] if row else "")
            name = clean(row[2] if len(row) > 2 else "")
            if not emp_no:
                continue
            rec = {
                "employee_no": emp_no,
                "employee_name": name,
                "arabic_name": clean(row[3] if len(row) > 3 else ""),
                "home_agency_code": code(row[10] if len(row) > 10 else "", 5),
                "home_agency_name": clean(row[11] if len(row) > 11 else ""),
            }
            employees[emp_no] = rec
            for alias in (rec["employee_name"], rec["arabic_name"]):
                if normalize_alias(alias):
                    employee_aliases.setdefault(normalize_alias(alias), []).append(emp_no)
        return agencies, employees, agency_aliases, employee_aliases
    finally:
        wb.close()


def _unique_alias(index: dict[str, list[str]], raw: str, kind: str, errors: list[str]) -> str:
    matches = sorted(set(index.get(normalize_alias(raw), [])))
    if len(matches) != 1:
        errors.append(f"{kind} alias {raw!r} resolved to {matches or 'nothing'}")
        return ""
    return matches[0]


def build_lookup(source_path: Path, master_path: Path) -> dict[str, Any]:
    """Parse the fixed workbook relationships and pre-resolve canonical codes."""
    source_path = source_path.resolve()
    master_path = master_path.resolve()
    agencies, employees, agency_aliases, _ = _master_indexes(master_path)
    wb = openpyxl.load_workbook(source_path, read_only=False, data_only=True)
    errors: list[str] = []
    warnings: list[str] = []
    rules: list[dict[str, Any]] = []
    try:
        if wb.sheetnames != ["Sheet1"]:
            errors.append(f"expected exactly ['Sheet1']; found {wb.sheetnames!r}")
        ws = wb[wb.sheetnames[0]]
        expected_headers = {"F4": "Agency", "G4": "Manager", "H4": "Employee NO", "F14": "BMX"}
        for cell, expected in expected_headers.items():
            if clean(ws[cell].value) != expected:
                errors.append(f"{cell} expected {expected!r}; found {clean(ws[cell].value)!r}")

        for row_no in range(5, 13):
            source_agency = clean(ws.cell(row_no, 6).value)
            manager_name = clean(ws.cell(row_no, 7).value)
            manager_no = code(ws.cell(row_no, 8).value)
            agency_code = _unique_alias(agency_aliases, source_agency, "agency", errors)
            manager = employees.get(manager_no)
            if not manager:
                errors.append(f"H{row_no} employee {manager_no!r} is absent from Manpower")
                continue
            if normalize_alias(manager_name) != normalize_alias(manager["employee_name"]):
                errors.append(
                    f"G{row_no} name {manager_name!r} does not exactly match Manpower employee {manager_no} ({manager['employee_name']!r})"
                )
            agency = agencies.get(agency_code)
            if not agency:
                continue
            if manager["home_agency_code"] != agency_code:
                warnings.append(
                    f"{agency_code} {agency['name']}: designated manager {manager_no} has Manpower home agency "
                    f"{manager['home_agency_code']} {manager['home_agency_name']}"
                )
            rules.append({
                "agency_code": agency_code,
                "agency_name": agency["name"],
                "agency_aliases": sorted(set([source_agency, agency["name"]]), key=lambda value: (normalize_alias(value), value)),
                "employee_strategy": "agency_manager",
                "manager": {
                    "employee_no": manager_no,
                    "employee_name": manager["employee_name"],
                    "employee_aliases": sorted(set([manager_name, manager["employee_name"]]), key=lambda value: (normalize_alias(value), value)),
                },
                "source_cells": {"agency": f"F{row_no}", "employee_name": f"G{row_no}", "employee_no": f"H{row_no}"},
            })

        bmx_code = _unique_alias(agency_aliases, "BMX", "agency", errors)
        bmx = agencies.get(bmx_code)
        heads: list[dict[str, Any]] = []
        for head_row in (15, 19, 23):
            if normalize_alias(ws.cell(head_row, 7).value) != "linehead":
                errors.append(f"G{head_row} expected 'Line Head'")
            head_name = clean(ws.cell(head_row, 8).value)
            head_no = code(ws.cell(head_row, 9).value)
            head = employees.get(head_no)
            if not head:
                errors.append(f"I{head_row} head employee {head_no!r} is absent from Manpower")
                continue
            if normalize_alias(head_name) != normalize_alias(head["employee_name"]):
                errors.append(f"H{head_row} head name does not match Manpower employee {head_no}")
            junior_name_row = head_row + 1
            junior_no_row = head_row + 2
            if normalize_alias(ws.cell(junior_name_row, 8).value) != "empolyeename":
                errors.append(f"H{junior_name_row} expected workbook label 'Empolyee Name'")
            if normalize_alias(ws.cell(junior_no_row, 8).value) != "empolyeeno":
                errors.append(f"H{junior_no_row} expected workbook label 'Empolyee No'")
            juniors: list[dict[str, Any]] = []
            for column in range(9, ws.max_column + 1):
                junior_name = clean(ws.cell(junior_name_row, column).value)
                junior_no = code(ws.cell(junior_no_row, column).value)
                if not junior_name and not junior_no:
                    continue
                junior = employees.get(junior_no)
                coordinate = ws.cell(junior_no_row, column).coordinate
                if not junior:
                    errors.append(f"{coordinate} junior employee {junior_no!r} is absent from Manpower")
                    continue
                if normalize_alias(junior_name) != normalize_alias(junior["employee_name"]):
                    errors.append(f"{ws.cell(junior_name_row, column).coordinate} junior name does not match Manpower employee {junior_no}")
                juniors.append({
                    "employee_no": junior_no,
                    "employee_name": junior["employee_name"],
                    "employee_aliases": sorted(set([junior_name, junior["employee_name"]]), key=lambda value: (normalize_alias(value), value)),
                    "source_cells": {
                        "employee_name": ws.cell(junior_name_row, column).coordinate,
                        "employee_no": coordinate,
                    },
                })
            heads.append({
                "employee_no": head_no,
                "employee_name": head["employee_name"],
                "employee_aliases": sorted(set([head_name, head["employee_name"]]), key=lambda value: (normalize_alias(value), value)),
                "source_cells": {"employee_name": f"H{head_row}", "employee_no": f"I{head_row}"},
                "juniors": juniors,
            })
        if bmx:
            rules.append({
                "agency_code": bmx_code,
                "agency_name": bmx["name"],
                "agency_aliases": ["BMX"],
                "employee_strategy": "bmx_junior_to_head",
                "heads": heads,
                "source_cells": {"agency": "F14", "instruction": "G14"},
            })
    finally:
        wb.close()

    rules.sort(key=lambda rule: rule["agency_code"])
    if errors:
        raise ProjectLookupValidationError("; ".join(errors))
    direct_count = sum(rule["employee_strategy"] == "agency_manager" for rule in rules)
    bmx_heads = sum(len(rule.get("heads", [])) for rule in rules)
    bmx_juniors = sum(len(head["juniors"]) for rule in rules for head in rule.get("heads", []))
    referenced_employee_codes = {
        rule["manager"]["employee_no"]
        for rule in rules
        if rule["employee_strategy"] == "agency_manager"
    }
    referenced_employee_codes.update(
        employee["employee_no"]
        for rule in rules
        for head in rule.get("heads", [])
        for employee in [head, *head.get("juniors", [])]
    )
    return {
        "schema_version": SCHEMA_VERSION,
        "mode": MODE,
        "description": "Labadi workbook allocation override for Asateel PROJECTS invoices only",
        "provenance": {
            "source_label": "Labadi Book1 Asateel projects allocation workbook",
            "source_sha256": _sha256(source_path),
            "source_size_bytes": source_path.stat().st_size,
            "source_sheet": "Sheet1",
            "source_cells": ["F4:H12", "F14:M25"],
            "master_label": "qc/master-data/Aljeel_Lookups-v2.xlsx",
            "master_sha256": _sha256(master_path),
            "master_size_bytes": master_path.stat().st_size,
        },
        "precedence": [
            "project mode applies only when invoice folder is PROJECTS",
            "agency: canonical code exact match",
            "agency: unique normalized alias exact match",
            "non-BMX employee: workbook manager selected by resolved agency",
            "BMX employee: exact junior/head code, then unique normalized employee-name alias",
            "unknown, conflicting, or ambiguous input: no guessed employee; emit project lookup review",
        ],
        "statistics": {
            "agency_rules": len(rules),
            "direct_agency_manager_rules": direct_count,
            "bmx_heads": bmx_heads,
            "bmx_junior_to_head_rules": bmx_juniors,
            "referenced_employee_codes": len(referenced_employee_codes),
        },
        "validation": {"errors": [], "ambiguities": [], "warnings": sorted(warnings)},
        "agency_rules": rules,
    }


def write_lookup(data: dict[str, Any], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(data, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def load_lookup(path: Path = DEFAULT_LOOKUP_PATH) -> dict[str, Any]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if data.get("schema_version") != SCHEMA_VERSION or data.get("mode") != MODE:
        raise ProjectLookupValidationError(f"unsupported project lookup schema/mode in {path}")
    if (data.get("validation") or {}).get("errors") or (data.get("validation") or {}).get("ambiguities"):
        raise ProjectLookupValidationError(f"project lookup contains validation failures: {path}")
    _runtime_indexes(data)  # validates all exact indexes and collisions eagerly
    return data


def _runtime_indexes(data: dict[str, Any]) -> dict[str, Any]:
    by_code: dict[str, dict[str, Any]] = {}
    by_alias: dict[str, dict[str, Any]] = {}
    employee_by_code: dict[str, tuple[dict[str, Any], dict[str, Any]]] = {}
    employee_by_alias: dict[str, tuple[dict[str, Any], dict[str, Any]]] = {}

    def unique_put(index: dict[str, Any], key: str, value: Any, label: str) -> None:
        if not key:
            return
        if key in index and index[key] != value:
            raise ProjectLookupValidationError(f"ambiguous {label} {key!r} in normalized lookup")
        index[key] = value

    for rule in data.get("agency_rules", []):
        unique_put(by_code, code(rule.get("agency_code"), 5), rule, "agency code")
        for alias in rule.get("agency_aliases", []):
            unique_put(by_alias, normalize_alias(alias), rule, "agency alias")
        if rule.get("employee_strategy") != "bmx_junior_to_head":
            continue
        for head in rule.get("heads", []):
            pair = (head, head)
            unique_put(employee_by_code, code(head.get("employee_no")), pair, "BMX employee code")
            for alias in head.get("employee_aliases", []):
                unique_put(employee_by_alias, normalize_alias(alias), pair, "BMX employee alias")
            for junior in head.get("juniors", []):
                pair = (junior, head)
                unique_put(employee_by_code, code(junior.get("employee_no")), pair, "BMX employee code")
                for alias in junior.get("employee_aliases", []):
                    unique_put(employee_by_alias, normalize_alias(alias), pair, "BMX employee alias")
    return {
        "agency_by_code": by_code,
        "agency_by_alias": by_alias,
        "bmx_employee_by_code": employee_by_code,
        "bmx_employee_by_alias": employee_by_alias,
    }


def resolve_override(
    data: dict[str, Any],
    *,
    agency_code: Any = "",
    agency_name: Any = "",
    employee_no: Any = "",
    employee_name: Any = "",
) -> dict[str, Any]:
    """Resolve an exact project override and return audit-ready fields."""
    indexes = _runtime_indexes(data)
    input_agency_code = code(agency_code, 5)
    input_agency_name = clean(agency_name)
    code_rule = indexes["agency_by_code"].get(input_agency_code) if input_agency_code else None
    alias_rule = indexes["agency_by_alias"].get(normalize_alias(input_agency_name)) if input_agency_name else None
    audit = {
        "mode": MODE,
        "status": "unresolved_agency",
        "agency_input": {"code": input_agency_code, "name": input_agency_name},
        "employee_input": {"employee_no": code(employee_no), "employee_name": clean(employee_name)},
        "agency_match_method": "",
        "employee_match_method": "",
        "agency_before": {"code": input_agency_code, "name": input_agency_name},
        "employee_before": {"employee_no": code(employee_no), "employee_name": clean(employee_name)},
        "agency_after": None,
        "employee_after": None,
        "explanation": "Project agency did not match an exact canonical code or unique normalized alias; no project override applied.",
    }
    if code_rule and alias_rule and code_rule["agency_code"] != alias_rule["agency_code"]:
        audit["status"] = "ambiguous_agency"
        audit["explanation"] = "Project agency code and alias resolve to different canonical agencies; no override applied."
        return audit
    rule = code_rule or alias_rule
    if not rule:
        return audit
    audit["agency_match_method"] = "canonical_code_exact" if code_rule else "normalized_alias_exact"
    audit["agency_after"] = {"code": rule["agency_code"], "name": rule["agency_name"]}

    if rule["employee_strategy"] == "agency_manager":
        manager = rule["manager"]
        audit.update({
            "status": "applied",
            "employee_match_method": "agency_manager",
            "employee_after": {"employee_no": manager["employee_no"], "employee_name": manager["employee_name"]},
            "explanation": f"Resolved agency {rule['agency_code']} first; employee overwritten with workbook-designated project manager.",
        })
        return audit

    input_employee_no = code(employee_no)
    input_employee_name = clean(employee_name)
    code_pair = indexes["bmx_employee_by_code"].get(input_employee_no) if input_employee_no else None
    alias_pair = indexes["bmx_employee_by_alias"].get(normalize_alias(input_employee_name)) if input_employee_name else None
    if code_pair and alias_pair and code_pair[1]["employee_no"] != alias_pair[1]["employee_no"]:
        audit["status"] = "ambiguous_employee"
        audit["explanation"] = "BMX employee code and alias resolve to different heads; agency resolved but employee was not guessed."
        return audit
    pair = code_pair or alias_pair
    if not pair:
        audit["status"] = "unresolved_employee"
        audit["explanation"] = "BMX agency resolved, but employee is not an explicit workbook junior/head; agency resolved and employee was not guessed."
        return audit
    matched, head = pair
    audit.update({
        "status": "applied",
        "employee_match_method": "employee_code_exact" if code_pair else "normalized_employee_alias_exact",
        "employee_after": {"employee_no": head["employee_no"], "employee_name": head["employee_name"]},
        "explanation": (
            f"Resolved BMX agency first; workbook employee {matched['employee_no']} mapped to designated head {head['employee_no']}."
        ),
    })
    return audit
