#!/usr/bin/env python3
"""Internal fire-and-forget Asateel trigger API.

Environment:
  ASATEEL_TRIGGER_KEY: shared secret required in X-Asateel-Trigger-Key.
  /usr/local/bin/gog: sends Gmail reports using service keyring auth.
  ASATEEL_REPORT_RECIPIENTS: comma-separated default report recipients.
"""
from __future__ import annotations

import json
import logging
import os
import queue
import re
import shutil
import subprocess
import threading
import uuid
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from flask import Blueprint, Response, jsonify, request

try:
    import yaml
except ImportError:
    yaml = None


ROOT = Path("/home/clawdbot/.openclaw/workspace/aljeel")
MATCHED = ROOT / "matched"
SO_DETAIL = ROOT / "reference" / "SO_Detail_Labadi_1_R21_AA.xlsx"
OPENAPI_SPEC = ROOT / "docs" / "asateel-trigger-api.openapi.yaml"
PIPELINE_TIMEOUT_SECONDS = 3 * 60 * 60
DEFAULT_RECIPIENTS = ["amr@accordpartners.ai"]
ALLOWED_REGIONS = {"CENTRAL", "PT_PROJECT", "PROJECTS", "ADMIN", "MAIN", "EASTERN", "WESTERN"}
REGION_ENGINE_FOLDERS = {
    "CENTRAL": "CENTRAL",
    "PT_PROJECT": "PROJECTS",
    "PROJECTS": "PROJECTS",
    "ADMIN": "ADMIN",
    "MAIN": "MAIN",
    "EASTERN": "EASTERN",
    "WESTERN": "WESTERN",
}
REGION_TITLES = {
    "CENTRAL": "Central",
    "PT_PROJECT": "P&T",
    "PROJECTS": "Projects",
    "ADMIN": "Admin",
    "MAIN": "Main",
    "EASTERN": "Eastern",
    "WESTERN": "Western",
}
GOG_BIN = "/usr/local/bin/gog"
GOG_ACCOUNT = "aljeel@accordpartners.ai"
HOUSE_CC = "amr@accordpartners.ai"

bp = Blueprint("asateel_trigger", __name__)
_logger = logging.getLogger("asateel_trigger")

_job_queue: queue.Queue["AsateelJob"] = queue.Queue()
_registry: dict[str, dict[str, Any]] = {}
_registry_lock = threading.Lock()
_worker_lock = threading.Lock()
_worker_thread: threading.Thread | None = None
_openapi_spec: dict[str, Any] | None = None
_openapi_yaml: bytes | None = None
# This registry/queue is intentionally in-memory; restarting Flask loses queued
# jobs and status history.


@dataclass(frozen=True)
class AsateelJob:
    run_id: str
    archive_date: str
    folder_name: str
    region: str
    batch_id: str
    recipients: list[str]


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _json_error(status: int, error: str) -> tuple[Response, int]:
    return jsonify({"error": error}), status


def _load_openapi_spec() -> tuple[dict[str, Any] | None, bytes | None]:
    global _openapi_spec, _openapi_yaml
    if _openapi_spec is not None and _openapi_yaml is not None:
        return _openapi_spec, _openapi_yaml
    if yaml is None:
        _logger.warning("Asateel OpenAPI spec unavailable: PyYAML is not installed")
        return None, None
    try:
        raw_yaml = OPENAPI_SPEC.read_bytes()
        spec = yaml.safe_load(raw_yaml)
    except OSError as exc:
        _logger.warning("Asateel OpenAPI spec unavailable: %s", exc)
        return None, None
    except yaml.YAMLError as exc:
        _logger.warning("Asateel OpenAPI spec unavailable: %s", exc)
        return None, None
    if not isinstance(spec, dict):
        _logger.warning("Asateel OpenAPI spec unavailable: expected YAML mapping")
        return None, None
    _openapi_spec = spec
    _openapi_yaml = raw_yaml
    return _openapi_spec, _openapi_yaml


def _authorized() -> bool:
    expected = os.environ.get("ASATEEL_TRIGGER_KEY", "")
    return bool(expected) and request.headers.get("X-Asateel-Trigger-Key", "") == expected


def _clean(value: Any) -> str:
    return str(value or "").strip()


def _validate_recipients(value: Any) -> list[str] | None:
    if value is None:
        return None
    if not isinstance(value, list):
        raise ValueError("recipients must be a list of email addresses")
    recipients = [_clean(item) for item in value]
    recipients = [item for item in recipients if item]
    if not recipients:
        raise ValueError("recipients must contain at least one email address")
    for item in recipients:
        if "@" not in item or any(ch.isspace() for ch in item):
            raise ValueError(f"invalid recipient: {item}")
    return recipients


def _parse_payload(payload: Any) -> dict[str, Any]:
    if not isinstance(payload, dict):
        raise ValueError("JSON body required")

    archive_date = _clean(payload.get("archive_date"))
    folder_name = _clean(payload.get("folder_name"))
    region = _clean(payload.get("region")).upper()
    batch_id = _clean(payload.get("batch_id"))
    recipients = _validate_recipients(payload.get("recipients"))

    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", archive_date):
        raise ValueError("archive_date must match YYYY-MM-DD")
    if region not in ALLOWED_REGIONS:
        raise ValueError(
            "region must be one of CENTRAL, PT_PROJECT, PROJECTS, ADMIN, MAIN, EASTERN, WESTERN"
        )
    if not folder_name:
        raise ValueError("folder_name is required")
    if not batch_id:
        raise ValueError("batch_id is required")
    if "/" in batch_id or "\\" in batch_id:
        raise ValueError("batch_id must not contain path separators")

    return {
        "archive_date": archive_date,
        "folder_name": folder_name,
        "region": region,
        "batch_id": batch_id,
        "recipients": recipients or _default_recipients(),
    }


def region_title_and_number(region: str, batch_id: str) -> tuple[str, str]:
    title = REGION_TITLES[region]
    match = re.search(r"(\d+)\s*$", batch_id)
    number = match.group(1).zfill(2) if match else "00"
    return title, number


def _batch_dir(batch_id: str) -> Path:
    return ROOT / "batches" / f"asateel-{batch_id}"


def stable_output_paths(region: str, batch_id: str) -> tuple[Path, Path]:
    title, number = region_title_and_number(region, batch_id)
    base = _batch_dir(batch_id) / f"{title}-{number}-2026"
    return base.with_name(f"{base.name}_Oracle-upload.xlsx"), base.with_name(f"{base.name}_Missing-JQs.xlsx")


def _default_recipients() -> list[str]:
    raw = os.environ.get("ASATEEL_REPORT_RECIPIENTS", "")
    recipients = [item.strip() for item in raw.split(",") if item.strip()]
    return recipients or DEFAULT_RECIPIENTS[:]


def _tail(text: str, limit: int = 4000) -> str:
    text = text or ""
    return text[-limit:]


def _run_command(args: list[str], timeout: int = PIPELINE_TIMEOUT_SECONDS) -> subprocess.CompletedProcess[str]:
    return subprocess.run(args, text=True, capture_output=True, timeout=timeout)


def _parse_master_path(stdout: str) -> Path | None:
    for line in stdout.splitlines():
        if line.startswith("Expenses-Format master: "):
            value = line.split(": ", 1)[1].strip()
            if value and value != "NOT FOUND":
                return Path(value)
    return None


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _money_text(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, (int, float)):
        return f"{float(value):.2f}"
    return str(value).strip()


def extract_jq(additional_information: Any) -> str:
    text = _cell_text(additional_information)
    if "." not in text:
        return ""
    return text.split(".", 1)[1].strip()


def build_missing_jq_sheet(oracle_xlsx: Path, output_xlsx: Path) -> int:
    wb = openpyxl.load_workbook(oracle_xlsx, data_only=True)
    try:
        ws = wb.active
        headers = {
            _cell_text(cell.value): idx
            for idx, cell in enumerate(ws[3], start=1)
            if _cell_text(cell.value)
        }
        required = [
            "Row Status",
            "SO_Detail Agency",
            "Cost Center",
            "Additional Information",
            "Employee No",
            "*Invoice Number",
            "*Amount",
        ]
        missing = [name for name in required if name not in headers]
        if missing:
            raise ValueError(f"Oracle sheet missing required columns: {', '.join(missing)}")

        rows: dict[str, tuple[str, str, str, str]] = {}
        for row_idx in range(4, ws.max_row + 1):
            row_status = _cell_text(ws.cell(row_idx, headers["Row Status"]).value).upper()
            so_detail_agency = _cell_text(ws.cell(row_idx, headers["SO_Detail Agency"]).value)
            cost_center = _cell_text(ws.cell(row_idx, headers["Cost Center"]).value)
            if row_status != "RED" or so_detail_agency or cost_center == "140040":
                continue
            jq = extract_jq(ws.cell(row_idx, headers["Additional Information"]).value)
            if not jq or jq in rows:
                continue
            rows[jq] = (
                jq,
                _cell_text(ws.cell(row_idx, headers["Employee No"]).value),
                _cell_text(ws.cell(row_idx, headers["*Invoice Number"]).value),
                _money_text(ws.cell(row_idx, headers["*Amount"]).value),
            )
    finally:
        wb.close()

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Missing JQs"
    out_ws.append(["JQ", "Employee No", "Invoice Number", "Amount SAR"])
    if rows:
        for record in sorted(rows.values(), key=lambda item: item[0]):
            out_ws.append(list(record))
    else:
        out_ws.append(["No missing JQs", "", "", ""])
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(output_xlsx)
    return len(rows)


def _summary_counts(summary: dict[str, Any], missing_jq_count: int) -> dict[str, Any]:
    row_counts = summary.get("row_status_counts") or {}
    return {
        "invoices": summary.get("invoice_count", 0),
        "allocation_rows": summary.get("allocation_lines", 0),
        "green": row_counts.get("GREEN", 0),
        "yellow": row_counts.get("YELLOW", 0),
        "red": row_counts.get("RED", 0),
        "reconciled": summary.get("reconciled_invoices", 0),
        "mismatched": summary.get("mismatched_invoices", 0),
        "reconciliation_rate": summary.get("reconciliation_rate", 0),
        "total_invoice_value_sar": summary.get("total_invoice_value_sar", 0),
        "missing_jq": missing_jq_count,
    }


def _email_configured() -> bool:
    return os.path.exists(GOG_BIN) and os.access(GOG_BIN, os.X_OK)


def _send_email(recipients: list[str], subject: str, body: str, attachments: list[Path] | None = None) -> bool:
    if not _email_configured():
        _logger.warning("Asateel gog mailer is not available; skipping email")
        return False

    clean_recipients = [recipient.strip() for recipient in recipients if recipient.strip()]
    recipient_keys = {recipient.lower() for recipient in clean_recipients}
    cc_recipients = [] if HOUSE_CC in recipient_keys else [HOUSE_CC]

    args = [
        GOG_BIN,
        "gmail",
        "send",
        "--account",
        GOG_ACCOUNT,
        "--to",
        ",".join(clean_recipients),
        "--subject",
        subject,
        "--body",
        body,
    ]
    if cc_recipients:
        args.extend(["--cc", ",".join(cc_recipients)])
    for attachment in attachments or []:
        args.extend(["--attach", str(attachment)])

    try:
        completed = subprocess.run(args, text=True, capture_output=True, timeout=120)
    except subprocess.TimeoutExpired as exc:
        _logger.warning("Asateel gog email timed out after %s seconds: %s", exc.timeout, exc)
        return False
    except OSError as exc:
        _logger.warning("Asateel gog email failed to start: %s", exc)
        return False

    if completed.returncode != 0:
        _logger.warning(
            "Asateel gog email failed with exit code %s; stdout=%r stderr=%r",
            completed.returncode,
            _tail(completed.stdout, 2000),
            _tail(completed.stderr, 2000),
        )
        return False
    return True


def _success_body(summary: dict[str, Any], missing_jq_count: int) -> str:
    counts = _summary_counts(summary, missing_jq_count)
    total = counts["reconciled"] + counts["mismatched"]
    return "\n".join([
        "Asateel reconciliation complete.",
        "",
        f"Invoices: {counts['invoices']}",
        f"Allocation rows: {counts['allocation_rows']}",
        f"GREEN/YELLOW/RED: {counts['green']}/{counts['yellow']}/{counts['red']}",
        f"Reconciled: {counts['reconciled']}/{total}",
        f"Reconciliation rate: {counts['reconciliation_rate']}%",
        f"Total invoice value SAR: {counts['total_invoice_value_sar']}",
        f"RED-JQ missing count: {counts['missing_jq']}",
    ])


def _failure_body(job: AsateelJob, detail: str) -> str:
    return "\n".join([
        "Asateel reconciliation failed.",
        "",
        f"Run ID: {job.run_id}",
        f"Region: {job.region}",
        f"Batch ID: {job.batch_id}",
        "",
        "Error tail:",
        _tail(detail),
    ])


def _set_run(run_id: str, **updates: Any) -> None:
    with _registry_lock:
        if run_id in _registry:
            _registry[run_id].update(updates)


def _fail_run(job: AsateelJob, title: str, number: str, detail: str) -> None:
    subject = f"Asateel {title} {number}-2026 — FAILED"
    email_sent = _send_email(job.recipients, subject, _failure_body(job, detail), [])
    _set_run(job.run_id, status="failed", finished_at=_utc_now(), error=_tail(detail), email_sent=email_sent)


def _execute_job(job: AsateelJob) -> None:
    title, number = region_title_and_number(job.region, job.batch_id)
    _set_run(job.run_id, status="running", started_at=_utc_now())

    stage_cmd = [
        "python3",
        str(ROOT / "scripts" / "asateel_stage_batch.py"),
        "--batch-id",
        job.batch_id,
        "--pre-staged",
    ]
    try:
        staged = _run_command(stage_cmd)
    except Exception as exc:
        _fail_run(job, title, number, str(exc))
        return
    if staged.returncode != 0:
        _fail_run(job, title, number, staged.stdout + "\n" + staged.stderr)
        return

    master_path = _parse_master_path(staged.stdout)
    if master_path is None:
        _fail_run(job, title, number, staged.stdout + "\nExpenses-Format master was not resolved")
        return

    src_dir = _batch_dir(job.batch_id) / "src"
    pipeline_cmd = [
        "python3",
        str(ROOT / "pipelines" / "asateel.py"),
        "--folder",
        REGION_ENGINE_FOLDERS[job.region],
        "--full",
        "--pdf-dir",
        str(src_dir),
        "--expenses-format",
        str(master_path),
        "--so-detail",
        str(SO_DETAIL),
    ]
    try:
        completed = _run_command(pipeline_cmd)
    except Exception as exc:
        _fail_run(job, title, number, str(exc))
        return
    if completed.returncode != 0:
        _fail_run(job, title, number, completed.stdout + "\n" + completed.stderr)
        return

    stable_oracle, missing_jq_xlsx = stable_output_paths(job.region, job.batch_id)
    stable_oracle.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(MATCHED / "asateel-oracle-upload.xlsx", stable_oracle)
    summary = json.loads((MATCHED / "asateel-summary.json").read_text(encoding="utf-8"))
    missing_jq_count = build_missing_jq_sheet(stable_oracle, missing_jq_xlsx)
    counts = _summary_counts(summary, missing_jq_count)
    subject = (
        f"Asateel {title} {number}-2026 — reconciliation complete "
        f"(G{counts['green']}/Y{counts['yellow']}/R{counts['red']})"
    )
    email_sent = _send_email(
        job.recipients,
        subject,
        _success_body(summary, missing_jq_count),
        [stable_oracle, missing_jq_xlsx],
    )
    _set_run(
        job.run_id,
        status="done",
        finished_at=_utc_now(),
        counts=counts,
        email_sent=email_sent,
    )


def _worker_loop() -> None:
    while True:
        job = _job_queue.get()
        try:
            _execute_job(job)
        except Exception as exc:
            title, number = region_title_and_number(job.region, job.batch_id)
            _logger.exception("Unhandled Asateel trigger failure for %s", job.run_id)
            _fail_run(job, title, number, str(exc))
        finally:
            _job_queue.task_done()


def start_worker() -> threading.Thread:
    global _worker_thread
    with _worker_lock:
        if _worker_thread and _worker_thread.is_alive():
            return _worker_thread
        _worker_thread = threading.Thread(target=_worker_loop, name="asateel-trigger-worker", daemon=True)
        _worker_thread.start()
        return _worker_thread


@bp.record_once
def _start_worker_on_registration(_state: Any) -> None:
    start_worker()


@bp.route("/asateel/openapi.json", methods=["GET"])
def get_openapi_json() -> tuple[Response, int]:
    spec, _raw_yaml = _load_openapi_spec()
    if spec is None:
        return _json_error(500, "openapi spec unavailable")
    return jsonify(spec), 200


@bp.route("/asateel/openapi.yaml", methods=["GET"])
def get_openapi_yaml() -> tuple[Response, int]:
    global _openapi_yaml
    if _openapi_yaml is None:
        try:
            _openapi_yaml = OPENAPI_SPEC.read_bytes()
        except OSError as exc:
            _logger.warning("Asateel OpenAPI spec unavailable: %s", exc)
            return _json_error(500, "openapi spec unavailable")
    return Response(_openapi_yaml, mimetype="application/yaml"), 200


@bp.route("/asateel/run", methods=["POST"])
def enqueue_run() -> tuple[Response, int]:
    if not _authorized():
        return _json_error(401, "unauthorized")
    try:
        payload = _parse_payload(request.get_json(silent=True))
    except ValueError as exc:
        return _json_error(400, str(exc))

    run_id = uuid.uuid4().hex
    created_at = _utc_now()
    job = AsateelJob(run_id=run_id, **payload)
    queue_position = _job_queue.qsize() + 1
    with _registry_lock:
        _registry[run_id] = {
            "run_id": run_id,
            "status": "queued",
            "region": job.region,
            "batch_id": job.batch_id,
            "created_at": created_at,
            "started_at": None,
            "finished_at": None,
            "email_sent": None,
        }
    _job_queue.put(job)
    return jsonify({"run_id": run_id, "status": "queued", "queue_position": queue_position}), 202


@bp.route("/asateel/run/<run_id>", methods=["GET"])
def get_run(run_id: str) -> tuple[Response, int]:
    if not _authorized():
        return _json_error(401, "unauthorized")
    with _registry_lock:
        run = dict(_registry.get(run_id) or {})
    if not run:
        return _json_error(404, "not found")
    return jsonify(run), 200
