#!/usr/bin/env python3
"""
Build a standalone Inconsistencies & Fraud-Watch report for a Jawal batch.

Usage:
    python3 scripts/build_inconsistencies_report.py J26-954

Outputs:
    batches/jawal-<BATCH_ID>/output/Inconsistencies-Report-<BATCH_ID>.xlsx
    batches/jawal-<BATCH_ID>/output/Inconsistencies-Report-<BATCH_ID>.md
"""
from __future__ import annotations

import json
import sys
from collections import Counter
from pathlib import Path
from typing import Any, Iterable

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
HEADER_ROW = 3
DATA_ROW_OFFSET = HEADER_ROW

NAVY = "1E40AF"
WHITE = "FFFFFF"
ERR = "DC2626"
WARN = "D97706"
INFO = "2563EB"
LINE = "E5E7EB"
SAGE = "F0F4F3"

SEVERITY_ORDER = ("HARD", "HIGH", "MEDIUM", "INFO")
GL_COVERAGE_COLUMNS = ("Account", "Cost Center", "DIV", "Solution", "Agency")
CONTEXT_COLUMNS = (
    "Employee No",
    "Description",
    "*Amount",
    *GL_COVERAGE_COLUMNS,
    "Row Status",
    "QC Catches",
)


def load_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def as_list(value: Any) -> list[Any]:
    if value is None:
        return []
    if isinstance(value, list):
        return value
    return [value]


def join_values(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return ", ".join(str(v) for v in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return str(value)


def amount_value(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return None


def severity_bucket(severity: Any) -> str:
    sev = str(severity or "").strip().upper()
    if sev in ("HARD", "HIGH", "MEDIUM", "INFO"):
        return sev
    return "INFO"


def severity_sort_key(row: dict[str, Any]) -> tuple[int, str, str]:
    sev = severity_bucket(row.get("severity"))
    return (SEVERITY_ORDER.index(sev), str(row.get("category", "")), str(row.get("sl_nos", "")))


def resolve_paths(batch_id: str) -> tuple[Path, Path]:
    batch_dir = ROOT / "batches" / f"jawal-{batch_id}"
    output_dir = batch_dir / "output"
    if not batch_dir.exists():
        raise FileNotFoundError(f"Batch directory not found: {batch_dir}")
    if not output_dir.exists():
        raise FileNotFoundError(f"Output directory not found: {output_dir}")
    return batch_dir, output_dir


def load_row_context(output_dir: Path, batch_id: str) -> dict[int, dict[str, Any]]:
    xlsx_path = output_dir / f"Spreadsheet-{batch_id}-FILLED-v30.xlsx"
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Filled v30 spreadsheet not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    headers = [ws.cell(HEADER_ROW, col).value for col in range(1, ws.max_column + 1)]
    hmap = {str(v).strip(): idx + 1 for idx, v in enumerate(headers) if v}

    context: dict[int, dict[str, Any]] = {}
    for sl_no in range(1, max(ws.max_row - DATA_ROW_OFFSET, 0) + 1):
        row_idx = sl_no + DATA_ROW_OFFSET
        row_context: dict[str, Any] = {"Sl. #": sl_no}
        for name in CONTEXT_COLUMNS:
            col_idx = hmap.get(name)
            row_context[name] = ws.cell(row_idx, col_idx).value if col_idx else ""
        context[sl_no] = row_context
    return context


def load_fraud_watch(output_dir: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    seen: set[str] = set()
    for filename in ("fraud-watch-v30.json", "fraud-watch-v24.json"):
        payload = load_json(output_dir / filename, {})
        for catch in payload.get("catches", []) if isinstance(payload, dict) else []:
            key = json.dumps(catch, ensure_ascii=False, sort_keys=True, default=str)
            if key in seen:
                continue
            seen.add(key)
            row = dict(catch)
            row["source_file"] = filename
            rows.append(row)
    return rows


def normalize_catch(raw: dict[str, Any], source: str) -> dict[str, Any]:
    return {
        "source": source,
        "category": raw.get("category", ""),
        "severity": raw.get("severity", ""),
        "tier": raw.get("tier", ""),
        "sl_nos": as_list(raw.get("sl_nos", raw.get("sl_no"))),
        "passenger": raw.get("passenger", ""),
        "ticket_no": raw.get("ticket_no", ""),
        "related_ticket": raw.get("related_ticket", ""),
        "route": raw.get("route", ""),
        "emp_no": raw.get("emp_no", ""),
        "account": raw.get("account", ""),
        "value_at_risk_sar": raw.get("value_at_risk_sar", ""),
        "detail": raw.get("detail", ""),
        "raw": raw,
    }


def load_all_catches(output_dir: Path) -> list[dict[str, Any]]:
    catches: list[dict[str, Any]] = []
    for filename, source in (
        ("catches-within-batch.json", "within_batch"),
        ("catches-cross-batch.json", "cross_batch"),
    ):
        for raw in load_json(output_dir / filename, []):
            if isinstance(raw, dict):
                catches.append(normalize_catch(raw, source))

    booking = load_json(output_dir / "catches-booking-groups-v30.json", {})
    if isinstance(booking, dict):
        for raw in booking.get("conflict_details", []) or []:
            if not isinstance(raw, dict):
                continue
            row = normalize_catch(raw, "booking_groups")
            row["category"] = row["category"] or "BOOKING_GROUP_CONFLICT"
            row["severity"] = row["severity"] or "HIGH"
            row["detail"] = row["detail"] or json.dumps(raw, ensure_ascii=False, sort_keys=True)
            catches.append(row)
    return catches


def expand_with_context(catches: list[dict[str, Any]], context_by_sl: dict[int, dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for catch in catches:
        sl_nos = [int(v) for v in catch.get("sl_nos") or [] if str(v).isdigit()]
        if not sl_nos:
            rows.append(build_catch_row(catch, None, {}))
            continue
        for sl_no in sl_nos:
            rows.append(build_catch_row(catch, sl_no, context_by_sl.get(sl_no, {})))
    return rows


def build_catch_row(catch: dict[str, Any], sl_no: int | None, context: dict[str, Any]) -> dict[str, Any]:
    return {
        "source": catch.get("source", ""),
        "category": catch.get("category", ""),
        "severity": catch.get("severity", ""),
        "severity_bucket": severity_bucket(catch.get("severity")),
        "tier": catch.get("tier", ""),
        "Sl. #": sl_no or "",
        "all_sl_nos": join_values(catch.get("sl_nos")),
        "passenger": catch.get("passenger", ""),
        "ticket_no": catch.get("ticket_no", ""),
        "related_ticket": catch.get("related_ticket", ""),
        "route": catch.get("route", ""),
        "emp_no": catch.get("emp_no", ""),
        "account": catch.get("account", ""),
        "value_at_risk_sar": catch.get("value_at_risk_sar", ""),
        "Employee No": context.get("Employee No", ""),
        "Description": context.get("Description", ""),
        "*Amount": context.get("*Amount", ""),
        "Account": context.get("Account", ""),
        "Row Status": context.get("Row Status", ""),
        "QC Catches": context.get("QC Catches", ""),
        "detail": catch.get("detail", ""),
    }


def hard_inconsistency_rows(all_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        row
        for row in all_rows
        if severity_bucket(row.get("severity")) in ("HARD", "HIGH")
        or str(row.get("category", "")).upper() == "BOOKING_GROUP_CONFLICT"
    ]


def category_counts(*row_groups: list[dict[str, Any]]) -> Counter[str]:
    counts: Counter[str] = Counter()
    for rows in row_groups:
        for row in rows:
            cat = str(row.get("category", "") or "UNKNOWN")
            counts[cat] += 1
    return counts


def severity_counts(*row_groups: list[dict[str, Any]]) -> Counter[str]:
    counts: Counter[str] = Counter({sev: 0 for sev in SEVERITY_ORDER})
    for rows in row_groups:
        for row in rows:
            counts[severity_bucket(row.get("severity"))] += 1
    return counts


def unique_flagged_amount_total(rows: list[dict[str, Any]]) -> tuple[int, float]:
    by_sl: dict[int, float] = {}
    for row in rows:
        sl_no = row.get("Sl. #")
        if not isinstance(sl_no, int):
            continue
        amount = amount_value(row.get("*Amount"))
        if amount is not None:
            by_sl[sl_no] = amount
    return len(by_sl), sum(by_sl.values())


def gl_allocation_coverage(rows: Iterable[dict[str, Any]]) -> dict[str, Any]:
    """Return completeness for Oracle GL allocation cells.

    Counted cells are Account, Cost Center, DIV, Solution, and Agency. Employee
    No/emp_no is intentionally excluded because sponsorship rows can validly
    have a blank or non-scorable employee number while still requiring the GL
    allocation combo to be complete.
    """
    rows = list(rows)
    total = len(rows) * len(GL_COVERAGE_COLUMNS)
    filled = 0
    for row in rows:
        for column in GL_COVERAGE_COLUMNS:
            if str(row.get(column) or "").strip():
                filled += 1
    blank = total - filled
    return {
        "gl_cells_total": total,
        "gl_cells_filled": filled,
        "gl_cells_blank": blank,
        "coverage_pct": round(filled / total * 100, 1) if total else None,
    }


def write_rows(ws: openpyxl.worksheet.worksheet.Worksheet, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([join_values(row.get(header)) for header in headers])
    style_sheet(ws)
    severity_col = None
    for idx, header in enumerate(headers, 1):
        if header == "severity":
            severity_col = idx
            break
    if severity_col:
        color_severity_cells(ws, severity_col)


def style_sheet(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    header_fill = PatternFill("solid", fgColor=NAVY)
    header_font = Font(color=WHITE, bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 60)


def color_severity_cells(ws: openpyxl.worksheet.worksheet.Worksheet, severity_col: int) -> None:
    fills = {
        "HARD": PatternFill("solid", fgColor=ERR),
        "HIGH": PatternFill("solid", fgColor=ERR),
        "MEDIUM": PatternFill("solid", fgColor=WARN),
        "INFO": PatternFill("solid", fgColor=INFO),
        "LOW": PatternFill("solid", fgColor=INFO),
    }
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row_idx, severity_col)
        sev = str(cell.value or "").strip().upper()
        bucket = sev if sev in fills else severity_bucket(sev)
        cell.fill = fills.get(bucket, fills["INFO"])
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def build_summary_sheet(
    wb: openpyxl.Workbook,
    batch_id: str,
    summary: dict[str, Any],
    severity_count: Counter[str],
    category_count: Counter[str],
    total_flagged_rows: int,
    at_risk_total: float,
    hard_count: int,
    fraud_count: int,
) -> None:
    ws = wb.active
    ws.title = "Summary"
    rows = [
        ("Metric", "Value"),
        ("Batch ID", batch_id),
        ("Total lines", summary.get("total_lines", "")),
        ("Clean lines", summary.get("clean_lines", "")),
        ("Soft flags", summary.get("soft_flags", "")),
        ("Total flagged rows", total_flagged_rows),
        ("SAR at-risk total", round(at_risk_total, 2)),
        ("Fraud-watch catches", fraud_count),
        ("Hard inconsistencies", hard_count),
        ("hard_failures", summary.get("hard_failures", "")),
        ("halt_queue_count", summary.get("halt_queue_count", 0)),
        ("", ""),
        ("Severity", "Count"),
    ]
    for sev in SEVERITY_ORDER:
        rows.append((sev, severity_count.get(sev, 0)))
    rows.append(("", ""))
    rows.append(("Category", "Count"))
    for category, count in category_count.most_common():
        rows.append((category, count))
    rows.append(("", ""))
    rows.append(("Pipeline flag_breakdown", "Count"))
    for category, count in Counter(summary.get("flag_breakdown", {}) or {}).most_common():
        rows.append((category, count))

    for row in rows:
        ws.append(row)
    style_sheet(ws)
    severity_fills = {
        "HARD": PatternFill("solid", fgColor=ERR),
        "HIGH": PatternFill("solid", fgColor=ERR),
        "MEDIUM": PatternFill("solid", fgColor=WARN),
        "INFO": PatternFill("solid", fgColor=INFO),
    }
    for row_idx in range(2, ws.max_row + 1):
        label = str(ws.cell(row_idx, 1).value or "")
        if label in severity_fills:
            ws.cell(row_idx, 1).fill = severity_fills[label]
            ws.cell(row_idx, 1).font = Font(color=WHITE, bold=True)
    for row_idx in range(1, ws.max_row + 1):
        if ws.cell(row_idx, 1).value in ("Severity", "Category", "Pipeline flag_breakdown"):
            for col_idx in (1, 2):
                ws.cell(row_idx, col_idx).fill = PatternFill("solid", fgColor=SAGE)
                ws.cell(row_idx, col_idx).font = Font(bold=True)


def write_workbook(
    output_dir: Path,
    batch_id: str,
    summary: dict[str, Any],
    fraud_rows: list[dict[str, Any]],
    hard_rows: list[dict[str, Any]],
    all_rows: list[dict[str, Any]],
    severity_count: Counter[str],
    category_count: Counter[str],
    total_flagged_rows: int,
    at_risk_total: float,
) -> Path:
    wb = openpyxl.Workbook()
    build_summary_sheet(
        wb,
        batch_id,
        summary,
        severity_count,
        category_count,
        total_flagged_rows,
        at_risk_total,
        len(hard_rows),
        len(fraud_rows),
    )

    fraud_headers = [
        "category",
        "severity",
        "passenger",
        "amount_sar",
        "trip_count",
        "distinct_routes",
        "sl_nos",
        "detail",
        "source_file",
    ]
    ws = wb.create_sheet("Fraud Watch")
    write_rows(ws, fraud_headers, fraud_rows)

    catch_headers = [
        "source",
        "category",
        "severity",
        "tier",
        "Sl. #",
        "all_sl_nos",
        "passenger",
        "ticket_no",
        "related_ticket",
        "route",
        "emp_no",
        "account",
        "value_at_risk_sar",
        "Employee No",
        "Description",
        "*Amount",
        "Account",
        "Row Status",
        "QC Catches",
        "detail",
    ]
    ws = wb.create_sheet("Hard Inconsistencies")
    write_rows(ws, catch_headers, sorted(hard_rows, key=severity_sort_key))

    ws = wb.create_sheet("All Catches")
    write_rows(ws, catch_headers, sorted(all_rows, key=severity_sort_key))

    out_path = output_dir / f"Inconsistencies-Report-{batch_id}.xlsx"
    wb.save(out_path)
    return out_path


def write_markdown(
    output_dir: Path,
    batch_id: str,
    fraud_rows: list[dict[str, Any]],
    hard_rows: list[dict[str, Any]],
    all_rows: list[dict[str, Any]],
    category_count: Counter[str],
    total_flagged_rows: int,
    at_risk_total: float,
) -> Path:
    fraud_headlines = []
    for row in fraud_rows[:5]:
        fraud_headlines.append(
            f"- {row.get('category', 'UNKNOWN')}: {row.get('passenger', '')} "
            f"(Sl. # {join_values(row.get('sl_nos'))}, SAR {row.get('amount_sar', '')})"
        )
    if not fraud_headlines:
        fraud_headlines.append("- None.")

    top_categories = category_count.most_common(8)
    top_category_lines = [f"- {category}: {count}" for category, count in top_categories] or ["- None."]

    text = "\n".join(
        [
            f"# Inconsistencies & Fraud-Watch Report - {batch_id}",
            "",
            f"- Total flagged rows: {total_flagged_rows}",
            f"- SAR at-risk total from flagged rows: {at_risk_total:,.2f}",
            f"- Fraud-watch catches: {len(fraud_rows)}",
            f"- Hard inconsistencies: {len(hard_rows)}",
            f"- Total catch rows in workbook: {len(all_rows)}",
            "",
            "## Fraud-Watch Headline",
            *fraud_headlines,
            "",
            "## Top Categories",
            *top_category_lines,
            "",
        ]
    )
    out_path = output_dir / f"Inconsistencies-Report-{batch_id}.md"
    out_path.write_text(text, encoding="utf-8")
    return out_path


def build_report(batch_id: str) -> dict[str, Any]:
    _, output_dir = resolve_paths(batch_id)
    summary = load_json(output_dir / "summary-v15.11.2.json", {})
    context_by_sl = load_row_context(output_dir, batch_id)
    fraud_rows = load_fraud_watch(output_dir)
    catches = load_all_catches(output_dir)
    all_rows = expand_with_context(catches, context_by_sl)
    hard_rows = hard_inconsistency_rows(all_rows)

    fraud_count_rows = [
        {"category": row.get("category", ""), "severity": row.get("severity", "")}
        for row in fraud_rows
    ]
    severity_count = severity_counts(all_rows, fraud_count_rows)
    category_count = category_counts(all_rows, fraud_count_rows)
    total_flagged_rows, at_risk_total = unique_flagged_amount_total(all_rows)
    coverage = gl_allocation_coverage(context_by_sl.values())

    xlsx_path = write_workbook(
        output_dir,
        batch_id,
        summary,
        fraud_rows,
        hard_rows,
        all_rows,
        severity_count,
        category_count,
        total_flagged_rows,
        at_risk_total,
    )
    md_path = write_markdown(
        output_dir,
        batch_id,
        fraud_rows,
        hard_rows,
        all_rows,
        category_count,
        total_flagged_rows,
        at_risk_total,
    )

    return {
        "batch_id": batch_id,
        "xlsx": str(xlsx_path),
        "markdown": str(md_path),
        "rows": {
            "Summary": 1,
            "Fraud Watch": len(fraud_rows),
            "Hard Inconsistencies": len(hard_rows),
            "All Catches": len(all_rows),
        },
        "total_flagged_rows": total_flagged_rows,
        "at_risk_total": round(at_risk_total, 2),
        **coverage,
    }


def main(argv: list[str]) -> int:
    if len(argv) != 2:
        print("Usage: python3 scripts/build_inconsistencies_report.py <BATCH_ID>", file=sys.stderr)
        return 2
    batch_id = argv[1].strip()
    try:
        result = build_report(batch_id)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    print(f"Wrote: {result['xlsx']}")
    print(f"Wrote: {result['markdown']}")
    print(
        "Rows: "
        + ", ".join(f"{sheet}={count}" for sheet, count in result["rows"].items())
        + f"; flagged_rows={result['total_flagged_rows']}; at_risk_sar={result['at_risk_total']:.2f}"
        + f"; coverage_pct={result['coverage_pct']}; gl_cells_total={result['gl_cells_total']}"
        + f"; gl_cells_filled={result['gl_cells_filled']}; gl_cells_blank={result['gl_cells_blank']}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
