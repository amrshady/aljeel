#!/usr/bin/env python3
"""
Build Oracle Fusion ingestion files for each deliverable.

Strategy: produce the file FORMAT Aljeel finance currently ingests into Oracle
Fusion, with agent annotations sitting in additional columns that can be
dropped/hidden when uploading. The agent's job isn't to replace Oracle's data
structure — it's to pre-resolve the exception lines so the upload posts cleanly.

Outputs:
- dashboard/public/files/oracle-jj-<inv>.xlsx       (mirrors Oracle PO-match report shape)
- dashboard/public/files/asateel-allocation-<period>.xlsx (mirrors Aljeel manual allocation Excel)
- dashboard/public/files/jawal-<invoice>-resolved.xlsx    (mirrors J26-640 Details sheet)
"""
from __future__ import annotations

import json
import shutil
from pathlib import Path
from datetime import datetime, timezone

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path('/home/clawdbot/.openclaw/workspace/aljeel')
# Was 'files' but that collides with the /files/* reverse-proxy route in the
# dashboard worker (which sends /files/<path> to the files-mirror project).
# Use 'outputs' so Excel downloads come straight from this Pages deployment.
DASH_FILES = ROOT / 'dashboard' / 'public' / 'outputs'
DASH_FILES.mkdir(parents=True, exist_ok=True)

# Colours
GREEN_FILL  = PatternFill('solid', fgColor='C6EFCE')   # matched / clean
AMBER_FILL  = PatternFill('solid', fgColor='FFEB9C')   # exception, action required
RED_FILL    = PatternFill('solid', fgColor='FFC7CE')   # exception, hard stop
GREY_FILL   = PatternFill('solid', fgColor='F2F2F2')   # informational
HDR_FILL    = PatternFill('solid', fgColor='1E40AF')
HDR_FONT    = Font(color='FFFFFF', bold=True)
BORDER_THIN = Border(left=Side(style='thin', color='D0D0D0'),
                     right=Side(style='thin', color='D0D0D0'),
                     top=Side(style='thin', color='D0D0D0'),
                     bottom=Side(style='thin', color='D0D0D0'))


# ---------------------------------------------------------------------------
# J&J — Annotate the Oracle PO-match report with agent results
# ---------------------------------------------------------------------------

def build_oracle_jj():
    """For each J&J invoice, take the original Oracle PO-match Excel and add
    three columns at the right: agent_match_status, agent_flags, agent_action.
    Output goes to dashboard/public/files/oracle-jj-<inv>.xlsx.

    Logic: read matched/<inv>.match.json; for each Oracle row, find the matching
    aggregated invoice item (by Inventory Item code OR description) and pull
    its match_status + flags.
    """
    matched_dir = ROOT / 'matched'
    raw_dir = ROOT / 'raw' / 'asn-25DEC' / '25DEC048-25DEC054'

    # Discovery: find Oracle reports
    from discover import _scan_jj_files, _extract_invoice_no
    inv_pdfs, dn_books, oracle_books = _scan_jj_files()

    outputs = []
    for pdf in inv_pdfs:
        inv_no = _extract_invoice_no(pdf)
        if not inv_no:
            continue
        match_path = matched_dir / f'{inv_no}.match.json'
        if not match_path.exists():
            continue
        oracle = next((o for o in oracle_books if inv_no in o.name), None)
        if not oracle:
            continue

        match_data = json.loads(match_path.read_text())
        # Build lookup: inventory_item -> aggregated item
        agg_by_code = {}
        agg_by_desc = {}
        for it in match_data['items']:
            for code in (it.get('vendor_codes') or []):
                agg_by_code[str(code).strip().upper()] = it
            desc_norm = (it.get('description') or '').strip().lower()
            if desc_norm:
                agg_by_desc[desc_norm] = it

        # Load Oracle into openpyxl (preserves formatting)
        out_path = DASH_FILES / f'oracle-jj-{inv_no}.xlsx'
        shutil.copy(oracle, out_path)
        wb = load_workbook(out_path)
        ws = wb.active

        # Find headers row (row 1 contains column names)
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        # Find columns we need: 'Inventory Item', 'Item Description', 'Unnamed: 3' (amount)
        inv_item_col = None
        desc_col = None
        for i, h in enumerate(headers):
            if h == 'Inventory Item':
                inv_item_col = i + 1
            elif h == 'Item Description':
                desc_col = i + 1

        # Add 3 new columns at the right
        next_col = ws.max_column + 1
        agent_status_col = next_col
        agent_flags_col = next_col + 1
        agent_action_col = next_col + 2

        ws.cell(row=1, column=agent_status_col, value='Agent Match Status').fill = HDR_FILL
        ws.cell(row=1, column=agent_status_col).font = HDR_FONT
        ws.cell(row=1, column=agent_flags_col, value='Agent Flags').fill = HDR_FILL
        ws.cell(row=1, column=agent_flags_col).font = HDR_FONT
        ws.cell(row=1, column=agent_action_col, value='Agent Action').fill = HDR_FILL
        ws.cell(row=1, column=agent_action_col).font = HDR_FONT

        # Walk rows
        for row in range(2, ws.max_row + 1):
            # Pull inv_item code if present
            agg = None
            if inv_item_col:
                code = ws.cell(row=row, column=inv_item_col).value
                if code:
                    agg = agg_by_code.get(str(code).strip().upper())
            if agg is None and desc_col:
                d = ws.cell(row=row, column=desc_col).value
                if d:
                    agg = agg_by_desc.get(str(d).strip().lower())
            if agg is None:
                continue
            status = agg.get('match_status', 'UNKNOWN')
            flags = ', '.join(agg.get('flags') or [])
            if status == 'MATCHED':
                action = 'Post to GL — no further review required'
                fill = GREEN_FILL
            elif 'NO_PO_MATCH' in (agg.get('flags') or []) or 'NO_GR' in (agg.get('flags') or []):
                action = 'HOLD — line missing PO or goods receipt'
                fill = RED_FILL
            elif 'QTY_BREAK' in (agg.get('flags') or []) or 'QTY_RECV_MISMATCH' in (agg.get('flags') or []):
                action = 'REVIEW — quantity variance vs PO or receipt'
                fill = AMBER_FILL
            else:
                action = 'REVIEW'
                fill = AMBER_FILL

            for c in (agent_status_col, agent_flags_col, agent_action_col):
                ws.cell(row=row, column=c).fill = fill
                ws.cell(row=row, column=c).border = BORDER_THIN
            ws.cell(row=row, column=agent_status_col, value=status)
            ws.cell(row=row, column=agent_flags_col, value=flags)
            ws.cell(row=row, column=agent_action_col, value=action)

        # Auto-size new columns
        for c in (agent_status_col, agent_flags_col, agent_action_col):
            ws.column_dimensions[get_column_letter(c)].width = 28

        wb.save(out_path)
        outputs.append({
            'invoice_no': inv_no,
            'path':       '/' + str(out_path.relative_to(ROOT / 'dashboard' / 'public')).replace('\\', '/'),
            'size_bytes': out_path.stat().st_size,
        })
        print(f'  Wrote {out_path.name} ({out_path.stat().st_size} bytes)')

    return outputs


# ---------------------------------------------------------------------------
# Asateel — Reproduce the manual allocation Excel
# ---------------------------------------------------------------------------

ASATEEL_COLS = [
    'Intercompany*Invoice Number', '*Invoice Currency', '*Invoice Amount', '*Invoice Date',
    '#', 'Expense Date', 'Employee Name', 'Employee Number', 'JQ', 'Description',
    'EXPENSE TYPE', 'Agency', 'Division', 'Solution', 'Cost Center', 'AMOUNT',
    'Unnamed: 16', 'Company', 'location', 'Account', 'GL', 'Cost Center.1', 'Cost Name',
    'DIV', 'Contribution', 'Solution.1', 'Solution Name', 'Agency.1', 'Agency Name',
    'Project', 'Intercompany', 'Future 1', 'Unnamed: 32', 'Unnamed: 33'
]


def build_oracle_asateel():
    """Reproduce the Aljeel finance allocation Excel shape from the matched data.

    Output mirrors the Details sheet structure exactly. Aljeel finance can drop
    this into Oracle Fusion's expense-allocation upload template.
    """
    matched_path = ROOT / 'matched' / 'asateel-allocation.json'
    catch_path = ROOT / 'matched' / 'asateel-catch.json'
    if not matched_path.exists():
        print('  [skip] no asateel-allocation.json')
        return []

    from discover import discover_asateel
    cfg = discover_asateel()

    recon = json.loads(matched_path.read_text())
    catches = json.loads(catch_path.read_text())
    # Build lookup of which invoices have catches
    catch_by_inv = {}
    for c in catches:
        inv = c.get('invoice_no')
        if inv:
            catch_by_inv.setdefault(inv, []).append(c)

    # Re-load the original allocation Excel so we have full per-row data
    orig = pd.read_excel(cfg.allocation_xlsx, sheet_name='Details', header=0)
    # Forward-fill invoice header columns to enrich each allocation row
    orig['invoice_no_filled'] = orig['Intercompany*Invoice Number'].ffill()
    orig['invoice_currency_filled'] = orig['*Invoice Currency'].ffill()
    orig['invoice_amount_filled'] = orig['*Invoice Amount'].ffill()
    orig['invoice_date_filled'] = orig['*Invoice Date'].ffill()

    out_path = DASH_FILES / 'asateel-allocation-ready.xlsx'
    # Date-stamp the period from the data — Excel serial-int detection needed
    def _safe_parse_date(v):
        if pd.isna(v):
            return pd.NaT
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            return pd.Timestamp('1899-12-30') + pd.Timedelta(days=int(v))
        try:
            return pd.to_datetime(v, errors='coerce')
        except Exception:
            return pd.NaT
    period_dates = orig['invoice_date_filled'].apply(_safe_parse_date).dropna()
    if not period_dates.empty:
        period_label = period_dates.min().strftime('%b-%Y')
    else:
        period_label = 'unknown'

    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        # Write Details sheet — same shape as input
        details_out = orig[[c for c in orig.columns if not c.endswith('_filled')]].copy()
        details_out.to_excel(writer, sheet_name='Details', index=False)
        # Add 'Agent Review' column at the end
        # Then add a Reconciliation Summary sheet with per-invoice status
        summary_rows = []
        for r in recon:
            inv_catches = catch_by_inv.get(r['invoice_no'], [])
            catch_cats = ','.join(sorted({c['category'] for c in inv_catches})) or 'CLEAN'
            summary_rows.append({
                'Invoice No':       r['invoice_no'],
                'Invoice Date':     r.get('invoice_date'),
                'Header Total SAR': r['header_total'],
                'Allocation Net':   r['allocation_sum'],
                'Expected Gross':   r['expected_gross_vat15'],
                'Delta SAR':        r['delta'],
                'Reconciled':       'YES' if r['reconciled'] else 'NO',
                'Allocation Lines': r['allocation_lines'],
                'Employees':        r['distinct_employees'],
                'Agent Flags':      catch_cats,
            })
        summary_df = pd.DataFrame(summary_rows)
        summary_df.to_excel(writer, sheet_name='Reconciliation Summary', index=False)

        # Add a Catches sheet
        catches_df = pd.DataFrame(catches)
        if not catches_df.empty:
            catches_df.to_excel(writer, sheet_name='Exceptions for Review', index=False)

    # Post-process: highlight the summary sheet
    wb = load_workbook(out_path)
    ws = wb['Reconciliation Summary']
    for row in range(2, ws.max_row + 1):
        reconciled = ws.cell(row=row, column=7).value
        flags = ws.cell(row=row, column=10).value
        if reconciled == 'YES' and flags == 'CLEAN':
            fill = GREEN_FILL
        elif reconciled == 'NO':
            fill = RED_FILL
        else:
            fill = AMBER_FILL
        for c in range(1, ws.max_column + 1):
            ws.cell(row=row, column=c).fill = fill
            ws.cell(row=row, column=c).border = BORDER_THIN
    # Header row
    for c in range(1, ws.max_column + 1):
        ws.cell(row=1, column=c).fill = HDR_FILL
        ws.cell(row=1, column=c).font = HDR_FONT
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    wb.save(out_path)

    print(f'  Wrote {out_path.name} ({out_path.stat().st_size} bytes)')
    return [{
        'name':       'Asateel allocation (Oracle-ready)',
        'period':     period_label,
        'path':       '/' + str(out_path.relative_to(ROOT / 'dashboard' / 'public')).replace('\\', '/'),
        'size_bytes': out_path.stat().st_size,
    }]


# ---------------------------------------------------------------------------
# Jawal — Exception-resolved version of J26-640
# ---------------------------------------------------------------------------

def build_oracle_jawal():
    """Mirror J26-640.xlsx Details sheet shape + add agent review columns.
    Aljeel finance can ingest this directly into Oracle Fusion."""
    from discover import discover_jawal
    cfg = discover_jawal()

    recon_path = ROOT / 'matched' / 'jawal-reconciliation.json'
    catch_path = ROOT / 'matched' / 'jawal-catch.json'
    recon = json.loads(recon_path.read_text())
    catches = json.loads(catch_path.read_text())

    # Build lookup of catches by ticket
    catches_by_tkt = {}
    for c in catches:
        tkt = c.get('ticket_no')
        if tkt:
            catches_by_tkt.setdefault(tkt, []).append(c)

    # Recon lookup
    recon_by_tkt_row = {(r['sl_no']): r for r in recon if r.get('sl_no')}

    # Copy the original xlsx and annotate
    out_path = DASH_FILES / f'jawal-{cfg.invoice_xlsx.stem}-resolved.xlsx'
    shutil.copy(cfg.invoice_xlsx, out_path)
    wb = load_workbook(out_path)
    ws = wb['Details']

    # Find Sl. No column (col 0 / A) and add agent columns at the end
    next_col = ws.max_column + 1
    flag_col = next_col
    action_col = next_col + 1
    folder_col = next_col + 2

    for c, label in [(flag_col, 'Agent Flags'), (action_col, 'Agent Action'), (folder_col, 'Folder Match')]:
        ws.cell(row=1, column=c, value=label).fill = HDR_FILL
        ws.cell(row=1, column=c).font = HDR_FONT

    # Walk rows; col 0 (A) is the Sl. No
    for row in range(2, ws.max_row + 1):
        sl_no = ws.cell(row=row, column=1).value
        if not isinstance(sl_no, (int, float)):
            continue
        r = recon_by_tkt_row.get(int(sl_no))
        if not r:
            continue
        ticket_catches = catches_by_tkt.get(r['ticket_no'], [])
        cats = sorted({c['category'] for c in ticket_catches})
        flag_str = ', '.join(cats) or 'CLEAN'
        if not cats:
            action = 'Post to GL'
            fill = GREEN_FILL
        elif 'NO_FOLDER' in cats:
            action = 'HOLD — request booking + approval from Jawal'
            fill = RED_FILL
        elif 'NO_APPROVAL' in cats:
            action = 'HOLD — manager approval missing'
            fill = RED_FILL
        elif 'PERSONAL_CONTRIB_SELF_APPROVAL' in cats:
            action = 'REVIEW — self-approved personal contribution; verify manager oversight'
            fill = AMBER_FILL
        elif 'EMD_FEE' in cats:
            action = 'Post — EMD fare adjustment, no separate folder expected'
            fill = GREY_FILL
        else:
            action = 'REVIEW'
            fill = AMBER_FILL

        for c in (flag_col, action_col, folder_col):
            ws.cell(row=row, column=c).fill = fill
            ws.cell(row=row, column=c).border = BORDER_THIN
        ws.cell(row=row, column=flag_col, value=flag_str)
        ws.cell(row=row, column=action_col, value=action)
        ws.cell(row=row, column=folder_col, value=r.get('folder_name') or '(no folder)')

    for c in (flag_col, action_col, folder_col):
        ws.column_dimensions[get_column_letter(c)].width = 36

    wb.save(out_path)
    print(f'  Wrote {out_path.name} ({out_path.stat().st_size} bytes)')
    return [{
        'name':       f'Jawal {cfg.invoice_xlsx.stem} (resolved)',
        'path':       '/' + str(out_path.relative_to(ROOT / 'dashboard' / 'public')).replace('\\', '/'),
        'size_bytes': out_path.stat().st_size,
    }]


# ---------------------------------------------------------------------------

def main():
    print('-> Building J&J Oracle-ready files...')
    jj = build_oracle_jj()
    print('-> Building Asateel allocation Excel...')
    asat = build_oracle_asateel()
    print('-> Building Jawal resolved invoice...')
    jawal = build_oracle_jawal()

    # Write manifest for dashboard to pick up
    manifest = {
        'generated_at': datetime.now(timezone.utc).isoformat(),
        'jj':           jj,
        'asateel':      asat,
        'jawal':        jawal,
    }
    (ROOT / 'dashboard' / 'public' / 'data' / 'oracle-files.json').write_text(
        json.dumps(manifest, indent=2), encoding='utf-8')
    print(f'\n=== Manifest written ===\n{json.dumps(manifest, indent=2)}')


if __name__ == '__main__':
    main()
