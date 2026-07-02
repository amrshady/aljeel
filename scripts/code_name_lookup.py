#!/usr/bin/env python3
"""
Code-to-Name Lookup Module for AlJeel AP (Rewired to master Aljeel_Lookups-v2.xlsx).

Loads 5 lookup tables from the master Aljeel_Lookups-v2.xlsx:
  - Account: Account code → GL name
  - Manpower (cols 13 & 14): CC code → CC name
  - DIV & Manpower (cols 9 & 10): DIV code → Division name (resolves DIV 194 perfectly!)
  - Solution: Solution code → Solution name
  - Agency: Agency code → Agency name

Returns '#N/A' (literal text) when lookup fails.
"""
from __future__ import annotations

from pathlib import Path
from functools import lru_cache
from typing import Dict, Optional

import openpyxl

# Default reference file pointing to Aljeel_Lookups-v2.xlsx
DEFAULT_REF = Path("/home/clawdbot/.openclaw/workspace/aljeel/qc/master-data/Aljeel_Lookups-v2.xlsx")

NA = "#N/A"


class CodeNameLookup:
    """Loads all 5 lookup tables once from Aljeel_Lookups-v2.xlsx and exposes per-segment lookup methods."""

    def __init__(self, reference_path: Path | str = DEFAULT_REF):
        # Always force Aljeel_Lookups-v2.xlsx as the reference
        self.reference_path = DEFAULT_REF
        self._account_to_gl: Dict[str, str] = {}
        self._cc_to_name: Dict[str, str] = {}
        self._div_to_name: Dict[str, str] = {}
        self._solution_to_name: Dict[str, str] = {}
        self._agency_to_name: Dict[str, str] = {}
        self._loaded = False

    def _ensure_loaded(self):
        if self._loaded:
            return
        
        # Load workbook without read_only=True to prevent hangs
        wb = openpyxl.load_workbook(self.reference_path, data_only=True)
        
        # 1. Load Account mappings
        ws_acct = wb["Account"]
        for r in range(2, ws_acct.max_row + 1):
            code = ws_acct.cell(r, 1).value
            desc = ws_acct.cell(r, 2).value
            if code is not None and desc is not None:
                self._account_to_gl[str(code).strip()] = str(desc).strip()
                
        # 2. Load Agency mappings
        ws_ag = wb["Agency"]
        for r in range(2, ws_ag.max_row + 1):
            code = ws_ag.cell(r, 1).value
            desc = ws_ag.cell(r, 2).value
            if code is not None and desc is not None:
                self._agency_to_name[str(code).strip()] = str(desc).strip()
                
        # 3. Load Solution mappings
        ws_sol = wb["Solution"]
        for r in range(2, ws_sol.max_row + 1):
            code = ws_sol.cell(r, 1).value
            desc = ws_sol.cell(r, 2).value
            if code is not None and desc is not None:
                self._solution_to_name[str(code).strip()] = str(desc).strip()
                
        # 4. Load Division mappings (from both DIV and Manpower for comprehensive coverage)
        ws_div = wb["DIV"]
        for r in range(2, ws_div.max_row + 1):
            code = ws_div.cell(r, 1).value
            desc = ws_div.cell(r, 2).value
            if code is not None and desc is not None:
                self._div_to_name[str(code).strip()] = str(desc).strip()
                
        # Add Manpower division names as comprehensive fallback (this covers 194 -> IVD Solutions)
        ws_mp = wb["Manpower"]
        for r in range(2, ws_mp.max_row + 1):
            code = ws_mp.cell(r, 9).value
            desc = ws_mp.cell(r, 10).value
            if code is not None and desc is not None:
                self._div_to_name.setdefault(str(code).strip(), str(desc).strip())
                
        # 5. Load Cost Center mappings directly from active Manpower rows (columns 13 & 14)
        for r in range(2, ws_mp.max_row + 1):
            code = ws_mp.cell(r, 13).value
            desc = ws_mp.cell(r, 14).value
            if code is not None and desc is not None:
                self._cc_to_name[str(code).strip()] = str(desc).strip()
                
        wb.close()
        self._loaded = True

    # ---- Public lookup methods ----

    def account_to_gl(self, code: str) -> str:
        self._ensure_loaded()
        return self._account_to_gl.get(str(code).strip(), NA)

    def cc_to_name(self, code: str) -> str:
        self._ensure_loaded()
        return self._cc_to_name.get(str(code).strip(), NA)

    def div_to_name(self, code: str) -> str:
        self._ensure_loaded()
        return self._div_to_name.get(str(code).strip(), NA)

    def solution_to_name(self, code: str) -> str:
        self._ensure_loaded()
        return self._solution_to_name.get(str(code).strip(), NA)

    def agency_to_name(self, code: str) -> str:
        self._ensure_loaded()
        return self._agency_to_name.get(str(code).strip(), NA)

    def _div_name_with_fallback(self, div_code: str, resolved_line=None) -> str:
        self._ensure_loaded()
        if resolved_line is not None and hasattr(resolved_line, 'manpower_div_name'):
            mp_name = resolved_line.manpower_div_name
            if mp_name and mp_name.strip():
                return mp_name.strip()
        return self._div_to_name.get(str(div_code).strip(), NA)

    def expand_combo(self, combo: str, resolved_line=None) -> dict:
        self._ensure_loaded()

        if resolved_line is not None:
            company = str(resolved_line.company)
            location = str(resolved_line.location)
            account = str(resolved_line.account)
            cost_center = str(resolved_line.cost_center)
            div_code = str(resolved_line.div)
            solution = str(resolved_line.solution)
            agency = str(resolved_line.agency)
            project = str(resolved_line.project)
            intercompany = str(resolved_line.intercompany)
            future1 = str(resolved_line.future1)
        elif combo and len(combo) >= 49:
            parts = combo.split("-")
            if len(parts) >= 10:
                company = parts[0]
                location = parts[1]
                account = parts[2]
                cost_center = parts[3]
                div_code = parts[4]
                solution = parts[5]
                agency = parts[6]
                project = parts[7]
                intercompany = parts[8]
                future1 = parts[9]
            else:
                return self._empty_expansion()
        else:
            return self._empty_expansion()

        return {
            "Company": company,
            "Location": location,
            "Account": account,
            "GL": self.account_to_gl(account),
            "Cost Center": cost_center,
            "Cost Name": self.cc_to_name(cost_center),
            "DIV": div_code,
            "Contribution": self._div_name_with_fallback(div_code, resolved_line),
            "Solution": solution,
            "Solution Name": self.solution_to_name(solution),
            "Agency": agency,
            "Agency Name": self.agency_to_name(agency),
            "Project": project,
            "Intercompany": intercompany,
            "Future 1": future1,
        }

    def _empty_expansion(self) -> dict:
        return {
            "Company": "", "Location": "", "Account": "", "GL": NA,
            "Cost Center": "", "Cost Name": NA, "DIV": "", "Contribution": NA,
            "Solution": "", "Solution Name": NA, "Agency": "", "Agency Name": NA,
            "Project": "", "Intercompany": "", "Future 1": "",
        }

    def stats(self) -> dict:
        self._ensure_loaded()
        return {
            "account_to_gl": len(self._account_to_gl),
            "cc_to_name": len(self._cc_to_name),
            "div_to_name": len(self._div_to_name),
            "solution_to_name": len(self._solution_to_name),
            "agency_to_name": len(self._agency_to_name),
        }


# Module-level singleton for convenience (bypasses caller-provided refs to always use v2)
_default_lookup: Optional[CodeNameLookup] = None


def get_lookup(reference_path: Path | str = DEFAULT_REF) -> CodeNameLookup:
    global _default_lookup
    if _default_lookup is None:
        _default_lookup = CodeNameLookup(DEFAULT_REF)
    return _default_lookup


if __name__ == "__main__":
    lookup = CodeNameLookup()
    print("Aljeel Lookups-v2 loaded successfully!")
    print("Lookup table sizes:", lookup.stats())
    print("Test DIV 194 lookup ➔", lookup.div_to_name("194"))
