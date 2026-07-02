import openpyxl, re
from pathlib import Path

ours_path = Path("/home/clawdbot/.openclaw/workspace/aljeel/batches/jawal-J26-788/output/Spreadsheet-J26-788-FILLED-v30.xlsx")
truth_path = Path("/home/clawdbot/.openclaw/media/inbound/Labadi_file_issues_j26-788---b977c06a-a3f6-4c21-b5a8-192bdadb863a.xlsx")

wb_ours = openpyxl.load_workbook(ours_path, data_only=True)
ws_ours = wb_ours.active
headers_ours = [ws_ours.cell(3, col).value for col in range(1, ws_ours.max_column+1)]
col_map_ours = {str(h).strip(): i for i, h in enumerate(headers_ours) if h}

wb_truth = openpyxl.load_workbook(truth_path, data_only=True)
ws_truth = wb_truth.active

TICKET_RE = re.compile(r"(?<!\d)(\d{10}|26-\d{3,4})(?!\d)")

ours_by_ticket = {}
for r in range(4, ws_ours.max_row+1):
    desc = str(ws_ours.cell(r, col_map_ours['Description']+1).value or "")
    m = TICKET_RE.search(desc)
    if m:
        ours_by_ticket[m.group(1)] = r

print(f"| Ticket | Passenger | Column | Ours (v30) | Labadi (Truth) | Status |")
print(f"|---|---|---|---|---|---|")

for r in range(2, ws_truth.max_row+1):
    p_name = str(ws_truth.cell(r, 1).value or "").split(" - ", 1)[0].strip()
    desc_val = str(ws_truth.cell(r, 1).value or "")
    m = TICKET_RE.search(desc_val)
    if not m:
        continue
    ticket = m.group(1)
    if ticket not in ours_by_ticket:
        print(f"| {ticket} | {p_name[:20]} | - | N/A in ours | Present in truth | Discrepancy |")
        continue
    
    ours_row = ours_by_ticket[ticket]
    
    # Check Account
    ours_acc = str(ws_ours.cell(ours_row, col_map_ours['Account']+1).value or "").strip()
    truth_acc = str(ws_truth.cell(r, 7).value or "").strip()
    if ours_acc != truth_acc:
        print(f"| {ticket} | {p_name[:20]} | Account | `{ours_acc}` | `{truth_acc}` | Discrepancy |")
        
    # Check Employee No
    ours_emp = str(ws_ours.cell(ours_row, col_map_ours['Employee No']+1).value or "").strip()
    truth_emp = str(ws_truth.cell(r, 4).value or "").strip()
    # Normalize 'None', empty, and '0'
    ours_emp_norm = "" if ours_emp in ("None", "0", "") else ours_emp
    truth_emp_norm = "" if truth_emp in ("None", "0", "") else truth_emp
    if ours_emp_norm != truth_emp_norm:
        print(f"| {ticket} | {p_name[:20]} | Employee No | `{ours_emp}` | `{truth_emp}` | Discrepancy |")
        
    # Check GL Description
    ours_gldesc = str(ws_ours.cell(ours_row, col_map_ours['GL Description']+1).value or "").strip()
    truth_gldesc = str(ws_truth.cell(r, 8).value or "").strip()
    # Normalize separators and spaces for comparing
    norm_ours = ours_gldesc.replace(" · ", " - ").replace(" ·", " -").replace("· ", "- ").replace("  ", " ").strip()
    norm_truth = truth_gldesc.replace(" · ", " - ").replace(" ·", " -").replace("· ", "- ").replace("  ", " ").strip()
    if norm_ours != norm_truth:
        print(f"| {ticket} | {p_name[:20]} | GL Description | `{ours_gldesc[:50]}...` | `{truth_gldesc[:50]}...` | Discrepancy |")
