#!/usr/bin/env python3
"""
Convert an uploaded Jawal TAX INVOICE workbook into the Stage 1
Spreadsheet-v4-input.xlsx workbook shape.
"""

import argparse
import re
import sys
from copy import copy
from datetime import datetime, date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side


SUPPLIER_NAME = "JAWWAL TRAVEL & TOURISM"
SUPPLIER_NUMBER = "1010122966"
SUPPLIER_SITE = "RIYADH"
BUSINESS_UNIT = "Al Jeel Medical BU"
CURRENCY = "SAR"
INVOICE_TYPE = "Standard"

GL_TRAVEL_SM = "60301003"
GL_TRAVEL_GA = "60301004"
GL_SPONSOR = "60307021"
GL_ANNUAL_TICKET = "21070229"

SM_DIVISIONS = {
    "IVD Solutions",
    "Capital Equipment",
    "Contribution",
    "Dental & Medical Solutions",
    "Technical Services",
    "S&M",
}

HEADERS = [
    "*Invoice Header Identifier",
    "*Business Unit",
    "*Invoice Number",
    "*Invoice Currency",
    "*Invoice Amount",
    "*Invoice Date",
    "**Supplier[..]",
    "**Supplier Number",
    "*Supplier Site[..]",
    "Invoice Type",
    "Description",
    "*Type",
    "*Amount",
    "Distribution Combination[..]",
    "Tax Classification Code[..]",
    "Employee No",
    "Invoice Ref No",
]

GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
HEADER_FILL = PatternFill("solid", fgColor="0070C0")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BORDER_THIN = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def _safe_float(value):
    try:
        if value is None or value == "":
            return 0.0
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _safe_int(value):
    try:
        if value is None or value == "":
            return None
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def _date_string(value):
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%d %b %Y"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return text


def _metadata_value(ws, label, preferred_rows):
    label_lc = label.lower()
    rows = list(preferred_rows) + [r for r in range(1, min(ws.max_row or 60, 60) + 1) if r not in preferred_rows]
    for r in rows:
        for c in range(1, min(ws.max_column or 60, 60) + 1):
            value = ws.cell(r, c).value
            if value is None or label_lc not in str(value).lower():
                continue
            for c2 in range(c + 1, min(ws.max_column or 60, 60) + 1):
                candidate = ws.cell(r, c2).value
                if candidate is None:
                    continue
                text = str(candidate).strip()
                if text and text != ":":
                    return candidate
    return None


def read_invoice_metadata(ws):
    invoice_no = _metadata_value(ws, "Invoice Number", [5]) or ws.cell(5, 3).value or ws.cell(5, 8).value
    invoice_date = _metadata_value(ws, "Invoice Date", [8]) or ws.cell(8, 3).value or ws.cell(8, 8).value
    supplier = _metadata_value(ws, "Name", [13]) or ws.cell(13, 3).value or ws.cell(13, 4).value
    return {
        "invoice_no": str(invoice_no or "").strip(),
        "invoice_date": _date_string(invoice_date),
        "supplier_name": str(supplier or SUPPLIER_NAME).strip() or SUPPLIER_NAME,
    }


def extract_invoice_lines(invoice_file):
    wb = load_workbook(invoice_file, data_only=True)
    ws = wb["Sheet"] if "Sheet" in wb.sheetnames else wb[wb.sheetnames[0]]
    metadata = read_invoice_metadata(ws)
    lines = []

    for r in range(28, (ws.max_row or 28) + 1):
        sl_no = _safe_int(ws.cell(r, 1).value)
        if sl_no is None:
            break

        ticket_raw = str(ws.cell(r, 12).value or "").strip()
        ticket_no = ticket_raw.split()[-1] if " " in ticket_raw else ticket_raw
        lines.append(
            {
                "sl_no": sl_no,
                "issue_date": _date_string(ws.cell(r, 2).value),
                "ref_no": str(ws.cell(r, 7).value or "").strip(),
                "ticket_no_raw": ticket_raw,
                "ticket_no": ticket_no,
                "passenger_name": str(ws.cell(r, 16).value or "").strip(),
                "route": str(ws.cell(r, 22).value or "").strip(),
                "service_date": _date_string(ws.cell(r, 30).value),
                "taxable_amt": _safe_float(ws.cell(r, 34).value),
                "vat_pct": _safe_float(ws.cell(r, 39).value),
                "vat_amt": _safe_float(ws.cell(r, 42).value),
                "inv_amt_incl_vat": _safe_float(ws.cell(r, 47).value),
            }
        )

    return metadata, lines


def _norm_header(value):
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def load_manpower(master_data):
    wb = load_workbook(master_data, data_only=True)
    if "Manpower" not in wb.sheetnames:
        raise ValueError(f"Manpower sheet not found in {master_data}")
    ws = wb["Manpower"]

    header_row = None
    for r in range(1, min(ws.max_row, 12) + 1):
        row_headers = [_norm_header(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if "empno" in row_headers and "name" in row_headers:
            header_row = r
            break
    if header_row is None:
        header_row = 7 if ws.max_row >= 8 else 1

    headers = {_norm_header(ws.cell(header_row, c).value): c for c in range(1, ws.max_column + 1)}

    def col(*names, fallback):
        for name in names:
            key = _norm_header(name)
            if key in headers:
                return headers[key]
        return fallback

    emp_col = col("Emp No", fallback=1)
    name_col = col("Name", fallback=3)
    div_col = col("New Division", fallback=10)
    cc_col = col("New cost center", fallback=13)
    cc_name_col = col("New cost center name", fallback=14)
    solution_col = col("Solution", fallback=16)

    employees = {}
    for r in range(header_row + 1, ws.max_row + 1):
        emp_no = _safe_int(ws.cell(r, emp_col).value)
        if emp_no is None:
            continue
        employees[emp_no] = {
            "emp_no": emp_no,
            "name": str(ws.cell(r, name_col).value or "").strip(),
            "division": str(ws.cell(r, div_col).value or "").strip(),
            "cost_center": _safe_int(ws.cell(r, cc_col).value),
            "cost_center_name": str(ws.cell(r, cc_name_col).value or "").strip(),
            "solution": str(ws.cell(r, solution_col).value or "").strip(),
        }
    return employees


def extract_emp_no_from_ref(ref_no):
    ref_str = str(ref_no or "").strip()
    match = re.match(r"(\d+)", ref_str)
    if not match:
        return None

    emp_no = int(match.group(1))
    if emp_no >= 100000:
        return emp_no
    return None


def classify_line(line, employees):
    emp_no = extract_emp_no_from_ref(line.get("ref_no"))
    emp = employees.get(emp_no) if emp_no else None
    ref_no = str(line.get("ref_no") or "").lower()

    sponsor_keywords = (
        "sponsor",
        "opex",
        "forum",
        "registration",
        "conference",
        "congress",
        "exhibition",
        "summit",
        "symposium",
        "iepc",
        "crm",
        "hotel",
        "reservation",
    )
    # "family" — invoice Ref. No.s like "1000862 family" mark family annual
    # travel groups (Labadi GL rule: employee+family annual leave → 21070229).
    annual_keywords = ("annual ticket", "annual leave", "vacation ticket", "family")

    if any(keyword in ref_no for keyword in sponsor_keywords):
        gl = GL_SPONSOR
    elif any(keyword in ref_no for keyword in annual_keywords):
        gl = GL_ANNUAL_TICKET
    elif emp:
        division = emp["division"]
        gl = GL_TRAVEL_SM if division in SM_DIVISIONS else GL_TRAVEL_GA if division == "G&A" else GL_TRAVEL_SM
    else:
        gl = GL_TRAVEL_SM

    return {
        "gl_account": gl,
        "cost_center": emp.get("cost_center") if emp else None,
        "emp_no": emp_no if emp else None,
    }


def _create_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for c, header in enumerate(HEADERS, start=1):
        cell = ws.cell(3, c, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER_THIN
    return wb


def _copy_cell_style(source, target):
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def prepare_output_workbook(output_path):
    if output_path.exists():
        wb = load_workbook(output_path)
        ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
        ws.title = "Sheet1"
    else:
        wb = _create_workbook()
        ws = wb["Sheet1"]

    for c, header in enumerate(HEADERS, start=1):
        if ws.cell(3, c).value is None:
            ws.cell(3, c, header)
        elif str(ws.cell(3, c).value).strip() != header:
            ws.cell(3, c, header)

    for r in range(4, ws.max_row + 1):
        for c in range(1, len(HEADERS) + 1):
            ws.cell(r, c).value = None

    return wb, ws


def write_spreadsheet(metadata, lines, employees, batch_dir):
    batch_dir.mkdir(parents=True, exist_ok=True)
    output_path = batch_dir / "Spreadsheet-v4-input.xlsx"
    wb, ws = prepare_output_workbook(output_path)

    # Labadi RULE 6 (2026-06-09): VAT placement fix.
    #   Column E (*Invoice Amount) = TOTAL invoice amount INCLUDING VAT — the
    #     grand total of the whole invoice (sum of every line incl VAT), repeated
    #     on every row as the Oracle invoice-header amount.
    #   Column M (*Amount) = per-line amount EXCLUDING VAT — taken from the source
    #     taxable (pre-VAT) column (col 34); falls back to incl-VAT / 1.15.
    invoice_total_incl_vat = round(sum(_safe_float(l.get("inv_amt_incl_vat")) for l in lines), 2)

    template_row = 4
    for idx, line in enumerate(lines, start=1):
        row_num = idx + 3
        cls = classify_line(line, employees)
        cost_center = cls.get("cost_center")
        dist_combo = f"{cls['gl_account']}.{int(cost_center)}" if cost_center else cls["gl_account"]
        tax_code = "KSA VAT STANDARD" if line.get("vat_pct") == 15 else "KSA VAT ZERO"
        desc = f"{line['passenger_name']} - {line['route']} ({line['ticket_no']})"

        # RULE 6 Column M: line item EXCLUDING VAT.
        line_ex_vat = _safe_float(line.get("taxable_amt"))
        if not line_ex_vat:
            # No ex-VAT source value — derive from incl-VAT line total.
            line_ex_vat = round(_safe_float(line.get("inv_amt_incl_vat")) / 1.15, 2)
        else:
            line_ex_vat = round(line_ex_vat, 2)

        values = [
            idx,
            BUSINESS_UNIT,
            metadata["invoice_no"],
            CURRENCY,
            invoice_total_incl_vat,   # RULE 6: col E = whole-invoice total incl VAT
            metadata["invoice_date"],
            metadata.get("supplier_name") or SUPPLIER_NAME,
            SUPPLIER_NUMBER,
            SUPPLIER_SITE,
            INVOICE_TYPE,
            desc,
            "Item",
            line_ex_vat,              # RULE 6: col M = line item ex VAT
            dist_combo,
            tax_code,
            cls.get("emp_no") or "",
            line.get("ref_no") or "",
        ]

        for c, value in enumerate(values, start=1):
            cell = ws.cell(row_num, c, value)
            if row_num != template_row:
                _copy_cell_style(ws.cell(template_row, c), cell)
            if not cell.fill or cell.fill.fill_type is None:
                cell.fill = GREEN_FILL
            if not cell.border or cell.border.left.style is None:
                cell.border = BORDER_THIN

    last_data_row = len(lines) + 3
    if ws.max_row > last_data_row:
        ws.delete_rows(last_data_row + 1, ws.max_row - last_data_row)

    ws.freeze_panes = "A4"
    wb.save(output_path)
    return output_path


def parse_args():
    parser = argparse.ArgumentParser(description="Convert Jawal TAX INVOICE xlsx to Spreadsheet-v4-input.xlsx")
    parser.add_argument("--invoice-file", required=True, type=Path)
    parser.add_argument("--batch-dir", required=True, type=Path)
    parser.add_argument("--master-data", required=True, type=Path)
    return parser.parse_args()


def main():
    args = parse_args()
    if not args.invoice_file.exists():
        raise FileNotFoundError(f"Invoice file not found: {args.invoice_file}")
    if not args.master_data.exists():
        raise FileNotFoundError(f"Master data not found: {args.master_data}")

    metadata, lines = extract_invoice_lines(args.invoice_file)
    if not metadata["invoice_no"]:
        raise ValueError("Could not read invoice number from uploaded invoice")
    if not metadata["invoice_date"]:
        raise ValueError("Could not read invoice date from uploaded invoice")
    if not lines:
        raise ValueError("No invoice line items found at row 28+")

    employees = load_manpower(args.master_data)
    output_path = write_spreadsheet(metadata, lines, employees, args.batch_dir)
    matched = sum(
        1
        for line in lines
        if (emp_no := extract_emp_no_from_ref(line.get("ref_no"))) is not None and emp_no in employees
    )
    unmatched = len(lines) - matched
    print(f"Converted invoice {metadata['invoice_no']} dated {metadata['invoice_date']}")
    print(f"Loaded {len(employees)} employees from {args.master_data}")
    print(f"Rows total: {len(lines)}")
    print(f"Matched by emp_no: {matched}")
    print(f"Unmatched (sponsorship/event): {unmatched}")
    print(f"Wrote {len(lines)} rows to {output_path}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(1)
