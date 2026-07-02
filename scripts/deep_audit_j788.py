#!/usr/bin/env python3
"""
Deep AP Audit Script - Jawwal J26-788 Batch
Programmatically loads J26-788 Spreadsheet-FILLED-v30_11.xlsx,
Aljeel_Lookups-v2.xlsx (Manpower), and the AI Fraud Detector output,
and performs a rigorous, multi-layered audit of all 103 rows.
"""
import openpyxl
import json
import re
from pathlib import Path
from collections import Counter

# Path definitions
WORKSPACE = Path("/home/clawdbot/.openclaw/workspace")
BATCH_DIR = WORKSPACE / "aljeel/batches/jawal-J26-788"
SPREADSHEET_PATH = BATCH_DIR / "output/Spreadsheet-J26-788-FILLED-v30_11.xlsx"
LOOKUPS_PATH = WORKSPACE / "aljeel/qc/master-data/Aljeel_Lookups-v2.xlsx"
FRAUD_AI_PATH = WORKSPACE / "aljeel/qc/ai-poc/j26-788-fraud-ai-v162.json"
REPORTS_DIR = WORKSPACE / "aljeel/qc/reports"

# Transliteration Map from employee_resolver_v2
TRANSLITERATION_MAP = {
    "MOHAMMAD": "MOHAMMED", "MOHAMED": "MOHAMMED", "MUHAMMED": "MOHAMMED",
    "MUHAMMAD": "MOHAMMED", "MOHAMAD": "MOHAMMED", "MHAMMED": "MOHAMMED",
    "BELAL": "BILAL", "KHALID": "KHALED", "HUSAYN": "HUSSAIN", 
    "HUSSEIN": "HUSSAIN", "HUSAIN": "HUSSAIN", "ABDULRAHMAN": "ABDELRAHMAN", 
    "ABDULLTAIF": "ABDELLATIF", "ABDEL": "ABDUL", "ABDAL": "ABDUL",
    "OUSAMA": "OSAMA", "OUSSAMA": "OSAMA", "USAMA": "OSAMA",
    "HESHAM": "HISHAM", "HOSNY": "HUSNI", "HOSNI": "HUSNI",
    "HAMMAM": "HAMAM", "ALAAELDIN": "ALAEDDIN", "ALAAEDDIN": "ALAEDDIN",
    "NJOUD": "NUJUD", "NUJOOD": "NUJUD", "WASEEM": "WASIM",
    "ISSAM": "ISAM", "ESSAM": "ISAM", "ADOLAY": "ADOLAI",
    "HOSSAM": "HUSAM", "HOSAM": "HUSAM", "HAMZAH": "HAMZEH", 
    "HAMZA": "HAMZEH", "ABDALLAH": "ABDULLAH", "AHMED": "AHMAD",
    "JUMAH": "JUMA", "SOBHI": "SUBHI", "ALANAZI": "ALENAZY",
    "ALNAAZI": "ALENAZY", "ALNAZI": "ALENAZY",
}

GDS_TITLES = {
    "MR", "MRS", "MS", "DR", "ENG", "MISS", "MASTER", "INF", "CHD",
    "MR(CHD)", "MS(CHD)", "MRS(CHD)", "MSTR",
}

def _normalize_name_tokens(name_str):
    if not name_str:
        return set()
    cleaned = str(name_str).upper()
    cleaned = re.sub(r"\((?:CHD|INF)\)", "", cleaned)
    for title in GDS_TITLES:
        cleaned = re.sub(r"\b" + re.escape(title) + r"\b", "", cleaned)
    # Split on non-alphabetic characters
    raw_tokens = re.split(r"[^A-Z]+", cleaned)
    tokens = set()
    for t in raw_tokens:
        if len(t) >= 2:
            # Strip article prefixes
            for art in ["EL", "AL", "UL"]:
                if t.startswith(art) and len(t) > len(art) + 1:
                    t = t[len(art):]
            # Apply transliteration mapping
            mapped = TRANSLITERATION_MAP.get(t, t)
            tokens.add(mapped)
    return tokens

def check_two_token_match(pax_name, master_name):
    pax_tokens = _normalize_name_tokens(pax_name)
    master_tokens = _normalize_name_tokens(master_name)
    overlap = pax_tokens & master_tokens
    return len(overlap) >= 2

def load_master_data():
    wb = openpyxl.load_workbook(LOOKUPS_PATH, read_only=True)
    sheet = wb["Manpower"]
    master_data = {}
    for r, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        if not row[0]:
            continue
        emp_no = int(float(str(row[0])))
        master_data[emp_no] = {
            "emp_no": emp_no,
            "old_emp_no": row[1],
            "name": row[2],
            "arabic_name": row[3],
            "location": row[4],
            "manager_no": row[5],
            "line_manager": row[6],
            "div_code": str(row[8]).zfill(3) if row[8] else None,
            "div_name": row[9],
            "agency_code": str(row[10]).zfill(5) if row[10] else None,
            "agency_name": row[11],
            "cost_center": str(row[12]).zfill(6) if row[12] else None,
            "cost_center_name": row[13],
            "status": row[14] if len(row) > 14 else "Unknown",
        }
    wb.close()
    return master_data

def load_fraud_ai_data():
    if not FRAUD_AI_PATH.exists():
        return {}
    with open(FRAUD_AI_PATH) as f:
        data = json.load(f)
    return {x["sl_no"]: x for x in data["ai_result"]["per_row_verdicts"]}

def extract_pax_from_description(desc):
    if not desc:
        return ""
    # Extract before the hyphen or parenthetical ticket
    parts = desc.split("-")
    if len(parts) > 1:
        pax_part = parts[0].strip()
    else:
        pax_part = desc.split("(")[0].strip()
    return pax_part

def run_audit():
    print("[Deep AP Audit] Loading master lookup...", flush=True)
    master_data = load_master_data()
    
    print("[Deep AP Audit] Loading AI Fraud Detector...", flush=True)
    fraud_data = load_fraud_ai_data()
    
    print("[Deep AP Audit] Loading J26-788 Spreadsheet...", flush=True)
    wb = openpyxl.load_workbook(SPREADSHEET_PATH, read_only=True)
    sheet = wb.active
    
    total_entries = 0
    challenges = []
    verified_count = 0
    
    # Iterate over data rows: starts from row 4 (Excel row 4, which is list index 3)
    # The header row is row 3 (list index 2)
    rows = list(sheet.iter_rows(values_only=True))
    data_rows = rows[3:106] # row 4 to 106
    
    print(f"[Deep AP Audit] Beginning analysis of {len(data_rows)} rows...", flush=True)
    
    for idx, row in enumerate(data_rows):
        total_entries += 1
        r_num = idx + 4  # Excel row number
        sl_no = idx + 1  # Serial number for AI Fraud lookup (1 to 100) and refunds
        
        desc = row[10]
        amount = row[12]
        combo = row[13]
        emp_no_val = row[15]
        acct = row[18]
        cc = row[20]
        div = row[22]
        agency = row[26]
        row_status = row[32]
        manpower_status = row[33]
        evidence_status = row[34]
        email_status = row[35]
        self_approval_status = row[36]
        human_note = row[37]
        qc_catches = row[38]
        agent_flags = row[39]
        
        pax_name = extract_pax_from_description(desc)
        
        row_challenges = []
        
        # 1. Match Passenger Name vs Employee No
        emp_no = None
        if emp_no_val is not None:
            try:
                emp_no = int(float(str(emp_no_val)))
            except (ValueError, TypeError):
                pass
        
        emp_master = None
        if emp_no:
            emp_master = master_data.get(emp_no)
            if emp_master:
                # Check name match
                matches = check_two_token_match(pax_name, emp_master["name"])
                if not matches:
                    row_challenges.append({
                        "type": "PASSENGER_NAME_MISMATCH",
                        "severity": "HIGH",
                        "detail": f"Passenger '{pax_name}' matched to Employee No {emp_no} ({emp_master['name']}) but less than 2 name tokens overlap. This is a potential mis-assignment.",
                    })
            else:
                row_challenges.append({
                    "type": "EMPLOYEE_NOT_IN_MASTER",
                    "severity": "HIGH",
                    "detail": f"Employee No {emp_no} assigned to row but was not found in the master manpower lookup sheet.",
                })
        else:
            # Check if this should have an Employee No
            # Non-employee sponsorships (60307021), training (60308009), or dependent (21070229) are allowed to be blank
            if acct in ["60301003", "60301004"]:
                row_challenges.append({
                    "type": "EMPLOYEE_NO_MISSING",
                    "severity": "HIGH",
                    "detail": f"Travel Tickets Expense row has no Employee No assigned, despite using employee account '{acct}'.",
                })
        
        # 2. Account-Division Alignment
        if emp_master:
            emp_div = emp_master["div_code"]
            emp_cc = emp_master["cost_center"]
            emp_agency = emp_master["agency_code"]
            
            # G&A Employee Check
            is_ga_employee = (emp_div == "888" or emp_master["div_name"] == "G&A")
            if is_ga_employee:
                if acct != "60301004":
                    row_challenges.append({
                        "type": "ACCOUNT_DIVISION_MISMATCH",
                        "severity": "HIGH",
                        "detail": f"G&A Employee {emp_no} ({emp_master['name']}) is charged to Business Account '{acct}' instead of G&A Employee Travel Account '60301004'.",
                    })
            else:
                if acct == "60301004":
                    row_challenges.append({
                        "type": "ACCOUNT_DIVISION_MISMATCH",
                        "severity": "HIGH",
                        "detail": f"Business/Sales Employee {emp_no} ({emp_master['name']}) is charged to G&A Travel Account '60301004' instead of Business Travel Account '60301003'.",
                    })
            
            # Segment Alignment Check (Manpower consistency)
            # Check if CC/DIV/Agency in spreadsheet aligns with employee's home CC/DIV/Agency in master
            # (Only applies to non-sponsorship/non-personal contribution, as they route differently)
            if acct in ["60301003", "60301004"]:
                if cc != emp_cc:
                    row_challenges.append({
                        "type": "COST_CENTER_MISMATCH",
                        "severity": "MEDIUM",
                        "detail": f"Row has CC '{cc}' but employee home CC is '{emp_cc}' ({emp_master['cost_center_name']}).",
                    })
                if div != emp_div:
                    row_challenges.append({
                        "type": "DIVISION_MISMATCH",
                        "severity": "MEDIUM",
                        "detail": f"Row has Division Code '{div}' but employee division code is '{emp_div}' ({emp_master['div_name']}).",
                    })
                if agency != emp_agency:
                    row_challenges.append({
                        "type": "AGENCY_MISMATCH",
                        "severity": "MEDIUM",
                        "detail": f"Row has Agency Code '{agency}' but employee agency code is '{emp_agency}' ({emp_master['agency_name']}).",
                    })
        
        # 3. Manpower Need Allocate Check
        if emp_master and emp_master["status"] == "Need to allocate":
            row_challenges.append({
                "type": "MANPOWER_NEED_ALLOCATE",
                "severity": "MEDIUM",
                "detail": f"Employee {emp_no} ({emp_master['name']}) is flagged as 'Need to allocate' in Master Manpower lookup. Home segments may be placeholder/outdated.",
            })

        # 4. Self-Approval Check
        if self_approval_status == "SELF_APPROVED":
            row_challenges.append({
                "type": "SELF_APPROVAL",
                "severity": "CRITICAL",
                "detail": f"Self-approval detected in approval emails. The passenger requested and self-approved their own flight or hotel booking.",
            })
            
        # 5. Missing Evidence / Corrupt Emails Check
        if evidence_status == "MISSING":
            # EMDs don't require evidence folders
            is_emd = "1936" in str(desc) or "EMD" in str(desc).upper()
            if not is_emd:
                row_challenges.append({
                    "type": "EVIDENCE_FOLDER_MISSING",
                    "severity": "HIGH",
                    "detail": f"No matching evidence folder found in gdrive-evidence for ticket.",
                })
        
        if email_status == "MISSING":
            is_emd = "1936" in str(desc) or "EMD" in str(desc).upper()
            if not is_emd:
                row_challenges.append({
                    "type": "APPROVAL_EMAIL_MISSING",
                    "severity": "HIGH",
                    "detail": f"Evidence folder exists, but no approval email (.msg) file is present.",
                })
        elif email_status == "CORRUPT_OR_EMPTY":
            row_challenges.append({
                "type": "APPROVAL_EMAIL_CORRUPT",
                "severity": "CRITICAL",
                "detail": f"Approval email exists but is corrupt or has empty content (body length < 10 characters).",
            })
            
        # 6. Specific Case Overrides (e.g. ALBASIRI)
        if "ALBASIRI" in str(desc).upper() and "6905428831" in str(desc):
            # Check if CC was corrected to EP/Contribution/Abbott
            if cc != "160014" or div != "170" or agency != "10072":
                row_challenges.append({
                    "type": "SPONSORSHIP_CC_MISMATCH",
                    "severity": "CRITICAL",
                    "detail": f"Saleh Albasiri (sponsorship row 8) is mapped to home/requestor CC Ansell ('160013'/'192'/'10060') instead of actual event CC Abbott/EP ('160014'/'170'/'10072') as documented in IEPC AF EP-2026-16.",
                })

        # 7. AI Fraud Detector Cross-Check
        ai_row = fraud_data.get(sl_no)
        if ai_row and ai_row.get("verdict") == "RED":
            # Check if this AI anomaly is already captured, else add it
            ai_category = ai_row.get("primary_category", "UNKNOWN")
            ai_reasoning = ai_row.get("reasoning", "")
            
            # Translate AI category to challenge type
            row_challenges.append({
                "type": f"AI_FRAUD_{ai_category}",
                "severity": "CRITICAL" if ai_category in ["DUPLICATE_BILLING", "REBOOKING_FRAUD", "APPROVAL_INVALID"] else "HIGH",
                "detail": f"AI Fraud Detector flagged as RED ({ai_category}): {ai_reasoning}",
                "ai_evidence": ai_row.get("evidence", [])
            })
            
        # Compile challenges
        if row_challenges:
            challenges.append({
                "excel_row": r_num,
                "sl_no": sl_no,
                "passenger": pax_name,
                "description": desc,
                "amount": amount,
                "acct": acct,
                "emp_no": emp_no,
                "row_status": row_status,
                "qc_catches": qc_catches,
                "findings": row_challenges
            })
        else:
            verified_count += 1
            
    wb.close()
    
    print(f"[Deep AP Audit] Finished. Found {len(challenges)} challenged rows and {verified_count} verified rows.", flush=True)
    return challenges, verified_count, total_entries

if __name__ == "__main__":
    challenges, verified, total = run_audit()
    
    # Save a temporary json report for analysis
    out_json = REPORTS_DIR / "J26-788-audit-findings.json"
    with open(out_json, "w") as f:
        json.dump({
            "total_rows": total,
            "verified_rows": verified,
            "challenged_rows": len(challenges),
            "challenges": challenges
        }, f, indent=2)
    print(f"[Deep AP Audit] Saved temporary JSON to {out_json}", flush=True)
