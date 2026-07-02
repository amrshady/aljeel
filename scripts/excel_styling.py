#!/usr/bin/env python3
"""
Excel Styling — Row-level color coding (GREEN / YELLOW / RED).

Classifies each row based on:
  - Hard gate failures → RED
  - Review-worthy soft flags → YELLOW
  - Low resolution confidence → YELLOW
  - Informational-only flags → GREEN

Also adds a "Row Status" column and a header legend row.
"""
from __future__ import annotations

from openpyxl.styles import PatternFill, Font

# ---------------------------------------------------------------------------
# Fill + Font definitions
# ---------------------------------------------------------------------------
ROW_GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE")
ROW_GREEN_FONT  = Font(color="006100")

ROW_YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
ROW_YELLOW_FONT = Font(color="9C5700")

ROW_RED_FILL    = PatternFill("solid", fgColor="FFC7CE")
ROW_RED_FONT    = Font(color="9C0006")


# ---------------------------------------------------------------------------
# Flag classification
# ---------------------------------------------------------------------------

# Review-worthy flags — forces YELLOW
REVIEW_WORTHY_FLAGS = {
    # Gate-level flags
    "FORM_EMP_NO_MISMATCH",
    "EMPLOYEE_NOT_IN_MASTER",
    "NEW_EMPLOYEE_NOT_IN_MASTER",
    "ALLOCATION_TARGET_MISSING",
    "MULTI_ALLOCATION_PENDING_REVIEW",
    "MANAGER_NOT_REALLOCATED",
    "EMPLOYEE_AS_SPONSORED",
    "EMPLOYEE_AS_SPONSORED_HARD",
    "SOLUTION_CODE_PENDING",
    "TRIP_PURPOSE_UNKNOWN",
    "MIXED_FAMILY_CLUSTER",
    "FORM_APPROVER_NOT_LINE_MANAGER",
    "ACCOUNT_DEFAULT_FALLBACK",
    "MEDSOURCE_ROUTE",
    "ERBE_EXCEPTION",
    "MAYBE_60301004_GA",
    "ALLOCATION_LOOP_DETECTED",
    "ALLOCATION_MISSING_FROM_EMAIL",
    "LLM_RETURNED_UNKNOWN_CODE",
    "PERSONAL_LOW_CONFIDENCE",
    "TRIP_PURPOSE_MISMATCH",
    # QC catches (within-batch + cross-batch) — added by qc modules
    "DUP_ROUTE_STRICT",
    "VAT_MISMATCH",
    "EMD_MISMATCH",
    "OVER_LIMIT",
    "ROUND_AMOUNT",
    "CROSS_BATCH_DUPLICATE_TICKET",
    "POTENTIAL_REBOOKING_FRAUD",
    "FREQUENT_TRAVELER_OVER_BUDGET",
    "PASSENGER_AMOUNT_PATTERN",
    "UNUSUAL_BOOKING_VELOCITY",
    # v15.3 location flags
    "LOCATION_INTERNATIONAL",
    "LOCATION_UNMAPPABLE",
    # v15.4 sponsorship origin flags
    "LOCATION_SPONSOR_ORIGIN_INTERNATIONAL",
    # v15.5 venue-text LLM flags
    "LOCATION_VENUE_INTERNATIONAL",
}

# Informational flags — GREEN-OK (don't force yellow)
INFORMATIONAL_FLAGS = {
    "FORM_AGREES_WITH_MANPOWER",
    "FORM_NOT_FOUND_IN_EMAIL",
    "FORM_TRIP_VALUE_DIFFERS",
    "FORM_FUSION_CODES_LOGGED",
    "MANPOWER_DIV_NOT_IN_MASTER",
    "ALLOCATION_RESOLVED_HIERARCHY",
    "ALLOCATION_RESOLVED_DETERMINISTIC",
    "ALLOCATION_RESOLVED_LLM",
    "OPEX_PDF_PARSED",
    "RESOLVED_VIA_LLM_EMAIL",
    "FAMILY_CLUSTER_DETECTED",
    "FAMILY_CLUSTER_CHD_CONFIRMED",
    "FAMILY_CLUSTER_NO_CHD",
    "ACCOUNT_OVERRIDE_APPLIED",
    "SPONSORSHIP_DETECTED",
    # v15.3 location resolution informational
    "LOCATION_FROM_IATA_MAP",
    "LOCATION_FROM_LLM_CITY_CLASS",
    "LOCATION_FALLBACK_TO_EMPLOYEE_HOME",
    # v15.5 venue-text LLM informational
    "LOCATION_FROM_VENUE_LLM",
    # Any RESOLVED_VIA_* flag
}


def _is_informational(flag: str) -> bool:
    """Check if a flag is informational (green-OK)."""
    if flag in INFORMATIONAL_FLAGS:
        return True
    if flag.startswith("RESOLVED_VIA_"):
        return True
    return False


def classify_row(gate_result, resolved_line, qc_catches: list[dict] | None = None) -> str:
    """
    Determine row state: GREEN, YELLOW, or RED.
    
    Args:
        gate_result: GateResult from qc_gates
        resolved_line: ResolvedLine with flags and attributes
        qc_catches: Optional list of QC catch dicts for this line (from within-batch + cross-batch)
    
    Returns:
        "GREEN", "YELLOW", or "RED"
    """
    # 1. RED CONDITIONS:
    # - Hard gate failure
    if gate_result.hard_failures:
        return "RED"
    
    # - QC catches with HARD severity
    if qc_catches:
        for catch in qc_catches:
            if catch.get("severity") == "HARD":
                return "RED"
    
    # - Employee is marked as "Need to allocate"
    sol_flag = getattr(resolved_line, 'sol_flag', '') or ''
    if sol_flag == "Need to allocate":
        return "RED"
    
    # Collect all flags from gate result
    all_flag_codes = set()
    for _, code in gate_result.soft_flags:
        all_flag_codes.add(code)
    
    # Also check resolved line flags directly
    for flag in resolved_line.flags:
        all_flag_codes.add(flag)
    
    # Add QC catch categories as flags
    if qc_catches:
        for catch in qc_catches:
            all_flag_codes.add(catch["category"])
            
    if "ALLOCATION_TARGET_MISSING" in all_flag_codes or "EMAIL_CORRUPT_OR_EMPTY" in all_flag_codes:
        return "RED"
    
    # Operator rule (2026-06-20): a missing/unclear approval is a RED hold, not a soft YELLOW.
    # The evidence folder exists (so GL is NOT blanked — that gate is separate and folder-based),
    # but without a clear approval the row must be escalated RED with column marking + Human
    # Review Note (both already produced by the NO_APPROVAL catch).
    if "NO_APPROVAL" in all_flag_codes:
        return "RED"
    
    # 2. YELLOW CONDITIONS:
    # - Review-worthy soft flags (excluding ALLOCATION_TARGET_MISSING)
    for flag in all_flag_codes:
        if flag in REVIEW_WORTHY_FLAGS and flag != "ALLOCATION_TARGET_MISSING":
            return "YELLOW"
    
    # - Match method uses LLM (starts with "v2_"), email allocation subordinate, or not_found
    match_method = getattr(resolved_line, 'emp_match_method', '') or ''
    if match_method.startswith("v2_") or match_method.startswith("allocation_email_subordinate") or match_method == "not_found":
        return "YELLOW"
    
    # - Check resolution confidence
    confidence = getattr(resolved_line, '_v2_confidence', 1.0)
    if confidence is not None:
        try:
            conf_val = float(confidence)
            if conf_val < 0.9:
                return "YELLOW"
        except (ValueError, TypeError):
            pass
    
    # 3. GREEN CONDITIONS:
    # Otherwise, clean deterministic script matching (emp_no_direct, name_fuzzy, msg_filename, family_cluster_unified)
    return "GREEN"


def get_row_style(row_status: str):
    """Return (fill, font) tuple for the given row status."""
    if row_status == "RED":
        return ROW_RED_FILL, ROW_RED_FONT
    elif row_status == "YELLOW":
        return ROW_YELLOW_FILL, ROW_YELLOW_FONT
    else:
        return ROW_GREEN_FILL, ROW_GREEN_FONT
