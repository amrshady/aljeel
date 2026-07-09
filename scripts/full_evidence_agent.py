#!/usr/bin/env python3
"""
Full Evidence LLM Agent — J26-593 BLIND TEST

Per-row workflow:
  1) Find ticket folder (by ticket# OR passenger name OR notes match)
  2) Collect ALL .msg bodies + ALL PDF text in that folder (NO TRUNCATION)
  3) Send everything to Gemini 2.5 Pro with strict JSON schema
  4) Cache by (ticket_no, row_idx) → re-runs are free
  5) Validate output against master data
  6) Concurrency: 5 parallel rows

Output: Spreadsheet-J26-593-LLMAGENT-v1.xlsx + comparison + cost summary
"""
from __future__ import annotations

import argparse
import concurrent.futures
import hashlib
import json
import os
import re
import sys
import time
import traceback
import urllib.request
from pathlib import Path
from typing import Any

import openpyxl

# Set up paths
ROOT = Path("/home/clawdbot/.openclaw/workspace/aljeel")
sys.path.insert(0, str(ROOT / "scripts"))

from msg_parser import parse_msg  # type: ignore

import fitz  # pymupdf

GEMINI_BASE_URL = os.environ.get("GEMINI_BASE_URL", "https://gateway.ai.cloudflare.com/v1/3724a3e71944b366a39b3735aa117a58/accord-aljeel-ap/google-ai-studio/v1beta")

TRUTH_XLSX = Path("/mnt/aljeel-ap_kb/current/J26-593/J26-593.xlsx")
BLIND_INPUT = Path("/home/clawdbot/.openclaw/workspace/aljeel/batches/jawal-J26-593-BLIND/Spreadsheet.xlsx")
RAW_ROOT = Path("/home/clawdbot/.openclaw/workspace/aljeel/batches/jawal-J26-593-BLIND/raw")
MASTER_XLSX = Path("/home/clawdbot/.openclaw/workspace/aljeel/qc/master-data/Aljeel_Lookups-v2.xlsx")
EMAIL_MASTER_XLSX = Path("/home/clawdbot/.openclaw/workspace/aljeel/qc/master-data/Employee_Emails_2026-05-26.xlsx")

CACHE_DIR = Path("/home/clawdbot/.openclaw/workspace/aljeel/extracted/full-evidence-agent-cache")
CACHE_DIR.mkdir(parents=True, exist_ok=True)

OUT_DIR = Path("/home/clawdbot/.openclaw/workspace/aljeel/batches/jawal-J26-593-BLIND/output/llmagent")
OUT_DIR.mkdir(parents=True, exist_ok=True)

# Canonical AP pipeline LLM cascade: Gemini 3 Pro (via gemini-pro-latest alias) -> 2.5 Pro -> 2.5 Flash
# Mirrors email_allocation_extractor.py / location_resolver.py / allocation_resolver.py.
# Comment in email_allocation_extractor.py v15.1: "gemini-3-pro 404; gemini-pro-latest resolves to best available"
GEMINI_MODEL_CASCADE = ["gemini-pro-latest", "gemini-2.5-pro", "gemini-2.5-flash"]
GEMINI_PRICING = {
    # VERIFIED Jun 20 2026 (OpenRouter/Verdent/MetaCTO). gemini-pro-latest -> gemini-3.1-pro-preview.
    # Thinking (thoughtsTokenCount) billed at OUTPUT rate. >200K context: all tokens at long-context rate.
    "gemini-pro-latest":      {"in": 2.00/1e6, "out": 12.00/1e6},
    "gemini-3.1-pro-preview": {"in": 2.00/1e6, "out": 12.00/1e6},
    "gemini-2.5-pro":         {"in": 1.25/1e6, "out": 10.00/1e6},
    "gemini-2.5-flash":       {"in": 0.30/1e6, "out": 2.50/1e6},
}

# Long-context (>200K input) doubles Gemini 3.x Pro rates; ALL tokens billed at long-context rate.
GEMINI_PRICING_LONG = {
    "gemini-pro-latest":      {"in": 4.00/1e6, "out": 18.00/1e6},
    "gemini-3.1-pro-preview": {"in": 4.00/1e6, "out": 18.00/1e6},
    "gemini-2.5-pro":         {"in": 2.50/1e6, "out": 15.00/1e6},
    "gemini-2.5-flash":       {"in": 0.30/1e6, "out": 2.50/1e6},
}
def _gemini_rate(model, in_tokens=0):
    if in_tokens and in_tokens > 200_000:
        return GEMINI_PRICING_LONG.get(model) or GEMINI_PRICING_LONG.get("gemini-2.5-pro")
    return GEMINI_PRICING.get(model) or GEMINI_PRICING.get("gemini-2.5-pro")
def _cf_custom_cost_header(model, in_tokens=0):
    import json as __j
    _p = _gemini_rate(model, in_tokens)
    return __j.dumps({"per_token_in": _p["in"], "per_token_out": _p["out"]})



def load_env_key(name: str) -> str:
    val = os.environ.get(name, "")
    if val:
        return val
    env_path = Path("/home/clawdbot/.openclaw/.env")
    if env_path.exists():
        for line in env_path.read_text().splitlines():
            line = line.strip()
            if line.startswith(name + "="):
                return line.split("=", 1)[1].strip().strip("'\"")
    return ""


GEMINI_API_KEY = load_env_key("GEMINI_API_KEY")
assert GEMINI_API_KEY, "GEMINI_API_KEY not found"


# ---------- master data loaders ----------
def load_manpower() -> dict[str, dict]:
    """Returns {emp_no: {name, arabic, location, manager, line_manager, division_code, division_name, agency_code, agency_name, cost_center, cc_name, solution}}."""
    wb = openpyxl.load_workbook(MASTER_XLSX, read_only=True, data_only=True)
    ws = wb["Manpower"]
    rows = list(ws.iter_rows(values_only=True))
    # Header is row index 6 (0-based)
    hdr = rows[0]
    # Columns: Emp No(0), Old Emp No(1), Name(2), Arabic Name(3), Location(4), Manager No(5), Line Manager(6),
    #  blank(7), Code(8)=DIV code, New Division(9), Code(10)=Agency, New agency(11), New cost center(12),
    #  New cost center name(13), blank(14), Solution(15)
    emps = {}
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        emp_no = str(row[0]).strip()
        if not emp_no or emp_no.lower() == "none":
            continue
        emps[emp_no] = {
            "emp_no": emp_no,
            "name": str(row[2] or "").strip(),
            "arabic": str(row[3] or "").strip(),
            "location": str(row[4] or "").strip(),
            "manager_no": str(row[5] or "").strip(),
            "line_manager": str(row[6] or "").strip(),
            "div_code": str(row[8] or "").strip(),
            "div_name": str(row[9] or "").strip(),
            "agency_code": str(row[10] or "").strip(),
            "agency_name": str(row[11] or "").strip(),
            "cost_center": str(row[12] or "").strip(),
            "cc_name": str(row[13] or "").strip(),
            "solution": str(row[15] or "").strip(),
        }
    wb.close()
    return emps


def load_email_master() -> dict[str, str]:
    """Returns {emp_no: email}."""
    wb = openpyxl.load_workbook(EMAIL_MASTER_XLSX, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    out = {}
    # detect header row
    hdr_idx = None
    for i, row in enumerate(rows[:5]):
        if row and any(v and "emp" in str(v).lower() for v in row):
            hdr_idx = i
            break
    if hdr_idx is None:
        hdr_idx = 0
    hdr = rows[hdr_idx]
    emp_col = None
    email_col = None
    for j, v in enumerate(hdr):
        if not v:
            continue
        s = str(v).lower()
        if "emp" in s and emp_col is None:
            emp_col = j
        if "email" in s or "mail" in s:
            email_col = j
    if emp_col is None or email_col is None:
        wb.close()
        return out
    for row in rows[hdr_idx + 1:]:
        if not row or not row[emp_col]:
            continue
        out[str(row[emp_col]).strip()] = str(row[email_col] or "").strip()
    wb.close()
    return out


def load_lookups() -> dict[str, dict[str, str]]:
    """Returns {'cc': {code: name}, 'div': {...}, 'solution': {...}, 'agency': {...}} rewired to Aljeel_Lookups-v2.xlsx."""
    from code_name_lookup import CodeNameLookup
    lk = CodeNameLookup()
    lk._ensure_loaded()
    return {
        "cc": lk._cc_to_name,
        "div": lk._div_to_name,
        "solution": lk._solution_to_name,
        "agency": lk._agency_to_name,
    }


# ---------- evidence collection ----------
def extract_pdf_text(path: Path) -> str:
    try:
        doc = fitz.open(str(path))
        parts = []
        for page in doc:
            parts.append(page.get_text())
        doc.close()
        return "\n".join(parts)
    except Exception as e:
        return f"[PDF_ERROR: {e}]"


_DATE_LAYER_RE = re.compile(r"^\d{2}[a-z]{3}$", re.IGNORECASE)
_EVIDENCE_SUFFIXES = {".msg", ".pdf", ".eml"}
_MERGED_EVIDENCE_FOLDERS: dict[str, list[Path]] = {}


def _has_direct_evidence_files(folder: Path) -> bool:
    try:
        return any(
            child.is_file() and child.suffix.lower() in _EVIDENCE_SUFFIXES
            for child in folder.iterdir()
        )
    except OSError:
        return False


def _single_evidence_child(folder: Path) -> Path | None:
    """Return the one descriptive child that holds evidence files, if present."""
    if _has_direct_evidence_files(folder):
        return None
    try:
        children = [child for child in folder.iterdir() if child.is_dir()]
    except OSError:
        return None
    if len(children) != 1:
        return None
    child = children[0]
    return child if _has_direct_evidence_files(child) else None


def _logical_ref_key(folder: Path) -> str:
    return folder.name.strip().casefold()


def iter_evidence_files(folder: Path):
    """Yield evidence files for a logical ref folder, including merged duplicates."""
    siblings = _MERGED_EVIDENCE_FOLDERS.get(str(folder.resolve(strict=False)), [folder])
    seen: set[tuple[str, int]] = set()
    for sibling in siblings:
        source = _single_evidence_child(sibling) or sibling
        try:
            children = sorted(source.iterdir())
        except OSError:
            continue
        for child in children:
            try:
                dedupe_key = (child.name.casefold(), child.stat().st_size)
            except OSError:
                dedupe_key = (child.name.casefold(), -1)
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)
            yield child


def _candidate_ref_folders(raw_root: Path) -> list[Path]:
    folders: list[Path] = []
    try:
        children = sorted(raw_root.iterdir())
    except OSError:
        return folders
    for child in children:
        if not child.is_dir():
            continue
        if _DATE_LAYER_RE.match(child.name):
            try:
                folders.extend(sub for sub in sorted(child.iterdir()) if sub.is_dir())
            except OSError:
                pass
        else:
            folders.append(child)
    return folders


def find_ticket_folder(ticket_no: str, passenger: str, notes: str, all_folders: list[Path]) -> Path | None:
    """Try multiple strategies to locate the evidence folder."""
    # 1) Exact ticket# match
    for f in all_folders:
        if f.name.strip() == ticket_no:
            return f
    # 2) Ticket# substring (handles "6905341979-80" ranges, "6905084618 change", etc.)
    for f in all_folders:
        # Match the leading 10-digit number with optional suffix
        m = re.match(r"^(\d{10,})", f.name.strip())
        if m and m.group(1).startswith(ticket_no):
            return f
        # Range folder: "6905341979-80"
        m2 = re.match(r"^(\d{10})-(\d{1,3})", f.name.strip())
        if m2:
            base = m2.group(1)
            end = m2.group(2)
            if ticket_no == base:
                return f
            # Check the range end (e.g., 6905341979-80 → 6905341980)
            if len(end) < 10:
                prefix = base[: 10 - len(end)]
                candidate = prefix + end
                if ticket_no == candidate:
                    return f
        # Match anywhere in folder name
        if ticket_no in f.name:
            return f
    # 3) Passenger name match REMOVED to prevent cross-contamination
    # 4) Notes match (sponsorship case): HF-2026-XX / J-2026-YY in notes → PDF name
    if notes:
        # Find HF-2026-XX or J-2026-YY tokens
        tokens = re.findall(r"[A-Z]{1,3}-\d{4}-\d{1,3}", notes.upper())
        for tok in tokens:
            for f in all_folders:
                if not f.is_dir():
                    continue
                try:
                    for child in iter_evidence_files(f):
                        if tok in child.name.upper():
                            return f
                except Exception:
                    pass
    return None


def collect_all_folders(raw_root: Path) -> list[Path]:
    """Return logical evidence ref folders, accepting flat or date-layer uploads."""
    buckets: dict[str, list[Path]] = {}
    for folder in _candidate_ref_folders(raw_root):
        if _has_direct_evidence_files(folder) or _single_evidence_child(folder):
            buckets.setdefault(_logical_ref_key(folder), []).append(folder)

    folders: list[Path] = []
    for siblings in buckets.values():
        primary = sorted(siblings, key=lambda p: str(p))[0]
        folders.append(primary)
        _MERGED_EVIDENCE_FOLDERS[str(primary.resolve(strict=False))] = sorted(
            siblings, key=lambda p: str(p)
        )
    return sorted(folders, key=lambda p: str(p))


def collect_evidence(folder: Path) -> dict:
    """Read every .msg and .pdf in folder. Return {'msgs': [{name, body, ...}], 'pdfs': [{name, text}], 'files': [name,...]}"""
    out = {"folder": str(folder), "msgs": [], "pdfs": [], "files": [], "total_chars": 0}
    if not folder or not folder.exists():
        return out
    for child in iter_evidence_files(folder):
        out["files"].append(child.name)
        if child.is_file():
            n = child.name.lower()
            if n.endswith(".msg"):
                try:
                    parsed = parse_msg(child)
                    body = parsed.get("body_text", parsed.get("body", ""))
                    out["msgs"].append({
                        "filename": child.name,
                        "subject": parsed.get("subject", ""),
                        "from": parsed.get("sender", parsed.get("from", "")),
                        "to": parsed.get("to", ""),
                        "cc": parsed.get("cc", ""),
                        "date": parsed.get("received_at", parsed.get("date", "")),
                        "body": body,
                    })
                    out["total_chars"] += len(body or "")
                except Exception as e:
                    out["msgs"].append({"filename": child.name, "error": str(e)})
            elif n.endswith(".pdf"):
                text = extract_pdf_text(child)
                out["pdfs"].append({"filename": child.name, "text": text})
                out["total_chars"] += len(text)
    return out


# ---------- gemini ----------
def call_gemini(prompt: str, model: str, retries: int = 3) -> dict:
    """Call Gemini with retry. Returns {'json': dict, 'in_tokens': int, 'out_tokens': int, 'model': str, 'error': str}."""
    url = f"{GEMINI_BASE_URL}/models/{model}:generateContent?key={GEMINI_API_KEY}"
    payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {
            "temperature": 0.05,
            "maxOutputTokens": 8192,
            "responseMimeType": "application/json",
        },
    }
    last_err = None
    for attempt in range(retries):
        try:
            req = urllib.request.Request(
                url, data=json.dumps(payload).encode(),
                headers={"Content-Type": "application/json", "User-Agent": "AlJeel-AP-Agent/1.0", "cf-aig-custom-cost": _cf_custom_cost_header(model)},
            )
            with urllib.request.urlopen(req, timeout=120) as resp:
                data = json.loads(resp.read())
            cand = (data.get("candidates") or [{}])[0]
            content = cand.get("content") or {}
            parts = content.get("parts") or [{}]
            text = parts[0].get("text", "")
            # Strip fences
            text = re.sub(r"^```(?:json)?\s*", "", text.strip())
            text = re.sub(r"\s*```$", "", text)
            text = text.strip()
            if not text.startswith("{"):
                m = re.search(r"\{.*\}", text, re.DOTALL)
                if m:
                    text = m.group()
            parsed = json.loads(text) if text else {}
            usage = data.get("usageMetadata") or {}
            return {
                "json": parsed,
                "in_tokens": usage.get("promptTokenCount", 0),
                "out_tokens": (usage.get("candidatesTokenCount", 0) + usage.get("thoughtsTokenCount", 0)),
                "model": model,
                "error": "",
            }
        except Exception as e:
            last_err = f"{type(e).__name__}: {e}"
            time.sleep(2 ** attempt)
    return {"json": {}, "in_tokens": 0, "out_tokens": 0, "model": model, "error": last_err or "unknown"}


def call_gemini_with_cascade(prompt: str) -> dict:
    for model in GEMINI_MODEL_CASCADE:
        res = call_gemini(prompt, model)
        if res["json"] and not res["error"]:
            return res
    return res


# ---------- per-row processing ----------
def build_prompt(row: dict, evidence: dict, master_snapshot: dict) -> str:
    def _s(v):
        if v is None:
            return ""
        if isinstance(v, list):
            return ", ".join(_s(x) for x in v)
        return str(v)
    msg_chunks = []
    for m in evidence["msgs"]:
        msg_chunks.append(
            f"=== EMAIL FILE: {_s(m.get('filename',''))} ===\n"
            f"Subject: {_s(m.get('subject',''))}\n"
            f"From: {_s(m.get('from',''))}\n"
            f"To: {_s(m.get('to',''))}\n"
            f"Cc: {_s(m.get('cc',''))}\n"
            f"Date: {_s(m.get('date',''))}\n"
            f"\n{_s(m.get('body','')) or _s(m.get('error','[no body]'))}\n"
        )
    pdf_chunks = []
    for p in evidence["pdfs"]:
        pdf_chunks.append(
            f"=== PDF FILE: {p.get('filename','')} ===\n{p.get('text','')}\n"
        )

    no_evidence_warning = ""
    if not evidence.get("total_chars"):
        no_evidence_warning = """
# NO EVIDENCE FOUND FOR THIS ROW
IMPORTANT: No evidence documents were found for this row. Do NOT guess from the master snapshot alone. Set confidence=very_low and flag this row for manual review. If you cannot find the requestor from the description text alone, set emp_no to the most likely requestor based only on the description — acknowledge the uncertainty.
"""

    prompt = f"""You are a senior Accounts Payable Audit Specialist for Al Jeel Medical (Saudi medical-equipment distributor).
Your job: given the invoice row + ALL evidence files for this ticket, analyze the documents and output the correct 6-field Oracle GL allocation.

# OUTPUT FORMAT (strict JSON, no extra fields, no markdown fences)
{{
  "emp_no": "<7-digit AlJeel employee number, or comma-separated numbers if multiple salesmen e.g. '1000477,1002037' — NEVER blank or 0000000; for sponsorship use the requestor's number(s), for employee travel use the traveler's number>",
  "account": "<8-digit GL account>",
  "cost_center": "<6-digit cost center>",
  "div": "<3-digit division code>",
  "solution": "<5-digit solution code>",
  "agency": "<5-digit agency code>",
  "opex_serial": "<OPEX serial code e.g. CRM-2026-30, CE-18-2026, HF-2026-23 — or empty string if not found>",
  "confidence": "high|medium|low|very_low",
  "reasoning": "<1-2 sentences: cite the exact file and form field used for this classification>"
}}

# AUDIT DECISION LOGIC (Determine which scenario this transaction represents)

## Scenario A: Employee Corporate Business Travel / Training
* **Criteria:** The passenger is an Al Jeel employee traveling for business (e.g., attending partner forums, corporate meetings, or training sessions).
  *(IMPORTANT CAVEAT: Al Jeel's Oracle ERP system automatically routes all employee travel approvals—including corporate business trips, seminars, and cash advances—with the subject 'Personal Contribution Approval Requested...'. Do NOT classify a row under Scenario B simply because the email subject says 'Personal Contribution'. You must inspect the actual workflow form inside the email body: if the transaction details list 'Business Trip', 'External Business Trip', or 'Partner Meeting', classify it as Scenario A, not Scenario B!)*
* **Account Code:** 
  - Use `60301003` (Travel Tickets Expense) for normal business travel.
  - Use `60308009` (Training Expenses) if the document explicitly shows a training course/registration fee.
* **Employee No:** Put the traveler's 7-digit **employee number** from the Manpower database.
* **Allocation segments (CC, DIV, Agency, Solution):** Map directly to the traveler's **home segments** from the Employee directory below.

## Scenario B: Employee Personal Travel / Personal Contribution
* **Criteria:** The trip is personal travel (annual leave, vacation) approved under Al Jeel's personal ticket / accrued allowance award.
  *(IMPORTANT CAVEAT: Only select Scenario B if the transaction details inside the email body explicitly list 'Annual Leave' or 'Personal Travel' awards. If the details list 'Business Trip' or 'External Business Trip', do NOT select Scenario B even if the email subject says 'Personal Contribution Approval Requested'.)*
* **Account Code:** Use `21070229` (Accrued Employee Annual Tickets).
* **Employee No:** Put the traveler's 7-digit **employee number** (from the Manpower database).
* **Allocation segments (CC, DIV, Agency, Solution):** Map directly to the traveler's **home segments** from the Employee directory below.

## Scenario C: Sponsoring Expenses (External HCP / Guest Travel)
* **Criteria:** The passenger is an external doctor or guest (not an employee), or the trip is sponsored under a specific vendor marketing event (e.g. Abbott, CRM, EP, HF).
* **Account Code:** Use `60307021` (Sponsoring Expenses).
* **Employee No:** Use the **requesting AlJeel employee's 7-digit employee number** — the person(s) who submitted the OPEX form. Find them in the "Salesman" / "Amount to Allocate" table, email thread, OPEX form header, or approval chain.
  - If ONE salesman: return their 7-digit employee number (e.g. "1000477")
  - If MULTIPLE salesmen in the allocation table: return ALL their employee numbers comma-separated, in the order they appear (e.g. "1000477,1002037")
  - Look up each name in the Employee directory below to confirm the number
  - Do NOT return "0000000" — if you cannot find any requestor, return your best match from the evidence
  - Do NOT use the external guest/client names (Client Name field) as the emp_no source
* **Allocation segments (CC, DIV, Agency, Solution):** 
  1. Find the **Requesting Al Jeel Employee (Sponsor)** named in the OPEX form or email thread.
  2. Copy that Requesting Employee's **home cost_center and div_code** from the Employee directory below.
  3. Map **Agency and Solution** to the specific vendor/marketing brand approved on the form (e.g., Abbott = `10072`, S&M = `10200`). Use the code hints below.

# OPEX SERIAL EXTRACTION
* Every OPEX/sponsorship submission carries a serial code. Look for it in ALL evidence files
  (PDF form headers, email subjects, email bodies, folder names).
* Format: <DEPT_PREFIX>-<NUMBER>-<YEAR> where prefix is 2-5 letters.
  Examples: CRM-2026-30, CE-18-2026, HF-2026-23, EP-2026-05, AATS-2026-11
  The number and year order can vary: both CE-18-2026 and CRM-2026-30 are valid.
* Known department prefixes: CE (Capital Equipment), CRM, HF (Heart Failure), EP, AATS
* If found: return it exactly as written in the document (e.g. "CRM-2026-30").
* If not found in any evidence: return empty string "".
* Only attempt extraction for Scenario C (sponsorship) rows. For Scenario A/B (employee
  travel), always return "".

# CRITICAL SEGMENT FORMATTING RULES
* cost_center: 6 digits ALWAYS (e.g., "160014", "160012")
* solution: 5 digits ALWAYS (e.g., "00000", "10050")
* agency: 5 digits ALWAYS (e.g., "10072", "10200")
* div: 3-digit code as in directory (e.g., "170", "194", "120")

# INVOICE ROW TO AUDIT
- Description: {row.get('description','')}
- Passenger / Service: {row.get('passenger','')}
- Route: {row.get('route','')}
- Ticket #: {row.get('ticket_no','')}
- Amount: {row.get('amount','')}
- Date: {row.get('date','')}
- Notes (if any): {row.get('notes','') or '(none)'}

# EVIDENCE FILES FOR THIS TICKET (FOLDER: {evidence.get('folder','MISSING')})
- Files present: {', '.join(evidence.get('files', [])) or 'NONE'}

## EMAILS (.msg files, full body)
{chr(10).join(msg_chunks) if msg_chunks else '(no email files)'}

## PDFs (full text)
{chr(10).join(pdf_chunks) if pdf_chunks else '(no PDF files)'}
{no_evidence_warning}
# MASTER DATA LOOKUPS

## Employee directory (Manpower snapshot of potential matches). Use these HOME values to allocate.
{json.dumps(master_snapshot.get('employees', {}), ensure_ascii=False, indent=1)}

## DIV codes
{json.dumps(master_snapshot.get('div', {}), ensure_ascii=False, indent=1)}

## Agency code hints
{json.dumps(master_snapshot.get('agency_hint', {}), ensure_ascii=False, indent=1)}

## Solution code hints
{json.dumps(master_snapshot.get('solution_hint', {}), ensure_ascii=False, indent=1)}

# OUTPUT COMMAND
Analyze the documents, make your classification choice, use the lookups, and return ONLY the strict JSON object. No prose. No code blocks.
"""
    return prompt


def build_master_snapshot(manpower: dict, lookups: dict, evidence: dict, row: dict) -> dict:
    """Inline only the relevant subset of master data (small to keep prompt tight)."""
    # Find employee names mentioned in evidence + row → include those in employee snapshot
    all_text = row.get("description", "") + " " + row.get("passenger", "") + " " + row.get("notes", "")
    def _s(v):
        if v is None:
            return ""
        if isinstance(v, list):
            return " ".join(_s(x) for x in v)
        return str(v)
    evidence_text = ""
    for m in evidence["msgs"]:
        evidence_text += " " + _s(m.get("body", "")) + " " + _s(m.get("subject", "")) + " " + _s(m.get("from", "")) + " " + _s(m.get("to", ""))
    for p in evidence["pdfs"]:
        evidence_text += " " + (p.get("text", "") or "")
    all_text += evidence_text
    
    def _emp_info(info):
        return {
            "name": info["name"],
            "div_code": info["div_code"],
            "div_name": info["div_name"],
            "agency_code": info["agency_code"],
            "agency_name": info["agency_name"],
            "cost_center": info["cost_center"],
            "cc_name": info["cc_name"],
            "solution": info["solution"],
        }
    emp_subset = {}
    # PRIORITY 1: any emp_no that appears as 7-digit numeric token in the text (highest signal)
    for tok in re.findall(r"\b\d{7}\b", all_text):
        if tok in manpower and tok not in emp_subset:
            emp_subset[tok] = _emp_info(manpower[tok])
    # PRIORITY 2: passenger name exact match (surname token, len>=4)
    passenger = row.get("passenger", "").upper()
    if passenger:
        for tok in re.split(r"[\s\-,/]+", passenger):
            tok = tok.strip()
            if len(tok) >= 4 and tok != "MR" and tok != "MS" and tok != "DR" and tok != "MRS":
                for emp_no, info in manpower.items():
                    name_upper = info.get("name", "").upper()
                    if tok in name_upper.split() or tok in re.split(r"[\s\-]+", name_upper):
                        if emp_no not in emp_subset:
                            emp_subset[emp_no] = _emp_info(info)
    # PRIORITY 3: other name token matches, with two guards (J26-870 row 30 fix):
    # - against the description, only the PASSENGER NAME segment (text before the
    #   first " - ") is searched — the route/hotel tail let OTHMAN match the hotel
    #   name "Kempinski Al Othman Hotel"
    # - tokens must match on word boundaries, not as substrings — plain `in`
    #   let HUSSEIN match inside ALHUSSEIN
    passenger_segment = (row.get("description", "") or "").split(" - ", 1)[0]
    p3_text = passenger_segment + " " + evidence_text
    if len(emp_subset) < 40:
        for emp_no, info in manpower.items():
            if emp_no in emp_subset:
                continue
            name = info.get("name", "")
            if not name:
                continue
            for tok in re.split(r"[\s\-,/]+", name):
                tok = tok.strip().upper()
                if len(tok) >= 5 and re.search(r"\b" + re.escape(tok) + r"\b", p3_text, re.IGNORECASE):
                    emp_subset[emp_no] = _emp_info(info)
                    break
            if len(emp_subset) >= 40:
                break

    # Compact lookup hints — the LLM doesn't need all 677 agencies, just the brand-code mapping it needs.
    # Hardcode known common mappings + give it the most-likely codes from employees in scope.
    agency_hint = {
        "ABBOTT": "10072", "GE": "10081", "NUBOMED": "10156", "BSC": "10101",
        "BMX": "10153", "Technical Services": "10206", "BIOTRONIK": "10058",
        "MEDTRONIC": "10133", "PHILIPS": "10168", "PENUMBRA": "10163",
        "KAVO": "10005", "SOLVENTUM": "10202", "3M": "10202",
        "GENERAL": "00000",
    }
    # OPEX-XX-PREFIX-2026 prefix maps to division/agency
    opex_division_hint = {
        "HF": {"div": "170", "cc": "160014", "sol": "10050", "agency_likely": "10072"},
        "EP": {"div": "170", "cc": "160014", "sol": "10064", "agency_likely": "10072"},
        "CRM": {"div": "170", "cc": "160014", "sol": "10094", "agency_likely": "10072"},
        "DMS": {"div": "192", "cc": "160013", "sol": "00000", "agency_likely": "10005"},  # Dental & Medical Solutions
        "IVD": {"div": "194", "cc": "160012", "sol": "00000", "agency_likely": "10153"},
        "S&M": {"div": "190", "cc": "140020", "sol": "00000", "agency_likely": "10200"},
        "GE":  {"div": "196", "cc": "160011", "sol": "00000", "agency_likely": "10081"},
        "BMX": {"div": "194", "cc": "160012", "sol": "00000", "agency_likely": "10153"},
    }
    solution_hint = {
        "HF": "10050", "EP": "10064", "CRM": "10017", "GENERAL": "00000",
        "BUSINESS_TRIP": "00000",
    }
    return {
        "employees": emp_subset,
        "div": lookups["div"],
        "agency_hint": agency_hint,
        "solution_hint": solution_hint,
        "opex_division_hint": opex_division_hint,
    }


def process_row(row_idx: int, row: dict, manpower: dict, lookups: dict, all_folders: list[Path]) -> dict:
    """Process a single row. Returns full result dict."""
    cache_key = f"row{row_idx:03d}-{row.get('ticket_no','noticket')}.json"
    cache_path = CACHE_DIR / cache_key
    if cache_path.exists():
        try:
            cached = json.loads(cache_path.read_text())
            cached["from_cache"] = True
            return cached
        except Exception:
            pass

    folder = find_ticket_folder(row.get("ticket_no", ""), row.get("passenger", ""), row.get("notes", ""), all_folders)
    evidence = collect_evidence(folder) if folder else {"folder": "", "msgs": [], "pdfs": [], "files": [], "total_chars": 0}
    snapshot = build_master_snapshot(manpower, lookups, evidence, row)
    prompt = build_prompt(row, evidence, snapshot)
    
    t0 = time.time()
    gemini_res = call_gemini_with_cascade(prompt)
    elapsed = time.time() - t0
    
    result = {
        "row_idx": row_idx,
        "row": row,
        "folder": str(folder) if folder else "",
        "evidence_files": evidence.get("files", []),
        "evidence_chars": evidence.get("total_chars", 0),
        "prompt_chars": len(prompt),
        "gemini": gemini_res,
        "elapsed_sec": elapsed,
        "from_cache": False,
    }
    try:
        cache_path.write_text(json.dumps(result, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    except Exception as e:
        result["cache_err"] = str(e)
    return result


# ---------- I/O ----------
def load_truth_rows() -> list[dict]:
    """Load all 160 rows from the truth file (Details sheet). For BLIND test, we keep only the input fields."""
    wb = openpyxl.load_workbook(TRUTH_XLSX, read_only=True, data_only=True)
    ws = wb["Details"]
    rows = list(ws.iter_rows(values_only=True))
    out = []
    # Header at row 0; data rows 1+
    for i, row in enumerate(rows[1:], start=1):
        if not row or row[0] is None:
            continue
        # Stop at footer
        if isinstance(row[0], str) and not str(row[0]).strip().isdigit():
            break
        try:
            int(str(row[0]).strip())
        except Exception:
            break
        ticket_raw = str(row[3] or "").strip()
        # ticket format "065 6905055971" — get the 10+ digit number
        m = re.search(r"\d{10,}", ticket_raw)
        ticket_no = m.group() if m else ticket_raw
        out.append({
            "row_idx": i,
            "sl_no": row[0],
            "date": str(row[1] or ""),
            "ref_no": str(row[2] or ""),
            "ticket_no_raw": ticket_raw,
            "ticket_no": ticket_no,
            "passenger": str(row[4] or ""),
            "route": str(row[5] or ""),
            "service_date": str(row[6] or ""),
            "amount": row[10],  # Inv. Amt. Incl. VAT
            "description": str(row[29] or "") if len(row) > 29 else "",
            "notes": str(row[12] or "") if len(row) > 12 else "",
            # GROUND TRUTH (kept separate; never shown to LLM)
            "truth_emp_no": str(row[11] or "").strip().rstrip("."),
            "truth_account": str(row[16] or "").strip(),
            "truth_cost_center": str(row[18] or "").strip(),
            "truth_div": str(row[20] or "").strip(),
            "truth_solution": str(row[22] or "").strip(),
            "truth_agency": str(row[24] or "").strip(),
        })
    wb.close()
    return out


def normalize_code(v: Any, width: int = 0) -> str:
    if v is None:
        return ""
    s = str(v).strip().rstrip(".")
    # Strip leading zeros for comparison? No — codes are zero-padded; keep as-is.
    # But normalize: ensure pure-digit codes with the same numeric value match.
    if width > 0 and s.isdigit():
        return s.zfill(width)
    return s


def score_results(results: list[dict]) -> dict:
    """Score each row against ground truth."""
    per_row = []
    total = 0
    fields = ["emp_no", "account", "cost_center", "div", "solution", "agency"]
    counts = {f: 0 for f in fields}
    all_five_count = 0  # 5 of 6 = account, CC, div, sol, agency
    sponsorship_idxs = []
    sponsorship_all5 = 0
    travel_idxs = []
    travel_all5 = 0
    for r in results:
        if r is None or "row" not in r:
            continue
        total += 1
        row = r["row"]
        gem = r.get("gemini", {}).get("json", {}) or {}
        match = {}
        # emp_no — compare as ints when possible
        truth = {f: row.get(f"truth_{f}", "") for f in fields}
        out = {f: str(gem.get(f, "") or "").strip().rstrip(".") for f in fields}
        # normalize
        def n(s):
            return str(s or "").strip().rstrip(".").lstrip("0") or "0"
        for f in fields:
            t = n(truth[f])
            o = n(out[f])
            if f == "emp_no":
                # emp_no scoring: only count rows where truth has a real emp_no (not "-" / blank / 0)
                if t in ("", "-", "0"):
                    match[f] = None  # skip
                else:
                    match[f] = (t == o and bool(o) and o != "0")
            else:
                match[f] = (t == o and bool(truth[f].strip()) and bool(out[f].strip()))
            if match[f] is True:
                counts[f] += 1
        # all 5 = account, cc, div, sol, agency  (not emp_no per Amr's original framework)
        all5 = all(match[f] for f in ["account", "cost_center", "div", "solution", "agency"])
        if all5:
            all_five_count += 1
        # Classify row
        is_sponsorship = truth["account"] == "60307021"
        is_travel = truth["account"] == "60301003"
        if is_sponsorship:
            sponsorship_idxs.append(r["row_idx"])
            if all5:
                sponsorship_all5 += 1
        if is_travel:
            travel_idxs.append(r["row_idx"])
            if all5:
                travel_all5 += 1
        per_row.append({
            "row_idx": r["row_idx"],
            "passenger": row.get("passenger", ""),
            "ticket": row.get("ticket_no", ""),
            "truth": truth,
            "predicted": out,
            "match": match,
            "all_five": all5,
            "category": "sponsorship" if is_sponsorship else ("travel" if is_travel else "other"),
            "confidence": gem.get("confidence", ""),
            "reasoning": gem.get("reasoning", "")[:200],
            "evidence_files_count": len(r.get("evidence_files", [])),
            "evidence_chars": r.get("evidence_chars", 0),
            "folder": r.get("folder", ""),
        })
    return {
        "total": total,
        "all_five_exact": all_five_count,
        "all_five_pct": 100 * all_five_count / total if total else 0,
        "field_counts": counts,
        "field_pct": {f: 100 * counts[f] / total for f in fields} if total else {},
        "emp_no_scorable": sum(1 for r in per_row if r['match']['emp_no'] is not None),
        "emp_no_correct": sum(1 for r in per_row if r['match']['emp_no'] is True),
        "sponsorship_n": len(sponsorship_idxs),
        "sponsorship_all5": sponsorship_all5,
        "sponsorship_pct": 100 * sponsorship_all5 / len(sponsorship_idxs) if sponsorship_idxs else 0,
        "travel_n": len(travel_idxs),
        "travel_all5": travel_all5,
        "travel_pct": 100 * travel_all5 / len(travel_idxs) if travel_idxs else 0,
        "per_row": per_row,
    }


def write_output_xlsx(results: list[dict], path: Path):
    """Write results in the same column structure as the v15.11.2 output (truncated to essential fields)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Match v15.11.2 format header rows
    ws.append([])  # row 1 blank
    section = ["ORACLE FUSION TEMPLATE"] + [""] * 15 + ["CODE & DESCRIPTION"] + [""] * 15 + ["LLM AGENT OUTPUT"]
    ws.append(section)  # row 2
    
    hdr = [
        "*Invoice Header Identifier", "*Business Unit", "*Invoice Number", "*Invoice Currency",
        "*Invoice Amount", "*Invoice Date", "**Supplier", "**Supplier Number", "*Supplier Site",
        "Invoice Type", "Description", "*Type", "*Amount", "Distribution Combination",
        "Tax Classification Code", "Employee No",
        "Company", "Location", "Account", "GL", "Cost Center", "Cost Name", "DIV", "Contribution",
        "Solution", "Solution Name", "Agency", "Agency Name", "Project", "Intercompany", "Future 1",
        "GL Description",
        # Agent fields
        "Agent Confidence", "Agent Reasoning", "Agent Model", "Agent Folder", "Agent EvFiles", "Agent EvChars",
        "Agent InTokens", "Agent OutTokens", "Agent Error",
    ]
    ws.append(hdr)  # row 3
    
    for r in results:
        if r is None:
            continue
        row = r["row"]
        gem = r.get("gemini", {}).get("json", {}) or {}
        gem_info = r.get("gemini", {})
        out_row = [
            row.get("sl_no", ""), "Al Jeel Medical BU", "J26-593", "SAR",
            row.get("amount", ""), row.get("date", ""),
            "شركة جوال للسفر والسياحة المحد", "10394", "شركة جوال للسفر",
            "Standard",
            f"{row.get('passenger','')} - {row.get('route','')}",
            "Item", row.get("amount", ""),
            f"03-40100-{gem.get('account','')}-{gem.get('cost_center','')}-{gem.get('div','')}-{gem.get('solution','')}-{gem.get('agency','')}-00000-00-000000",
            "KSA VAT STANDARD",
            gem.get("emp_no", ""),
            # Code & description columns
            "03", "40100", gem.get("account", ""), "", gem.get("cost_center", ""), "", gem.get("div", ""), "",
            gem.get("solution", ""), "", gem.get("agency", ""), "", "00000", "00", "000000",
            "",
            # Agent
            gem.get("confidence", ""), gem.get("reasoning", ""), gem_info.get("model", ""),
            r.get("folder", ""), len(r.get("evidence_files", [])), r.get("evidence_chars", 0),
            gem_info.get("in_tokens", 0), gem_info.get("out_tokens", 0), gem_info.get("error", ""),
        ]
        ws.append(out_row)
    
    wb.save(str(path))


# ---------- main ----------
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=0, help="0 = all rows")
    parser.add_argument("--workers", type=int, default=5)
    parser.add_argument("--no-cache", action="store_true")
    args = parser.parse_args()
    
    if args.no_cache:
        # Clear cache
        for f in CACHE_DIR.glob("*.json"):
            f.unlink()
    
    print(f"[load] reading truth file: {TRUTH_XLSX}", flush=True)
    truth_rows = load_truth_rows()
    if args.limit > 0:
        truth_rows = truth_rows[: args.limit]
    print(f"[load] {len(truth_rows)} rows to process", flush=True)
    
    print(f"[load] master data...", flush=True)
    manpower = load_manpower()
    lookups = load_lookups()
    print(f"[load] {len(manpower)} employees, {len(lookups['cc'])} CC, {len(lookups['div'])} DIV, {len(lookups['solution'])} Solution, {len(lookups['agency'])} Agency", flush=True)
    
    print(f"[scan] collecting all folder candidates...", flush=True)
    all_folders = collect_all_folders(RAW_ROOT)
    print(f"[scan] {len(all_folders)} ticket-folder candidates", flush=True)
    
    results = [None] * len(truth_rows)
    t_start = time.time()
    done = 0
    
    def worker(idx_row):
        idx, row = idx_row
        try:
            return idx, process_row(row["row_idx"], row, manpower, lookups, all_folders)
        except Exception as e:
            tb = traceback.format_exc()
            return idx, {"row_idx": row["row_idx"], "row": row, "error": str(e), "traceback": tb, "gemini": {"json": {}, "in_tokens": 0, "out_tokens": 0, "model": "", "error": str(e)}}
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=args.workers) as ex:
        futs = [ex.submit(worker, (i, r)) for i, r in enumerate(truth_rows)]
        for fut in concurrent.futures.as_completed(futs):
            idx, res = fut.result()
            results[idx] = res
            done += 1
            elapsed = time.time() - t_start
            cached = " [cached]" if res.get("from_cache") else ""
            err = res.get("gemini", {}).get("error", "")
            err_marker = f" ERR={err[:30]}" if err else ""
            print(f"[{done:3d}/{len(truth_rows)}] row {res.get('row_idx','?'):3d}  t={elapsed:5.1f}s  chars={res.get('evidence_chars',0):6d}  files={len(res.get('evidence_files',[]))}{cached}{err_marker}", flush=True)
    
    total_elapsed = time.time() - t_start
    
    # Cost summary
    total_in = sum((r.get("gemini", {}).get("in_tokens", 0) or 0) for r in results if r)
    total_out = sum((r.get("gemini", {}).get("out_tokens", 0) or 0) for r in results if r)
    # use first non-cached model
    model_used = "gemini-2.5-pro"
    for r in results:
        if r and r.get("gemini", {}).get("model"):
            model_used = r["gemini"]["model"]
            break
    pricing = GEMINI_PRICING.get(model_used, GEMINI_PRICING["gemini-2.5-pro"])
    # Per-call pricing so the >200K long-context rate applies correctly per request, not per batch.
    cost_usd = 0.0
    for _r in results:
        if not _r:
            continue
        _g = _r.get("gemini", {}) or {}
        _it = _g.get("in_tokens", 0) or 0
        _ot = _g.get("out_tokens", 0) or 0
        _m = _g.get("model") or model_used
        _rate = _gemini_rate(_m, _it)
        cost_usd += _it * _rate["in"] + _ot * _rate["out"]
    
    # Score
    score = score_results(results)
    
    ts = time.strftime("%Y%m%dT%H%M%SZ", time.gmtime())
    
    # Write Excel output
    xlsx_path = OUT_DIR / f"Spreadsheet-J26-593-LLMAGENT-v1.xlsx"
    write_output_xlsx([r for r in results if r], xlsx_path)
    print(f"[out] wrote {xlsx_path}", flush=True)
    
    # Write summary
    cost_summary = {
        "timestamp_utc": ts,
        "rows_processed": len(results),
        "total_in_tokens": total_in,
        "total_out_tokens": total_out,
        "model_used": model_used,
        "cost_usd": round(cost_usd, 4),
        "total_runtime_sec": round(total_elapsed, 1),
        "rate_per_min": round(60 * len(results) / total_elapsed, 1) if total_elapsed else 0,
    }
    (OUT_DIR / f"cost-summary.json").write_text(json.dumps(cost_summary, indent=2), encoding="utf-8")
    print(f"[out] cost: ${cost_usd:.4f} ({total_in:,} in / {total_out:,} out tokens, model={model_used})", flush=True)
    
    # Write score JSON
    score_path = OUT_DIR / f"score-llmagent.json"
    score_path.write_text(json.dumps(score, indent=2, default=str), encoding="utf-8")
    print(f"[out] wrote {score_path}", flush=True)
    
    # Print top-line score
    print()
    print("=" * 60)
    print(f"  ALL-5-EXACT:     {score['all_five_exact']}/{score['total']}  ({score['all_five_pct']:.1f}%)")
    print(f"  Sponsorships:    {score['sponsorship_all5']}/{score['sponsorship_n']}  ({score['sponsorship_pct']:.1f}%)")
    print(f"  Travel:          {score['travel_all5']}/{score['travel_n']}  ({score['travel_pct']:.1f}%)")
    print(f"  Emp_no:          {score['emp_no_correct']}/{score['emp_no_scorable']}  (where truth set)")
    print(f"  Cost / runtime:  ${cost_usd:.4f}  /  {total_elapsed/60:.1f} min")
    print(f"  Per-field:")
    for f, pct in score["field_pct"].items():
        print(f"    {f:14s} {pct:5.1f}%   ({score['field_counts'][f]}/{score['total']})")
    print("=" * 60)


if __name__ == "__main__":
    main()
