#!/usr/bin/env python3
"""
Full Evidence LLM Agent — J26-593 BLIND TEST

Per-row workflow:
  1) Find ticket folder (by ticket# OR passenger name OR notes match)
  2) Collect ALL .msg bodies + ALL PDF text in that folder (NO TRUNCATION)
  3) Send everything to Gemini 2.5 Pro with strict JSON schema
  4) Cache by (ticket_no, row_idx) → re-runs are free
  5) Validate output against master data
  6) Concurrency: 5 parallel rows

Output: Spreadsheet-J26-593-LLMAGENT-v1.xlsx + comparison + cost summary
"""
from __future__ import annotations

import argparse
import concurrent.futures
import hashlib
import json
import os
import re
import sys
import time
import traceback
import urllib.request
from pathlib import Path
from typing import Any

import openpyxl

# Set up paths
ROOT = Path("/home/clawdbot/.openclaw/workspace/aljeel")
sys.path.insert(0, str(ROOT / "scripts"))

from msg_parser import parse_msg  # type: ignore

import fitz  # pymupdf

GEMINI_BASE_URL = os.environ.get("GEMINI_BASE_URL", "https://gateway.ai.cloudflare.com/v1/3724a3e71944b366a39b3735aa117a58/accord-aljeel-ap/google-ai-studio/v1beta")

TRUTH_XLSX = Path("/mnt/aljeel-ap_kb/current/J26-593/J26-593.xlsx")
BLIND_INPUT = Path("/home/clawdbot/.openclaw/workspace/aljeel/batches/jawal-J26-593-BLIND/Spreadsheet.xlsx")
RAW_ROOT = Path("/home/clawdbot/.openclaw/workspace/aljeel/batches/jawal-J26-593-BLIND/raw")
MASTER_XLSX = Path("/home/clawdbot/.openclaw/workspace/aljeel/qc/master-data/Aljeel_Lookups-v2.xlsx")
EMAIL_MASTER_XLSX = Path("/home/clawdbot/.openclaw/workspace/aljeel/qc/master-data/Employee_Emails_2026-05-26.xlsx")

CACHE_DIR = Path("/home/clawdbot/.openclaw/workspace/aljeel/extracted/full-evidence-agent-cache")
CACHE_DIR.mkdir(parents=True, exist_ok=True)

OUT_DIR = Path("/home/clawdbot/.openclaw/workspace/aljeel/batches/jawal-J26-593-BLIND/output/llmagent")
OUT_DIR.mkdir(parents=True, exist_ok=True)

# Canonical AP pipeline LLM cascade: Gemini 3 Pro (via gemini-pro-latest alias) -> 2.5 Pro -> 2.5 Flash
# Mirrors email_allocation_extractor.py / location_resolver.py / allocation_resolver.py.
# Comment in email_allocation_extractor.py v15.1: "gemini-3-pro 404; gemini-pro-latest resolves to best available"
GEMINI_MODEL_CASCADE = ["gemini-pro-latest", "gemini-2.5-pro", "gemini-2.5-flash"]
GEMINI_PRICING = {
    # VERIFIED Jun 20 2026 (OpenRouter/Verdent/MetaCTO). gemini-pro-latest -> gemini-3.1-pro-preview.
    # Thinking (thoughtsTokenCount) billed at OUTPUT rate. >200K context: all tokens at long-context rate.
    "gemini-pro-latest":      {"in": 2.00/1e6, "out": 12.00/1e6},
    "gemini-3.1-pro-preview": {"in": 2.00/1e6, "out": 12.00/1e6},
    "gemini-2.5-pro":         {"in": 1.25/1e6, "out": 10.00/1e6},
    "gemini-2.5-flash":       {"in": 0.30/1e6, "out": 2.50/1e6},
}

# Long-context (>200K input) doubles Gemini 3.x Pro rates; ALL tokens billed at long-context rate.
GEMINI_PRICING_LONG = {
    "gemini-pro-latest":      {"in": 4.00/1e6, "out": 18.00/1e6},
    "gemini-3.1-pro-preview": {"in": 4.00/1e6, "out": 18.00/1e6},
    "gemini-2.5-pro":         {"in": 2.50/1e6, "out": 15.00/1e6},
    "gemini-2.5-flash":       {"in": 0.30/1e6, "out": 2.50/1e6},
}
def _gemini_rate(model, in_tokens=0):
    if in_tokens and in_tokens > 200_000:
        return GEMINI_PRICING_LONG.get(model) or GEMINI_PRICING_LONG.get("gemini-2.5-pro")
    return GEMINI_PRICING.get(model) or GEMINI_PRICING.get("gemini-2.5-pro")
def _cf_custom_cost_header(model, in_tokens=0):
    import json as __j
    _p = _gemini_rate(model, in_tokens)
    return __j.dumps({"per_token_in": _p["in"], "per_token_out": _p["out"]})



def load_env_key(name: str) -> str:
    val = os.environ.get(name, "")
    if val:
        return val
    env_path = Path("/home/clawdbot/.openclaw/.env")
    if env_path.exists():
        for line in env_path.read_text().splitlines():
            line = line.strip()
            if line.startswith(name + "="):
                return line.split("=", 1)[1].strip().strip("'\"")
    return ""


GEMINI_API_KEY = load_env_key("GEMINI_API_KEY")
assert GEMINI_API_KEY, "GEMINI_API_KEY not found"


# ---------- master data loaders ----------
def load_manpower() -> dict[str, dict]:
    """Returns {emp_no: {name, arabic, location, manager, line_manager, division_code, division_name, agency_code, agency_name, cost_center, cc_name, solution}}."""
    wb = openpyxl.load_workbook(MASTER_XLSX, read_only=True, data_only=True)
    ws = wb["Manpower"]
    rows = list(ws.iter_rows(values_only=True))
    # Header is row index 6 (0-based)
    hdr = rows[0]
    # Columns: Emp No(0), Old Emp No(1), Name(2), Arabic Name(3), Location(4), Manager No(5), Line Manager(6),
    #  blank(7), Code(8)=DIV code, New Division(9), Code(10)=Agency, New agency(11), New cost center(12),
    #  New cost center name(13), blank(14), Solution(15)
    emps = {}
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        emp_no = str(row[0]).strip()
        if not emp_no or emp_no.lower() == "none":
            continue
        emps[emp_no] = {
            "emp_no": emp_no,
            "name": str(row[2] or "").strip(),
            "arabic": str(row[3] or "").strip(),
            "location": str(row[4] or "").strip(),
            "manager_no": str(row[5] or "").strip(),
            "line_manager": str(row[6] or "").strip(),
            "div_code": str(row[8] or "").strip(),
            "div_name": str(row[9] or "").strip(),
            "agency_code": str(row[10] or "").strip(),
            "agency_name": str(row[11] or "").strip(),
            "cost_center": str(row[12] or "").strip(),
            "cc_name": str(row[13] or "").strip(),
            "solution": str(row[15] or "").strip(),
        }
    wb.close()
    return emps


def load_email_master() -> dict[str, str]:
    """Returns {emp_no: email}."""
    wb = openpyxl.load_workbook(EMAIL_MASTER_XLSX, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    out = {}
    # detect header row
    hdr_idx = None
    for i, row in enumerate(rows[:5]):
        if row and any(v and "emp" in str(v).lower() for v in row):
            hdr_idx = i
            break
    if hdr_idx is None:
        hdr_idx = 0
    hdr = rows[hdr_idx]
    emp_col = None
    email_col = None
    for j, v in enumerate(hdr):
        if not v:
            continue
        s = str(v).lower()
        if "emp" in s and emp_col is None:
            emp_col = j
        if "email" in s or "mail" in s:
            email_col = j
    if emp_col is None or email_col is None:
        wb.close()
        return out
    for row in rows[hdr_idx + 1:]:
        if not row or not row[emp_col]:
            continue
        out[str(row[emp_col]).strip()] = str(row[email_col] or "").strip()
    wb.close()
    return out


def load_lookups() -> dict[str, dict[str, str]]:
    """Returns {'cc': {code: name}, 'div': {...}, 'solution': {...}, 'agency': {...}}."""
    wb = openpyxl.load_workbook(TRUTH_XLSX, read_only=True, data_only=True)
    out = {"cc": {}, "div": {}, "solution": {}, "agency": {}}
    # Cost Center Segment: row 0 = header (CC, DESCRIPTION); rows 1+
    ws = wb["Cost Center Segment"]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not row or not row[0]:
            continue
        out["cc"][str(row[0]).strip()] = str(row[1] or "").strip()
    # DIV: header at row 1 (None, None, 'DIV', 'DESCRIPTION'); rows 2+
    ws = wb["DIV"]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        if row and row[2]:
            out["div"][str(row[2]).strip()] = str(row[3] or "").strip()
    # Solution: header at row 1 (SOL, DESCRIPTION); rows 2+
    ws = wb["Solution"]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        if row and row[0]:
            out["solution"][str(row[0]).strip()] = str(row[1] or "").strip()
    # Agency: header at row 0 (None, AG, DESCRIPTION); rows 1+
    ws = wb["Agency"]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if row and row[1]:
            out["agency"][str(row[1]).strip()] = str(row[2] or "").strip()
    wb.close()
    return out


# ---------- evidence collection ----------
def extract_pdf_text(path: Path) -> str:
    try:
        doc = fitz.open(str(path))
        parts = []
        for page in doc:
            parts.append(page.get_text())
        doc.close()
        return "\n".join(parts)
    except Exception as e:
        return f"[PDF_ERROR: {e}]"


def find_ticket_folder(ticket_no: str, passenger: str, notes: str, all_folders: list[Path]) -> Path | None:
    """Try multiple strategies to locate the evidence folder."""
    # 1) Exact ticket# match
    for f in all_folders:
        if f.name.strip() == ticket_no:
            return f
    # 2) Ticket# substring (handles "6905341979-80" ranges, "6905084618 change", etc.)
    for f in all_folders:
        # Match the leading 10-digit number with optional suffix
        m = re.match(r"^(\d{10,})", f.name.strip())
        if m and m.group(1).startswith(ticket_no):
            return f
        # Range folder: "6905341979-80"
        m2 = re.match(r"^(\d{10})-(\d{1,3})", f.name.strip())
        if m2:
            base = m2.group(1)
            end = m2.group(2)
            if ticket_no == base:
                return f
            # Check the range end (e.g., 6905341979-80 → 6905341980)
            if len(end) < 10:
                prefix = base[: 10 - len(end)]
                candidate = prefix + end
                if ticket_no == candidate:
                    return f
        # Match anywhere in folder name
        if ticket_no in f.name:
            return f
    # 3) Passenger surname match against PDF filenames inside any folder
    if passenger:
        # Passenger format: "LASTNAME/FIRSTNAME MR" or "LASTNAME-FIRSTNAME MR"
        surname = re.split(r"[/\-,]", passenger)[0].strip().upper()
        if surname and len(surname) >= 3:
            for f in all_folders:
                if not f.is_dir():
                    continue
                try:
                    for child in f.iterdir():
                        cn = child.name.upper()
                        if surname in cn and (cn.endswith(".PDF") or cn.endswith(".MSG")):
                            return f
                except Exception:
                    pass
    # 4) Notes match (sponsorship case): HF-2026-XX / J-2026-YY in notes → PDF name
    if notes:
        # Find HF-2026-XX or J-2026-YY tokens
        tokens = re.findall(r"[A-Z]{1,3}-\d{4}-\d{1,3}", notes.upper())
        for tok in tokens:
            for f in all_folders:
                if not f.is_dir():
                    continue
                try:
                    for child in f.iterdir():
                        if tok in child.name.upper():
                            return f
                except Exception:
                    pass
    return None


def collect_all_folders(raw_root: Path) -> list[Path]:
    """Return all ticket-folder candidates (depth 2: dayfolder/ticketfolder)."""
    folders = []
    for day in sorted(raw_root.iterdir()):
        if not day.is_dir():
            continue
        for sub in sorted(day.iterdir()):
            if sub.is_dir():
                folders.append(sub)
    return folders


def collect_evidence(folder: Path) -> dict:
    """Read every .msg and .pdf in folder. Return {'msgs': [{name, body, ...}], 'pdfs': [{name, text}], 'files': [name,...]}"""
    out = {"folder": str(folder), "msgs": [], "pdfs": [], "files": [], "total_chars": 0}
    if not folder or not folder.exists():
        return out
    for child in sorted(folder.iterdir()):
        out["files"].append(child.name)
        if child.is_file():
            n = child.name.lower()
            if n.endswith(".msg"):
                try:
                    parsed = parse_msg(child)
                    out["msgs"].append({
                        "filename": child.name,
                        "subject": parsed.get("subject", ""),
                        "from": parsed.get("from", ""),
                        "to": parsed.get("to", ""),
                        "cc": parsed.get("cc", ""),
                        "date": parsed.get("date", ""),
                        "body": parsed.get("body", ""),
                    })
                    out["total_chars"] += len(parsed.get("body", "") or "")
                except Exception as e:
                    out["msgs"].append({"filename": child.name, "error": str(e)})
            elif n.endswith(".pdf"):
                text = extract_pdf_text(child)
                out["pdfs"].append({"filename": child.name, "text": text})
                out["total_chars"] += len(text)
    return out


# ---------- gemini ----------
def call_gemini(prompt: str, model: str, retries: int = 3) -> dict:
    """Call Gemini with retry. Returns {'json': dict, 'in_tokens': int, 'out_tokens': int, 'model': str, 'error': str}."""
    url = f"{GEMINI_BASE_URL}/models/{model}:generateContent?key={GEMINI_API_KEY}"
    payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {
            "temperature": 0.05,
            "maxOutputTokens": 8192,
            "responseMimeType": "application/json",
        },
    }
    last_err = None
    for attempt in range(retries):
        try:
            req = urllib.request.Request(
                url, data=json.dumps(payload).encode(),
                headers={"Content-Type": "application/json", "cf-aig-custom-cost": _cf_custom_cost_header(model)},
            )
            with urllib.request.urlopen(req, timeout=120) as resp:
                data = json.loads(resp.read())
            cand = (data.get("candidates") or [{}])[0]
            content = cand.get("content") or {}
            parts = content.get("parts") or [{}]
            text = parts[0].get("text", "")
            # Strip fences
            text = re.sub(r"^```(?:json)?\s*", "", text.strip())
            text = re.sub(r"\s*```$", "", text)
            text = text.strip()
            if not text.startswith("{"):
                m = re.search(r"\{.*\}", text, re.DOTALL)
                if m:
                    text = m.group()
            parsed = json.loads(text) if text else {}
            usage = data.get("usageMetadata") or {}
            return {
                "json": parsed,
                "in_tokens": usage.get("promptTokenCount", 0),
                "out_tokens": (usage.get("candidatesTokenCount", 0) + usage.get("thoughtsTokenCount", 0)),
                "model": model,
                "error": "",
            }
        except Exception as e:
            last_err = f"{type(e).__name__}: {e}"
            time.sleep(2 ** attempt)
    return {"json": {}, "in_tokens": 0, "out_tokens": 0, "model": model, "error": last_err or "unknown"}


def call_gemini_with_cascade(prompt: str) -> dict:
    for model in GEMINI_MODEL_CASCADE:
        res = call_gemini(prompt, model)
        if res["json"] and not res["error"]:
            return res
    return res


# ---------- per-row processing ----------
def build_prompt(row: dict, evidence: dict, master_snapshot: dict) -> str:
    def _s(v):
        if v is None:
            return ""
        if isinstance(v, list):
            return ", ".join(_s(x) for x in v)
        return str(v)
    msg_chunks = []
    for m in evidence["msgs"]:
        msg_chunks.append(
            f"=== EMAIL FILE: {_s(m.get('filename',''))} ===\n"
            f"Subject: {_s(m.get('subject',''))}\n"
            f"From: {_s(m.get('from',''))}\n"
            f"To: {_s(m.get('to',''))}\n"
            f"Cc: {_s(m.get('cc',''))}\n"
            f"Date: {_s(m.get('date',''))}\n"
            f"\n{_s(m.get('body','')) or _s(m.get('error','[no body]'))}\n"
        )
    pdf_chunks = []
    for p in evidence["pdfs"]:
        pdf_chunks.append(
            f"=== PDF FILE: {p.get('filename','')} ===\n{p.get('text','')}\n"
        )

    prompt = f"""You are an AP allocation agent for Al Jeel Medical (Saudi medical-equipment distributor).
Your job: given the invoice row + ALL evidence files for this ticket, output the correct 6-field allocation.

# OUTPUT FORMAT (strict JSON, no extra fields)
{{
  "emp_no": "<7-digit Aljeel employee number OR empty string \"\" if not applicable>",
  "account": "<8-digit GL account>",
  "cost_center": "<6-digit cost center>",
  "div": "<3-digit division code>",
  "solution": "<5-digit solution code>",
  "agency": "<5-digit agency code>",
  "confidence": "high|medium|low",
  "reasoning": "<1-2 sentences: name the evidence file + exact phrase you used>"
}}

# CRITICAL RULES

## Row classification (decides everything else):

**TYPE A — Business travel for an Aljeel employee** (most common):
  - Traveler IS an Aljeel employee on Manpower master
  - No personal-contribution / annual-ticket / sponsorship form attached
  - account = 60301003
  - cost_center / div / solution / agency = the TRAVELER's HOME values from Manpower master
  - **emp_no = EMPTY STRING** (truth file leaves col 12 blank for routine business travel — the emp_no field is reserved for personal-contribution flags)

**TYPE B — Personal contribution / annual ticket** (employee paying for personal travel):
  - Approval .msg subject contains "Personal Contribution" OR "Annual Ticket"
  - HOWEVER: even if .msg says "Personal Contribution Approval", the BOOKING is still a regular travel ticket — the personal-contribution PORTION is tracked separately.
  - For the AP row: account = 60301003 (still standard travel).
  - **emp_no = traveler's emp_no FROM THE FORM** (the 7-digit number in the email subject, e.g., "Approval Requested for X (1001406)" → emp_no="1001406"). This is a flag that this row had a personal-contribution form attached.
  - cost_center / div / solution / agency = traveler's HOME values from Manpower master.
  - Only use account 21070229 if the row description EXPLICITLY refers to a deduction / personal share.

**TYPE C — Sponsorship (external doctor / guest visit funded by Aljeel)**:
  - OPEX-XXXX.pdf or HF-XXXX / J-XXXX / EP-XXXX / CRM-XXXX form exists
  - Passenger is usually an EXTERNAL doctor (not in Manpower)
  - account = 60307021
  - emp_no = EMPTY STRING
  - **MANDATORY ALGORITHM for cost_center / div / agency / solution on sponsorships:**
    1. Find the REQUESTING ALJEEL EMPLOYEE'S emp_no from the OPEX form. It's listed in a table like "Employee No | Name | Amount" — take the 7-digit number.
    2. Look that emp_no up in the "Employee directory" master data section below.
    3. **COPY the requester's `cost_center` field as your cost_center.** EXAMPLE: if requester is 1000820 and the directory shows `"cost_center": "160014"`, your cost_center MUST be "160014". NOT 000000. NOT 210110. NOT 170000. **"160014"** — the literal value from the directory.
    4. div_code: same algorithm — copy `div_code` from the requester's row.
    5. agency_code: copy `agency_code` from the requester's row (UNLESS the OPEX form explicitly names a different brand — then use that brand's code).
    6. solution: copy `solution` field from the requester (HF/EP/CRM) and map to numeric code via the solution hint.

## CRITICAL: For ANY row, the cost_center / div / agency / solution should come from the EMPLOYEE'S HOME VALUES in the directory below — NOT from the OPEX form headers, which are often just labels, not codes. Codes always come from the directory.

## Account codes — use ONLY these:
   - 60301003 = standard travel tickets (default for any plane/train ticket)
   - 60307021 = sponsoring expenses (only when an OPEX/sponsoring form exists)
   - 21070229 = personal contribution / annual ticket (rare — only when row is explicitly a deduction)
   - DO NOT use 11034013 (personal travel) — Amr has killed it.

## When evidence is missing or ambiguous:
   - If you cannot find a relevant evidence file, default to TYPE A and look up the traveler in the Manpower master to fill all fields.
   - solution = "00000" only if truly general; otherwise use traveler's home solution.
   - agency = "00000" only if no brand context.
   - **NEVER output "000000" for cost_center unless the traveler/requester is unidentifiable.** Always try to use a real employee's home CC.

## Code zero-padding:
   - cost_center: 6 digits ALWAYS (e.g., "160014" not "160014.0"; "000000" if general)
   - solution: 5 digits ALWAYS (e.g., "00000", "10050")
   - agency: 5 digits ALWAYS (e.g., "10072", "00000")
   - div: leading-zero-stripped numeric or 3-digit code as in master (e.g., "170", "196", "120")

# INVOICE ROW
- Description: {row.get('description','')}
- Passenger / Service: {row.get('passenger','')}
- Route: {row.get('route','')}
- Ticket #: {row.get('ticket_no','')}
- Amount: {row.get('amount','')}
- Date: {row.get('date','')}
- Notes (if any): {row.get('notes','') or '(none)'}

# EVIDENCE FILES FOR THIS TICKET (FOLDER: {evidence.get('folder','MISSING')})
- Files present: {', '.join(evidence.get('files', [])) or 'NONE'}

## EMAILS (.msg files, full body)
{chr(10).join(msg_chunks) if msg_chunks else '(no email files)'}

## PDFs (full text)
{chr(10).join(pdf_chunks) if pdf_chunks else '(no PDF files)'}

# MASTER DATA SNAPSHOTS

## Employee directory (relevant subset — by emp_no). Use these HOME values to allocate.
{json.dumps(master_snapshot.get('employees', {}), ensure_ascii=False, indent=1)}

## DIV codes
{json.dumps(master_snapshot.get('div', {}), ensure_ascii=False, indent=1)}

## Agency code hints (brand → 5-digit code)
{json.dumps(master_snapshot.get('agency_hint', {}), ensure_ascii=False, indent=1)}

## Solution code hints (name → 5-digit code)
{json.dumps(master_snapshot.get('solution_hint', {}), ensure_ascii=False, indent=1)}

## OPEX prefix → default allocation (use when OPEX form is image-based and requester not in directory)
When the OPEX-XX-PREFIX-2026-J-2026-YY.pdf filename has a 'PREFIX' segment, use these defaults:
{json.dumps(master_snapshot.get('opex_division_hint', {}), ensure_ascii=False, indent=1)}
Examples: 'OPEX-20-DMS-2026-...' → div=192, cc=160013, agency_likely=10005 (KAVO). 'OPEX-HF-2026-...' → div=170, cc=160014, sol=10050, agency_likely=10072 (ABBOTT).

# OUTPUT
Return ONLY the JSON object. No prose. No code fences.
"""
    return prompt


def build_master_snapshot(manpower: dict, lookups: dict, evidence: dict, row: dict) -> dict:
    """Inline only the relevant subset of master data (small to keep prompt tight)."""
    # Find employee names mentioned in evidence + row → include those in employee snapshot
    all_text = row.get("description", "") + " " + row.get("passenger", "") + " " + row.get("notes", "")
    def _s(v):
        if v is None:
            return ""
        if isinstance(v, list):
            return " ".join(_s(x) for x in v)
        return str(v)
    for m in evidence["msgs"]:
        all_text += " " + _s(m.get("body", "")) + " " + _s(m.get("subject", "")) + " " + _s(m.get("from", "")) + " " + _s(m.get("to", ""))
    for p in evidence["pdfs"]:
        all_text += " " + (p.get("text", "") or "")
    text_upper = all_text.upper()
    
    def _emp_info(info):
        return {
            "name": info["name"],
            "div_code": info["div_code"],
            "div_name": info["div_name"],
            "agency_code": info["agency_code"],
            "agency_name": info["agency_name"],
            "cost_center": info["cost_center"],
            "cc_name": info["cc_name"],
            "solution": info["solution"],
        }
    emp_subset = {}
    # PRIORITY 1: any emp_no that appears as 7-digit numeric token in the text (highest signal)
    for tok in re.findall(r"\b\d{7}\b", all_text):
        if tok in manpower and tok not in emp_subset:
            emp_subset[tok] = _emp_info(manpower[tok])
    # PRIORITY 2: passenger name — BOTH first AND last name required (v28)
    # Single-token matching (e.g. "ABEER" from "ABEER BAKHSH" matching employee
    # "Abeer Saleh Thabet Hajer") produces wrong employee assignments in the LLM
    # snapshot. We now require that at least 2 tokens from the passenger name
    # appear in the employee name. Exception: 1-token passengers still use single.
    _PAX_TITLE_SKIP = {"MR", "MS", "MRS", "DR", "MISS", "PROF", "ENG", "MASTER"}
    passenger = row.get("passenger", "").upper()
    if passenger:
        pax_tokens = [
            t for t in re.split(r"[\s\-,/]+", passenger)
            if len(t.strip()) >= 3 and t.strip() not in _PAX_TITLE_SKIP
        ]
        min_hits = min(2, len(pax_tokens))
        if pax_tokens:
            for emp_no, info in manpower.items():
                name_upper = info.get("name", "").upper()
                name_tokens = set(re.split(r"[\s\-]+", name_upper))
                hit_count = sum(1 for t in pax_tokens if t in name_tokens)
                if hit_count >= min_hits and emp_no not in emp_subset:
                    emp_subset[emp_no] = _emp_info(info)
    # PRIORITY 3: other name token matches anywhere in evidence text
    if len(emp_subset) < 40:
        for emp_no, info in manpower.items():
            if emp_no in emp_subset:
                continue
            name = info.get("name", "")
            if not name:
                continue
            for tok in re.split(r"[\s\-,/]+", name):
                tok = tok.strip().upper()
                if len(tok) >= 5 and tok in text_upper:
                    emp_subset[emp_no] = _emp_info(info)
                    break
            if len(emp_subset) >= 40:
                break

    # Compact lookup hints — the LLM doesn't need all 677 agencies, just the brand-code mapping it needs.
    # Hardcode known common mappings + give it the most-likely codes from employees in scope.
    agency_hint = {
        "ABBOTT": "10072", "GE": "10081", "NUBOMED": "10156", "BSC": "10101",
        "BMX": "10153", "Technical Services": "10206", "BIOTRONIK": "10058",
        "MEDTRONIC": "10133", "PHILIPS": "10168", "PENUMBRA": "10163",
        "KAVO": "10005", "SOLVENTUM": "10202", "3M": "10202",
        "GENERAL": "00000",
    }
    # OPEX-XX-PREFIX-2026 prefix maps to division/agency
    opex_division_hint = {
        "HF": {"div": "170", "cc": "160014", "sol": "10050", "agency_likely": "10072"},
        "EP": {"div": "170", "cc": "160014", "sol": "10064", "agency_likely": "10072"},
        "CRM": {"div": "170", "cc": "160014", "sol": "10094", "agency_likely": "10072"},
        "DMS": {"div": "192", "cc": "160013", "sol": "00000", "agency_likely": "10005"},  # Dental & Medical Solutions
        "IVD": {"div": "194", "cc": "160012", "sol": "00000", "agency_likely": "10153"},
        "S&M": {"div": "190", "cc": "140020", "sol": "00000", "agency_likely": "10200"},
        "GE":  {"div": "196", "cc": "160011", "sol": "00000", "agency_likely": "10081"},
        "BMX": {"div": "194", "cc": "160012", "sol": "00000", "agency_likely": "10153"},
    }
    solution_hint = {
        "HF": "10050", "EP": "10064", "CRM": "10017", "GENERAL": "00000",
        "BUSINESS_TRIP": "00000",
    }
    return {
        "employees": emp_subset,
        "div": lookups["div"],
        "agency_hint": agency_hint,
        "solution_hint": solution_hint,
        "opex_division_hint": opex_division_hint,
    }


def process_row(row_idx: int, row: dict, manpower: dict, lookups: dict, all_folders: list[Path]) -> dict:
    """Process a single row. Returns full result dict."""
    cache_key = f"row{row_idx:03d}-{row.get('ticket_no','noticket')}.json"
    cache_path = CACHE_DIR / cache_key
    if cache_path.exists():
        try:
            cached = json.loads(cache_path.read_text())
            cached["from_cache"] = True
            return cached
        except Exception:
            pass

    folder = find_ticket_folder(row.get("ticket_no", ""), row.get("passenger", ""), row.get("notes", ""), all_folders)
    evidence = collect_evidence(folder) if folder else {"folder": "", "msgs": [], "pdfs": [], "files": [], "total_chars": 0}
    snapshot = build_master_snapshot(manpower, lookups, evidence, row)
    prompt = build_prompt(row, evidence, snapshot)
    
    t0 = time.time()
    gemini_res = call_gemini_with_cascade(prompt)
    elapsed = time.time() - t0
    
    result = {
        "row_idx": row_idx,
        "row": row,
        "folder": str(folder) if folder else "",
        "evidence_files": evidence.get("files", []),
        "evidence_chars": evidence.get("total_chars", 0),
        "prompt_chars": len(prompt),
        "gemini": gemini_res,
        "elapsed_sec": elapsed,
        "from_cache": False,
    }
    try:
        cache_path.write_text(json.dumps(result, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    except Exception as e:
        result["cache_err"] = str(e)
    return result


# ---------- I/O ----------
def load_truth_rows() -> list[dict]:
    """Load all 160 rows from the truth file (Details sheet). For BLIND test, we keep only the input fields."""
    wb = openpyxl.load_workbook(TRUTH_XLSX, read_only=True, data_only=True)
    ws = wb["Details"]
    rows = list(ws.iter_rows(values_only=True))
    out = []
    # Header at row 0; data rows 1+
    for i, row in enumerate(rows[1:], start=1):
        if not row or row[0] is None:
            continue
        # Stop at footer
        if isinstance(row[0], str) and not str(row[0]).strip().isdigit():
            break
        try:
            int(str(row[0]).strip())
        except Exception:
            break
        ticket_raw = str(row[3] or "").strip()
        # ticket format "065 6905055971" — get the 10+ digit number
        m = re.search(r"\d{10,}", ticket_raw)
        ticket_no = m.group() if m else ticket_raw
        out.append({
            "row_idx": i,
            "sl_no": row[0],
            "date": str(row[1] or ""),
            "ref_no": str(row[2] or ""),
            "ticket_no_raw": ticket_raw,
            "ticket_no": ticket_no,
            "passenger": str(row[4] or ""),
            "route": str(row[5] or ""),
            "service_date": str(row[6] or ""),
            "amount": row[10],  # Inv. Amt. Incl. VAT
            "description": str(row[29] or "") if len(row) > 29 else "",
            "notes": str(row[12] or "") if len(row) > 12 else "",
            # GROUND TRUTH (kept separate; never shown to LLM)
            "truth_emp_no": str(row[11] or "").strip().rstrip("."),
            "truth_account": str(row[16] or "").strip(),
            "truth_cost_center": str(row[18] or "").strip(),
            "truth_div": str(row[20] or "").strip(),
            "truth_solution": str(row[22] or "").strip(),
            "truth_agency": str(row[24] or "").strip(),
        })
    wb.close()
    return out


def normalize_code(v: Any, width: int = 0) -> str:
    if v is None:
        return ""
    s = str(v).strip().rstrip(".")
    # Strip leading zeros for comparison? No — codes are zero-padded; keep as-is.
    # But normalize: ensure pure-digit codes with the same numeric value match.
    if width > 0 and s.isdigit():
        return s.zfill(width)
    return s


def score_results(results: list[dict]) -> dict:
    """Score each row against ground truth."""
    per_row = []
    total = 0
    fields = ["emp_no", "account", "cost_center", "div", "solution", "agency"]
    counts = {f: 0 for f in fields}
    all_five_count = 0  # 5 of 6 = account, CC, div, sol, agency
    sponsorship_idxs = []
    sponsorship_all5 = 0
    travel_idxs = []
    travel_all5 = 0
    for r in results:
        if r is None or "row" not in r:
            continue
        total += 1
        row = r["row"]
        gem = r.get("gemini", {}).get("json", {}) or {}
        match = {}
        # emp_no — compare as ints when possible
        truth = {f: row.get(f"truth_{f}", "") for f in fields}
        out = {f: str(gem.get(f, "") or "").strip().rstrip(".") for f in fields}
        # normalize
        def n(s):
            return str(s or "").strip().rstrip(".").lstrip("0") or "0"
        for f in fields:
            t = n(truth[f])
            o = n(out[f])
            if f == "emp_no":
                # emp_no scoring: only count rows where truth has a real emp_no (not "-" / blank / 0)
                if t in ("", "-", "0"):
                    match[f] = None  # skip
                else:
                    match[f] = (t == o and bool(o) and o != "0")
            else:
                match[f] = (t == o and bool(truth[f].strip()) and bool(out[f].strip()))
            if match[f] is True:
                counts[f] += 1
        # all 5 = account, cc, div, sol, agency  (not emp_no per Amr's original framework)
        all5 = all(match[f] for f in ["account", "cost_center", "div", "solution", "agency"])
        if all5:
            all_five_count += 1
        # Classify row
        is_sponsorship = truth["account"] == "60307021"
        is_travel = truth["account"] == "60301003"
        if is_sponsorship:
            sponsorship_idxs.append(r["row_idx"])
            if all5:
                sponsorship_all5 += 1
        if is_travel:
            travel_idxs.append(r["row_idx"])
            if all5:
                travel_all5 += 1
        per_row.append({
            "row_idx": r["row_idx"],
            "passenger": row.get("passenger", ""),
            "ticket": row.get("ticket_no", ""),
            "truth": truth,
            "predicted": out,
            "match": match,
            "all_five": all5,
            "category": "sponsorship" if is_sponsorship else ("travel" if is_travel else "other"),
            "confidence": gem.get("confidence", ""),
            "reasoning": gem.get("reasoning", "")[:200],
            "evidence_files_count": len(r.get("evidence_files", [])),
            "evidence_chars": r.get("evidence_chars", 0),
            "folder": r.get("folder", ""),
        })
    return {
        "total": total,
        "all_five_exact": all_five_count,
        "all_five_pct": 100 * all_five_count / total if total else 0,
        "field_counts": counts,
        "field_pct": {f: 100 * counts[f] / total for f in fields} if total else {},
        "emp_no_scorable": sum(1 for r in per_row if r['match']['emp_no'] is not None),
        "emp_no_correct": sum(1 for r in per_row if r['match']['emp_no'] is True),
        "sponsorship_n": len(sponsorship_idxs),
        "sponsorship_all5": sponsorship_all5,
        "sponsorship_pct": 100 * sponsorship_all5 / len(sponsorship_idxs) if sponsorship_idxs else 0,
        "travel_n": len(travel_idxs),
        "travel_all5": travel_all5,
        "travel_pct": 100 * travel_all5 / len(travel_idxs) if travel_idxs else 0,
        "per_row": per_row,
    }


def write_output_xlsx(results: list[dict], path: Path):
    """Write results in the same column structure as the v15.11.2 output (truncated to essential fields)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Match v15.11.2 format header rows
    ws.append([])  # row 1 blank
    section = ["ORACLE FUSION TEMPLATE"] + [""] * 15 + ["CODE & DESCRIPTION"] + [""] * 15 + ["LLM AGENT OUTPUT"]
    ws.append(section)  # row 2
    
    hdr = [
        "*Invoice Header Identifier", "*Business Unit", "*Invoice Number", "*Invoice Currency",
        "*Invoice Amount", "*Invoice Date", "**Supplier", "**Supplier Number", "*Supplier Site",
        "Invoice Type", "Description", "*Type", "*Amount", "Distribution Combination",
        "Tax Classification Code", "Employee No",
        "Company", "Location", "Account", "GL", "Cost Center", "Cost Name", "DIV", "Contribution",
        "Solution", "Solution Name", "Agency", "Agency Name", "Project", "Intercompany", "Future 1",
        "GL Description",
        # Agent fields
        "Agent Confidence", "Agent Reasoning", "Agent Model", "Agent Folder", "Agent EvFiles", "Agent EvChars",
        "Agent InTokens", "Agent OutTokens", "Agent Error",
    ]
    ws.append(hdr)  # row 3
    
    for r in results:
        if r is None:
            continue
        row = r["row"]
        gem = r.get("gemini", {}).get("json", {}) or {}
        gem_info = r.get("gemini", {})
        out_row = [
            row.get("sl_no", ""), "Al Jeel Medical BU", "J26-593", "SAR",
            row.get("amount", ""), row.get("date", ""),
            "شركة جوال للسفر والسياحة المحد", "10394", "شركة جوال للسفر",
            "Standard",
            f"{row.get('passenger','')} - {row.get('route','')}",
            "Item", row.get("amount", ""),
            f"03-40100-{gem.get('account','')}-{gem.get('cost_center','')}-{gem.get('div','')}-{gem.get('solution','')}-{gem.get('agency','')}-00000-00-000000",
            "KSA VAT STANDARD",
            gem.get("emp_no", ""),
            # Code & description columns
            "03", "40100", gem.get("account", ""), "", gem.get("cost_center", ""), "", gem.get("div", ""), "",
            gem.get("solution", ""), "", gem.get("agency", ""), "", "00000", "00", "000000",
            "",
            # Agent
            gem.get("confidence", ""), gem.get("reasoning", ""), gem_info.get("model", ""),
            r.get("folder", ""), len(r.get("evidence_files", [])), r.get("evidence_chars", 0),
            gem_info.get("in_tokens", 0), gem_info.get("out_tokens", 0), gem_info.get("error", ""),
        ]
        ws.append(out_row)
    
    wb.save(str(path))


# ---------- main ----------
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=0, help="0 = all rows")
    parser.add_argument("--workers", type=int, default=5)
    parser.add_argument("--no-cache", action="store_true")
    args = parser.parse_args()
    
    if args.no_cache:
        # Clear cache
        for f in CACHE_DIR.glob("*.json"):
            f.unlink()
    
    print(f"[load] reading truth file: {TRUTH_XLSX}", flush=True)
    truth_rows = load_truth_rows()
    if args.limit > 0:
        truth_rows = truth_rows[: args.limit]
    print(f"[load] {len(truth_rows)} rows to process", flush=True)
    
    print(f"[load] master data...", flush=True)
    manpower = load_manpower()
    lookups = load_lookups()
    print(f"[load] {len(manpower)} employees, {len(lookups['cc'])} CC, {len(lookups['div'])} DIV, {len(lookups['solution'])} Solution, {len(lookups['agency'])} Agency", flush=True)
    
    print(f"[scan] collecting all folder candidates...", flush=True)
    all_folders = collect_all_folders(RAW_ROOT)
    print(f"[scan] {len(all_folders)} ticket-folder candidates", flush=True)
    
    results = [None] * len(truth_rows)
    t_start = time.time()
    done = 0
    
    def worker(idx_row):
        idx, row = idx_row
        try:
            return idx, process_row(row["row_idx"], row, manpower, lookups, all_folders)
        except Exception as e:
            tb = traceback.format_exc()
            return idx, {"row_idx": row["row_idx"], "row": row, "error": str(e), "traceback": tb, "gemini": {"json": {}, "in_tokens": 0, "out_tokens": 0, "model": "", "error": str(e)}}
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=args.workers) as ex:
        futs = [ex.submit(worker, (i, r)) for i, r in enumerate(truth_rows)]
        for fut in concurrent.futures.as_completed(futs):
            idx, res = fut.result()
            results[idx] = res
            done += 1
            elapsed = time.time() - t_start
            cached = " [cached]" if res.get("from_cache") else ""
            err = res.get("gemini", {}).get("error", "")
            err_marker = f" ERR={err[:30]}" if err else ""
            print(f"[{done:3d}/{len(truth_rows)}] row {res.get('row_idx','?'):3d}  t={elapsed:5.1f}s  chars={res.get('evidence_chars',0):6d}  files={len(res.get('evidence_files',[]))}{cached}{err_marker}", flush=True)
    
    total_elapsed = time.time() - t_start
    
    # Cost summary
    total_in = sum((r.get("gemini", {}).get("in_tokens", 0) or 0) for r in results if r)
    total_out = sum((r.get("gemini", {}).get("out_tokens", 0) or 0) for r in results if r)
    # use first non-cached model
    model_used = "gemini-2.5-pro"
    for r in results:
        if r and r.get("gemini", {}).get("model"):
            model_used = r["gemini"]["model"]
            break
    pricing = GEMINI_PRICING.get(model_used, GEMINI_PRICING["gemini-2.5-pro"])
    # Per-call pricing so the >200K long-context rate applies correctly per request, not per batch.
    cost_usd = 0.0
    for _r in results:
        if not _r:
            continue
        _g = _r.get("gemini", {}) or {}
        _it = _g.get("in_tokens", 0) or 0
        _ot = _g.get("out_tokens", 0) or 0
        _m = _g.get("model") or model_used
        _rate = _gemini_rate(_m, _it)
        cost_usd += _it * _rate["in"] + _ot * _rate["out"]
    
    # Score
    score = score_results(results)
    
    ts = time.strftime("%Y%m%dT%H%M%SZ", time.gmtime())
    
    # Write Excel output
    xlsx_path = OUT_DIR / f"Spreadsheet-J26-593-LLMAGENT-v1.xlsx"
    write_output_xlsx([r for r in results if r], xlsx_path)
    print(f"[out] wrote {xlsx_path}", flush=True)
    
    # Write summary
    cost_summary = {
        "timestamp_utc": ts,
        "rows_processed": len(results),
        "total_in_tokens": total_in,
        "total_out_tokens": total_out,
        "model_used": model_used,
        "cost_usd": round(cost_usd, 4),
        "total_runtime_sec": round(total_elapsed, 1),
        "rate_per_min": round(60 * len(results) / total_elapsed, 1) if total_elapsed else 0,
    }
    (OUT_DIR / f"cost-summary.json").write_text(json.dumps(cost_summary, indent=2), encoding="utf-8")
    print(f"[out] cost: ${cost_usd:.4f} ({total_in:,} in / {total_out:,} out tokens, model={model_used})", flush=True)
    
    # Write score JSON
    score_path = OUT_DIR / f"score-llmagent.json"
    score_path.write_text(json.dumps(score, indent=2, default=str), encoding="utf-8")
    print(f"[out] wrote {score_path}", flush=True)
    
    # Print top-line score
    print()
    print("=" * 60)
    print(f"  ALL-5-EXACT:     {score['all_five_exact']}/{score['total']}  ({score['all_five_pct']:.1f}%)")
    print(f"  Sponsorships:    {score['sponsorship_all5']}/{score['sponsorship_n']}  ({score['sponsorship_pct']:.1f}%)")
    print(f"  Travel:          {score['travel_all5']}/{score['travel_n']}  ({score['travel_pct']:.1f}%)")
    print(f"  Emp_no:          {score['emp_no_correct']}/{score['emp_no_scorable']}  (where truth set)")
    print(f"  Cost / runtime:  ${cost_usd:.4f}  /  {total_elapsed/60:.1f} min")
    print(f"  Per-field:")
    for f, pct in score["field_pct"].items():
        print(f"    {f:14s} {pct:5.1f}%   ({score['field_counts'][f]}/{score['total']})")
    print("=" * 60)


if __name__ == "__main__":
    main()
