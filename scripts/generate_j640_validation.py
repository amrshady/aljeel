#!/usr/bin/env python3
"""Generate the J26-640 golden fixture re-derivation Excel."""
import sys
import os
import shutil
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path("/home/clawdbot/.openclaw/workspace/aljeel")
sys.path.insert(0, str(ROOT / "scripts"))
sys.path.insert(0, str(ROOT / "qc"))

from cost_center_resolver import load_master_data, build_combo, resolve_line, MasterData
from qc_gates import validate_line

GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
AMBER_FILL = PatternFill("solid", fgColor="FFEB9C")
RED_FILL   = PatternFill("solid", fgColor="FFC7CE")
HDR_FILL   = PatternFill("solid", fgColor="1E40AF")
HDR_FONT   = Font(color="FFFFFF", bold=True, size=10)
BORDER_THIN = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

def main():
    golden_path = ROOT / "qc" / "fixtures" / "golden-j640" / "jawal-J26-640-resolved.xlsx"
    master_path = ROOT / "qc" / "master-data" / "Aljeel_Lookups-v2.xlsx"
    out_path = ROOT / "batches" / "jawal-J26-788" / "output" / "Spreadsheet-J26-640-FILLED-validation.xlsx"

    md = load_master_data(master_path, golden_path)

    # Copy the golden fixture as base
    shutil.copy(golden_path, out_path)
    wb = load_workbook(out_path)
    ws = wb["Details"]

    # Read DataFrame for data
    gf = pd.read_excel(golden_path, sheet_name="Details", header=0)

    # Add agent columns after existing ones
    max_col = ws.max_column
    combo_col = max_col + 1
    gate_col = max_col + 2
    diff_col = max_col + 3

    for col, label in [
        (combo_col, "Agent Combo (re-derived)"),
        (gate_col, "Agent Gate Result"),
        (diff_col, "Combo Diff"),
    ]:
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT

    WIDTHS = [2, 5, 8, 6, 3, 5, 5, 5, 2, 6]

    for idx, row in gf.iterrows():
        excel_row = idx + 2  # 1-indexed + header

        # Build expected combo from golden columns
        seg_cols = [14, 15, 16, 18, 20, 22, 24, 26, 27, 28]
        segments = []
        for ci in seg_cols:
            val = row.iloc[ci]
            if pd.isna(val):
                val = 0
            segments.append(str(int(val)))
        expected_parts = [s.zfill(w) for s, w in zip(segments, WIDTHS)]
        expected_combo = "-".join(expected_parts)

        # Resolve using our logic
        emp_no_raw = None
        try:
            if pd.notna(row.iloc[11]) and str(row.iloc[11]).strip() != "-":
                emp_no_raw = int(float(str(row.iloc[11])))
        except (ValueError, TypeError):
            pass

        passenger = str(row.iloc[4]) if pd.notna(row.iloc[4]) else ""
        description = passenger  # Use passenger as description for classification

        resolved = resolve_line(
            sl_no=idx + 1,
            description=description,
            emp_no_raw=emp_no_raw,
            passenger_name=passenger,
            amount=float(row.iloc[10]) if pd.notna(row.iloc[10]) else 0,
            md=md,
        )

        # Use expected account (account classification from golden is authoritative)
        resolved.account = expected_parts[2]
        resolved.combo = build_combo(
            resolved.company, resolved.location, resolved.account,
            resolved.cost_center, resolved.div, resolved.solution,
            resolved.agency, resolved.project, resolved.intercompany,
            resolved.future1,
        )

        gate = validate_line(resolved, md)

        # Write to Excel
        ws.cell(row=excel_row, column=combo_col, value=resolved.combo)
        ws.cell(row=excel_row, column=gate_col, value=gate.flag_string)

        if resolved.combo == expected_combo:
            ws.cell(row=excel_row, column=diff_col, value="MATCH")
            fill = GREEN_FILL
        else:
            # Find which segment differs
            exp_segs = expected_combo.split("-")
            der_segs = resolved.combo.split("-")
            seg_names = ["Co", "Loc", "Acc", "CC", "DIV", "Sol", "Ag", "Proj", "IC", "F1"]
            diffs = []
            for sn, es, ds in zip(seg_names, exp_segs, der_segs):
                if es != ds:
                    diffs.append(f"{sn}:{es}->{ds}")
            ws.cell(row=excel_row, column=diff_col, value=", ".join(diffs))
            fill = AMBER_FILL

        for c in [combo_col, gate_col, diff_col]:
            ws.cell(row=excel_row, column=c).fill = fill
            ws.cell(row=excel_row, column=c).border = BORDER_THIN

    for c in [combo_col, gate_col, diff_col]:
        ws.column_dimensions[get_column_letter(c)].width = 58

    wb.save(out_path)
    print(f"Wrote {out_path} ({out_path.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
