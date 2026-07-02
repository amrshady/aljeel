#!/usr/bin/env python3
"""
Golden Field Diff - compares pipeline output against the human-resolved golden.

The golden is the Details sheet in jawal-J26-640-resolved.xlsx with columns:
  Col 14(O): Company, Col 15(P): Location, Col 16(Q): Account,
  Col 18(S): CostCenter, Col 20(U): DIV, Col 22(W): Solution, Col 24(Y): Agency

Pipeline output is the Spreadsheet-J26-640-FILLED-*.xlsx (Oracle template format)
with the 10-segment combo in col N (index 13).

Usage:
    python3 scripts/golden_field_diff.py [--pipeline PATH] [--version LABEL]
"""
import argparse
import json
import sys
from pathlib import Path
from collections import Counter
import openpyxl
import pandas as pd

ROOT = Path("/home/clawdbot/.openclaw/workspace/aljeel")
GOLDEN = ROOT / "qc" / "fixtures" / "golden-j640" / "jawal-J26-640-resolved.xlsx"

SEGMENT_NAMES = ["Company", "Location", "Account", "CostCenter", "DIV", "Solution", "Agency"]
SEGMENT_WIDTHS = {"Company": 2, "Location": 5, "Account": 8, "CostCenter": 6, "DIV": 3, "Solution": 5, "Agency": 5}


def load_golden():
    """Load golden segments from the Details sheet."""
    wb = openpyxl.load_workbook(GOLDEN, data_only=True, read_only=True)
    ws = wb["Details"]
    lines = []
    for r in range(2, ws.max_row + 1):
        sl = ws.cell(r, 1).value
        if sl is None:
            continue
        lines.append({
            "sl": int(sl) if sl else 0,
            "Company": str(ws.cell(r, 15).value or "").strip(),
            "Location": str(ws.cell(r, 16).value or "").strip(),
            "Account": str(ws.cell(r, 17).value or "").strip(),
            "CostCenter": str(ws.cell(r, 19).value or "").strip(),
            "DIV": str(ws.cell(r, 21).value or "").strip(),
            "Solution": str(ws.cell(r, 23).value or "").strip(),
            "Agency": str(ws.cell(r, 25).value or "").strip(),
        })
    wb.close()
    return lines


def load_pipeline(pipeline_path):
    """Load pipeline combo segments from Oracle template output."""
    df = pd.read_excel(pipeline_path, sheet_name=0, header=None)
    header_row = None
    for i in range(10):
        if "Invoice Header Identifier" in str(df.iloc[i, 0]):
            header_row = i
            break
    if header_row is None:
        raise ValueError("No header row found")

    lines = []
    for i in range(header_row + 1, len(df)):
        if pd.isna(df.iloc[i, 0]):
            continue
        combo = str(df.iloc[i, 13]) if pd.notna(df.iloc[i, 13]) else ""
        parts = combo.split("-") if "-" in combo else []
        if len(parts) >= 7:
            lines.append({
                "Company": parts[0],
                "Location": parts[1],
                "Account": parts[2],
                "CostCenter": parts[3],
                "DIV": parts[4],
                "Solution": parts[5],
                "Agency": parts[6],
            })
        else:
            lines.append({s: "" for s in SEGMENT_NAMES})
    return lines


def compare(golden, pipeline, version_label=""):
    """Compare golden vs pipeline, return per-segment stats."""
    n = min(len(golden), len(pipeline))
    seg_match = {s: 0 for s in SEGMENT_NAMES}
    full_match = 0
    mismatches = []

    for i in range(n):
        g = golden[i]
        p = pipeline[i]
        all_ok = True
        line_diffs = {}
        for seg in SEGMENT_NAMES:
            w = SEGMENT_WIDTHS[seg]
            gv = g[seg].zfill(w) if g[seg] else ""
            pv = p[seg]
            if gv == pv:
                seg_match[seg] += 1
            else:
                all_ok = False
                line_diffs[seg] = pv + "->" + gv
        if all_ok:
            full_match += 1
        elif line_diffs:
            mismatches.append({"line": i + 1, "sl": g.get("sl", i + 1), "diffs": line_diffs})

    result = {"version": version_label, "compared": n}
    for seg in SEGMENT_NAMES:
        pct = round(100 * seg_match[seg] / max(n, 1), 1)
        result[seg] = {"match": seg_match[seg], "total": n, "pct": pct}
    result["FullCombo"] = {"match": full_match, "total": n, "pct": round(100 * full_match / max(n, 1), 1)}
    result["mismatches"] = mismatches
    return result


def print_summary(result):
    sep = "=" * 60
    print()
    print(sep)
    print("Golden Field Diff -- " + result["version"])
    print(sep)
    print("Lines compared: " + str(result["compared"]))
    print()
    print("{:<14} {:>6} / {:>6}  {:>7}".format("Segment", "Match", "Total", "%"))
    print("-" * 40)
    for seg in SEGMENT_NAMES + ["FullCombo"]:
        d = result[seg]
        print("{:<14} {:>6} / {:>6}  {:>6.1f}%".format(seg, d["match"], d["total"], d["pct"]))

    print()
    print("Mismatches ({} lines):".format(len(result["mismatches"])))
    for mm in result["mismatches"][:30]:
        diff_str = ", ".join(k + ":" + v for k, v in mm["diffs"].items())
        print("  Line {}: {}".format(mm["sl"], diff_str))
    if len(result["mismatches"]) > 30:
        print("  ... and {} more".format(len(result["mismatches"]) - 30))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--pipeline", help="Path to pipeline output xlsx")
    parser.add_argument("--version", default="", help="Version label")
    parser.add_argument("--json-out", help="Write results to JSON file")
    args = parser.parse_args()

    golden = load_golden()

    if args.pipeline:
        pipeline_path = Path(args.pipeline)
    else:
        candidates = sorted(
            (ROOT / "batches" / "jawal-J26-640" / "output").glob("Spreadsheet-J26-640-FILLED-v*.xlsx"),
            reverse=True,
        )
        if not candidates:
            print("No pipeline output found")
            sys.exit(1)
        pipeline_path = candidates[0]

    pipeline = load_pipeline(pipeline_path)
    version = args.version or pipeline_path.name

    result = compare(golden, pipeline, version)
    print_summary(result)

    if args.json_out:
        with open(args.json_out, "w") as f:
            json.dump(result, f, indent=2)
        print("\nJSON: " + args.json_out)


if __name__ == "__main__":
    main()
