#!/usr/bin/env python3
import sys
import json
import openpyxl
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def inject_fraud(batch_id):
    print(f"\n[inject_fraud] Starting AI Consistency Check injection for {batch_id}...")
    
    root = Path("/home/clawdbot/.openclaw/workspace/aljeel")
    
    # 1. Load JSON file
    json_path = root / "qc" / "ai-poc" / f"{batch_id.lower()}-fraud-ai-v162.json"
    if not json_path.exists():
        print(f"❌ JSON file not found: {json_path}")
        return False
        
    try:
        with open(json_path, "r") as f:
            data = json.load(f)
    except Exception as e:
        print(f"❌ Failed to parse JSON: {e}")
        return False
        
    # Map sl_no -> verdict dict.
    # Schema v16.2 nests verdicts under "ai_result"; older files had them top-level.
    _rows = []
    ai_result = data.get("ai_result")
    if isinstance(ai_result, dict) and isinstance(ai_result.get("per_row_verdicts"), list):
        _rows = ai_result["per_row_verdicts"]
    elif isinstance(data.get("per_row_verdicts"), list):
        _rows = data["per_row_verdicts"]
    verdicts = {}
    for item in _rows:
        try:
            verdicts[item["sl_no"]] = item
        except (KeyError, TypeError):
            continue
    print(f"[inject_fraud] Loaded {len(verdicts)} per-row verdicts")

    # 2. Load Excel file
    excel_path = root / "batches" / f"jawal-{batch_id}" / "output" / f"Spreadsheet-{batch_id}-FILLED-v30.xlsx"
    if not excel_path.exists():
        print(f"❌ Excel file not found: {excel_path}")
        return False
        
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
    except Exception as e:
        print(f"❌ Failed to load Excel: {e}")
        return False

    # Fills & Fonts
    GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
    AMBER_FILL = PatternFill("solid", fgColor="FFEB9C")
    RED_FILL   = PatternFill("solid", fgColor="FFC7CE")
    
    BLOCK3_HDR_FILL = PatternFill("solid", fgColor="1E40AF")
    HDR_FONT_DARK   = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    HDR_ALIGN       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    BORDER_THIN     = Border(
        left=Side(style="thin", color="D3D3D3"),
        right=Side(style="thin", color="D3D3D3"),
        top=Side(style="thin", color="D3D3D3"),
        bottom=Side(style="thin", color="D3D3D3")
    )

    # 3. Find end of header row (Row 3)
    header_row = 3
    last_col = ws.max_column
    
    new_cols = [
        "AI Consistency Verdict",
        "AI Consistency Category",
        "AI Consistency Notes"
    ]
    
    # Write new headers
    for i, col_name in enumerate(new_cols):
        col_idx = last_col + 1 + i
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.fill = BLOCK3_HDR_FILL
        cell.font = HDR_FONT_DARK
        cell.alignment = HDR_ALIGN
        
    # 4. Write data rows (starting at Row 4)
    DATA_START_ROW = 4
    for r in range(DATA_START_ROW, ws.max_row + 1):
        sl_no = r - DATA_START_ROW + 1   # 1-based sequential index matching consistency-check sl_no
        verdict_data = verdicts.get(sl_no, {})
        
        # Write values
        v_cell = ws.cell(row=r, column=last_col + 1, value=verdict_data.get("verdict", "CLEAN"))
        c_cell = ws.cell(row=r, column=last_col + 2, value=verdict_data.get("primary_category", "NONE"))
        r_cell = ws.cell(row=r, column=last_col + 3, value=verdict_data.get("reasoning", "No flags identified by AI auditor."))
        
        # Borders
        v_cell.border = BORDER_THIN
        c_cell.border = BORDER_THIN
        r_cell.border = BORDER_THIN
        
        # Style verdict cell
        v_str = str(v_cell.value).upper()
        if v_str == "RED":
            v_cell.fill = RED_FILL
        elif v_str == "YELLOW":
            v_cell.fill = AMBER_FILL
        elif v_str == "CLEAN":
            v_cell.fill = GREEN_FILL

    # 5. Set column widths
    ws.column_dimensions[openpyxl.utils.get_column_letter(last_col + 1)].width = 18
    ws.column_dimensions[openpyxl.utils.get_column_letter(last_col + 2)].width = 25
    ws.column_dimensions[openpyxl.utils.get_column_letter(last_col + 3)].width = 75

    # Save
    wb.save(excel_path)
    print(f"✅ Successfully injected AI Consistency Check data and saved: {excel_path}")
    return True

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python3 inject_fraud_to_excel.py <batch_id>")
        sys.exit(1)
    inject_fraud(sys.argv[1])
