#!/usr/bin/env python3
"""Pre-flight evidence scan for Jawal AP batches."""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path("/home/clawdbot/.openclaw/workspace/aljeel")
VOLUME_BASE = Path("/mnt/aljeel_ap_kb/current")
TICKET_RE = re.compile(r"\b(\d{10})\b")
TICKET_SCAN_RE = re.compile(r"\b(\d{10})\b")
SHORT_REF_SCAN_RE = re.compile(r"\b(\d{2}-\d{3,})\b")
PNR_SCAN_RE = re.compile(r"(?<![A-Z0-9])([A-Z0-9]{6})(?![A-Z0-9])", re.IGNORECASE)
TRAILING_REF_RE = re.compile(r"\(([^)]+)\)\s*$")
INVOICE_BASENAME_RE = re.compile(r"INV(?:OICE)?", re.IGNORECASE)


def normalize_reference_token(token: str) -> str:
    token = str(token or "").strip()
    if re.fullmatch(r"\d{10}", token):
        return token
    if re.fullmatch(r"\d{2}-\d{3,}", token):
        return token
    if (
        re.fullmatch(r"[A-Z0-9]{6}", token, re.IGNORECASE)
        and re.search(r"[A-Z]", token, re.IGNORECASE)
        and re.search(r"\d", token)
    ):
        return token.upper()
    return ""


def trailing_description_reference(description: str) -> str:
    match = TRAILING_REF_RE.search(str(description or ""))
    return normalize_reference_token(match.group(1)) if match else ""


def reference_tokens_in_text(text: str) -> set[str]:
    refs = {match.group(1) for match in TICKET_SCAN_RE.finditer(text)}
    refs.update(match.group(1) for match in SHORT_REF_SCAN_RE.finditer(text))
    for match in PNR_SCAN_RE.finditer(text):
        token = normalize_reference_token(match.group(1))
        if token:
            refs.add(token)
    return refs


def normalize_batch_id(batch_id: str) -> str:
    match = re.search(r"J26-\d+", str(batch_id or ""), re.IGNORECASE)
    return match.group(0).upper() if match else str(batch_id or "").strip().upper()


def discover_evidence_dir(batch_id: str, raw_dir: str | None = None) -> Path:
    if raw_dir:
        candidate = Path(raw_dir).expanduser().resolve(strict=False)
        if candidate.is_dir():
            return candidate
        raise FileNotFoundError(f"raw dir not found: {candidate}")

    normalized = normalize_batch_id(batch_id)
    volume_candidate = VOLUME_BASE / normalized
    if volume_candidate.is_dir():
        return volume_candidate

    local_candidate = ROOT / "batches" / f"jawal-{normalized}" / "raw"
    if local_candidate.is_dir():
        return local_candidate

    raise FileNotFoundError(
        f"evidence dir not found for {normalized}: checked {volume_candidate} and {local_candidate}"
    )


def emit(prefix: str, payload: dict | None = None) -> None:
    if payload is None:
        print(prefix, flush=True)
        return
    print(f"{prefix} {json.dumps(payload, ensure_ascii=False)}", flush=True)


def relative_path(path: Path, root: Path) -> str:
    try:
        return path.relative_to(root).as_posix()
    except ValueError:
        return path.as_posix()


def nearest_folder(path: Path, root: Path) -> str:
    ticket_pat = re.compile(r"^\d{10}$")
    for parent in [path.parent, *path.parents]:
        if parent == root.parent:
            break
        if ticket_pat.fullmatch(parent.name):
            return parent.name
    try:
        rel_parent = path.parent.relative_to(root)
    except ValueError:
        return path.parent.name
    return rel_parent.parts[-1] if rel_parent.parts else ""


def scan_files(evidence_dir: Path) -> tuple[list[dict], list[dict], int, int]:
    zero_byte: list[dict] = []
    unreadable: list[dict] = []
    total_files = 0
    total_bytes = 0

    for dirpath, _dirnames, filenames in os.walk(evidence_dir):
        for filename in filenames:
            path = Path(dirpath) / filename
            rel = relative_path(path, evidence_dir)
            try:
                size = os.path.getsize(path)
            except OSError as exc:
                unreadable.append({"path": rel, "error": str(exc)})
                continue

            total_files += 1
            total_bytes += size

            if size == 0:
                zero_byte.append({
                    "path": rel,
                    "folder": nearest_folder(path, evidence_dir),
                })

            if not os.access(path, os.R_OK):
                unreadable.append({"path": rel, "error": "Permission denied"})
                continue

            try:
                with open(path, "rb") as handle:
                    handle.read(1)
            except OSError as exc:
                unreadable.append({"path": rel, "error": str(exc)})

    return zero_byte, unreadable, total_files, total_bytes


def read_invoice_tickets(batch_id: str) -> tuple[list[dict], str | None]:
    normalized = normalize_batch_id(batch_id)
    invoice_path = ROOT / "batches" / f"jawal-{normalized}" / "Spreadsheet-v4-input.xlsx"
    if not invoice_path.exists():
        return [], f"input xlsx not found: {invoice_path}"

    workbook = load_workbook(invoice_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        description_idx = None
        header_row = None
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            values = [str(cell or "").strip() for cell in row]
            for idx, value in enumerate(values):
                if value.lower() == "description":
                    description_idx = idx
                    header_row = row_idx
                    break
            if description_idx is not None:
                break

        if description_idx is None or header_row is None:
            return [], "Description column not found in input xlsx"

        tickets: dict[str, dict] = {}
        for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            description = str(row[description_idx] or "").strip() if description_idx < len(row) else ""
            ref_token = trailing_description_reference(description)
            if not ref_token:
                continue
            passenger = description.split(" - ", 1)[0].strip() or description
            tickets.setdefault(ref_token, {"ticket_no": ref_token, "passenger": passenger})
        return list(tickets.values()), None
    finally:
        workbook.close()


def collect_ticket_folder_names(evidence_dir: Path) -> set[str]:
    """Return supported Jawal refs found anywhere in the evidence tree.

    Matches:
    - Folder names that contain a ref: "6905992249 family", "26-765", "ELDS5J"
    - File basenames that contain a ref: "6905929962.pdf", "YOUSEF_AL DIGHRIR_O6IV3Y.pdf"
    - Nested, non-invoice PDF bodies that contain one of the supported refs
    """
    ticket_numbers: set[str] = set()
    pdf_text_cache: dict[Path, set[str]] = {}
    invoice_skip_paths, invoice_skip_stems = invoice_pdf_skip_set(evidence_dir)
    for root, dirs, files in os.walk(evidence_dir):
        # Check folder name
        basename = os.path.basename(root)
        ticket_numbers.update(reference_tokens_in_text(basename))
        # Check file names
        root_path = Path(root)
        for fname in files:
            file_path = root_path / fname
            if (
                file_path.resolve(strict=False) in invoice_skip_paths
                or file_path.stem.lower() in invoice_skip_stems
                or INVOICE_BASENAME_RE.search(file_path.stem)
            ):
                continue
            ticket_numbers.update(reference_tokens_in_text(file_path.stem))
        # Check body text in nested non-invoice PDFs only.
        for fname in files:
            if not fname.lower().endswith(".pdf"):
                continue
            pdf_path = root_path / fname
            if not should_scan_ticket_body_pdf(
                pdf_path, evidence_dir, invoice_skip_paths, invoice_skip_stems
            ):
                continue
            ticket_numbers.update(pdf_ticket_body_numbers(pdf_path, pdf_text_cache))
    return ticket_numbers


def batch_id_from_path(path: Path) -> str:
    for part in path.parts:
        match = re.search(r"J26-\d+", part, re.IGNORECASE)
        if match:
            return match.group(0).upper()
    return ""


def invoice_pdf_skip_set(evidence_root: Path) -> tuple[set[Path], set[str]]:
    """Known source invoice PDFs/stems that must never seed ticket evidence."""
    skip_paths: set[Path] = set()
    skip_stems: set[str] = set()
    if not evidence_root:
        return skip_paths, skip_stems

    batch_id = batch_id_from_path(evidence_root)
    candidate_dirs: list[Path] = []
    for directory in [evidence_root, *evidence_root.parents]:
        candidate_dirs.append(directory)
        if batch_id and directory.name.upper() == batch_id:
            break
    if batch_id:
        candidate_dirs.extend([
            ROOT / "batches" / f"jawal-{batch_id}",
            VOLUME_BASE / batch_id,
        ])
    invoice_sources: list[Path] = []
    seen_dirs: set[Path] = set()
    for directory in candidate_dirs:
        directory = directory.resolve(strict=False)
        if directory in seen_dirs:
            continue
        seen_dirs.add(directory)
        if not directory.is_dir():
            continue
        invoice_sources.extend(
            p for p in directory.iterdir()
            if p.is_file()
            and p.suffix.lower() in {".xlsx", ".xls", ".pdf"}
            and (
                p.name == "invoice-source.xlsx"
                or INVOICE_BASENAME_RE.search(p.stem)
            )
        )

    for source in invoice_sources:
        skip_stems.add(source.stem.lower())
        if source.suffix.lower() == ".pdf":
            skip_paths.add(source.resolve(strict=False))
            continue
        for directory in candidate_dirs:
            if directory.is_dir():
                skip_paths.add((directory / f"{source.stem}.pdf").resolve(strict=False))
    return skip_paths, skip_stems


def nested_below_root(path: Path, evidence_root: Path) -> bool:
    try:
        rel = path.relative_to(evidence_root)
    except ValueError:
        return False
    return len(rel.parts) > 1


def should_scan_ticket_body_pdf(
    pdf_path: Path,
    evidence_root: Path,
    invoice_skip_paths: set[Path],
    invoice_skip_stems: set[str],
) -> bool:
    if not nested_below_root(pdf_path, evidence_root):
        return False
    if pdf_path.resolve(strict=False) in invoice_skip_paths:
        return False
    if pdf_path.stem.lower() in invoice_skip_stems:
        return False
    if INVOICE_BASENAME_RE.search(pdf_path.stem):
        return False
    return True


def pdf_ticket_body_numbers(pdf_path: Path, pdf_text_cache: dict[Path, set[str]]) -> set[str]:
    if pdf_path in pdf_text_cache:
        return pdf_text_cache[pdf_path]
    numbers: set[str] = set()
    try:
        import fitz
        doc = fitz.open(str(pdf_path))
        try:
            text = "\n".join(page.get_text() for page in doc)
        finally:
            doc.close()
        numbers = reference_tokens_in_text(text)
    except Exception as exc:
        print(f"PREFLIGHT_WARNING {json.dumps({'pdf_read_failed': str(pdf_path), 'error': str(exc)}, ensure_ascii=False)}", flush=True)
    pdf_text_cache[pdf_path] = numbers
    return numbers


def run_scan(batch_id: str, raw_dir: str | None = None, emit_lines: bool = False) -> dict:
    normalized = normalize_batch_id(batch_id)
    evidence_dir = discover_evidence_dir(normalized, raw_dir)

    if emit_lines:
        emit("PREFLIGHT_START", {"batch_id": normalized, "evidence_dir": str(evidence_dir)})

    zero_byte, unreadable, total_files, total_bytes = scan_files(evidence_dir)
    for item in zero_byte:
        if emit_lines:
            emit("PREFLIGHT_ZERO_BYTE", item)
    for item in unreadable:
        if emit_lines:
            emit("PREFLIGHT_UNREADABLE", item)

    tickets, ticket_note = read_invoice_tickets(normalized)
    folder_names = collect_ticket_folder_names(evidence_dir) if tickets else set()
    missing_folders = [
        ticket for ticket in tickets
        if ticket["ticket_no"] not in folder_names
    ]
    for item in missing_folders:
        if emit_lines:
            emit("PREFLIGHT_MISSING_FOLDER", item)

    summary = {
        "zero_byte": len(zero_byte),
        "missing_folders": len(missing_folders),
        "unreadable": len(unreadable),
        "total_files": total_files,
        "total_bytes": total_bytes,
    }
    if ticket_note:
        summary["missing_folder_check"] = "skipped"
        summary["missing_folder_note"] = ticket_note

    if emit_lines:
        emit("PREFLIGHT_SUMMARY", summary)
        emit("PREFLIGHT_DONE")

    return {
        "batch_id": normalized,
        "evidence_dir": str(evidence_dir),
        "zero_byte": zero_byte,
        "missing_folders": missing_folders,
        "unreadable": unreadable,
        "total_files": total_files,
        "total_bytes": total_bytes,
        "has_issues": bool(zero_byte or missing_folders or unreadable),
        "missing_folder_check": "skipped" if ticket_note else "complete",
        "missing_folder_note": ticket_note or "",
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Scan evidence directory before running the AP pipeline.")
    parser.add_argument("batch_id", help="Batch id, e.g. J26-788")
    parser.add_argument("--raw-dir", default=None, help="Override evidence directory")
    args = parser.parse_args(argv)

    try:
        run_scan(args.batch_id, raw_dir=args.raw_dir, emit_lines=True)
        return 0
    except Exception as exc:
        emit("PREFLIGHT_ERROR", {"error": str(exc)})
        return 1


if __name__ == "__main__":
    sys.exit(main())
