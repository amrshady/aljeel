import openpyxl, re
from pathlib import Path

ours_path = Path("/home/clawdbot/.openclaw/workspace/aljeel/batches/jawal-J26-550/output/Spreadsheet-J26-550-FILLED-v30.xlsx")
truth_path = Path("/home/clawdbot/.openclaw/media/inbound/J26-550_-_validation---620a131c-94f0-416c-a0e0-c612be9e6667.xlsx")

wb_ours = openpyxl.load_workbook(ours_path, data_only=True)
ws_ours = wb_ours.active
headers_ours = [ws_ours.cell(3, col).value for col in range(1, ws_ours.max_column+1)]
col_map_ours = {str(h).strip(): i for i, h in enumerate(headers_ours) if h}

wb_truth = openpyxl.load_workbook(truth_path, data_only=True)
ws_truth = wb_truth.active

TICKET_RE = re.compile(r"(?<!\d)(\d{10}|26-\d{3,4})(?!\d)")

ours_by_ticket = {}
for r in range(4, ws_ours.max_row+1):
    desc = str(ws_ours.cell(r, col_map_ours['Description']+1).value or "")
    m = TICKET_RE.search(desc)
    if m:
        ours_by_ticket[m.group(1)] = r

print("=== LOCATION DIFFS ===")
for r in range(2, ws_truth.max_row+1):
    ticket_m = TICKET_RE.search(ws_truth.cell(r, 4).value or "") # Ticket No is column 4 (index 3)
    if not ticket_m:
        continue
    ticket = ticket_m.group(1)
    if ticket not in ours_by_ticket:
        continue
    
    ours_row = ours_by_ticket[ticket]
    ours_loc = str(ws_ours.cell(ours_row, col_map_ours['Location']+1).value or "").strip()
    
    # Truth combo is in Column AF (index 31, 1-based index 32)
    truth_combo = str(ws_truth.cell(r, 32).value or "").strip()
    truth_parts = truth_combo.split('-')
    truth_loc = truth_parts[1] if len(truth_parts) > 1 else ""
    
    if ours_loc != truth_loc:
        print(f"Ticket {ticket} | Passenger: {ws_truth.cell(r, 5).value} | Route: {ws_truth.cell(r, 6).value} | Ours Location: {ours_loc} | Truth Location: {truth_loc}")
