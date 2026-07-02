#!/usr/bin/env python3
"""
Batch Processor v5 — Full 10-segment Distribution Combination.

Reads an Oracle Fusion upload template (Spreadsheet.xlsx) with columns:
  [0]  *Invoice Header Identifier
  [1]  *Business Unit
  [2]  *Invoice Number
  [3]  *Invoice Currency
  [4]  *Invoice Amount
  [5]  *Invoice Date
  [6]  **Supplier[..]
  [7]  **Supplier Number
  [8]  *Supplier Site[..]
  [9]  Invoice Type
  [10] Description
  [11] *Type
  [12] *Amount
  [13] Distribution Combination[..]  ← WE FILL THIS
  [14] Tax Classification Code[..]
  [15] Employee No

Resolves each line to the full 10-segment combo per cost-center-rulebook-v1.md.

Usage:
    python3 scripts/process_batch.py --batch batches/jawal-J26-788
    python3 scripts/process_batch.py --batch batches/jawal-J26-788 --master-data qc/master-data/master-data-003.xlsx
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import os
import math
import shutil
from pathlib import Path
from datetime import datetime, timezone
from collections import Counter, defaultdict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# v15.11 (Amr May 25) — Pipeline version + static-field constants.
PIPELINE_VERSION = "v15.11"
JAWAL_SUPPLIER_AR = "شركة جوال للسفر والسياحة المحدودة"  # شركة جوال للسفر والسياحة المحدودة
JAWAL_SUPPLIER_NUMBER = "10394"
JAWAL_SUPPLIER_SITE_AR = "شركة جوال للسفر"  # شركة جوال للسفر
# v15.11.1 (Amr May 25 15:49 UTC): Invoice Type is row-aware.
# Refund rows -> "Debit Memo"; everything else -> "Standard".
# Detection: negative Invoice Amount OR Invoice Number starts with "CN-".
JAWAL_INVOICE_TYPE_REFUND = "Debit Memo"
JAWAL_INVOICE_TYPE_DEFAULT = "Standard"

def _v15_11_1_invoice_type(ws_row_inv_amount, ws_row_inv_no):
    """Decide Invoice Type per row. Refund = Debit Memo, else Standard."""
    try:
        amt = float(ws_row_inv_amount) if ws_row_inv_amount is not None else 0.0
    except (TypeError, ValueError):
        amt = 0.0
    inv_no = str(ws_row_inv_no or '').strip().upper()
    if amt < 0 or inv_no.startswith('CN-') or inv_no.startswith('CN '):
        return JAWAL_INVOICE_TYPE_REFUND
    return JAWAL_INVOICE_TYPE_DEFAULT
JAWAL_INVOICE_HEADER_ID = 1  # literal 1, every row

# Mai approval — HR business partner whose approval must be present on
# training trips + annual tickets. Detect by name token in approver chain.
MAI_APPROVER_TOKENS = ("mai", "مي")  # mai / مي
ANNUAL_TICKET_DIV = "888"
ANNUAL_TICKET_SOLUTION = "00000"
ANNUAL_TICKET_AGENCY = "88888"

# =============================================================================
# Labadi QC rule changes (2026-06-09)
# =============================================================================
# RULE 4 — OPEX serial numbers. Every OPEX submission carries a serial code of
# the form <DEPT_PREFIX>-<NUMBER>-<YEAR>; NUMBER and YEAR appear in either order
# (e.g. "CE-18-2026" and "CRM-2026-30"). Laith to provide the full prefix list;
# this is the partial set confirmed on the QC call.
KNOWN_OPEX_PREFIXES = {"CE", "CRM", "HF", "EP", "AATS"}

# Generic dept-prefix serial: 2-4 letters + two numeric groups, one of which is
# a 4-digit 20xx year. The generic prefix lets us also flag UNKNOWN prefixes.
_OPEX_SERIAL_RE = re.compile(r"\b([A-Za-z]{2,4})-(\d{1,4})-(\d{1,4})\b")


def _extract_opex_serial(text):
    """Return (serial_as_written, UPPER_prefix) for the first OPEX serial in
    `text`, else (None, None).

    Format: <DEPT_PREFIX>-<NUMBER>-<YEAR> with NUMBER/YEAR in either order. To
    avoid matching unrelated hyphenated tokens, one of the two numeric groups
    must look like a year (4 digits, 20xx).
    """
    if not text:
        return None, None
    for m in _OPEX_SERIAL_RE.finditer(str(text)):
        prefix, n1, n2 = m.group(1), m.group(2), m.group(3)
        if re.fullmatch(r"20\d{2}", n1) or re.fullmatch(r"20\d{2}", n2):
            return m.group(0), prefix.upper()
    return None, None


def _labadi_check_sanad_approval(form_data, *extra_texts):
    """RULE 5 — return True if a Sanad (or substitute Abdullah) approval is
    present. Checks approver-chain / approval-text / attachment-name style
    fields on the OPEX form plus any extra text blobs (subject/body), all
    case-insensitive in English and Arabic ("سند").

    Distinct from the existing `_v15_11_sanad_checks` scaffold (which flags
    missing screenshots); this is a presence check for the dual-approval rule.
    """
    parts = list(extra_texts)
    if isinstance(form_data, dict):
        for k in (
            "approver_name", "approver", "approver_chain", "approvers",
            "approval_text", "hr_approver", "hr_business_partner",
            "attachment_names", "body_text", "subject",
        ):
            v = form_data.get(k)
            if not v:
                continue
            if isinstance(v, (list, tuple)):
                parts.extend(str(x) for x in v)
            else:
                parts.append(str(v))
    blob = " ".join(p for p in parts if p)
    if not blob:
        return False
    low = blob.lower()
    return ("sanad" in low) or ("abdullah" in low) or ("سند" in blob)


def _labadi_is_opex_context(r):
    """RULE 4 — True if a row has an OPEX-form context worth a serial number:
    sponsorship rows (account 60307021), rows where an OPEX form was found, or
    rows whose description/folder text reads like an event. Employee direct-travel
    rows (60301003 / 60301004 with no OPEX form/event signal) are excluded.
    """
    acct = str(getattr(r, "account", "") or "").strip()
    if acct == "60307021":  # sponsorship
        return True
    if getattr(r, "_had_opex_form", False):
        return True
    blob = (getattr(r, "_opex_text_blob", "") or "").lower()
    event_kw = (
        "sponsor", "opex", "forum", "registration", "conference", "congress",
        "exhibition", "summit", "symposium", "iepc",
    )
    return any(k in blob for k in event_kw)

# Add scripts/ to path
ROOT = Path("/home/clawdbot/.openclaw/workspace/aljeel")
sys.path.insert(0, str(ROOT / "scripts"))
sys.path.insert(0, str(ROOT / "qc"))

from cost_center_resolver import (
    load_master_data, resolve_line, ResolvedLine, MasterData,
    sync_row_derived_fields,
    classify_account, _fuzzy_name_match, derive_sponsor_segments, _is_sponsorship,
)
from qc_gates import validate_line, GateResult
from allocation_resolver import resolve_allocation
from msg_parser import parse_msg, find_msgs_for_ticket
from oracle_form_parser import parse_form
from sponsorship_detector import detect_sponsorship, find_requesting_employee_from_form
from employee_resolver_v2 import resolve_employee, enrich_cache, _normalize_gds_name, ResolutionResult
from trip_purpose_classifier import classify_trip, detect_family_clusters, TripClassification
try:
    from location_resolver import resolve_location_v15_7 as _resolve_location_v15_7
    _LOCATION_RESOLVER_AVAILABLE = True
except ImportError:
    def _resolve_location_v15_7(description):
        return '20100', 'fallback', 'LOCATION_DEFAULT_CENTRAL_20100'
    _LOCATION_RESOLVER_AVAILABLE = False
from email_resolver import (
    extract_employee_email, enrich_email_cache, detect_manpower_email_column,
    generate_email_report, update_derived_batch, _load_derived_cache, resolve_by_email,
)
from load_employee_email_master import load_employee_email_master
from excel_styling import classify_row, get_row_style, REVIEW_WORTHY_FLAGS
from qc_catches_within_batch import run_within_batch_catches, _parse_route_corridor, _extract_ticket_no
from cross_batch_fraud import (
    run_cross_batch_fraud, update_cross_batch_history,
    load_history as load_cross_batch_history,
    save_history as save_cross_batch_history,
)


# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE")
AMBER_FILL  = PatternFill("solid", fgColor="FFEB9C")
RED_FILL    = PatternFill("solid", fgColor="FFC7CE")
GREY_FILL   = PatternFill("solid", fgColor="F2F2F2")
HDR_FILL    = PatternFill("solid", fgColor="1E40AF")
HDR_FONT    = Font(color="FFFFFF", bold=True, size=10)
BORDER_THIN = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

# ---------------------------------------------------------------------------
# Column indices in the Oracle template (0-based, header at row 2)
# ---------------------------------------------------------------------------
COL_HEADER_ID   = 0   # A
COL_BU          = 1   # B
COL_INV_NO      = 2   # C
COL_CURRENCY    = 3   # D
COL_AMOUNT_HDR  = 4   # E
COL_INV_DATE    = 5   # F
COL_SUPPLIER    = 6   # G
COL_SUPP_NO     = 7   # H
COL_SUPP_SITE   = 8   # I
COL_INV_TYPE    = 9   # J
COL_DESC        = 10  # K
COL_TYPE        = 11  # L
COL_AMOUNT      = 12  # M
COL_DIST_COMBO  = 13  # N  ← target
COL_TAX_CLASS   = 14  # O
COL_EMP_NO      = 15  # P
HEADER_ROW      = 2   # 0-indexed; row 3 in Excel (1-indexed)
DATA_START_ROW  = 3   # 0-indexed; row 4 in Excel


def _safe_float(v, default=0.0):
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return default
    try:
        return float(str(v).replace(",", ""))
    except (ValueError, TypeError):
        return default


def _safe_int(v, default=None):
    if v is None:
        return default
    try:
        return int(float(str(v).strip()))
    except (ValueError, TypeError):
        return default


def _extract_empno_from_desc(desc: str) -> int | None:
    """Try to extract 7-digit emp_no from Description field parenthetical."""
    if not desc:
        return None
    # Pattern: (1000539) or similar 7-digit in parens
    m = re.search(r"\((\d{7,8})\)", desc)
    if m:
        return int(m.group(1))
    return None


def _configure_stage1_no_cache(no_cache: bool):
    """Make known stage-1 cache files invisible to this process."""
    if not no_cache:
        return

    if getattr(Path, "_aljeel_stage1_no_cache", False):
        return

    original_exists = Path.exists
    cache_roots = tuple(
        p.resolve(strict=False)
        for p in (
            # msg-cache is deterministic (same file = same output) — never bust it
            # ROOT / "extracted" / "msg-cache",
            ROOT / "extracted" / "allocation-llm-cache",
            ROOT / "extracted" / "allocation-llm-email-cache",
            ROOT / "extracted" / "location-llm-cache",
            ROOT / "extracted" / "opex-pdf-cache",
        )
    )
    cache_files = {
        (ROOT / "cache" / "passenger_to_empno.json").resolve(strict=False),
        (ROOT / "cache" / "email_to_empno.json").resolve(strict=False),
    }

    def no_cache_exists(self):
        try:
            resolved = self.resolve(strict=False)
            if resolved in cache_files:
                return False
            resolved_str = str(resolved)
            for root in cache_roots:
                root_str = str(root)
                if resolved_str == root_str or resolved_str.startswith(root_str + os.sep):
                    return False
        except Exception:
            pass
        return original_exists(self)

    Path.exists = no_cache_exists
    Path._aljeel_stage1_no_cache = True




def _find_msgs_for_line(ticket_no, description, raw_dir):
    """Find .msg files for a line by ticket number or folder name."""
    if not raw_dir:
        return []
    from pathlib import Path
    raw_dir = Path(raw_dir)
    results = []
    
    # By ticket number
    if ticket_no:
        results = find_msgs_for_ticket(ticket_no, raw_dir)
    
    # If no results, try broader search by folder name patterns
    if not results and description:
        # Extract potential folder-name identifiers from description
        import os
        for dirpath, dirnames, filenames in os.walk(raw_dir):
            for f in filenames:
                if f.lower().endswith(".msg"):
                    results.append(Path(dirpath) / f)
    
    return results


def _segment_value(value, width: int, default: str = "") -> str:
    """Format numeric-ish master-data segment values without leaking blanks."""
    if value is None:
        return default
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return default
    try:
        text = str(int(float(text)))
    except (TypeError, ValueError):
        pass
    return text.zfill(width)


def _emp_no_from_text(text: str) -> int | None:
    """Extract a valid-looking AlJeel employee number from free text."""
    for m in re.finditer(r"\b(10\d{5})\b", str(text or "")):
        try:
            return int(m.group(1))
        except ValueError:
            continue
    return None


def _employee_from_emp_no(emp_no, md):
    try:
        emp_no_int = int(str(emp_no).strip())
    except (TypeError, ValueError):
        return None
    return md.employees.get(emp_no_int)


def _sender_email(sender: str) -> str:
    m = re.search(r"([\w.+-]+@aljeel\.com)", str(sender or ""), re.IGNORECASE)
    if not m:
        return ""
    email = m.group(1).lower()
    if email in {"sfalshammari@aljeel.com", "info@aljeel.com"}:
        return ""
    return email


def _sender_display_name(sender: str) -> str:
    text = re.sub(r"<[^>]+>", " ", str(sender or ""))
    text = re.sub(r"[\w.+-]+@[\w.-]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip(" '\"")
    return text


def _folder_ref_key(path_text: str) -> str:
    m = re.search(
        r"\b(?:OPEX[-_\s]*)?((?:CRM|HF|EP)-\d{4}-\d+|SIS-\d+-\d{4}|DMS-\d{4})\b",
        str(path_text or ""),
        re.IGNORECASE,
    )
    return m.group(1).upper() if m else ""


def _find_sponsoring_employee_from_shared_opex(folder_path, md, manpower_emails):
    """Resolve the staff sponsor from the shared OPEX evidence folder."""
    folder = Path(folder_path or "")
    if not folder.is_dir():
        return None, "folder_missing"

    folder_key = _folder_ref_key(str(folder))
    try:
        msg_files = sorted(folder.glob("*.msg"))
    except Exception:
        msg_files = []

    for msg_path in msg_files:
        parsed = parse_msg(msg_path, use_cache=True)
        if parsed.get("parse_method") == "failed":
            continue

        body = parsed.get("body_text", "") or ""
        subject = parsed.get("subject", "") or ""
        sender = parsed.get("sender", "") or ""
        attachment_names = parsed.get("attachment_names", []) or []

        form_candidates = []
        parsed_form = parse_form(body)
        if parsed_form:
            form_candidates.append(parsed_form)
        try:
            from opex_pdf_parser import parse_opex_from_msg
            opex_form = parse_opex_from_msg(str(msg_path))
            if opex_form:
                form_candidates.append(opex_form)
        except Exception:
            pass

        for form_data in form_candidates:
            form_blob = json.dumps(form_data, ensure_ascii=False, default=str)
            form_source = folder_key if folder_key and folder_key in form_blob.upper() else msg_path.name
            lookup_info = find_requesting_employee_from_form(form_data, body)
            emp = _employee_from_emp_no(lookup_info.get("requesting_emp_no"), md)
            if emp:
                return emp, f"form_requester:{form_source}"
            for key in ("emp_no", "employee_no", "employee_id", "requester_emp_no", "requesting_emp_no", "applicant_emp_no"):
                emp = _employee_from_emp_no(form_data.get(key), md)
                if emp:
                    return emp, f"form_emp:{key}"

        sender_addr = _sender_email(sender)
        if sender_addr:
            by_email = resolve_by_email(sender_addr, md, manpower_emails=manpower_emails)
            if by_email:
                emp = _employee_from_emp_no(by_email.get("emp_no"), md)
                if emp:
                    return emp, f"msg_sender_email:{sender_addr}"

        sender_name = _sender_display_name(sender)
        if sender_name:
            matched_emp_no = _fuzzy_name_match(sender_name, md.employees, threshold=0.72)
            emp = _employee_from_emp_no(matched_emp_no, md)
            if emp:
                return emp, f"msg_sender_name:{sender_name}"

        lookup_info = find_requesting_employee_from_form(
            {"body_text": body, "subject": subject, "attachment_names": attachment_names},
            body,
        )
        emp = _employee_from_emp_no(lookup_info.get("requesting_emp_no"), md)
        if emp:
            return emp, "msg_body_requester"

        emp = _employee_from_emp_no(_emp_no_from_text(body), md)
        if emp:
            return emp, "msg_body_emp_no"

        extracted = extract_employee_email(body, msg_sender=sender, msg_to=parsed.get("to") or [])
        by_email = resolve_by_email(extracted, md, manpower_emails=manpower_emails) if extracted else None
        if by_email:
            emp = _employee_from_emp_no(by_email.get("emp_no"), md)
            if emp:
                return emp, f"msg_body_email:{extracted}"

    # Standalone OPEX PDFs sitting loose in the shared folder (not .msg attachments)
    try:
        opex_pdfs = sorted(p for p in folder.glob("*.pdf") if p.name.upper().startswith("OPEX-"))
    except Exception:
        opex_pdfs = []

    for pdf in opex_pdfs:
        try:
            import fitz  # PyMuPDF
            doc = fitz.open(str(pdf))
            try:
                text = "\n".join(doc[i].get_text() for i in range(len(doc)))
            finally:
                doc.close()
        except Exception:
            continue

        if not text.strip():
            continue

        # Same OPEX form-field parsing path used for .msg bodies: requesting
        # employee number, then any bare employee number in the form text.
        lookup_info = find_requesting_employee_from_form(
            {"body_text": text, "subject": pdf.name, "attachment_names": [pdf.name]},
            text,
        )
        emp = _employee_from_emp_no(lookup_info.get("requesting_emp_no"), md)
        if emp:
            return emp, f"opex_pdf_standalone:{pdf.name}"

        emp = _employee_from_emp_no(_emp_no_from_text(text), md)
        if emp:
            return emp, f"opex_pdf_standalone:{pdf.name}"

        # Fall back to the requesting employee name on the form
        req_name = lookup_info.get("requesting_emp_name")
        if req_name:
            matched_emp_no = _fuzzy_name_match(req_name, md.employees, threshold=0.72)
            emp = _employee_from_emp_no(matched_emp_no, md)
            if emp:
                return emp, f"opex_pdf_standalone:{pdf.name}"

    return None, "not_found"


def _apply_shared_opex_sponsor_segments(r, shared_opex, md, manpower_emails):
    """Replace inherited passenger segments for SHARED_OPEX_SPONSORSHIP rows."""
    folder_path = shared_opex.get("shared_folder_path", "") if shared_opex else ""
    sponsor_emp, source = _find_sponsoring_employee_from_shared_opex(folder_path, md, manpower_emails)

    if sponsor_emp:
        r.location = _segment_value(getattr(sponsor_emp, "location", None), 5, "20100")
        r.cost_center = _segment_value(getattr(sponsor_emp, "cost_center", None), 6, "000000")
        r.div = _segment_value(getattr(sponsor_emp, "div_code", None), 3, "170")
        r.agency = _segment_value(getattr(sponsor_emp, "agency_code", None), 5, "00000")
        try:
            from cost_center_resolver import resolve_solution_code
            r.solution = resolve_solution_code(sponsor_emp)[0]
        except Exception:
            r.solution = "00000"
        # Labadi RULE 1 (2026-06-09): sponsorship rows carry the requestor's
        # emp_no (the AlJeel employee who submitted the OPEX form) — never blank.
        _sponsor_emp_no = getattr(sponsor_emp, "emp_no", None)
        if _sponsor_emp_no:
            r.emp_no = _sponsor_emp_no
        if "SHARED_OPEX_SPONSOR_SEGMENTS" not in r.flags:
            r.flags.append("SHARED_OPEX_SPONSOR_SEGMENTS")
        return source

    sponsor_segs = derive_sponsor_segments(str(folder_path))
    r.location = "20100"
    r.cost_center = "000000"
    r.div = "170"
    r.solution = sponsor_segs["solution"] if sponsor_segs.get("opex_key") != "_DEFAULT" else "00000"
    r.agency = "00000"
    if "SHARED_OPEX_DEFAULT_SEGMENTS" not in r.flags:
        r.flags.append("SHARED_OPEX_DEFAULT_SEGMENTS")
    return source


def _resolved_emp_tokens(emp_no):
    """Return normalized 7-digit employee tokens from a scalar or comma list."""
    if emp_no is None:
        return []
    text = str(emp_no).strip()
    if not text or text in ("0", "-", "nan", "None"):
        return []
    tokens = []
    for part in re.split(r"[,\s]+", text):
        part = part.strip()
        if not part:
            continue
        if re.fullmatch(r"\d+\.0", part):
            part = part[:-2]
        if re.fullmatch(r"\d{7}", part):
            tokens.append(part)
    return tokens


def _clear_stale_employee_not_in_master(r, gate, md):
    """Clear stale S7 when a later post-resolution stage assigned a valid emp_no."""
    resolved_emp = None
    for emp_no in _resolved_emp_tokens(getattr(r, "emp_no", None)):
        resolved_emp = md.employees.get(emp_no)
        if resolved_emp is not None:
            break
    if resolved_emp is None:
        return False

    changed = False
    if "EMPLOYEE_NOT_IN_MASTER" in (getattr(r, "flags", None) or []):
        r.flags = [f for f in r.flags if f != "EMPLOYEE_NOT_IN_MASTER"]
        changed = True

    if gate is not None:
        old_soft = list(getattr(gate, "soft_flags", []) or [])
        gate.soft_flags = [
            item for item in old_soft
            if not (len(item) >= 2 and item[1] == "EMPLOYEE_NOT_IN_MASTER")
        ]
        changed = changed or len(gate.soft_flags) != len(old_soft)
        r.action = gate.action

    if getattr(resolved_emp, "sol_flag", None):
        r.sol_flag = resolved_emp.sol_flag

    return changed


def _validate_with_email_form(resolved, description, ticket_no, amount, raw_dir, md, no_cache=False):
    """Run email-as-validator: parse Oracle form from .msg, compare with Manpower-derived segments.
    
    Returns dict with validator columns.
    """
    result = {
        "emp_match_source": "manpower",
        "form_emp_no": None,
        "form_approver": None,
        "form_division": None,
        "form_agency": None,
        "form_solution": None,
        "form_cost_center_ref": None,
        "validator_status": "FORM_MISSING",
        "discrepancy_detail": "",
        "form_value_sar": None,
        "form_trip_goal": None,
        "form_travel_method": None,
        "form_from_city": None,
        "form_to_city": None,
        "form_perdiem_class": None,
        "form_job_title": None,
        "form_grade": None,
    }
    
    if not raw_dir:
        return result, []
    
    flags = []
    
    # Find .msg files for this line
    msg_files = []
    if ticket_no:
        msg_files = find_msgs_for_ticket(ticket_no, raw_dir)
    
    if not msg_files:
        return result, ["FORM_NOT_FOUND_IN_EMAIL"]
    
    # Try to parse Oracle form from each .msg
    form_data = None
    for mf in msg_files:
        parsed_msg = parse_msg(mf, use_cache=True)
        if parsed_msg.get("parse_method") == "failed":
            continue
        body = parsed_msg.get("body_text", "")
        form = parse_form(body)
        if form:
            form_data = form
            break
    
    if not form_data:
        # Try OPEX PDF parsing (import here to avoid circular at module level)
        try:
            from opex_pdf_parser import parse_opex_from_msg
            for mf in msg_files:
                opex = parse_opex_from_msg(str(mf))
                if opex:
                    form_data = opex
                    flags.append("OPEX_PDF_PARSED")
                    break
        except Exception:
            pass
    
    if not form_data:
        # Try sponsorship detection
        for mf in msg_files:
            parsed_msg = parse_msg(mf, use_cache=True)
            if parsed_msg.get("parse_method") == "failed":
                continue
            body = parsed_msg.get("body_text", "")
            subject = parsed_msg.get("subject", "")
            atts = parsed_msg.get("attachment_names", [])
            
            spons = detect_sponsorship(
                msg_body=body,
                subject=subject,
                attachment_names=atts,
                emp_no_from_form=None,
                emp_no_from_manpower=resolved.emp_no,
                has_oracle_form=False,
                has_opex_pdf=False,
            )
            if spons:
                result["validator_status"] = "SPONSORSHIP_DETECTED"
                result["discrepancy_detail"] = f"confidence={spons.get('confidence',0):.2f}, host_hint={spons.get('host_name_hint','?')}"
                flags.append("SPONSORSHIP_DETECTED")
                return result, flags
        
        flags.append("FORM_NOT_FOUND_IN_EMAIL")
        return result, flags
    
    # Populate form fields
    result["form_emp_no"] = form_data.get("emp_no")
    result["form_approver"] = form_data.get("approver_name")
    result["form_division"] = form_data.get("form_division")
    result["form_agency"] = form_data.get("form_agency")
    result["form_solution"] = form_data.get("form_solution")
    result["form_cost_center_ref"] = form_data.get("trip_cost_for") or form_data.get("form_cost_center")
    result["form_value_sar"] = form_data.get("value_sar")
    result["form_trip_goal"] = form_data.get("trip_goal")
    result["form_travel_method"] = form_data.get("travel_method")
    result["form_from_city"] = form_data.get("from_city")
    result["form_to_city"] = form_data.get("to_city")
    result["form_perdiem_class"] = form_data.get("perdiem_class")
    result["form_job_title"] = form_data.get("job_title")
    result["form_grade"] = form_data.get("grade")
    
    # === VALIDATION COMPARISONS ===
    discrepancies = []
    
    # 1. Employee number match
    form_emp = form_data.get("emp_no")
    if form_emp and resolved.emp_no:
        if str(form_emp) == str(resolved.emp_no):
            result["emp_match_source"] = "both"
        else:
            result["emp_match_source"] = "form_disagrees"
            discrepancies.append(f"emp_no: form={form_emp} / manpower={resolved.emp_no}")
            flags.append("FORM_EMP_NO_MISMATCH")

            # FIX 3: the form has corrected the employee number. Apply the
            # correction and RE-DERIVE the GL segments from the corrected
            # employee's Manpower record. Otherwise the row keeps the wrong
            # (e.g. fuzzy-matched) employee's cost_center/div/agency/solution/
            # location and only the emp_no cell is right.
            corrected_emp = _employee_from_emp_no(form_emp, md)
            if corrected_emp is not None:
                resolved.emp_no = corrected_emp.emp_no
                resolved.cost_center = _segment_value(
                    getattr(corrected_emp, "cost_center", None), 6, resolved.cost_center)
                resolved.div = _segment_value(
                    getattr(corrected_emp, "div_code", None), 3, resolved.div)
                resolved.agency = _segment_value(
                    getattr(corrected_emp, "agency_code", None), 5, resolved.agency)
                resolved.location = _segment_value(
                    getattr(corrected_emp, "location", None), 5, resolved.location)
                try:
                    from cost_center_resolver import resolve_solution_code
                    resolved.solution = resolve_solution_code(corrected_emp)[0]
                except Exception:
                    pass
                # Re-run account classification against the corrected employee.
                try:
                    _acct, _acct_rule = classify_account(
                        description, corrected_emp, md,
                        approver_name=form_data.get("approver_name", "") or "")
                    resolved.account = _acct
                    resolved.account_rule = _acct_rule
                except Exception:
                    pass
                # Trace note so reviewers can see segments were re-applied.
                _reapply_note = f"form_disagrees_segments_reapplied: emp={corrected_emp.emp_no}"
                _prev_trace = getattr(resolved, "routing_reason", "") or ""
                resolved.routing_reason = (
                    f"{_prev_trace} | {_reapply_note}" if _prev_trace else _reapply_note)
                if "FORM_DISAGREES_SEGMENTS_REAPPLIED" not in flags:
                    flags.append("FORM_DISAGREES_SEGMENTS_REAPPLIED")
                result["discrepancy_detail_extra"] = _reapply_note
    elif form_emp:
        result["emp_match_source"] = "form"
    
    # 2. Approver vs line manager
    if form_data.get("approver_name") and resolved.emp_no:
        emp = md.employees.get(resolved.emp_no)
        if emp and hasattr(emp, "manager_name") and emp.manager_name:
            from cost_center_resolver import _fuzzy_name_match
            if not _fuzzy_name_match(form_data["approver_name"], emp.manager_name):
                discrepancies.append(f"approver: form={form_data['approver_name']} / manpower_mgr={emp.manager_name}")
                flags.append("FORM_APPROVER_NOT_LINE_MANAGER")
    
    # 3. Trip value vs invoice amount
    form_value = form_data.get("value_sar")
    if form_value and amount and abs(form_value - amount) > 1.0:
        discrepancies.append(f"value: form={form_value} / invoice={amount}")
        flags.append("FORM_TRIP_VALUE_DIFFERS")
    
    # 4. Log Fusion internal codes (no gate, just informational)
    if form_data.get("form_division") or form_data.get("form_agency") or form_data.get("form_solution"):
        flags.append("FORM_FUSION_CODES_LOGGED")
    
    # Set validator status
    if discrepancies:
        result["validator_status"] = "FORM_DISAGREES"
        result["discrepancy_detail"] = " | ".join(discrepancies)
    else:
        result["validator_status"] = "FORM_AGREES"
        flags.append("FORM_AGREES_WITH_MANPOWER")
    
    return result, flags


def get_human_action_recommendation(flags, sol_flag, row_status):
    """Generate a helpful, friendly sentence for the human reviewer describing what to do."""
    all_flags = set(flags or [])
    
    if "ALLOCATION_TARGET_MISSING" in all_flags or sol_flag == "Need to allocate":
        return "This employee is marked 'Need to allocate' in lookups. Please manually confirm their correct Cost Center/Agency with HR/Finance."
        
    if "EMAIL_CORRUPT_OR_EMPTY" in all_flags:
        return "The approval email (.msg) file is empty (0 KB) or corrupt. Please obtain a valid email copy from the traveler/manager."

    if "SHARED_OPEX_SPONSORSHIP" in all_flags:
        return "Shared OPEX sponsorship evidence exists in the date folder. Use that folder as the approval evidence for this ticket."
        
    if "NO_FOLDER" in all_flags:
        return "No evidence folder exists on disk for this ticket. Please locate the traveler's approval and create the folder."
        
    if "NO_APPROVAL" in all_flags:
        return "An evidence folder exists, but it contains zero approval email (.msg) files. Please upload the approval email."
        
    if "PERSONAL_CONTRIB_SELF_APPROVAL" in all_flags:
        return "Self-approval detected (requester is the approver in the email). Please verify manager sign-off for audit compliance."
        
    if "FORM_NOT_FOUND_IN_EMAIL" in all_flags:
        return "Approval email exists, but the standard Oracle/Workday form was not found inside it. Please check if the email text contains valid authorization."
        
    if "TRIP_PURPOSE_UNKNOWN" in all_flags:
        return "The trip purpose could not be identified. Please verify the email body text or passenger description to assign the correct GL account."
        
    if "MANPOWER_DIV_NOT_IN_MASTER" in all_flags:
        return "The employee's division is missing or invalid in lookups. Please verify and map their division in the master lookups file."
        
    if "FORM_EMP_NO_MISMATCH" in all_flags:
        return "The employee ID on the approval form differs from the matched employee. Please verify which employee actually traveled."
        
    if "FORM_TRIP_VALUE_DIFFERS" in all_flags:
        return "The amount on the approval form differs from the invoiced amount. Please check for fare adjustments or flight changes."

    if row_status == "GREEN":
        return "Clean match. No action required."
        
    return "Soft flags identified. Please review the Agent Flags and QC Catches column."


def process_batch(
    batch_dir: Path,
    master_data_path: Path,
    reference_path: Path,
    output_suffix: str = "v5-rulebook",
    raw_dir: Path = None,
    no_cache: bool = False,
    invoice_file: Path = None,
) -> dict:
    """Process a Jawal batch end-to-end.

    Args:
        batch_dir: Path to the batch directory (e.g., batches/jawal-J26-788)
        master_data_path: Path to Master Data (003).xlsx
        reference_path: Path to J26-640-resolved.xlsx (reference tabs)
        output_suffix: suffix for output filename
        raw_dir: path to raw .msg evidence directory
        no_cache: bypass persistent stage-1 cache reads
        invoice_file: explicit path to invoice xlsx; overrides disk glob when provided

    Returns:
        dict with summary stats
    """
    _configure_stage1_no_cache(no_cache)

    # --- Find the Spreadsheet template ---
    # Prefer the canonical batch input when present; generated output/ files are
    # legacy fallbacks and may be stale after a no-cache rerun.
    if (batch_dir / "Spreadsheet-v4-input.xlsx").exists():
        spreadsheet = batch_dir / "Spreadsheet-v4-input.xlsx"
    elif (batch_dir / "Spreadsheet.xlsx").exists():
        spreadsheet = batch_dir / "Spreadsheet.xlsx"
    elif (filled_candidates := sorted(batch_dir.glob("output/Spreadsheet-*-FILLED-v4*.xlsx"), reverse=True)):
        spreadsheet = filled_candidates[0]
    else:
        candidates = list(batch_dir.glob("Spreadsheet*.xlsx"))
        if not candidates:
            raise FileNotFoundError(f"No Spreadsheet.xlsx found in {batch_dir}")
        spreadsheet = candidates[0]

    print(f"[process_batch] Loading master data from {master_data_path.name}...")
    md = load_master_data(master_data_path, reference_path)
    print(f"  {len(md.employees)} employees loaded")
    # Detect Manpower Email column
    manpower_emails = detect_manpower_email_column(master_data_path)
    if manpower_emails:
        print(f"  Manpower Email column detected: {len(manpower_emails)} emails")
    else:
        print(f"  No Manpower Email column — using learned email cache")

    # v15.10.1 (May 25, 2026): Layer the standalone Employee Email Master on top.
    # AlJeel finance team supplied a separate xlsx with emp_no->email mappings
    # (677 of 679 staff have @aljeel/@elajou emails). This is the canonical
    # email->emp_no source for fuzzy-name resolution failures.
    email_master_path = ROOT / 'qc' / 'master-data' / 'Employee_Email_Master_2026-05-25.xlsx'
    extra_emails = load_employee_email_master(email_master_path)
    if extra_emails:
        # Newer file wins on conflict (overwrites any stale Manpower entry)
        before = len(manpower_emails)
        manpower_emails.update(extra_emails)
        print(f"  Employee Email Master loaded: +{len(extra_emails)} emails "
              f"({before} -> {len(manpower_emails)} total)")

    print(f"  {len(md.valid_accounts)} valid accounts, {len(md.valid_divs)} DIVs, "
          f"{len(md.valid_agencies)} agencies, {len(md.valid_solutions)} solutions, "
          f"{len(md.valid_cost_centers)} cost centers")

    # --- Read the spreadsheet ---
    print(f"[process_batch] Reading {spreadsheet.name}...")
    df = pd.read_excel(spreadsheet, sheet_name=0, header=None)

    # Find header row (contains "*Invoice Header Identifier")
    header_row_idx = None
    for i in range(min(10, len(df))):
        cell_val = str(df.iloc[i, 0]).strip()
        if "Invoice Header Identifier" in cell_val:
            header_row_idx = i
            break
    if header_row_idx is None:
        raise ValueError("Cannot find header row with *Invoice Header Identifier")

    print(f"  Header at row {header_row_idx}, data starts at row {header_row_idx + 1}")



    # --- Process each data row ---
    results = []
    gate_results = []
    halt_queue = []

    _cache_enrichment_pairs = []
    _trip_batch_lines = []
    _email_enrichment_pairs = []
    _email_extraction_log = []

    for i in range(header_row_idx + 1, len(df)):
        row = df.iloc[i]
        header_id = row.iloc[COL_HEADER_ID]
        if pd.isna(header_id):
            continue

        sl_no = _safe_int(header_id, 0)
        description = str(row.iloc[COL_DESC]) if pd.notna(row.iloc[COL_DESC]) else ""
        amount = _safe_float(row.iloc[COL_AMOUNT])
        emp_no_raw = _safe_int(row.iloc[COL_EMP_NO])

        # Also try extracting emp_no from description if not in col P
        if emp_no_raw is None:
            emp_no_raw = _extract_empno_from_desc(description)
        
        # Try ticket number from description for context
        # GDS 10-digit tickets: (6905264364) or voucher format: (26-659)
        ticket_match = re.search(r"\((\d{10})\)", description or "")
        voucher_match = re.search(r"\((\d{2}-\d{3})\)", description or "")
        ticket_no = ticket_match.group(1) if ticket_match else None
        voucher_no = voucher_match.group(1) if voucher_match else None

        # Read existing Distribution Combination from v4 (Account.CC format)
        existing_combo = str(row.iloc[COL_DIST_COMBO]) if pd.notna(row.iloc[COL_DIST_COMBO]) else ""
        existing_account = None
        if "." in existing_combo:
            parts = existing_combo.split(".")
            if len(parts[0]) == 8 and parts[0].isdigit():
                existing_account = parts[0]

        # Extract passenger name from Description
        # Format: "LASTNAME/FIRSTNAME TITLE - ROUTE (TICKET)"
        passenger = ""
        if description:
            # Take everything before the first " - " as passenger
            dash_idx = description.find(" - ")
            if dash_idx > 0:
                passenger = description[:dash_idx].strip()
            else:
                passenger = description.strip()

        # --- Email extraction from .msg ---
        extracted_email = None
        email_extract_reason = "no_msg_files"
        is_corrupt_or_empty = False
        if raw_dir and ticket_no:
            from msg_parser import find_msgs_for_ticket
            email_msg_files = find_msgs_for_ticket(ticket_no, raw_dir)
            for mf in email_msg_files:
                if not os.path.exists(mf) or os.path.getsize(mf) == 0:
                    is_corrupt_or_empty = True
                    continue
                parsed_for_email = parse_msg(mf, use_cache=True)
                if parsed_for_email.get("parse_method") == "failed":
                    is_corrupt_or_empty = True
                    continue
                body_for_email = parsed_for_email.get("body_text", "")
                extracted_email = extract_employee_email(
                    msg_body=body_for_email,
                    msg_sender=parsed_for_email.get("sender", ""),
                    msg_to=parsed_for_email.get("to", []),
                )
                if extracted_email:
                    email_extract_reason = "extracted"
                    break
            if not extracted_email and email_msg_files:
                email_extract_reason = "no_employee_email_in_msg"

        # --- Employee Resolution v2 (9-layer cascade) ---
        # First, get form data if available (for L1, L7)
        v2_form_emp = None
        v2_form_approver = None
        v2_msg_filenames = []
        if raw_dir and ticket_no:
            v2_msg_files = find_msgs_for_ticket(ticket_no, raw_dir)
            v2_msg_filenames = [str(f) for f in v2_msg_files]
            # Try parsing Oracle form from msgs for form_emp_no
            for mf in v2_msg_files:
                parsed = parse_msg(mf, use_cache=True)
                if parsed.get("parse_method") == "failed":
                    continue
                body = parsed.get("body_text", "")
                form = parse_form(body)
                if form:
                    v2_form_emp = form.get("emp_no")
                    v2_form_approver = form.get("approver_name")
                    break
            # Also try OPEX PDF for form emp_no
            if not v2_form_emp:
                try:
                    from opex_pdf_parser import parse_opex_from_msg
                    for mf in v2_msg_files:
                        opex = parse_opex_from_msg(str(mf))
                        if opex and opex.get("emp_no"):
                            v2_form_emp = opex["emp_no"]
                            break
                except Exception:
                    pass
        
        # Run the 10-layer cascade (L0-L9 including L1.5 email)
        # L8 cache validation now handles poisoned employee entries generally,
        # so no row-specific L0 clearing is needed here.

        v2_result = resolve_employee(
            passenger_name=passenger,
            description=description,
            emp_no_raw=emp_no_raw,
            form_emp_no=v2_form_emp,
            form_approver=v2_form_approver,
            msg_filenames=v2_msg_filenames,
            ticket_no=ticket_no,
            raw_dir=raw_dir,
            md=md,
            extracted_email=extracted_email,
            manpower_emails=manpower_emails,
            no_cache=no_cache,
        )
        
        # Ref. No. trust rule (2026-06-09): the invoice Ref. No. column already
        # carries the responsible employee's emp_no (converter extracts it into
        # col P). When that value is a valid employee in master it is
        # authoritative for ALL row types — employee, CHD, spouse, family —
        # and must not be overridden downstream.
        emp_from_ref_no = bool(emp_no_raw and emp_no_raw in md.employees)

        # Use v2 result for emp_no resolution
        resolved_emp_no = v2_result.emp_no if v2_result.emp_no else emp_no_raw
        if emp_from_ref_no:
            resolved_emp_no = emp_no_raw

        # Resolve the line with the v2-resolved emp_no
        resolved = resolve_line(
            sl_no=sl_no,
            description=description,
            emp_no_raw=resolved_emp_no,
            passenger_name=passenger,
            amount=amount,
            md=md,
            msg_filenames=v2_msg_filenames,
        )
        
        # Override match method with v2 layer info
        if v2_result.layer != "not_resolved" and v2_result.layer != "L0":
            resolved.emp_match_method = f"v2_{v2_result.layer}"
            resolved.routing_reason = v2_result.trace

        # Fix E v15.1: L7.7 emp_no propagation
        # If L7.7 found a Person Number from the Oracle form but the person is
        # not in md.employees (e.g. manager-only row like MERHEB 1002576),
        # resolve_line() would have set resolved.emp_no=None. Force it here.
        if (
            getattr(v2_result, "layer", "") in ("v3_L7.7_email_llm", "v3_L8_manager_cc_unanimous", "v3_L_overrides")
            and v2_result.emp_no is not None
            and (resolved.emp_no is None or resolved.emp_no != v2_result.emp_no)
        ):
            resolved.emp_no = v2_result.emp_no
            print(f"[Fix E v15.8] {v2_result.layer} emp_no {v2_result.emp_no} forced to primary (was {resolved.emp_no!r}, not in md.employees)", flush=True)

        # EMP_FROM_REF_NO enforcement: re-assert the Ref. No. emp_no after
        # resolve_line / Fix E so no downstream layer silently replaces it.
        if emp_from_ref_no:
            if resolved.emp_no != emp_no_raw:
                resolved.emp_no = emp_no_raw
            if "EMP_FROM_REF_NO" not in resolved.flags:
                resolved.flags.append("EMP_FROM_REF_NO")

        # Handle sponsorship auto-routing from L9
        if v2_result.is_sponsorship and v2_result.sponsorship_meta:
            resolved.account = v2_result.sponsorship_meta.get("auto_account", "60307021")
            resolved.account_rule = f"v2_L9_sponsorship: {v2_result.sponsorship_meta.get('reason', 'auto')}"
            # v15.6: location already set by cost_center_resolver via resolve_location_v15_6
            if v2_result.flag_code:
                resolved.flags.append(v2_result.flag_code)
        elif v2_result.flag_code == "NEW_EMPLOYEE_NOT_IN_MASTER":
            resolved.flags.append("NEW_EMPLOYEE_NOT_IN_MASTER")
        elif v2_result.flag_code and v2_result.flag_code not in resolved.flags:
            resolved.flags.append(v2_result.flag_code)

        # v15.2: extra_flags from multi-flag results (e.g. MANAGER_CC_FRAGMENTED)
        for _xf in getattr(v2_result, "extra_flags", []):
            if _xf and _xf not in resolved.flags:
                resolved.flags.append(_xf)
        
        # Segment overrides: apply pooled overrides from L7.5/L8/PaxOverrides
        if hasattr(v2_result, "segment_overrides") and v2_result.segment_overrides:
            seg = v2_result.segment_overrides
            resolved.cost_center = seg.get("cost_center", resolved.cost_center)
            resolved.div = seg.get("div", resolved.div)
            resolved.agency = seg.get("agency", resolved.agency)
            resolved.solution = seg.get("solution", resolved.solution)
            # v15.9 Fix 3b: apply explicit account from seg (e.g. PaxOverrides 60308007)
            _seg_account = seg.get("account")
            if _seg_account:
                resolved.account = _seg_account
                resolved.account_rule = f"segment_override_account: {_seg_account}"
            # v15.6: location from itinerary (only if employee home location is not found)
            if not resolved.location or resolved.location in ("", "None", "0"):
                _l75_loc, _l75_layer, _l75_flag = _resolve_location_v15_7(description)
                resolved.location = _l75_loc
                if _l75_flag and _l75_flag not in resolved.flags:
                    resolved.flags.append(_l75_flag)
            # v15.9 Fix 3b: only reclassify account by DIV when:
            #   - seg does NOT contain an explicit account, AND
            #   - resolution layer is NOT v3_L_overrides (PaxOverrides trusts its own account)
            _is_overrides_layer = getattr(v2_result, "layer", "") == "v3_L_overrides"
            if not _seg_account and not _is_overrides_layer:
                if resolved.div in ("888", "190"):
                    resolved.account = "60301004"  # G&A travel
                    resolved.account_rule = "L7.5_reverse_mgr_ga_travel"
                else:
                    resolved.account = "60301003"  # S&M travel
                    resolved.account_rule = "L7.5_reverse_mgr_sm_travel"
            # Remove EMPLOYEE_NOT_IN_MASTER since we resolved via manager pool
            resolved.flags = [f for f in resolved.flags if f != "EMPLOYEE_NOT_IN_MASTER"]
            resolved.action = "Post to GL"
        
        # Store v2 resolution metadata for Excel output
        # Store raw data for QC catches
        resolved._description = description
        resolved._amount = amount
        resolved._ticket_no = ticket_no
        resolved._inv_date = None  # Will be extracted from invoice
        resolved._v2_layer = v2_result.layer
        resolved._v2_confidence = v2_result.confidence
        resolved._v2_trace = v2_result.trace
        resolved._v2_flag = v2_result.flag_code or ""
        if emp_from_ref_no:
            resolved._v2_flag = (resolved._v2_flag + " EMP_FROM_REF_NO").strip()
            resolved._v2_trace = (resolved._v2_trace or "") + f" | EMP_FROM_REF_NO: invoice Ref. No. emp_no={emp_no_raw} trusted"
        
        # Track email extraction for report + cache enrichment
        _email_extraction_log.append({
            "sl_no": sl_no,
            "passenger": passenger,
            "extracted_email": extracted_email,
            "resolved_via_email": v2_result.layer == "L1.5",
            "emp_no": v2_result.emp_no,
            "no_email_reason": email_extract_reason if not extracted_email else None,
        })
        
        # Enrich email cache: if resolved AND email extracted, pair them
        if v2_result.emp_no and extracted_email:
            _email_enrichment_pairs.append((
                extracted_email, v2_result.emp_no,
                f"msg_body_via_{v2_result.layer}", v2_result.confidence,
            ))
        
        # Also store matched_email on resolved line for Excel output
        resolved._v2_email = extracted_email or (v2_result.matched_email if hasattr(v2_result, 'matched_email') else "")
        resolved.is_corrupt_or_empty_email = is_corrupt_or_empty
        if is_corrupt_or_empty:
            resolved.flags.append("EMAIL_CORRUPT_OR_EMPTY")
        
        # Track for cache enrichment
        # GUARD: never cache passenger→emp_no for sponsorship rows.
        # On sponsorship rows the emp_no comes from the OPEX form requestor,
        # not from the passenger themselves. Caching it would poison future
        # batches by assigning the requestor's emp_no to whoever travels under
        # that name. (Bug: UTHMAN ALUTHMAN J26-593 requestor 1001986 wrongly
        # inherited by ALUTHMAN/UTHMAN MR in J26-589.)
        if v2_result.emp_no and not v2_result.is_sponsorship:
            given, surname = _normalize_gds_name(passenger)
            _cache_enrichment_pairs.append((f"{given} {surname}".strip().upper(), v2_result.emp_no))

        # If our classifier defaulted (Rule 6/99) but v4 had a specific account,
        # trust the v4 account (it was derived with more context)
        if existing_account and existing_account in md.valid_accounts:
            if any(x in resolved.account_rule for x in ("Rule 6", "Rule 99", "L8_sm_travel", "L7_ga_travel")):
                resolved.account = existing_account
                resolved.account_rule = f"v4_preserved: {existing_account}"

        # === OPEX-ref segment override (v11-labadi) ===
        # When sponsorship is detected from OPEX/event refs in the description,
        # ALWAYS use OPEX-derived segments for CC/DIV/Agency/Solution/Location
        # regardless of whether an employee was found in Manpower.
        # Rationale: sponsorship budget codes are determined by the EVENT
        # (CRM/HF/EP/SIS/DMS), not by the traveling employee's home department.
        # v15.9: SKIP when v3_L_overrides was used — PaxOverrides has explicit segments
        # (e.g., AL MAHROUQ Fujifilm agency=10041 must not be clobbered by SIS key agency=10043)
        _skip_opex = getattr(v2_result, "layer", "") == "v3_L_overrides"
        is_sponsor, _sp = _is_sponsorship(description)
        if is_sponsor and not _skip_opex:
            sponsor_segs = derive_sponsor_segments(description)
            resolved.account = "60307021"
            resolved.cost_center = sponsor_segs["cost_center"]
            resolved.div = sponsor_segs["div"]
            resolved.agency = sponsor_segs["agency"]
            resolved.solution = sponsor_segs["solution"]
            # v15.6: location already set by cost_center_resolver via resolve_location_v15_6
            resolved.account_rule = f"L1_sponsor_override: opex_key={sponsor_segs['opex_key']}"
            if "EMPLOYEE_NOT_IN_MASTER" not in resolved.flags:
                resolved.flags.append("OPEX_SEGMENT_OVERRIDE")


        # --- Allocation resolution for "Need to allocate" employees ---
        allocation_result = None
        if "ALLOCATION_TARGET_MISSING" in resolved.flags and resolved.emp_no:
            # Find .msg files for this ticket
            msg_bodies = []
            if raw_dir and ticket_no:
                msg_files = find_msgs_for_ticket(ticket_no, raw_dir)
                for mf in msg_files:
                    parsed = parse_msg(mf, use_cache=True)
                    if parsed.get("body_text"):
                        msg_bodies.append(parsed["body_text"])

            alloc = resolve_allocation(msg_bodies, resolved.emp_no, md)
            allocation_result = alloc

            if alloc.resolved and alloc.subordinate_emp_no:
                # Rebuild combo using subordinate's Manpower data
                sub_emp = md.employees.get(alloc.subordinate_emp_no)
                if sub_emp:
                    sub_account, sub_account_rule = classify_account(description, sub_emp, md)
                    _alloc_loc, _, _alloc_flag = _resolve_location_v15_7(description)
                    resolved.location = _alloc_loc
                    if _alloc_flag and _alloc_flag not in resolved.flags:
                        resolved.flags.append(_alloc_flag)
                    resolved.cost_center = str(sub_emp.cost_center).zfill(6)
                    resolved.div = str(sub_emp.div_code).zfill(3)
                    resolved.agency = str(sub_emp.agency_code).zfill(5)
                    resolved.account = sub_account
                    resolved.account_rule = f"allocation_to_{sub_emp.emp_no}: {sub_account_rule} (mgr={resolved.emp_no})"
                    resolved.emp_match_method = f"allocation_email_subordinate ({alloc.method})"

                    # Resolve solution for subordinate
                    from cost_center_resolver import resolve_solution_code
                    sub_sol, sub_sol_flags = resolve_solution_code(sub_emp)
                    resolved.solution = sub_sol

                    # Replace ALLOCATION_TARGET_MISSING with the resolution flag
                    resolved.flags = [f for f in resolved.flags if f != "ALLOCATION_TARGET_MISSING"]
                    resolved.flags.append(alloc.flag_code)
            elif alloc.flag_code == "MULTI_ALLOCATION_PENDING_REVIEW":
                resolved.flags = [f for f in resolved.flags if f != "ALLOCATION_TARGET_MISSING"]
                resolved.flags.append("MULTI_ALLOCATION_PENDING_REVIEW")
            elif alloc.flag_code == "ALLOCATION_LOOP_DETECTED":
                resolved.flags = [f for f in resolved.flags if f != "ALLOCATION_TARGET_MISSING"]
                resolved.flags.append("ALLOCATION_LOOP_DETECTED")
            # else: keep ALLOCATION_TARGET_MISSING (no regression)


        # --- Email Validator Step (A4) ---
        validator_result = {}
        validator_flags = []
        if raw_dir:
            validator_result, validator_flags = _validate_with_email_form(
                resolved, description, ticket_no, amount, raw_dir, md, no_cache=no_cache
            )
            for vf in validator_flags:
                if vf not in resolved.flags:
                    resolved.flags.append(vf)


        # --- Trip Purpose Classification (Step between resolver and QC gates) ---
        trip_cls = None
        trip_subject = ""
        trip_body = ""
        trip_form = None
        
        if raw_dir and ticket_no:
            trip_msg_files = find_msgs_for_ticket(ticket_no, raw_dir)
            for mf in trip_msg_files:
                trip_parsed = parse_msg(mf, use_cache=True)
                if trip_parsed.get("parse_method") == "failed":
                    continue
                trip_subject = trip_parsed.get("subject", "")
                trip_body = trip_parsed.get("body_text", "")
                trip_form_candidate = parse_form(trip_body)
                if trip_form_candidate:
                    trip_form = trip_form_candidate
                if trip_subject:
                    break  # Use first msg with a subject
        
        trip_cls = classify_trip(
            subject=trip_subject,
            body=trip_body,
            form_data=trip_form,
            description=description,
            passenger_name=passenger,
        )
        
        # Store classification on resolved line
        resolved._trip_purpose = trip_cls.trip_purpose
        resolved._trip_confidence = trip_cls.confidence
        resolved._trip_signals = ", ".join(trip_cls.signals_used)
        resolved._trip_trace = trip_cls.trace
        resolved._trip_account_override = trip_cls.account_override or ""

        # --- Labadi RULE 4 & 5 capture (2026-06-09) ---
        # Stash the OPEX form context + approval signals on the row now, while the
        # parsed form/subject/body are in scope. The OPEX serial, Mai (HR) and
        # Sanad approval *statuses* are finalised after all account mutations
        # (family-cluster + shared-OPEX) settle, just before row-status classify.
        _opex_parts = [description or "", trip_subject or "", trip_body or ""]
        if isinstance(trip_form, dict):
            _opex_parts.append(json.dumps(trip_form, ensure_ascii=False, default=str))
        resolved._opex_text_blob = " ".join(p for p in _opex_parts if p)
        resolved._had_opex_form = bool(trip_form)
        _approver_for_mai = None
        if isinstance(trip_form, dict):
            _approver_for_mai = trip_form.get("approver_name") or trip_form.get("approver")
        # Reuse the existing Mai detector (RULE 5). Raw booleans only here.
        resolved._mai_ok_raw = _v15_11_check_mai_approval(trip_form, _approver_for_mai)
        resolved._sanad_ok_raw = _labadi_check_sanad_approval(trip_form, trip_subject, trip_body)

        # Apply account override if classifier has high enough confidence
        if trip_cls.account_override and trip_cls.confidence >= 0.7:
            old_account = resolved.account
            resolved.account = trip_cls.account_override
            resolved.account_rule = f"trip_classifier_{trip_cls.trip_purpose}: {trip_cls.trace[:80]}"
            # v15.6: location already set by cost_center_resolver, no override needed
            resolved.flags.append("ACCOUNT_OVERRIDE_APPLIED")
            # Current 21070229 rule: home CC + annual-ticket fixed segments.
            _v15_11_apply_personal_or_annual_rollup(resolved, md)

            # v15.11 (Change #4): HR/Mai approval required for training + annual-ticket rows.
            # Sponsorships are NOT subject to this rule (Mai is HR-only).
            if trip_cls.trip_purpose in ("TRAINING", "ANNUAL_LEAVE_TICKET"):
                approver_name_for_check = None
                if trip_form:
                    approver_name_for_check = trip_form.get("approver_name") or trip_form.get("approver")
                if not _v15_11_check_mai_approval(trip_form, approver_name_for_check):
                    if "MISSING_HR_APPROVAL" not in resolved.flags:
                        resolved.flags.append("MISSING_HR_APPROVAL")

            # v15.11 (Change #6): Sanad scaffold flags
            sanad_flags = _v15_11_sanad_checks(trip_form, raw_dir, ticket_no)
            for f in sanad_flags:
                if f not in resolved.flags:
                    resolved.flags.append(f)

            # v15.11 (Change #6): 50/50 shared sponsorship
            if trip_cls.trip_purpose == "SPONSORSHIP" and _v15_11_shared_sponsorship_check(trip_form):
                if "SHARED_SPONSORSHIP_50_50" not in resolved.flags:
                    resolved.flags.append("SHARED_SPONSORSHIP_50_50")

            # v15.11 (Change #5): Sponsorship form lookup — find the AlJeel staff
            # member who requested/sponsored the guest, verify approval completion.
            # Triggered when the row is classified as SPONSORSHIP AND the resolver
            # has no employee (no emp_no on the line).
            if trip_cls.trip_purpose == "SPONSORSHIP" and not resolved.emp_no:
                lookup_info = find_requesting_employee_from_form(trip_form, trip_body)
                for f in lookup_info.get("flags", []):
                    if f not in resolved.flags:
                        resolved.flags.append(f)
                req_emp = lookup_info.get("requesting_emp_no")
                if req_emp and req_emp.isdigit():
                    # Resolve the requesting employee through Manpower to get GL segments.
                    req_emp_int = int(req_emp)
                    if req_emp_int in md.employees:
                        emp_record = md.employees[req_emp_int]
                        # Charge to requesting employee's GL — overwrite segments.
                        resolved.emp_no = req_emp_int
                        # Use cost_center_resolver to derive segments cleanly.
                        from cost_center_resolver import resolve_line as _resolve_for_sponsor
                        # Reuse simpler segment copy: take CC/DIV/Solution/Agency from emp record.
                        # Note: detailed field names depend on master-data layout.
                        if hasattr(emp_record, "cost_center") and emp_record.cost_center:
                            resolved.cost_center = str(emp_record.cost_center).zfill(6)
                        if hasattr(emp_record, "div") and emp_record.div is not None:
                            resolved.div = str(emp_record.div).zfill(3)
                        if hasattr(emp_record, "solution") and emp_record.solution is not None:
                            resolved.solution = str(emp_record.solution).zfill(5)
                        if hasattr(emp_record, "agency") and emp_record.agency is not None:
                            resolved.agency = str(emp_record.agency).zfill(5)
                        if "SPONSORSHIP_CHARGED_TO_REQUESTER_v15.11" not in resolved.flags:
                            resolved.flags.append("SPONSORSHIP_CHARGED_TO_REQUESTER_v15.11")
            
            # Check for mismatch between subject and form
            if trip_form:
                form_type = trip_form.get("form_type", "")
                perdiem = trip_form.get("perdiem_class", "") or ""
                if trip_cls.trip_purpose == "PERSONAL" and form_type == "business_trip" and ".Personal" not in perdiem:
                    resolved.flags.append("TRIP_PURPOSE_MISMATCH")

                elif trip_cls.trip_purpose == "BUSINESS_TRIP" and ".Personal" in perdiem:
                    resolved.flags.append("TRIP_PURPOSE_MISMATCH")
            
            # Check for low confidence personal
            if trip_cls.trip_purpose == "PERSONAL" and trip_cls.confidence < 0.8 and len(trip_cls.signals_used) <= 1:
                resolved.flags.append("PERSONAL_LOW_CONFIDENCE")
        
        elif trip_cls.trip_purpose == "UNKNOWN":
            resolved.flags.append("TRIP_PURPOSE_UNKNOWN")

        # v15.11 (Amr May 25): Apply personal/annual rollup unconditionally at end-of-row.
        # Catches L5_personal from cost_center_resolver even when trip classifier did NOT override.
        _v15_11_apply_personal_or_annual_rollup(resolved, md)

        # Collect for family cluster detection
        _trip_batch_lines.append({
            "sl_no": sl_no,
            "passenger_name": passenger,
            "description": description,
            "trip_purpose": trip_cls.trip_purpose if trip_cls else "UNKNOWN",
            "idx": len(results),  # index into results list (before append)
        })

        # Save trip classifier flags before gate reset
        _pre_gate_trip_flags = [f for f in resolved.flags if f in (
            "ACCOUNT_OVERRIDE_APPLIED", "TRIP_PURPOSE_MISMATCH",
            "PERSONAL_LOW_CONFIDENCE", "TRIP_PURPOSE_UNKNOWN",
            "FAMILY_CLUSTER_DETECTED", "MIXED_FAMILY_CLUSTER",
            "EMPLOYEE_AS_SPONSORED_HARD",
            # v15.11: preserve through gate reset
            "MISSING_HR_APPROVAL",
            "INCOMPLETE_SPONSORSHIP_APPROVAL",
            "EXTERNAL_SPONSORSHIP_NO_EMP_NUM",
            "SHARED_SPONSORSHIP_50_50",
            "SANAD_NO_SCREENSHOT",
            "SANAD_TIMING_MISMATCH",
            "ANNUAL_TICKET_HOME_SEGMENTS",
            "ANNUAL_TICKET_HOME_CC_MISSING",
            "SPONSORSHIP_CHARGED_TO_REQUESTER_v15.11",
        )]

        # Run QC gates
        gate = validate_line(resolved, md)
        gate_results.append(gate)

        # Override the line's action/flags with gate results
        resolved.flags = []  # Reset — gates are authoritative
        for _, code in gate.hard_failures:
            resolved.flags.append(code)
        for _, code in gate.soft_flags:
            resolved.flags.append(code)
        # Re-add trip classifier flags (survived gate reset)
        for tf in _pre_gate_trip_flags:
            if tf not in resolved.flags:
                resolved.flags.append(tf)
        if getattr(resolved, 'is_corrupt_or_empty_email', False):
            if "EMAIL_CORRUPT_OR_EMPTY" not in resolved.flags:
                resolved.flags.append("EMAIL_CORRUPT_OR_EMPTY")
        resolved.action = gate.action

        results.append(resolved)
        # Store validator result for Excel output
        if not hasattr(resolved, '_validator'):
            resolved._validator = validator_result

        if not gate.passed_hard:
            halt_queue.append({
                "sl_no": sl_no,
                "passenger": passenger,
                "description": description,
                "amount": amount,
                "combo": resolved.combo,
                "failures": [(gid, code) for gid, code in gate.hard_failures],
            })


    # --- Family Cluster Detection ---
    family_clusters = detect_family_clusters(_trip_batch_lines)
    for sl_no, cluster_info in family_clusters.items():
        for idx, r in enumerate(results):
            if r.sl_no == sl_no:
                if cluster_info.get("mixed"):
                    r.flags.append("MIXED_FAMILY_CLUSTER")
                elif cluster_info.get("should_flip_personal", False):
                    # v10: only flip to PERSONAL if cluster has CHD markers
                    # v15.11: PERSONAL routes to 21070229 (not 11034013).
                    r.flags.append("FAMILY_CLUSTER_DETECTED")
                    r.flags.append("FAMILY_CLUSTER_CHD_CONFIRMED")
                    if getattr(r, "_trip_purpose", "") != "PERSONAL":
                        r.account = "21070229"  # v15.11: was 11034013
                        r.account_rule = f"family_cluster_{cluster_info['cluster_id']}: CHD+cluster→PERSONAL (v15.11)"
                        # Current 21070229 rule: home CC + annual-ticket fixed segments.
                        _v15_11_apply_personal_or_annual_rollup(r, md)
                        r._trip_purpose = "PERSONAL"
                        r._trip_confidence = 0.9
                        r._trip_signals = "family_cluster_chd"
                        r._trip_trace = f"family cluster {cluster_info['cluster_id']} with CHD"
                        r._trip_account_override = "21070229"  # v15.11: was 11034013
                        r.flags.append("ACCOUNT_OVERRIDE_APPLIED")
                else:
                    # Cluster detected but no CHD — flag only, don't flip
                    r.flags.append("FAMILY_CLUSTER_DETECTED")
                    r.flags.append("FAMILY_CLUSTER_NO_CHD")
                break

    # --- Family Cluster emp_no Unification (v15.11.2, May 25) ---
    # Within a single cluster, all rows should resolve to ONE canonical
    # sponsor emp_no. Sponsorship_detector was guessing parent per-row by
    # walking Manpower line-managers, which produced split sponsors
    # (e.g. SALEH cluster: 2x emp 1002066 + 1x emp 1000074 — same family,
    # two different "sponsors").
    #
    # Rule (only applies when the cluster was promoted to PERSONAL via the
    # CHD path, i.e. should_flip_personal=True): pick the modal emp_no
    # across the cluster. Tie-break order:
    #   1. Adult passenger emp_no wins over child emp_no
    #     (CHD/INF passenger title indicates the row is the child; the OTHER
    #     row in a 1-adult/1-child tie must be the adult and supplies the parent)
    #   2. Higher manpower grade wins
    #   3. Lower emp_no (stable, deterministic) wins
    # Any row whose emp_no differs from canonical is reassigned, its segments
    # are recomputed from the canonical employee's Manpower record (then
    # re-rolled-up via _v15_11_apply_personal_or_annual_rollup since the
    # account stays 21070229), and the row gets a FAMILY_CLUSTER_UNIFIED flag
    # that downstream QC (NO_APPROVAL) reads as proof of sponsor coverage.
    _seen_clusters = set()
    for sl_no, cluster_info in family_clusters.items():
        cid = cluster_info.get("cluster_id")
        if not cid or cid in _seen_clusters:
            continue
        _seen_clusters.add(cid)
        if not cluster_info.get("should_flip_personal"):
            continue  # only unify CHD-confirmed clusters (where account flipped to 21070229)
        member_sls = set(cluster_info.get("cluster_members", []))
        if not member_sls:
            continue

        # Collect (sl_no, emp_no, is_adult, grade) for each cluster row
        members = []
        for r in results:
            if r.sl_no not in member_sls:
                continue
            pax_upper = (r.passenger_name or "").upper()
            is_child = ("(CHD)" in pax_upper) or ("(INF)" in pax_upper)
            grade_val = 0
            if r.emp_no and r.emp_no in md.employees:
                emp_rec = md.employees[r.emp_no]
                try:
                    grade_val = int(getattr(emp_rec, "grade", 0) or 0)
                except (TypeError, ValueError):
                    grade_val = 0
            members.append({
                "sl_no": r.sl_no,
                "emp_no": r.emp_no,
                "is_adult": not is_child,
                "grade": grade_val,
                "row_ref": r,
            })

        # Only emp_nos that actually resolve in Manpower are candidates for
        # the canonical sponsor. Rows with unresolved emp_no (None) can
        # still be reassigned to the canonical, but cannot supply it.
        emp_candidates = [m for m in members if m["emp_no"] and m["emp_no"] in md.employees]
        if not emp_candidates:
            continue  # nothing to unify against

        # Modal vote on emp_no across all members (including children).
        counts = {}
        for m in emp_candidates:
            counts[m["emp_no"]] = counts.get(m["emp_no"], 0) + 1
        max_votes = max(counts.values())
        leaders = [e for e, c in counts.items() if c == max_votes]

        if len(leaders) == 1:
            canonical_emp = leaders[0]
            tiebreak_reason = "modal"
        else:
            # Tie-break: adult > child, then higher grade, then lower emp_no
            tied_members = [m for m in emp_candidates if m["emp_no"] in leaders]
            # Reduce tied emp_nos to a single representative member each
            # (so per-emp_no attributes can be compared cleanly)
            best_per_emp = {}
            for m in tied_members:
                cur = best_per_emp.get(m["emp_no"])
                if cur is None:
                    best_per_emp[m["emp_no"]] = m
                else:
                    # within same emp_no, prefer the adult representative
                    if m["is_adult"] and not cur["is_adult"]:
                        best_per_emp[m["emp_no"]] = m
            ranked = sorted(
                best_per_emp.values(),
                key=lambda m: (
                    0 if m["is_adult"] else 1,    # adults first
                    -m["grade"],                  # higher grade first
                    m["emp_no"],                  # lower emp_no first
                ),
            )
            canonical_emp = ranked[0]["emp_no"]
            tiebreak_reason = "adult>child/grade/emp_no"

        canonical_record = md.employees.get(canonical_emp)
        if canonical_record is None:
            continue

        # Canonical employee name tokens for dependent detection
        _canon_name_tokens = set(
            t for t in canonical_record.name.upper().split() if len(t) >= 3
        )

        # v29 family guard helper — returns True if this passenger is a
        # dependent (CHD/INF title) OR a non-employee family member.
        # Logic: CHD/INF markers are definitive.  For adults without CHD/INF
        # markers (e.g. a spouse listed as MRS), we check whether the
        # passenger's GDS-family-name token appears in the canonical employee's
        # name tokens.  If the family-name matches but the given-name does NOT
        # overlap at all, the row is a family member, not the employee.
        def _is_dependent_pax(pax_name: str) -> bool:
            _pu = (pax_name or "").upper()
            # Explicit child/infant marker → always dependent
            if "(CHD)" in _pu or "(INF)" in _pu:
                return True
            # Strip GDS title suffixes and get meaningful name tokens
            _strip = re.sub(
                r"\b(MR|MRS|MS|DR|ENG|MISS|MASTER|INF|CHD|MSTR)\b", "", _pu
            ).strip()
            # GDS format: FAMILY/GIVEN  → split on '/' if present
            if "/" in _strip:
                _parts = _strip.split("/", 1)
                _family_tok = _parts[0].strip()
                _given_toks = set(t for t in _parts[1].split() if len(t) >= 3)
            else:
                _toks = [t for t in _strip.split() if len(t) >= 3]
                _family_tok = _toks[0] if _toks else ""
                _given_toks = set(_toks[1:]) if len(_toks) > 1 else set()
            # If family name is not in the canonical employee's name tokens,
            # this isn't the same person at all — could be unrelated, don't
            # suppress based on name mismatch alone (too risky; leave to CHD check).
            if _family_tok and _family_tok not in _canon_name_tokens:
                return False  # different family entirely — not our cluster's employee
            # Same family name.  Now check given name overlap.
            # If there is NO given-name token overlap with canonical, it's a
            # family member sharing the same surname.
            if _given_toks and not _given_toks.intersection(_canon_name_tokens):
                return True  # same surname, different given name → dependent
            return False

        # Reassign any member whose emp_no != canonical
        for m in members:
            r = m["row_ref"]
            _pax_up = (r.passenger_name or "").upper()

            # Labadi RULE 1 (2026-06-09): dependents/CHD rows must KEEP an emp_no —
            # specifically the employee whose family members are travelling. The
            # prior v29 guard blanked emp_no for dependent/family passengers on
            # account=21070229; that is now REVERSED. Dependent rows are no longer
            # special-cased here — they fall through to the normal canonical
            # reassignment below, which assigns the sponsoring employee's emp_no
            # AND propagates CC/DIV/Solution/Agency segments. Tag them so the
            # assignment reason is visible in the debug output.
            if r.account == "21070229" and _is_dependent_pax(r.passenger_name or ""):
                flag_dep = "FAMILY_CLUSTER_DEPENDENT_EMP_ASSIGNED"
                if flag_dep not in r.flags:
                    r.flags.append(flag_dep)

            if r.emp_no == canonical_emp:
                continue

            old_emp = r.emp_no
            r.emp_no = canonical_emp
            # Pull canonical employee's segments (CC/DIV/Solution/Agency).
            # Location stays as resolved from the itinerary.
            try:
                r.cost_center = str(canonical_record.cost_center).zfill(6)
                r.div = str(canonical_record.div_code).zfill(3)
                r.agency = str(canonical_record.agency_code).zfill(5)
                # Solution: use solution_name on the record if present,
                # otherwise default to 00000.
                # cost_center_resolver's resolve_solution_code expects the
                # full employee record; reuse it for parity.
                try:
                    from cost_center_resolver import resolve_solution_code as _resolve_sol
                    new_sol, _sol_flags = _resolve_sol(canonical_record)
                    r.solution = new_sol
                except Exception:
                    r.solution = "00000"
            except Exception:
                pass
            # Account 21070229 (PERSONAL/family-cluster rollup): stamp the
            # canonical sponsor's home CC plus annual-ticket fixed segments.
            if r.account == "21070229":
                _v15_11_apply_personal_or_annual_rollup(r, md)
            r.emp_match_method = "family_cluster_unified"
            r.account_rule = (
                f"{r.account_rule} | family_cluster_unified: "
                f"emp {old_emp}→{canonical_emp} ({tiebreak_reason})"
            )
            flag_str = f"FAMILY_CLUSTER_UNIFIED(emp_no={canonical_emp},{tiebreak_reason})"
            if flag_str not in r.flags:
                r.flags.append(flag_str)

        # Also stamp the canonical-aligned rows with the unified flag (so
        # downstream NO_APPROVAL clearing treats every cluster member as
        # covered by the sponsor's approval).
        for m in members:
            r = m["row_ref"]
            if any(f.startswith("FAMILY_CLUSTER_UNIFIED") for f in r.flags):
                continue
            flag_str = f"FAMILY_CLUSTER_UNIFIED(emp_no={canonical_emp},already_canonical)"
            r.flags.append(flag_str)

    # --- Enrich cross-batch cache ---
    if _cache_enrichment_pairs and not no_cache:
        enrich_cache(_cache_enrichment_pairs)

    # --- Write output Excel ---
    output_dir = batch_dir / "output"
    output_dir.mkdir(parents=True, exist_ok=True)

    # Derive invoice number and date from the data
    inv_no = ""
    inv_date_str = None
    for i in range(header_row_idx + 1, min(header_row_idx + 5, len(df))):
        v = df.iloc[i, COL_INV_NO]
        if pd.notna(v) and str(v).strip():
            inv_no = str(v).strip()
            break
    for i in range(header_row_idx + 1, min(header_row_idx + 5, len(df))):
        v = df.iloc[i, COL_INV_DATE]
        if pd.notna(v):
            try:
                if isinstance(v, datetime):
                    inv_date_str = v.strftime("%Y-%m-%d")
                else:
                    inv_date_str = str(v).strip()[:10]
            except Exception:
                pass
            break

    out_filename = f"Spreadsheet-{inv_no}-FILLED-{output_suffix}.xlsx" if inv_no else f"Spreadsheet-FILLED-{output_suffix}.xlsx"
    out_path = output_dir / out_filename

    # --- Enrich email cache ---
    if _email_enrichment_pairs and not no_cache:
        enrich_email_cache(_email_enrichment_pairs)
        if inv_no:
            update_derived_batch(inv_no)

    # --- Generate email report ---
    report_dir = ROOT / "qc" / "reports" / "2026-05-21-resolver-v2"
    email_report_stats = generate_email_report(
        extraction_log=_email_extraction_log,
        derived_cache=_load_derived_cache(),
        report_dir=report_dir,
        batch_ids=[inv_no or "unknown"],
    )
    print(f"  Email report: {report_dir}")
    print(f"    Emails extracted: {email_report_stats['with_email']}/{email_report_stats['total_lines']}")
    print(f"    Resolved via email (L1.5): {email_report_stats['resolved_via_email']}")
    print(f"    Unique derived emails: {email_report_stats['unique_derived']}")

    # --- Within-batch QC catches ---
    print(f"\n[process_batch] Running within-batch QC catches...")
    within_batch_catches = run_within_batch_catches(
        results, gate_results, raw_dir=raw_dir, md=md, inv_date=inv_date_str,
    )
    print(f"  Within-batch catches: {len(within_batch_catches)}")
    for cat, count in Counter(c['category'] for c in within_batch_catches).most_common():
        print(f"    {cat}: {count}")

    # Write within-batch catches
    catches_within_path = output_dir / "catches-within-batch.json"
    with open(catches_within_path, "w") as f:
        json.dump(within_batch_catches, f, indent=2, default=str)

    # --- Cross-batch fraud detection ---
    print(f"\n[process_batch] Running cross-batch fraud detection...")
    cross_history = load_cross_batch_history()
    
    # Build batch_lines for cross-batch checks
    _cross_batch_lines = []
    for r in results:
        desc = getattr(r, '_description', '')
        _cross_batch_lines.append({
            "ticket_no": getattr(r, '_ticket_no', None) or _extract_ticket_no(desc),
            "passenger": r.passenger_name,
            "amount": getattr(r, '_amount', 0),
            "route_corridor": _parse_route_corridor(desc),
            "inv_date": inv_date_str,
            "emp_no": r.emp_no,
            "grade": None,  # Grade not in Manpower master — OVER_LIMIT and OVER_BUDGET catches skip gracefully
            "sl_no": r.sl_no,
        })
    
    cross_batch_catches = run_cross_batch_fraud(inv_no, _cross_batch_lines, cross_history)
    print(f"  Cross-batch catches: {len(cross_batch_catches)}")
    for cat, count in Counter(c['category'] for c in cross_batch_catches).most_common():
        print(f"    {cat}: {count}")
    
    # Write cross-batch catches
    catches_cross_path = output_dir / "catches-cross-batch.json"
    with open(catches_cross_path, "w") as f:
        json.dump(cross_batch_catches, f, indent=2, default=str)

    # Update cross-batch history with current batch AFTER catches
    update_cross_batch_history(inv_no, _cross_batch_lines, cross_history)
    save_cross_batch_history(cross_history)
    print(f"  Cross-batch history updated: {len(cross_history.get('tickets', {}))} tickets tracked")

    # --- Build per-line QC catch index ---
    line_catches = defaultdict(list)  # sl_no -> list of catches
    for catch in within_batch_catches + cross_batch_catches:
        for sn in catch.get("sl_nos", []):
            if sn is not None:
                line_catches[sn].append(catch)

    for r in results:
        catches_for_line = line_catches.get(r.sl_no, [])
        shared_opex = next(
            (c for c in catches_for_line if c.get("category") == "SHARED_OPEX_SPONSORSHIP"),
            None,
        )
        if not shared_opex:
            continue

        r.account = "60307021"
        # Labadi RULE 1 (2026-06-09): do NOT blank emp_no on shared-OPEX
        # sponsorship rows. _apply_shared_opex_sponsor_segments sets emp_no to the
        # sponsoring (requestor) employee when resolvable from the OPEX folder;
        # otherwise the resolver-assigned emp_no is left intact.
        segment_source = _apply_shared_opex_sponsor_segments(r, shared_opex, md, manpower_emails)
        method = str(getattr(r, "_agent_method", "") or "")
        if "+shared_opex_sponsorship" not in method:
            setattr(r, "_agent_method", method + "+shared_opex_sponsorship")
        detail = (
            f"shared_opex_sponsorship: {shared_opex.get('shared_folder_path', '')}; "
            f"segments={segment_source}"
        )
        r.account_rule = f"{r.account_rule} | {detail}" if r.account_rule else detail

    stale_emp_master_cleared = 0
    for r, gate in zip(results, gate_results):
        if _clear_stale_employee_not_in_master(r, gate, md):
            stale_emp_master_cleared += 1
    if stale_emp_master_cleared:
        print(
            f"  Stale EMPLOYEE_NOT_IN_MASTER cleared after post-resolution: "
            f"{stale_emp_master_cleared}",
            flush=True,
        )

    # --- Labadi RULE 4 & 5 finalisation (2026-06-09) ---
    # Account routing is now settled (trip override + family-cluster + shared-OPEX).
    # Compute the OPEX serial and the Mai/Sanad dual-approval statuses per row,
    # generate the new catches, and merge them into the within-batch catch set and
    # the per-line index so they surface in the QC Catches column + row status.
    labadi_catches = []
    for r in results:
        # RULE 4 — OPEX serial number (OPEX-context rows only).
        if _labadi_is_opex_context(r):
            # Priority: LLM-extracted serial > regex from text blob > MISSING
            llm_serial = getattr(r, '_opex_serial', None)  # set by run_v30.py from LLM result
            if llm_serial and llm_serial not in ("MISSING", "N/A", ""):
                serial = llm_serial
                prefix = llm_serial.split("-")[0].upper() if "-" in llm_serial else ""
            else:
                # Fallback: regex
                serial, prefix = _extract_opex_serial(getattr(r, "_opex_text_blob", "") or "")
            if serial:
                r._opex_serial = serial
                if prefix not in KNOWN_OPEX_PREFIXES:
                    labadi_catches.append({
                        "category": "OPEX_SERIAL_UNKNOWN_PREFIX",
                        "severity": "LOW",
                        "passenger": r.passenger_name,
                        "account": r.account,
                        "sl_nos": [r.sl_no],
                        "detail": (
                            f"OPEX serial {serial!r} has unknown dept prefix "
                            f"{prefix!r} (known: {sorted(KNOWN_OPEX_PREFIXES)})."
                        ),
                    })
            else:
                r._opex_serial = "MISSING"
                labadi_catches.append({
                    "category": "OPEX_SERIAL_MISSING",
                    "severity": "MEDIUM",
                    "passenger": r.passenger_name,
                    "account": r.account,
                    "sl_nos": [r.sl_no],
                    "detail": "OPEX-context row has no extractable OPEX serial number.",
                })
        else:
            r._opex_serial = "N/A"

        # RULE 5 — annual/family travel dual approval (Mai + Sanad).
        if str(getattr(r, "account", "") or "").strip() == "21070229":
            mai_ok = bool(getattr(r, "_mai_ok_raw", False))
            sanad_ok = bool(getattr(r, "_sanad_ok_raw", False))
            r._mai_approval_status = "OK" if mai_ok else "MISSING"
            r._sanad_approval_status = "OK" if sanad_ok else "MISSING"
            if not mai_ok:
                labadi_catches.append({
                    "category": "ANNUAL_HR_APPROVAL_MISSING",
                    "severity": "HIGH",
                    "passenger": r.passenger_name,
                    "account": r.account,
                    "sl_nos": [r.sl_no],
                    "detail": "Annual/family travel (21070229) missing HR (Mai) approval.",
                })
            if not sanad_ok:
                labadi_catches.append({
                    "category": "ANNUAL_SANAD_APPROVAL_MISSING",
                    "severity": "HIGH",
                    "passenger": r.passenger_name,
                    "account": r.account,
                    "sl_nos": [r.sl_no],
                    "detail": "Annual/family travel (21070229) missing Sanad/Abdullah approval.",
                })
        else:
            r._mai_approval_status = "N/A"
            r._sanad_approval_status = "N/A"

    # Merge Labadi catches into the within-batch set + per-line index.
    if labadi_catches:
        within_batch_catches.extend(labadi_catches)
        for catch in labadi_catches:
            for sn in catch.get("sl_nos", []):
                if sn is not None:
                    line_catches[sn].append(catch)
        print(f"  Labadi RULE 4/5 catches: {len(labadi_catches)}")
        for cat, count in Counter(c["category"] for c in labadi_catches).most_common():
            print(f"    {cat}: {count}")

    # --- Classify row status ---
    row_statuses = []
    for i, (r, g) in enumerate(zip(results, gate_results)):
        catches_for_line = line_catches.get(r.sl_no, [])
        status = classify_row(g, r, catches_for_line)
        row_statuses.append(status)
        # Add QC catch flags to the resolved line for Excel output
        for catch in catches_for_line:
            cat = catch["category"]
            if cat not in r.flags:
                r.flags.append(cat)

    status_counter = Counter(row_statuses)
    print(f"\n[process_batch] Row status breakdown:")
    print(f"  🟢 GREEN:  {status_counter.get('GREEN', 0)}")
    print(f"  🟡 YELLOW: {status_counter.get('YELLOW', 0)}")
    print(f"  🔴 RED:    {status_counter.get('RED', 0)}")

    from code_name_lookup import get_lookup
    lookup = get_lookup(reference_path)

    # Finalize derived fields once all segment mutations have completed.
    for r in results:
        sync_row_derived_fields(r, lookup)

    # Copy original and modify
    shutil.copy(spreadsheet, out_path)
    wb = load_workbook(out_path)
    ws = wb.active

    # openpyxl is 1-indexed
    excel_header_row = header_row_idx + 1  # 1-indexed

    # =========================================================================
    # OUTPUT RESTRUCTURE v11: 3 blocks — Oracle Fusion | Code/Name | Debug
    # =========================================================================
    # --- Block 1: Oracle Fusion template (cols A-P = 1-16) — UNCHANGED ---
    # These 16 columns already exist in the spreadsheet from the copy.
    # We just write combo + emp_no into them (as before).

    # --- Block 2: Code/Description expansion (cols Q-AF = 17-32, 16 cols) ---
    # v15.11 (Amr May 25): Added "Full GL String + Description" as the final
    # column AFTER the existing 15 segment+name columns. Existing green
    # code+description columns preserved unchanged (AlJeel uses them for QA).
    BLOCK2_START = 17  # col Q (1-indexed)
    BLOCK2_HEADERS = [
        "Company", "Location", "Account", "GL",
        "Cost Center", "Cost Name", "DIV", "Contribution",
        "Solution", "Solution Name", "Agency", "Agency Name",
        "Project", "Intercompany", "Future 1",
        "GL Description",  # v15.11.2: renamed from 'Full GL String + Description'; now description-only
    ]

    # --- Block 3: Debug/Agent columns (cols AG onwards = 33+) ---
    # v15.11: shifted from 32 -> 33 because BLOCK2 grew by 1 column ("Full GL String + Description")
    BLOCK3_START = 33  # col AG (1-indexed)
    BLOCK3_HEADERS = [
        # v15.11: column letters shifted right by 1 (BLOCK2 grew by 1 column).
        "Row Status",                # AG
        "Manpower Allocation Status", # AH
        "Evidence Folder Status",    # AI
        "Approval Email Status",     # AJ
        "Self-Approval Status",      # AK
        "Human Review Note",         # AL
        "QC Catches",                # AM
        "Agent Flags",               # AN
        "Agent Action",              # AO
        "Agent Emp Match Source",    # AP
        "Agent Match Method",        # AQ
        "Agent Account Rule",        # AR
        "Agent Segments Breakdown",  # AS
        "Form Emp No",               # AT
        "Form Approver",             # AU
        "Form Division (Fusion code)",   # AV
        "Form Agency (Fusion code)",     # AW
        "Form Solution (Fusion code)",   # AX
        "Form Cost-Center-Ref (Fusion 15-digit)",  # AY
        "Validator Status",          # AZ
        "Discrepancy Detail",        # BA
        "Resolution Layer",          # BB
        "Resolution Confidence",     # BC
        "Resolution Trace",          # BD
        "Resolution Flag",           # BE
        "Email Match",               # BF
        "Trip Purpose",              # BG
        "Trip Purpose Confidence",   # BH
        "Trip Purpose Signals",      # BI
        "Trip Purpose Trace",        # BJ
        "Trip Account Override",     # BK
        # Labadi RULE 4 & 5 (2026-06-09): OPEX serial + dual-approval columns.
        "OPEX Serial",               # BL  (RULE 4)
        "HR Approval (Mai)",         # BM  (RULE 5)
        "Sanad Approval",            # BN  (RULE 5)
    ]

    # Clear any existing columns beyond P (col 16) to avoid leftover data
    for col in range(17, ws.max_column + 1):
        for row in range(1, ws.max_row + 1):
            ws.cell(row=row, column=col).value = None
            ws.cell(row=row, column=col).fill = PatternFill()  # clear fill

    # --- Header fills ---
    BLOCK1_HDR_FILL = PatternFill("solid", fgColor="D9E1F2")  # light blue
    BLOCK2_HDR_FILL = PatternFill("solid", fgColor="C6EFCE")  # light green
    BLOCK3_HDR_FILL = PatternFill("solid", fgColor="E7E6E6")  # light gray
    LABEL_FONT = Font(bold=True, size=11)
    HDR_FONT_DARK = Font(bold=True, size=10)
    HDR_ALIGN = Alignment(horizontal="center")

    # --- Row 2 (excel_header_row - 1): Block labels ---
    label_row = excel_header_row - 1  # row 2 in Excel (1-indexed)
    if label_row >= 1:
        # Block 1 label: A-P
        cell = ws.cell(row=label_row, column=1, value="ORACLE FUSION TEMPLATE")
        cell.font = LABEL_FONT
        cell.fill = BLOCK1_HDR_FILL
        cell.alignment = HDR_ALIGN
        ws.merge_cells(start_row=label_row, start_column=1, end_row=label_row, end_column=16)

        # Block 2 label: Q-AE
        cell = ws.cell(row=label_row, column=BLOCK2_START, value="CODE & DESCRIPTION")
        cell.font = LABEL_FONT
        cell.fill = BLOCK2_HDR_FILL
        cell.alignment = HDR_ALIGN
        ws.merge_cells(start_row=label_row, start_column=BLOCK2_START, end_row=label_row, end_column=BLOCK2_START + len(BLOCK2_HEADERS) - 1)

        # Block 3 label: AF onwards
        cell = ws.cell(row=label_row, column=BLOCK3_START, value="DEBUG (delete before posting)")
        cell.font = LABEL_FONT
        cell.fill = BLOCK3_HDR_FILL
        cell.alignment = HDR_ALIGN
        ws.merge_cells(start_row=label_row, start_column=BLOCK3_START, end_row=label_row, end_column=BLOCK3_START + len(BLOCK3_HEADERS) - 1)

    # --- Row 3 (excel_header_row): Block 1 headers (already exist, just re-style) ---
    for col in range(1, 17):
        cell = ws.cell(row=excel_header_row, column=col)
        cell.fill = BLOCK1_HDR_FILL
        cell.font = HDR_FONT_DARK
        cell.alignment = HDR_ALIGN

    # --- Row 3: Block 2 headers ---
    for i, hdr in enumerate(BLOCK2_HEADERS):
        col = BLOCK2_START + i
        cell = ws.cell(row=excel_header_row, column=col, value=hdr)
        cell.fill = BLOCK2_HDR_FILL
        cell.font = HDR_FONT_DARK
        cell.alignment = HDR_ALIGN

    # --- Row 3: Block 3 headers ---
    for i, hdr in enumerate(BLOCK3_HEADERS):
        col = BLOCK3_START + i
        cell = ws.cell(row=excel_header_row, column=col, value=hdr)
        cell.fill = BLOCK3_HDR_FILL
        cell.font = HDR_FONT_DARK
        cell.alignment = HDR_ALIGN

    # --- Fill data rows ---
    last_col = BLOCK3_START + len(BLOCK3_HEADERS) - 1  # v15.11: was col BE, now col BF

    data_row_idx = 0
    for excel_row in range(excel_header_row + 1, ws.max_row + 1):
        header_id = ws.cell(row=excel_row, column=COL_HEADER_ID + 1).value
        if header_id is None:
            continue
        if data_row_idx >= len(results):
            break

        r = results[data_row_idx]
        g = gate_results[data_row_idx]
        data_row_idx += 1

        # --- Block 1: Write combo + emp_no ---
        ws.cell(row=excel_row, column=COL_DIST_COMBO + 1, value=r.combo)
        # Labadi RULE 1 (2026-06-09): emp_no is ALWAYS written — never blank, even
        # for sponsorship/dependent/CHD rows. It represents the AlJeel employee the
        # booking is assigned to (the requestor for sponsorship, the sponsoring
        # employee for dependents). This REVERSES the prior v15.13 rule (which kept
        # the Oracle Employee No column blank) and the v29 dependent-blanking guard.
        ws.cell(row=excel_row, column=COL_EMP_NO + 1, value=r.emp_no)

        # v15.11 (Amr May 25): Static-field overrides per Jawal Oracle ingestion spec.
        ws.cell(row=excel_row, column=COL_HEADER_ID  + 1, value=JAWAL_INVOICE_HEADER_ID)   # col A = 1
        ws.cell(row=excel_row, column=COL_SUPPLIER   + 1, value=JAWAL_SUPPLIER_AR)         # col G (Arabic)
        ws.cell(row=excel_row, column=COL_SUPP_NO    + 1, value=JAWAL_SUPPLIER_NUMBER)     # col H = 10394
        ws.cell(row=excel_row, column=COL_SUPP_SITE  + 1, value=JAWAL_SUPPLIER_SITE_AR)    # col I (Arabic)
        # v15.11.1: Invoice Type now refund-aware
        _inv_amt_val = ws.cell(row=excel_row, column=COL_AMOUNT_HDR + 1).value
        _inv_no_val  = ws.cell(row=excel_row, column=COL_INV_NO     + 1).value
        ws.cell(row=excel_row, column=COL_INV_TYPE   + 1, value=_v15_11_1_invoice_type(_inv_amt_val, _inv_no_val))  # col J

        # --- Block 2: Code/Name expansion ---
        expansion = lookup.expand_combo(r.combo, resolved_line=r)
        # v15.11.2: inject GL Description into expansion dict so the
        # generic loop below writes it like every other Block 2 column.
        expansion["GL Description"] = r.gl_description
        for i, key in enumerate(BLOCK2_HEADERS):
            col = BLOCK2_START + i
            ws.cell(row=excel_row, column=col, value=expansion.get(key, ""))

        # --- Block 3: Debug/Agent columns ---
        row_status = row_statuses[data_row_idx - 1] if (data_row_idx - 1) < len(row_statuses) else "YELLOW"
        flag_str = g.flag_string
        # v15.11: merge custom flags that sit on r.flags but are not in gate output
        # (MISSING_HR_APPROVAL, annual-ticket segment flags, sponsorship+sanad flags, FAMILY_CLUSTER_*).
        _V15_11_EXTRA_FLAGS = {
            "MISSING_HR_APPROVAL",
            "INCOMPLETE_SPONSORSHIP_APPROVAL",
            "EXTERNAL_SPONSORSHIP_NO_EMP_NUM",
            "SHARED_SPONSORSHIP_50_50",
            "SANAD_NO_SCREENSHOT",
            "SANAD_TIMING_MISMATCH",
            "ANNUAL_TICKET_HOME_SEGMENTS",
            "ANNUAL_TICKET_HOME_CC_MISSING",
            "SPONSORSHIP_CHARGED_TO_REQUESTER_v15.11",
            "FAMILY_CLUSTER_DETECTED",
            "FAMILY_CLUSTER_CHD_CONFIRMED",
            "FAMILY_CLUSTER_NO_CHD",
            "MIXED_FAMILY_CLUSTER",
            "EMP_FROM_REF_NO",
            "EMP_FROM_FAMILY_FOLDER",
            "DEPENDENT_USES_SPONSOR_EMP",
            "FAMILY_CLUSTER_CHD_USES_SPONSOR_EMP",
        }
        existing = {f.strip() for f in (flag_str or "").split("|")}
        # v15.11.2: also surface parameterized FAMILY_CLUSTER_UNIFIED flag
        # (e.g. FAMILY_CLUSTER_UNIFIED(emp_no=1002066,modal)).
        def _is_v15_11_extra(f):
            return (
                f in _V15_11_EXTRA_FLAGS
                or (isinstance(f, str) and f.startswith("FAMILY_CLUSTER_UNIFIED"))
            )
        extra = [f for f in (r.flags or []) if _is_v15_11_extra(f) and f not in existing]
        if extra:
            if flag_str and flag_str != "CLEAN":
                flag_str = flag_str + " | " + " | ".join(extra)
            else:
                flag_str = " | ".join(extra)
        action_str = g.action
        vr = getattr(r, '_validator', {}) or {}

        # Check for NO_FOLDER, NO_APPROVAL, FORM_NOT_FOUND_IN_EMAIL, PERSONAL_CONTRIB_SELF_APPROVAL, EMAIL_CORRUPT_OR_EMPTY
        is_no_folder = False
        is_no_approval = False
        is_form_missing = False
        is_self_approved = False
        is_corrupt_or_empty_file = False
        is_shared_opex = False
        
        line_catch_list = line_catches.get(r.sl_no, [])
        all_line_flag_codes = set(r.flags or [])
        for c in line_catch_list:
            all_line_flag_codes.add(c["category"])
            
        if "NO_FOLDER" in all_line_flag_codes:
            is_no_folder = True
        if "NO_APPROVAL" in all_line_flag_codes:
            is_no_approval = True
        if "FORM_NOT_FOUND_IN_EMAIL" in all_line_flag_codes:
            is_form_missing = True
        if "PERSONAL_CONTRIB_SELF_APPROVAL" in all_line_flag_codes:
            is_self_approved = True
        if "EMAIL_CORRUPT_OR_EMPTY" in all_line_flag_codes:
            is_corrupt_or_empty_file = True
        if "SHARED_OPEX_SPONSORSHIP" in all_line_flag_codes:
            is_shared_opex = True

        evidence_folder_status = "SHARED_OPEX" if is_shared_opex else ("MISSING" if is_no_folder else "OK")
        approval_email_status = "CORRUPT_OR_EMPTY" if is_corrupt_or_empty_file else ("MISSING" if is_no_approval else ("FORM_MISSING" if is_form_missing else "OK"))
        self_approval_status = "SELF_APPROVED" if is_self_approved else "OK"
        
        # Human Action Note
        human_review_note = get_human_action_recommendation(all_line_flag_codes, getattr(r, 'sol_flag', ''), row_status)

        block3_values = [
            row_status,                                     # Row Status
            getattr(r, 'sol_flag', ''),                     # Manpower Allocation Status
            evidence_folder_status,                         # Evidence Folder Status
            approval_email_status,                          # Approval Email Status
            self_approval_status,                           # Self-Approval Status
            human_review_note,                              # Human Review Note
            None,                                           # QC Catches (filled below, index 6)
            flag_str,                                       # Agent Flags
            action_str,                                     # Agent Action
            vr.get("emp_match_source", ""),                 # Agent Emp Match Source
            r.emp_match_method,                             # Agent Match Method
            r.account_rule,                                 # Agent Account Rule
            f"Co={r.company} Loc={r.location} Acc={r.account} CC={r.cost_center} DIV={r.div} Sol={r.solution} Ag={r.agency} Proj={r.project} IC={r.intercompany} F1={r.future1}",  # Agent Segments Breakdown
            vr.get("form_emp_no", ""),                      # Form Emp No
            vr.get("form_approver", ""),                    # Form Approver
            vr.get("form_division", ""),                    # Form Division
            vr.get("form_agency", ""),                      # Form Agency
            vr.get("form_solution", ""),                    # Form Solution
            vr.get("form_cost_center_ref", ""),             # Form Cost-Center-Ref
            vr.get("validator_status", ""),                 # Validator Status
            vr.get("discrepancy_detail", ""),               # Discrepancy Detail
            getattr(r, '_v2_layer', ''),                    # Resolution Layer
            getattr(r, '_v2_confidence', ''),               # Resolution Confidence
            getattr(r, '_v2_trace', ''),                    # Resolution Trace
            getattr(r, '_v2_flag', ''),                     # Resolution Flag
            getattr(r, '_v2_email', ''),                    # Email Match
            getattr(r, '_trip_purpose', ''),                # Trip Purpose
            getattr(r, '_trip_confidence', ''),             # Trip Purpose Confidence
            getattr(r, '_trip_signals', ''),                # Trip Purpose Signals
            getattr(r, '_trip_trace', ''),                  # Trip Purpose Trace
            getattr(r, '_trip_account_override', ''),       # Trip Account Override
            getattr(r, '_opex_serial', 'N/A'),              # OPEX Serial (RULE 4)
            getattr(r, '_mai_approval_status', 'N/A'),      # HR Approval (Mai) (RULE 5)
            getattr(r, '_sanad_approval_status', 'N/A'),    # Sanad Approval (RULE 5)
        ]
        # Block 3 value-column indices for the new Labadi columns (for colouring).
        _idx_opex_serial = len(block3_values) - 3
        _idx_mai = len(block3_values) - 2
        _idx_sanad = len(block3_values) - 1

        # QC catches summary
        if line_catch_list:
            catch_summary = "; ".join(f"{c['category']}({c.get('severity','?')})" for c in line_catch_list)
            block3_values[6] = catch_summary  # QC Catches slot is index 6

        for i, val in enumerate(block3_values):
            col = BLOCK3_START + i
            cell = ws.cell(row=excel_row, column=col, value=val)
            # Custom cells highlight
            if i == 2:  # Evidence Folder Status
                if val == "MISSING":
                    cell.fill = RED_FILL
                elif val == "SHARED_OPEX":
                    cell.fill = AMBER_FILL
                elif val == "OK":
                    cell.fill = GREEN_FILL
            elif i == 3:  # Approval Email Status
                if val == "MISSING" or val == "CORRUPT_OR_EMPTY":
                    cell.fill = RED_FILL
                elif val == "FORM_MISSING":
                    cell.fill = AMBER_FILL
                elif val == "OK":
                    cell.fill = GREEN_FILL
            elif i == 4:  # Self-Approval Status
                if val == "SELF_APPROVED":
                    cell.fill = RED_FILL
                elif val == "OK":
                    cell.fill = GREEN_FILL

        # --- Row color across ALL columns (A through last_col) ---
        row_fill, row_font = get_row_style(row_status)
        for col in range(1, last_col + 1):
            cell = ws.cell(row=excel_row, column=col)
            cell.fill = row_fill
            cell.border = BORDER_THIN

        # --- Labadi RULE 4 & 5 cell colouring (applied AFTER the row fill so the
        # OPEX-serial / Mai / Sanad statuses stay visible): OK=green, MISSING=red,
        # N/A or unknown=grey. ---
        def _labadi_fill(value):
            v = str(value or "")
            if v == "OK":
                return GREEN_FILL
            if v == "MISSING":
                return RED_FILL
            if v in ("N/A", ""):
                return GREY_FILL
            return None  # a real serial value — leave the row colour
        for _li in (_idx_opex_serial, _idx_mai, _idx_sanad):
            _f = _labadi_fill(block3_values[_li])
            if _f is not None:
                ws.cell(row=excel_row, column=BLOCK3_START + _li).fill = _f
                ws.cell(row=excel_row, column=BLOCK3_START + _li).border = BORDER_THIN

    # --- Column widths ---
    # Block 1: Distribution Combination gets extra width
    ws.column_dimensions[get_column_letter(COL_DIST_COMBO + 1)].width = 58
    # Block 2: reasonable widths
    for i, hdr in enumerate(BLOCK2_HEADERS):
        col = BLOCK2_START + i
        if hdr == "GL Description":  # v15.11.2: description-only column
            ws.column_dimensions[get_column_letter(col)].width = 70
        elif hdr in ("GL", "Cost Name", "Contribution", "Solution Name", "Agency Name"):
            ws.column_dimensions[get_column_letter(col)].width = 30
        else:
            ws.column_dimensions[get_column_letter(col)].width = 14
    # Block 3: all get 28
    for i in range(len(BLOCK3_HEADERS)):
        col = BLOCK3_START + i
        ws.column_dimensions[get_column_letter(col)].width = 28

    wb.save(out_path)
    print(f"\n[process_batch] Output: {out_path}")
    print(f"  {out_path.stat().st_size} bytes")

    # --- Write halt queue ---
    if halt_queue:
        halt_path = ROOT / "qc" / "halt-queue"
        halt_path.mkdir(parents=True, exist_ok=True)
        halt_file = halt_path / f"{inv_no or 'batch'}.json"
        with open(halt_file, "w") as f:
            json.dump(halt_queue, f, indent=2, default=str)
        print(f"  Halt queue: {halt_file} ({len(halt_queue)} lines)")

    # --- Summary ---
    n_total = len(results)
    n_clean = sum(1 for g in gate_results if g.flag_string == "CLEAN")
    n_hard_fail = sum(1 for g in gate_results if not g.passed_hard)
    n_soft_flag = sum(1 for g in gate_results if g.passed_hard and g.soft_flags)
    flag_counter = Counter()
    for g in gate_results:
        for _, code in g.hard_failures + g.soft_flags:
            flag_counter[code] += 1

    match_counter = Counter(r.emp_match_method for r in results)
    account_counter = Counter(r.account for r in results)

    summary = {
        "invoice_no": inv_no,
        "processed_at": datetime.now(timezone.utc).isoformat(),
        "total_lines": n_total,
        "clean_lines": n_clean,
        "hard_failures": n_hard_fail,
        "soft_flags": n_soft_flag,
        "flag_breakdown": dict(flag_counter.most_common()),
        "match_method_breakdown": dict(match_counter),
        "account_breakdown": dict(account_counter),
        "output_file": str(out_path),
        "halt_queue_count": len(halt_queue),
        "row_status_breakdown": dict(status_counter),
        "within_batch_catches": len(within_batch_catches),
        "cross_batch_catches": len(cross_batch_catches),
        "catches_by_category": dict(Counter(
            c["category"] for c in within_batch_catches + cross_batch_catches
        ).most_common()),
    }

    summary_path = output_dir / f"summary-{output_suffix}.json"
    with open(summary_path, "w") as f:
        json.dump(summary, f, indent=2)
    print(f"  Summary: {summary_path}")

    print(f"\n=== BATCH SUMMARY ===")
    print(f"  Total lines:    {n_total}")
    print(f"  Clean (Post):   {n_clean}")
    print(f"  Hard failures:  {n_hard_fail}")
    print(f"  Soft flags:     {n_soft_flag}")
    print(f"  Flag breakdown: {dict(flag_counter.most_common())}")
    print(f"  Match methods:  {dict(match_counter)}")

    # Labadi QC self-check (2026-06-09) — verify RULE 1/4/5/6 on the written output.
    try:
        qc = run_labadi_qc_check(batch_dir, output_path=out_path)
        summary["labadi_qc_check"] = qc
        with open(summary_path, "w") as f:
            json.dump(summary, f, indent=2)
    except Exception as exc:
        print(f"[labadi-qc] WARNING — QC check failed: {exc}")

    return summary



# =============================================================================
# v15.11 (Amr May 25): helper utilities for static fields, HR approval flag,
# annual-ticket / personal segment rollup, and final GL-string column.
# =============================================================================

def _v15_11_check_mai_approval(form_data, approver_name):
    """Return True if Mai is present in the approval chain.

    Mai is the HR business partner. Looks ONLY at approver-name-like fields
    (NOT free-form email bodies — too many false positives from email
    addresses like "mailto:" / "domain.com"). Uses word-boundary regex.

    Approval discipline: training trips and annual-ticket rows MUST carry
    a Mai approval; missing = soft flag.
    """
    import re as _re
    parts = []
    if approver_name:
        parts.append(str(approver_name))
    if isinstance(form_data, dict):
        # Only inspect explicit approver/HR fields — never the full body.
        for k in ("approver_name", "approver", "hr_approver", "approver_chain", "approvers", "hr_business_partner"):
            v = form_data.get(k)
            if v:
                if isinstance(v, list):
                    parts.extend(str(x) for x in v)
                else:
                    parts.append(str(v))
    if not parts:
        return False
    blob = " ".join(parts)
    # Word-bounded match: "Mai" as a standalone name (not inside "email", "domain", "mailto").
    if _re.search(r"\bmai\b", blob, _re.IGNORECASE):
        return True
    # Arabic "مي" (standalone word, surrounded by Arabic word boundaries)
    if _re.search(r"(?<![\u0600-\u06ff])\u0645\u064a(?![\u0600-\u06ff])", blob):
        return True
    return False


def _v15_11_apply_personal_or_annual_rollup(r, md=None):
    """Stamp annual-ticket rows with the employee home annual-ticket segments.

    Current finance rule for 21070229 rows:
      CC = resolved employee's home cost center from Manpower
      DIV = 888, Solution = 00000, Agency = 88888

    This replaces the stale v15.11 General-segment rollup, which incorrectly
    zeroed CC/DIV/Solution/Agency for rows that can ship directly from
    process_batch without a later run_v30 correction pass.
    """
    if r.account != "21070229":
        return

    emp = None
    if md is not None and getattr(r, "emp_no", None):
        try:
            emp = md.employees.get(int(r.emp_no))
        except (TypeError, ValueError):
            emp = None

    if emp is not None and getattr(emp, "cost_center", None):
        r.cost_center = str(emp.cost_center).zfill(6)
        flag = "ANNUAL_TICKET_HOME_SEGMENTS"
    else:
        flag = "ANNUAL_TICKET_HOME_CC_MISSING"

    r.div = ANNUAL_TICKET_DIV
    r.solution = ANNUAL_TICKET_SOLUTION
    r.agency = ANNUAL_TICKET_AGENCY

    if "ACCOUNT_GENERAL_SEGMENTS_v15.11" in r.flags:
        r.flags = [f for f in r.flags if f != "ACCOUNT_GENERAL_SEGMENTS_v15.11"]
    if flag not in r.flags:
        r.flags.append(flag)


def _v15_11_sanad_checks(form_data, raw_dir, ticket_no):
    """Sanad-related soft flags. Detection scaffold; returns flag list.

    SANAD_NO_SCREENSHOT  — Sanad approval mentioned but no attached screenshot.
    SANAD_TIMING_MISMATCH — Sanad approval flight timing != vendor invoice timing.

    NOTE (v15.11 May 25): full implementation requires OCR on attached
    approval screenshots; this is a SCAFFOLD that fires SANAD_NO_SCREENSHOT
    when the form references Sanad but attachment_names has no image, and
    leaves SANAD_TIMING_MISMATCH as always-False until OCR is wired in.
    """
    flags = []
    if not isinstance(form_data, dict):
        return flags
    body = (form_data.get("body_text") or "") + " " + (form_data.get("subject") or "")
    body_low = body.lower()
    is_sanad = ("sanad" in body_low) or ("سند" in body)  # English + Arabic "sanad"
    if not is_sanad:
        return flags
    attachment_names = form_data.get("attachment_names") or []
    has_image = any(
        str(n).lower().endswith((".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp"))
        for n in attachment_names
    )
    if not has_image:
        flags.append("SANAD_NO_SCREENSHOT")
    # SANAD_TIMING_MISMATCH stub: always False pending OCR wiring.
    return flags


def _v15_11_shared_sponsorship_check(form_data):
    """Detect 50/50 shared sponsorship splits indicated in the form body.

    Returns True if form body explicitly mentions 50/50 split, partner share,
    or shared sponsorship language.
    """
    if not isinstance(form_data, dict):
        return False
    body = (form_data.get("body_text") or "") + " " + (form_data.get("subject") or "")
    body_low = body.lower()
    patterns = (
        "50/50", "50 / 50", "50:50", "fifty fifty", "fifty-fifty",
        "shared sponsorship", "shared sponsor", "co-sponsor", "co sponsor",
        "joint sponsorship", "partner share", "partner contribution",
        "مشاركة",  # Arabic "sharing"
    )
    return any(p in body_low for p in patterns)



def run_labadi_qc_check(batch_id, output_path=None):
    """Labadi QC self-check (2026-06-09). Inspects the produced Oracle output
    workbook and prints four verification counts to stdout:

      1. emp_no blank rows                     -> must be 0 (RULE 1)
      2. OPEX Serial == MISSING                -> count (RULE 4)
      3. account==21070229 rows missing Mai or Sanad approval -> count (RULE 5)
      4. |*Amount (col M, ex-VAT)| exceeding |*Invoice Amount (col E, incl-VAT
         invoice total)| -> count (RULE 6; expected 0)

    Returns a dict of the counts. Reads column anchors by the fixed Oracle
    indices (COL_EMP_NO / COL_AMOUNT / COL_AMOUNT_HDR) and the Block-3 columns by
    header text so it survives Block-3 reordering.
    """
    path = Path(output_path) if output_path else None
    if path is None:
        bdir = Path(batch_id)
        if not bdir.is_absolute():
            bdir = ROOT / bdir
        out_glob = sorted((bdir / "output").glob("*.xlsx")) if (bdir / "output").is_dir() else []
        path = out_glob[-1] if out_glob else None
    if not path or not Path(path).exists():
        print(f"[labadi-qc] SKIP — output workbook not found for {batch_id}")
        return {}

    wb = load_workbook(path, data_only=True)
    ws = wb.active

    # Locate the header row (row carrying "Employee No" + an "Amount" column).
    header_row = None
    for rr in range(1, min(ws.max_row or 1, 8) + 1):
        rowvals = [str(ws.cell(rr, c).value or "") for c in range(1, (ws.max_column or 1) + 1)]
        if any("Employee No" in v for v in rowvals) and any("Amount" in v for v in rowvals):
            header_row = rr
            break
    if header_row is None:
        header_row = 3

    hdr_map = {}
    for c in range(1, (ws.max_column or 1) + 1):
        h = str(ws.cell(header_row, c).value or "").strip()
        if h:
            hdr_map[h] = c
    col_opex_serial = hdr_map.get("OPEX Serial")
    col_mai = hdr_map.get("HR Approval (Mai)")
    col_sanad = hdr_map.get("Sanad Approval")

    blank_emp = opex_serial_missing = annual_appr_missing = amount_violation = rows_checked = 0
    for rr in range(header_row + 1, (ws.max_row or header_row) + 1):
        header_id = ws.cell(rr, COL_HEADER_ID + 1).value
        if header_id is None or str(header_id).strip() == "":
            continue
        rows_checked += 1

        emp_val = ws.cell(rr, COL_EMP_NO + 1).value
        if emp_val is None or str(emp_val).strip() in ("", "0", "-", "0000000", "None", "nan"):
            blank_emp += 1

        if col_opex_serial and str(ws.cell(rr, col_opex_serial).value or "").strip().upper() == "MISSING":
            opex_serial_missing += 1

        mai_v = str(ws.cell(rr, col_mai).value or "").strip().upper() if col_mai else ""
        sanad_v = str(ws.cell(rr, col_sanad).value or "").strip().upper() if col_sanad else ""
        if mai_v == "MISSING" or sanad_v == "MISSING":
            annual_appr_missing += 1

        amt = _safe_float(ws.cell(rr, COL_AMOUNT + 1).value)
        inv_amt = _safe_float(ws.cell(rr, COL_AMOUNT_HDR + 1).value)
        # Magnitude comparison so refund/credit (negative) rows are handled too.
        # Equality is allowed (zero-rated lines: ex-VAT == incl-VAT). Only a line
        # ex-VAT amount LARGER than the whole-invoice total is a real violation.
        if abs(amt) - abs(inv_amt) > 0.01:
            amount_violation += 1

    print("\n=== LABADI QC CHECK (2026-06-09) ===")
    print(f"  Workbook: {path}")
    print(f"  Rows checked: {rows_checked}")
    print(f"  1. Blank emp_no rows: {blank_emp}  (expected 0)  {'FAIL ⚠️' if blank_emp else 'OK ✓'}")
    print(f"  2. OPEX Serial MISSING: {opex_serial_missing}")
    print(f"  3. Annual/family (21070229) rows missing Mai/Sanad approval: {annual_appr_missing}")
    print(f"  4. Rows where |*Amount(M)| > |*Invoice Amount(E)|: {amount_violation}  (expected 0)  {'FAIL ⚠️' if amount_violation else 'OK ✓'}")

    return {
        "rows_checked": rows_checked,
        "blank_emp_no": blank_emp,
        "opex_serial_missing": opex_serial_missing,
        "annual_approval_missing": annual_appr_missing,
        "amount_exceeds_invoice": amount_violation,
    }


def main():
    parser = argparse.ArgumentParser(description="Process Jawal batch with full 10-segment combo")
    parser.add_argument("--batch", required=True, help="Path to batch directory")
    parser.add_argument("--master-data", default=str(ROOT / "qc" / "master-data" / "Aljeel_Lookups-v2.xlsx"),
                        help="Path to Master Data xlsx")
    parser.add_argument("--reference", default=str(ROOT / "qc" / "master-data" / "Aljeel_Lookups-v2.xlsx"),
                        help="Path to reference xlsx (INDEX/DIV/Agency/Solution/CC tabs)")
    parser.add_argument("--suffix", default="v5-rulebook", help="Output filename suffix")
    parser.add_argument("--raw-dir", default=None, help="Path to raw .msg files directory (for allocation resolution)")
    parser.add_argument("--no-cache", action="store_true",
                        help="Bypass persistent stage-1 cache reads")
    parser.add_argument("--invoice-file", default=None,
                        help="Explicit path to invoice xlsx; overrides disk glob")
    args = parser.parse_args()

    batch_dir = Path(args.batch)
    if not batch_dir.is_absolute():
        batch_dir = ROOT / batch_dir

    raw_dir = Path(args.raw_dir) if args.raw_dir else None
    if raw_dir and not raw_dir.is_absolute():
        raw_dir = ROOT / raw_dir

    invoice_file = Path(args.invoice_file) if args.invoice_file else None

    process_batch(
        batch_dir=batch_dir,
        master_data_path=Path(args.master_data),
        reference_path=Path(args.reference),
        output_suffix=args.suffix,
        raw_dir=raw_dir,
        no_cache=args.no_cache,
        invoice_file=invoice_file,
    )


if __name__ == "__main__":
    main()
