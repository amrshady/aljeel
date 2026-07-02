#!/usr/bin/env python3
"""
Read-only holistic QA agent for completed Jawal batches.

Loads compact batch inputs, asks Gemini for a structured audit, saves the raw
JSON report, and streams a human-readable report to stdout.
"""
from __future__ import annotations

import argparse
import io
import json
import os
import re
import socket
import time
import urllib.error
import urllib.request
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
GEMINI_BASE_URL = os.environ.get("GEMINI_BASE_URL", "https://gateway.ai.cloudflare.com/v1/3724a3e71944b366a39b3735aa117a58/accord-aljeel-ap/google-ai-studio/v1beta")


# Long-context (>200K input) doubles Gemini 3.x Pro rates; ALL tokens billed at long-context rate.
GEMINI_PRICING_LONG = {
    "gemini-pro-latest":      {"in": 4.00/1e6, "out": 18.00/1e6},
    "gemini-3.1-pro-preview": {"in": 4.00/1e6, "out": 18.00/1e6},
    "gemini-2.5-pro":         {"in": 2.50/1e6, "out": 15.00/1e6},
    "gemini-2.5-flash":       {"in": 0.30/1e6, "out": 2.50/1e6},
}
def _gemini_rate(model, in_tokens=0):
    if in_tokens and in_tokens > 200_000:
        return GEMINI_PRICING_LONG.get(model) or GEMINI_PRICING_LONG.get("gemini-2.5-pro")
    return GEMINI_PRICING.get(model) or GEMINI_PRICING.get("gemini-2.5-pro")
def _cf_custom_cost_header(model, in_tokens=0):
    import json as __j
    _p = _gemini_rate(model, in_tokens)
    return __j.dumps({"per_token_in": _p["in"], "per_token_out": _p["out"]})
MODEL = "gemini-3.1-pro-preview"
GEMINI_MAX_OUTPUT_TOKENS = 8192
GEMINI_THINKING_BUDGET = 4096
GEMINI_REQUEST_TIMEOUT_SECONDS = 180
GEMINI_MAX_ATTEMPTS = 3
GEMINI_RETRY_DELAY_SECONDS = 5
GEMINI_RETRYABLE_HTTP_CODES = {500, 503}
MAX_ROWS = 200

KNOWLEDGE_FILES = [
    "catch-rules.md",
    "vendor-handlers.md",
    "known-issues.md",
    "master-data.md",
    "gl-account-map.md",
    "escalation-paths.md",
]

GL_RULES = """\
- 60301003: Travel Tickets (S&M/Business)
- 60301004: Travel G&A (DIV 888 / DIV 190)
- 60307021: Sponsoring Expenses (external parties, emp_no must be blank)
- 60308007: Recruitment Fees
- 60308009: Training Expenses
- 21070227: Accrued Project & Warranty
- 21070229: Accrued Employee Annual Tickets (personal/family clusters)
"""


def qa_print(message: str = "") -> None:
    print(f"[QA] {message}", flush=True)


def _load_dotenv_key(name: str) -> str:
    val = os.environ.get(name, "")
    if val:
        return val
    env_path = Path("/home/clawdbot/.openclaw/.env")
    if env_path.exists():
        for line in env_path.read_text(errors="replace").splitlines():
            if line.startswith(f"{name}="):
                return line.split("=", 1)[1].strip().strip('"').strip("'")
    return ""


def _is_timeout_urlerror(exc: urllib.error.URLError) -> bool:
    reason = getattr(exc, "reason", None)
    return (
        isinstance(reason, socket.timeout)
        or isinstance(reason, TimeoutError)
        or "timed out" in str(reason).lower()
        or "timeout" in str(reason).lower()
    )


def _call_gemini_direct(prompt_text: str, api_key: str, model: str) -> dict:
    """Call Gemini API directly. Self-contained copy of the project pattern."""
    url = (
        f"{GEMINI_BASE_URL}/"
        f"models/{model}:generateContent?key={api_key}"
    )
    payload = json.dumps({
        "contents": [{"parts": [{"text": prompt_text}]}],
        "generationConfig": {
            "temperature": 0.1,
            "maxOutputTokens": GEMINI_MAX_OUTPUT_TOKENS,
            "thinkingConfig": {"thinkingBudget": GEMINI_THINKING_BUDGET},
        },
    }).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=payload,
        headers={"Content-Type": "application/json", "User-Agent": "AlJeel-AP-Agent/1.0", "cf-aig-custom-cost": _cf_custom_cost_header(model)},
    )
    for attempt in range(1, GEMINI_MAX_ATTEMPTS + 1):
        try:
            with urllib.request.urlopen(req, timeout=GEMINI_REQUEST_TIMEOUT_SECONDS) as resp:
                if resp.status in GEMINI_RETRYABLE_HTTP_CODES:
                    body = resp.read()
                    raise urllib.error.HTTPError(
                        url, resp.status, "Retryable Gemini response", resp.headers, io.BytesIO(body)
                    )
                return json.loads(resp.read())
        except socket.timeout:
            if attempt == GEMINI_MAX_ATTEMPTS:
                raise
            qa_print(
                f"WARNING: Gemini timed out after {GEMINI_REQUEST_TIMEOUT_SECONDS}s "
                f"(attempt {attempt}/{GEMINI_MAX_ATTEMPTS}); retrying..."
            )
            time.sleep(GEMINI_RETRY_DELAY_SECONDS)
        except urllib.error.HTTPError as exc:
            if exc.code not in GEMINI_RETRYABLE_HTTP_CODES or attempt == GEMINI_MAX_ATTEMPTS:
                raise
            qa_print(
                f"WARNING: Gemini returned HTTP {exc.code} "
                f"(attempt {attempt}/{GEMINI_MAX_ATTEMPTS}); retrying..."
            )
            time.sleep(GEMINI_RETRY_DELAY_SECONDS)
        except urllib.error.URLError as exc:
            if not _is_timeout_urlerror(exc) or attempt == GEMINI_MAX_ATTEMPTS:
                raise
            qa_print(
                f"WARNING: Gemini timed out after {GEMINI_REQUEST_TIMEOUT_SECONDS}s "
                f"(attempt {attempt}/{GEMINI_MAX_ATTEMPTS}); retrying..."
            )
            time.sleep(GEMINI_RETRY_DELAY_SECONDS)
    raise RuntimeError("Gemini request retry loop exited unexpectedly")


def _gemini_text(resp: dict) -> str:
    candidates = resp.get("candidates") or []
    if not candidates:
        return ""
    parts = candidates[0].get("content", {}).get("parts", [])
    return " ".join(part.get("text", "") for part in parts if "text" in part).strip()


def _extract_json(text: str) -> dict:
    cleaned = text.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned, flags=re.I)
        cleaned = re.sub(r"\s*```$", "", cleaned)
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        start = cleaned.find("{")
        end = cleaned.rfind("}")
        if start >= 0 and end > start:
            return json.loads(cleaned[start:end + 1])
        raise


def _normalise_header(value) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def _find_header_row(ws) -> tuple[int, list[str]]:
    best_row = 1
    best_headers: list[str] = []
    for idx, row in enumerate(ws.iter_rows(max_row=20, values_only=True), 1):
        headers = [_normalise_header(cell) for cell in row]
        header_set = set(headers)
        score = len(header_set & {
            "invoice_header_identifier",
            "business_unit",
            "invoice_number",
            "description",
            "amount",
            "employee_no",
            "account",
            "cost_center",
        })
        if score > len(set(best_headers) & {
            "invoice_header_identifier",
            "business_unit",
            "invoice_number",
            "description",
            "amount",
            "employee_no",
            "account",
            "cost_center",
        }):
            best_row = idx
            best_headers = headers
    return best_row, best_headers


def _load_excel_rows(path: Path, max_rows: int = MAX_ROWS) -> list[dict]:
    """Load xlsx rows as list of dicts. Returns at most max_rows."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header_row, headers = _find_header_row(ws)
    rows: list[dict] = []
    for excel_row_num, values in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
        if len(rows) >= max_rows:
            break
        if not any(v is not None and str(v).strip() != "" for v in values):
            continue
        row = {"_excel_row": excel_row_num, "_row_num": len(rows) + 1}
        for idx, header in enumerate(headers):
            if not header:
                continue
            if idx < len(values):
                row[header] = values[idx]
        rows.append(row)
    return rows


def _val(row: dict, *names: str) -> str:
    for name in names:
        value = row.get(name)
        if value is not None and str(value).strip() != "":
            return str(value).strip()
    return ""


def _amount(row: dict) -> str:
    raw = _val(row, "amount", "invoice_amount")
    try:
        return f"{float(raw):.2f}"
    except (TypeError, ValueError):
        return raw


def _parse_description(desc: str) -> tuple[str, str, str]:
    ticket_match = re.search(r"\(([^()]+)\)\s*$", desc or "")
    ticket_id = ticket_match.group(1).strip() if ticket_match else ""
    body = re.sub(r"\s*\([^()]+\)\s*$", "", desc or "").strip()
    if " - " in body:
        passenger, route = body.split(" - ", 1)
    else:
        passenger, route = body, ""
    return passenger.strip(), route.strip(), ticket_id


def _build_input_tsv(rows: list[dict]) -> str:
    """Compact TSV of input columns."""
    lines = ["row_num\temp_no\tpassenger_name\tticket_id\troute\tamount_sar\tref_no"]
    for row in rows[:MAX_ROWS]:
        desc = _val(row, "description")
        passenger, route, ticket_id = _parse_description(desc)
        lines.append("\t".join([
            str(row.get("_row_num", "")),
            _val(row, "employee_no", "emp_no"),
            passenger,
            ticket_id,
            route,
            _amount(row),
            _val(row, "invoice_number", "ref_no") or ticket_id,
        ]))
    if len(rows) > MAX_ROWS:
        lines.append(f"[truncated after {MAX_ROWS} rows]")
    return "\n".join(lines)


def _fraud_lookup(fraud_json: dict) -> dict[int, dict]:
    out: dict[int, dict] = {}
    verdicts = fraud_json.get("ai_result", {}).get("per_row_verdicts", [])
    for item in verdicts:
        try:
            out[int(item.get("sl_no"))] = item
        except (TypeError, ValueError):
            continue
    return out


def _portal_lookup(portal_json: dict) -> dict[int, dict]:
    out: dict[int, dict] = {}
    for item in portal_json.get("rows", []) if isinstance(portal_json, dict) else []:
        try:
            out[int(item.get("row_idx"))] = item
        except (TypeError, ValueError):
            continue
    return out


def _build_output_tsv(rows: list[dict], portal_json: dict, fraud_json: dict) -> str:
    """Compact TSV of output columns."""
    portal = _portal_lookup(portal_json)
    fraud = _fraud_lookup(fraud_json)
    lines = [
        "row_num\temp_no\tpassenger_name\tgl_account\tcost_center\tdivision\tagency\tcolor\tflags\tai_fraud_verdict\tai_fraud_category"
    ]
    for row in rows[:MAX_ROWS]:
        row_num = int(row.get("_row_num", len(lines)))
        desc = _val(row, "description")
        passenger, _, _ = _parse_description(desc)
        p = portal.get(row_num, {})
        f = fraud.get(row_num, {})
        flags = p.get("qc_catches") or ""
        if isinstance(flags, list):
            flags = "|".join(str(x) for x in flags)
        lines.append("\t".join([
            str(row_num),
            _val(row, "employee_no", "emp_no") or str(p.get("emp_no", "") or ""),
            passenger,
            _val(row, "account") or str(p.get("account", "") or ""),
            _val(row, "cost_center") or str(p.get("cost_center", "") or ""),
            _val(row, "div", "division") or str(p.get("div", "") or ""),
            _val(row, "agency") or str(p.get("agency", "") or ""),
            str(p.get("row_status", "") or p.get("color", "")),
            str(flags).replace("\t", " "),
            str(f.get("verdict", "")),
            str(f.get("primary_category", "")),
        ]))
    if len(rows) > MAX_ROWS:
        lines.append(f"[truncated after {MAX_ROWS} rows]")
    return "\n".join(lines)


def _load_fraud_summary(fraud_json_path: Path) -> str:
    """Return a short text summary from the fraud JSON."""
    if not fraud_json_path.exists():
        return f"Fraud JSON missing: {fraud_json_path}"
    data = json.loads(fraud_json_path.read_text(encoding="utf-8"))
    verdicts = data.get("ai_result", {}).get("per_row_verdicts", [])
    verdict_counts = Counter(v.get("verdict", "UNKNOWN") for v in verdicts)
    categories = Counter(v.get("primary_category", "UNKNOWN") for v in verdicts if v.get("primary_category") != "NONE")
    risk = sum(float(v.get("value_at_risk_sar") or 0) for v in verdicts if v.get("verdict") != "CLEAN")
    top = ", ".join(f"{cat}={count}" for cat, count in categories.most_common(3)) or "none"
    return (
        f"RED={verdict_counts.get('RED', 0)}, YELLOW={verdict_counts.get('YELLOW', 0)}, "
        f"CLEAN={verdict_counts.get('CLEAN', 0)}\n"
        f"Top categories: {top}\n"
        f"Total SAR at risk: {risk:.2f}"
    )


def _load_evidence_index(raw_dir: Path) -> str:
    """Walk raw/ folder, return folder name + file count lines."""
    if not raw_dir.exists():
        return f"Raw evidence folder missing: {raw_dir}"
    lines: list[str] = []
    for dirpath, _, filenames in os.walk(raw_dir):
        rel = Path(dirpath).relative_to(raw_dir)
        if rel == Path("."):
            continue
        file_count = len([name for name in filenames if not name.startswith(".")])
        lines.append(f"{rel.as_posix()}\t{file_count} files")
        if len(lines) >= 100:
            lines.append("[truncated after 100 folders]")
            break
    return "\n".join(lines) if lines else "(no subfolders found)"


def _load_knowledge() -> tuple[str, str]:
    kb_dir = ROOT / "knowledge"
    parts = []
    catch_rules = ""
    for name in KNOWLEDGE_FILES:
        path = kb_dir / name
        content = path.read_text(encoding="utf-8", errors="replace") if path.exists() else f"[missing: {name}]"
        if name == "catch-rules.md":
            catch_rules = content
        else:
            parts.append(f"## {name}\n{content}")
    return catch_rules, "\n\n".join(parts)


def _batch_compact(batch_id: str) -> str:
    return batch_id.replace("-", "").lower()


def _portal_json_path(batch_id: str) -> Path:
    data_dir = ROOT / "dashboard" / "public" / "data"
    lower = batch_id.lower()
    compact = _batch_compact(batch_id)
    candidates = [
        data_dir / f"{lower}-rows-v30.json",
        data_dir / f"{compact}-rows-v30.json",
    ]
    short = re.sub(r"^j26", "j", compact)
    candidates.append(data_dir / f"{short}-rows-v30.json")
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0]


def _output_spreadsheet_path(batch_dir: Path, batch_id: str) -> Path:
    output_dir = batch_dir / "output"
    exact = output_dir / f"Spreadsheet-{batch_id}-FILLED-v30.xlsx"
    if exact.exists():
        return exact
    matches = sorted(output_dir.glob("Spreadsheet-*-FILLED-*.xlsx"), key=lambda p: p.stat().st_mtime)
    if not matches:
        return exact
    return matches[-1]


def _fraud_json_path(batch_id: str) -> Path:
    fraud_dir = ROOT / "qc" / "ai-poc"
    compact = _batch_compact(batch_id)
    candidates = [
        fraud_dir / f"{compact}-fraud-ai-v162.json",
        fraud_dir / f"{batch_id.lower()}-fraud-ai-v162.json",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0]


def _build_prompt(
    batch_id: str,
    input_tsv: str,
    output_tsv: str,
    fraud_summary: str,
    evidence_index: str,
    catch_rules: str,
    additional_kb: str,
    portal_stats: str,
) -> str:
    return f"""\
SECTION 1 - CONTEXT
You are a senior AP auditor reviewing a completed Jawal Travel AP batch for AlJeel Medical.
You have been given:
- The cleaned input invoice spreadsheet (all rows)
- The final output spreadsheet with all resolved fields
- The AI fraud detector results
- The catch rules and GL account mapping rules
- A summary of evidence folders

Your job is to do a HOLISTIC QA review and identify:
1. GL account assignment errors
2. Cost center / division assignment errors
3. Employee lookup errors or suspicious matches
4. Fraud risk items not caught by the automated pipeline
5. Cross-row patterns (same employee, duplicate routes, unusual clusters)
6. Missing evidence that should block payment
7. Any other anomalies that warrant human review

Be specific: cite row numbers, employee names, and amounts. Do not repeat what the pipeline already flagged - focus on what it may have MISSED or got WRONG.

Batch: {batch_id}

SECTION 2 - GL RULES
{GL_RULES}

SECTION 3 - CATCH RULES
{catch_rules}

ADDITIONAL KNOWLEDGE BASE FILES
{additional_kb}

SECTION 4 - INPUT ROWS
{input_tsv}

SECTION 5 - OUTPUT ROWS
{output_tsv}

SECTION 6 - AI FRAUD SUMMARY
{fraud_summary}

Portal row stats:
{portal_stats}

SECTION 7 - EVIDENCE FOLDER INDEX
{evidence_index}

SECTION 8 - OUTPUT FORMAT
Respond with valid JSON ONLY (no markdown, no code fences):
{{
  "summary": "2-3 sentence overview of QA findings",
  "total_issues_found": <int>,
  "issues": [
    {{
      "severity": "HIGH|MEDIUM|LOW",
      "category": "GL_ERROR|CC_ERROR|LOOKUP_ERROR|FRAUD_RISK|PATTERN|MISSING_EVIDENCE|OTHER",
      "title": "short title",
      "detail": "specific detail with row numbers, names, and amounts",
      "rows_affected": [<row numbers>],
      "recommended_action": "what reviewer should do"
    }}
  ],
  "cross_row_patterns": [
    {{
      "pattern_type": "DUPLICATE_ROUTE|SAME_EMPLOYEE_CLUSTER|AMOUNT_CLUSTER|OTHER",
      "description": "...",
      "rows_affected": [...]
    }}
  ],
  "clean_signal": "statement about rows that look correct and why",
  "model_used": "{MODEL}"
}}
"""


def _cost(prompt_tokens: int, response_tokens: int) -> float:
    # Approximate Gemini Pro blended public-rate style estimate; exact billing is provider-side.
    return (prompt_tokens / 1_000_000 * 1.25) + (response_tokens / 1_000_000 * 10.00)


def _format_rows(rows) -> str:
    if not rows:
        return "[]"
    return "[" + ", ".join(str(x) for x in rows) + "]"


def _print_report(report: dict, batch_id: str, usage: dict) -> None:
    issues = report.get("issues") if isinstance(report.get("issues"), list) else []
    patterns = report.get("cross_row_patterns") if isinstance(report.get("cross_row_patterns"), list) else []
    qa_print("────────────────────────────────────")
    qa_print(f"📋 QA REPORT — {batch_id}")
    qa_print("────────────────────────────────────")
    qa_print(f"Summary: {report.get('summary', '')}")
    qa_print(f"Total issues found: {report.get('total_issues_found', len(issues))}")
    qa_print()
    severity_meta = [
        ("HIGH", "🔴 HIGH SEVERITY"),
        ("MEDIUM", "🟡 MEDIUM SEVERITY"),
        ("LOW", "🟢 LOW SEVERITY"),
    ]
    for severity, label in severity_meta:
        sev_issues = [i for i in issues if str(i.get("severity", "")).upper() == severity]
        qa_print(f"{label} ({len(sev_issues)})")
        for issue in sev_issues:
            qa_print(f"  [{severity}] {issue.get('category', 'OTHER')} — {issue.get('title', '')}")
            qa_print(f"    Rows: {_format_rows(issue.get('rows_affected', []))}")
            qa_print(f"    Detail: {issue.get('detail', '')}")
            qa_print(f"    Action: {issue.get('recommended_action', '')}")
        qa_print()
    qa_print(f"📊 CROSS-ROW PATTERNS ({len(patterns)})")
    for pattern in patterns:
        qa_print(
            f"  {pattern.get('pattern_type', 'OTHER')}: {pattern.get('description', '')} "
            f"(rows: {_format_rows(pattern.get('rows_affected', []))})"
        )
    qa_print()
    qa_print(f"✅ Clean signal: {report.get('clean_signal', '')}")
    qa_print("────────────────────────────────────")
    prompt_tokens = int(usage.get("promptTokenCount") or 0)
    resp_tokens = int(usage.get("candidatesTokenCount") or usage.get("outputTokenCount") or 0)
    qa_print(f"Cost: ~${_cost(prompt_tokens, resp_tokens):.2f} | Prompt: {prompt_tokens}t | Response: {resp_tokens}t")


def main() -> int:
    parser = argparse.ArgumentParser(description="Run holistic QA review for a Jawal batch")
    parser.add_argument("--batch-id", required=True)
    args = parser.parse_args()
    batch_id = args.batch_id.strip()
    batch_dir = ROOT / "batches" / f"jawal-{batch_id}"
    output_dir = batch_dir / "output"

    try:
        qa_print("Loading input spreadsheet...")
        input_path = batch_dir / "Spreadsheet-v4-input.xlsx"
        if not input_path.exists():
            input_path = batch_dir / "Spreadsheet.xlsx"
        input_rows = _load_excel_rows(input_path, MAX_ROWS)

        qa_print("Loading output spreadsheet...")
        output_path = _output_spreadsheet_path(batch_dir, batch_id)
        output_rows = _load_excel_rows(output_path, MAX_ROWS)

        qa_print("Loading AI fraud JSON...")
        fraud_path = _fraud_json_path(batch_id)
        fraud_json = json.loads(fraud_path.read_text(encoding="utf-8")) if fraud_path.exists() else {}
        fraud_summary = _load_fraud_summary(fraud_path)

        qa_print("Loading portal review JSON...")
        portal_path = _portal_json_path(batch_id)
        portal_json = json.loads(portal_path.read_text(encoding="utf-8")) if portal_path.exists() else {}
        portal_rows = portal_json.get("rows", []) if isinstance(portal_json, dict) else []
        portal_stats = (
            f"path={portal_path}; total_rows={portal_json.get('total_rows', len(portal_rows))}; "
            f"llm_rows={portal_json.get('llm_rows', 'unknown')}"
            if portal_json else f"path={portal_path}; missing"
        )

        qa_print("Loading knowledge base...")
        catch_rules, additional_kb = _load_knowledge()

        qa_print("Loading evidence folder index...")
        evidence_index = _load_evidence_index(batch_dir / "raw")

        input_tsv = _build_input_tsv(input_rows)
        output_tsv = _build_output_tsv(output_rows, portal_json, fraud_json)
        prompt = _build_prompt(
            batch_id=batch_id,
            input_tsv=input_tsv,
            output_tsv=output_tsv,
            fraud_summary=fraud_summary,
            evidence_index=evidence_index,
            catch_rules=catch_rules,
            additional_kb=additional_kb,
            portal_stats=portal_stats,
        )
        qa_print(f"Building prompt... (~{len(prompt)} chars)")

        api_key = _load_dotenv_key("GEMINI_API_KEY") or _load_dotenv_key("GOOGLE_API_KEY")
        if not api_key:
            raise RuntimeError("GEMINI_API_KEY not found in environment")

        qa_print(f"Calling Gemini ({MODEL})...")
        raw = _call_gemini_direct(prompt, api_key, MODEL)
        qa_print("Response received. Parsing...")
        text = _gemini_text(raw)
        report = _extract_json(text)
        report.setdefault("model_used", MODEL)

        output_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        report_path = output_dir / f"qa-report-{batch_id}-{timestamp}.json"
        report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

        _print_report(report, batch_id, raw.get("usageMetadata", {}))
        qa_print(f"Saved raw JSON: {report_path}")
        print("[END]", flush=True)
        return 0
    except Exception as exc:
        qa_print(f"ERROR: {exc}")
        print("[END]", flush=True)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
