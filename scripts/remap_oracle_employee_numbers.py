from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


SOURCE = Path(
    "/home/clawdbot/.openclaw/workspace/aljeel/batches/"
    "asateel-مشاريع 17/P&T-17-2026_Oracle-upload.xlsx"
)
OUTPUT = Path(
    "/home/clawdbot/.openclaw/workspace/aljeel/batches/"
    "asateel-مشاريع 17/P&T-17-2026_Oracle-upload_empno-remap.xlsx"
)
SHEET_NAME = "Sheet"
FIRST_DATA_ROW = 4
EXPECTED_DATA_ROWS = 276
EMPNO_COL = 16
AGENCY_COL = 28

MAPPING = {
    "bio-rad": ("Bio-Rad", "1001982"),
    "solventum": ("Solventum", "1000375"),
    "deroyal": ("Deroyal", "1000157"),
    "abbott": ("Abbott", "1000593"),
    "dirui": ("Dirui", "1001982"),
    "medsource": ("Medsource", "1000157"),
    "steris": ("Steris", "1000320"),
    "vygon": ("Vygon", "1001061"),
}


def normalized(value):
    return value.strip().casefold() if isinstance(value, str) else None


def value_bytes(value):
    """Stable byte representation used to verify unchanged Employee No values."""
    if value is None:
        return b"<None>"
    return str(value).encode("utf-8")


def main():
    if SOURCE.resolve() == OUTPUT.resolve():
        raise RuntimeError("Output must not overwrite the source workbook")
    if OUTPUT.exists():
        raise FileExistsError(f"Refusing to overwrite existing output: {OUTPUT}")

    wb = load_workbook(SOURCE, read_only=False)
    ws = wb[SHEET_NAME]
    last_data_row = FIRST_DATA_ROW + EXPECTED_DATA_ROWS - 1
    if ws.max_row != last_data_row:
        raise AssertionError(
            f"Expected {EXPECTED_DATA_ROWS} data rows ending at row "
            f"{last_data_row}; sheet max_row is {ws.max_row}"
        )

    source_values = tuple(
        tuple(cell.value for cell in row)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column)
    )
    changed = []
    counts = Counter()
    bmx_rows = []

    for row_number in range(FIRST_DATA_ROW, last_data_row + 1):
        agency_value = ws.cell(row=row_number, column=AGENCY_COL).value
        agency_key = normalized(agency_value)
        if agency_key == "bmx":
            bmx_rows.append(row_number)
        if agency_key not in MAPPING:
            continue

        display_name, new_empno = MAPPING[agency_key]
        empno_cell = ws.cell(row=row_number, column=EMPNO_COL)
        old_empno = empno_cell.value
        empno_cell.value = new_empno
        changed.append((row_number, str(agency_value).strip(), old_empno, new_empno))
        counts[display_name] += 1

    wb.save(OUTPUT)

    # Re-open both workbooks and verify saved values, types, and change scope.
    src_wb = load_workbook(SOURCE, read_only=False, data_only=False)
    out_wb = load_workbook(OUTPUT, read_only=False, data_only=False)
    src_ws = src_wb[SHEET_NAME]
    out_ws = out_wb[SHEET_NAME]

    verified_mapped = 0
    verified_other_empno_bytes = 0
    for row_number in range(FIRST_DATA_ROW, last_data_row + 1):
        agency_key = normalized(src_ws.cell(row_number, AGENCY_COL).value)
        src_empno = src_ws.cell(row_number, EMPNO_COL)
        out_empno = out_ws.cell(row_number, EMPNO_COL)
        if agency_key in MAPPING:
            expected = MAPPING[agency_key][1]
            assert out_empno.value == expected
            assert out_empno.data_type == "s"
            verified_mapped += 1
        else:
            assert value_bytes(out_empno.value) == value_bytes(src_empno.value)
            verified_other_empno_bytes += 1

    # Assert that no cell value outside mapped Employee No cells changed.
    allowed = {(row_number, EMPNO_COL) for row_number, *_ in changed}
    value_differences = []
    for row_number in range(1, src_ws.max_row + 1):
        for col_number in range(1, src_ws.max_column + 1):
            before = source_values[row_number - 1][col_number - 1]
            after = out_ws.cell(row_number, col_number).value
            if before != after:
                value_differences.append((row_number, col_number))
    assert set(value_differences).issubset(allowed)
    assert len(value_differences) == sum(
        old != new for _, _, old, new in changed
    )

    print(f"Output file: {OUTPUT}")
    print(f"Total rows changed: {len(changed)}")
    print("Changed rows by agency:")
    for _, (display_name, _) in MAPPING.items():
        print(f"  {display_name}: {counts[display_name]}")
    print(f"BMX rows left unchanged: {len(bmx_rows)} (confirmed)")
    print("Sample of 5 changed rows:")
    for row_number, agency, old_empno, new_empno in changed[:5]:
        print(f"  Row {row_number}: {agency}: {old_empno!r} -> {new_empno!r}")
    print("Re-open verification:")
    print(f"  Mapped rows equal required Employee No: {verified_mapped}")
    print(
        "  BMX/other rows byte-identical to source in col 16: "
        f"{verified_other_empno_bytes}"
    )
    print(
        "  Only Employee No (col 16) changed; all other cell values unchanged: "
        "confirmed"
    )
    print("  Remapped Employee No cells stored as plain strings: confirmed")


if __name__ == "__main__":
    main()
