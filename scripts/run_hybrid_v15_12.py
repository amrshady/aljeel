#!/usr/bin/env python3
"""
AlJeel AP Pipeline v15.12 - Hybrid (Cascade + LLM Agent)

Runs ON TOP of the v15.11.2 cascade output. For each row:
  1. If routing rule matches -> re-resolve via full_evidence_agent (Gemini cascade)
  2. Otherwise -> keep cascade result
  3. Edge-case overlays (DIV=888 G&A travel, solution-from-route) -> cascade wins
  4. Family-cluster emp_no unification -> harmonize across same-surname/route clusters

Output: Spreadsheet-<BATCH>-FILLED-v15.12.xlsx with `Agent Method` column.

Usage:
    python3 scripts/run_hybrid_v15_12.py <BATCH_ID>
    e.g. python3 scripts/run_hybrid_v15_12.py J26-640

Hard rules:
  - Does NOT modify process_batch.py (v15.11.2 stays locked)
  - Uses the AP pipeline canonical Gemini cascade:
        gemini-pro-latest (= Gemini 3 Pro) -> gemini-2.5-pro -> gemini-2.5-flash
  - Cache shared with full_evidence_agent: extracted/full-evidence-agent-cache/
  - Cost ceiling: $3/batch on Gemini - kills run if exceeded
  - Hard rule: GEMINI_API_KEY direct, no LiteLLM proxy
"""
from __future__ import annotations

import argparse
import concurrent.futures
import json
import os
import re
import shutil
import sys
import time
import traceback
from collections import Counter
from pathlib import Path
from typing import Any

import openpyxl

# Reuse the evidence agent's machinery
ROOT = Path("/home/clawdbot/.openclaw/workspace/aljeel")
sys.path.insert(0, str(ROOT / "scripts"))

import full_evidence_agent as fea  # type: ignore
from code_name_lookup import get_lookup
from cost_center_resolver import sync_row_derived_fields

V15_12 = "v15.12"
COST_CEILING = 3.0  # USD per batch on Gemini (3 Pro pricing tier)

# Sponsorship hint keywords - if cascade said travel but description has these, route to LLM
SPONSORSHIP_KEYWORDS = [
    "HF", "AATS", "CRM-", "EP-", "IEPC", "SIS-", "DDW",
    "conference", "meeting room", "sponsoring", "sponsor",
]
SPONSORSHIP_TEXT_REGEX = re.compile(r"\b(" + "|".join(re.escape(k) for k in SPONSORSHIP_KEYWORDS) + r")\b", re.IGNORECASE)


def utc_ts() -> str:
    return time.strftime("%Y%m%dT%H%M%SZ", time.gmtime())


def backup_file(path: Path):
    if path.exists():
        bak = path.with_suffix(path.suffix + f".bak-pre-v15.12-{utc_ts()}")
        shutil.copy2(path, bak)
        print(f"[backup] {path.name} -> {bak.name}", flush=True)


SEGMENT_WIDTHS = {
    "cost_center": 6,
    "div": 3,
    "solution": 5,
    "agency": 5,
}


def normalize_output_segment(value, width: int) -> str:
    text = str(value or "").strip()
    if not text or text.lower() in {"none", "nan", "null"}:
        return "0" * width
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return text.zfill(width) if text.isdigit() and len(text) <= width else text


def normalize_combo_segments(row: dict) -> None:
    for key, width in SEGMENT_WIDTHS.items():
        row[key] = normalize_output_segment(row.get(key, ""), width)


# ---------- routing rules ----------
def folder_has_opex(folder: Path | None) -> bool:
    if not folder or not folder.exists():
        return False
    for _ in folder.rglob("OPEX-*.pdf"):
        return True
    return False


def cascade_says_sponsorship_keyword(account: str, description: str) -> bool:
    """Cascade said account=60301003 (travel) but description has sponsorship keyword -> route to LLM."""
    if str(account).strip() != "60301003":
        return False
    desc_upper = (description or "").upper()
    for kw in SPONSORSHIP_KEYWORDS:
        if kw.upper() in desc_upper:
            return True
    return False


def cascade_not_resolved(cost_center: str, account: str) -> bool:
    """If CC is 000000/999999/blank, treat as not-resolved -> route to LLM."""
    cc = str(cost_center or "").strip()
    if cc in ("", "000000", "999999"):
        return True
    return False


def cascade_is_shared_opex_sponsorship(row: dict) -> bool:
    """FIX 1b: row was deterministically classified as SHARED_OPEX_SPONSORSHIP by
    process_batch.py (account 60307021, segments from the sponsor employee). Such a
    row must NOT be re-routed to the LLM, which sees a travel ticket and flips the
    account back to 60301003, destroying the sponsorship classification."""
    qc = str(row.get("QC Catches", "") or "").upper()
    if "SHARED_OPEX_SPONSORSHIP" in qc:
        return True
    method = str(row.get("Agent Match Method", "") or "").lower()
    if "shared_opex_sponsorship" in method:
        return True
    for col in ("Agent Flags", "Agent Segments Breakdown"):
        val = str(row.get(col, "") or "").upper()
        if "SHARED_OPEX_SPONSOR_SEGMENTS" in val or "SHARED_OPEX_DEFAULT_SEGMENTS" in val:
            return True
    return False


# ---------- cascade xlsx reader ----------
def read_cascade_xlsx(xlsx_path: Path) -> tuple[list[dict], list[str], int]:
    """Read v15.11.2 cascade output. Returns (rows, headers, header_row_idx)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    header_row_idx = None
    headers: list[str] = []
    for ri in range(1, min(20, ws.max_row + 1)):
        cells = [ws.cell(row=ri, column=c).value for c in range(1, ws.max_column + 1)]
        if any(str(c or "").strip() == "*Invoice Header Identifier" for c in cells):
            header_row_idx = ri
            headers = [str(c or "").strip() for c in cells]
            break
    if header_row_idx is None:
        raise ValueError(f"No header row found in {xlsx_path}")

    rows = []
    for ri in range(header_row_idx + 1, ws.max_row + 1):
        cells = [ws.cell(row=ri, column=c).value for c in range(1, ws.max_column + 1)]
        if cells[0] in (None, ""):
            continue
        first = str(cells[0]).strip()
        if not first or first.lower().startswith(("total", "sum", "grand")):
            break
        row = {h: v for h, v in zip(headers, cells)}
        row["_row_idx"] = ri
        rows.append(row)

    return rows, headers, header_row_idx


def write_v15_12_xlsx(src_path: Path, dst_path: Path, hybrid_rows: list[dict], cascade_rows: list[dict], header_row_idx: int):
    """Copy v15.11.2 xlsx -> modify rows where hybrid overrode -> add Agent Method column."""
    shutil.copy2(src_path, dst_path)
    wb = openpyxl.load_workbook(dst_path)
    ws = wb.active

    # Find or append "Agent Method" column
    method_col = None
    for ci in range(1, ws.max_column + 2):
        v = ws.cell(row=header_row_idx, column=ci).value
        if v in (None, "") and method_col is None:
            method_col = ci
            break
        if str(v or "").strip() == "Agent Method":
            method_col = ci
            break
    if method_col is None:
        method_col = ws.max_column + 1
    ws.cell(row=header_row_idx, column=method_col, value="Agent Method")
    lookup = get_lookup()

    # Map header -> column
    col_map = {}
    for ci in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row_idx, column=ci).value
        if v:
            col_map[str(v).strip()] = ci

    OVERWRITE_KEYS = {
        "Employee No": "emp_no",
        "Account": "account",
        "Cost Center": "cost_center",
        "DIV": "div",
        "Solution": "solution",
        "Agency": "agency",
    }
    DISTRIBUTION_KEY = next(
        (str(v).strip() for v in col_map
         if v and "distribution combination" in str(v).lower()),
        "Distribution Combination"
    )

    for i, hyb in enumerate(hybrid_rows):
        excel_row = hyb["_row_idx"]
        method = hyb.get("_agent_method", "cascade")
        ws.cell(row=excel_row, column=method_col, value=method)

        if method in ("llm_agent", "hybrid_overlay", "cluster_unified"):
            for key, hdr in [
                ("company", "Company"),
                ("location", "Location"),
                ("project", "Project"),
                ("intercompany", "Intercompany"),
                ("future1", "Future 1"),
            ]:
                if not hyb.get(key):
                    hyb[key] = cascade_rows[i].get(hdr, "")
            normalize_combo_segments(hyb)
            for hdr, key in OVERWRITE_KEYS.items():
                if hdr in col_map:
                    val = hyb.get(key, "")
                    if val is not None:
                        ws.cell(row=excel_row, column=col_map[hdr]).value = val
            # Rebuild Location column if it exists in col_map
            if "Location" in col_map:
                loc = hyb.get("location", "") or cascade_rows[i].get("Location", "") or "20100"
                hyb["location"] = loc
                ws.cell(row=excel_row, column=col_map["Location"], value=loc)

            sync_row_derived_fields(hyb, lookup)
            if DISTRIBUTION_KEY in col_map:
                ws.cell(row=excel_row, column=col_map[DISTRIBUTION_KEY], value=hyb["combo"])
            if "GL Description" in col_map:
                ws.cell(row=excel_row, column=col_map["GL Description"], value=hyb["gl_description"])

    wb.save(str(dst_path))


# ---------- LLM resolution ----------
def llm_resolve_row(row_idx: int, cascade_row: dict, batch_id: str, raw_root: Path,
                    all_folders: list[Path], manpower: dict, lookups: dict) -> dict:
    """Build pseudo-row from cascade, run LLM, return resolution dict.

    Cache key = (batch_id, row_idx, ticket_no). Stores under the shared evidence-agent cache.
    """
    desc = str(cascade_row.get("Description", "") or "")
    notes = str(cascade_row.get("Notes", "") or "")
    passenger = desc.split(" - ", 1)[0].strip() if " - " in desc else ""
    route = desc.split(" - ", 1)[1] if " - " in desc else ""
    amount = cascade_row.get("*Amount", "")
    date_val = str(cascade_row.get("*Invoice Date", "") or "")
    m = re.search(r"\b(\d{10,})\b", desc + " " + notes)
    ticket_no = m.group(1) if m else ""

    pseudo_row = {
        "row_idx": row_idx,
        "ticket_no": ticket_no,
        "passenger": passenger,
        "route": route,
        "amount": amount,
        "date": date_val,
        "notes": notes,
        "description": desc,
    }

    folder = fea.find_ticket_folder(ticket_no, passenger, notes, all_folders)
    evidence = fea.collect_evidence(folder) if folder else {
        "folder": "", "msgs": [], "pdfs": [], "files": [], "total_chars": 0
    }
    snapshot = fea.build_master_snapshot(manpower, lookups, evidence, pseudo_row)
    prompt = fea.build_prompt(pseudo_row, evidence, snapshot)

    # Cache - include batch in key to avoid cross-batch collisions on same row_idx
    cache_key = f"hybrid-{batch_id}-row{row_idx:03d}-{ticket_no or 'noticket'}.json"
    cache_path = fea.CACHE_DIR / cache_key
    if cache_path.exists():
        try:
            cached = json.loads(cache_path.read_text())
            cached["_from_cache"] = True
            return cached
        except Exception:
            pass

    t0 = time.time()
    res = fea.call_gemini_with_cascade(prompt)
    elapsed = time.time() - t0
    j = res.get("json", {}) or {}
    out = {
        "emp_no": str(j.get("emp_no", "") or "").strip(),
        "account": str(j.get("account", "") or "").strip(),
        "cost_center": str(j.get("cost_center", "") or "").strip(),
        "div": str(j.get("div", "") or "").strip(),
        "solution": str(j.get("solution", "") or "").strip(),
        "agency": str(j.get("agency", "") or "").strip(),
        "confidence": str(j.get("confidence", "") or "").strip(),
        "reasoning": str(j.get("reasoning", "") or "").strip(),
        "_llm_model": res.get("model", ""),
        "_llm_in_tokens": res.get("in_tokens", 0),
        "_llm_out_tokens": res.get("out_tokens", 0),
        "_llm_error": res.get("error", ""),
        "_elapsed_sec": elapsed,
        "_folder": str(folder) if folder else "",
        "_evidence_chars": evidence.get("total_chars", 0),
        "_evidence_files_count": len(evidence.get("files", [])),
        "_from_cache": False,
    }
    # Cache writing disabled permanently for zero-cache policy
    # try:
    #     cache_path.write_text(json.dumps(out, ensure_ascii=False, indent=2, default=str))
    # except Exception as e:
    #     out["_cache_err"] = str(e)
    return out


# ---------- overlays ----------
def apply_overlays(llm_out: dict, cascade_row: dict) -> tuple[dict, str]:
    """Apply edge-case overlays where cascade should override LLM."""
    method = "llm_agent"
    final = dict(llm_out)

    cascade_account = str(cascade_row.get("Account", "") or "").strip()
    cascade_solution = str(cascade_row.get("Solution", "") or "").strip()
    cascade_div = str(cascade_row.get("DIV", "") or "").strip()

    llm_account = str(llm_out.get("account", "") or "").strip()
    llm_solution = str(llm_out.get("solution", "") or "").strip()

    # Overlay 1: DIV=888 G&A travel -> cascade said 60301004, LLM said 60301003 -> keep cascade
    if cascade_account == "60301004" and llm_account == "60301003":
        final["account"] = cascade_account
        if cascade_div == "888":
            method = "hybrid_overlay"

    # Overlay 2: cascade has a deterministic solution from route mapping; LLM blanked it -> keep cascade
    if cascade_solution and cascade_solution not in ("00000", "0", "") and not llm_solution:
        final["solution"] = cascade_solution
        method = "hybrid_overlay"

    # Overlay 3: for travel rows, cascade solution-from-route beats LLM guess
    if (cascade_solution and llm_solution and cascade_solution != llm_solution
            and cascade_account == "60301003"):
        if llm_solution in ("00000", "0", "") or (
            cascade_solution.isdigit() and cascade_solution != "00000"
            and len(cascade_solution.zfill(5)) == 5
        ):
            final["solution"] = cascade_solution
            method = "hybrid_overlay"

    return final, method


def apply_family_cluster_unification(hybrid_rows: list[dict], cascade_rows: list[dict]) -> int:
    """Unify emp_no within family clusters (same surname + route + month)."""
    by_cluster: dict[str, list[int]] = {}
    for i, c in enumerate(cascade_rows):
        desc = str(c.get("Description", "") or "")
        passenger = desc.split(" - ", 1)[0].strip().upper() if " - " in desc else ""
        surname = passenger.split()[-1] if passenger else ""
        date = str(c.get("*Invoice Date", "") or "")[:7]
        route = desc.split(" - ", 1)[1] if " - " in desc else ""
        key = f"{surname}|{route}|{date}"
        if surname:
            by_cluster.setdefault(key, []).append(i)

    unified_count = 0
    for key, idxs in by_cluster.items():
        if len(idxs) < 2:
            continue
        cascade_empnos = [str(cascade_rows[i].get("Employee No", "") or "").strip() for i in idxs]
        cascade_empnos = [e for e in cascade_empnos if e and e != "0"]
        if not cascade_empnos:
            continue
        emp_winner, n = Counter(cascade_empnos).most_common(1)[0]
        if n >= 2:
            for i in idxs:
                if hybrid_rows[i].get("_agent_method") in ("llm_agent", "hybrid_overlay"):
                    if str(hybrid_rows[i].get("emp_no", "")).strip() != emp_winner:
                        hybrid_rows[i]["emp_no"] = emp_winner
                        hybrid_rows[i]["_agent_method"] = "cluster_unified"
                        unified_count += 1
    return unified_count


# ---------- main ----------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("batch_id", help="e.g. J26-640")
    ap.add_argument("--workers", type=int, default=5)
    ap.add_argument("--limit", type=int, default=0, help="0 = all rows")
    ap.add_argument("--cost-ceiling", type=float, default=COST_CEILING)
    ap.add_argument("--input-suffix", default="v15.11", help="Cascade input suffix prefix (matches v15.11, v15.11.2, etc.)")
    args = ap.parse_args()

    batch_id = args.batch_id
    batch_dir = ROOT / "batches" / f"jawal-{batch_id}"
    raw_root = batch_dir / "raw"
    if not batch_dir.exists():
        print(f"[fatal] batch dir not found: {batch_dir}", file=sys.stderr)
        sys.exit(2)

    cascade_pattern = f"Spreadsheet-{batch_id}-FILLED-{args.input_suffix}*.xlsx"
    candidates = sorted((batch_dir / "output").glob(cascade_pattern),
                        key=lambda p: p.stat().st_mtime, reverse=True)
    if not candidates:
        print(f"[fatal] cascade output missing: {batch_dir}/output/{cascade_pattern}", file=sys.stderr)
        print(f"[hint] run: bash PROCESS/run_jawal_batch.sh {batch_id}", file=sys.stderr)
        sys.exit(3)
    cascade_xlsx = candidates[0]
    print(f"[load] cascade output: {cascade_xlsx.name}", flush=True)

    out_dir = batch_dir / "output"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_xlsx = out_dir / f"Spreadsheet-{batch_id}-FILLED-{V15_12}.xlsx"
    if out_xlsx.exists():
        backup_file(out_xlsx)

    cascade_rows, headers, hdr_row = read_cascade_xlsx(cascade_xlsx)
    print(f"[load] {len(cascade_rows)} cascade rows", flush=True)
    if args.limit > 0:
        cascade_rows = cascade_rows[:args.limit]

    print(f"[load] master data...", flush=True)
    manpower = fea.load_manpower()
    lookups = fea.load_lookups()
    all_folders = fea.collect_all_folders(raw_root)
    print(f"[scan] {len(all_folders)} ticket-folder candidates in {raw_root}", flush=True)

    # Routing pass
    routed: list[tuple[int, str]] = []
    shared_opex_guarded: set[int] = set()  # FIX 1b: deterministic sponsorship rows, never sent to LLM
    for i, row in enumerate(cascade_rows):
        cascade_account = str(row.get("Account", "") or "").strip()
        cascade_cc = str(row.get("Cost Center", "") or "").strip()
        desc = str(row.get("Description", "") or "")
        notes = str(row.get("Notes", "") or "")
        m = re.search(r"\b(\d{10,})\b", desc + " " + notes)
        ticket_no = m.group(1) if m else ""
        passenger = desc.split(" - ", 1)[0].strip() if " - " in desc else ""
        folder = fea.find_ticket_folder(ticket_no, passenger, notes, all_folders)

        # FIX 1b: protect SHARED_OPEX_SPONSORSHIP rows from LLM re-routing.
        # Keep the deterministic cascade result; method is set after hybrid_rows init.
        if cascade_is_shared_opex_sponsorship(row):
            shared_opex_guarded.add(i)
            continue

        reason = None
        if folder_has_opex(folder):
            reason = "OPEX_PDF"
        elif cascade_says_sponsorship_keyword(cascade_account, desc):
            reason = "SPONSORSHIP_KEYWORD"
        elif cascade_not_resolved(cascade_cc, cascade_account):
            reason = "NOT_RESOLVED"

        if reason:
            routed.append((i, reason))

    reason_counts: dict[str, int] = {}
    for _, r in routed:
        reason_counts[r] = reason_counts.get(r, 0) + 1
    print(f"[route] {len(routed)} rows routed to LLM / {len(cascade_rows) - len(routed)} stay cascade", flush=True)
    for r, n in reason_counts.items():
        print(f"        {r}: {n}", flush=True)
    if shared_opex_guarded:
        print(f"        SHARED_OPEX_SPONSORSHIP (guarded, kept cascade): {len(shared_opex_guarded)}", flush=True)

    # Initialize hybrid_rows
    hybrid_rows = []
    for i, c in enumerate(cascade_rows):
        h = {
            "_row_idx": c["_row_idx"],
            "_agent_method": "cascade",
            "emp_no": c.get("Employee No", "") or "",
            "account": c.get("Account", "") or "",
            "cost_center": c.get("Cost Center", "") or "",
            "div": c.get("DIV", "") or "",
            "solution": c.get("Solution", "") or "",
            "agency": c.get("Agency", "") or "",
            "location": c.get("Location", "") or "",
            "company": c.get("Company", "") or "03",
            "project": c.get("Project", "") or "00000",
            "intercompany": c.get("Intercompany", "") or "00",
            "future1": c.get("Future 1", "") or "000000",
        }
        hybrid_rows.append(h)

    # FIX 1b: tag guarded sponsorship rows so they are recorded as the deterministic
    # method (not llm_agent) while keeping their cascade segments untouched.
    for i in shared_opex_guarded:
        hybrid_rows[i]["_agent_method"] = "shared_opex_sponsorship"

    # LLM resolve routed rows
    total_in = 0
    total_out = 0
    llm_errors = 0

    def worker(idx_reason):
        idx, reason = idx_reason
        try:
            row = cascade_rows[idx]
            llm = llm_resolve_row(idx, row, batch_id, raw_root, all_folders, manpower, lookups)
            return idx, reason, llm
        except Exception as e:
            return idx, reason, {"_llm_error": f"{type(e).__name__}: {e}", "_traceback": traceback.format_exc()}

    t_start = time.time()
    done = 0
    pricing_pro = fea.GEMINI_PRICING["gemini-pro-latest"]

    with concurrent.futures.ThreadPoolExecutor(max_workers=args.workers) as ex:
        futs = [ex.submit(worker, (i, r)) for i, r in routed]
        try:
            for fut in concurrent.futures.as_completed(futs):
                idx, reason, llm = fut.result()
                done += 1
                total_in += llm.get("_llm_in_tokens", 0) or 0
                total_out += llm.get("_llm_out_tokens", 0) or 0
                # Use 3 Pro pricing as worst-case estimate (Flash is 5x cheaper)
                cost_est = (total_in * pricing_pro["in"]) + (total_out * pricing_pro["out"])
                if llm.get("_llm_error"):
                    llm_errors += 1
                    cached = ""
                    err_marker = f" ERR={llm.get('_llm_error', '')[:50]}"
                else:
                    cached = " [cached]" if llm.get("_from_cache") else ""
                    err_marker = ""
                elapsed = time.time() - t_start
                print(f"[{done:3d}/{len(routed)}] row {idx:3d} ({reason})  "
                      f"t={elapsed:5.1f}s  cost~=${cost_est:.3f}{cached}{err_marker}", flush=True)

                if not llm.get("_llm_error"):
                    final, method = apply_overlays(llm, cascade_rows[idx])
                    hybrid_rows[idx].update({
                        "_agent_method": method,
                        "emp_no": final.get("emp_no", ""),
                        "account": final.get("account", ""),
                        "cost_center": final.get("cost_center", ""),
                        "div": final.get("div", ""),
                        "solution": final.get("solution", ""),
                        "agency": final.get("agency", ""),
                        "_llm_model": llm.get("_llm_model", ""),
                        "_llm_reason": llm.get("reasoning", "")[:200],
                        "_route_reason": reason,
                    })
                else:
                    hybrid_rows[idx]["_agent_method"] = "cascade_fallback"
                    hybrid_rows[idx]["_llm_error"] = llm.get("_llm_error", "")

                if cost_est > args.cost_ceiling:
                    print(f"[fatal] cost ceiling exceeded: ${cost_est:.2f} > ${args.cost_ceiling:.2f}", file=sys.stderr)
                    for f in futs:
                        f.cancel()
                    sys.exit(4)
        except KeyboardInterrupt:
            print("\n[interrupted] saving partial progress...", flush=True)

    # Family cluster unification
    unified = apply_family_cluster_unification(hybrid_rows, cascade_rows)
    print(f"[cluster] {unified} rows emp_no-harmonized within family clusters", flush=True)

    # Write output
    write_v15_12_xlsx(cascade_xlsx, out_xlsx, hybrid_rows, cascade_rows, hdr_row)
    print(f"[out] wrote {out_xlsx}", flush=True)

    # Cost summary - use the most-common model from this run for billing accuracy
    cost_usd = (total_in * pricing_pro["in"]) + (total_out * pricing_pro["out"])
    method_counts: dict[str, int] = {}
    for h in hybrid_rows:
        m = h["_agent_method"]
        method_counts[m] = method_counts.get(m, 0) + 1

    summary = {
        "batch_id": batch_id,
        "version": V15_12,
        "timestamp_utc": utc_ts(),
        "total_rows": len(cascade_rows),
        "routed_to_llm": len(routed),
        "routing_reasons": reason_counts,
        "method_counts": method_counts,
        "cluster_unified_rows": unified,
        "llm_in_tokens": total_in,
        "llm_out_tokens": total_out,
        "llm_cascade": fea.GEMINI_MODEL_CASCADE,
        "cost_usd_est_at_3pro_pricing": round(cost_usd, 4),
        "cost_ceiling": args.cost_ceiling,
        "llm_errors": llm_errors,
        "runtime_sec": round(time.time() - t_start, 1),
        "cascade_input": str(cascade_xlsx),
        "output": str(out_xlsx),
    }
    (out_dir / f"summary-{V15_12}.json").write_text(json.dumps(summary, indent=2, default=str))

    print()
    print("=" * 60)
    print(f"  Hybrid {V15_12} done")
    print(f"  Rows:           {summary['total_rows']}  ({len(routed)} LLM-routed)")
    print(f"  Methods:        {method_counts}")
    print(f"  Cluster unif.:  {unified}")
    print(f"  Cost (est):     ${cost_usd:.4f}  ({total_in:,} in / {total_out:,} out)")
    print(f"  Runtime:        {summary['runtime_sec']:.1f}s")
    print(f"  Output:         {out_xlsx}")
    print("=" * 60)


if __name__ == "__main__":
    main()
