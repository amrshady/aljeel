#!/usr/bin/env python3
"""
AlJeel AP Pipeline v16 - Hybrid with Employee/Sponsorship Classifier

Builds on v15.12. Key addition:
  NEW ROUTING TRIGGER: TRAVEL_NO_EVIDENCE
    Fires when cascade=travel (60301003/60301004) AND FORM_NOT_FOUND_IN_EMAIL.
    These rows got no email evidence so the cascade defaulted to travel — but many
    are actually sponsored externals going to group-booked conferences.

  NEW DECISION TREE for TRAVEL_NO_EVIDENCE rows:
    Step 1 — CLASSIFY (Flash, cheap): Employee trip or Sponsored external?
              Passes ALL day-range folders so group-booking evidence is visible.
    Step 2a — If SPONSORED: full LLM resolve using classifier-identified folder
    Step 2b — If EMPLOYEE: full LLM resolve (existing path, broader folder hint)
    Step 2c — If UNKNOWN: full LLM resolve with all day folders as context

  All other routing triggers from v15.12 are unchanged.

Output: Spreadsheet-<INVOICE>-FILLED-v16.xlsx with `Agent Method` column.

Usage:
    python3 scripts/run_hybrid_v16.py <BATCH_ID>
    python3 scripts/run_hybrid_v16.py J26-593-BLIND J26-593   (blind test)
"""
from __future__ import annotations

import argparse
import concurrent.futures
import json
import os
import re
import shutil
import sys
import time
import traceback
from pathlib import Path

import openpyxl

ROOT = Path("/home/clawdbot/.openclaw/workspace/aljeel")
sys.path.insert(0, str(ROOT / "scripts"))

import full_evidence_agent as fea       # core LLM machinery (root-owned, unchanged)
import fea_v16_ext as fea16             # v16 additions: classify, get_day_folders, etc.

V16 = "v16"
COST_CEILING = 5.0   # slightly higher than v15.12 — classify calls add ~$0.01/row

# ── re-export routing helpers from v15.12 (same logic) ──────────────────────
import run_hybrid_v15_12 as v1512
folder_has_opex             = v1512.folder_has_opex
cascade_says_sponsorship_keyword = v1512.cascade_says_sponsorship_keyword
cascade_not_resolved        = v1512.cascade_not_resolved
read_cascade_xlsx           = v1512.read_cascade_xlsx
apply_overlays              = v1512.apply_overlays
apply_cluster_unification   = v1512.apply_family_cluster_unification
SPONSORSHIP_TEXT_REGEX      = v1512.SPONSORSHIP_TEXT_REGEX


def utc_ts() -> str:
    return time.strftime("%Y%m%dT%H%M%SZ", time.gmtime())


def backup_file(path: Path):
    if path.exists():
        bak = path.with_suffix(path.suffix + f".bak-pre-v16-{utc_ts()}")
        shutil.copy2(path, bak)
        print(f"[backup] {path.name} -> {bak.name}", flush=True)


def write_v16_xlsx(src_path: Path, dst_path: Path,
                   hybrid_rows: list[dict], cascade_rows: list[dict],
                   header_row_idx: int):
    """Copy cascade xlsx → apply hybrid overrides → add Agent Method column."""
    shutil.copy2(src_path, dst_path)
    wb = openpyxl.load_workbook(dst_path)
    ws = wb.active

    # Find or create "Agent Method" column
    method_col = None
    for ci in range(1, ws.max_column + 2):
        v = ws.cell(row=header_row_idx, column=ci).value
        if v in (None, "") and method_col is None:
            method_col = ci
            break
        if str(v or "").strip() == "Agent Method":
            method_col = ci
            break
    if method_col is None:
        method_col = ws.max_column + 1
    ws.cell(row=header_row_idx, column=method_col, value="Agent Method")

    col_map = {}
    for ci in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row_idx, column=ci).value
        if v:
            col_map[str(v).strip()] = ci

    OVERWRITE_KEYS = {
        "Employee No": "emp_no",
        "Account":      "account",
        "Cost Center":  "cost_center",
        "DIV":          "div",
        "Solution":     "solution",
        "Agency":       "agency",
        "Location":     "location",
    }

    # Find distribution combo column — header may be truncated differently per file
    combo_col = None
    for k, ci in col_map.items():
        if k.startswith("Distribution Combination"):
            combo_col = ci
            break

    for h_row, c_row in zip(hybrid_rows, cascade_rows):
        excel_row = c_row["_row_idx"]
        method = h_row.get("_agent_method", "cascade")
        ws.cell(row=excel_row, column=method_col, value=method)

        if method == "cascade":
            continue

        # Write overridden fields
        for col_name, key in OVERWRITE_KEYS.items():
            ci = col_map.get(col_name)
            if ci and h_row.get(key):
                ws.cell(row=excel_row, column=ci, value=h_row[key])

        # Rebuild combo from fields
        if combo_col:
            acc = h_row.get("account") or c_row.get("Account", "") or ""
            cc  = h_row.get("cost_center") or c_row.get("Cost Center", "") or ""
            div = h_row.get("div") or c_row.get("DIV", "") or ""
            sol = h_row.get("solution") or c_row.get("Solution", "") or ""
            ag  = h_row.get("agency") or c_row.get("Agency", "") or ""
            loc = h_row.get("location") or c_row.get("Location", "") or "40100"
            combo = f"03-{loc}-{acc}-{cc}-{div}-{sol}-{ag}-00000-00-000000"
            ws.cell(row=excel_row, column=combo_col, value=combo)

    wb.save(str(dst_path))


# ── routing ──────────────────────────────────────────────────────────────────
def route_rows(cascade_rows: list[dict], all_folders: list[Path]) -> list[tuple[int, str]]:
    """Return list of (row_idx, reason) for rows to send to LLM."""
    routed = []
    for i, row in enumerate(cascade_rows):
        desc  = str(row.get("Description", "") or "")
        notes = str(row.get("Notes", "") or "")
        cascade_account = str(row.get("Account", "") or "").strip()
        cascade_cc      = str(row.get("Cost Center", "") or "").strip()
        cascade_agency  = str(row.get("Agency", "") or "").strip()
        cascade_flags   = str(row.get("Agent Flags", "") or "")
        cascade_trip    = str(row.get("Trip Purpose", "") or "").strip().upper()

        m = re.search(r"\b(\d{10,})\b", desc + " " + notes)
        ticket_no  = m.group(1) if m else ""
        passenger  = desc.split(" - ", 1)[0].strip() if " - " in desc else ""
        folder = fea.find_ticket_folder(ticket_no, passenger, notes, all_folders)

        reason = None

        # ── v15.12 triggers (unchanged) ─────────────────────────────────────
        if folder_has_opex(folder):
            cc_ok  = cascade_cc not in ("", "000000", "999999")
            ag_ok  = cascade_agency not in ("", "00000")
            if not (cascade_account == "60307021" and cc_ok and ag_ok):
                reason = "OPEX_PDF_UNRESOLVED"
        if not reason and cascade_says_sponsorship_keyword(cascade_account, desc, notes):
            reason = "SPONSORSHIP_KEYWORD"
        if not reason and cascade_not_resolved(cascade_cc, cascade_account):
            reason = "NOT_RESOLVED"
        if not reason and cascade_trip == "UNKNOWN" and cascade_account in ("60301003","60301004"):
            if SPONSORSHIP_TEXT_REGEX.search(desc + " " + notes):
                reason = "SPONSORSHIP_SIGNAL_UNKNOWN_TRIP"
        if not reason and "EMPLOYEE_NOT_IN_MASTER" in cascade_flags:
            reason = "EMPLOYEE_NOT_IN_MASTER"

        # ── v16 NEW trigger ──────────────────────────────────────────────────
        # Travel with no email evidence found — classify before deciding
        if not reason and cascade_account in ("60301003","60301004"):
            if "FORM_NOT_FOUND_IN_EMAIL" in cascade_flags:
                reason = "TRAVEL_NO_EVIDENCE"

        if reason:
            routed.append((i, reason))

    return routed


# ── LLM resolution ───────────────────────────────────────────────────────────
def llm_resolve_row_v16(
    row_idx: int,
    cascade_row: dict,
    batch_id: str,
    raw_root: Path,
    all_folders: list[Path],
    manpower: dict,
    lookups: dict,
    routing_reason: str,
) -> dict:
    """
    v16 resolution. For TRAVEL_NO_EVIDENCE rows: classify first (Flash),
    then do full LLM resolve with the right folder context.
    All other reasons: delegate to v15.12 logic directly.
    """
    desc  = str(cascade_row.get("Description", "") or "")
    notes = str(cascade_row.get("Notes", "") or "")
    passenger = desc.split(" - ", 1)[0].strip() if " - " in desc else ""
    route = desc.split(" - ", 1)[1] if " - " in desc else ""
    amount = cascade_row.get("*Amount", "")
    date_val = str(cascade_row.get("*Invoice Date", "") or "")
    m = re.search(r"\b(\d{10,})\b", desc + " " + notes)
    ticket_no = m.group(1) if m else ""

    pseudo_row = {
        "row_idx": row_idx, "ticket_no": ticket_no, "passenger": passenger,
        "route": route, "amount": amount, "date": date_val,
        "notes": notes, "description": desc,
    }

    # ── v16 branch: TRAVEL_NO_EVIDENCE ──────────────────────────────────────
    if routing_reason == "TRAVEL_NO_EVIDENCE":
        # Step 1: classify (cheap Flash call)
        clf = fea16.classify_employee_or_sponsorship(
            pseudo_row, date_val, all_folders, batch_id, row_idx
        )
        row_type     = clf.get("row_type", "TYPE_UNKNOWN")
        folder_hint  = clf.get("evidence_folder")

        # Step 2: find the best evidence folder
        if folder_hint:
            folder = fea16.find_folder_by_name(folder_hint, all_folders)
        else:
            folder = fea.find_ticket_folder(ticket_no, passenger, notes, all_folders)

        # If still no folder and sponsored → look at all day-range folders as context
        if not folder and row_type == "TYPE_SPONSORED":
            day_folders = fea16.get_day_folders(date_val, all_folders)
            # Use the first folder that has an OPEX PDF as a last resort
            for df in day_folders:
                for _ in df.rglob("OPEX-*.pdf"):
                    folder = df
                    break
                if folder:
                    break

        # Collect evidence and resolve
        evidence = fea.collect_evidence(folder) if folder else {
            "folder": "", "msgs": [], "pdfs": [], "files": [], "total_chars": 0
        }

        # Annotate pseudo_row with classifier hint for the prompt
        pseudo_row["_v16_row_type"]  = row_type
        pseudo_row["_v16_reasoning"] = clf.get("reasoning", "")

        snapshot = fea.build_master_snapshot(manpower, lookups, evidence, pseudo_row)
        prompt   = _build_v16_prompt(pseudo_row, evidence, snapshot, clf)

        cache_key = "v16-resolve-{}-row{:03d}-{}.json".format(
            batch_id, row_idx, ticket_no or "noticket"
        )
        cache_path = fea.CACHE_DIR / cache_key
        if cache_path.exists():
            try:
                cached = json.loads(cache_path.read_text())
                cached["_from_cache"] = True
                return cached
            except Exception:
                pass

        t0  = time.time()
        res = fea.call_gemini_with_cascade(prompt)
        elapsed = time.time() - t0
        j = res.get("json", {}) or {}

        out = {
            "emp_no":      str(j.get("emp_no", "") or "").strip(),
            "account":     str(j.get("account", "") or "").strip(),
            "cost_center": str(j.get("cost_center", "") or "").strip(),
            "div":         str(j.get("div", "") or "").strip(),
            "solution":    str(j.get("solution", "") or "").strip(),
            "agency":      str(j.get("agency", "") or "").strip(),
            "confidence":  str(j.get("confidence", "") or "").strip(),
            "reasoning":   str(j.get("reasoning", "") or "").strip(),
            "_llm_model":           res.get("model", ""),
            "_llm_in_tokens":       res.get("in_tokens", 0),
            "_llm_out_tokens":      res.get("out_tokens", 0),
            "_llm_error":           res.get("error", ""),
            "_elapsed_sec":         elapsed,
            "_folder":              str(folder) if folder else "",
            "_evidence_chars":      evidence.get("total_chars", 0),
            "_evidence_files_count":len(evidence.get("files", [])),
            "_v16_row_type":        row_type,
            "_v16_classify_reason": clf.get("reasoning", ""),
            "_from_cache":          False,
        }
        # Classify tokens go into summary (approximate add)
        out["_classify_in_tokens"]  = clf.get("_in_tokens", 0)
        out["_classify_out_tokens"] = clf.get("_out_tokens", 0)

        try:
            cache_path.write_text(json.dumps(out, ensure_ascii=False, indent=2, default=str))
        except Exception as e:
            out["_cache_err"] = str(e)
        return out

    # ── all other reasons: delegate to v15.12 path ──────────────────────────
    return v1512.llm_resolve_row(
        row_idx, cascade_row, batch_id, raw_root, all_folders, manpower, lookups
    )


def _build_v16_prompt(row: dict, evidence: dict, master_snapshot: dict, clf: dict) -> str:
    """Build v16 prompt.
    For TYPE_SPONSORED: use targeted prompt that pre-commits account=60307021
    and only asks LLM to identify requester for CC/DIV/solution/agency.
    For TYPE_EMPLOYEE / TYPE_UNKNOWN: augment standard prompt with classifier hint.
    """
    row_type   = clf.get("row_type", "TYPE_UNKNOWN")
    ev_folder  = clf.get("evidence_folder") or "not identified"
    reasoning  = clf.get("reasoning", "")
    requester  = clf.get("requester_emp_no") or ""

    if row_type == "TYPE_SPONSORED":
        # Targeted sponsorship resolver — account is committed, just fill the rest
        emp_dir = json.dumps(master_snapshot.get("employees", {}), ensure_ascii=False, indent=1)
        div_dir = json.dumps(master_snapshot.get("div", {}), ensure_ascii=False, indent=1)
        ag_dir  = json.dumps(master_snapshot.get("agency_hint", {}), ensure_ascii=False, indent=1)
        sol_dir = json.dumps(master_snapshot.get("solution_hint", {}), ensure_ascii=False, indent=1)
        opex_dir= json.dumps(master_snapshot.get("opex_division_hint", {}), ensure_ascii=False, indent=1)

        def _s(v):
            if v is None: return ""
            if isinstance(v, list): return ", ".join(str(x) for x in v)
            return str(v)
        msg_chunks = []
        for m in evidence.get("msgs", []):
            msg_chunks.append(
                "=== EMAIL: {} ===\nSubject: {}\nFrom: {}\n{}\n".format(
                    _s(m.get("filename","")), _s(m.get("subject","")),
                    _s(m.get("from","")), (_s(m.get("body","")) or "[no body]")[:3000]
                )
            )
        pdf_chunks = [
            "=== PDF: {} ===\n{}\n".format(p.get("filename",""), (p.get("text","") or "")[:3000])
            for p in evidence.get("pdfs", [])
        ]

        return (
            "You are an AlJeel AP sponsorship specialist. This invoice row has been CONFIRMED as\n"
            "a SPONSORSHIP expense (external doctor/guest funded by AlJeel).\n\n"
            "CONFIRMED: account = 60307021 (Sponsoring Expenses). DO NOT change this.\n\n"
            "YOUR TASK: Identify the AlJeel requester/organizer from the evidence below,\n"
            "then look up their home cost_center / div / solution / agency from the directory.\n\n"
            "## INVOICE ROW\n"
            "Passenger:    {}\n"
            "Route:        {}\n"
            "Ticket #:     {}\n"
            "Amount (SAR): {}\n"
            "Evidence folder: {}\n"
            "Classifier reasoning: {}\n\n"
            "## EVIDENCE FILES\n"
            "{}\n"
            "{}\n"
            "## EMPLOYEE DIRECTORY (look up AlJeel REQUESTER's home values here)\n"
            "{}\n\n"
            "## DIV codes\n{}\n## Agency hints\n{}\n## Solution hints\n{}\n"
            "## OPEX prefix defaults (use when OPEX PDF found but requester unclear)\n{}\n\n"
            "## ALGORITHM\n"
            "1. Find the AlJeel employee who REQUESTED this sponsorship from the emails/PDFs.\n"
            "   Look for: OPEX form requester table, email sender/from field, 'Employee No' in PDF.\n"
            "   Hint from pre-classifier: requester_emp_no hint = '{}'.\n"
            "2. Look up their emp_no in the Employee Directory above.\n"
            "3. Copy EXACTLY: cost_center, div_code, agency_code, solution from their row.\n"
            "4. If requester not found: use OPEX prefix defaults or 000000/00000/00000.\n\n"
            "## OUTPUT — return ONLY this JSON:\n"
            "{{\n"
            '  "emp_no": "",\n'
            '  "account": "60307021",\n'
            '  "cost_center": "<6-digit from requester home>",\n'
            '  "div": "<div from requester home>",\n'
            '  "solution": "<5-digit from requester home>",\n'
            '  "agency": "<5-digit from requester home>",\n'
            '  "confidence": "high|medium|low",\n'
            '  "reasoning": "<requester emp_no found, copied CC/DIV/sol/agency from their row>"\n'
            "}}\n"
        ).format(
            row.get("passenger",""), row.get("route",""), row.get("ticket_no",""),
            row.get("amount",""), ev_folder, reasoning,
            "\n".join(msg_chunks) or "(no emails)",
            "\n".join(pdf_chunks) or "(no PDFs)",
            emp_dir, div_dir, ag_dir, sol_dir, opex_dir, requester or "not identified"
        )

    # TYPE_EMPLOYEE or TYPE_UNKNOWN: use standard prompt with classifier hint appended
    base_prompt = fea.build_prompt(row, evidence, master_snapshot)
    classifier_block = (
        "\n\n# V16 CLASSIFIER PRE-ANALYSIS\n"
        "A lightweight classifier already examined this row and nearby folders:\n"
        "  Row type:      {}\n"
        "  Evidence hint: {}\n"
        "  Reasoning:     {}\n\n"
        "Use this as a strong prior but override if the full evidence contradicts it.\n"
        "TYPE_EMPLOYEE  → resolve account/CC/DIV/solution/agency from manpower master.\n"
        "TYPE_UNKNOWN   → default to TYPE A (employee travel) if no OPEX evidence found.\n"
    ).format(row_type, ev_folder, reasoning)

    if "# OUTPUT" in base_prompt:
        return base_prompt.replace("# OUTPUT", classifier_block + "# OUTPUT", 1)
    return base_prompt + classifier_block


# ── main ─────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(description="AlJeel AP Hybrid v16")
    ap.add_argument("batch_id")
    ap.add_argument("invoice_id", nargs="?")
    ap.add_argument("--workers", type=int, default=4)
    ap.add_argument("--limit",   type=int, default=0)
    ap.add_argument("--cost-ceiling", type=float, default=COST_CEILING)
    ap.add_argument("--input-suffix", default="v15.11", help="Cascade input suffix")
    args = ap.parse_args()

    batch_id  = args.batch_id
    batch_dir = ROOT / "batches" / f"jawal-{batch_id}"
    raw_root  = batch_dir / "raw"
    invoice_id = args.invoice_id or batch_id

    if not batch_dir.exists():
        print(f"[fatal] batch dir not found: {batch_dir}", file=sys.stderr)
        sys.exit(2)

    cascade_pattern = f"Spreadsheet-{invoice_id}-FILLED-{args.input_suffix}*.xlsx"
    candidates = sorted(
        (batch_dir / "output").glob(cascade_pattern),
        key=lambda p: p.stat().st_mtime, reverse=True
    )
    if not candidates:
        print(f"[fatal] cascade output missing: {batch_dir}/output/{cascade_pattern}", file=sys.stderr)
        sys.exit(3)
    cascade_xlsx = candidates[0]
    print(f"[load] cascade output: {cascade_xlsx.name}", flush=True)

    out_dir  = batch_dir / "output"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_xlsx = out_dir / f"Spreadsheet-{invoice_id}-FILLED-{V16}.xlsx"
    if out_xlsx.exists():
        backup_file(out_xlsx)

    cascade_rows, headers, hdr_row = read_cascade_xlsx(cascade_xlsx)
    print(f"[load] {len(cascade_rows)} cascade rows", flush=True)
    if args.limit > 0:
        cascade_rows = cascade_rows[:args.limit]

    print("[load] master data...", flush=True)
    manpower = fea.load_manpower()
    lookups  = fea.load_lookups()
    all_folders = fea.collect_all_folders(raw_root)
    print(f"[scan] {len(all_folders)} ticket-folder candidates in {raw_root}", flush=True)

    # Routing pass
    routed = route_rows(cascade_rows, all_folders)
    reason_counts: dict[str, int] = {}
    for _, r in routed:
        reason_counts[r] = reason_counts.get(r, 0) + 1
    print(f"[route] {len(routed)} rows routed / {len(cascade_rows)-len(routed)} stay cascade", flush=True)
    for r, n in reason_counts.items():
        print(f"        {r}: {n}", flush=True)

    # Init hybrid rows
    hybrid_rows = []
    for c in cascade_rows:
        hybrid_rows.append({
            "_row_idx":     c["_row_idx"],
            "_agent_method":"cascade",
            "emp_no":       c.get("Employee No", "") or "",
            "account":      c.get("Account", "") or "",
            "cost_center":  c.get("Cost Center", "") or "",
            "div":          c.get("DIV", "") or "",
            "solution":     c.get("Solution", "") or "",
            "agency":       c.get("Agency", "") or "",
            "location":     c.get("Location", "") or "",
        })

    # LLM resolution
    total_in = total_out = llm_errors = 0
    cost_so_far = 0.0
    pricing_pro = fea.GEMINI_PRICING["gemini-pro-latest"]
    classify_in_tokens = classify_out_tokens = 0

    def worker(idx_reason):
        idx, reason = idx_reason
        try:
            row = cascade_rows[idx]
            result = llm_resolve_row_v16(
                idx, row, batch_id, raw_root, all_folders, manpower, lookups, reason
            )
            return idx, reason, result
        except Exception as e:
            return idx, reason, {"_llm_error": f"{type(e).__name__}: {e}", "_traceback": traceback.format_exc()}

    t_start = time.time()
    done = 0

    with concurrent.futures.ThreadPoolExecutor(max_workers=args.workers) as ex:
        futs = [ex.submit(worker, (i, r)) for i, r in routed]
        for fut in concurrent.futures.as_completed(futs):
            idx, reason, llm = fut.result()
            done += 1

            in_tok  = llm.get("_llm_in_tokens", 0) + llm.get("_classify_in_tokens", 0)
            out_tok = llm.get("_llm_out_tokens", 0) + llm.get("_classify_out_tokens", 0)
            total_in  += in_tok
            total_out += out_tok
            classify_in_tokens  += llm.get("_classify_in_tokens", 0)
            classify_out_tokens += llm.get("_classify_out_tokens", 0)

            if llm.get("_llm_error") and not llm.get("_from_cache"):
                llm_errors += 1

            cost_so_far = (total_in * pricing_pro["in"] + total_out * pricing_pro["out"]) / 1_000_000
            elapsed = time.time() - t_start
            print(
                f"[{done:3d}/{len(routed)}] row {idx:3d} ({reason})"
                f"  t={elapsed:5.1f}s  cost~=${cost_so_far:.3f}"
                f"  v16_type={llm.get('_v16_row_type','')}"
                f"  from_cache={llm.get('_from_cache',False)}",
                flush=True
            )

            if cost_so_far > args.cost_ceiling:
                print(f"[ABORT] cost ceiling ${args.cost_ceiling} exceeded", file=sys.stderr)
                for f in futs:
                    f.cancel()
                break

            if llm.get("_llm_error"):
                print(f"  [error] {llm['_llm_error']}", flush=True)
                continue

            # Apply overlays
            final, method = apply_overlays(llm, cascade_rows[idx])
            hybrid_rows[idx]["_agent_method"] = method
            for k in ("emp_no","account","cost_center","div","solution","agency","location"):
                if final.get(k):
                    hybrid_rows[idx][k] = final[k]

    # Family cluster unification (same as v15.12)
    cluster_unified = apply_cluster_unification(hybrid_rows, cascade_rows)

    # Write output
    write_v16_xlsx(cascade_xlsx, out_xlsx, hybrid_rows, cascade_rows, hdr_row)
    print(f"[out] wrote {out_xlsx}", flush=True)

    # Summary
    method_counts = {}
    for h in hybrid_rows:
        m = h.get("_agent_method","cascade")
        method_counts[m] = method_counts.get(m,0) + 1

    summary = {
        "batch_id":            batch_id,
        "version":             V16,
        "timestamp_utc":       utc_ts(),
        "total_rows":          len(cascade_rows),
        "routed_to_llm":       len(routed),
        "routing_reasons":     reason_counts,
        "method_counts":       method_counts,
        "cluster_unified_rows":cluster_unified,
        "classify_in_tokens":  classify_in_tokens,
        "classify_out_tokens": classify_out_tokens,
        "llm_in_tokens":       total_in,
        "llm_out_tokens":      total_out,
        "llm_cascade":         fea.GEMINI_MODEL_CASCADE,
        "cost_usd_est":        round(cost_so_far, 4),
        "cost_ceiling":        args.cost_ceiling,
        "llm_errors":          llm_errors,
        "runtime_sec":         round(time.time() - t_start, 1),
        "cascade_input":       str(cascade_xlsx),
        "output":              str(out_xlsx),
    }
    summary_path = out_dir / f"summary-{V16}.json"
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2))

    print()
    print("=" * 60)
    print(f"  Hybrid v16 done")
    print(f"  Rows:           {len(cascade_rows)}  ({len(routed)} LLM-routed)")
    print(f"  Methods:        {method_counts}")
    print(f"  TRAVEL_NO_EVIDENCE routed: {reason_counts.get('TRAVEL_NO_EVIDENCE',0)}")
    print(f"  Classify tokens: {classify_in_tokens} in / {classify_out_tokens} out")
    print(f"  Total cost (est): ${cost_so_far:.4f}")
    print(f"  Runtime:        {summary['runtime_sec']}s")
    print(f"  Output:         {out_xlsx}")
    print("=" * 60)


if __name__ == "__main__":
    main()
