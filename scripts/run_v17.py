#!/usr/bin/env python3
"""
AlJeel AP Pipeline v17 — Booking Group Detection

Layers on top of v16 output as a deterministic post-processing pass.
No LLM calls. Zero additional cost.

PROBLEM BEING SOLVED
────────────────────
Family trips produce multiple Jawal tickets (adult + spouse + children) with
consecutive ticket numbers and the same itinerary string. When the .msg evidence
is corrupt, v16 can assign different employees to different family members — or
leave dependents at 999999 — because it resolves each ticket in isolation.

v17 detects these booking groups and enforces consistency:
  1. Group tickets by consecutive number (within TICKET_GAP=5) + identical route
  2. Find the group anchor: row with the most reliable resolution
  3. Propagate anchor's full GL combo + emp_no to unresolved members
  4. CHD/INF passengers always inherit from anchor (never standalone)
  5. FAMILY_CONFLICT flag when two rows in the same group have different valid combos

EXAMPLES FIXED
──────────────
J26-550 Mahmoud Elkilany Hussein (1002466) family:
  6904732429 HUSSEIN/TALIA MS(CHD)  → inherited 1002466 / 160014 / 170 / 10239
  6904732430 HUSSEIN/YOUSSEF MR     → inherited 1002466 / 160014 / 170 / 10239
  6904732431 SALEM/EFFAT MRS        → inherited 1002466 / 160014 / 170 / 10239

Usage:
    python3 scripts/run_v17.py J26-550
    python3 scripts/run_v17.py J26-640
    python3 scripts/run_v17.py J26-593
    python3 scripts/run_v17.py J26-589
    python3 scripts/run_v17.py J26-788 --input-suffix v15.11.2
"""
from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
import time
from collections import Counter
from pathlib import Path
from typing import Optional

import openpyxl

ROOT = Path("/home/clawdbot/.openclaw/workspace/aljeel")
sys.path.insert(0, str(ROOT / "scripts"))

VERSION = "v17.1"

# ─── constants ────────────────────────────────────────────────────────────────

TICKET_GAP_MAX = 5      # max gap between consecutive ticket numbers in same booking
UNRESOLVED_CC  = {"999999", "000000", "", "0"}
DEPENDENT_RE   = re.compile(r"\b(CHD|INF|INFANT)\b", re.IGNORECASE)
TICKET_10_RE   = re.compile(r"\((\d{10})\)")
ROUTE_RE       = re.compile(r" - (.+?) \(\d{10}\)")
# Oracle GL combo pattern: 03-LOCATION-ACCOUNT-CC-DIV-SOLUTION-AGENCY-00000-00-000000
COMBO_RE       = re.compile(r"^\d{2}-\d{5}-\d{8}-(\d{6})-")


# ─── helpers ──────────────────────────────────────────────────────────────────

def utc_ts() -> str:
    return time.strftime("%Y%m%dT%H%M%SZ", time.gmtime())


def extract_ticket_no(desc: str) -> Optional[int]:
    m = TICKET_10_RE.search(desc)
    return int(m.group(1)) if m else None


def extract_route(desc: str) -> str:
    """Returns the itinerary string, e.g. 'RUH JED HBE RUH'."""
    m = ROUTE_RE.search(desc)
    return m.group(1).strip() if m else ""


def is_dependent(desc: str) -> bool:
    """True if the passenger is explicitly marked as a child or infant."""
    return bool(DEPENDENT_RE.search(desc))


def combo_cc(combo: str) -> str:
    """Extract CC segment from GL combo string; '' if unparseable."""
    m = COMBO_RE.match(combo or "")
    return m.group(1) if m else ""


def is_resolved(combo: str) -> bool:
    cc = combo_cc(combo)
    return cc not in UNRESOLVED_CC


def combo_emp_key(combo: str, emp_no: str) -> str:
    """Stable key to detect conflicting resolutions within a group."""
    return f"{combo or ''}||{emp_no or ''}"


# ─── GL breakdown column names ──────────────────────────────────────────
# These are the individual segment columns that live alongside the combo string
# and must stay in sync with it.

# Columns whose values are parsed directly from the combo string segments
COMBO_SEGMENT_COLS = [
    "Company",       # combo[0]
    "Location",      # combo[1]
    "Account",       # combo[2]
    "Cost Center",   # combo[3]
    "DIV",           # combo[4]
    "Solution",      # combo[5]
    "Agency",        # combo[6]
    "Project",       # combo[7]
    "Intercompany",  # combo[8]
    "Future 1",      # combo[9]
]

# Columns whose values are human-readable lookups — copy from anchor row
COMBO_LOOKUP_COLS = [
    "GL",             # account name
    "Cost Name",      # CC name
    "Contribution",   # DIV name
    "Solution Name",  # solution name
    "Agency Name",    # agency name
    "GL Description", # composite text
]

ALL_GL_BREAKDOWN_COLS = COMBO_SEGMENT_COLS + COMBO_LOOKUP_COLS


def parse_combo_segments(combo: str) -> dict[str, str]:
    """
    Parse GL combo string into individual segment values.
    Format: 03-10100-60301003-160014-170-00000-10239-00000-00-000000
    """
    parts = (combo or "").split("-")
    # Pad to at least 10 segments
    while len(parts) < 10:
        parts.append("")
    return {
        "Company":      parts[0],
        "Location":     parts[1],
        "Account":      parts[2],
        "Cost Center":  parts[3],
        "DIV":          parts[4],
        "Solution":     parts[5],
        "Agency":       parts[6],
        "Project":      parts[7],
        "Intercompany": parts[8],
        "Future 1":     parts[9],
    }


# ─── XLSX read / write ────────────────────────────────────────────────────────

HEADER_ROW_IDX = 3   # 1-based; row 3 in the Oracle Fusion template

def read_xlsx_rows(path: Path) -> tuple[list[str], list[dict], int]:
    """
    Returns (headers, rows_as_dicts, header_row_1indexed).
    rows_as_dicts includes '_row_1idx' (1-based row index in sheet).
    """
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Auto-detect header row: first row where >3 cells are non-None
    hdr_row = HEADER_ROW_IDX
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), start=1):
        non_null = sum(1 for v in row if v is not None)
        if non_null > 3:
            hdr_row = i
            break

    headers = [c.value for c in ws[hdr_row]]
    rows = []
    for r in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        if all(v is None for v in r):
            continue
        d = dict(zip(headers, r))
        d["_row_1idx"] = ws.min_row + (hdr_row - ws.min_row) + len(rows) + 1
        rows.append(d)
    return headers, rows, hdr_row


def write_xlsx_patches(src: Path, dst: Path, patches: dict[int, dict]) -> None:
    """
    Copy src → dst, then apply patches.
    patches: {row_1idx: {header_name: new_value}}
    """
    shutil.copy2(src, dst)
    wb = openpyxl.load_workbook(dst)
    ws = wb.active

    # Build col index from header row
    hdr_row = HEADER_ROW_IDX
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), start=1):
        if sum(1 for v in row if v is not None) > 3:
            hdr_row = i
            break
    headers = [c.value for c in ws[hdr_row]]
    col_map = {h: j + 1 for j, h in enumerate(headers) if h is not None}

    for row_1idx, field_vals in patches.items():
        for field, val in field_vals.items():
            if field in col_map:
                ws.cell(row=row_1idx, column=col_map[field]).value = val

    wb.save(dst)


# ─── booking group detection ──────────────────────────────────────────────────

def build_booking_groups(rows: list[dict]) -> list[list[int]]:
    """
    Returns list of groups; each group is a list of indices into `rows`.
    Only groups with 2+ members are returned.

    Grouping criteria:
      - Ticket numbers within TICKET_GAP_MAX of each other (consecutive booking)
      - Identical route string (same itinerary — guards against coincidental proximity)
    """
    # Collect (ticket_int, route, row_index)
    indexed: list[tuple[int, str, int]] = []
    for i, row in enumerate(rows):
        desc   = str(row.get("Description", "") or "")
        ticket = extract_ticket_no(desc)
        route  = extract_route(desc)
        if ticket and route:
            indexed.append((ticket, route, i))

    indexed.sort(key=lambda x: x[0])

    groups: list[list[int]] = []
    current: list[tuple[int, str, int]] = []

    for ticket, route, idx in indexed:
        if not current:
            current = [(ticket, route, idx)]
        else:
            prev_ticket, prev_route, _ = current[-1]
            same_route = (route == prev_route)
            close      = (ticket - prev_ticket) <= TICKET_GAP_MAX
            if same_route and close:
                current.append((ticket, route, idx))
            else:
                if len(current) >= 2:
                    groups.append([x[2] for x in current])
                current = [(ticket, route, idx)]

    if len(current) >= 2:
        groups.append([x[2] for x in current])

    return groups


def resolve_combo_col(rows: list[dict]) -> Optional[str]:
    """
    Resolve the Distribution Combination column from the actual row headers.
    Source files vary between 2 and 3 dots in the literal ("[..]" vs "[...]"),
    so a hardcoded name silently no-ops on half the batches.
    """
    headers = rows[0].keys() if rows else []
    return next((h for h in headers if "distribution combination" in str(h).lower()), None)


def _pick_anchor(group_indices: list[int], rows: list[dict]) -> Optional[int]:
    """
    Select the best anchor row index from a group.
    Priority:
      1. Non-dependent (MR/MS adult) with resolved combo + non-empty emp_no
      2. Any resolved combo with emp_no
      3. Any resolved combo
    Returns row index into `rows`, or None if nothing resolved.
    """
    combo_col = resolve_combo_col(rows)
    empno_col = "Employee No"

    candidates: list[tuple[int, int, int]] = []  # (priority, ticket_no, row_idx)
    for i in group_indices:
        row   = rows[i]
        desc  = str(row.get("Description", "") or "")
        combo = str(row.get(combo_col, "") or "")
        emp   = str(row.get(empno_col, "") or "").strip()

        if not is_resolved(combo):
            continue

        is_dep   = is_dependent(desc)
        has_emp  = bool(emp and emp not in ("0", "None"))
        ticket   = extract_ticket_no(desc) or 0

        if not is_dep and has_emp:
            priority = 0
        elif has_emp:
            priority = 1
        elif not is_dep:
            priority = 2
        else:
            priority = 3

        # v24: rows whose combo was rebuilt from the Account column (PC override)
        # have high-fidelity CC/agency from master. Boost them to priority=0
        # so they win over rows resolved from contaminated evidence folders.
        if rows[i].get("_pc_resynced"):
            priority = 0

        candidates.append((priority, ticket, i))

    if not candidates:
        return None
    candidates.sort()
    return candidates[0][2]


# ─── family group filters ─────────────────────────────────────────────────────

SURNAME_RE = re.compile(r'^([A-Z][A-Z\s]+)/')


def _group_has_dependent(group_indices: list[int], rows: list[dict]) -> bool:
    """True if any member in the group is CHD/INF."""
    return any(is_dependent(str(rows[i].get("Description", "") or "")) for i in group_indices)


def _group_has_shared_surname(group_indices: list[int], rows: list[dict]) -> bool:
    """True if 2+ members share the same surname (LASTNAME/ format)."""
    surnames: list[str] = []
    for i in group_indices:
        desc = str(rows[i].get("Description", "") or "")
        m = SURNAME_RE.match(desc)
        if m:
            surnames.append(m.group(1).strip())
    counts = Counter(surnames)
    return any(v >= 2 for v in counts.values())


def _group_has_shared_empno(group_indices: list[int], rows: list[dict]) -> bool:
    """
    True if 2+ members carry the same non-blank Employee No. That column comes
    from the invoice Ref. No. (e.g. "1000862 family"), so a shared value marks
    a family booking even with no CHD/INF rows and different surnames
    (employee + spouse, spouse-only travel).
    """
    emps = [str(rows[i].get("Employee No", "") or "").strip() for i in group_indices]
    emps = [e for e in emps if e and e not in ("0", "None")]
    counts = Counter(emps)
    return any(v >= 2 for v in counts.values())


def _is_family_group(group_indices: list[int], rows: list[dict]) -> bool:
    """
    A group is a genuine family booking — not just colleagues on the same flight —
    when it contains a CHD/INF passenger, OR 2+ members share the same surname,
    OR 2+ members share the same Ref. No. emp_no. A CHD row is one signal, not a
    requirement: spouse-only travel is also a family booking.

    Colleagues traveling together (ALMUTAIRI + ALUTHMAN on JED JFK, etc.) will have
    different surnames, no dependents, and distinct emp_nos, so they are skipped.
    """
    return (
        _group_has_dependent(group_indices, rows)
        or _group_has_shared_surname(group_indices, rows)
        or _group_has_shared_empno(group_indices, rows)
    )


def _row_family_annual_locked(row: dict) -> bool:
    """run_v30 stage 3i finalised this row's account + segments (family annual
    ticket → 21070229 + G&A segments). The lock travels as the in-memory key
    on hybrid rows and as FAMILY_ANNUAL_LOCKED in the Agent Flags column on
    rows re-read from the output xlsx. Locked rows may anchor a group but must
    never be overwritten."""
    if row.get("_family_annual_locked"):
        return True
    return "FAMILY_ANNUAL_LOCKED" in str(row.get("Agent Flags", "") or "")


def apply_booking_group_propagation(
    rows: list[dict],
    verbose: bool = True,
) -> tuple[int, int, list[str]]:
    """
    Detect booking groups and propagate anchor allocations.
    Modifies rows in-place (adds _v17_method marker).

    Returns:
        (propagated_count, conflict_count, conflict_descriptions)
    """
    combo_col = resolve_combo_col(rows)
    empno_col = "Employee No"
    if combo_col is None:
        print("  [v17] WARNING: no 'Distribution Combination' column found — propagation skipped", flush=True)
        return 0, 0, []

    groups     = build_booking_groups(rows)
    propagated = 0
    conflicts  = 0
    conflict_descs: list[str] = []

    for group_indices in groups:
        # Skip colleague groups — only act on genuine family bookings
        if not _is_family_group(group_indices, rows):
            if verbose:
                descs = [str(rows[i].get("Description","") or "")[:40] for i in group_indices]
                print(f"  [skip-colleagues] {' | '.join(descs)}", flush=True)
            continue
        group_pc_wins = False   # set True when PC-wins resolution was applied
        # ── check for conflicts in this group ──────────────────────────────
        resolved_keys: dict[str, list[int]] = {}
        for i in group_indices:
            combo = str(rows[i].get(combo_col, "") or "")
            emp   = str(rows[i].get(empno_col, "") or "").strip()
            if is_resolved(combo):
                key = combo_emp_key(combo, emp)
                resolved_keys.setdefault(key, []).append(i)

        if len(resolved_keys) > 1:
            # Multiple different resolutions — check if one is PC (21070229)
            # v22: PC account is the strongest signal in a family group.
            # If any row has account=21070229 (personal contribution), all family members
            # should inherit it — the PC row is the anchor, travel rows are overridden.
            pc_key = next((k for k in resolved_keys if "21070229" in k.split("||")[0]), None)
            if pc_key:
                # Promote the PC row to the only anchor and clear the conflict
                pc_combo, pc_emp = pc_key.split("||")
                # Re-write resolved_keys to only the PC resolution
                resolved_keys = {pc_key: resolved_keys[pc_key]}
                # Flag: when pc-wins, override ALL non-anchor rows (even if "resolved" with wrong CC)
                group_pc_wins = True
                if verbose:
                    print(f"  [pc-wins] PC account 21070229 elected as family anchor — travel entries overridden", flush=True)
            else:
                conflicts += 1
                descs = [str(rows[i].get("Description",""))[:60] for i in group_indices]
                msg   = f"FAMILY_CONFLICT tickets=[{', '.join(descs)}]"
                conflict_descs.append(msg)
                if verbose:
                    print(f"  [conflict] {msg}", flush=True)
            # Continue propagation with whichever anchor was selected.

        # ── pick anchor ────────────────────────────────────────────────────
        anchor_idx = _pick_anchor(group_indices, rows)
        if anchor_idx is None:
            continue   # entire group unresolved; nothing to propagate

        anchor_combo = str(rows[anchor_idx].get(combo_col, "") or "")
        anchor_emp   = str(rows[anchor_idx].get(empno_col, "") or "").strip()

        # Sync anchor's own breakdown columns to its combo string
        # (manual col-N patches may have left Q-AF stale)
        if not _row_family_annual_locked(rows[anchor_idx]):
            anchor_segments = parse_combo_segments(anchor_combo)
            for col, val in anchor_segments.items():
                if str(rows[anchor_idx].get(col, "") or "") != str(val):
                    rows[anchor_idx][col] = val

        # ── propagate to unresolved + all dependents, sync breakdowns for all ──
        for i in group_indices:
            if i == anchor_idx:
                continue

            if _row_family_annual_locked(rows[i]):
                if verbose:
                    desc = str(rows[i].get("Description", "") or "")[:40]
                    print(f"  [lock-skip] family-annual locked, not propagating onto: {desc}", flush=True)
                continue

            row   = rows[i]
            desc  = str(row.get("Description", "") or "")
            combo = str(row.get(combo_col, "") or "")
            is_dep = is_dependent(desc)
            was_resolved = is_resolved(combo)

            # Check if breakdown columns are stale (combo was manually patched
            # in a prior version but Q-AF were never updated)
            combo_cc_val = combo_cc(combo)
            breakdown_cc = str(row.get("Cost Center", "") or "")
            breakdown_stale = (was_resolved and str(combo_cc_val) != str(breakdown_cc))

            # When PC-wins fired for this group, override ALL non-anchor members
            # (including resolved adults with wrong CC from contaminated evidence)
            needs_fix = (not was_resolved) or is_dep or breakdown_stale or group_pc_wins

            if needs_fix:
                old_combo = combo
                old_emp   = str(row.get(empno_col, "") or "")

                # ─ update combo string + emp_no ────────────────────────────────
                rows[i][combo_col] = anchor_combo
                rows[i][empno_col] = anchor_emp if anchor_emp else old_emp

                # ─ update individual segment columns from combo ───────────────
                # Parse directly from the combo string so we never inherit
                # stale breakdown values from the anchor row.
                segments = parse_combo_segments(anchor_combo)
                for col, val in segments.items():
                    rows[i][col] = val

                # ─ copy lookup/name columns from anchor row ────────────────
                anchor_row = rows[anchor_idx]
                for col in COMBO_LOOKUP_COLS:
                    if col in anchor_row and anchor_row[col] is not None:
                        rows[i][col] = anchor_row[col]

                rows[i]["_v17_method"]       = "group_dep_fix" if is_dep else "group_propagation"
                rows[i]["_v17_anchor_row"]   = anchor_idx
                rows[i]["_v17_was_resolved"] = was_resolved
                propagated += 1
                if verbose:
                    fix_type = "dep_fix" if is_dep else "propagation"
                    print(
                        f"  [{fix_type}] {desc[:60]}"
                        f"\n              combo: {old_combo} → {anchor_combo}"
                        f"\n              emp:   {old_emp} → {anchor_emp}",
                        flush=True,
                    )

    return propagated, conflicts, conflict_descs


# ─── XLSX patch writer ────────────────────────────────────────────────────────

def collect_patches(
    rows: list[dict],
    orig_rows: list[dict],
) -> dict[int, dict]:
    """
    Compare modified rows against original rows.
    Return {row_1idx: {field: new_value}} for any changed cell.
    Tracks combo string, emp_no, and all individual GL breakdown columns.
    """
    combo_col = resolve_combo_col(rows)
    empno_col = "Employee No"
    tracked   = ([combo_col] if combo_col else []) + [empno_col] + ALL_GL_BREAKDOWN_COLS

    patches: dict[int, dict] = {}
    for row, orig in zip(rows, orig_rows):
        row_1idx = row["_row_1idx"]
        for col in tracked:
            new_val = row.get(col)
            old_val = orig.get(col)
            if str(new_val or "") != str(old_val or ""):
                patches.setdefault(row_1idx, {})[col] = new_val
    return patches


# ─── QC catch writer ─────────────────────────────────────────────────────────

def write_catches(
    batch_dir: Path,
    propagated: int,
    conflicts: int,
    conflict_descs: list[str],
    groups: list[list[int]],
    rows: list[dict],
) -> Path:
    combo_col = resolve_combo_col(rows)
    catches: list[dict] = []

    # Propagation catches
    for i, row in enumerate(rows):
        if row.get("_v17_method") in ("group_propagation", "group_dep_fix"):
            desc  = str(row.get("Description","") or "")
            combo = str(row.get(combo_col,"") or "")
            catches.append({
                "catch_type":   "BOOKING_GROUP_PROPAGATED",
                "severity":     "INFO",
                "description":  desc,
                "new_combo":    combo,
                "method":       row.get("_v17_method",""),
                "was_resolved": row.get("_v17_was_resolved", False),
            })

    # Conflict catches
    for msg in conflict_descs:
        catches.append({
            "catch_type":  "FAMILY_CONFLICT",
            "severity":    "HIGH",
            "description": msg,
            "new_combo":   "",
            "method":      "conflict",
            "was_resolved": False,
        })

    out_path = batch_dir / "output" / f"catches-{VERSION}.json"
    out_path.write_text(json.dumps(catches, indent=2, default=str))
    return out_path


# ─── scoring (if ground truth available) ─────────────────────────────────────

def score_if_truth_available(out_xlsx: Path, batch_dir: Path, batch_id: str) -> Optional[float]:
    truth_candidates = list(batch_dir.glob(f"{batch_id}.xlsx"))
    if not truth_candidates:
        return None
    try:
        import subprocess
        qc_script = ROOT / "qc" / "score_against_truth.py"
        result = subprocess.run(
            ["python3", str(qc_script), str(out_xlsx), str(truth_candidates[0])],
            capture_output=True, text=True, timeout=120,
        )
        # Parse score from last line: "All-5: 84.7%"
        for line in reversed(result.stdout.splitlines()):
            m = re.search(r"All-5:\s*([\d.]+)%", line)
            if m:
                return float(m.group(1))
    except Exception:
        pass
    return None


# ─── main ─────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description="AlJeel AP Pipeline v17 — Booking Group Detection")
    ap.add_argument("batch_id")
    ap.add_argument("--invoice-id", default=None)
    ap.add_argument("--input-suffix", default="v16",
                    help="Version suffix of the input XLSX (default: v16)")
    args = ap.parse_args()

    batch_id   = args.batch_id
    invoice_id = args.invoice_id or batch_id
    batch_dir  = ROOT / "batches" / f"jawal-{batch_id}"
    out_dir    = batch_dir / "output"

    if not batch_dir.exists():
        print(f"[fatal] batch dir not found: {batch_dir}", file=sys.stderr)
        sys.exit(2)

    # ── find input (v16 output) ───────────────────────────────────────────
    # Prefer exact suffix match (e.g. v16.xlsx) before falling back to v16*.xlsx
    # This prevents v16.1 / v16.2 manual sheet-fix files from being used as input.
    exact_name = f"Spreadsheet-{invoice_id}-FILLED-{args.input_suffix}.xlsx"
    exact_path = out_dir / exact_name
    if exact_path.exists():
        in_xlsx = exact_path
    else:
        pattern    = f"Spreadsheet-{invoice_id}-FILLED-{args.input_suffix}*.xlsx"
        candidates = sorted(out_dir.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
        # Exclude already-generated v17 files and manual sheet-fix variants (v16.1, v16.2...)
        candidates = [c for c in candidates if "v17" not in c.name
                      and c.name == exact_name  # only exact
                      or ("v17" not in c.name and "." not in c.stem.replace(f"-{args.input_suffix}", ""))]
        if not candidates:
            print(f"[fatal] no exact input found: {exact_path}", file=sys.stderr)
            sys.exit(3)
        in_xlsx = candidates[0]
    print(f"[load] input: {in_xlsx.name}", flush=True)

    # ── output path ───────────────────────────────────────────────────────
    out_xlsx = out_dir / f"Spreadsheet-{invoice_id}-FILLED-{VERSION}.xlsx"
    if out_xlsx.exists():
        bak = out_xlsx.with_suffix(out_xlsx.suffix + f".bak-pre-v17-{utc_ts()}")
        shutil.copy2(out_xlsx, bak)
        print(f"[backup] {bak.name}", flush=True)

    # ── read rows ─────────────────────────────────────────────────────────
    headers, rows, hdr_row = read_xlsx_rows(in_xlsx)
    print(f"[load] {len(rows)} rows, header row {hdr_row}", flush=True)

    # Save original state for patch diff
    import copy
    orig_rows = copy.deepcopy(rows)

    # ── detect groups (dry run for stats) ─────────────────────────────────
    groups = build_booking_groups(rows)
    print(f"[groups] found {len(groups)} booking groups across {sum(len(g) for g in groups)} rows", flush=True)
    for g in groups:
        desc_list = [str(rows[i].get("Description","") or "")[:50] for i in g]
        print(f"  group({len(g)}): {' | '.join(desc_list)}", flush=True)

    # ── apply propagation ─────────────────────────────────────────────────
    print("\n[v17] applying booking group propagation...", flush=True)
    propagated, conflicts, conflict_descs = apply_booking_group_propagation(rows, verbose=True)
    print(f"\n[v17] propagated={propagated}  conflicts={conflicts}", flush=True)

    # ── collect patches and write output XLSX ────────────────────────────
    patches = collect_patches(rows, orig_rows)
    if patches:
        write_xlsx_patches(in_xlsx, out_xlsx, patches)
        print(f"[out] {out_xlsx.name} ({len(patches)} rows patched)", flush=True)
    else:
        # No changes — copy input as-is (still rename to v17)
        shutil.copy2(in_xlsx, out_xlsx)
        print(f"[out] {out_xlsx.name} (no changes; copied from {in_xlsx.name})", flush=True)

    # ── write catches JSON ────────────────────────────────────────────────
    catches_path = write_catches(batch_dir, propagated, conflicts, conflict_descs, groups, rows)
    print(f"[catches] {catches_path.name}", flush=True)

    # ── score if ground truth available ───────────────────────────────────
    score = score_if_truth_available(out_xlsx, batch_dir, batch_id)
    score_str = f"{score:.1f}%" if score is not None else "N/A (no truth file)"

    # ── write summary JSON ────────────────────────────────────────────────
    summary = {
        "batch_id":       batch_id,
        "version":        VERSION,
        "timestamp_utc":  utc_ts(),
        "total_rows":     len(rows),
        "groups_found":   len(groups),
        "rows_in_groups": sum(len(g) for g in groups),
        "propagated":     propagated,
        "conflicts":      conflicts,
        "conflict_details": conflict_descs,
        "patches_written": len(patches),
        "score_all5":     score_str,
        "input":          str(in_xlsx),
        "output":         str(out_xlsx),
    }
    (out_dir / f"summary-{VERSION}.json").write_text(json.dumps(summary, indent=2, default=str))

    print()
    print("=" * 65)
    print(f"  v17 done — {batch_id}")
    print(f"  Groups found:  {len(groups)}  ({summary['rows_in_groups']} rows)")
    print(f"  Propagated:    {propagated}")
    print(f"  Conflicts:     {conflicts}  {'⚠️  review catches JSON' if conflicts else '✅'}")
    print(f"  Score (all-5): {score_str}")
    print(f"  Output:        {out_xlsx.name}")
    print("=" * 65)


if __name__ == "__main__":
    main()
