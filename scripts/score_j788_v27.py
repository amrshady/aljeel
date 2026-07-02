#!/usr/bin/env python3
"""
Direct field comparison: v27 output vs J26-788 ground truth.
Both files are Oracle template format; matches on ticket from Description col.
"""
import re, sys
from pathlib import Path
import openpyxl

BATCH = Path("/home/clawdbot/.openclaw/workspace/aljeel/batches/jawal-J26-788/output")
V27   = BATCH / "Spreadsheet-J26-788-FILLED-v27.xlsx"
TRUTH = BATCH / "J26-788-GROUND-TRUTH.xlsx"
V26   = BATCH / "Spreadsheet-J26-788-FILLED-v26.xlsx"

SCORE_FIELDS = ["Employee No", "Account", "Cost Center", "DIV", "Solution", "Agency"]
TICKET_RE    = re.compile(r"(?<![\d])(\d{10}|26-\d{3,4})(?![\d])")

def load_file(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    # Row 3 = headers
    headers = [str(c.value).strip() if c.value else "" for c in ws[3]]
    col_map  = {h: i for i, h in enumerate(headers) if h}
    desc_col = col_map.get("Description") or next((col_map[h] for h in col_map if "description" in h.lower()), None)
    rows = {}
    for r in ws.iter_rows(min_row=4, values_only=True):
        if not r[desc_col]:
            continue
        desc = str(r[desc_col])
        for m in TICKET_RE.finditer(desc):
            ticket = m.group(1)
            row_dict = {h: r[i] for h, i in col_map.items()}
            rows[ticket] = row_dict
            break
    return rows, col_map

def val(row, field):
    v = row.get(field, "")
    if v is None: return ""
    return str(v).strip().lstrip("0") or "0"

def main():
    truth_rows, _ = load_file(TRUTH)
    v27_rows, _   = load_file(V27)
    v26_rows, _   = load_file(V26)

    common = sorted(set(truth_rows) & set(v27_rows))
    print(f"Truth rows: {len(truth_rows)} | v27 rows: {len(v27_rows)} | Common: {len(common)}\n")

    total = len(common)
    v27_all5 = 0
    v26_all5 = 0
    per_field_v27 = {f: 0 for f in SCORE_FIELDS}
    per_field_v26 = {f: 0 for f in SCORE_FIELDS}
    diffs = []

    for ticket in common:
        t  = truth_rows[ticket]
        p7 = v27_rows[ticket]
        p6 = v26_rows.get(ticket, {})

        v27_match = all(val(p7, f) == val(t, f) for f in SCORE_FIELDS)
        v26_match = all(val(p6, f) == val(t, f) for f in SCORE_FIELDS)
        if v27_match: v27_all5 += 1
        if v26_match: v26_all5 += 1

        for f in SCORE_FIELDS:
            if val(p7, f) == val(t, f): per_field_v27[f] += 1
            if val(p6, f) == val(t, f): per_field_v26[f] += 1

        if not v27_match:
            row_diffs = [(f, val(t, f), val(p7, f)) for f in SCORE_FIELDS if val(p7, f) != val(t, f)]
            diffs.append((ticket, str(t.get("Description",""))[:55], row_diffs))

    print("=" * 68)
    print(f"  SCORE vs GROUND TRUTH  (n={total})")
    print("=" * 68)
    print(f"  v26  all-5-exact: {v26_all5}/{total} = {100*v26_all5/total:.1f}%")
    print(f"  v27  all-5-exact: {v27_all5}/{total} = {100*v27_all5/total:.1f}%")
    delta = v27_all5 - v26_all5
    print(f"  Δ v27 vs v26: {'+' if delta>=0 else ''}{delta} rows  ({'+' if delta>=0 else ''}{100*delta/total:.1f}pp)")
    print()
    print("  Per-field match rate (v26 → v27):")
    print(f"  {'Field':<15} {'v26':>6} {'v27':>6}")
    print(f"  {'-'*30}")
    for f in SCORE_FIELDS:
        pct6 = 100*per_field_v26[f]/total
        pct7 = 100*per_field_v27[f]/total
        arrow = "✅" if pct7 > pct6 else ("➡" if pct7==pct6 else "⚠️")
        print(f"  {f:<15} {pct6:5.1f}% {pct7:5.1f}% {arrow}")
    print()

    if diffs:
        print(f"  Remaining mismatches in v27 ({len(diffs)} rows):")
        print()
        for ticket, desc, row_diffs in diffs:
            print(f"  [{ticket}] {desc}")
            for f, tv, pv in row_diffs:
                print(f"      {f}: truth={tv!r}  v27={pv!r}")
        print()

    # Write score file
    score_md = BATCH / "score-v27.md"
    with open(score_md, "w") as fh:
        fh.write(f"# J26-788 v27 Score vs Ground Truth\n\n")
        fh.write(f"- Total rows: {total}\n")
        fh.write(f"- v26 all-5-exact: {v26_all5}/{total} = {100*v26_all5/total:.1f}%\n")
        fh.write(f"- **v27 all-5-exact: {v27_all5}/{total} = {100*v27_all5/total:.1f}%**\n")
        fh.write(f"- Δ vs v26: {'+' if delta>=0 else ''}{delta} rows ({'+' if delta>=0 else ''}{100*delta/total:.1f}pp)\n\n")
        fh.write("## Per-field\n\n| Field | v26 | v27 |\n|---|---|---|\n")
        for f in SCORE_FIELDS:
            fh.write(f"| {f} | {100*per_field_v26[f]/total:.1f}% | {100*per_field_v27[f]/total:.1f}% |\n")
        if diffs:
            fh.write(f"\n## Remaining v27 mismatches ({len(diffs)} rows)\n\n")
            for ticket, desc, row_diffs in diffs:
                fh.write(f"### `{ticket}` — {desc}\n")
                for f, tv, pv in row_diffs:
                    fh.write(f"- **{f}**: truth=`{tv}` v27=`{pv}`\n")
                fh.write("\n")
    print(f"  Score written to: {score_md}")

if __name__ == "__main__":
    main()
