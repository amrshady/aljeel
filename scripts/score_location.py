import openpyxl
import re

wb_t = openpyxl.load_workbook('/home/clawdbot/.openclaw/media/inbound/J26-550_-_validation---620a131c-94f0-416c-a0e0-c612be9e6667.xlsx', data_only=True)
ws_t = wb_t.active
wb_o = openpyxl.load_workbook('/home/clawdbot/.openclaw/workspace/aljeel/dashboard/public/outputs/Spreadsheet-J26-550-FILLED-v30.xlsx', data_only=True)
ws_o = wb_o.active

TICKET_RE = re.compile(r'\d{10}')

t_map = {}
for r in range(2, ws_t.max_row+1):
    val = ws_t.cell(r, 4).value
    if val and TICKET_RE.search(str(val)):
        t_map[TICKET_RE.search(str(val)).group()] = r

o_map = {}
for r in range(4, ws_o.max_row+1):
    val = ws_o.cell(r, 11).value
    if val and TICKET_RE.search(str(val)):
        o_map[TICKET_RE.search(str(val)).group()] = r

shared = set(t_map.keys()) & set(o_map.keys())

loc_matches = 0
loc_mismatches = 0
diffs = []

for tkt in shared:
    o_val = str(ws_o.cell(o_map[tkt], 14).value)
    t_val = str(ws_t.cell(t_map[tkt], 32).value) if ws_t.cell(t_map[tkt], 32).value else ''
    
    o_loc = o_val.split('-')[1] if len(o_val.split('-')) > 1 else 'ERR'
    t_loc = t_val.split('-')[1] if len(t_val.split('-')) > 1 else 'ERR'
    
    if o_loc == t_loc:
        loc_matches += 1
    else:
        loc_mismatches += 1
        diffs.append((tkt, o_loc, t_loc))

print(f"Total Paired Evaluated: {len(shared)}")
print(f"Location Match: {loc_matches}")
print(f"Location Mismatch: {loc_mismatches}")
print(f"Location Match %: {round(loc_matches/len(shared)*100, 2)}%")
print("\nMismatches (Ticket, Pipe_Loc, Truth_Loc):")
for d in diffs:
    print(d)

