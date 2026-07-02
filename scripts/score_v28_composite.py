#!/usr/bin/env python3
"""
Score v28 against composite ground truth:
  - v26 output is truth for ALL rows (all fields)
  - Labadi's file OVERRIDES Account + Employee No for the 12 flagged rows

Usage:
    python3 scripts/score_v28_composite.py
"""
from __future__ import annotations
import re, sys
from pathlib import Path
import openpyxl

ROOT   = Path("/home/clawdbot/.openclaw/workspace/aljeel")
BATCH  = ROOT / "batches/jawal-J26-788"
OUTPUT = BATCH / "output"

V28    = OUTPUT / "Spreadsheet-J26-788-FILLED-v28.xlsx"
V26    = OUTPUT / "Spreadsheet-J26-788-FILLED-v26.xlsx"
LABADI = BATCH  / "Labadi-issues-v28-review.xlsx"

SCORE_FIELDS = ["Account", "Employee No", "Cost Center", "DIV", "Solution", "Agency"]
TICKET_RE    = re.compile(r"(?<!\d)(\d{10}|26-\d{3,4})(?!\d)")


# ── Loaders ──────────────────────────────────────────────────────────────────

def load_oracle(path: Path) -> dict[str, dict]:
    """Load Oracle Fusion template xlsx → dict keyed by ticket number."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    # Row index 2 (0-based) = header row in our template
    header = rows[2]
    hmap = {str(v).strip(): i for i, v in enumerate(header) if v}
    
    result = {}
    for row in rows[3:]:
        if not row[0]:
            continue
        desc_idx = hmap.get("Description", 10)
        desc = str(row[desc_idx] or "")
        for m in TICKET_RE.finditer(desc):
            ticket = m.group(1)
            result[ticket] = {
                "description": desc,
                "Account":     str(row[hmap["Account"]]  or "").strip(),
                "Employee No": str(row[hmap["Employee No"]] or "").strip(),
                "Cost Center": str(row[hmap["Cost Center"]] or "").strip(),
                "DIV":         str(row[hmap["DIV"]]      or "").strip(),
                "Solution":    str(row[hmap["Solution"]] or "").strip(),
                "Agency":      str(row[hmap["Agency"]]   or "").strip(),
            }
            break
    return result


def load_labadi(path: Path) -> dict[str, dict]:
    """
    Load Labadi's corrections file.
    Columns: Description | *Type | *Amount | Employee No | Company | Location | Account | GL Description
    Returns dict keyed by ticket → {Account, Employee No}
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    result = {}
    for row in rows[1:]:  # skip header
        if not row[0]:
            continue
        desc = str(row[0]).strip()
        emp_no = str(row[3]).strip() if row[3] else ""
        account = str(row[6]).strip() if row[6] else ""
        for m in TICKET_RE.finditer(desc):
            ticket = m.group(1)
            result[ticket] = {
                "description": desc,
                "Account":     account,
                "Employee No": emp_no,
            }
            break
    return result


def norm(v: str) -> str:
    """Normalise: strip leading zeros, lowercase."""
    v = str(v).strip()
    # Don't strip leading zeros from emp_no (they matter), but strip from numeric codes
    try:
        return str(int(v))
    except ValueError:
        return v.lower() if v else ""


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("Loading files...")
    v26 = load_oracle(V26)
    v28 = load_oracle(V28)
    labadi = load_labadi(LABADI)

    print(f"  v26: {len(v26)} rows | v28: {len(v28)} rows | Labadi overrides: {len(labadi)} rows")

    # Build composite truth:
    # Start with v26, then apply Labadi overrides (Account + Employee No only)
    composite: dict[str, dict] = {}
    for ticket, v26_row in v26.items():
        ct_row = dict(v26_row)
        if ticket in labadi:
            lab = labadi[ticket]
            if lab.get("Account"):
                ct_row["Account"] = lab["Account"]
            if lab.get("Employee No"):
                ct_row["Employee No"] = lab["Employee No"]
        composite[ticket] = ct_row

    # Rows common to composite truth and v28
    common = sorted(set(composite) & set(v28))
    total = len(common)
    print(f"  Rows to score: {total}\n")

    # Score
    all6_ok = 0
    acc_emp_ok = 0  # just Account + Emp_No (the Labadi-controlled fields)
    per_field = {f: 0 for f in SCORE_FIELDS}
    mismatches = []
    labadi_mismatches = []

    for ticket in sorted(common):
        ct  = composite[ticket]
        p28 = v28[ticket]
        is_labadi = ticket in labadi

        field_match = {f: norm(p28[f]) == norm(ct[f]) for f in SCORE_FIELDS}
        row_ok = all(field_match.values())
        acc_emp = field_match["Account"] and field_match["Employee No"]

        if row_ok:
            all6_ok += 1
        if acc_emp:
            acc_emp_ok += 1
        for f in SCORE_FIELDS:
            if field_match[f]:
                per_field[f] += 1

        if not row_ok:
            diffs = [(f, ct[f], p28[f]) for f in SCORE_FIELDS if not field_match[f]]
            entry = (ticket, ct["description"][:60], diffs, is_labadi)
            mismatches.append(entry)
            if is_labadi:
                labadi_mismatches.append(entry)

    # Print report
    print("=" * 72)
    print(f"  J26-788 v28 SCORE vs COMPOSITE TRUTH  (n={total})")
    print(f"  Composite = v26 baseline + {len(labadi)} Labadi overrides (Account + Emp No)")
    print("=" * 72)
    print(f"  All-6-exact:      {all6_ok}/{total} = {100*all6_ok/total:.1f}%")
    print(f"  Acct+EmpNo exact: {acc_emp_ok}/{total} = {100*acc_emp_ok/total:.1f}%")
    print()
    print(f"  Per-field match rate:")
    print(f"  {'Field':<15} {'Match':>6} {'%':>7}")
    print(f"  {'-'*32}")
    for f in SCORE_FIELDS:
        pct = 100 * per_field[f] / total
        print(f"  {f:<15} {per_field[f]:>5}/{total:<4} {pct:>6.1f}%")
    print()

    # Labadi override rows summary
    print(f"  === Labadi-flagged rows ({len(labadi)}) — v28 performance ===")
    labadi_acc = sum(1 for t in labadi if t in v28 and norm(v28[t]["Account"]) == norm(labadi[t].get("Account","")))
    labadi_emp = sum(1 for t in labadi if t in v28 and norm(v28[t]["Employee No"]) == norm(labadi[t].get("Employee No","")))
    n_lab = sum(1 for t in labadi if t in v28)
    print(f"  Account match:   {labadi_acc}/{n_lab}")
    print(f"  Employee match:  {labadi_emp}/{n_lab}")
    print()

    # Detail mismatches on Labadi rows
    if labadi_mismatches:
        print(f"  Labadi rows with remaining mismatches ({len(labadi_mismatches)}):")
        for ticket, desc, diffs, _ in labadi_mismatches:
            print(f"\n  [{ticket}] {desc}")
            lab_row = labadi.get(ticket, {})
            for f, tv, pv in diffs:
                lab_override = lab_row.get(f, "—")
                source = f"Labadi={lab_override}" if f in ("Account","Employee No") else f"v26={tv}"
                print(f"    {f}: truth({source})  v28={pv!r}  {'❌' if tv != pv else '✅'}")
    print()

    # All mismatches summary (non-Labadi)
    non_labadi_mm = [(t,d,fs,_) for t,d,fs,_ in mismatches if not _]
    if non_labadi_mm:
        print(f"  Non-Labadi rows with mismatches vs v26 baseline ({len(non_labadi_mm)}):")
        for ticket, desc, diffs, _ in non_labadi_mm[:20]:
            diff_str = ", ".join(f"{f}:{tv!r}→{pv!r}" for f,tv,pv in diffs)
            print(f"  [{ticket}] {desc[:50]}  |  {diff_str}")
        if len(non_labadi_mm) > 20:
            print(f"  ... and {len(non_labadi_mm)-20} more")
    print()

    # Write score file
    score_path = OUTPUT / "score-v28-composite.md"
    lines = [
        f"# J26-788 v28 Score vs Composite Truth\n",
        f"\nComposite truth = v26 baseline + {len(labadi)} Labadi overrides (Account + Employee No)\n",
        f"\n## Headline\n",
        f"- Rows scored: {total}\n",
        f"- **All-6-exact: {all6_ok}/{total} = {100*all6_ok/total:.1f}%**\n",
        f"- Account+EmpNo exact: {acc_emp_ok}/{total} = {100*acc_emp_ok/total:.1f}%\n",
        f"\n## Per-field\n\n| Field | Match | % |\n|---|---|---|\n",
    ]
    for f in SCORE_FIELDS:
        lines.append(f"| {f} | {per_field[f]}/{total} | {100*per_field[f]/total:.1f}% |\n")

    lines.append(f"\n## Labadi-flagged rows ({n_lab})\n")
    lines.append(f"- Account match: {labadi_acc}/{n_lab}\n")
    lines.append(f"- Employee No match: {labadi_emp}/{n_lab}\n")

    if labadi_mismatches:
        lines.append(f"\n## Labadi row mismatches ({len(labadi_mismatches)})\n")
        for ticket, desc, diffs, _ in labadi_mismatches:
            lines.append(f"\n### `{ticket}` — {desc}\n")
            lab_row = labadi.get(ticket, {})
            for f, tv, pv in diffs:
                lab_override = lab_row.get(f, "—")
                source = f"Labadi={lab_override}" if f in ("Account","Employee No") else f"v26={tv}"
                lines.append(f"- **{f}**: truth({source}) → v28=`{pv}`\n")

    if non_labadi_mm:
        lines.append(f"\n## Non-Labadi row mismatches vs v26 ({len(non_labadi_mm)} rows)\n\n")
        for ticket, desc, diffs, _ in non_labadi_mm:
            lines.append(f"### `{ticket}` — {desc}\n")
            for f, tv, pv in diffs:
                lines.append(f"- **{f}**: v26=`{tv}` → v28=`{pv}`\n")
            lines.append("\n")

    with open(score_path, "w") as fh:
        fh.writelines(lines)
    print(f"Score written to: {score_path}")


if __name__ == "__main__":
    main()
