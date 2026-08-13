#!/usr/bin/env python3
"""Build a Solventum chargeback workbook from sales lines backed by POD PDFs."""

from __future__ import annotations

import argparse
import re
from io import BytesIO
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook, load_workbook


OUTPUT_COLUMNS = (
    "TRX #",
    "TRX Date",
    "Order Type",
    "Account Name",
    "Ship Address",
    "Item Description",
    "Manufacturer",
    "Agency",
    "Lot Number",
    "Quantity",
    "UOM",
)
TRX_PATTERN = re.compile(r"(?<!\d)(2600\d{6})(?!\d)")
MANUFACTURER_PREFIX = re.compile(r"^(?:3MOC-|3MOR-)")


def collect_pod_trx_numbers(pods: Iterable[str | Path]) -> set[str]:
    """Return every 10-digit, 2600-prefixed TRX token found in PDF filenames."""
    trx_numbers: set[str] = set()
    for pod in pods:
        path = Path(pod)
        if path.suffix.lower() != ".pdf":
            continue
        trx_numbers.update(TRX_PATTERN.findall(path.name))
    return trx_numbers


def expand_pod_arguments(pods: Sequence[str | Path]) -> list[Path]:
    """Expand CLI POD arguments, accepting PDF paths and/or directories."""
    expanded: list[Path] = []
    for raw in pods:
        path = Path(raw)
        if path.is_dir():
            expanded.extend(sorted(p for p in path.iterdir() if p.is_file() and p.suffix.lower() == ".pdf"))
        else:
            expanded.append(path)
    return expanded


def _normalise_trx(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _open_sales_workbook(path: Path):
    # Passing a binary stream lets openpyxl read OOXML workbooks whose upstream
    # filename happens to use .xls. True legacy BIFF .xls files remain outside
    # openpyxl's format support and produce a clear InvalidFileException.
    return load_workbook(BytesIO(path.read_bytes()), read_only=True, data_only=True)


def generate_chargeback(
    sales_path: str | Path,
    pod_paths: Iterable[str | Path],
    output_path: str | Path,
) -> int:
    """Generate the chargeback and return the number of POD-backed sales rows."""
    sales_path = Path(sales_path)
    output_path = Path(output_path)
    pod_trx = collect_pod_trx_numbers(pod_paths)

    workbook = _open_sales_workbook(sales_path)
    try:
        if "Sheet2" not in workbook.sheetnames:
            raise ValueError("Sales workbook must contain a sheet named 'Sheet2'")
        sheet = workbook["Sheet2"]
        header_values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_values:
            raise ValueError("Sales workbook Sheet2 is empty")
        header_map = {str(value).strip(): index for index, value in enumerate(header_values) if value is not None}
        missing = [column for column in OUTPUT_COLUMNS if column not in header_map]
        if missing:
            raise ValueError(f"Sales workbook Sheet2 is missing required columns: {', '.join(missing)}")

        output = Workbook()
        output_sheet = output.active
        output_sheet.title = "Sheet1"
        output_sheet.append(OUTPUT_COLUMNS)

        row_count = 0
        for source_row in sheet.iter_rows(min_row=2, values_only=True):
            trx = _normalise_trx(source_row[header_map["TRX #"]])
            if trx not in pod_trx:
                continue
            values = [source_row[header_map[column]] for column in OUTPUT_COLUMNS]
            manufacturer_index = OUTPUT_COLUMNS.index("Manufacturer")
            manufacturer = values[manufacturer_index]
            if isinstance(manufacturer, str):
                values[manufacturer_index] = MANUFACTURER_PREFIX.sub("", manufacturer)
            output_sheet.append(values)
            row_count += 1

        output_path.parent.mkdir(parents=True, exist_ok=True)
        output.save(output_path)
        output.close()
        return row_count
    finally:
        workbook.close()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--sales", required=True, type=Path, help="JUNE SALES workbook")
    parser.add_argument(
        "--pods",
        required=True,
        nargs="+",
        type=Path,
        help="One or more POD PDFs and/or directories containing POD PDFs",
    )
    parser.add_argument("--out", required=True, type=Path, help="Output .xlsx path")
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    pods = expand_pod_arguments(args.pods)
    row_count = generate_chargeback(args.sales, pods, args.out)
    print(f"Generated {row_count} POD-backed chargeback rows: {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
