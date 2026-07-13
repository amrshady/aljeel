#!/usr/bin/env python3
"""Stage 5: Multi-employee row splitter.

Rows whose Employee No column holds multiple comma-separated employee numbers
(e.g. "1000477,1002037" on Euro Anesthesia rows) are split into one row per
employee in a NEW output file. Normal travel rows retain the legacy equal split.
Sponsorship rows use the exact employee/amount tuples persisted from the OPEX
allocation table and preserve the parent event segments. The original v30 file
is left untouched.

Usage:
    python3 scripts/split_multi_emp.py <input_xlsx> <output_xlsx> [lookups_xlsx]
"""

import sys
import json
from copy import copy
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
import re

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from code_name_lookup import get_lookup  # noqa: E402
from run_v16 import _norm_master_code  # noqa: E402

DEFAULT_LOOKUPS = "/home/clawdbot/.openclaw/workspace/aljeel/qc/master-data/Aljeel_Lookups-v2.xlsx"

# Combo order: Company-Location-Account-CostCenter-DIV-Solution-Agency-Project-IC-Future1
COMBO_SEGMENTS = ["Company", "Location", "Account", "Cost Center", "DIV",
                  "Solution", "Agency", "Project", "Intercompany", "Future 1"]
NUMERIC_SEGMENT_WIDTHS = {
    "Cost Center": 6,
    "DIV": 3,
    "Solution": 5,
    "Agency": 5,
}

# Header → accepted aliases (v30 sheet uses "Cost Name" for the CC name and
# "Contribution" for the DIV name).
HEADER_ALIASES = {
    "Employee No": ["Employee No"],
    "*Amount": ["*Amount"],
    "*Invoice Amount": ["*Invoice Amount"],
    "Distribution Combination": ["Distribution Combination"],  # prefix match
    "Company": ["Company"],
    "Location": ["Location"],
    "Account": ["Account"],
    "Cost Center": ["Cost Center"],
    "Cost Center Name": ["Cost Name", "Cost Center Name"],
    "DIV": ["DIV"],
    "DIV Name": ["Contribution", "DIV Name"],
    "Solution": ["Solution"],
    "Solution Name": ["Solution Name"],
    "Agency": ["Agency"],
    "Agency Name": ["Agency Name"],
    "Project": ["Project"],
    "Intercompany": ["Intercompany"],
    "Future 1": ["Future 1"],
    "GL Description": ["GL Description"],
    "Agent Flags": ["Agent Flags"],
    "OPEX Allocation Details": ["OPEX Allocation Details"],
}

EVENT_SEGMENTS_TO_PRESERVE = (
    "Account",
    "Cost Center",
    "Cost Center Name",
    "DIV",
    "DIV Name",
    "Solution",
    "Solution Name",
    "Agency",
    "Agency Name",
)


# ---------- master data loaders (same pattern as full_evidence_agent_v30) ----------
def load_manpower(lookups_path: str) -> dict:
    """Returns {emp_no: {location, div_code, div_name, agency_code, agency_name, cost_center, cc_name, solution}}."""
    wb = openpyxl.load_workbook(lookups_path, read_only=True, data_only=True)
    ws = wb["Manpower"]
    rows = list(ws.iter_rows(values_only=True))
    # Columns: Emp No(0), Old Emp No(1), Name(2), Arabic Name(3), Location(4), Manager No(5), Line Manager(6),
    #  blank(7), Code(8)=DIV code, New Division(9), Code(10)=Agency, New agency(11), New cost center(12),
    #  New cost center name(13), blank(14), Solution(15)
    emps = {}
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        emp_no = str(row[0]).strip()
        if not emp_no or emp_no.lower() == "none":
            continue
        emps[emp_no] = {
            "emp_no": emp_no,
            "name": str(row[2] or "").strip(),
            "location": str(row[4] or "").strip(),
            "div_code": str(row[8] or "").strip(),
            "div_name": str(row[9] or "").strip(),
            "agency_code": str(row[10] or "").strip(),
            "agency_name": str(row[11] or "").strip(),
            "cost_center": str(row[12] or "").strip(),
            "cc_name": str(row[13] or "").strip(),
            "solution": str(row[15] or "").strip() if len(row) > 15 else "",
        }
    wb.close()
    return emps


def load_solution_names(lookups_path: str) -> dict:
    """Returns {solution_code: name} from the Solution tab."""
    wb = openpyxl.load_workbook(lookups_path, read_only=True, data_only=True)
    ws = wb["Solution"]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        out[str(row[0]).strip()] = str(row[1] or "").strip()
    wb.close()
    return out


# ---------- helpers ----------
def find_header_row(ws) -> int:
    for r in range(1, 7):
        for cell in ws[r]:
            if cell.value is not None and str(cell.value).strip() == "Employee No":
                return r
    raise ValueError("Header row with 'Employee No' not found in first 6 rows")


def build_col_map(ws, header_row: int) -> dict:
    headers = {}
    for cell in ws[header_row]:
        if cell.value is not None:
            headers[cell.column] = str(cell.value).strip()
    col_map = {}
    for key, aliases in HEADER_ALIASES.items():
        for col, hdr in headers.items():
            if hdr in aliases or any(hdr.startswith(a) for a in aliases):
                col_map[key] = col
                break
    missing = [k for k in ("Employee No", "*Amount", "Distribution Combination") if k not in col_map]
    if missing:
        raise ValueError(f"Required columns not found: {missing}")
    return col_map


def parse_amount(value):
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def split_amount(total: float, n: int) -> list:
    """Equal split rounded to 2dp; rounding remainder goes to the last row."""
    share = round(total / n, 2)
    shares = [share] * (n - 1)
    shares.append(round(total - share * (n - 1), 2))
    return shares


def allocate_line_amount(total: float, form_amounts: list[float | None]) -> list[float]:
    """Allocate a billed line across form employees without changing its value.

    Form amounts are weights only.  All-present amounts produce a proportional
    split; all-missing amounts produce an even split.  The first child absorbs
    the cent-rounding remainder so the returned children exactly conserve the
    parent line amount.
    """
    if not form_amounts:
        return []
    try:
        line_amount = Decimal(str(total))
    except (InvalidOperation, ValueError, TypeError):
        return []

    explicit = [amount is not None for amount in form_amounts]
    if any(explicit) and not all(explicit):
        return []
    if all(explicit):
        weights = [Decimal(str(amount)) for amount in form_amounts]
        weight_total = sum(weights, Decimal("0"))
        if weight_total == 0:
            return []
        shares = [
            (line_amount * weight / weight_total).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )
            for weight in weights
        ]
    else:
        # Retain the established equal-split calculation, then move its
        # rounding remainder from the last child to the first for sponsorship.
        shares = [Decimal(str(value)) for value in split_amount(float(line_amount), len(form_amounts))]
        if len(shares) > 1:
            remainder = shares[-1] - shares[0]
            shares[0] += remainder
            shares[-1] -= remainder

    shares[0] += line_amount - sum(shares, Decimal("0"))
    return [float(share) for share in shares]


def parse_opex_allocations(value) -> list[dict]:
    """Validate employee tuples and optional ratio amounts serialized by v30."""
    if not value:
        return []
    try:
        raw = json.loads(str(value))
    except (TypeError, ValueError, json.JSONDecodeError):
        return []
    if not isinstance(raw, list):
        return []
    out = []
    for item in raw:
        if not isinstance(item, dict):
            return []
        emp_no = seg(item.get("emp_no"))
        raw_amount = item.get("amount")
        amount = parse_amount(raw_amount)
        if not re.fullmatch(r"1\d{6}", emp_no):
            return []
        if raw_amount not in (None, "") and amount is None:
            return []
        out.append({"emp_no": emp_no, "name": seg(item.get("name")), "amount": amount})
    explicit = [item["amount"] is not None for item in out]
    if any(explicit) and not all(explicit):
        return []
    return out


def _cell_text(ws, row_i: int, col: dict, header: str) -> str:
    if header not in col:
        return ""
    return seg(ws.cell(row=row_i, column=col[header]).value)


def _is_event_sponsorship_row(ws, row_i: int, col: dict) -> bool:
    account = _cell_text(ws, row_i, col, "Account")
    serial = _cell_text(ws, row_i, col, "OPEX Serial").upper()
    form_values = [
        _cell_text(ws, row_i, col, header)
        for header in (
            "Form Division (Fusion code)",
            "Form Agency (Fusion code)",
            "Form Solution (Fusion code)",
            "Form Cost-Center-Ref (Fusion 15-digit)",
        )
    ]
    return account == "60307021" or (serial and serial not in {"MISSING", "N/A", "NONE"}) or any(form_values)


def copy_row(ws, src_row: int, dst_row: int, max_col: int):
    for c in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=c)
        dst = ws.cell(row=dst_row, column=c, value=src.value)
        if src.has_style:
            dst._style = copy(src._style)
    if src_row in ws.row_dimensions:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def seg(value) -> str:
    return str(value).strip() if value is not None else ""


def normalize_output_segment(value, width: int) -> str:
    text = seg(value)
    if not text or text.lower() in {"none", "nan", "null"}:
        return "0" * width
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return text.zfill(width) if text.isdigit() and len(text) <= width else text


def normalize_row_segments(ws, row_i: int, col: dict) -> dict:
    normalized = {}
    for header, width in NUMERIC_SEGMENT_WIDTHS.items():
        if header not in col:
            continue
        value = normalize_output_segment(
            ws.cell(row=row_i, column=col[header]).value,
            width,
        )
        ws.cell(row=row_i, column=col[header], value=value)
        normalized[header] = value
    return normalized


def build_gl_desc(combo: str, lookup) -> str:
    """Concatenated segment names for the GL Description column.

    Same expand_combo + " · " join as run_v30's sync_final_gl_descriptions,
    but a segment with no name match becomes "" instead of crashing.
    """
    expansion = lookup.expand_combo(combo)

    def name(key) -> str:
        v = expansion.get(key)
        s = str(v).strip() if v is not None else ""
        return "" if s == "#N/A" else s

    parts = [name("GL"), name("Cost Name"), name("Contribution"),
             name("Solution Name"), name("Agency Name"),
             "00000", "00", "000000"]
    return " · ".join(parts)


# ---------- main ----------
def split_multi_emp(input_xlsx: str, output_xlsx: str, lookups_path: str = DEFAULT_LOOKUPS) -> dict:
    """Split multi-employee rows into one row per employee in a new file.

    Returns: {"total_rows": N, "split_rows": M, "output_rows": K}
    """
    manpower = load_manpower(lookups_path)
    solution_names = load_solution_names(lookups_path)
    lookup = get_lookup(lookups_path)

    wb = openpyxl.load_workbook(input_xlsx)
    ws = wb.active
    header_row = find_header_row(ws)
    col = build_col_map(ws, header_row)
    max_col = ws.max_column

    emp_col = col["Employee No"]
    total_rows = 0
    split_rows = 0

    # Bottom-up so row insertion never shifts rows we have yet to visit.
    for r in range(ws.max_row, header_row, -1):
        row_has_data = any(ws.cell(row=r, column=c).value not in (None, "") for c in range(1, max_col + 1))
        if not row_has_data:
            continue
        total_rows += 1
        emp_val = ws.cell(row=r, column=emp_col).value
        sponsorship = _cell_text(ws, r, col, "Account") == "60307021"
        allocations = parse_opex_allocations(
            ws.cell(row=r, column=col["OPEX Allocation Details"]).value
            if sponsorship and "OPEX Allocation Details" in col else None
        )
        if sponsorship and allocations:
            stamped_emps = [e.strip() for e in str(emp_val or "").split(",") if e.strip()]
            allocation_emps = [item["emp_no"] for item in allocations]
            if stamped_emps != allocation_emps:
                # Never resurrect stale tuples that a later folder-review pass
                # invalidated by blanking Employee No.
                allocations = []
        if sponsorship:
            if not allocations:
                ws.cell(row=r, column=emp_col, value="")
                if "Agent Flags" in col:
                    flag_cell = ws.cell(row=r, column=col["Agent Flags"])
                    flags = seg(flag_cell.value)
                    review = "SPONSORSHIP_ALLOCATION_TABLE_REVIEW"
                    if review not in flags:
                        flag_cell.value = (flags + " | " + review).strip(" |")
                continue
            emps = [item["emp_no"] for item in allocations]
            line_amount = parse_amount(ws.cell(row=r, column=col["*Amount"]).value)
            amount_shares = (
                allocate_line_amount(line_amount, [item["amount"] for item in allocations])
                if line_amount is not None else None
            )
        else:
            if emp_val is None or "," not in str(emp_val):
                continue
            emps = [e.strip() for e in str(emp_val).split(",") if e.strip()]
            if len(emps) < 2:
                continue
            amount = parse_amount(ws.cell(row=r, column=col["*Amount"]).value)
            amount_shares = split_amount(amount, len(emps)) if amount is not None else None
        n = len(emps)
        if n > 1:
            split_rows += 1

        # *Invoice Amount is the header-level invoice total — do NOT divide, copy as-is to all split rows
        inv_amount = parse_amount(ws.cell(row=r, column=col["*Invoice Amount"]).value) if "*Invoice Amount" in col else None
        # inv_shares removed — *Invoice Amount is copied as-is (not divided)
        preserve_event_segments = _is_event_sponsorship_row(ws, r, col)
        parent_event_segments = {
            key: ws.cell(row=r, column=col[key]).value
            for key in EVENT_SEGMENTS_TO_PRESERVE
            if preserve_event_segments and key in col
        }

        # openpyxl's insert_rows(..., amount=0) corrupts the used range, so a
        # valid one-employee form allocation updates the row in place.
        if n > 1:
            ws.insert_rows(r + 1, n - 1)
            for k in range(1, n):
                copy_row(ws, r, r + k, max_col)

        for k, emp in enumerate(emps):
            row_i = r + k
            ws.cell(row=row_i, column=emp_col, value=emp)
            if amount_shares is not None:
                ws.cell(row=row_i, column=col["*Amount"], value=amount_shares[k])
            if inv_amount is not None and "*Invoice Amount" in col:
                ws.cell(row=row_i, column=col["*Invoice Amount"], value=inv_amount)

            if preserve_event_segments:
                for key, value in parent_event_segments.items():
                    ws.cell(row=row_i, column=col[key], value=value)

            rec = manpower.get(emp)
            if rec is None:
                # Unknown employee: keep original segments, flag the row.
                if "Agent Flags" in col:
                    flags_cell = ws.cell(row=row_i, column=col["Agent Flags"])
                    existing = seg(flags_cell.value)
                    flags_cell.value = (existing + " | EMP_NOT_IN_MANPOWER") if existing else "EMP_NOT_IN_MANPOWER"
                continue

            # Normal travel splits re-derive personal segments from Manpower;
            # event sponsorship splits keep the v30 parent row's event segments.
            def set_col(key, value):
                if key in col and value:
                    ws.cell(row=row_i, column=col[key], value=value)

            if not preserve_event_segments:
                solution = _norm_master_code(rec["solution"], "solution") or "00000"
                if not solution.isdigit():
                    original_solution = (seg(ws.cell(row=row_i, column=col["Solution"]).value) if "Solution" in col else "")
                    solution = original_solution if original_solution.isdigit() else "00000"
                solution = solution.zfill(5)
                set_col("Location", rec["location"])
                set_col("Cost Center", normalize_output_segment(rec["cost_center"], 6))
                set_col("Cost Center Name", rec["cc_name"])
                set_col("DIV", normalize_output_segment(rec["div_code"], 3))
                set_col("DIV Name", rec["div_name"])
                set_col("Solution", solution)
                set_col("Solution Name", solution_names.get(solution, ""))
                set_col("Agency", normalize_output_segment(rec["agency_code"], 5))
                set_col("Agency Name", rec["agency_name"])

            normalize_row_segments(ws, row_i, col)
            combo = "-".join(seg(ws.cell(row=row_i, column=col[name]).value) if name in col else ""
                             for name in COMBO_SEGMENTS)
            ws.cell(row=row_i, column=col["Distribution Combination"], value=combo)
            if "GL Description" in col:
                ws.cell(row=row_i, column=col["GL Description"], value=build_gl_desc(combo, lookup))

    output_rows = 0
    for r in range(header_row + 1, ws.max_row + 1):
        if any(ws.cell(row=r, column=c).value not in (None, "") for c in range(1, max_col + 1)):
            output_rows += 1

    Path(output_xlsx).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)
    wb.close()
    return {"total_rows": total_rows, "split_rows": split_rows, "output_rows": output_rows}


def main():
    if len(sys.argv) < 3:
        print("Usage: python3 split_multi_emp.py <input_xlsx> <output_xlsx> [lookups_xlsx]")
        sys.exit(1)
    input_xlsx, output_xlsx = sys.argv[1], sys.argv[2]
    lookups = sys.argv[3] if len(sys.argv) > 3 else DEFAULT_LOOKUPS
    if not Path(input_xlsx).exists():
        print(f"❌ Input file not found: {input_xlsx}")
        sys.exit(1)
    result = split_multi_emp(input_xlsx, output_xlsx, lookups)
    print(f"✅ Split complete: {result['total_rows']} rows in, "
          f"{result['split_rows']} split, {result['output_rows']} rows out → {output_xlsx}")


if __name__ == "__main__":
    main()
