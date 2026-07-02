import openpyxl

wb_o = openpyxl.load_workbook('/home/clawdbot/.openclaw/workspace/aljeel/batches/jawal-J26-550/output/Spreadsheet-J26-550-FILLED-v30.xlsx', data_only=True)
ws_o = wb_o.active

def rewrite_locations():
    # Load manpower
    wb_m = openpyxl.load_workbook('/home/clawdbot/.openclaw/workspace/aljeel/qc/master-data/Aljeel_Lookups-v2.xlsx', data_only=True)
    ws_m = wb_m['Manpower']
    loc_map = {}
    for r in range(2, ws_m.max_row+1):
        emp_raw = ws_m.cell(r, 1).value
        loc_raw = ws_m.cell(r, 5).value
        if emp_raw:
            try:
                emp = str(int(emp_raw))
                loc = str(int(loc_raw)) if loc_raw else ""
                loc_map[emp] = loc
            except:
                pass
                
    for r in range(4, ws_o.max_row+1):
        combo = str(ws_o.cell(r, 14).value)
        emp_no = str(ws_o.cell(r, 16).value) if ws_o.cell(r, 16).value else ""
        if combo and len(combo.split('-')) == 10:
            parts = combo.split('-')
            new_loc = "20100"
            if emp_no in loc_map:
                l = loc_map[emp_no]
                if l == "10100":
                    new_loc = "20100"
                elif l in ("30100", "40100"):
                    new_loc = l
                else:
                    new_loc = parts[1]
            else:
                new_loc = parts[1]
                
            if new_loc == "10100":
                new_loc = "20100"
                
            parts[1] = new_loc
            ws_o.cell(r, 14).value = "-".join(parts)
            
    wb_o.save('/home/clawdbot/.openclaw/workspace/aljeel/batches/jawal-J26-550/output/Spreadsheet-J26-550-FILLED-v30.xlsx')
    
rewrite_locations()
print("Locations rewritten")
