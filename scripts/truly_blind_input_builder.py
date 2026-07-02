#!/usr/bin/env python3
"""
truly_blind_input_builder.py — Build Spreadsheet.xlsx ONLY from the Jawal supplier
invoice xlsx. No leak from Aljeel's prepared workbook.

Inputs (truly blind):
  - Jawal invoice xlsx  (supplier side: Sl#, Issue Date, Ticket No., Pax, Route,
    Service Date, Taxable Amt, VAT%, Inv. Amt. Incl. VAT)
  - (Optional) Credit-note xlsx for refund rows -> Debit Memo invoice type

Banned (intentionally NOT touched):
  - Aljeel's prepared J26-XXX.xlsx (Details sheet has emp_no, GL, CC, DIV, etc.)

Output:
  Spreadsheet.xlsx with the standard Oracle Fusion AP template:
    col A  *Invoice Header Identifier       = 1, 2, 3, ...
    col B  *Business Unit                   = "Al Jeel Medical BU"
    col C  *Invoice Number                  = <BATCH_ID>
    col D  *Invoice Currency                = SAR
    col E  *Invoice Amount                  = Jawal Inv. Amt. Incl. VAT
    col F  *Invoice Date                    = Jawal Issue Date (yyyy-mm-dd)
    col G  **Supplier (Arabic)              = "شركة جوال للسفر والسياحة المحدودة"
    col H  **Supplier Number                = 10394   (Aljeel-side Jawal supplier code)
    col I  *Supplier Site (Arabic)          = "شركة جوال للسفر"
    col J  Invoice Type                     = BLANK (pipeline fills Standard/Debit Memo)
    col K  Description                      = "PAX - ROUTE (TICKET)[/Notes if any]"
    col L  *Type                            = "Item"
    col M  *Amount (line)                   = same as col E for single-line entries
    col N  Distribution Combination         = BLANK  (no leak)
    col O  Tax Classification Code          = derived from Jawal VAT%
                                              15 -> "KSA VAT STANDARD"
                                              0  -> "KSA VAT ZERO"
    col P  Employee No                      = BLANK  (no leak)

Usage:
  python3 truly_blind_input_builder.py \
      --jawal-xlsx <path/to/AL JEEL ... INV.xlsx> \
      --batch-id  J26-640 \
      --out       <out_path>/Spreadsheet.xlsx

Notes:
  - Skips Jawal rows with Inv.Amt. == 0 (consistent with how Aljeel excluded
    ticket 6905397691 / Sl 87 of J26-640: a "sponse" line zeroed out).
  - For duplicate voucher tickets (e.g. 26-689 appears 3x as 3 OPEX lines),
    each Jawal row becomes a separate Spreadsheet row.
  - Notes from Jawal Ref column are appended to Description when they look like
    real notes (e.g. "EP-2026-14", "SIS-07-2026", "change", "sponsor").
"""
from __future__ import annotations

import argparse
import re
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Constants pulled from Aljeel/Jawal canonical setup
SUPPLIER_NAME_AR = "شركة جوال للسفر والسياحة المحدودة"
SUPPLIER_NUMBER  = "10394"
SUPPLIER_SITE_AR = "شركة جوال للسفر"
BUSINESS_UNIT    = "Al Jeel Medical BU"
CURRENCY         = "SAR"


# ── Jawal invoice parsing ────────────────────────────────────────────────────

JAWAL_COL = {
    'sl':       1,
    'issue':    2,
    'ref':      7,
    'ticket':  12,
    'pax':     16,
    'route':   22,
    'service': 30,
    'taxable': 34,
    'vat_pct': 39,
    'vat_amt': 42,
    'inv_amt': 47,
}


def _norm_date(v) -> str:
    """Coerce date-ish value -> 'YYYY-MM-DD' string."""
    if v is None or v == '':
        return ''
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    # Already 'YYYY-MM-DD'?
    m = re.match(r'^(\d{4}-\d{2}-\d{2})', s)
    if m:
        return m.group(1)
    # Try parse
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return s  # give up, pass through


def _clean(v) -> str:
    if v is None:
        return ''
    return str(v).strip()


def _ticket_short(raw: str) -> str:
    """Extract 10-digit airline ticket or 26-XXX voucher from Jawal Ticket column.
    Jawal format examples:
      '065 6905264364'  -> '6905264364'
      'sponsor payment form' + ticket col '390 6905264365' -> '6905264365'
      '26-706' -> '26-706'
    """
    s = _clean(raw)
    m = re.search(r'(?<!\d)(\d{10})(?!\d)', s)
    if m:
        return m.group(1)
    m = re.search(r'\b(26-\d{3})\b', s)
    if m:
        return m.group(1)
    # Fall back to last whitespace-separated token if it has digits
    parts = s.split()
    for p in reversed(parts):
        if any(c.isdigit() for c in p):
            return p
    return s


def _build_description(pax: str, route: str, ticket_short: str, ref: str) -> str:
    """Compose description in pipeline-friendly form:
       'PAX - ROUTE (TICKET)' optionally with notes appended.
    Mirrors what 593-BLIND Spreadsheet.xlsx looked like.
    """
    desc = f"{pax} - {route} ({ticket_short})"

    # Heuristic: append ref text only if it looks like a real note
    # (not a plain numeric reference like "1002576")
    ref_clean = _clean(ref)
    if ref_clean and not re.fullmatch(r'\d+\.?', ref_clean):
        # Examples to keep:
        #   "Re: OPEX / Prague Rhythm 2026 / EP-2026-14"
        #   "sponsor payment form"
        #   "change 1002405"
        # Examples to drop (pure numeric refs):
        #   "1002576", "1002576."
        # Append after a slash
        desc = f"{desc} {ref_clean}".strip()
    return desc


def _tax_code(vat_pct, inv_amt) -> str:
    """Derive Oracle tax classification from Jawal VAT% column.
       15% -> KSA VAT STANDARD
       0%  -> KSA VAT ZERO
    Robust to floats, ints, and string representations.
    """
    try:
        pct = float(vat_pct)
    except (TypeError, ValueError):
        return ''
    if pct >= 1:  # 15% or higher
        return 'KSA VAT STANDARD'
    return 'KSA VAT ZERO'


def parse_jawal_invoice(path: Path) -> list[dict]:
    """Parse a Jawal supplier invoice xlsx into structured line items.

    Returns list of dicts (one per non-zero data row), keyed by:
      sl, issue, ref, ticket_raw, ticket_short, pax, route, service,
      taxable, vat_pct, vat_amt, inv_amt
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Find the header row of the line-items table (look for 'Sl. #' / 'Ticket No.')
    hdr_row = None
    for r in range(1, min(ws.max_row + 1, 50)):
        a = _clean(ws.cell(r, JAWAL_COL['sl']).value)
        l = _clean(ws.cell(r, JAWAL_COL['ticket']).value)
        if 'Sl' in a and 'Ticket' in l:
            hdr_row = r
            break
    if hdr_row is None:
        raise ValueError(f"Could not find line-item header row in {path}")

    rows = []
    for r in range(hdr_row + 1, ws.max_row + 1):
        sl  = ws.cell(r, JAWAL_COL['sl']).value
        # Stop at first blank Sl# row that is NOT followed by data
        if sl is None or sl == '':
            # Continue scanning a couple more rows; if all blank, stop
            continue
        if not isinstance(sl, (int, float)):
            # Likely "Prepared By" / totals row — stop
            break

        inv_amt = ws.cell(r, JAWAL_COL['inv_amt']).value
        if inv_amt is None:
            continue
        try:
            inv_amt_f = float(inv_amt)
        except (TypeError, ValueError):
            continue

        # Truly-blind builder MUST skip zero-amount rows.
        # (Aljeel does too: Sl 87 of J26-640 with Inv=0 is not in Details.)
        if inv_amt_f == 0:
            continue

        rows.append({
            'sl':           int(sl),
            'issue':        _norm_date(ws.cell(r, JAWAL_COL['issue']).value),
            'ref':          _clean(ws.cell(r, JAWAL_COL['ref']).value),
            'ticket_raw':   _clean(ws.cell(r, JAWAL_COL['ticket']).value),
            'ticket_short': _ticket_short(ws.cell(r, JAWAL_COL['ticket']).value),
            'pax':          _clean(ws.cell(r, JAWAL_COL['pax']).value),
            'route':        _clean(ws.cell(r, JAWAL_COL['route']).value),
            'service':      _norm_date(ws.cell(r, JAWAL_COL['service']).value),
            'taxable':      ws.cell(r, JAWAL_COL['taxable']).value,
            'vat_pct':      ws.cell(r, JAWAL_COL['vat_pct']).value,
            'vat_amt':      ws.cell(r, JAWAL_COL['vat_amt']).value,
            'inv_amt':      inv_amt_f,
        })
    return rows


# ── Spreadsheet.xlsx writing ─────────────────────────────────────────────────

HEADERS = [
    '*Invoice Header Identifier',  # A
    '*Business Unit',              # B
    '*Invoice Number',             # C
    '*Invoice Currency',           # D
    '*Invoice Amount',             # E
    '*Invoice Date',               # F
    '**Supplier[..]',              # G
    '**Supplier Number',           # H
    '*Supplier Site[..]',          # I
    'Invoice Type',                # J
    'Description',                 # K
    '*Type',                       # L
    '*Amount',                     # M
    'Distribution Combination[..]',# N
    'Tax Classification Code[..]', # O
    'Employee No',                 # P
]


def write_spreadsheet(rows: list[dict], batch_id: str, invoice_date: str, out_path: Path) -> None:
    """Write the Oracle Fusion AP-style Spreadsheet.xlsx.
       Headers live in row 3 (per pipeline convention).
       Data starts at row 4.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # Header row 3
    for ci, h in enumerate(HEADERS, start=1):
        ws.cell(3, ci).value = h
        ws.cell(3, ci).font = Font(bold=True)

    # Data rows starting r4
    for i, row in enumerate(rows, start=1):
        r_excel = 3 + i

        ticket_short = row['ticket_short']
        desc = _build_description(row['pax'], row['route'], ticket_short, row['ref'])
        tax = _tax_code(row['vat_pct'], row['inv_amt'])

        ws.cell(r_excel,  1).value = i
        ws.cell(r_excel,  2).value = BUSINESS_UNIT
        ws.cell(r_excel,  3).value = batch_id
        ws.cell(r_excel,  4).value = CURRENCY
        ws.cell(r_excel,  5).value = row['inv_amt']
        ws.cell(r_excel,  6).value = invoice_date or row['issue']
        ws.cell(r_excel,  7).value = SUPPLIER_NAME_AR
        ws.cell(r_excel,  8).value = SUPPLIER_NUMBER
        ws.cell(r_excel,  9).value = SUPPLIER_SITE_AR
        ws.cell(r_excel, 10).value = None                    # Invoice Type — BLANK
        ws.cell(r_excel, 11).value = desc
        ws.cell(r_excel, 12).value = 'Item'
        ws.cell(r_excel, 13).value = row['inv_amt']
        ws.cell(r_excel, 14).value = None                    # Distribution Combination — BLANK (no leak)
        ws.cell(r_excel, 15).value = tax
        ws.cell(r_excel, 16).value = None                    # Employee No — BLANK (no leak)

    # Column widths (cosmetic)
    for ci, w in enumerate([8, 18, 12, 8, 12, 12, 35, 14, 25, 14, 60, 8, 12, 26, 22, 12], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--jawal-xlsx', required=True, help='Jawal supplier invoice xlsx')
    ap.add_argument('--batch-id', required=True, help='e.g. J26-640')
    ap.add_argument('--invoice-date', default='', help='Invoice header date (YYYY-MM-DD). If empty, uses first row issue date')
    ap.add_argument('--out', required=True, help='Output Spreadsheet.xlsx path')
    args = ap.parse_args()

    jawal_path = Path(args.jawal_xlsx)
    out_path = Path(args.out)
    if not jawal_path.exists():
        print(f'ERROR: Jawal invoice not found: {jawal_path}', file=sys.stderr)
        sys.exit(2)

    # Safety: refuse if the Jawal path looks like an Aljeel-prepared workbook
    # (their files are named J26-XXX.xlsx with Details sheet — Jawal files
    # are named "AL JEEL ... INV.xlsx" or "INV-... AL JEEL.xlsx").
    fname = jawal_path.name.upper()
    if re.match(r'^J26-\d+\.', fname):
        print(f'ERROR: input looks like an Aljeel-prepared workbook (J26-XXX.xlsx): {jawal_path}', file=sys.stderr)
        print('       refusing to use — pass the Jawal supplier invoice instead.', file=sys.stderr)
        sys.exit(3)

    rows = parse_jawal_invoice(jawal_path)
    if not rows:
        print('ERROR: no non-zero data rows parsed from Jawal invoice', file=sys.stderr)
        sys.exit(4)

    invoice_date = args.invoice_date or rows[0]['issue']
    write_spreadsheet(rows, args.batch_id, invoice_date, out_path)

    print(f'OK: wrote {out_path}')
    print(f'    {len(rows)} line items (Jawal supplier invoice only — no Aljeel leak)')
    print(f'    Invoice date: {invoice_date}')
    print(f'    Invoice total: {sum(r["inv_amt"] for r in rows):.2f} SAR')


if __name__ == '__main__':
    main()
