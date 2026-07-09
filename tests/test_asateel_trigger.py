from pathlib import Path

import openpyxl
from flask import Flask

from scripts import asateel_trigger


def _client():
    app = Flask(__name__)
    app.register_blueprint(asateel_trigger.bp)
    return app.test_client()


def test_auth_rejects_missing_trigger_key(monkeypatch):
    monkeypatch.setenv("ASATEEL_TRIGGER_KEY", "secret")
    response = _client().post("/asateel/run", json={})
    assert response.status_code == 401
    assert response.get_json() == {"error": "unauthorized"}


def test_bad_region_validation(monkeypatch):
    monkeypatch.setenv("ASATEEL_TRIGGER_KEY", "secret")
    response = _client().post(
        "/asateel/run",
        headers={"X-Asateel-Trigger-Key": "secret"},
        json={
            "archive_date": "2026-07-02",
            "folder_name": "مشاريع 16",
            "region": "WEST",
            "batch_id": "projects-16",
        },
    )
    assert response.status_code == 400
    assert response.get_json()["error"] == "region must be one of CENTRAL, PROJECTS, ADMIN"


def test_region_title_and_number_derivation():
    assert asateel_trigger.region_title_and_number("CENTRAL", "central-3") == ("Central", "03")
    assert asateel_trigger.region_title_and_number("PROJECTS", "projects-16") == ("Projects", "16")
    assert asateel_trigger.region_title_and_number("ADMIN", "admin") == ("Admin", "00")


def test_missing_jq_extraction_builds_distinct_list(tmp_path: Path):
    oracle = tmp_path / "oracle.xlsx"
    missing = tmp_path / "missing.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([])
    ws.append([])
    ws.append([
        "Row Status",
        "SO_Detail Agency",
        "Cost Center",
        "Additional Information",
        "Employee No",
        "*Invoice Number",
        "*Amount",
    ])
    ws.append(["RED", "", "200010", "12345.JQ-00000001", "12345", "03912", 100.5])
    ws.append(["RED", "", "200010", "12345.JQ-00000001", "12345", "03912", 100.5])
    ws.append(["RED", "", "140040", "22222.JQ-WAREHOUSE", "22222", "03913", 25])
    ws.append(["RED", "50123", "200010", "33333.JQ-HASAGENCY", "33333", "03914", 40])
    ws.append(["YELLOW", "", "200010", "44444.JQ-YELLOW", "44444", "03915", 60])
    wb.save(oracle)

    count = asateel_trigger.build_missing_jq_sheet(oracle, missing)

    assert count == 1
    out_wb = openpyxl.load_workbook(missing, data_only=True)
    try:
        rows = list(out_wb.active.iter_rows(values_only=True))
    finally:
        out_wb.close()
    assert rows == [
        ("JQ", "Employee No", "Invoice Number", "Amount SAR"),
        ("JQ-00000001", "12345", "03912", "100.50"),
    ]


def test_missing_jq_sheet_has_no_missing_row(tmp_path: Path):
    oracle = tmp_path / "oracle.xlsx"
    missing = tmp_path / "missing.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([])
    ws.append([])
    ws.append([
        "Row Status",
        "SO_Detail Agency",
        "Cost Center",
        "Additional Information",
        "Employee No",
        "*Invoice Number",
        "*Amount",
    ])
    ws.append(["GREEN", "", "200010", "12345.JQ-00000001", "12345", "03912", 100.5])
    wb.save(oracle)

    count = asateel_trigger.build_missing_jq_sheet(oracle, missing)

    assert count == 0
    out_wb = openpyxl.load_workbook(missing, data_only=True)
    try:
        rows = list(out_wb.active.iter_rows(values_only=True))
    finally:
        out_wb.close()
    assert rows == [
        ("JQ", "Employee No", "Invoice Number", "Amount SAR"),
        ("No missing JQs", None, None, None),
    ]
