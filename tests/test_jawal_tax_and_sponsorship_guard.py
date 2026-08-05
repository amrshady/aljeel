from pathlib import Path
import sys

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from convert_jawal_invoice import derive_tax_classification
from run_v30 import enforce_final_sponsorship_account_guard


def _line(route, amount=100, vat_pct=15, ticket_no="4860000000"):
    return {
        "route": route,
        "taxable_amt": amount,
        "vat_pct": vat_pct,
        "ticket_no": ticket_no,
    }


def test_tax_code_uses_whole_route_and_zero_amount_precedence():
    assert derive_tax_classification(_line("RUH JED RUH")) == ("KSA VAT STANDARD", [])
    assert derive_tax_classification(_line("RUH JED ZRH RUH", vat_pct=15)) == (
        "KSA VAT ZERO", ["VENDOR_VAT_MISMATCH"]
    )
    assert derive_tax_classification(_line("JED EAM JED", amount=0, vat_pct=15)) == (
        "KSA VAT ZERO", ["VENDOR_VAT_MISMATCH"]
    )


def test_tax_code_no_route_is_reviewed_and_unknown_route_token_is_flagged():
    assert derive_tax_classification(_line(
        "ECS MUNICH - REGISTRATION", vat_pct=0, ticket_no="26-1046"
    )) == ("KSA VAT ZERO", ["TAX_CODE_NEEDS_REVIEW"])
    assert derive_tax_classification(_line(
        "Crowne Plaza Riyadh RDC Hotel", vat_pct=15, ticket_no="26-1047"
    )) == ("KSA VAT STANDARD", ["TAX_CODE_NEEDS_REVIEW"])
    assert derive_tax_classification(_line("RUH XYZ RUH", vat_pct=0)) == (
        "KSA VAT ZERO", ["TAX_CODE_UNKNOWN_TOKEN"]
    )


def test_final_sponsorship_guard_keeps_employee_number(tmp_path):
    path = tmp_path / "out.xlsx"
    wb = Workbook()
    ws = wb.active
    headers = ["Account", "Distribution Combination[..]", "Employee No", "Agent Flags"]
    for col, value in enumerate(headers, 1):
        ws.cell(3, col, value)
    ws.append(["21070229", "03-10100-21070229-160014-170-10017-10072-00000-00-000000", "1000640", "CLEAN"])
    wb.save(path)
    hybrid = [{"_row_idx": 4, "account": "21070229", "emp_no": "1000640", "_flags": "SPONSORSHIP_DETECTED"}]

    assert enforce_final_sponsorship_account_guard(path, hybrid, [{}], 3) == 1
    ws = load_workbook(path, data_only=True).active
    assert ws.cell(4, 1).value == "60307021"
    assert ws.cell(4, 2).value.split("-")[2] == "60307021"
    assert ws.cell(4, 3).value == "1000640"
    assert "SPONSORSHIP_ANNUAL_OVERRIDE_BLOCKED" in ws.cell(4, 4).value
    assert hybrid[0]["account"] == "60307021"
    assert hybrid[0]["emp_no"] == "1000640"
