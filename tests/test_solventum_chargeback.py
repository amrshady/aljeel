from pathlib import Path

from openpyxl import load_workbook

from scripts.solventum_chargeback import OUTPUT_COLUMNS, collect_pod_trx_numbers, generate_chargeback


ROOT = Path(__file__).resolve().parents[1]
SAMPLES = ROOT / "batches" / "solventum"


def _sample_rows(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["Sheet1"]
        rows = list(sheet.iter_rows(values_only=True))
        header_index = next(index for index, row in enumerate(rows) if row[0] == "TRX #")
        return [tuple(row[: len(OUTPUT_COLUMNS)]) for row in rows[header_index + 1:] if row[0] is not None]
    finally:
        workbook.close()


def _headers(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        rows = workbook["Sheet1"].iter_rows(values_only=True)
        return next(tuple(row[: len(OUTPUT_COLUMNS)]) for row in rows if row[0] == "TRX #")
    finally:
        workbook.close()


def test_collects_all_trx_tokens_from_bundled_pdf_name():
    pods = [Path("2600014513, 2600015192 checked.pdf"), Path("ignore 26125380.pdf")]
    assert collect_pod_trx_numbers(pods) == {"2600014513", "2600015192"}


def test_generator_against_both_sample_waves(tmp_path):
    sales = SAMPLES / "JUNE_SALES_2026.xlsx"
    expected_counts = {
        "chargeback_1st_wave_SAMPLE.xlsx": {"sample": 52, "derived": 52},
        # The sample omits four Sheet2 lines for covered TRX 2600015875. The
        # locked derivation keeps every covered sales line, hence 90 (not 86).
        "chargeback_2nd_wave_SAMPLE.xlsx": {"sample": 86, "derived": 90},
    }

    for sample_name, counts in expected_counts.items():
        expected = _sample_rows(SAMPLES / sample_name)
        assert len(expected) == counts["sample"]
        trx_numbers = sorted({str(row[0]) for row in expected})
        pods = [tmp_path / f"{trx}.pdf" for trx in trx_numbers]
        output = tmp_path / f"generated-{sample_name}"

        row_count = generate_chargeback(sales, pods, output)
        actual = _sample_rows(output)

        assert row_count == counts["derived"]
        assert len(actual) == counts["derived"]
        sample_headers = list(_headers(SAMPLES / sample_name))
        # The second-wave reference uses the legacy label "TRX Type" in its
        # third cell. The locked build brief supersedes it with "Order Type".
        if sample_headers[2] == "TRX Type":
            sample_headers[2] = "Order Type"
        assert _headers(output) == tuple(sample_headers) == OUTPUT_COLUMNS
        assert {str(row[0]) for row in actual} == set(trx_numbers)
        assert all(not str(row[6]).startswith(("3MOC-", "3MOR-")) for row in actual)
