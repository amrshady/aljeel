import json
from pathlib import Path
import sys

import openpyxl


ROOT = Path("/home/clawdbot/.openclaw/workspace/aljeel")
sys.path.insert(0, str(ROOT / "scripts"))

import run_v30
import split_multi_emp as splitter
from run_v16 import enforce_sponsorship_rules, resolve_sponsorship_from_master


def _home(cc="999999", div="999", solution="99999", agency="99998"):
    return {
        "cost_center": cc, "div_code": div, "solution": solution,
        "agency_code": agency, "location": "10100",
    }


def test_enforcement_does_not_blank_sponsorship_employee_number():
    final = {"account": "60307021", "emp_no": "1001422"}
    assert enforce_sponsorship_rules(final, {"row_type": "sponsorship"}, {})["emp_no"] == "1001422"


def test_sponsorship_master_resolution_keeps_provisional_requester(monkeypatch):
    monkeypatch.setattr(
        "run_v16.resolve_sponsorship_codes_from_agency",
        lambda agency, records: (None, "AGENCY_FROM_FORM_MISSING"),
    )
    result = resolve_sponsorship_from_master("1000433", {"1000433": _home()}, {})
    assert result["account"] == "60307021"
    assert result["emp_no"] == "1000433"


def test_nested_opex_pdf_discovery_and_sis_serial(tmp_path):
    nested = tmp_path / "event" / "descriptive child"
    nested.mkdir(parents=True)
    expected = nested / "OPEX-SIS-15-2026.PDF"
    expected.write_bytes(b"fixture")
    (nested / "itinerary.pdf").write_bytes(b"fixture")
    assert run_v30._find_opex_pdfs(tmp_path) == [expected]
    assert run_v30._extract_serial_from_text(expected.name) == "SIS-15-2026"


def test_normalized_event_serial_variants_are_equivalent_and_prefix_safe():
    assert run_v30._canonical_event_serial("CE-202-26") == "CE20"
    assert run_v30._canonical_event_serial("CE-20-2026") == "CE20"
    assert run_v30._canonical_event_serial("OPEX-CE-202-26-J-2026-120-new.pdf") == "CE20"
    assert run_v30._canonical_event_serial("OPEX-LAB-16-2026J-2026-126.pdf") == "LAB16"
    assert run_v30._canonical_event_serial("lab-16") == "LAB16"
    assert run_v30._canonical_event_serial("SIS-14-2026") == "SIS14"
    assert run_v30._canonical_event_serial("SIS-15-2026") == "SIS15"


def test_event_form_index_handles_missing_dash_case_and_new_suffix(tmp_path):
    ce = tmp_path / "CE-20-2026" / "OPEX-CE-202-26-J-2026-120-new.pdf"
    lab = tmp_path / "lab-16-2026" / "OPEX-LAB-16-2026J-2026-126.pdf"
    ce.parent.mkdir(); lab.parent.mkdir()
    ce.write_bytes(b"fixture"); lab.write_bytes(b"fixture")
    index = run_v30._build_opex_event_pdf_index([tmp_path])
    assert index["CE20"] == [ce]
    assert index["LAB16"] == [lab]


def test_no_orphan_line_when_normalized_event_form_exists(tmp_path, monkeypatch):
    form_dir = tmp_path / "SIS-14-2026"
    form_dir.mkdir()
    pdf = form_dir / "OPEX-SIS-14-2026-J-2026-123.pdf"
    pdf.write_bytes(b"fixture")
    allocations = [
        {"emp_no": "1001422", "name": "One", "amount": "10,000"},
        {"emp_no": "1002169", "name": "Two", "amount": "10,000"},
        {"emp_no": "1001530", "name": "Three", "amount": "10,000"},
    ]
    monkeypatch.setattr(
        run_v30, "_extract_sponsorship_allocations_from_opex_pdf",
        lambda path, *args: ([item["emp_no"] for item in allocations], allocations),
    )
    rows = [
        {"account": "60307021", "emp_no": ""},
        {"account": "60307021", "emp_no": ""},
    ]
    cascades = [
        {"Invoice Ref No": "SIS-14-2026", "Description": "NAWAF (4860205136)"},
        {"Invoice Ref No": "sis-14", "Description": "ABDULLAH (26-908)"},
    ]
    assert run_v30.apply_sponsorship_allocations(
        rows, cascades, tmp_path, [tmp_path], {}, {}
    ) == (2, 0)
    assert {row["emp_no"] for row in rows} == {"1001422,1002169,1001530"}
    assert all(row["_sponsorship_allocations"] == allocations for row in rows)


def test_real_sis15_form_extracts_full_exact_allocation_table():
    pdf = Path("/mnt/aljeel_ap_kb/current/J26-1029/28jun/SIS-15-2026/payment form for conference on July 1-2/OPEX for MOVIVA Expert Training.pdf")
    salesmen, allocations = run_v30._extract_sponsorship_allocations_from_opex_pdf(pdf)
    assert salesmen == ["1001422", "1002169", "1001530"]
    assert [(a["emp_no"], a["name"], a["amount"]) for a in allocations] == [
        ("1001422", "Wasseem Mustafa", "24,000"),
        ("1002169", "Abdallah Amoudi", "24,000"),
        ("1001530", "Belal Ahmed", "24,000"),
    ]


def test_short_unresolved_form_label_is_not_guessed_by_substring():
    assert run_v30._lookup_segment_code_by_name("solution", "GI") == ""


def test_short_exact_form_label_resolves_from_master():
    assert run_v30._lookup_segment_code_by_name("div", "SIS") == "130"


def test_agency_ocr_matching_is_token_safe_and_exact():
    assert run_v30._unique_agency_from_text("Division: Capital equipment") == ""
    assert run_v30._unique_agency_from_text("Agency: Getinge") == "10100"


def test_real_sis14_form_keeps_its_own_exact_amounts():
    pdf = Path("/mnt/aljeel_ap_kb/current/J26-1029/30jun/SIS-14-2026/DBE - Training for Qassim change/OPEX-SIS-14-2026-J-2026-123.pdf")
    _salesmen, allocations = run_v30._extract_sponsorship_allocations_from_opex_pdf(pdf)
    assert [(a["emp_no"], a["amount"]) for a in allocations] == [
        ("1001422", "10,000"),
        ("1002169", "10,000"),
        ("1001530", "10,000"),
    ]


def test_real_scanned_ce_and_lab_forms_extract_all_allocation_employees():
    manpower = run_v30.fea.load_manpower()
    ce = Path("/mnt/aljeel_ap_kb/current/J26-1029/29jun/CE-20-2026/Getinge Service training sponsorship/OPEX-CE-202-26-J-2026-120-new.pdf")
    lab = Path("/mnt/aljeel_ap_kb/current/J26-1029/27jun/lab-16-2026/RE 24 June event in RDC crown plaza Riyadh/OPEX-LAB-16-2026J-2026-126.pdf")
    _, ce_allocations = run_v30._extract_sponsorship_allocations_from_opex_pdf(ce, manpower)
    _, lab_allocations = run_v30._extract_sponsorship_allocations_from_opex_pdf(lab, manpower)
    assert [(a["emp_no"], a["amount"]) for a in ce_allocations] == [
        ("1002075", "30,000.00"), ("1001986", "10,000.00"),
    ]
    assert [a["emp_no"] for a in lab_allocations] == [
        "1000862", "1002144", "1001059", "1000414", "1001256",
    ]
    assert all(a["amount"] == "" for a in lab_allocations)


def test_participant_folder_overrides_ref_folder_and_flags_mismatch(tmp_path, monkeypatch):
    sis14 = tmp_path / "SIS-14-2026"; sis14.mkdir()
    sis15 = tmp_path / "SIS-15-2026"; sis15.mkdir()
    row = {"_evidence_folder": str(sis14), "_invoice_ref_folder_status": "REF_FOLDER"}
    cascade = {"Description": "SIS-14-2026-ADEEB - TRAIN TICKET", "Invoice Ref No": "SIS-14-2026"}
    monkeypatch.setattr(
        run_v30, "_event_folder_by_participant",
        lambda cascade_row, all_folders: (sis15, "PARTICIPANT_FOLDER"),
    )
    found = run_v30._sponsorship_event_folder_for_row(
        row, cascade, tmp_path, [sis14, sis15], {}
    )
    assert found == sis15
    assert "SPONSORSHIP_REF_FOLDER_MISMATCH" in row["_agent_flag_details"]


def test_exact_non_sis_ref_folder_is_not_overridden_by_common_participant(tmp_path, monkeypatch):
    crm31 = tmp_path / "CRM-2026-31"; crm31.mkdir()
    crm38 = tmp_path / "CRM-2026-38"; crm38.mkdir()
    row = {"_evidence_folder": str(crm31), "_invoice_ref_folder_status": "REF_FOLDER"}
    cascade = {"Description": "KHALED ELENIZI - REGISTRATION", "Invoice Ref No": "CRM-2026-31"}
    monkeypatch.setattr(
        run_v30, "_event_folder_by_participant",
        lambda cascade_row, all_folders: (crm38, "PARTICIPANT_FOLDER"),
    )
    assert run_v30._sponsorship_event_folder_for_row(
        row, cascade, tmp_path, [crm31, crm38], {}
    ) == crm31
    assert "_agent_flag_details" not in row


def test_ambiguous_participant_folder_never_falls_back_by_serial(tmp_path, monkeypatch):
    sis14 = tmp_path / "SIS-14-2026"; sis14.mkdir()
    pdf = sis14 / "OPEX-SIS-14-2026.pdf"; pdf.write_bytes(b"fixture")
    final = {"account": "60307021", "emp_no": "1001422", "opex_serial": "SIS-14-2026"}
    monkeypatch.setattr(run_v30, "_find_shared_opex_pdfs", lambda serial: [pdf])
    run_v30._apply_multi_salesman_from_opex(
        final,
        {"_folder": "", "opex_code": "SIS-14-2026", "_folder_status": "PARTICIPANT_FOLDER_AMBIGUOUS"},
        1,
    )
    assert final["emp_no"] == ""
    assert "SPONSORSHIP_ALLOCATION_TABLE_REVIEW" in final["_agent_flag_details"]


def test_apply_allocation_replaces_requester_with_all_table_employees(tmp_path, monkeypatch):
    pdf = tmp_path / "OPEX-CRM-2026-31.pdf"
    pdf.write_bytes(b"fixture")
    allocations = [
        {"emp_no": "1001762", "name": "One", "amount": "1,250.00"},
        {"emp_no": "1000433", "name": "Two", "amount": "750.00"},
    ]
    monkeypatch.setattr(
        run_v30, "_extract_sponsorship_allocations_from_opex_pdf",
        lambda path: ([a["emp_no"] for a in allocations], allocations),
    )
    final = {"account": "60307021", "emp_no": "1001762"}
    run_v30._apply_multi_salesman_from_opex(final, {"_folder": str(tmp_path)}, 7)
    assert final["emp_no"] == "1001762,1000433"
    assert final["_sponsorship_allocations"] == allocations


def test_apply_allocation_accepts_employee_list_without_form_amounts(tmp_path, monkeypatch):
    pdf = tmp_path / "OPEX-CRM-2026-31.pdf"
    pdf.write_bytes(b"fixture")
    allocations = [
        {"emp_no": "1001762", "name": "One", "amount": ""},
        {"emp_no": "1000433", "name": "Two", "amount": ""},
    ]
    monkeypatch.setattr(
        run_v30, "_extract_sponsorship_allocations_from_opex_pdf",
        lambda path: ([a["emp_no"] for a in allocations], allocations),
    )
    final = {"account": "60307021", "emp_no": "1001762"}
    run_v30._apply_multi_salesman_from_opex(final, {"_folder": str(tmp_path)}, 7)
    assert final["emp_no"] == "1001762,1000433"
    assert final["_sponsorship_allocations"] == allocations
    assert "SPONSORSHIP_ALLOCATION_AMOUNT_REVIEW" not in final.get("_agent_flag_details", "")


def test_explicit_form_segments_are_shared_by_allocated_employees(tmp_path, monkeypatch):
    row = {
        "account": "60307021", "emp_no": "1001422,1002169,1001530",
        "cost_center": "999999", "div": "999", "solution": "99999", "agency": "99998",
        "_evidence_folder": str(tmp_path), "_sponsorship_requesting_emp_no": "1001422",
    }
    cascade = {
        "Form Cost-Center-Ref (Fusion 15-digit)": "160011",
        "Form Division (Fusion code)": "130",
        "Form Solution (Fusion code)": "00000",
        "Form Agency (Fusion code)": "10043",
    }
    monkeypatch.setattr(
        run_v30, "resolve_sponsorship_codes_from_agency",
        lambda agency, manpower: (None, "AGENCY_CODES_INCONSISTENT"),
    )
    run_v30.apply_sponsorship_event_segments(
        [row], [cascade], {"1001422": _home()}, tmp_path, [tmp_path], {}
    )
    assert (row["cost_center"], row["div"], row["solution"], row["agency"]) == (
        "160011", "130", "00000", "10043"
    )
    assert row["emp_no"] == "1001422,1002169,1001530"


def _patch_split_dependencies(monkeypatch):
    monkeypatch.setattr(splitter, "load_manpower", lambda path: {})
    monkeypatch.setattr(splitter, "load_solution_names", lambda path: {})
    monkeypatch.setattr(splitter, "get_lookup", lambda path: object())


def test_splitter_conserves_line_amount_with_equal_form_ratios_and_preserves_event_segments(tmp_path, monkeypatch):
    source, output = tmp_path / "source.xlsx", tmp_path / "split.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    headers = [
        "Employee No", "*Amount", "Distribution Combination", "Account", "Agent Flags",
        "Cost Center", "DIV", "Solution", "Agency", "OPEX Allocation Details",
    ]
    ws.append(headers)
    allocations = [
        {"emp_no": "1001422", "name": "Wasseem Mustafa", "amount": "24,000"},
        {"emp_no": "1002169", "name": "Abdallah Amoudi", "amount": "24,000"},
        {"emp_no": "1001530", "name": "Belal Ahmed", "amount": "24,000"},
    ]
    ws.append([
        "1001422,1002169,1001530", 100,
        "03-40100-60307021-160011-130-00000-10043-00000-00-000000",
        "60307021", "CLEAN", "160011", "130", "00000", "10043", json.dumps(allocations),
    ])
    wb.save(source); wb.close()
    _patch_split_dependencies(monkeypatch)
    result = splitter.split_multi_emp(str(source), str(output), "unused.xlsx")
    wb = openpyxl.load_workbook(output, data_only=True); ws = wb.active
    assert result == {"total_rows": 1, "split_rows": 1, "output_rows": 3}
    assert [(ws.cell(r, 1).value, ws.cell(r, 2).value) for r in range(2, 5)] == [
        ("1001422", 33.34), ("1002169", 33.33), ("1001530", 33.33),
    ]
    assert sum(ws.cell(r, 2).value for r in range(2, 5)) == 100
    assert {tuple(ws.cell(r, c).value for c in range(6, 10)) for r in range(2, 5)} == {
        ("160011", "130", "00000", "10043")
    }
    wb.close()


def test_splitter_conserves_line_amount_with_proportional_form_ratios(tmp_path, monkeypatch):
    source, output = tmp_path / "source.xlsx", tmp_path / "split.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append([
        "Employee No", "*Amount", "Distribution Combination", "Account", "Agent Flags",
        "OPEX Allocation Details",
    ])
    allocations = [
        {"emp_no": "1000001", "name": "One", "amount": "3,000"},
        {"emp_no": "1000002", "name": "Two", "amount": "2,000"},
        {"emp_no": "1000003", "name": "Three", "amount": "1,000"},
    ]
    ws.append([
        "1000001,1000002,1000003", 100,
        "03-40100-60307021-160011-130-00000-10043-00000-00-000000",
        "60307021", "CLEAN", json.dumps(allocations),
    ])
    wb.save(source); wb.close()
    _patch_split_dependencies(monkeypatch)
    splitter.split_multi_emp(str(source), str(output), "unused.xlsx")
    wb = openpyxl.load_workbook(output, data_only=True); ws = wb.active
    amounts = [ws.cell(r, 2).value for r in range(2, 5)]
    assert amounts == [50, 33.33, 16.67]
    assert sum(amounts) == 100
    wb.close()


def test_splitter_conserves_line_amount_with_no_form_amounts(tmp_path, monkeypatch):
    source, output = tmp_path / "source.xlsx", tmp_path / "split.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append([
        "Employee No", "*Amount", "Distribution Combination", "Account", "Agent Flags",
        "OPEX Allocation Details",
    ])
    allocations = [
        {"emp_no": "1000001", "name": "One", "amount": ""},
        {"emp_no": "1000002", "name": "Two", "amount": ""},
        {"emp_no": "1000003", "name": "Three", "amount": ""},
    ]
    ws.append([
        "1000001,1000002,1000003", 10,
        "03-40100-60307021-160011-130-00000-10043-00000-00-000000",
        "60307021", "CLEAN", json.dumps(allocations),
    ])
    wb.save(source); wb.close()
    _patch_split_dependencies(monkeypatch)
    splitter.split_multi_emp(str(source), str(output), "unused.xlsx")
    wb = openpyxl.load_workbook(output, data_only=True); ws = wb.active
    amounts = [ws.cell(r, 2).value for r in range(2, 5)]
    assert amounts == [3.34, 3.33, 3.33]
    assert sum(amounts) == 10
    wb.close()


def test_splitter_flags_sponsorship_without_allocation_table(tmp_path, monkeypatch):
    source, output = tmp_path / "source.xlsx", tmp_path / "split.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["Employee No", "*Amount", "Distribution Combination", "Account", "Agent Flags"])
    ws.append(["1001422", 5000, "03-40100-60307021-160011-130-00000-10043-00000-00-000000", "60307021", "CLEAN"])
    wb.save(source); wb.close()
    _patch_split_dependencies(monkeypatch)
    splitter.split_multi_emp(str(source), str(output), "unused.xlsx")
    wb = openpyxl.load_workbook(output, data_only=True); ws = wb.active
    assert ws["A2"].value in (None, "")
    assert "SPONSORSHIP_ALLOCATION_TABLE_REVIEW" in ws["E2"].value
    wb.close()


def test_splitter_does_not_resurrect_invalidated_stale_allocations(tmp_path, monkeypatch):
    source, output = tmp_path / "source.xlsx", tmp_path / "split.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append([
        "Employee No", "*Amount", "Distribution Combination", "Account", "Agent Flags",
        "OPEX Allocation Details",
    ])
    stale = [{"emp_no": "1001422", "name": "Wasseem", "amount": "24,000"}]
    ws.append([
        "", 9800, "03-40100-60307021-160011-130-00000-10043-00000-00-000000",
        "60307021", "SPONSORSHIP_REF_FOLDER_MISMATCH", json.dumps(stale),
    ])
    wb.save(source); wb.close()
    _patch_split_dependencies(monkeypatch)
    splitter.split_multi_emp(str(source), str(output), "unused.xlsx")
    wb = openpyxl.load_workbook(output, data_only=True); ws = wb.active
    assert ws["A2"].value in (None, "")
    assert ws["B2"].value == 9800
    assert "SPONSORSHIP_ALLOCATION_TABLE_REVIEW" in ws["E2"].value
    wb.close()


def test_non_sponsorship_multi_employee_keeps_legacy_equal_split(tmp_path, monkeypatch):
    source, output = tmp_path / "source.xlsx", tmp_path / "split.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["Employee No", "*Amount", "Distribution Combination", "Account", "Agent Flags"])
    ws.append(["1000001,1000002", 100, "03-40100-60301003-160011-130-00000-10043-00000-00-000000", "60301003", "CLEAN"])
    wb.save(source); wb.close()
    _patch_split_dependencies(monkeypatch)
    splitter.split_multi_emp(str(source), str(output), "unused.xlsx")
    wb = openpyxl.load_workbook(output, data_only=True); ws = wb.active
    assert [(ws.cell(r, 1).value, ws.cell(r, 2).value) for r in (2, 3)] == [
        ("1000001", 50), ("1000002", 50)
    ]
    wb.close()


def test_late_allocation_pass_restores_non_sponsorship_employee_number(tmp_path):
    rows = [{"account": "60301003", "emp_no": "1001422,1002169,1001530"}]
    cascades = [{"Employee No": "1000986"}]
    assert run_v30.apply_sponsorship_allocations(
        rows, cascades, tmp_path, [], {}, {}
    ) == (0, 0)
    assert rows[0]["emp_no"] == "1000986"
