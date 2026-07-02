from pathlib import Path
import sys

ROOT = Path("/home/clawdbot/.openclaw/workspace/aljeel")
sys.path.insert(0, str(ROOT / "scripts"))

from run_v30 import (
    OPEX_EMAIL_ONLY_FLAG,
    apply_opex_email_only_review_state,
    apply_sibling_leg_trip_purpose_inheritance,
    collect_family_annual_rows,
    inherit_ancillary_event_allocations,
    row_missing_evidence,
    stamp_verified_annual_home_segments,
)
from split_multi_emp import _is_event_sponsorship_row


def test_chd_emp_group_requires_family_evidence():
    hybrid_rows = [
        {"emp_no": "1000001", "account": "60301003", "_evidence_folder": "/evidence/adult"},
        {"emp_no": "1000001", "account": "60301003", "_evidence_folder": "/evidence/1234567890 family"},
    ]
    cascade_rows = [
        {"Description": "UNRELATED/ADULT MR - AAA BBB (1000000001)"},
        {"Description": "FAMILY/CHILD MR(CHD) - AAA BBB (1000000002)"},
    ]

    assert collect_family_annual_rows(hybrid_rows, cascade_rows, {}) == {}


def test_family_folder_child_uses_folder_owner_without_emp_contamination():
    hybrid_rows = [
        {"emp_no": "1000001", "account": "60301003", "_evidence_folder": "/evidence/1234567890 family"},
        {"emp_no": "1009999", "account": "60301003", "_evidence_folder": "/evidence/1234567890 family"},
    ]
    cascade_rows = [
        {"Description": "OWNER/PARENT MR - AAA BBB (1234567890)"},
        {"Description": "OWNER/CHILD MR(CHD) - AAA BBB (1234567891)"},
    ]

    fam = collect_family_annual_rows(
        hybrid_rows,
        cascade_rows,
        {"1234567890": "1000001"},
    )

    assert fam == {
        0: ("1000001", "family_folder:1234567890"),
        1: ("1000001", "family_folder:1234567890"),
    }


def test_verified_annual_rows_stamp_home_segments_but_keep_general_solution():
    hybrid_rows = [{
        "emp_no": "1000001",
        "account": "21070229",
        "cost_center": "",
        "div": "",
        "solution": "",
        "agency": "",
        "_agent_method": "cascade",
    }]
    cascade_rows = [{"Description": "OWNER/PERSON MR - AAA BBB (1234567890)"}]
    manpower = {
        "1000001": {
            "cost_center": "123456",
            "div_code": "321",
            "solution": "99999",
            "agency_code": "54321",
        }
    }

    assert stamp_verified_annual_home_segments(hybrid_rows, cascade_rows, manpower) == 1
    assert hybrid_rows[0]["cost_center"] == "123456"
    assert hybrid_rows[0]["div"] == "321"
    assert hybrid_rows[0]["solution"] == "00000"
    assert hybrid_rows[0]["agency"] == "54321"


def test_splitter_detects_event_rows_from_structured_opex_evidence():
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    headers = [
        "Account",
        "OPEX Serial",
        "Form Division (Fusion code)",
        "Form Agency (Fusion code)",
        "Form Solution (Fusion code)",
        "Form Cost-Center-Ref (Fusion 15-digit)",
    ]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(1, col_idx, header)
    col = {header: idx for idx, header in enumerate(headers, 1)}

    ws.cell(2, col["Account"], "60307021")
    assert _is_event_sponsorship_row(ws, 2, col)

    ws.cell(3, col["OPEX Serial"], "DEPT-2026-1")
    assert _is_event_sponsorship_row(ws, 3, col)

    ws.cell(4, col["Form Agency (Fusion code)"], "12345")
    assert _is_event_sponsorship_row(ws, 4, col)


def test_missing_ancillary_row_inherits_resolved_event_by_attendee():
    hybrid_rows = [
        {
            "account": "60307021",
            "emp_no": "1000001",
            "cost_center": "111111",
            "div": "222",
            "solution": "33333",
            "agency": "44444",
            "_opex_serial": "EVT-2026-1",
            "_evidence_folder": "/evidence/event",
        },
        {
            "_missing_evidence": True,
            "_flags": "MISSING_EVIDENCE",
            "account": "",
            "emp_no": "",
            "cost_center": "",
            "div": "",
            "solution": "",
            "agency": "",
        },
    ]
    cascade_rows = [
        {"Description": "ATTENDEE/PERSON MR - AAA BBB (1234567890)"},
        {"Description": "ATTENDEE PERSON - AIRPORT TRANSFER (26-001)"},
    ]

    assert inherit_ancillary_event_allocations(hybrid_rows, cascade_rows) == 1
    assert hybrid_rows[1]["account"] == "60307021"
    assert hybrid_rows[1]["cost_center"] == "111111"
    assert hybrid_rows[1]["_missing_evidence"] is False
    assert row_missing_evidence(hybrid_rows[1]) is False


def test_sibling_leg_inherits_trip_purpose_from_shared_itinerary_pdf(tmp_path, monkeypatch):
    folder = tmp_path / "1234567890"
    folder.mkdir()
    pdf = folder / "itinerary.pdf"
    pdf.write_text("placeholder", encoding="utf-8")

    def fake_pdf_text(path):
        assert path == pdf
        return (
            "Booking ref:\nABC123\n"
            "Traveler\nMr Sample Person\n"
            "E-ticket XY 1234567890 for Mr Sample Person\n"
            "E-ticket XY 1234567891 for Mr Sample Person\n"
            "Airline Booking Reference(s)\nXY: ZX9YQ1\n"
        )

    monkeypatch.setattr("run_v30._pdf_text", fake_pdf_text)

    hybrid_rows = [
        {
            "account": "60308009",
            "emp_no": "1000001",
            "cost_center": "111111",
            "div": "222",
            "solution": "33333",
            "agency": "44444",
            "_own_form_trip_purpose": {
                "account_override": "60308009",
                "folder": str(folder),
            },
        },
        {"account": "60301003", "_agent_method": "cascade"},
    ]
    cascade_rows = [
        {
            "Description": "PERSON/SAMPLE MR - AAA BBB (1234567890)",
            "Trip Purpose": "TRAINING",
            "Trip Purpose Confidence": 0.92,
            "Trip Account Override": "60308009",
        },
        {
            "Description": "PERSON/SAMPLE MR - BBB AAA (1234567891)",
            "Trip Purpose": "UNKNOWN",
            "Trip Purpose Confidence": 0,
            "Trip Account Override": "",
        },
    ]

    assert apply_sibling_leg_trip_purpose_inheritance(hybrid_rows, cascade_rows) == 1
    assert hybrid_rows[1]["account"] == "60308009"
    assert hybrid_rows[1]["emp_no"] == "1000001"
    assert hybrid_rows[1]["cost_center"] == "111111"
    assert hybrid_rows[1]["div"] == "222"
    assert hybrid_rows[1]["solution"] == "33333"
    assert hybrid_rows[1]["agency"] == "44444"
    assert cascade_rows[1]["Trip Purpose"] == "TRAINING"


def test_opex_email_only_flags_review_without_exact_segments(tmp_path):
    folder = tmp_path / "OPEX City EP-2026-1"
    folder.mkdir()
    (folder / "Re_ OPEX City EP-2026-1.msg").write_text("OPEX City EP-2026-1 approved", encoding="utf-8")
    hybrid_rows = [{
        "account": "60301004",
        "cost_center": "111111",
        "div": "222",
        "solution": "33333",
        "agency": "44444",
        "_evidence_folder": str(folder),
    }]
    cascade_rows = [{"Description": "PERSON/SAMPLE MR - AAA BBB (1234567890)", "Notes": ""}]

    assert apply_opex_email_only_review_state(
        hybrid_rows,
        cascade_rows,
        tmp_path,
        [folder],
        {},
    ) == 1
    assert hybrid_rows[0]["account"] == "60307021"
    assert hybrid_rows[0]["agency"] == ""
    assert OPEX_EMAIL_ONLY_FLAG in hybrid_rows[0]["_flags"]
